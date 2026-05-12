import os
import re
import json
import urllib.request
import urllib.parse
import logging
from flask import Flask, request, abort, jsonify
from linebot.v3 import WebhookHandler
from linebot.v3.messaging import (
    Configuration, ApiClient, MessagingApi, MessagingApiBlob,
    ReplyMessageRequest, TextMessage, FlexMessage, FlexContainer,
    QuickReply, QuickReplyItem, MessageAction, PushMessageRequest,
    MulticastRequest,
)
try:
    from linebot.v3.messaging import (
        TemplateMessage, ConfirmTemplate, ButtonsTemplate,
        CarouselTemplate, CarouselColumn,
        PostbackAction, URIAction as MsgURIAction,
        ImagemapMessage, ImagemapBaseSize, ImagemapArea,
        MessageImagemapAction, URIImagemapAction,
        BroadcastRequest,
    )
except ImportError:
    TemplateMessage = None
    BroadcastRequest = None
try:
    from linebot.v3.messaging import ShowLoadingAnimationRequest
except ImportError:
    ShowLoadingAnimationRequest = None
try:
    from linebot.v3.messaging import MarkMessagesAsReadRequest
except ImportError:
    MarkMessagesAsReadRequest = None
try:
    from linebot.v3.messaging import (
        DatetimePickerAction as MsgDatetimePickerAction,
        CameraAction as MsgCameraAction,
        CameraRollAction as MsgCameraRollAction,
        LocationAction as MsgLocationAction,
        ClipboardAction as MsgClipboardAction,
    )
except ImportError:
    MsgDatetimePickerAction = None
    MsgCameraAction = None
    MsgCameraRollAction = None
    MsgLocationAction = None
    MsgClipboardAction = None
try:
    from linebot.v3.messaging import ValidateMessageRequest
except ImportError:
    ValidateMessageRequest = None
try:
    from linebot.v3.messaging import Sender as MessageSender
except ImportError:
    MessageSender = None
try:
    from linebot.v3.messaging import (
        RichMenuRequest, RichMenuArea, RichMenuBounds, RichMenuSize,
        CreateRichMenuAliasRequest, URIAction,
    )
except ImportError:
    RichMenuRequest = None
from linebot.v3.webhooks import MessageEvent, TextMessageContent, ImageMessageContent, AudioMessageContent
try:
    from linebot.v3.webhooks import VideoMessageContent
except ImportError:
    VideoMessageContent = None
try:
    from linebot.v3.webhooks import FileMessageContent
except ImportError:
    FileMessageContent = None
try:
    from linebot.v3.webhooks import LocationMessageContent
except ImportError:
    LocationMessageContent = None
try:
    from linebot.v3.webhooks import StickerMessageContent
except ImportError:
    StickerMessageContent = None
try:
    from linebot.v3.webhooks import JoinEvent
except ImportError:
    JoinEvent = None
try:
    from linebot.v3.webhooks import MemberJoinedEvent
except ImportError:
    MemberJoinedEvent = None
try:
    from linebot.v3.webhooks import MemberLeftEvent
except ImportError:
    MemberLeftEvent = None
try:
    from linebot.v3.webhooks import FollowEvent, UnfollowEvent
except ImportError:
    FollowEvent = None
    UnfollowEvent = None
try:
    from linebot.v3.webhooks import LeaveEvent as BotLeaveEvent
except ImportError:
    BotLeaveEvent = None
try:
    from linebot.v3.webhooks import PostbackEvent
except ImportError:
    PostbackEvent = None
try:
    from linebot.v3.webhooks import UnsendEvent
except ImportError:
    UnsendEvent = None
try:
    from linebot.v3.webhooks import VideoPlayCompleteEvent
except ImportError:
    VideoPlayCompleteEvent = None
# v3.8: Reaction event for instant translation feedback.
# Members react with 👍/❤️ → auto-add to examples.
# Members react with 😢/😡 → auto-mark as wrong.
try:
    from linebot.v3.webhooks import MessageReactionEvent
except ImportError:
    MessageReactionEvent = None
from linebot.v3.exceptions import InvalidSignatureError
from openai import OpenAI
import base64
import tempfile
import time
import uuid

app = Flask(__name__)
# v3.9.30c B20 修補: 限制 upload 大小防 OOM
# 沒設的話 Flask default 沒限,惡意/誤傳大檔會吃光 Render 256-512MB RAM
# Storage Excel 通常 <2MB,Rich Menu image 1040x1040 PNG 通常 <1MB,給 8MB 緩衝
app.config['MAX_CONTENT_LENGTH'] = 8 * 1024 * 1024  # 8 MB
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

VERSION = "v3.9.30e-0504-22bugs+night-shift-allowance"

LINE_TOKEN = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN", "")
LINE_SECRET = os.environ.get("LINE_CHANNEL_SECRET", "")
OPENAI_KEY = os.environ.get("OPENAI_API_KEY", "")
ADMIN_KEY = os.environ.get("ADMIN_KEY", "changeme")
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = "onerkk/line-translator-bot"
LIFF_ID = os.environ.get("LIFF_ID", "")
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")

configuration = Configuration(access_token=LINE_TOKEN)
handler = WebhookHandler(LINE_SECRET)
oai = OpenAI(api_key=OPENAI_KEY, timeout=30.0) if OPENAI_KEY else None  # v3.9.24: 全域 30 秒 timeout

group_settings = {}
# Target language for Chinese translation per group, default "id"
group_target_lang = {}
# Image translation toggle per group, default True
group_img_settings = {}
# v3.9.18: 記錄最後 5 張收到的 LINE 圖片 message_id(供 /debug/recent-images 使用)
_last_image_received_msgs = []  # list of dicts: {msg_id, group_id, ts}

# v3.9.19: 全 webhook 事件磁碟 log,跨 worker 共享
def _event_log_path():
    for d in ("/tmp", "/var/data", "/data"):
        if os.path.isdir(d) and os.access(d, os.W_OK):
            return os.path.join(d, "bot_event_log.json")
    return "bot_event_log.json"

_EVENT_LOG_FILE = None
_EVENT_LOG_MAX = 200  # 最多保留 200 筆

# v3.9.30c B17 修補: LINE 訊息長度限制(5000 字)截斷工具
# LINE 平台規定 text message 最多 5000 字,超過會被 reject。
# 翻譯結果加上注解 / hint 可能超 5000;集中處理避免散落各處 [:4990]。
LINE_TEXT_MAX = 5000
def _clip_line_text(text, suffix="\n...(已截斷)"):
    """Clip text to LINE's 5000-char limit. Adds suffix to indicate truncation."""
    if not text:
        return text
    if len(text) <= LINE_TEXT_MAX:
        return text
    cap = LINE_TEXT_MAX - len(suffix)
    return text[:cap] + suffix


# v3.9.30c B19 修補: 安全 int 轉換
# 之前 admin endpoint 用 int(request.args.get("xxx", 100)),如果 query 帶 ?xxx=abc
# 會炸 ValueError → 500。改用此函式。
def _safe_int(val, default=0, min_val=None, max_val=None):
    """Safely convert val to int, return default on failure. Optionally clamp."""
    try:
        n = int(val)
    except (ValueError, TypeError):
        return default
    if min_val is not None and n < min_val:
        return min_val
    if max_val is not None and n > max_val:
        return max_val
    return n


# v3.9.30c B15 修補: LINE webhook 重發去重
# LINE 在 server 1 秒內沒回 200 OK 時會重發 webhook。
# OpenAI 翻譯通常 3-10 秒,LINE 會以為失敗 → 重發 → 同訊息翻譯兩次扣兩次錢。
# 雙重保險:
#   1. 用 event.delivery_context.is_redelivery 直接判斷(SDK 已暴露)
#   2. 後備:用 message_id 去重(60 秒 TTL),即使 SDK 沒帶 redelivery flag 也能擋
import collections as _collections_dedup
_processed_msg_ids = _collections_dedup.OrderedDict()  # message_id → ts
_PROCESSED_MSG_MAX = 1000
_PROCESSED_MSG_TTL = 60  # 60 秒 TTL,LINE 重發通常在 30 秒內

def _is_duplicate_message(message_id):
    """判斷此 message_id 是否近期已處理過(LINE 重發保護)"""
    if not message_id:
        return False
    now = int(time.time())
    # 清過期
    expired = [k for k, v in _processed_msg_ids.items() if now - v > _PROCESSED_MSG_TTL]
    for k in expired:
        _processed_msg_ids.pop(k, None)
    # 檢查
    if message_id in _processed_msg_ids:
        return True
    # 記錄
    _processed_msg_ids[message_id] = now
    if len(_processed_msg_ids) > _PROCESSED_MSG_MAX:
        _processed_msg_ids.popitem(last=False)  # FIFO 移除最舊
    return False

def _is_redelivery(event):
    """判斷 LINE webhook 是不是重發(SDK + message_id 雙重檢查)"""
    try:
        dc = getattr(event, 'delivery_context', None)
        if dc and getattr(dc, 'is_redelivery', False):
            return True
    except Exception:
        pass
    return False


def _event_log_write(event_type, data):
    """寫一筆事件到磁碟 log。失敗安靜略過,不影響主流程。"""
    global _EVENT_LOG_FILE
    try:
        if _EVENT_LOG_FILE is None:
            _EVENT_LOG_FILE = _event_log_path()
        # 讀現有 log
        existing = []
        try:
            if os.path.exists(_EVENT_LOG_FILE):
                with open(_EVENT_LOG_FILE, "r", encoding="utf-8") as f:
                    existing = json.load(f)
                if not isinstance(existing, list):
                    existing = []
        except Exception:
            existing = []
        # 追加
        entry = {
            "ts": int(time.time()),
            "type": event_type,
            "data": data,
        }
        existing.append(entry)
        # 截尾保留最後 _EVENT_LOG_MAX 筆
        if len(existing) > _EVENT_LOG_MAX:
            existing = existing[-_EVENT_LOG_MAX:]
        # 原子寫
        tmp = _EVENT_LOG_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False)
        os.replace(tmp, _EVENT_LOG_FILE)
    except Exception:
        pass  # 永遠不阻擋主流程
# v3.9.10: 圖片翻譯詢問模式 — 收到圖片不自動翻,先問使用者要不要翻
# 結構:group_img_settings=False AND group_img_ask_settings=True → 詢問模式
#       group_img_settings=False AND group_img_ask_settings=False → 完全不翻
#       group_img_settings=True → 自動翻譯(原本行為)
group_img_ask_settings = {}
# v3.9.13: pending image 跨 worker 共享 — 寫到本地檔案(避免 gunicorn 多 worker 看不到彼此)
# 1 分鐘自動清掉沒按的(短 TTL 避免誤觸,需要翻譯會立刻按)
_PENDING_IMG_TTL = 60
def _pending_img_path():
    """選擇可寫入的暫存路徑 — 同 worker 共用,deploy 重啟即清空(這就是要的行為)"""
    for d in ("/var/data", "/data", "/tmp"):
        if os.path.isdir(d) and os.access(d, os.W_OK):
            return os.path.join(d, "pending_img_translate.json")
    return "pending_img_translate.json"
_PENDING_IMG_FILE = None  # lazy init,等 logger 準備好後設

def _load_pending_imgs():
    """讀檔。檔案不存在 / 壞掉 → 回 {}。每次讀都過濾掉過期的。"""
    global _PENDING_IMG_FILE
    if _PENDING_IMG_FILE is None:
        _PENDING_IMG_FILE = _pending_img_path()
    try:
        if not os.path.exists(_PENDING_IMG_FILE):
            return {}
        with open(_PENDING_IMG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {}
        # 過濾過期
        now = int(time.time())
        return {k: v for k, v in data.items()
                if isinstance(v, dict) and now - v.get("ts", 0) <= _PENDING_IMG_TTL}
    except Exception:
        return {}

def _save_pending_imgs(data):
    """原子寫入 — 用 write+rename 避免被讀到一半"""
    global _PENDING_IMG_FILE
    if _PENDING_IMG_FILE is None:
        _PENDING_IMG_FILE = _pending_img_path()
    try:
        tmp = _PENDING_IMG_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f)
        os.replace(tmp, _PENDING_IMG_FILE)
        return True
    except Exception as e:
        try:
            logger.warning("[ImgAsk] save pending failed: %s", e)
        except Exception:
            pass
        return False

def _pending_img_set(message_id, info):
    """加入一筆 pending,先 load 再 add 再 save(避免覆蓋其他 worker 寫的)"""
    data = _load_pending_imgs()
    data[message_id] = info
    _save_pending_imgs(data)

def _pending_img_pop(message_id):
    """取出並刪除一筆 pending(類似 dict.pop)"""
    data = _load_pending_imgs()
    info = data.pop(message_id, None)
    if info is not None:
        _save_pending_imgs(data)
    return info
# Audio/voice translation toggle per group, default True
group_audio_settings = {}
# Work order photo detection toggle per group, default True
group_wo_settings = {}
# Per-group command toggles: {group_id: {"pw1": bool, "pw2": bool, ...}}
group_cmd_enabled = {}
# Command definitions: (key, emoji+label_on, emoji+label_off, default)
CMD_DEFS = [
    ("pw1",    "🔑密碼1開", "🔑密碼1關", True),
    ("pw2",    "🏭密碼2開", "🏭密碼2關", True),
    ("pkg",    "📦包裝開",  "📦包裝關",  True),
    ("scrap",  "🎨廢料開",  "🎨廢料關",  True),
    ("qry",    "🔍儲區開",  "🔍儲區關",  True),
    ("notice", "📢公告開",  "📢公告關",  True),
]

def is_cmd_enabled(group_id, cmd_key):
    """Check if a command is enabled for a group."""
    cmds = group_cmd_enabled.get(group_id, {})
    # Find default from CMD_DEFS
    for key, _, _, default in CMD_DEFS:
        if key == cmd_key:
            return cmds.get(cmd_key, default)
    return True

# Skip list: set of user_ids per group whose messages won't be translated
group_skip_users = {}
# Track user display names per group: {group_id: {user_id: display_name}}
group_user_names = {}
# Group tracking: {group_id: {"name": str, "joined_at": float}}
group_tracking = {}

# DM (private message) target language per user, default "id"
dm_target_lang = {}
# DM master toggle (global on/off for all DM)
dm_master_enabled = True
# DM whitelist: set of user_ids allowed to DM when master is off
dm_whitelist = set()
# DM known users: {user_id: display_name} for anyone who has DM'd the bot
dm_known_users = {}

# Bot start time for uptime tracking
bot_start_time = time.time()

# Stats counters (resets on restart — Render free tier)
bot_stats = {
    "text_translations": 0,
    "image_translations": 0,
    "voice_translations": 0,
    "work_order_detections": 0,
    "commands": 0,
    "tokens_prompt": 0,
    "tokens_completion": 0,
}


def track_tokens(response):
    """Track token usage from OpenAI API response.
    
    v3.9.30 B8 修補: 加入 reasoning_tokens 記錄
    GPT-5 系列(o1/o3/gpt-5*)會把推理 token 放在
    response.usage.completion_tokens_details.reasoning_tokens,
    這部分 OpenAI 同樣計費(以 completion 費率),不記錄會造成成本黑洞。
    """
    try:
        if response and hasattr(response, 'usage') and response.usage:
            bot_stats["tokens_prompt"] += response.usage.prompt_tokens or 0
            bot_stats["tokens_completion"] += response.usage.completion_tokens or 0
            # v3.9.30: reasoning_tokens 額外追蹤(已包含在 completion_tokens 中,
            # 但需要分開知道才能算實際輸出 vs reasoning 比例)
            try:
                _details = getattr(response.usage, 'completion_tokens_details', None)
                if _details:
                    _rt = getattr(_details, 'reasoning_tokens', 0) or 0
                    bot_stats["tokens_reasoning"] = bot_stats.get("tokens_reasoning", 0) + _rt
                # cached prompt tokens(prompt cache hit 部分,計費較低)
                _pt_details = getattr(response.usage, 'prompt_tokens_details', None)
                if _pt_details:
                    _ct = getattr(_pt_details, 'cached_tokens', 0) or 0
                    bot_stats["tokens_prompt_cached"] = bot_stats.get("tokens_prompt_cached", 0) + _ct
            except Exception:
                pass
    except Exception:
        pass


def track_group_usage(group_id, before_prompt, before_completion):
    """Calculate token diff since snapshot and attribute to group."""
    dp = bot_stats.get("tokens_prompt", 0) - before_prompt
    dc = bot_stats.get("tokens_completion", 0) - before_completion
    if group_id and (dp > 0 or dc > 0):
        if group_id not in group_api_usage:
            group_api_usage[group_id] = {"tokens_prompt": 0, "tokens_completion": 0}
        group_api_usage[group_id]["tokens_prompt"] += dp
        group_api_usage[group_id]["tokens_completion"] += dc


def calc_group_cost_twd(group_id):
    """Calculate cost in TWD for a group.
    
    v3.9.30 B9 修補: 舊版寫死 GPT-4 系列價格 (input $0.15/M, output $0.60/M),
    用 GPT-5 系列時計費全錯。新版改用「當前主模型」的價格估算。
    
    這是估算值(因為單一群組可能混用多個模型),要精準計費需逐次記錄。
    使用 model_default 作為估算基準(歐那實際使用情境符合)。
    """
    u = group_api_usage.get(group_id, {})
    tp = u.get("tokens_prompt", 0)
    tc = u.get("tokens_completion", 0)
    # 依當前 model_default 估算單價(USD per token)
    md = (model_default or "").lower()
    # OpenAI 公開價格(2026-05);若未來價格有變更新此表即可
    # 格式: (input_per_M, output_per_M) USD
    PRICE_PER_M = {
        "gpt-5-nano":      (0.05,  0.40),
        "gpt-5-mini":      (0.25,  2.00),
        "gpt-5":           (1.25, 10.00),
        "gpt-5.4-nano":    (0.20,  1.25),
        "gpt-5.4-mini":    (0.75,  4.50),
        "gpt-5.4":         (2.50, 15.00),
        "gpt-5.5":         (5.00, 30.00),
        "gpt-4.1-nano":    (0.10,  0.40),
        "gpt-4.1-mini":    (0.40,  1.60),
        "gpt-4.1":         (2.00,  8.00),
        "gpt-4o-mini":     (0.15,  0.60),
        "gpt-4o":          (2.50, 10.00),
    }
    # 模糊匹配(歐那 UI 上的選項可能完整也可能 prefix)
    # v3.9.30b: 修匹配優先級 — 完全相等優先;否則用「最長 prefix」匹配,避免 "gpt-5.4-mini" 誤匹到 "gpt-5"
    in_rate, out_rate = (0.25, 2.00)  # 預設 gpt-5-mini
    # 1. 完全相等
    if md in PRICE_PER_M:
        in_rate, out_rate = PRICE_PER_M[md]
    else:
        # 2. 最長 prefix 匹配(避免 gpt-5.4-mini 被 gpt-5 吃掉)
        best_match = ""
        for mk in PRICE_PER_M.keys():
            if md.startswith(mk) and len(mk) > len(best_match):
                best_match = mk
        if best_match:
            in_rate, out_rate = PRICE_PER_M[best_match]
    # 換算:per_M → per_token
    usd = (tp * in_rate / 1_000_000) + (tc * out_rate / 1_000_000)
    return round(usd * USD_TO_TWD, 2)

# Admin users tracking: {user_id: {"is_admin": bool}}
admin_users = {}

# User language cache from LINE profile: {user_id: "id"|"zh-TW"|"en"|...}
user_languages = {}

# Per-group API usage tracking: {group_id: {"tokens_prompt": int, "tokens_completion": int}}
group_api_usage = {}

# ── Admin-controllable feature settings ──
# Welcome message: {enabled: bool, text_zh: str, text_id: str}
welcome_settings = {
    "enabled": True,
    "text_zh": "👋 歡迎新成員加入！\n本群組有 AI 翻譯助手，中文和印尼文會自動互譯。",
    "text_id": "👋 Selamat datang!\nGrup ini memiliki asisten penerjemah AI, bahasa Mandarin dan Indonesia akan diterjemahkan otomatis.",
}
# Flex message ON/OFF (True = Flex card, False = plain text)
flex_enabled = True
# Quick Reply buttons ON/OFF
quick_reply_enabled = True
# Silent mode: translation messages don't buzz the phone
silent_mode = False
# Video OCR translation ON/OFF
video_ocr_enabled = True
# Location translation ON/OFF
location_translate_enabled = True
# Mark-as-read ON/OFF (shows 'read' indicator in chat)
mark_read_enabled = True
# X-Line-Retry-Key ON/OFF (idempotent message sending)
retry_key_enabled = True
# Camera Quick Reply button ON/OFF
camera_qr_enabled = True
# Clipboard Quick Reply button ON/OFF (copy storage zone etc.)
clipboard_qr_enabled = False
# Camera Roll Quick Reply button ON/OFF
camera_roll_qr_enabled = False
# Location Quick Reply button ON/OFF
location_qr_enabled = False
# Per-group feature overrides (group_id -> bool), global values above are defaults
group_flex_settings = {}      # per-group flex card toggle
group_qr_settings = {}        # per-group quick reply toggle
group_silent_settings = {}    # per-group silent mode toggle
group_video_settings = {}     # per-group video OCR toggle
group_location_settings = {}  # per-group location translate toggle
group_mark_read_settings = {} # per-group mark-as-read toggle
group_retry_key_settings = {} # per-group retry key toggle
group_camera_qr_settings = {} # per-group camera QR button toggle
group_clipboard_qr_settings = {} # per-group clipboard QR button toggle
group_camera_roll_qr_settings = {} # per-group camera roll QR button toggle
group_location_qr_settings = {} # per-group location QR button toggle
group_welcome_settings = {}   # per-group welcome: {group_id: {"enabled": bool, "text_zh": str, "text_id": str}}
# Translation tone settings
TONE_PRESETS = {
    "casual": "Translate casually like real people talk at work. Use everyday slang and informal language.",
    "natural": "Translate like a native speaker would naturally say it in daily factory conversation. Use the most natural, fluent, mother-tongue phrasing. Prefer colloquial expressions over textbook ones (e.g. Indonesian: prefer 'belum' over 'tidak' for not-yet-done actions, prefer 'udah' over 'sudah').",
    "formal": "Translate in formal, polite, professional language suitable for official announcements or documents.",
    "factory": (
        "你是工廠現場中印翻譯助理，專門處理台灣工廠工作群組、現場指令、品質異常、站別流轉、工單、TAG、PMI、包裝、混料、入站、資料輸入等訊息。"
        "你的任務不是逐字翻譯，而是把中文準確翻成「印尼工廠現場員工一看就懂、實際會用」的自然印尼文。請嚴格遵守以下規則："
        "【最高優先規則】"
        "1. 只要範例表/內建範例已有對應詞、對應句、固定說法，必須優先套用，不可改寫，不可換同義詞，不可自創新翻法。"
        "2. 若輸入句子中包含範例表已有的片語或句型，優先沿用範例表的翻法，再補齊其他部分。"
        "3. 若範例表已有專有詞定義，例如 PMI、工單、TAG、混料、站別等專有詞，必須依範例表理解，不可用一般字典義亂翻；像「入完了」「不擋」「無主」「放料」這類工廠動作慣用語，也必須依範例表的工廠脈絡解讀，不可從字面直翻。"
        "【翻譯風格規則】"
        "4. 一律使用印尼工廠現場自然口吻，簡短、直接、清楚，像主管或同事在工作群組講話。"
        "5. 不要逐字直譯，不要翻成書面公文，不要翻成教科書語氣，不要華麗修飾。"
        "6. 句子要以「現場人員看得懂、能立刻執行」為優先，不以文法漂亮為優先。"
        "7. 若原文是提醒、警告、要求、異常通報，語氣要明確、有執行感，但不要粗暴失禮。"
        "8. 若原文很口語、省略主詞或量詞混亂，必須依工廠作業脈絡補足正確意思後再翻，不可照字面硬翻。"
        "【工廠語境理解規則】"
        "9. 在本系統中，很多中文詞語不是日常字面義，必須依工廠語境理解，例如："
        "- 「入了 / 入完了」要依上下文判斷：預設是「入庫」(masuk gudang)；如果上下文在講站別、製程、登錄，才理解為「入站」(masuk stasiun)或「資料登錄完成」(data sudah dimasukkan)。不要直接用無語境的 sudah masuk。"
        "- 「把」作為量詞時(數字+把，如「6把」「兩把」) = bundel(棒材捆數)；作為介詞時(把+受詞，如「把工單入完」「把資料登錄」) = 中文文法輔助詞，不要翻成 bundel，而是依完整句子重組成印尼文語序。"
        "- 「PMI」不是泛稱，必須依範例表定義理解(分光檢測棒材鋼種)"
        "- 「混料」固定理解為 material tercampur"
        "- 「工單」固定用 work order"
        "- 「TAG」固定保留 TAG"
        "- 「站別」固定理解為 stasiun"
        "10. 遇到工廠專有語、現場省略句、短句、代號、站號、料號、ID、數字、批號時，優先保留原資訊完整，不可漏掉站號、數量、ID、重量、長度、尺寸、編號。"
        "【輸出規則】"
        "11. 只輸出最終譯文，不要解釋，不要加註解，不要說明原因，不要列出其他可能翻法。"
        "12. 不要擅自補充原文沒有的資訊。"
        "13. 不要省略數字、站號、代碼、重量、尺寸、長度、批號。"
        "14. 若原文是短句，就翻成短句；若原文是群組公告，就翻成可直接貼群組的公告語氣。"
        "15. 若原文已經是明顯命令句、提醒句、異常句，翻譯後必須保留同等強度，不可弱化。"
        "【品質優先規則】"
        "16. 翻譯優先順序為：範例表固定翻法 > 工廠現場正確語意 > 印尼員工可懂度 > 文法自然 > 字面對應"
        "17. 寧可翻得直白清楚，也不要翻得漂亮卻不符合現場。"
        "18. 若一句中文有歧義，請優先依「工廠製造、站別流轉、工單、包裝、品質異常、鋼材/棒材、分光檢測」脈絡判斷最合理意思後再翻。"
        "19. 若輸入是繁中口語群組訊息，請預設場景為工廠工作群組，不要用日常聊天語境解讀。"
        "20. 全程維持同一套翻譯標準，不因句長變動口吻，不因句子簡短就隨便翻。"
        "21. CRITICAL: 翻譯前先判斷說話者的情感方向(道歉/感謝/請求/警告/抱怨/承諾/通報)，確認後再翻；不要被 emoji(尤其 🙏)誤導，不要看到第一個合理選項就停。"
        "22. THINK-BEFORE-TRANSLATE (Chain-of-thought, internal): Before producing the final translation, internally consider these checks (don't output them, only output final translation): "
        "(a) Who is the speaker? (manager/worker/admin) "
        "(b) What is the intent? (apology/request/warning/announcement/complaint/promise/report) "
        "(c) Are there any factory-specific terms requiring exact mapping? (PMI/工單/混料/站別/不擋/入完了/台車...) "
        "(d) Are there ellipsed subjects/objects to fill in from context? "
        "(e) Could the translation be misread as the OPPOSITE meaning by an Indonesian worker? If yes, reword. "
        "(f) Is there an emoji that might mislead? (🙏 might suggest 'tolong' but with 'maaf' it means apology, not request) "
        "After these internal checks, output ONLY the final translation, no explanation."
    ),
}
translation_tone = "factory"      # global default: factory / casual / natural / formal
translation_tone_custom = ""      # global custom tone text (overrides preset if non-empty)
group_tone_settings = {}          # per-group: {gid: {"tone": str, "custom": str}}

# Model auto-switch: use gpt-4.1 for long messages, gpt-4.1-mini for short
# v3.2-0426d: upgraded from gpt-4o family to gpt-4.1 family.
# gpt-4.1 ranks #1 for translation in Intento State of Translation Automation 2025.
model_default = "gpt-4.1-mini"    # model for short messages
model_upgrade = "gpt-4.1"         # model for long messages
model_threshold = 0               # char count threshold (0 = always use default, no auto-switch)

# v3.9.7 (2026-05): 模型能力對照表 — 讓使用者亂勾任何進階設定都不會 400。
# 後端永遠根據實際模型過濾參數;UI 端會把不相容的設定自動還原成中性值。
def _model_family(model_name):
    """v3.9.8: GPT-5 series splits into two sub-families based on lowest reasoning effort:
      - 'gpt5_minimal': gpt-5, gpt-5.4, gpt-5.5 — supports reasoning_effort='minimal'
      - 'gpt5_none':    gpt-5.1, gpt-5.2 — supports reasoning_effort='none'
    """
    if not model_name:
        return "unknown"
    m = model_name.lower()
    # GPT-5.1 / 5.2 family — supports 'none' as lowest reasoning effort
    if m.startswith(("gpt-5.2", "gpt-5.1")):
        return "gpt5_none"
    # GPT-5 / 5.4 / 5.5 family — supports 'minimal' as lowest
    if m.startswith(("gpt-5.5", "gpt-5.4", "gpt-5")):
        return "gpt5_minimal"
    if m.startswith(("o1", "o3", "o4")):
        return "oseries_reasoning"
    if m.startswith(("gpt-4.1", "gpt-4o", "gpt-4")):
        return "gpt4_classic"
    return "unknown"

# v3.9.8: optimal reasoning_effort for translation per family.
# Based on OpenAI official migration guide:
#   "gpt-4.1: gpt-5.2 with `none` reasoning"
#   "minimal performs especially well in coding and instruction following scenarios"
# Translation = lightweight task that needs faithful instruction following, NOT reasoning.
# Confirmed by arxiv 2505.14810: "improving reasoning capability often comes at the
# cost of instruction adherence" — exactly what we don't want for translation.
TRANSLATION_OPTIMAL_REASONING = {
    "gpt5_minimal": "minimal",
    "gpt5_none": "none",
    "oseries_reasoning": "low",  # o-series doesn't have minimal/none
    "gpt4_classic": None,
    "unknown": None,
}

MODEL_CAPABILITIES = {
    # GPT-5 / 5.4 / 5.5 — reasoning models with minimal/low/medium/high effort
    "gpt5_minimal": {
        "temperature": False, "top_p": False, "max_tokens": False,
        "max_completion_tokens": True, "logprobs": False, "logit_bias": False,
        "stop": False, "structured_output": True, "metadata": True,
        "prompt_cache_key": True, "reasoning_effort": True, "seed": False,
        "verbosity": True,
    },
    # GPT-5.1 / 5.2 — newer reasoning models with none/low/medium/high
    "gpt5_none": {
        "temperature": False, "top_p": False, "max_tokens": False,
        "max_completion_tokens": True, "logprobs": False, "logit_bias": False,
        "stop": False, "structured_output": True, "metadata": True,
        "prompt_cache_key": True, "reasoning_effort": True, "seed": False,
        "verbosity": True,
    },
    "oseries_reasoning": {
        "temperature": False, "top_p": False, "max_tokens": False,
        "max_completion_tokens": True, "logprobs": False, "logit_bias": False,
        "stop": False, "structured_output": False, "metadata": True,
        "prompt_cache_key": False, "reasoning_effort": True, "seed": False,
        "verbosity": False,
    },
    "gpt4_classic": {
        "temperature": True, "top_p": True, "max_tokens": True,
        "max_completion_tokens": False, "logprobs": True, "logit_bias": True,
        "stop": True, "structured_output": True, "metadata": True,
        "prompt_cache_key": True, "reasoning_effort": False, "seed": True,
        "verbosity": False,
    },
    "unknown": {
        "temperature": True, "top_p": True, "max_tokens": True,
        "max_completion_tokens": False, "logprobs": True, "logit_bias": True,
        "stop": True, "structured_output": True, "metadata": True,
        "prompt_cache_key": True, "reasoning_effort": False, "seed": True,
        "verbosity": False,
    },
}

def model_supports(model_name, capability):
    family = _model_family(model_name)
    return MODEL_CAPABILITIES.get(family, MODEL_CAPABILITIES["unknown"]).get(capability, False)

def optimal_reasoning_for_translation(model_name):
    """v3.9.8: Get optimal reasoning_effort for translation tasks.
    Based on OpenAI official guidance: translation = lightweight,
    instruction-following task. Higher reasoning_effort HURTS instruction
    adherence and adds latency/cost without quality gain on this task type.
    """
    family = _model_family(model_name)
    return TRANSLATION_OPTIMAL_REASONING.get(family)

# v3.2-0426d: New translation parameters (admin-controllable)
translation_temperature = 0.0     # 0.0 = deterministic, 0.3 = slight variety. Translation should be 0~0.3.
translation_top_p = 1.0           # Nucleus sampling. Keep 1.0 unless you know what you're doing.
translation_seed = 0              # 0 = random; non-zero = (mostly) reproducible outputs for same input.
# v3.2-0426d Batch B: Round-trip verification (double-check)
# Mode: "off" / "smart" (long msgs + apology keywords) / "all" / "keywords_only"
double_check_mode = "smart"
double_check_threshold = 0.55     # Similarity threshold (0~1). Below = warning. Lower = more permissive.
double_check_keywords = "maaf,maafkan,sori,ampun,salah,gak akan,bukan saya,jangan salah,對不起,抱歉,請原諒,誤會,搞錯,放錯,弄錯,搞混"
# v3.2-0426d Batch B: Few-shot examples format mode
# "system_prompt" (legacy) | "messages" (OpenAI standard, better quality)
fewshot_mode = "messages"
# v3.2-0426d Batch C: Logprobs + Structured Outputs
logprobs_enabled = True             # When True, request logprobs to compute confidence
confidence_threshold = 0.85         # Below this = prepend ⚠️ to translation
structured_output_enabled = False   # When True, force JSON {translation, confidence, alternatives}
prompt_caching_enabled = True       # Use prefix-stable system prompts for 75% discount
# v3.8: prompt_cache_key sticky-routing.
# OpenAI hashes the first ~256 prompt tokens to pick a backend; if many
# requests share the same prefix but spread across machines, cache hit
# rate drops. Passing a stable prompt_cache_key per logical "stream"
# (group + direction) raises sticky-routing chance, taking cache hit
# rate from ~60% → ~87% in OpenAI's published cookbook benchmark.
# Kept independently toggleable; default ON because it's free upside.
prompt_cache_key_enabled = True


def _build_cache_key(group_id="", src="", tgt="", kind="trans"):
    """v3.8: Stable per-stream cache routing key.
    Same group + same direction + same kind always yields the same key,
    so OpenAI routes those requests to the same backend with warm KV cache.
    Falls back to a generic key if group_id is missing.
    """
    if not prompt_cache_key_enabled:
        return None
    gid = (group_id or "default")[-20:]  # last 20 chars enough for uniqueness
    return f"{kind}:{src}-{tgt}:{gid}"


def _pick_aux_model(purpose="utility"):
    """v3.9.9: Pick a fast, cheap model in the same family as the user's main model.
    Used for auxiliary tasks (back-translation check, ID normalization, OCR, pivot).
    Auto-adapts so if user upgrades to GPT-5, aux calls also use GPT-5 (no API mismatch).
    
    v3.9.30 修正: 之前 startswith("gpt-5") 對 gpt-5.4-mini / gpt-5.5 等都回 gpt-5-nano,
    跨代家族 — gpt-5.4 主模型配 gpt-5 aux 雖然能呼叫,但不是同代最佳搭配。
    現在精準對應同代 nano。
    """
    md = (model_default or "").lower()
    # GPT-5 系列(精細到子家族)
    if md.startswith("gpt-5.5"):
        return "gpt-5.5-nano" if md != "gpt-5.5-nano" else "gpt-5-nano"
    if md.startswith("gpt-5.4"):
        return "gpt-5.4-nano"  # 同代最便宜
    if md.startswith("gpt-5.2"):
        return "gpt-5.2-nano" if md != "gpt-5.2-nano" else "gpt-5-nano"
    if md.startswith("gpt-5.1"):
        return "gpt-5.1-nano" if md != "gpt-5.1-nano" else "gpt-5-nano"
    if md.startswith("gpt-5"):
        return "gpt-5-nano"   # 通用 GPT-5 nano
    if md.startswith("gpt-4.1"):
        return "gpt-4.1-nano"
    return "gpt-4o-mini"


def _build_aux_kwargs(model_name, messages, max_out_tokens=500, temperature=0.0, cache_key=None):
    """v3.9.9: Build OpenAI kwargs for an auxiliary call (NOT the main translation).
    
    Auto-filters parameters per model capability so the same call works on:
      - GPT-4 family (uses temperature, max_tokens)
      - GPT-5 family (uses max_completion_tokens, reasoning_effort=minimal, verbosity=low)
      - o-series   (uses max_completion_tokens, reasoning_effort=low)
    
    Use this for: back-translation checks, normalization, pivot intermediate, etc.
    Do NOT use for the MAIN translate_openai call (that has its own richer logic).
    
    v3.9.30b B11 修補: reasoning model (GPT-5 系列) max_completion_tokens 包含
    reasoning tokens,即使 effort=minimal 也會用一些。對 aux 任務(回譯/normalize)
    額外加 2K reasoning 緩衝,避免短輸出被 reasoning 吃光。
    """
    kwargs = {"model": model_name, "messages": messages, "timeout": 30}
    if model_supports(model_name, "temperature"):
        kwargs["temperature"] = temperature
    if model_supports(model_name, "max_completion_tokens"):
        # v3.9.30b: reasoning model 額外給 2K reasoning 預算(aux 任務用 minimal effort 夠了)
        _is_reasoning = not model_supports(model_name, "temperature")
        _budget = max_out_tokens + (2000 if _is_reasoning else 0)
        kwargs["max_completion_tokens"] = _budget
    elif model_supports(model_name, "max_tokens"):
        kwargs["max_tokens"] = max_out_tokens
    _opt = optimal_reasoning_for_translation(model_name)
    if model_supports(model_name, "reasoning_effort") and _opt:
        kwargs["reasoning_effort"] = _opt
    if model_supports(model_name, "verbosity"):
        kwargs["verbosity"] = "low"
    if cache_key and model_supports(model_name, "prompt_cache_key"):
        kwargs["prompt_cache_key"] = cache_key
    return kwargs


# v3.2-0426d Batch D: Translation logging + monitoring
translation_logging_enabled = True   # Record every translation
translation_log = []                 # In-memory ring buffer + persisted JSON file
TRANSLATION_LOG_MAX = 500            # Cap at 500 to prevent memory blowup

# v3.9.5: Debug snapshot for the last translation. Used by /admin/debug/last-translate
# to diagnose why custom examples may not be reaching the model. Stores the
# actual messages array sent to OpenAI, the path taken, model, and raw response.
last_translate_debug = {}

# v3.8: Reaction-event feedback pipeline.
# Maps: bot's sent message_id → {"entry_id": ..., "ts": ..., "group_id": ...}
# When a user reacts to a translation, look up which entry it was and apply
# the feedback (positive → add to examples; negative → mark wrong).
# Bounded LRU-style: capped at 2000 entries, oldest evicted.
import collections as _collections
sent_message_to_entry = _collections.OrderedDict()
SENT_MSG_MAP_MAX = 2000

# Reaction → action mapping. Reaction names follow LINE's API enum.
# Positive reactions teach the bot "this translation was good";
# Negative ones flag the entry as wrong (operators can later add a correction).
REACTION_POSITIVE = {"love", "like", "happy"}
REACTION_NEGATIVE = {"sad", "surprise"}  # 😢 sad, 😮 surprise (user-confused)
# `all` (👍 thumbs-down equivalent depending on platform) is left out
# because LINE's reaction set is small and ambiguous; tune as needed.

def _resolve_translation_log_path():
    """Pick a persistence path that survives Render deploys.

    Priority:
      1. TRANSLATION_LOG_FILE env var (explicit override)
      2. /var/data/translation_log.json   (Render persistent disk default mount)
      3. /data/translation_log.json       (common alt mount)
      4. ./translation_log.json           (last resort; lost on each deploy)
    """
    env = os.environ.get("TRANSLATION_LOG_FILE", "").strip()
    if env:
        return env
    for candidate_dir in ("/var/data", "/data"):
        if os.path.isdir(candidate_dir) and os.access(candidate_dir, os.W_OK):
            return os.path.join(candidate_dir, "translation_log.json")
    return "translation_log.json"

TRANSLATION_LOG_FILE = _resolve_translation_log_path()
logger.info("[TLOG] persistence path = %s", TRANSLATION_LOG_FILE)

def _load_translation_log_from_disk():
    """Load persisted translation log on startup. Best effort; never blocks bot startup."""
    try:
        path = TRANSLATION_LOG_FILE
        if not path or not os.path.exists(path):
            return
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            translation_log[:] = data[-TRANSLATION_LOG_MAX:]
            logger.info("Loaded %d translation log entries from %s", len(translation_log), path)
    except Exception as e:
        logger.warning("Failed to load translation log: %s", e)

def _save_translation_log_to_disk():
    """Persist translation log to JSON. Best effort; avoids losing logs on normal restarts."""
    try:
        path = TRANSLATION_LOG_FILE
        if not path:
            return
        folder = os.path.dirname(os.path.abspath(path))
        if folder and not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(translation_log[-TRANSLATION_LOG_MAX:], f, ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except Exception as e:
        logger.warning("Failed to save translation log: %s", e)

ab_test_enabled = False              # A/B prompt testing (legacy, deprecated)
ab_test_variant_b_prompt = ""        # Custom variant B prompt (legacy, deprecated)

# ★ v3.4 ID→ZH 翻譯品質強化(對症下藥:印尼文→中文常翻錯)
# 三項學術論文驗證有效的技巧,可獨立開關
id_zh_cot_enabled = True             # Chain-of-Thought 二階段翻譯(ID→ZH 專用)
id_zh_cod_enabled = True             # Chain-of-Dictionary 動態術語注入
id_zh_pivot_enabled = False          # Pivot via English(成本較高,預設關閉)
id_zh_pivot_threshold = 80           # 訊息長度超過此值才啟用 pivot(避免短句浪費)
id_zh_double_translation = False     # 雙翻 ensemble(翻兩次選較完整的)成本高

# ★ v3.6 印尼文預處理 + 多路徑反譯 + 監控儀表板
id_preprocessing_enabled = True      # 印尼俚語/簡寫標準化(純詞表,0 API 呼叫)
id_preprocessing_nano = False        # 額外用 nano 模型做語意級規範化(會多一次 API)
multi_path_backtrans_enabled = False # 多路徑反譯(zh→id 平行 zh→en→id),成本翻倍
multi_path_min_chars = 60            # 多路徑反譯的觸發字數
quality_metrics_enabled = True       # 收集品質指標(catastrophic_compression / missing_terms 等)

# ★ v3.7 段落結構保留(對症下藥:長公告翻譯後不再連成一坨)
preserve_paragraphs_enabled = True   # 在 prompt 中要求保留段落
paragraph_split_translate = True     # 分段翻譯模式(雙保險)
paragraph_split_threshold = 50       # 訊息長度超過此值且含分段時,走分段翻譯路徑

# v3.2-0426e: New official OpenAI features
stop_sequences_enabled = True        # Use stop sequences to prevent GPT adding explanations
forbidden_words_zh = "註：,(註,(備註,以下是,翻譯如下,Translation:"  # zh forbidden phrases (comma-separated)
forbidden_words_id = "Catatan:,(Catatan,Terjemahan:,Penjelasan:"  # id forbidden phrases
reasoning_effort = "low"             # v3.9.8: was 'medium'. For translation tasks, lower=more faithful. Auto-overridden by optimal_reasoning_for_translation().
send_user_id_to_openai = True        # For abuse tracking compliance
send_metadata_to_openai = True       # Tag each request with group_id for filtering

import threading as _threading
_tl = _threading.local()          # thread-local for passing tone / audit metadata into translate_openai
_load_translation_log_from_disk()

def get_group_tone(group_id):
    """Return (preset, custom_text) for a group."""
    if group_id and group_id in group_tone_settings:
        gs = group_tone_settings[group_id]
        return gs.get("tone", translation_tone), gs.get("custom", "")
    return translation_tone, translation_tone_custom


def pick_model(text):
    """Pick OpenAI model based on text length and threshold setting."""
    if model_threshold > 0 and len(text) >= model_threshold:
        return model_upgrade
    return model_default


def _translation_quality_check(original, translation, back_translation):
    """v3.3 多維度翻譯品質檢查,取代純 Jaccard 相似度。
    
    回傳 (passes, reason, details_dict)
    
    檢查項目:
      1. 譯文長度比 - 譯文/原文,< 0.33 視為災難性壓縮
      2. 反譯長度比 - 反譯/原文,< 0.4 視為譯文丟失資訊
      3. Jaccard 字元相似度
    """
    orig_len = len(re.sub(r"\s+", "", original or ""))
    trans_len = len(re.sub(r"\s+", "", translation or ""))
    back_len = len(re.sub(r"\s+", "", back_translation or ""))
    
    if orig_len == 0:
        return True, "empty_source", {}
    
    details = {"orig_len": orig_len, "trans_len": trans_len, "back_len": back_len}
    
    # 檢查 1:譯文長度比過低 → 必定丟資訊
    trans_ratio = trans_len / orig_len
    details["trans_ratio"] = round(trans_ratio, 3)
    if trans_ratio < 0.33:
        return False, f"catastrophic_compression(ratio={trans_ratio:.2f})", details
    
    # 檢查 2:反譯長度比 - 譯文資訊已丟失
    if back_len > 0:
        back_ratio = back_len / orig_len
        details["back_ratio"] = round(back_ratio, 3)
        if back_ratio < 0.4:
            return False, f"backtrans_too_short(ratio={back_ratio:.2f})", details
    
    # 檢查 3:Jaccard 字元相似度
    if back_translation:
        sim = _text_similarity(original, back_translation)
        details["jaccard"] = round(sim, 3)
        if sim < double_check_threshold:
            return False, f"low_similarity(jaccard={sim:.2f})", details
    
    return True, "ok", details


def _round_trip_check(original_text, translated_text, src_lang, tgt_lang):
    """v3.3 反譯檢查 - 加入多維度品質檢查
    回傳 (passes_check, similarity, back_translation)
    
    Returns:
      passes_check: bool - 是否通過所有檢查
      similarity: float - Jaccard 相似度(沒有反譯時為 1.0)
      back_translation: str - 反譯結果
    """
    if not oai or not translated_text:
        return True, 1.0, ""
    try:
        # v3.9.9: aux helper handles all model families automatically.
        check_model = _pick_aux_model("backcheck")
        if tgt_lang == "zh":
            back_prompt = "Translate this Chinese to Indonesian. Output ONLY the translation, no explanation:"
        elif tgt_lang == "id":
            back_prompt = "Translate this Indonesian to Chinese. Output ONLY the translation, no explanation:"
        else:
            return True, 1.0, ""
        r = oai.chat.completions.create(**_build_aux_kwargs(
            check_model,
            [{"role": "system", "content": back_prompt},
             {"role": "user", "content": translated_text}],
            max_out_tokens=500
        ))
        track_tokens(r)
        # v3.9.30 B7 修補: GPT-5 reasoning model 在 max_completion_tokens 不足時
        # 會回 content=None + finish_reason='length',直接 .strip() 會 AttributeError
        _content = r.choices[0].message.content if r.choices else None
        if _content is None:
            _finish = r.choices[0].finish_reason if r.choices else "unknown"
            logger.warning(
                "[back-trans] empty content from %s, finish_reason=%s — skip back check",
                check_model, _finish
            )
            # 反譯失敗就跳過反譯檢查(不擋住主翻譯流程)
            return True, 1.0, ""
        back = _content.strip()
        
        # ★ v3.3:多維度檢查
        ok, reason, details = _translation_quality_check(original_text, translated_text, back)
        sim = details.get("jaccard", 1.0)
        
        if not ok:
            logger.warning(
                "Round-trip FAILED: %s | orig=%s | trans=%s | back=%s | details=%s",
                reason, original_text[:60], translated_text[:60], back[:60], details
            )
            # 把失敗原因記到 thread-local,供日誌使用
            try:
                _tl.rtc_fail_reason = reason
            except Exception:
                pass
        
        return ok, sim, back
    except Exception as e:
        logger.error("Round-trip check error: %s", e)
        return True, 1.0, ""


# =====================================================================
# v3.4 ID→ZH 翻譯品質強化:三大技巧
# =====================================================================

# ★ 高風險誤譯詞典(印尼文→中文)
# 這些詞是常見的「直譯陷阱」,GPT 看到容易翻錯
# 編譯來源:歐那實際使用 + 工廠語境常見錯誤
ID_ZH_HIGH_RISK_TERMS = {
    # 工廠用語(非直譯)
    "barang": "料件(NOT 物品/東西)",
    "batang": "棒材(NOT 棍子/枝)",
    "bahan": "材料(在工廠語境)",
    "rusak": "損傷/異常(NOT 壞掉)",
    "bengkok": "彎曲(料件變形)",
    "cacat": "異常/不良品(NOT 殘廢)",
    "tercampur": "混料(NOT 混合飲料/食物)",
    "pemotongan": "切斷區/切斷工序",
    "penanda": "標記/識別(NOT 標誌牌)",
    "tali pemisah": "束帶/分隔帶",
    "lecet": "擦傷(料件表面)",
    "gores": "刮傷",
    "patah": "斷裂",
    
    # 工廠站別/動作
    "ambil": "領取/拿(具體看上下文)",
    "simpan": "存放/保管",
    "kasih tau": "通知/告知(NOT 給知識)",
    "bertugas": "值班/負責",
    "ketahuan": "讓人知道/被發現",
    "pekerja": "員工/作業員",
    "atasan": "主管/上司",
    
    # 易錯日常
    "kasih": "給(口語,NOT 愛/同情)",
    "udah": "已經(口語=sudah)",
    "gak": "不(口語=tidak)",
    "bisa": "可以/能",
    "harus": "必須/要",
    "jangan": "不要/別",
    "biar": "讓/以便(NOT 讓它去)",
    "karna": "因為(=karena)",
    "buang": "丟棄",
    "langsung": "直接/立刻",
    "terutama": "特別是/尤其",
    "perhatikan": "注意",
    "pemberitahuan": "公告/通知",
    "pengumuman": "公告",
    
    # 中印雙向高頻誤譯
    "dicuri": "被偷(慎用,工廠語境通常指「被先拿走」)",
    "mencuri": "偷(慎用)",
    "bereaksi": "反應(化學,NOT 回報)",
    "tertelan": "吞下去(NOT 被吃掉痕跡)",
}


def detect_id_zh_risk_terms(text):
    """偵測印尼文中的高風險誤譯詞,回傳「字典片段」字串供注入 prompt
    
    只對句中實際出現的詞建立字典,避免污染 prompt
    """
    if not text:
        return ""
    t = text.lower()
    found = []
    for word, hint in ID_ZH_HIGH_RISK_TERMS.items():
        # 用 word boundary 避免誤判子字串
        if re.search(r"(?<![a-z])" + re.escape(word) + r"(?![a-z])", t):
            found.append(f"  - {word} → {hint}")
    if not found:
        return ""
    return "【翻譯字典(必須遵守)】\n" + "\n".join(found) + "\n"


# =====================================================================
# v3.6 印尼文預處理 (Bahasa Gaul / 簡寫標準化)
# =====================================================================
# 來源:NusaMT-7B 論文 + 印尼語常見口語縮寫詞典
# 原理:LLM 對標準印尼文比口語熟,把 gak→tidak / udah→sudah 等先還原,
#       翻譯品質會明顯提升。整個過程是純詞表替換,0 API 呼叫,即時生效。

ID_NORMALIZATION_MAP = {
    # ===== 高頻簡寫(SMS/WhatsApp 風格)=====
    "gak": "tidak", "ga": "tidak", "nggak": "tidak", "ngga": "tidak", "g": "tidak",
    "gk": "tidak", "ndak": "tidak", "gada": "tidak ada", "gpp": "tidak apa-apa",
    "gapapa": "tidak apa-apa",
    "udah": "sudah", "udh": "sudah", "uda": "sudah", "dah": "sudah",
    "blm": "belum", "blum": "belum", "blom": "belum",
    "bgt": "banget", "bgtt": "banget",
    "yg": "yang", "y": "ya",
    "krn": "karena", "krna": "karena", "karna": "karena",
    "tdk": "tidak", "tdak": "tidak",
    "dg": "dengan", "dgn": "dengan",
    "hrs": "harus", "hrus": "harus",
    "bs": "bisa", "bsa": "bisa",
    "lg": "lagi",
    "org": "orang", "orng": "orang",
    "kalo": "kalau", "klo": "kalau", "klau": "kalau", "kl": "kalau",
    "tau": "tahu", "taw": "tahu",
    "jd": "jadi", "jdi": "jadi",
    "sm": "sama", "sma": "sama",
    "tlg": "tolong", "tlng": "tolong", "tolg": "tolong",
    "cepet": "cepat", "cpt": "cepat", "cpet": "cepat",
    "dpt": "dapat", "dpat": "dapat",
    "kpd": "kepada", "kpda": "kepada",
    "utk": "untuk", "untk": "untuk",
    "dr": "dari", "dri": "dari",
    "tp": "tapi", "tpi": "tapi",
    "skrg": "sekarang", "skg": "sekarang", "skrng": "sekarang",
    "trs": "terus", "trus": "terus",
    "nih": "ini",
    "tuh": "itu",
    "msh": "masih",
    "udeh": "sudah",  # v3.9.30: 移除原本與 line 938 衝突的 "udh":"sudah" 重複定義
    "knp": "kenapa", "knpa": "kenapa",
    "gmn": "bagaimana", "gimana": "bagaimana", "gmna": "bagaimana",
    "bnr": "benar", "bnar": "benar",
    "ntr": "nanti", "ntar": "nanti",
    "smua": "semua",  # v3.9.29: 移除原本 "sma":"semua" 跟 line 952 衝突
    "krj": "kerja", "krja": "kerja",
    "kerjaan": "pekerjaan",
    "msk": "masuk",
    "klr": "keluar",
    "dh": "sudah",
    
    # ===== 雅加達/爪哇方言 =====
    "gw": "saya", "gue": "saya", "gua": "saya", "aku": "saya",
    "lu": "kamu", "lo": "kamu", "loe": "kamu",
    "ane": "saya", "ente": "kamu",
    "doang": "saja", "aja": "saja",
    "bener": "benar",
    "ngapain": "sedang apa", "ngapa": "kenapa",
    "kok": "mengapa",
    "nyari": "mencari",
    "ngambil": "mengambil",
    "ngomong": "berbicara",
    "ngerti": "mengerti",
    "kasian": "kasihan",  # v3.9.30: 移除原本與 line 935 衝突的 "ngga":"tidak" 重複定義
    "duit": "uang",
    "sip": "baik",
    "ngecek": "memeriksa", "cek": "memeriksa",
    
    # ===== 工廠常見變體 =====
    "kerjaaan": "pekerjaan",  # 重複字
    "stp": "setiap",
    # v3.9.30: 移除原本與 line 953 衝突的 "tlg" / "tlng" 重複定義
    "msl": "misal", "msl-nya": "misalnya",
    "biar": "agar",  # 工廠語境,biar = agar/supaya 比 supaya 更常見
}

# 大寫專有名詞保護(機台代號、工廠縮寫不能被改)
ID_PRESERVE_TOKENS = re.compile(
    r"\b(?:[A-Z]{2,}\d*|BF\d?|CYA|CYB|QC|QA|PM|PMI|MI|SS|SUS\d+|SAE\d+|"
    r"PK\d+|S\d+|F\d+|"
    r"[A-Z]\d[A-Z0-9-]+|\d+[A-Z][A-Z0-9-]+)\b"
)


def normalize_indonesian_text(text):
    """將印尼俚語/簡寫還原為標準印尼文(純詞表替換,0 API 呼叫)
    
    保護機制:
      - 大寫專有名詞(BF2/CYA/QC 等)不會被改
      - 處理單字邊界,避免誤改子字串
      - 大小寫不敏感比對,但保留首字母大寫
    
    回傳 (normalized_text, replacements_count)
    """
    if not text:
        return text, 0
    
    if not id_preprocessing_enabled:
        return text, 0
    
    # 1. 找出需要保護的大寫專有名詞,用 placeholder 暫存
    protected = {}
    counter = [0]
    
    def _protect(m):
        token = m.group(0)
        ph = f"__PROT_{counter[0]}__"
        protected[ph] = token
        counter[0] += 1
        return ph
    
    text_protected = ID_PRESERVE_TOKENS.sub(_protect, text)
    
    # 2. 詞表替換 - 用 word boundary,大小寫不敏感
    replacements = 0
    
    # 按長度由長到短排序,避免短詞先匹配吃掉長詞
    sorted_map = sorted(ID_NORMALIZATION_MAP.items(), key=lambda x: -len(x[0]))
    
    for slang, standard in sorted_map:
        pattern = re.compile(r"(?<![a-zA-Z])" + re.escape(slang) + r"(?![a-zA-Z])", re.IGNORECASE)
        new_text, n = pattern.subn(standard, text_protected)
        if n > 0:
            text_protected = new_text
            replacements += n
    
    # 3. 還原保護的專有名詞
    for ph, token in protected.items():
        text_protected = text_protected.replace(ph, token)
    
    return text_protected, replacements


def normalize_indonesian_text_with_nano(text):
    """進階模式:額外用 nano 模型做語意級規範化
    
    只在 id_preprocessing_nano = True 時使用,會多一次 API 呼叫
    用途:處理詞表抓不到的特殊用法、混雜爪哇語等
    """
    if not id_preprocessing_nano or not oai or not text:
        return text
    try:
        prompt = (
            "Convert this Indonesian text to standard Bahasa Indonesia, "
            "expanding all abbreviations and slang. "
            "Do NOT translate to other languages. Do NOT change machine codes "
            "(like BF2, CYA, CYB) or proper nouns. "
            "Output ONLY the normalized Indonesian text, no explanation."
        )
        # v3.9.9: aux helper auto-adapts to any model family
        r = oai.chat.completions.create(**_build_aux_kwargs(
            _pick_aux_model("normalize"),
            [{"role": "system", "content": prompt},
             {"role": "user", "content": text}],
            max_out_tokens=600
        ))
        track_tokens(r)
        normalized = (r.choices[0].message.content or "").strip()
        return normalized if normalized else text
    except Exception as e:
        logger.warning("ID nano normalization failed: %s", e)
        return text


# =====================================================================
# v3.6 多路徑反譯 (Multi-Path Back-Translation)
# =====================================================================
# 來源:LLM-BT-Terms 論文,證實平行多路徑反譯能找出 90%+ 翻譯錯誤
# 原理:單一反譯路徑可能因模型偏誤而漏抓錯誤,
#       兩條獨立路徑(直譯 + 經英語)兩者都通過才算真的 OK

def _multi_path_back_translation(original, translation, src_lang, tgt_lang):
    """多路徑反譯檢查
    
    回傳 (passes, details_dict)
    
    details = {
      'path1_passes': bool,    # 直接反譯
      'path1_jaccard': float,
      'path2_passes': bool,    # 經英語反譯
      'path2_jaccard': float,
      'final_decision': str,   # 'all_pass' / 'partial' / 'all_fail'
    }
    """
    if not oai:
        return True, {"final_decision": "no_oai"}
    
    details = {}
    
    try:
        # v3.9.9: aux helper picks model in same family as user's main translator,
        # so multi-path also works on GPT-5 family without errors.
        check_model = _pick_aux_model("multipath")
        
        # ===== Path 1: 直接反譯 =====
        if tgt_lang == "zh":
            back_prompt1 = "Translate this Chinese to Indonesian. Output ONLY translation:"
        elif tgt_lang == "id":
            back_prompt1 = "Translate this Indonesian to Chinese. Output ONLY translation:"
        else:
            return True, {"final_decision": "unsupported"}
        
        r1 = oai.chat.completions.create(**_build_aux_kwargs(
            check_model,
            [{"role": "system", "content": back_prompt1},
             {"role": "user", "content": translation}],
            max_out_tokens=500
        ))
        track_tokens(r1)
        back1 = (r1.choices[0].message.content or "").strip()
        sim1 = _text_similarity(original, back1)
        details["path1_jaccard"] = round(sim1, 3)
        ok1, _, _ = _translation_quality_check(original, translation, back1)
        details["path1_passes"] = ok1
        
        # ===== Path 2: 經英語反譯 (translation → English → back to source lang) =====
        # 第一段:translation → English
        r2a = oai.chat.completions.create(**_build_aux_kwargs(
            check_model,
            [{"role": "system", "content": "Translate to English. Output ONLY translation:"},
             {"role": "user", "content": translation}],
            max_out_tokens=500
        ))
        track_tokens(r2a)
        english = (r2a.choices[0].message.content or "").strip()
        
        # 第二段:English → source language
        if tgt_lang == "zh":
            back2_prompt = "Translate this English to Indonesian. Output ONLY translation:"
        else:
            back2_prompt = "Translate this English to Chinese. Output ONLY translation:"
        
        r2b = oai.chat.completions.create(**_build_aux_kwargs(
            check_model,
            [{"role": "system", "content": back2_prompt},
             {"role": "user", "content": english}],
            max_out_tokens=500
        ))
        track_tokens(r2b)
        back2 = (r2b.choices[0].message.content or "").strip()
        sim2 = _text_similarity(original, back2)
        details["path2_jaccard"] = round(sim2, 3)
        ok2, _, _ = _translation_quality_check(original, translation, back2)
        details["path2_passes"] = ok2
        
        # ===== 最終判斷 =====
        if ok1 and ok2:
            details["final_decision"] = "all_pass"
            return True, details
        elif ok1 or ok2:
            details["final_decision"] = "partial"
            # 部分通過 - 嚴格模式下視為失敗,但只給警告
            return False, details
        else:
            details["final_decision"] = "all_fail"
            return False, details
    
    except Exception as e:
        logger.warning("Multi-path back-translation error: %s", e)
        details["final_decision"] = f"exception:{e}"
        return True, details  # 不阻擋


def build_id_zh_cot_instruction(text):
    """建構 CoT 二階段翻譯指令(只在 ID→ZH 用)
    
    原理:讓 GPT 在翻譯前先「展開思考」,降低直譯誤譯率
    來源:ACL 2025 論文 "Reasoning for Translation"
    """
    return (
        "【翻譯思考流程(內部,不要輸出)】"
        "你必須先在心中完成以下三步,再輸出最終翻譯:"
        "(1) 識別句子類型:這是公告/異常通報/閒聊/指令/問句? "
        "(2) 列出關鍵資訊點:對象/動作/地點/原因/時間/操作要求(每個都要保留) "
        "(3) 注意俚語、簡寫、混雜爪哇語的部分,還原成標準語意 "
        "完成思考後,直接輸出最終的繁體中文翻譯,不要解釋過程,不要列點。"
    )


def translate_id_zh_with_pivot(text, src, tgt):
    """v3.4 Pivot via English: ID → EN → ZH 三段式翻譯
    
    原理:GPT 對 ID-EN 和 EN-ZH 都很強(EN 是樞紐),
    比直接 ID→ZH 通常更準確,代價是多一次 API 呼叫
    
    來源:ICLR 2025 論文 - 中等資源語言英語樞紐證實有效
    
    回傳譯文(失敗時 None,呼叫端 fallback 到原本流程)
    """
    if not oai or src != "id" or tgt != "zh":
        return None
    try:
        # 第 1 段:ID → EN(用 helper,自動配合主模型家族)
        en_prompt = (
            "Translate this Indonesian to fluent, complete English. "
            "Preserve ALL information including names, numbers, machine codes, "
            "place markers (BF, BF2, CYA, CYB etc). Output ONLY the English "
            "translation, no notes."
        )
        # v3.9.9: aux helper auto-adapts to GPT-5 if user upgrades
        _pivot_aux = _pick_aux_model("pivot")
        r1 = oai.chat.completions.create(**_build_aux_kwargs(
            _pivot_aux,
            [{"role": "system", "content": en_prompt},
             {"role": "user", "content": text}],
            max_out_tokens=800
        ))
        track_tokens(r1)
        english = (r1.choices[0].message.content or "").strip()
        if not english:
            return None
        
        # 第 2 段:EN → ZH(用主模型,品質更好)— v3.9.9 用 helper
        zh_prompt = (
            "Translate this English to Traditional Chinese (繁體中文) "
            "as used in Taiwan factories. Preserve all factory terminology. "
            "Machine codes and English acronyms should stay in original. "
            "Output ONLY the Chinese translation."
        )
        r2 = oai.chat.completions.create(**_build_aux_kwargs(
            pick_model(text),
            [{"role": "system", "content": zh_prompt},
             {"role": "user", "content": english}],
            max_out_tokens=800
        ))
        track_tokens(r2)
        return (r2.choices[0].message.content or "").strip()
    except Exception as e:
        logger.warning("Pivot translation failed: %s", e)
        return None


def should_use_pivot(text, src, tgt):
    """判斷是否該用 pivot 模式"""
    if not id_zh_pivot_enabled:
        return False
    if src != "id" or tgt != "zh":
        return False
    char_len = len(re.sub(r"\s+", "", text or ""))
    return char_len >= id_zh_pivot_threshold


def select_better_translation(t1, t2, original):
    """雙翻 ensemble:在兩個翻譯中選資訊更完整的那個
    
    判斷標準:
      1. 較長者通常資訊較完整
      2. 但若一者明顯過長(> 1.8x 另一者),選較短者(避免囉嗦)
      3. 若一者含 ⚠️ 警告,優先選另一者
    """
    if not t1: return t2
    if not t2: return t1
    
    has_warn1 = t1.startswith("⚠️")
    has_warn2 = t2.startswith("⚠️")
    if has_warn1 and not has_warn2: return t2
    if has_warn2 and not has_warn1: return t1
    
    l1 = len(re.sub(r"\s+", "", t1))
    l2 = len(re.sub(r"\s+", "", t2))
    if l1 == 0: return t2
    if l2 == 0: return t1
    
    ratio = max(l1, l2) / min(l1, l2)
    if ratio > 1.8:
        # 太懸殊,選較短(較長者可能在亂掰)
        return t1 if l1 < l2 else t2
    # 否則選較長(資訊較完整)
    return t1 if l1 > l2 else t2


# v3.2-0426e: tiktoken integration for real logit_bias
_tiktoken_encoders = {}  # Cache encoder per model
_tiktoken_failed = False  # If tiktoken unavailable, give up gracefully


def _get_tiktoken_encoder(model):
    """Lazy-load tiktoken encoder for the given model. Returns None on failure."""
    global _tiktoken_failed
    if _tiktoken_failed:
        return None
    if model in _tiktoken_encoders:
        return _tiktoken_encoders[model]
    try:
        import tiktoken
        try:
            enc = tiktoken.encoding_for_model(model)
        except KeyError:
            # Unknown model - use o200k_base (gpt-4o, gpt-4.1, gpt-5 family)
            enc = tiktoken.get_encoding("o200k_base")
        _tiktoken_encoders[model] = enc
        return enc
    except Exception as e:
        logger.warning("tiktoken unavailable, logit_bias disabled: %s", e)
        _tiktoken_failed = True
        return None


def _build_logit_bias(tgt_lang, model):
    """v3.2-0426e: REAL logit_bias to FORBID specific phrases.
    Encodes forbidden phrases to token IDs using tiktoken,
    then sets their bias to -100 (effectively banned).
    Returns dict of {token_id: -100} or {} if tiktoken unavailable.

    Forbidden phrases for translation:
    - zh target: 防止「麻煩你了」「請您」等錯誤搭配,以及機翻常見痕跡
    - id target: 防止「Catatan:」「Penjelasan:」等多餘解釋
    """
    if not stop_sequences_enabled:
        # If user disabled stop sequences, also disable logit_bias as a paired protection
        return {}
    enc = _get_tiktoken_encoder(model)
    if enc is None:
        return {}
    # Get forbidden phrase list based on target language
    if tgt_lang == "zh":
        phrases_str = forbidden_words_zh or ""
    elif tgt_lang == "id":
        phrases_str = forbidden_words_id or ""
    else:
        return {}
    phrases = [p.strip() for p in phrases_str.split(",") if p.strip()]
    if not phrases:
        return {}
    bias = {}
    try:
        for phrase in phrases:
            # Encode phrase - get all token IDs that make up this phrase
            try:
                token_ids = enc.encode(phrase)
                for tid in token_ids:
                    # OpenAI logit_bias keys must be strings, values -100..100
                    bias[str(tid)] = -100
            except Exception:
                continue
        # OpenAI limits logit_bias to max 300 entries
        if len(bias) > 300:
            # Keep first 300 (most likely the most critical short phrases)
            keys = list(bias.keys())[:300]
            bias = {k: bias[k] for k in keys}
        return bias
    except Exception as e:
        logger.warning("logit_bias build error: %s", e)
        return {}


def _build_stop_sequences(tgt_lang):
    """v3.2-0426e: Stop sequences to terminate GPT before it adds explanations.
    These are common patterns GPT uses when adding extra commentary."""
    if not stop_sequences_enabled:
        return None
    if tgt_lang == "zh":
        return ["\n註:", "\n（註", "\n注:", "\nTranslation:", "\n翻譯:"]
    if tgt_lang == "id":
        return ["\nCatatan:", "\n(Catatan", "\nTerjemahan:", "\nPenjelasan:"]
    return None


def _example_relevance_score(example_text, query_text):
    """Score how relevant an example is to the query, by character n-gram overlap.
    Lightweight, no external deps. Higher score = more relevant.
    Used to pick the most useful few-shot examples when total examples grow large.
    """
    if not example_text or not query_text:
        return 0.0
    et = example_text.lower()
    qt = query_text.lower()
    # 1. Exact substring match: massive bonus
    if et in qt or qt in et:
        return 100.0 + min(len(et), len(qt))
    # 2. Character bigram overlap
    def bigrams(s):
        s = re.sub(r'\s+', '', s)
        return set(s[i:i+2] for i in range(len(s)-1)) if len(s) >= 2 else set()
    bg_e = bigrams(et)
    bg_q = bigrams(qt)
    if not bg_e or not bg_q:
        return 0.0
    overlap = len(bg_e & bg_q)
    union = len(bg_e | bg_q)
    return (overlap / union) * 10.0 if union else 0.0


def _build_messages_with_fewshot(sys_prompt, user_msg, src, tgt):
    """v3.2-0426e: Build messages array using OpenAI standard few-shot format.
    Inserts BUILTIN_EXAMPLES + custom_translation_examples as
    {role: "system", name: "example_user"/"example_assistant"} pairs.
    This is OpenAI's recommended way (better than embedding examples in system prompt).

    v3.4: When custom_examples grows large (we now allow up to 5000),
    pick the FEWSHOT_INJECT_MAX most relevant examples by character bigram overlap
    with the user message, instead of just taking the last N.
    """
    msgs = [{"role": "system", "content": sys_prompt}]
    all_examples = list(BUILTIN_EXAMPLES) + list(custom_translation_examples or [])
    # Pick examples matching translation direction (only show direction-relevant pairs)
    # If translating zh->id, show zh2id examples; reverse for id->zh
    direction_key = "zh2id" if (src == "zh" and tgt != "zh") else "id2zh" if (src != "zh" and tgt == "zh") else None
    relevant = [ex for ex in all_examples if ex.get("dir", "zh2id") == direction_key] if direction_key else []

    # Score each example by relevance to user message and pick top N.
    # Score against the source-language side of the example (what GPT will pattern-match).
    if len(relevant) > FEWSHOT_INJECT_MAX:
        scored = []
        for ex in relevant:
            src_side = ex.get("zh", "") if direction_key == "zh2id" else ex.get("id", "")
            score = _example_relevance_score(src_side, user_msg)
            scored.append((score, ex))
        # Sort: highest relevance first, then most recent (we approximate recency by list order)
        # Stable sort preserves original order for ties, so recent examples win on tie.
        scored.sort(key=lambda x: x[0], reverse=True)
        # Always include at least 2 highest-relevance + recent fallbacks if relevance is weak
        top = [ex for score, ex in scored[:FEWSHOT_INJECT_MAX] if score > 0]
        # If fewer than FEWSHOT_INJECT_MAX matched, fill with most recent
        if len(top) < FEWSHOT_INJECT_MAX:
            recent_fill = [ex for ex in relevant[-FEWSHOT_INJECT_MAX*2:] if ex not in top]
            top.extend(recent_fill[:FEWSHOT_INJECT_MAX - len(top)])
        chosen = top[:FEWSHOT_INJECT_MAX]
    else:
        chosen = relevant[-FEWSHOT_INJECT_MAX:]

    for ex in chosen:
        zh = ex.get("zh", "").strip()
        idn = ex.get("id", "").strip()
        if not zh or not idn:
            continue
        if direction_key == "zh2id":
            user_eg, assistant_eg = zh, idn
        else:
            user_eg, assistant_eg = idn, zh
        # v3.9.3 (2026-05): OpenAI 早期文件曾推薦用
        # {"role": "system", "name": "example_user/example_assistant"} 格式做 few-shot,
        # 但這個格式已被棄用,GPT-5 系列(reasoning model)幾乎完全忽略它,
        # GPT-4 系列也不一定可靠。改用 OpenAI 現在所有官方文件、Cookbook 都使用的標準格式:
        # 直接用 user/assistant 對話對。
        # 這是讓範例真正進入 prompt 並影響翻譯的關鍵修復。
        msgs.append({"role": "user", "content": user_eg})
        msgs.append({"role": "assistant", "content": assistant_eg})
    msgs.append({"role": "user", "content": user_msg})
    return msgs


def _calc_confidence_from_logprobs(logprobs_obj):
    """v3.3 計算翻譯整體信心度
    
    使用「平均 logprob 後再取 exp」(等同幾何平均的對數空間版本),
    比舊版的算術平均更穩定,長譯文不會因 token 數變多而虛高。
    
    Returns float 0~1 (1.0 = 完美信心, 0.5 = 不確定)
    
    注意:單個 token 的低信心常見且無害;真正該關注的是整體序列的
    幾何平均信心過低 → 表示模型整段都在猜。
    """
    try:
        if not logprobs_obj or not getattr(logprobs_obj, 'content', None):
            return 1.0
        import math
        logprobs = []
        for tok in logprobs_obj.content:
            if tok and getattr(tok, 'logprob', None) is not None:
                logprobs.append(tok.logprob)
        if not logprobs:
            return 1.0
        # 幾何平均 = exp(mean(logprob))。比算術平均更不易被單一高機率字蓋過低機率字。
        avg_logprob = sum(logprobs) / len(logprobs)
        return math.exp(avg_logprob)
    except Exception:
        return 1.0


def _log_translation(src_text, tgt_text, src_lang, tgt_lang, model, tokens, confidence, double_checked, similarity, group_id=""):
    """Append a translation event to the in-memory + persisted log.

    Also records factory semantic audit metadata so the admin panel can show
    auto-detected factory translation issues even when the final message was
    auto-corrected before being sent.
    """
    if not translation_logging_enabled:
        return
    try:
        audit = getattr(_tl, 'factory_audit', None)
        entry = {
            "id": f"{int(time.time())}-{uuid.uuid4().hex[:8]}",
            "ts": int(time.time()),
            "src": src_text[:500],
            "tgt": tgt_text[:500],
            "src_lang": src_lang,
            "tgt_lang": tgt_lang,
            "model": model,
            "tokens": tokens,
            "confidence": round(confidence, 3) if confidence is not None else None,
            "double_checked": bool(double_checked),
            "similarity": round(similarity, 3) if similarity is not None else None,
            "group_id": group_id or getattr(_tl, 'group_id', ''),
            "marked_wrong": False,
            "warned": bool(tgt_text.startswith("⚠️")),
        }
        if audit and isinstance(audit, dict):
            # Only attach current audit if it belongs to this source text.
            if not audit.get("src") or audit.get("src") == src_text:
                entry.update({
                    "warned": True,
                    "factory_warning": True,
                    "warning_type": audit.get("type", "factory_semantic_warning"),
                    "warning_reason": audit.get("reason", ""),
                    "raw_translation": audit.get("raw_translation", ""),
                    "auto_corrected": bool(audit.get("auto_corrected")),
                    "corrected_translation": audit.get("corrected_translation", tgt_text),
                    "factory_domain": audit.get("domain", []),
                })
                try:
                    delattr(_tl, 'factory_audit')
                except Exception:
                    pass
        # ★ v3.6 收集品質指標到 entry
        try:
            # 預處理紀錄
            _prep = getattr(_tl, 'id_preprocessing', None)
            if _prep and isinstance(_prep, dict):
                entry["preprocessing"] = {
                    "original": _prep.get("original", "")[:200],
                    "normalized": _prep.get("normalized", "")[:200],
                    "count": _prep.get("count", 0),
                }
                try:
                    delattr(_tl, 'id_preprocessing')
                except Exception:
                    pass
            
            # 多路徑反譯紀錄
            _mp = getattr(_tl, 'multi_path_fail', None)
            if _mp and isinstance(_mp, dict):
                entry["multi_path_fail"] = _mp
                try:
                    delattr(_tl, 'multi_path_fail')
                except Exception:
                    pass
            
            # Self-check 失敗紀錄
            _sc = getattr(_tl, 'struct_self_check', None)
            if _sc:
                entry["struct_self_check"] = str(_sc)
                try:
                    delattr(_tl, 'struct_self_check')
                except Exception:
                    pass
            
            # 反譯失敗類型(catastrophic_compression / backtrans_too_short / low_similarity)
            _rtc_reason = getattr(_tl, 'rtc_fail_reason', None)
            if _rtc_reason:
                entry["rtc_fail_reason"] = str(_rtc_reason)
                try:
                    delattr(_tl, 'rtc_fail_reason')
                except Exception:
                    pass
            
            # 譯文/原文長度比(永遠記錄)
            _orig_l = len(re.sub(r"\s+", "", src_text or ""))
            _trans_l = len(re.sub(r"\s+", "", tgt_text or ""))
            if _orig_l > 0:
                entry["length_ratio"] = round(_trans_l / _orig_l, 3)
        except Exception as _e:
            logger.error("[TLOG] metrics enrichment failed: %s", _e)
        
        translation_log.append(entry)
        # Ring buffer
        if len(translation_log) > TRANSLATION_LOG_MAX:
            del translation_log[:len(translation_log) - TRANSLATION_LOG_MAX]
        _save_translation_log_to_disk()
        # v3.8: expose the just-logged entry id for reply-message-id binding.
        try:
            _tl.last_entry_id = entry.get("id")
        except Exception:
            pass
        logger.info("[TLOG] saved entry id=%s src=%r tgt=%r total=%d",
                    entry.get("id"), (src_text or "")[:40], (tgt_text or "")[:40],
                    len(translation_log))
    except Exception as e:
        import traceback
        logger.error("[TLOG] Log translation FAILED: %s\n%s", e, traceback.format_exc())


def _text_similarity(a, b):
    """Compute Jaccard similarity on character bigrams. Range 0~1."""
    if not a or not b:
        return 0.0
    def bigrams(s):
        s = s.lower().strip()
        s = re.sub(r'\s+', '', s)
        if len(s) < 2:
            return {s}
        return {s[i:i+2] for i in range(len(s)-1)}
    A = bigrams(a)
    B = bigrams(b)
    if not A or not B:
        return 0.0
    inter = len(A & B)
    union = len(A | B)
    return inter / union if union else 0.0


def _should_double_check(text, src_lang):
    """Decide whether to perform round-trip verification based on mode."""
    if double_check_mode == "off":
        return False
    if double_check_mode == "all":
        return True
    if double_check_mode == "keywords_only":
        kws = [k.strip().lower() for k in double_check_keywords.split(",") if k.strip()]
        tl = text.lower()
        return any(k in tl for k in kws)
    if double_check_mode == "smart":
        # Long messages OR keyword present
        if len(text) > 30:
            return True
        kws = [k.strip().lower() for k in double_check_keywords.split(",") if k.strip()]
        tl = text.lower()
        return any(k in tl for k in kws)
    return False


def _build_custom_examples_prompt():
    """Build custom examples string. Includes BUILTIN_EXAMPLES + custom."""
    all_examples = list(BUILTIN_EXAMPLES) + list(custom_translation_examples or [])
    if not all_examples:
        return " "
    zh2id = []
    id2zh = []
    for ex in all_examples:
        zh = ex.get("zh", "").strip()
        idn = ex.get("id", "").strip()
        if not zh or not idn:
            continue
        d = ex.get("dir", "zh2id")
        if d == "id2zh":
            id2zh.append(idn + " → " + zh)
        else:
            zh2id.append(zh + " → " + idn)
    parts = []
    if zh2id:
        parts.append(" 【自訂中→印尼】" + " ".join(zh2id))
    if id2zh:
        parts.append(" 【自訂印尼→中文】" + " ".join(id2zh))
    return " ".join(parts) if parts else " "


def _check_custom_example_exact(text, src, tgt):
    """Check if text exactly matches a custom or builtin example."""
    all_examples_local = list(BUILTIN_EXAMPLES) + list(custom_translation_examples or [])
    if not all_examples_local:
        return None
    t = text.strip()
    tl = t.lower()
    # Common prefixes that don't change meaning
    _zh_prefixes = ["先", "幫", "麻煩", "請", "幫忙", "再", "趕快", "快"]
    _id_prefixes = ["tolong ", "coba ", "mohon ", "bisa ", "mau "]
    # Build variants for matching
    zh_variants = [tl]
    id_variants = [tl]
    for p in _zh_prefixes:
        if tl.startswith(p) and len(tl) > len(p):
            zh_variants.append(tl[len(p):].strip())
    for p in _id_prefixes:
        if tl.startswith(p) and len(tl) > len(p):
            id_variants.append(tl[len(p):].strip())
    for ex in all_examples_local:
        zh = ex.get("zh", "").strip()
        idn = ex.get("id", "").strip()
        if not zh or not idn:
            continue
        zl = zh.lower()
        il = idn.lower()
        if src == "zh":
            # Check if input matches zh side (with prefix stripping on both sides)
            zh_ex_variants = [zl]
            for p in _zh_prefixes:
                if zl.startswith(p) and len(zl) > len(p):
                    zh_ex_variants.append(zl[len(p):].strip())
            for v in zh_variants:
                if v in zh_ex_variants:
                    return idn
        elif src == "id":
            # Check if input matches id side
            id_ex_variants = [il]
            for p in _id_prefixes:
                if il.startswith(p) and len(il) > len(p):
                    id_ex_variants.append(il[len(p):].strip())
            for v in id_variants:
                if v in id_ex_variants:
                    return zh
    return None

# Custom sender name/icon for translation messages
sender_name = "翻譯小助手"
sender_icon = ""  # URL to icon image, empty = default
# User profile pictures cache: {user_id: url}
user_pictures = {}

# ── Password settings (editable from admin) ──
pw1_text = "班長工號密碼：(尚未設定)\nPassword shift leader: (not set)"
pw2_text = "儲運工號密碼：(尚未設定)\nPassword gudang: (not set)"

# ── Scrap color text ──
scrap_text = (
    "🎨 廢料鋼種顏色 / Warna Scrap\n"
    "==================\n"
    "U物料(廢料) / U Material:\n"
    "  303 → 白/Putih\n"
    "  304 → 黃/Kuning\n"
    "  316 → 桃/Pink\n"
    "  209 → 特藍/Biru Khusus\n"
    "  174 → 紫羅蘭/Ungu\n"
    "  400系列 → 紅/Merah\n"
    "\n"
    "委外代工 / Outsource:\n"
    "  303 → 白/Putih\n"
    "  304 → 黃/Kuning\n"
    "  316 → 桃/Pink\n"
    "  403 → 紅/Merah\n"
    "=================="
)

# ── Packaging code lookup ──
PACKAGING_LOOKUP = {}

# ── Custom translation examples (editable from admin panel) ──
# List of {"zh": "中文", "id": "Indonesian", "dir": "zh2id"|"id2zh"}
# v3.4: Raised cap from 200 to 5000.
#   - Few-shot injection still uses only top 8 most relevant per request (see _build_messages_with_fewshot).
#   - All examples are kept as the long-term training corpus for future fine-tuning.
#   - 5000 entries is well within OpenAI's context window even if all were injected (they aren't).
CUSTOM_EXAMPLES_MAX = 5000
# Few-shot only uses N most semantically relevant examples per request:
FEWSHOT_INJECT_MAX = 8
custom_translation_examples = []


# v3.9.29: 修補隱性 bug — 之前 reaction 升級為 example 後沒存磁碟,重啟丟失
def _resolve_examples_path():
    """同 translation_log 邏輯,Render persistent disk 優先。"""
    env = os.environ.get("CUSTOM_EXAMPLES_FILE", "").strip()
    if env:
        return env
    for candidate_dir in ("/var/data", "/data"):
        if os.path.isdir(candidate_dir) and os.access(candidate_dir, os.W_OK):
            return os.path.join(candidate_dir, "custom_examples.json")
    return "custom_examples.json"

CUSTOM_EXAMPLES_FILE = _resolve_examples_path()


def _load_examples_from_disk():
    """Load persisted custom examples on startup. Best effort."""
    try:
        path = CUSTOM_EXAMPLES_FILE
        if not path or not os.path.exists(path):
            return
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            custom_translation_examples[:] = data[-CUSTOM_EXAMPLES_MAX:]
            logger.info("Loaded %d custom examples from %s",
                        len(custom_translation_examples), path)
    except Exception as e:
        logger.warning("Failed to load custom examples: %s", e)


def _save_examples_to_disk():
    """Persist custom examples to JSON. Atomic write."""
    try:
        path = CUSTOM_EXAMPLES_FILE
        if not path:
            return
        folder = os.path.dirname(os.path.abspath(path))
        if folder and not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(custom_translation_examples[-CUSTOM_EXAMPLES_MAX:],
                      f, ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except Exception as e:
        logger.warning("Failed to save custom examples: %s", e)


# 啟動時 load 一次
_load_examples_from_disk()

# ── Built-in factory examples (hardcoded, NOT visible in admin panel) ──
BUILTIN_EXAMPLES = [
    # ID->ZH hard factory examples: these are also enforced by the semantic layer below.
    {"zh": "料件後端損傷", "id": "Barang rusak dari belakang", "dir": "id2zh"},
    {"zh": "棒材後端損傷", "id": "Batang rusak dari belakang", "dir": "id2zh"},
    {"zh": "料件後端有損傷", "id": "Barang bagian belakang rusak", "dir": "id2zh"},
    {"zh": "前端損傷", "id": "Bagian depan rusak", "dir": "id2zh"},
    {"zh": "後端損傷", "id": "Bagian belakang rusak", "dir": "id2zh"},
    {"zh": "再強調一次", "id": "Saya tegaskan lagi", "dir": "zh2id"},
    {"zh": "一定要確實執行", "id": "harus benar-benar dijalankan", "dir": "zh2id"},
    {"zh": "一定要標示清楚", "id": "harus diberi penandaan yang jelas", "dir": "zh2id"},
    {"zh": "PMI作業", "id": "pemeriksaan grade baja batang dengan spektrometer", "dir": "zh2id"},
    {"zh": "台車", "id": "troli angkut batang", "dir": "zh2id"},
    {"zh": "台車滿了", "id": "troli angkut batang sudah penuh", "dir": "zh2id"},
    {"zh": "麻煩一下", "id": "tolong bantu", "dir": "zh2id"},
    {"zh": "台車再幫忙一下", "id": "Troli angkut batang sudah penuh, tolong bantu turunkan batangnya lagi.", "dir": "zh2id"},
    {"zh": "削皮那邊還需要一台", "id": "Bagian peeling masih butuh satu troli lagi.", "dir": "zh2id"},
    {"zh": "太凸", "id": "terlalu panjang", "dir": "zh2id"},
    
    # ===== v3.5 新增:印尼文公告/長句的失敗案例修正 =====
    # 來源:歐那實際遇到的災難性壓縮 (Martin 訊息)
    # 原則:讓 GPT 看到「公告類訊息要完整翻譯,不能簡化成短語」
    {
        "id": "@All Pemberitahuan Bagi pekerja BF, terutama BF2, jika kalian mengambil bahan baru, dan ada tali pemisah, jangan langsung di buang talinya, karna itu penanda orang yang bekerja di pemotongan (cya, cyb) karna banyak ditemukan barang bengkok. biar ketahuan siapa yang bertugas di mesin CYA atau CYB.",
        "zh": "@All 公告:給 BF 區員工,特別是 BF2,如果你們領取新材料時有附束帶,不要直接把束帶丟掉,因為那是切斷區(CYA、CYB)作業者的識別標記,因為發現很多料件彎曲。為了能知道是誰在 CYA 或 CYB 機台值班。",
        "dir": "id2zh"
    },
    # ===== 印尼文常見錯譯修正 =====
    {
        "id": "Pemberitahuan untuk semua: besok lembur sampai jam 8 malam.",
        "zh": "公告:明天加班到晚上 8 點。",
        "dir": "id2zh"
    },
    {
        "id": "Tolong perhatikan, jangan ambil bahan dari area QC tanpa konfirmasi.",
        "zh": "請注意,沒有確認過不要從品保區拿材料。",
        "dir": "id2zh"
    },
    {
        "id": "Barang yang sudah di-grinding harus segera di-cek dimensinya.",
        "zh": "已研磨的料件要立即檢查尺寸。",
        "dir": "id2zh"
    },
    {
        "id": "Mesin CYA bermasalah lagi, tolong panggil maintenance.",
        "zh": "CYA 機台又有問題,請叫保養人員過來。",
        "dir": "id2zh"
    },
    {
        "id": "Bahan tercampur dengan grade lain, harus dipisahkan dulu.",
        "zh": "材料混到其他等級了,要先分開。",
        "dir": "id2zh"
    },
    {
        "id": "biar gak salah, tolong tanya dulu ke kepala sebelum proses.",
        "zh": "為了不要出錯,加工前請先問組長。",
        "dir": "id2zh"
    },
    {
        "id": "Kerjaan udah selesai semua, tinggal nunggu pengiriman.",
        "zh": "工作已經全部完成,只等出貨。",
        "dir": "id2zh"
    },
    # ===== ZH→ID 公告類補強(同樣道理)=====
    {
        "zh": "公告:今天 BF2 下午要保養,請大家提前完成手上的料件。",
        "id": "Pemberitahuan: Hari ini BF2 sore akan ada perawatan, mohon semua menyelesaikan barang yang sedang dikerjakan lebih awal.",
        "dir": "zh2id"
    },
    {
        "zh": "請大家注意:看到料件後端有損傷要馬上回報組長,不要自己處理。",
        "id": "Mohon perhatian semua: kalau lihat barang ada kerusakan di bagian belakang, harus segera lapor ke kepala, jangan tangani sendiri.",
        "dir": "zh2id"
    },
]

# ── LIFF Form System ──────────────────────────────────
# forms_data: {form_id: {id, title_zh, title_id, fields:[{id,type,label_zh,label_id,options,required}], created, status, target_groups:[]}}
# forms_submissions: {form_id: {user_id: {fields..., submitted_at, user_name, approved}}}
forms_data = {}
forms_submissions = {}
_forms_loaded = False

# USD to TWD rate (approximate)
USD_TO_TWD = 32.0


def get_group_feature(group_id, feature):
    """Get per-group feature setting with global fallback."""
    _map = {
        'flex': (group_flex_settings, 'flex_enabled'),
        'quick_reply': (group_qr_settings, 'quick_reply_enabled'),
        'silent': (group_silent_settings, 'silent_mode'),
        'video_ocr': (group_video_settings, 'video_ocr_enabled'),
        'location': (group_location_settings, 'location_translate_enabled'),
        'mark_read': (group_mark_read_settings, 'mark_read_enabled'),
        'retry_key': (group_retry_key_settings, 'retry_key_enabled'),
        'camera_qr': (group_camera_qr_settings, 'camera_qr_enabled'),
        'clipboard_qr': (group_clipboard_qr_settings, 'clipboard_qr_enabled'),
        'camera_roll_qr': (group_camera_roll_qr_settings, 'camera_roll_qr_enabled'),
        'location_qr': (group_location_qr_settings, 'location_qr_enabled'),
    }
    if feature not in _map:
        return True
    d, global_key = _map[feature]
    if group_id and group_id in d:
        return d[group_id]
    return globals().get(global_key, True)


def get_group_welcome(group_id):
    """Get per-group welcome settings with global fallback."""
    if group_id and group_id in group_welcome_settings:
        gw = group_welcome_settings[group_id]
        # Merge with global defaults for missing keys
        return {
            "enabled": gw.get("enabled", welcome_settings.get("enabled", True)),
            "text_zh": gw.get("text_zh", welcome_settings.get("text_zh", "")),
            "text_id": gw.get("text_id", welcome_settings.get("text_id", "")),
        }
    return welcome_settings

# Translation cache: key = (text, src, tgt), value = (result, timestamp)
translation_cache = {}
_cache_lock = _threading.Lock()
CACHE_MAX_SIZE = 500
CACHE_TTL = 3600  # 1 hour

# Message cache for quoted message context: {message_id: {"text": str, "ts": float}}
message_cache = {}
MESSAGE_CACHE_MAX = 200

LANG_FLAGS = {
    "zh": "\U0001f1f9\U0001f1fc",
    "id": "\U0001f1ee\U0001f1e9",
}

LANG_NAMES = {
    "zh": "Traditional Chinese",
    "id": "Indonesian",
}

LANG_NAMES_ZH = {
    "id": "\u5370\u5c3c\u6587",
}

# Valid target languages
VALID_TARGETS = ["id"]


def extract_mentions(text):
    """Extract @mentions from text. Skip @Indonesian_word (not real mentions)."""
    _id_skip = {
        'tolong','semua','untuk','yang','dan','ini','itu','ada','tidak','akan',
        'sudah','bisa','juga','saya','kami','kita','mereka','dia','apa','belum',
        'sedang','harus','boleh','mau','bukan','jangan','terima','kasih','baik',
        'bagus','benar','salah','kerja','pulang','pergi','karena','tapi','atau',
        'kalau','masih','lagi','nanti','sekarang','siap','izin','minta','cepat',
        'capek','sakit','gak','udah','gimana','dong','banget','kipas','mesin',
        'rusak','bocor','macet','stok','habis','ganti','pasang','gudang','masuk',
        'keluar','tutup','buka','material','selesai','beres','datang','besok',
        'kemarin','libur','lembur','cuti','proses','produksi','diperhatikan',
        'selalu','mohon','pakai','pake','cek','lihat','bilang','ambil','kirim',
        'tunggu','bantu','butuh','perlu','panggil','suruh','hati','awas',
        'bahaya','lantai','mesin','pompa','pipa','oli','besi','baja','batang',
    }
    mentions = []
    # English @mentions: grab @word + up to 2 more, trim Indonesian words from end
    for m in re.finditer(r'@([A-Za-z0-9][A-Za-z0-9_.-]*)(?:\s+([A-Za-z0-9_.-]+))?(?:\s+([A-Za-z0-9_.-]+))?', text):
        first = m.group(1)
        if first.lower() in _id_skip:
            continue
        parts = [first]
        for g in [m.group(2), m.group(3)]:
            if g and g.lower() not in _id_skip:
                parts.append(g)
            else:
                break
        mention = '@' + ' '.join(parts)
        if mention not in mentions:
            mentions.append(mention)
    # Chinese @mentions
    for m in re.findall(r'@[\u4e00-\u9fff\u3040-\u30ff]+(?:\s*[\uff08(][^\uff09)]*[\uff09)])?', text):
        m = m.rstrip()
        if m and len(m) > 1 and m not in mentions:
            mentions.append(m)
    # @All
    for m in re.findall(r'@[Aa][Ll][Ll]', text):
        if m not in mentions:
            mentions.append(m)
    return list(dict.fromkeys(mentions))


def extract_line_mentions(text, message):
    """Extract @mention strings using LINE's actual mention data (the blue text).
    Returns list of exact mention strings from the message."""
    mentions = []
    try:
        mention_data = getattr(message, 'mention', None)
        if mention_data and hasattr(mention_data, 'mentionees'):
            for m in mention_data.mentionees:
                idx = m.index
                length = m.length
                mention_text = text[idx:idx+length]
                if mention_text and mention_text not in mentions:
                    mentions.append(mention_text)
    except Exception:
        pass
    return mentions


def protect_mentions(text, line_mentions=None):
    # Use LINE's actual mention data if available (100% accurate)
    # Fall back to regex extraction if not
    if line_mentions:
        mentions = line_mentions
    else:
        mentions = extract_mentions(text)
    protected = text
    placeholders = {}
    for i, m in enumerate(mentions):
        ph = f"__MENTION_{i}__"
        if m in protected:
            placeholders[ph] = m
            protected = protected.replace(m, ph, 1)
    return protected, placeholders


def restore_mentions(text, placeholders):
    restored = text or ""
    for ph, original in placeholders.items():
        idx = ph.replace("__MENTION_", "").replace("__", "")
        variants = [
            ph,
            ph.replace("_", " "),
            ph.replace("__", ""),
            f"MENTION_{idx}",
            f"MENTION {idx}",
            f"__MENTION {idx}__",
            f"[[MENTION_{idx}]]",
            # v3.9.30c B21 修補: GPT 偶爾會把 __MENTION_ 字面翻譯成中文「提及」
            # 歐那實際案例:_提及_0__ 殘留在輸出
            f"__提及_{idx}__",
            f"提及_{idx}",
            f"_提及_{idx}_",
            f"提及{idx}",
            # 印尼語可能的字面翻譯
            f"__SEBUTAN_{idx}__",
            f"SEBUTAN_{idx}",
            f"__sebutan_{idx}__",
        ]
        for v in variants:
            restored = restored.replace(v, original)

    # v3.9.30d B21 升級:萬一 GPT 用了我們沒列的變體,用正則兜底清掉所有 __提及_X__ / __MENTION_X__ 殘留
    # 這個只在還有 placeholders 時跑,避免誤殺正常文字
    if placeholders:
        # 把所有 placeholder index 收集起來,任何含這些 index 的「提及」字樣都還原
        indices = [ph.replace("__MENTION_", "").replace("__", "") for ph in placeholders.keys()]
        # 取第一個 placeholder 對應的 mention 字串作為兜底還原值(通常只有一個 @All / @user)
        fallback_mention = next(iter(placeholders.values())) if placeholders else ""
        # 用正則清殘留
        # 模式 1: __提及_N__ / _提及_N_ / 提及N / 提及_N
        def _replace_residual(m):
            return fallback_mention
        restored = re.sub(r'_{0,2}提及[_\s]?\d+_{0,2}', _replace_residual, restored)
        # 模式 2: __MENTION_N__ 殘留(英文未還原,通常已被上面 variants 處理,但保險)
        restored = re.sub(r'_{0,2}MENTION[_\s]?\d+_{0,2}', _replace_residual, restored)
        # 模式 3: 印尼文 sebutan 殘留
        restored = re.sub(r'_{0,2}[Ss][Ee][Bb][Uu][Tt][Aa][Nn][_\s]?\d+_{0,2}', _replace_residual, restored)

    # Final safety net: if any original @mention disappeared during translation,
    # prepend it back so the tagged person is not lost.
    missing = [original for original in placeholders.values() if original not in restored]
    if missing:
        prefix = " ".join(missing)
        restored = (prefix + " " + restored).strip()
    return restored


def strip_mentions_for_detect(text, line_mentions=None):
    """Strip @mentions for language detection."""
    if line_mentions:
        # Use LINE's actual mention data - most accurate
        clean = text
        for m in line_mentions:
            clean = clean.replace(m, ' ')
        return clean
    _id_skip = {
        'tolong','semua','untuk','yang','dan','ini','itu','ada','tidak','akan',
        'sudah','bisa','juga','saya','kami','kita','mereka','dia','apa','belum',
        'sedang','harus','boleh','mau','bukan','jangan','terima','kasih','baik',
        'kerja','pulang','pergi','karena','tapi','atau','kalau','masih','lagi',
        'siap','izin','minta','capek','sakit','gak','udah','gimana','dong',
        'kipas','mesin','rusak','bocor','macet','stok','habis','ganti','pasang',
        'gudang','masuk','keluar','tutup','buka','material','selesai','beres',
        'datang','besok','kemarin','libur','lembur','cuti','proses','produksi',
        'selalu','mohon','pakai','pake','cek','lihat','bilang','ambil','kirim',
        'tunggu','bantu','butuh','perlu','panggil','suruh','hati','awas',
    }
    def _replace_en(m):
        first_word = re.match(r'@([A-Za-z0-9]+)', m.group(0))
        if first_word and first_word.group(1).lower() in _id_skip:
            return m.group(0)  # Keep: not a real @mention
        return ' '
    clean = re.sub(r'@[A-Za-z0-9][A-Za-z0-9 _.-]*(?:\s+[\u4e00-\u9fff]{1,4})?(?=(?:\s|[\n,\uff0c\u3002!\uff01?\uff1f:\uff1a;\uff1b()\uff08\uff09\[\]{}<>\u201c\u201d]|$))', _replace_en, text)
    # Strip Chinese @mentions
    clean = re.sub(r'@[\u4e00-\u9fff]+(?:\s*[\uff08(][^\uff09)]*[\uff09)])?', ' ', clean)
    return clean


def has_chinese(text):
    return len(re.findall(r'[\u4e00-\u9fff]', text)) >= 2


def has_japanese(text):
    hira = len(re.findall(r'[\u3040-\u309f]', text))
    kata = len(re.findall(r'[\u30a0-\u30ff]', text))
    return (hira + kata) >= 2


def has_korean(text):
    return len(re.findall(r'[\uac00-\ud7af]', text)) >= 2


def has_thai(text):
    return len(re.findall(r'[\u0e00-\u0e7f]', text)) >= 2


def has_vietnamese(text):
    vi_special = re.findall(r'[\u01a0\u01a1\u01af\u01b0\u0110\u0111]', text)
    vi_chars = re.findall(r'[\u00e0-\u00ff\u1ea0-\u1ef9]', text.lower())
    vi_marks = re.findall(r'[\u0300-\u036f]', text)
    words = text.lower().split()
    vi_markers = set([
        'cua', 'nhung', 'trong', 'duoc', 'khong', 'nhu', 'mot',
        'toi', 'ban', 'anh', 'chi', 'em', 'ong', 'ba',
        'la', 'va', 'cac', 'cho', 'voi', 'tai', 'nay', 'khi',
        'con', 'roi', 'lam', 'biet', 'muon', 'den', 'di',
        'xin', 'cam', 'chao', 'dep', 'ngon', 'tot', 'xau',
    ])
    marker_count = sum(1 for w in words if w in vi_markers)
    if len(vi_special) >= 1:
        return True
    if len(vi_chars) >= 3 and marker_count >= 1:
        return True
    if len(vi_marks) >= 2 and marker_count >= 1:
        return True
    return False


def has_indonesian(text):
    if has_chinese(text) or has_thai(text) or has_korean(text) or has_japanese(text):
        return False
    words = re.findall(r'[a-zA-Z]+', text.lower())
    if len(words) < 2:
        return False
    id_words = set([
        # ── Pronouns / titles ──
        'saya', 'aku', 'gue', 'gw', 'kamu', 'lu', 'elo', 'dia', 'mereka',
        'kami', 'kita', 'kalian', 'bapak', 'ibu', 'pak', 'bu', 'mas', 'mbak',
        'bang', 'kak', 'om', 'tante', 'bos', 'boss', 'gan',
        # ── Particles / fillers ──
        'ya', 'lah', 'loh', 'dong', 'sih', 'nih', 'kok', 'deh', 'kan',
        'tuh', 'nah', 'wah', 'aduh', 'masa', 'emang', 'kayak', 'kayaknya',
        'soalnya', 'makanya', 'jadinya', 'aja', 'doang', 'cuma', 'gitu',
        'gini', 'sini', 'sana', 'situ', 'mana', 'iya', 'oke',
        # ── Prepositions / conjunctions ──
        'di', 'ke', 'dari', 'pada', 'oleh', 'untuk', 'dengan', 'supaya',
        'agar', 'karena', 'tetapi', 'tapi', 'namun', 'sehingga', 'meskipun',
        'walaupun', 'sebelum', 'sesudah', 'setelah', 'selama', 'ketika',
        'sambil', 'tanpa', 'antara', 'tentang', 'terhadap', 'atau', 'dan',
        'jika', 'kalau', 'biar', 'sampai',
        # ── Question words ──
        'apa', 'siapa', 'dimana', 'kapan', 'kenapa', 'bagaimana', 'berapa',
        'gimana', 'mana', 'mengapa',
        # ── Common verbs ──
        'ada', 'adalah', 'ambil', 'angkat', 'antar', 'atur', 'bangun',
        'bantu', 'bawa', 'bayar', 'beli', 'berangkat', 'berhenti', 'bicara',
        'bilang', 'bisa', 'bikin', 'boleh', 'buat', 'buang', 'buka', 'butuh',
        'cari', 'catat', 'cek', 'coba', 'cuci', 'dapat', 'datang', 'duduk',
        'ganti', 'hapus', 'hitung', 'hubungi', 'ikut', 'ingat', 'isi',
        'jaga', 'jalan', 'jawab', 'jemput', 'jual', 'kasih', 'kejar',
        'keluar', 'kembali', 'kirim', 'kurang', 'lari', 'lepas', 'lewat',
        'lihat', 'lupa', 'makan', 'masak', 'masuk', 'mau', 'minum', 'minta',
        'naik', 'paham', 'pakai', 'pake', 'panggil', 'pasang', 'perbaiki',
        'pergi', 'periksa', 'pindah', 'potong', 'pulang', 'selesai',
        'sembuh', 'simpan', 'suruh', 'tahu', 'tau', 'tambah', 'tanya',
        'taruh', 'tiba', 'tidur', 'tinggal', 'tolong', 'tukar', 'tulis',
        'tunggu', 'turun', 'tutup', 'ngerti', 'paham', 'ngomong', 'ngobrol',
        'nyari', 'nyoba', 'nunggu', 'ngitung', 'ngirim', 'ngecek', 'ngangkat',
        'ingin', 'harus', 'boleh', 'perlu', 'wajib',
        # ── Common adjectives ──
        'bagus', 'baik', 'baru', 'benar', 'berat', 'besar', 'bersih',
        'buruk', 'cepat', 'dingin', 'gampang', 'gelap', 'jelek', 'kecil',
        'keras', 'kotor', 'kuat', 'lambat', 'lama', 'lebar', 'lemah',
        'lurus', 'mahal', 'miring', 'murah', 'panas', 'panjang', 'pendek',
        'penuh', 'rata', 'ringan', 'salah', 'sehat', 'sempit', 'susah',
        'tajam', 'tebal', 'terang', 'tipis', 'tua', 'muda', 'lembut',
        'kasar', 'kosong', 'basah', 'kering',
        # ── Nouns (general) ──
        'air', 'api', 'asap', 'barang', 'batu', 'biaya', 'botol', 'cat',
        'dinding', 'ember', 'gelas', 'helm', 'kabel', 'kaca', 'kain',
        'kamar', 'kayu', 'kertas', 'kotak', 'kursi', 'lampu', 'listrik',
        'meja', 'mobil', 'motor', 'obat', 'paku', 'papan', 'pintu',
        'plastik', 'rak', 'roda', 'sabun', 'sapu', 'selang', 'sepatu',
        'surat', 'tangga', 'tali', 'tas', 'tiang', 'topi', 'truk',
        # ── Nouns (work / factory) ──
        'absen', 'alat', 'atasan', 'bahan', 'bengkel', 'bor', 'crane',
        'debu', 'forklift', 'gaji', 'gerinda', 'gudang', 'jadwal',
        'kartu', 'kecelakaan', 'kerusakan', 'kualitas', 'laporan',
        'las', 'limbah', 'lini', 'logam', 'lubang', 'mandor', 'masalah',
        'meter', 'mutu', 'pabrik', 'pekerja', 'pelindung', 'peralatan',
        'perbaikan', 'peraturan', 'permukaan', 'produksi', 'produk',
        'rapat', 'sabuk', 'sampel', 'shift', 'sisa', 'sparepart',
        'supervisor', 'tabung', 'tekanan', 'timbangan', 'toleransi',
        'tungku', 'upah', 'wadah',
        # ── Factory equipment / materials ──
        'kipas', 'angin', 'mesin', 'pompa', 'kunci', 'baut', 'mur',
        'pipa', 'oli', 'besi', 'baja', 'batang', 'stok', 'material',
        'lantai', 'atas', 'bawah', 'ukuran', 'nomor',
        'bocor', 'macet', 'mati', 'hidup', 'nyala', 'jalan',
        # ── Factory actions ──
        'ukur', 'timbang', 'sortir', 'pisah', 'gabung', 'campur', 'cetak',
        'press', 'poles', 'tekuk', 'lipat', 'gulung', 'tarik', 'dorong',
        'geser', 'putar', 'balik', 'susun', 'tumpuk', 'bungkus', 'ikat',
        'segel', 'proses', 'bagian', 'tempat',
        # ── Safety / quality ──
        'bahaya', 'aman', 'keselamatan', 'cidera', 'luka', 'awas', 'hati',
        'peringatan', 'darurat', 'masker', 'kacamata', 'rompi', 'cacat',
        'retak', 'gores', 'bengkok', 'penyok', 'standar', 'inspeksi',
        'audit', 'lapor',
        # ── Time ──
        'detik', 'menit', 'jam', 'hari', 'minggu', 'bulan', 'tahun',
        'pagi', 'siang', 'sore', 'malam', 'subuh', 'kemarin', 'sekarang',
        'besok', 'lusa', 'nanti', 'dulu', 'tadi', 'segera', 'selalu',
        'sering', 'kadang', 'jarang',
        # ── Numbers / quantity ──
        'satu', 'dua', 'tiga', 'empat', 'lima', 'enam', 'tujuh',
        'delapan', 'sembilan', 'sepuluh', 'puluh', 'ratus', 'ribu',
        'juta', 'setengah', 'cukup', 'terlalu', 'sekitar', 'kira',
        'banyak', 'sedikit', 'semua', 'beberapa',
        # ── States / emotions ──
        'senang', 'sedih', 'marah', 'takut', 'capek', 'cape', 'males',
        'lapar', 'haus', 'sakit', 'sehat', 'ngantuk', 'bosan', 'bingung',
        'kaget', 'malu', 'bangga', 'puas', 'kecewa', 'khawatir', 'tenang',
        'sibuk', 'santai', 'mantap', 'keren', 'asik',
        # ── Daily / HR ──
        'izin', 'cuti', 'libur', 'lembur', 'istirahat', 'kerja', 'masuk',
        'pulang', 'absen', 'telat', 'terlambat', 'ijin', 'sakit', 'mangkir',
        'resign', 'kontrak', 'tetap', 'harian', 'bulanan', 'THR',
        # ── Negation / affirmation ──
        'tidak', 'bukan', 'belum', 'jangan', 'sudah', 'akan', 'sedang',
        'masih', 'lagi', 'saja', 'juga', 'pernah', 'tidak', 'tanpa',
        'hanya', 'bahkan', 'sangat', 'amat', 'sekali', 'paling',
        # ── Slang abbreviations ──
        'gak', 'nggak', 'ga', 'gk', 'udah', 'udh', 'uda',
        'gmn', 'bgt', 'org', 'yg', 'tdk', 'dg', 'dgn', 'krn',
        'blm', 'bs', 'sy', 'trs', 'tp', 'tpi', 'sm', 'lg',
        'dl', 'skrg', 'hr', 'msh', 'brp', 'dpt', 'hrs', 'kmrn',
        'bsk', 'wkwk', 'otw', 'gpp', 'jgn', 'tlg', 'cb', 'emg',
        'stlh', 'sblm', 'tgl', 'mksd', 'kl', 'krj', 'plg', 'msk',
        'klr', 'btw', 'fyi', 'cmn', 'drpd', 'blg', 'klo', 'knp',
        'dmn', 'gmna', 'bkn', 'sbg', 'ttg', 'scr', 'utk',
        # ── Common responses ──
        'siap', 'beres', 'selesai', 'oke', 'sip', 'mantap', 'lanjut',
        'betul', 'benar', 'setuju', 'mengerti', 'paham', 'jelas',
        'terima', 'kasih', 'makasih', 'maaf', 'permisi', 'selamat',
        'halo', 'hai', 'assalamualaikum', 'waalaikumsalam',
        # ── Misc common ──
        'orang', 'baru', 'lain', 'beda', 'sama', 'sendiri', 'bersama',
        'bareng', 'duluan', 'belakangan', 'awal', 'akhir', 'mulai',
        'setiap', 'tiap', 'per', 'masing', 'soal', 'hal', 'cara',
        'jenis', 'tipe', 'macam', 'warna', 'bentuk', 'sisi', 'ujung',
        'tengah', 'tepi', 'pinggir', 'depan', 'belakang', 'kiri', 'kanan',
        'dalam', 'luar', 'atas', 'bawah', 'samping', 'sebelah',
        'dekat', 'jauh', 'sini', 'sana', 'situ',
        'yang', 'ini', 'itu', 'rumah', 'kantor', 'uang', 'harga',
        'yuk', 'ayo', 'banget', 'ruang', 'baca', 'ujian', 'terakhir',
        'punya', 'jadi', 'mohon', 'saat', 'secara', 'harap', 'rusak',
        'habis', 'bulat', 'kamu',
        # ── Food / break ──
        'nasi', 'ayam', 'ikan', 'sayur', 'teh', 'kopi', 'susu', 'roti',
        'mie', 'goreng', 'rebus', 'pedas', 'manis', 'asin', 'pahit',
        'warung', 'kantin',
        # ── Missing common words (comprehensive) ──
        'kata', 'sandi', 'nama', 'alamat', 'telepon', 'email', 'buku',
        'dengar', 'pikir', 'rasa', 'cinta', 'suka', 'benci',
        'teman', 'musuh', 'keluarga', 'anak', 'istri', 'suami',
        'adik', 'kakak', 'ayah', 'nenek', 'kakek', 'paman', 'bibi',
        # ── Places ──
        'negara', 'kota', 'desa', 'gedung', 'toko', 'sekolah',
        'masjid', 'gereja', 'pasar', 'bandara', 'stasiun', 'terminal',
        'rumah', 'hotel', 'restoran', 'warnet', 'bengkel',
        # ── Transport ──
        'sepeda', 'pesawat', 'kapal', 'kereta', 'bis', 'taksi', 'ojek',
        # ── Food detail ──
        'makanan', 'minuman', 'buah', 'sayuran', 'daging', 'beras',
        'garam', 'gula', 'minyak', 'tepung', 'bumbu', 'sambal',
        'telur', 'tempe', 'tahu', 'soto', 'bakso', 'sate',
        # ── Body ──
        'tangan', 'kaki', 'kepala', 'mata', 'telinga', 'mulut',
        'hidung', 'perut', 'punggung', 'dada', 'bahu', 'jari',
        'lutut', 'siku', 'leher', 'pinggang', 'tumit', 'bibir',
        # ── Abstract / reasoning ──
        'milik', 'hak', 'kewajiban', 'tugas', 'tanggung', 'jawab',
        'solusi', 'metode', 'alasan', 'tujuan', 'maksud', 'arti',
        'makna', 'contoh', 'info', 'informasi', 'berita', 'pesan',
        # ── Ability / certainty ──
        'mampu', 'sanggup', 'berani', 'gembira', 'sulit', 'mudah',
        'lebih', 'hampir', 'nyaris', 'mungkin', 'pasti', 'tentu',
        'yakin', 'ragu', 'percaya',
        # ── ber- prefix verbs ──
        'bekerja', 'belajar', 'bermain', 'berlari', 'berjalan',
        'berbicara', 'berpikir', 'berharap', 'berdoa', 'bernyanyi',
        'beristirahat', 'berbelanja', 'bertemu', 'bercerita',
        'berdiri', 'berbaring', 'berputar', 'bergerak', 'berhenti',
        'bergabung', 'berpisah', 'bertugas', 'bertanya', 'berubah',
        # ── me- prefix verbs ──
        'membuat', 'membeli', 'menjual', 'membawa', 'mengambil',
        'memberikan', 'menerima', 'mengirim', 'menyimpan', 'membuang',
        'mencari', 'menemukan', 'menunggu', 'melihat', 'mendengar',
        'menulis', 'membaca', 'menghitung', 'mengukur', 'memotong',
        'membuka', 'menutup', 'menyalakan', 'mematikan', 'menghubungi',
        'menelepon', 'mengecek', 'memeriksa', 'memperbaiki', 'mengganti',
        'memasang', 'melepas', 'mengisi', 'mengosongkan', 'membersihkan',
        'mencuci', 'membantu', 'meminta', 'memakai', 'memasak',
        'memilih', 'memiliki', 'mengerti', 'mengetahui', 'memulai',
        'menyelesaikan', 'mengerjakan', 'melapor', 'melaporkan',
        'mengatur', 'mengantar', 'menjaga', 'menjemput', 'menaruh',
        'memindahkan', 'mengangkat', 'menurunkan', 'mendorong',
        'menarik', 'memutar', 'menekan', 'mengunci', 'merasa',
        # ── di- prefix (passive) ──
        'dibuat', 'dibeli', 'dijual', 'dibawa', 'diambil', 'dikirim',
        'disimpan', 'dibuang', 'dicari', 'ditemukan', 'dilihat',
        'ditulis', 'dibaca', 'dihitung', 'diukur', 'dipotong',
        'dibuka', 'ditutup', 'dinyalakan', 'dimatikan', 'dicek',
        'diperiksa', 'diperbaiki', 'diganti', 'dipasang', 'dilepas',
        'diisi', 'dibersihkan', 'dicuci', 'diminta', 'dipakai',
        'dipilih', 'diketahui', 'dikerjakan', 'dilaporkan', 'diatur',
        'dijaga', 'ditaruh', 'dipindahkan', 'diangkat', 'diturunkan',
        'diperhatikan', 'disampaikan', 'dilakukan', 'diberikan',
        'diterima', 'digunakan', 'disediakan', 'dibutuhkan',
        # ── Documents / admin ──
        'formulir', 'dokumen', 'berkas', 'file', 'data', 'rekening',
        'tabungan', 'pinjaman', 'hutang', 'bayaran', 'diskon', 'gratis',
        'untung', 'rugi', 'modal', 'surat', 'izin', 'tanda', 'tangan',
        'stempel', 'cap', 'kuitansi', 'faktur', 'invoice', 'nota',
        # ── Dimensions / colors ──
        'tinggi', 'rendah', 'dangkal', 'halus', 'cair', 'padat', 'lunak',
        'merah', 'kuning', 'hijau', 'biru', 'putih', 'hitam',
        'coklat', 'abu', 'emas', 'perak', 'ungu', 'oranye', 'pink',
        # ── Weather / nature ──
        'hujan', 'awan', 'mendung', 'cerah', 'badai', 'banjir',
        'gempa', 'petir', 'kabut', 'embun',
        # ── Speed / manner ──
        'bentar', 'sebentar', 'langsung', 'pelan', 'keras', 'kencang',
        'terburu', 'santai', 'harian', 'mingguan', 'bulanan', 'tahunan',
        # ── ke-...-an nouns ──
        'keselamatan', 'kecelakaan', 'kerusakan', 'kebersihan',
        'keamanan', 'kesehatan', 'kecepatan', 'keterlambatan',
        'kekurangan', 'kelebihan', 'kesalahan', 'keberhasilan',
        'kemampuan', 'kebutuhan', 'keperluan', 'keterangan',
        # ── per-...-an nouns ──
        'perbaikan', 'perubahan', 'perbedaan', 'perhatian',
        'perkembangan', 'pertemuan', 'perjanjian', 'perusahaan',
        'pekerjaan', 'peralatan', 'peraturan', 'perlengkapan',
        'permintaan', 'pengiriman', 'penggantian', 'pemasangan',
        'pemeriksaan', 'pembersihan', 'pengisian', 'pengecekan',
        'penggunaan', 'pemakaian', 'pelaksanaan', 'pelaporan',
        # ── Common endings -kan / -i ──
        'pastikan', 'perhatikan', 'sampaikan', 'lakukan', 'berikan',
        'gunakan', 'sediakan', 'siapkan', 'selesaikan', 'kerjakan',
        'beritahukan', 'hubungi', 'temui', 'cari', 'ambilkan',
        'tolong', 'mohon', 'harap', 'silakan', 'silahkan',
    ])
    count = sum(1 for w in words if w in id_words)
    if count >= 2:
        return True
    if count >= 1 and len(words) >= 2 and count / len(words) >= 0.4:
        return True
    return False


def has_english(text):
    if has_chinese(text) or has_thai(text) or has_korean(text) or has_japanese(text):
        return False
    if has_vietnamese(text) or has_indonesian(text):
        return False
    words = re.findall(r'[a-zA-Z]+', text.lower())
    if len(words) < 3:
        return False
    en_words = set([
        'the', 'is', 'are', 'was', 'were', 'have', 'has', 'had',
        'will', 'would', 'could', 'should', 'can', 'may', 'might',
        'this', 'that', 'these', 'those', 'what', 'which', 'who',
        'where', 'when', 'how', 'why', 'not', 'but', 'and', 'or',
        'for', 'with', 'from', 'about', 'into', 'your', 'you',
        'we', 'they', 'she', 'him', 'her', 'its', 'our', 'their',
        'just', 'also', 'very', 'much', 'more', 'most', 'some',
        'any', 'all', 'each', 'every', 'been', 'being', 'does',
        'did', 'doing', 'going', 'want', 'need', 'know', 'think',
        'come', 'make', 'like', 'time', 'good', 'new', 'first',
        'please', 'thank', 'thanks', 'sorry', 'hello', 'okay',
        'yes', 'yeah', 'already', 'still', 'here', 'there',
    ])
    count = sum(1 for w in words if w in en_words)
    if count >= 2:
        return True
    if len(words) > 0 and count / len(words) > 0.25:
        return True
    return False


def detect_language(text):
    """Detect language: Chinese → 'zh', Latin text → 'id'.
    For mixed messages (factory codes + Chinese), Chinese dominates."""
    clean = strip_mentions_for_detect(text).strip()
    if not clean or len(clean) < 2:
        return None
    zh_count = len(re.findall(r'[\u4e00-\u9fff]', clean))
    latin_words = re.findall(r'[a-zA-Z]{2,}', clean.lower())
    # Has Chinese characters — if Chinese dominates or Latin is minimal, it's Chinese
    if zh_count >= 2 and zh_count >= len(latin_words):
        return "zh"
    # No Chinese but has Latin → Indonesian
    if zh_count == 0 and latin_words:
        return "id"
    # Both exist but Latin dominates → Indonesian
    if latin_words and len(latin_words) > zh_count:
        return "id"
    # Only Chinese (1+ chars)
    if zh_count >= 1:
        return "zh"
    # Has some Latin words
    if latin_words:
        return "id"
    return None


def contains_source_script_outside_placeholders(text, src):
    cleaned = re.sub(r'__MENTION_\d+__', ' ', text or '')
    cleaned = re.sub(r'__CUST_\d+__', ' ', cleaned)
    # Also strip known customer names (they are kept in original language intentionally)
    for name in CUSTOMER_NAMES:
        if name in cleaned:
            cleaned = cleaned.replace(name, ' ')
    patterns = {
        "zh": r'[\u4e00-\u9fff]',
        "ja": r'[\u3040-\u30ff\u4e00-\u9fff]',
        "ko": r'[\uac00-\ud7af]',
        "th": r'[\u0e00-\u0e7f]',
    }
    pattern = patterns.get(src)
    if not pattern:
        return False
    return len(re.findall(pattern, cleaned)) >= 2


def is_translation_valid(result, src, tgt):
    if not result or not result.strip():
        return False
    if src != tgt and contains_source_script_outside_placeholders(result, src):
        return False
    return True



# === Factory semantic translation engine (ID -> ZH) ===
# Purpose: do not rely on prompt-only translation for factory chat.  This layer
# detects factory domains, normalizes Indonesian shop-floor terms into semantic
# slots, validates Chinese output, and repairs literal translations before reply.

FACTORY_ID_ZH_OBJECTS = {
    "barang": "料件",
    "barangnya": "料件",
    "material": "材料",
    "materialnya": "材料",
    "bahan": "材料",
    "batang": "棒材",
    "batangnya": "棒材",
    "batang baja": "棒材",
    "bundel": "這把",
    "ikat": "這捆",
    "lot": "這批料",
    "order": "這張單",
    "work order": "工單",
    "wo": "工單",
    "tag": "TAG",
    "label": "標籤",
    "mesin": "機台",
    "mesinnya": "機台",
    "troli": "台車",
    "trolley": "台車",
    "gudang": "倉庫",
}

FACTORY_ID_ZH_DEFECTS = {
    "rusak": "損傷",
    "cacat": "異常",
    "lecet": "擦傷",
    "gores": "刮傷",
    "goresan": "刮傷",
    "tergores": "刮傷",
    "retak": "裂傷",
    "patah": "斷裂",
    "bengkok": "彎曲",
    "penyok": "凹傷",
    "aus": "磨損",
    "kasar": "表面粗糙",
    "kotor": "髒污",
    "karat": "生鏽",
    "berkarat": "生鏽",
    "bocor": "漏油",
    "macet": "卡住",
    "habis": "用完",
    "kurang": "短少",
    "salah": "錯誤",
    "tercampur": "混料",
}

FACTORY_ID_ZH_POSITIONS = {
    "dari belakang": "後端",
    "bagian belakang": "後端",
    "ujung belakang": "後端",
    "belakang": "後端",
    "dari depan": "前端",
    "bagian depan": "前端",
    "ujung depan": "前端",
    "depan": "前端",
    # 研磨工廠專用:自由端 / 夾頭端(rod 兩端在機台上的位置)
    "ujung bebas": "自由端",
    "ujung bebasnya": "自由端",
    "bebas": "自由端",  # 在研磨語境下單獨出現的「bebas」幾乎都是指自由端
    "ujung jepit": "夾頭端",
    "ujung pencekam": "夾頭端",
    "jepit": "夾頭端",
    # 拼字容錯:ujing / ujong / ujug 都是 ujung 的常見誤寫
    "ujing": "端",
    "ujong": "端",
    "ujug": "端",
    "bagian tengah": "中段",
    "tengah": "中段",
    "sebelah samping": "側邊",
    "samping": "側邊",
    "bagian atas": "上方",
    "atas": "上方",
    "bagian bawah": "下方",
    "bawah": "下方",
    "permukaan": "表面",
    "luar": "外側",
    "dalam": "內側",
}

# v3.9.30: 期限/排程/積壓詞彙表
# 這張表針對工廠最常被誤譯的時間/排程詞彙(印尼班長公告幾乎句句出現)
# 重點: jatuh tempo 在工廠語境 ≠ 字面「到期」(機器壽命到期),而是「該工序積壓未處理」
# 案例:歐那回報 "30 ton mesin pemoles yang jatuh tempo bulan lalu"
#       → 錯翻成「30 噸的拋光機到期」(把料量翻成機器重量)
#       → 正確「上月積壓 30 噸料(待拋光)」
FACTORY_ID_ZH_TIME = {
    "jatuh tempo": "積壓到期",
    "sudah jatuh tempo": "已積壓",
    "belum jatuh tempo": "尚未到期",
    "jatuh tempo bulan lalu": "上月積壓未處理",
    "jatuh tempo bulan ini": "本月到期要處理",
    "jatuh tempo minggu lalu": "上週積壓未處理",
    "jatuh tempo minggu ini": "本週到期要處理",
    "jatuh tempo hari ini": "今日到期",
    "tertunda": "遞延",
    "tunggakan": "積壓量",
    "menunggak": "積壓",
    "deadline": "交期",
    "tenggat": "期限",
    "jadwal": "排程",
    "schedule": "排程",
    "overdue": "已逾期",
    "tepat waktu": "準時",
    "telat": "延遲",
    "terlambat": "延遲",
    "sebelum": "之前",
    "setelah": "之後",
    "mendesak": "急迫",
    "urgent": "急迫",
}

FACTORY_DOMAIN_KEYWORDS_ID = {
    "quality_issue": [
        "rusak", "cacat", "lecet", "gores", "goresan", "tergores", "retak", "patah",
        "bengkok", "penyok", "aus", "kasar", "karat", "berkarat", "visual", "qc",
        "lulus", "tidak lulus", "reject", "rework", "toleransi", "diameter", "ukuran",
    ],
    "material_flow": [
        "barang", "material", "bahan", "batang", "bundel", "lot", "work order", "wo",
        "masuk gudang", "gudang", "packing", "di-packing", "station", "stasiun", "line",
        "tag", "label", "heat number", "masuk", "keluar", "taruh", "letakkan",
        # v3.9.30: 工序/排程相關詞(jatuh tempo / tertunda 等也算工廠語境)
        "jatuh tempo", "tertunda", "tunggakan", "deadline", "urgent",
    ],
    "equipment": [
        "mesin", "batu gerinda", "gerinda", "bearing", "kopel", "as ", "roda", "pompa",
        "pipa", "oli", "bocor", "macet", "maintenance", "perbaiki", "mati", "jalan",
        # v3.9.30: 工序動詞 / 機器名稱(讓「30 ton pemoles」這種短句也能命中工廠 domain)
        "pemoles", "polishing", "pemolesan", "grinding", "sanding",
        "drawing", "annealing", "straightening", "peeling",
    ],
    "safety": [
        "bahaya", "awas", "hati-hati", "pelindung", "interlock", "crane", "forklift",
        "dilarang", "safety", "kacamata", "helm", "sarung tangan",
    ],
}

FACTORY_ZH_LITERAL_RISK = {
    "物品": "料件",
    "東西": "料件",
    "從後面": "後端",
    "從前面": "前端",
    "後面損壞": "後端損傷",
    "前面損壞": "前端損傷",
    "後面壞了": "後端損傷",
    "前面壞了": "前端損傷",
    "損壞": "損傷",
    "壞掉": "損傷",
    "壞了": "損傷",
}

FACTORY_BAD_ZH_PATTERNS = [
    r"物品從(後面|前面)(損壞|壞了|壞掉)",
    r"東西從(後面|前面)(損壞|壞了|壞掉)",
    r"材料從(後面|前面)(損壞|壞了|壞掉)",
    r"棒材從(後面|前面)(損壞|壞了|壞掉)",
]


def _clean_factory_id(text):
    t = (text or "").strip().lower()
    t = re.sub(r"[。．.！!？?，,：:；;()（）\[\]{}]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def detect_factory_domain(text, src=None, tgt=None):
    """Return a lightweight factory-domain profile used before/after translation."""
    t = _clean_factory_id(text) if src == "id" else (text or "").lower()
    domains = set()
    if src == "id":
        for domain, kws in FACTORY_DOMAIN_KEYWORDS_ID.items():
            if any(kw in t for kw in kws):
                domains.add(domain)
        if any(p in t for p in FACTORY_ID_ZH_POSITIONS):
            domains.add("material_direction")
    else:
        zh_kws = ["料", "棒材", "工單", "包裝", "入庫", "站", "機台", "品保", "異常", "損傷", "尺寸", "公差", "研磨", "削皮"]
        if any(k in (text or "") for k in zh_kws):
            domains.add("factory")
    if domains:
        return {"is_factory": True, "domains": sorted(domains)}
    return {"is_factory": False, "domains": []}


def _find_longest_phrase(table, text):
    for k in sorted(table.keys(), key=len, reverse=True):
        if re.search(r"(?<![a-z])" + re.escape(k) + r"(?![a-z])", text):
            return k, table[k]
    return None, None


# ===== v3.3 訊息分類器(工廠翻譯架構 v2)=====
# 公告/廣播訊號詞 - 命中即視為 announcement,絕不走槽位拼接
ANNOUNCEMENT_SIGNALS = [
    # 印尼文公告詞
    "@all", "pemberitahuan", "pengumuman", "peringatan",
    "bagi pekerja", "bagi semua", "untuk semua", "kepada semua",
    "diharapkan", "wajib", "mohon perhatian", "harap diingat",
    "tolong perhatikan", "perhatian:", "info:", "informasi:",
    # 繁中公告詞
    "公告", "通知", "全體", "請注意", "敬請", "通報", "宣達",
    "請大家", "各位", "全部人員", "所有同仁",
]

# 現場簡訊長度上限
SHORT_INCIDENT_MAX_LEN = 30


def classify_factory_message(text, src=None):
    """訊息分類器:announcement / incident / general
    
    這是工廠翻譯架構 v2 的入口。所有舊版 factory_semantic_*
    函數現在都會先呼叫這個分類器,只有 incident 才走嚴格槽位邏輯。
    
    回傳 dict { type, reason, len }
    """
    raw = (text or "").strip()
    if not raw:
        return {"type": "general", "reason": "empty", "len": 0}
    
    low = raw.lower()
    char_len = len(re.sub(r"\s+", "", raw))
    
    # 規則 1:任一公告訊號詞 → announcement
    for sig in ANNOUNCEMENT_SIGNALS:
        if sig in low:
            return {"type": "announcement", "reason": f"signal:{sig}", "len": char_len}
    
    # 規則 2:長度 > 50 且含廣播符號 → announcement
    if char_len > 50 and ("@" in raw or " all " in (" " + low + " ")):
        return {"type": "announcement", "reason": "long+broadcast_marker", "len": char_len}
    
    # 規則 3:長度 > 80 → 一般長訊息(不該被槽位吃掉)
    if char_len > 80:
        return {"type": "general", "reason": "too_long_for_incident", "len": char_len}
    
    # 規則 4+5:只有印尼文 + 短/中等長度 + 命中工廠術語才視為 incident
    if char_len <= 50 and src == "id":
        t = _clean_factory_id(raw)
        has_obj = any(re.search(r"(?<![a-z])" + re.escape(k) + r"(?![a-z])", t) 
                      for k in FACTORY_ID_ZH_OBJECTS)
        has_defect = any(re.search(r"(?<![a-z])" + re.escape(k) + r"(?![a-z])", t) 
                         for k in FACTORY_ID_ZH_DEFECTS)
        if has_obj and has_defect:
            tag = "short" if char_len <= SHORT_INCIDENT_MAX_LEN else "medium"
            return {"type": "incident", "reason": f"{tag}+obj+defect", "len": char_len}
    
    return {"type": "general", "reason": "default", "len": char_len}


def factory_semantic_translate_id_zh(text):
    """現場簡訊的語義槽位拼接(工廠翻譯架構 v2)
    
    v3.3 變更:
      - 只有 incident 類別才執行
      - 長度上限 SHORT_INCIDENT_MAX_LEN
      - 拼接結果若 < 原文 1/4 視為災難性壓縮,放棄
    """
    raw = text or ""
    t = _clean_factory_id(raw)
    if not t:
        return None
    
    # ★ v3.3:分類保護 - 公告/一般訊息不走槽位
    cls = classify_factory_message(raw, src="id")
    if cls["type"] != "incident":
        return None
    
    # ★ v3.3:硬長度上限
    if cls["len"] > SHORT_INCIDENT_MAX_LEN:
        return None

    leading_codes = re.findall(r"\b(?:[A-Z]\d[A-Z0-9-]{3,}|\d+[A-Z][A-Z0-9-]{2,})\b", raw)
    prefix = (" ".join(dict.fromkeys(leading_codes)) + " ") if leading_codes else ""

    obj_key, obj_zh = _find_longest_phrase(FACTORY_ID_ZH_OBJECTS, t)
    defect_key, defect_zh = _find_longest_phrase(FACTORY_ID_ZH_DEFECTS, t)
    pos_key, pos_zh = _find_longest_phrase(FACTORY_ID_ZH_POSITIONS, t)

    result = None
    if obj_zh and defect_zh:
        if pos_zh:
            result = (prefix + f"{obj_zh}{pos_zh}{defect_zh}").strip()
        elif obj_key in ("barang", "barangnya", "material", "materialnya",
                         "bahan", "batang", "batangnya", "batang baja"):
            result = (prefix + f"{obj_zh}{defect_zh}").strip()
    elif pos_zh and defect_zh and not obj_zh:
        result = (prefix + f"{pos_zh}{defect_zh}").strip()
    
    # ★ v3.3:資訊密度檢查 - 譯文 < 原文 1/4 = 災難性壓縮
    if result and len(result) * 4 < cls["len"]:
        return None
    if result:
        return result

    phrase_map = {
        "barang rusak dari belakang": "料件後端損傷",
        "barang bagian belakang rusak": "料件後端有損傷",
        "batang rusak dari belakang": "棒材後端損傷",
        "batang bagian belakang rusak": "棒材後端有損傷",
        "bagian belakang rusak": "後端損傷",
        "bagian depan rusak": "前端損傷",
        "barang rusak depan": "料件前端損傷",
        "barang rusak belakang": "料件後端損傷",
    }
    if t in phrase_map:
        return (prefix + phrase_map[t]).strip()

    return None


def build_factory_context_hint(text, src, tgt):
    """產生注入 GPT 的工廠語境提示(工廠翻譯架構 v2)
    
    v3.3 變更:
      - 公告類:純術語表 + 「必須完整逐句翻譯」要求,不給 deterministic 結論
      - incident 類:給術語表 + deterministic 結論(明確標示為「參考」)
      - 一般訊息:純術語表
    """
    if src == "zh" and tgt == "id":
        return build_factory_context_hint_zh_id(text)
    if src != "id" or tgt != "zh":
        return ""
    
    domain = detect_factory_domain(text, src, tgt)
    if not domain["is_factory"]:
        return ""
    
    cls = classify_factory_message(text, src="id")
    t = _clean_factory_id(text)
    
    # 收集出現的術語(供 hint 使用)
    # v3.9.30: 加入 FACTORY_ID_ZH_TIME(jatuh tempo / deadline / tertunda 等)
    terms = []
    for source, zh in {**FACTORY_ID_ZH_OBJECTS, **FACTORY_ID_ZH_DEFECTS,
                       **FACTORY_ID_ZH_POSITIONS, **FACTORY_ID_ZH_TIME}.items():
        if re.search(r"(?<![a-z])" + re.escape(source) + r"(?![a-z])", t):
            terms.append(f"{source}={zh}")
    
    # v3.9.30: 句型歧義消解
    # 工廠口語常見省略主詞,例如 "30 ton mesin pemoles yang jatuh tempo"
    # 字面意思可被誤解為「30 噸的拋光機到期」(機器本身重量)
    # 真實意思是「拋光機要處理的 30 噸料(已積壓上月)」
    # 這個 hint 會在 announcement 跟 incident 都注入
    has_ton_machine_pattern = bool(
        re.search(r'\d+\s*ton\s+mesin\s+\w+', t) or
        re.search(r'\d+\s*ton\s+(?:bahan|material|barang)\s+(?:mesin|untuk)', t)
    )
    has_jatuh_tempo = "jatuh tempo" in t
    
    pattern_hint = ""
    if has_ton_machine_pattern:
        pattern_hint += (
            " 【關鍵句型1】「[數量] ton mesin [工序]」(例:30 ton mesin pemoles)在工廠語境下,"
            "指該工序待處理/積壓的料量,**不是**機器本身的重量。"
            "請翻成「[工序]要處理的[數量]噸料」或「積壓[數量]噸料(待[工序])」,"
            "**禁止**翻成「[數量]噸的拋光機」「[數量]噸的研磨機」這類字面譯法。"
        )
    if has_jatuh_tempo:
        pattern_hint += (
            " 【關鍵句型2】「jatuh tempo」在工廠排程語境=「該完成卻未完成/積壓未處理」,"
            "**不是**字面的「設備到期/壽命到期」。"
            "「jatuh tempo bulan lalu」=上月該做沒做完(積壓單);"
            "「jatuh tempo bulan ini」=本月排程要做。"
            "翻成「積壓」「到期未處理」「該完成」皆可,**禁止**翻成「設備到期」「機器到期」。"
        )
    
    if cls["type"] == "announcement":
        # 公告:絕不給 deterministic 結論,要求完整逐句翻譯
        hint = (
            "【印尼→繁中工廠語境】這是台灣不鏽鋼棒材工廠的群組公告，"
            "必須完整逐句翻譯每一個資訊點(對象、原因、操作要求、相關站別、目的)，"
            "不可只翻關鍵詞、不可簡化成短語。輸出繁體中文現場用語。"
        )
        if terms:
            hint += " 術語對應：" + "、".join(terms) + "。"
        if pattern_hint:
            hint += pattern_hint
        return hint
    
    elif cls["type"] == "incident":
        # 現場簡訊:可給 deterministic,但明確標示為「參考」
        hint = (
            "【印尼→繁中工廠語義提示】這是台灣不鏽鋼棒材工廠的現場異常簡訊，"
            "barang 在工廠品質語境=料件/材料，不要翻物品；"
            "rusak 在品質語境=損傷/異常，不要只翻壞掉；"
            "belakang/depan 描述料件方向=後端/前端，不要翻從後面/前面；"
            "輸出繁體中文現場用語，短句保持短句。"
        )
        if terms:
            hint += " 強制術語：" + "、".join(terms) + "。"
        if pattern_hint:
            hint += pattern_hint
        deterministic = factory_semantic_translate_id_zh(text)
        if deterministic:
            hint += f" 槽位參考(僅供確認術語，實際翻譯仍需保留原文所有資訊):「{deterministic}」"
        return hint
    
    else:  # general
        hint = "【印尼→繁中工廠語境】請使用台灣工廠現場用語。"
        if terms:
            hint += " 術語對應：" + "、".join(terms) + "。"
        if pattern_hint:
            hint += pattern_hint
        return hint


def post_fix_factory_id_to_zh(src_text, zh_text):
    """Fix literal Chinese outputs in factory ID->ZH translations."""
    if not zh_text:
        return zh_text
    domain = detect_factory_domain(src_text, "id", "zh")
    result = zh_text.strip()
    if not domain["is_factory"]:
        return result

    high_risk = {
        "物品從後面損壞": "料件後端損傷",
        "物品從後面壞了": "料件後端損傷",
        "物品從後面壞掉": "料件後端損傷",
        "東西從後面損壞": "料件後端損傷",
        "東西從後面壞了": "料件後端損傷",
        "材料從後面損壞": "材料後端損傷",
        "棒材從後面損壞": "棒材後端損傷",
        "物品從前面損壞": "料件前端損傷",
        "東西從前面損壞": "料件前端損傷",
        "材料從前面損壞": "材料前端損傷",
        "棒材從前面損壞": "棒材前端損傷",
    }
    for wrong, correct in sorted(high_risk.items(), key=lambda x: -len(x[0])):
        result = result.replace(wrong, correct)

    for wrong, correct in sorted(FACTORY_ZH_LITERAL_RISK.items(), key=lambda x: -len(x[0])):
        result = result.replace(wrong, correct)

    # v3.9.30: 「噸 + 機器/工序」誤譯修補
    # 對應 Bug B3: GPT 把 "30 ton mesin pemoles" 翻成「30 噸的拋光機」
    # 只要原文是 "[N] ton mesin [工序]" 句型,目標應該是料量,不是機器本身
    src_lower = (src_text or "").lower()
    if re.search(r'\d+\s*ton\s+mesin\s+\w+', src_lower):
        # 把「N噸的[X]機」「N噸[X]機」改成「[X]要處理的N噸料」風格
        # 規則:出現「噸的拋光機/研磨機/砂光機/壓光機/矯直機」就修
        machine_words_zh = ["拋光機", "研磨機", "砂光機", "壓光機", "矯直機", "標籤機"]
        for mword in machine_words_zh:
            # 模式 A:「30噸的拋光機」「30 噸的拋光機」「30噸拋光機」「30 噸拋光機」
            # 改寫為:「30噸料(待[拋光])」
            ton_pattern = re.compile(
                r'(\d+)\s*噸\s*(?:的)?\s*' + re.escape(mword)
            )
            # 對應「料」的描述:拋光機→拋光料,研磨機→研磨料 …
            material_word = mword.replace("機", "料")  # 拋光機→拋光料
            replacement_text = r'\1噸料(待' + mword.replace("機", "") + r')'
            new_result, n_sub = ton_pattern.subn(replacement_text, result)
            if n_sub > 0:
                logger.info(
                    "[post_fix v3.9.30] ton+%s pattern: %d match(es) repaired",
                    mword, n_sub
                )
                result = new_result

    # v3.9.30: 「jatuh tempo」字面誤譯修補
    # 對應 Bug B3: GPT 把 jatuh tempo 翻成字面「到期」+ 機器名,造成「拋光機到期」歧義
    if "jatuh tempo" in src_lower:
        # (a) 固定字串修補
        jt_fixes = {
            "拋光機到期": "拋光積壓",
            "研磨機到期": "研磨積壓",
            "砂光機到期": "砂光積壓",
            "壓光機到期": "壓光積壓",
            "矯直機到期": "矯直積壓",
            "拋光機已到期": "拋光已積壓",
            "研磨機已到期": "研磨已積壓",
            # 班長公告變體
            "上個月到期的": "上月積壓的",
            "本月到期的": "本月到期要處理的",
            "上月到期的": "上月積壓的",
        }
        for wrong, correct in sorted(jt_fixes.items(), key=lambda x: -len(x[0])):
            if wrong in result:
                logger.info(
                    "[post_fix v3.9.30] jatuh_tempo fix: %r -> %r",
                    wrong, correct
                )
                result = result.replace(wrong, correct)
        
        # (b) 正則修補 — 涵蓋「拋光機到期 30 噸」「研磨機已到期,還有 20 噸」這類變體
        # 抓「(機名)+到期/已到期」即使中間有空白/標點/數字也修
        regex_jt_fixes = [
            (re.compile(r'(拋光|研磨|砂光|壓光|矯直)機\s*(?:已)?到期'),
             r'\1積壓'),
            # 「N 噸的拋光機到期」這類雙重錯(若 ton+機器修補沒抓到,這裡兜底)
            (re.compile(r'(\d+)\s*噸\s*(?:的)?\s*(拋光|研磨|砂光|壓光|矯直)機\s*(?:已)?到期'),
             r'\2積壓\1噸料'),
        ]
        for pat, repl in regex_jt_fixes:
            new_result, n_sub = pat.subn(repl, result)
            if n_sub > 0:
                logger.info(
                    "[post_fix v3.9.30] jatuh_tempo regex fix: %d match(es), pattern=%s",
                    n_sub, pat.pattern
                )
                result = new_result

    result = re.sub(r"[，,。．\s]+$", "", result)
    return result.strip()


def detect_factory_semantic_error(src_text, zh_text, src="id", tgt="zh"):
    """Detect factory-context semantic errors that normal translation confidence will miss.

    Example: 'Barang rusak dari belakang' -> '物品從後面損壞' is linguistically
    plausible but invalid in Taiwan factory quality context. This detector is
    deliberately rule-based, not model-confidence-based.
    """
    if src != "id" or tgt != "zh":
        return False, "", []
    if not zh_text:
        return True, "empty", []
    domain = detect_factory_domain(src_text, src, tgt)
    domains = domain.get("domains", [])
    if not domain.get("is_factory"):
        return False, "", domains

    t = _clean_factory_id(src_text)
    has_obj = any(re.search(r"(?<![a-z])" + re.escape(k) + r"(?![a-z])", t) for k in FACTORY_ID_ZH_OBJECTS.keys())
    has_defect = any(re.search(r"(?<![a-z])" + re.escape(k) + r"(?![a-z])", t) for k in FACTORY_ID_ZH_DEFECTS.keys())
    has_pos = any(re.search(r"(?<![a-z])" + re.escape(k) + r"(?![a-z])", t) for k in FACTORY_ID_ZH_POSITIONS.keys())
    quality_context = has_defect and (has_obj or has_pos or "quality" in domains or "material_direction" in domains)

    # Hard invalid patterns in quality/material direction context.
    for pat in FACTORY_BAD_ZH_PATTERNS:
        if re.search(pat, zh_text):
            return True, "factory_literal_direction_error", domains

    if quality_context:
        bad_terms = {
            "物品": "factory_object_literal",
            "東西": "factory_object_literal",
            "從後面": "factory_direction_literal",
            "從前面": "factory_direction_literal",
            "壞掉": "factory_defect_literal",
            "壞了": "factory_defect_literal",
        }
        for bad, reason in bad_terms.items():
            if bad in zh_text:
                return True, reason + ":" + bad, domains

    # Direction words in source should usually appear as 前端/後端/尾端/側邊 in target.
    if quality_context and has_pos:
        if ("belakang" in t or "ujung belakang" in t) and ("後端" not in zh_text and "尾端" not in zh_text):
            return True, "missing_factory_back_end_direction", domains
        if ("depan" in t or "ujung depan" in t) and "前端" not in zh_text:
            return True, "missing_factory_front_end_direction", domains

    return False, "", domains


def validate_factory_translation(src_text, zh_text, src, tgt):
    """Validate Chinese translation against factory literal-risk rules."""
    bad, reason, _domains = detect_factory_semantic_error(src_text, zh_text, src, tgt)
    if bad:
        return False, reason
    return True, ""


def repair_factory_translation_openai(src_text, bad_result, reason):
    """v3.3 修復直譯錯誤的 ID->ZH 翻譯
    
    v3.3 變更:
      - 移除 deterministic 強斷言注入(那會讓 GPT 直接複製短答案)
      - 公告類訊息要求完整翻譯,max_tokens 提升到 600
      - 依分類調整 user message 口徑
    """
    if not oai:
        return None
    try:
        cls = classify_factory_message(src_text, src="id")
        hint = build_factory_context_hint(src_text, "id", "zh")
        
        sys_prompt = (
            "你是台灣不鏽鋼棒材工廠的印尼文→繁體中文現場翻譯審核器。"
            "你的任務是修正直譯錯誤，輸出現場人員會用的繁體中文。"
            "不要解釋，不要加註解，只輸出修正後譯文。"
            "規則：barang 在工廠品質語境=料件/材料；batang=棒材；rusak=損傷/異常；"
            "belakang/depan 描述料件方向=後端/前端，不可翻成從後面/從前面。"
        )
        
        # ★ v3.3:依分類調整 user message
        if cls["type"] == "announcement":
            sys_prompt += "重要：這是公告訊息，必須完整翻譯每個資訊點，不可簡化成短語或只翻關鍵詞。"
            user_msg = (
                f"原印尼文：{src_text}\n"
                f"錯誤中文(可能因簡化或漏譯)：{bad_result}\n"
                f"錯誤原因：{reason}\n"
                f"術語提示：{hint}\n"
                f"請輸出完整逐句翻譯後的繁體中文(必須涵蓋原文所有資訊):"
            )
            max_tok = 600  # 公告需要長譯文
        else:
            user_msg = (
                f"原印尼文：{src_text}\n"
                f"錯誤中文：{bad_result}\n"
                f"錯誤原因：{reason}\n"
                f"語境提示：{hint}\n"
                f"請輸出修正後繁體中文："
            )
            max_tok = 300
        
        # v3.9.8: build kwargs with model_supports() filter so GPT-5 series doesn't 400.
        repair_kwargs = {
            "model": pick_model(src_text),
            "messages": [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": user_msg}
            ],
        }
        _rmodel = repair_kwargs["model"]
        if model_supports(_rmodel, "temperature"):
            repair_kwargs["temperature"] = 0.0
        if model_supports(_rmodel, "top_p"):
            repair_kwargs["top_p"] = 1.0
        # v3.9.30b B11 修補: reasoning model 要加 reasoning 預算
        if model_supports(_rmodel, "max_completion_tokens"):
            # reasoning model: max_tok(輸出) + 4000(reasoning 緩衝)
            _is_reasoning = not model_supports(_rmodel, "temperature")
            _budget = max_tok + (4000 if _is_reasoning else 0)
            repair_kwargs["max_completion_tokens"] = _budget
        elif model_supports(_rmodel, "max_tokens"):
            repair_kwargs["max_tokens"] = max_tok
        # Translation-optimal reasoning effort for GPT-5 family
        _opt = optimal_reasoning_for_translation(_rmodel)
        if model_supports(_rmodel, "reasoning_effort") and _opt:
            repair_kwargs["reasoning_effort"] = _opt
        if model_supports(_rmodel, "verbosity"):
            repair_kwargs["verbosity"] = "low"
        try:
            # `prediction` is for low-latency edits — only on classical models.
            if bad_result and model_supports(_rmodel, "stop"):  # stop is a good proxy for "classical model"
                repair_kwargs["prediction"] = {
                    "type": "content",
                    "content": bad_result,
                }
            _ck_r = _build_cache_key(getattr(_tl, 'group_id', ''), "id", "zh", "repair")
            if _ck_r and model_supports(_rmodel, "prompt_cache_key"):
                repair_kwargs["prompt_cache_key"] = _ck_r
        except Exception:
            pass
        r = oai.chat.completions.create(**repair_kwargs)
        track_tokens(r)
        return (r.choices[0].message.content or "").strip()
    except Exception as e:
        logger.error("Factory repair error: %s", e)
        return None


def finalize_factory_translation(src_text, result, src, tgt):
    """Post-process, validate, repair; deterministic fallback for covered factory shapes."""
    if not result:
        return result
    if src == "id" and tgt == "zh":
        result = post_fix_factory_id_to_zh(src_text, result)
        ok, reason = validate_factory_translation(src_text, result, src, tgt)
        if ok:
            return result
        repaired = repair_factory_translation_openai(src_text, result, reason)
        if repaired:
            repaired = post_fix_factory_id_to_zh(src_text, repaired)
            ok2, _ = validate_factory_translation(src_text, repaired, src, tgt)
            if ok2:
                return repaired
        # ★ v3.3:fallback 槽位拼接只用於 incident 類別
        # 公告/一般訊息寧可回傳 repaired 或原 result(可能不完美),
        # 也不要被槽位拼接吃掉所有資訊
        cls = classify_factory_message(src_text, src="id")
        if cls["type"] == "incident":
            fallback = factory_semantic_translate_id_zh(src_text)
            if fallback:
                return fallback
        # 公告/一般訊息:回傳 repaired 或原 result
        if repaired:
            return repaired
    if src == "zh" and tgt == "id":
        raw = result
        bad, reason, domains = detect_factory_semantic_error_zh_id(src_text, raw)
        result = post_fix_factory_zh_to_id(src_text, raw)
        bad2, reason2, _ = detect_factory_semantic_error_zh_id(src_text, result)
        if bad or bad2:
            _tl.factory_audit = {
                "src": src_text,
                "type": "factory_semantic_auto_detected_zh_id",
                "reason": reason2 or reason,
                "raw_translation": raw,
                "corrected_translation": result,
                "domain": domains,
                "auto_corrected": True,
            }
        if bad2:
            repaired = repair_factory_translation_openai_zh_id(src_text, result, reason2)
            if repaired:
                result = post_fix_factory_zh_to_id(src_text, repaired)
        fallback = factory_semantic_translate_zh_id(src_text)
        bad3, _, _ = detect_factory_semantic_error_zh_id(src_text, result)
        if bad3 and fallback:
            result = fallback
    return result


# === Factory semantic audit for Chinese -> Indonesian ===
FACTORY_ZH_ID_BAD_PATTERNS = {
    "dicuri": "factory_theft_literal",
    "mencuri": "factory_theft_literal",
    "pencuri": "factory_theft_literal",
    "derek QC": "factory_crane_qc_literal",
    "bereaksi": "factory_response_literal",
    "tertelan": "factory_swallowed_literal",
}

FACTORY_ZH_ID_POST_FIX = {
    "dicuri oleh derek QC": "dibawa oleh QC duluan",
    "dicuri oleh QC": "dibawa oleh QC duluan",
    "ditemukan dicuri": "ditemukan sudah dibawa duluan",
    "mencuri": "mengambil duluan",
    "dicuri": "dibawa duluan",
    "pencuri": "pengambilan duluan",
    "Anda perlu bereaksi": "harus segera lapor",
    "perlu bereaksi": "harus segera lapor",
    "harus bereaksi": "harus segera lapor",
    "bereaksi": "lapor",
    "akan tertelan setelah dibersihkan": "bisa hilang atau tertutup setelah dibersihkan",
    "tertelan setelah dibersihkan": "hilang atau tertutup setelah dibersihkan",
    "tertelan": "hilang atau tertutup",
    # v3.9.29: 廠區地名統一交給 post_fix 函式裡的 regex 處理(非詞庫)
    # 原因:詞庫機制會被觸發兩次,導致「(鹽水廠) (鹽水廠)」重複註解
    # 法規/管理術語(印尼工友更熟悉的講法)
    "Dinas K3": "Departemen Keselamatan Kerja (職安署)",
    "Biro K3": "Departemen Keselamatan Kerja (職安署)",
    "Otoritas K3": "Departemen Keselamatan Kerja (職安署)",
    "akan kena pengawasan": "akan diawasi ketat oleh",
    "kena pengawasan": "diawasi ketat oleh",
    # 罰單
    "tilang": "surat denda",  # tilang 是交通罰單,不是工廠罰單
    "kena 2 SP berat": "sudah dapat 2 surat peringatan berat",
    "kena 1 SP berat": "sudah dapat 1 surat peringatan berat",
    "SP berat": "surat peringatan berat",
    # interlock / bypass(技術術語)
    "interlock di-bypass": "interlock dipintas",
    "bypass interlock": "memintas interlock",
    "kena tangkap": "ketahuan",  # 抓到 ≠ 逮捕
    "tertangkap": "ketahuan",
    # 班組/管理階層
    "kepala shift patroli": "kepala regu patroli",
    "kepala shift": "kepala regu / 班長",
    "班長": "ketua shift / kepala regu",
    # v3.9.30d B22:夜點費 ≠ uang lembur malam(夜班加班費)的修補在 post_fix_factory_zh_to_id 內條件式處理
    # 不放這裡是因為合法的「夜班加班費」也會翻成 uang lembur malam,不能無條件改
}


# v3.9.27: 台灣→印尼專有名詞翻譯前置詞庫(在 prompt 中強化)
TAIWAN_FACTORY_GLOSSARY_HINT = """
重要術語對照(中文 → 印尼文):
- 鹽水廠 → Pabrik Yanshui (鹽水廠)
- 台中廠 → Pabrik Taichung (台中廠)
- 冷精棒冷抽課 → Bagian Cold-Drawn Bar
- 職安署 → Departemen Keselamatan Kerja (職安署)
- 職災 → Kecelakaan kerja serius
- 罰單 → Surat denda / Surat peringatan
- SP / 警告單 → Surat Peringatan
- 重大職災 → Kecelakaan kerja parah
- 列管 → Diawasi ketat
- interlock → Interlock (保留原英文)
- bypass → Memintas / Di-bypass
- 班長 → Ketua shift / Kepala regu
- 副總 → Wakil Direktur / Vice President
- 帽扣 → Tali helm / Strap helm
- 違規作業 → Pelanggaran prosedur kerja
- 違規操作 → Pelanggaran operasi
- 工安 → Keselamatan kerja
- 巡視設備 → Inspeksi peralatan / Patroli mesin
- 記過 → Surat peringatan / Catatan pelanggaran
"""


def detect_factory_semantic_error_zh_id(src_text, id_text):
    """Detect Chinese->Indonesian factory semantic errors that normal confidence misses."""
    if not id_text:
        return True, "empty", []
    src = src_text or ""
    low = id_text.lower()
    factory_src = any(k in src for k in ["料", "品保", "清洗", "研磨", "進料", "刮傷", "吊", "偷跑", "工單", "包裝", "站別"])
    if not factory_src:
        return False, "", []
    domains = ["factory"]
    if any(k in src for k in ["品保", "刮傷", "異常", "損傷"]):
        domains.append("quality_issue")
    if any(k in src for k in ["清洗", "研磨", "吊", "偷跑", "進料"]):
        domains.append("material_flow")
    if "偷跑" in src and any(x in low for x in ["dicuri", "mencuri", "pencuri"]):
        return True, "factory_theft_literal:偷跑不可翻成偷竊", domains
    if "吊" in src and "derek qc" in low:
        return True, "factory_crane_qc_literal:品保不是吊車", domains
    if "反應" in src and "bereaksi" in low:
        return True, "factory_response_literal:反應應為回報/通報", domains
    if "吃掉" in src and "tertelan" in low:
        return True, "factory_swallowed_literal:被吃掉不可直譯吞掉", domains
    for bad, reason in FACTORY_ZH_ID_BAD_PATTERNS.items():
        b = bad.lower()
        if b in low:
            if b in ("dicuri", "mencuri", "pencuri") and "偷跑" not in src:
                continue
            if b == "bereaksi" and "反應" not in src:
                continue
            if b == "tertelan" and "吃掉" not in src:
                continue
            return True, reason + ":" + bad, domains
    return False, "", domains


def post_fix_factory_zh_to_id(src_text, id_text):
    """Fix literal Indonesian outputs for Taiwan factory Chinese->ID translation.
    
    v3.9.29: 擴大 trigger keyword 範圍。之前只有「料/品保/清洗/研磨」等 11 個工廠製程詞,
    導致「鹽水廠/職安署/警告/重大職災/interlock」這類安全/法規類訊息**完全沒套用詞庫**。
    現在加入廠區地名、安全術語、法規詞,確保詞庫能命中 v3.9.27 加的全部規則。
    
    v3.9.29 修補 8: 加入 idempotent 處理 — 廠區名「Yanshui」可能已經被 GPT 翻成
    「Pabrik Yanshui」或還是「Yanshui」,要避免重複套用「(鹽水廠)」標註。
    """
    if not id_text:
        return id_text
    src = src_text or ""
    result = id_text.strip()
    # v3.9.29: 大幅擴充 trigger 詞庫
    factory_keywords = [
        # 製程(原本就有)
        "料", "品保", "清洗", "研磨", "進料", "刮傷", "吊", "偷跑", "工單", "包裝", "站別",
        # v3.9.29 新增:廠區/地名
        "鹽水廠", "鹽水", "台中廠", "冷精棒", "冷抽課",
        # v3.9.29 新增:安全/法規
        "職安", "工安", "職災", "重大職災", "警告", "罰單", "記過", "違規", "列管", "停工",
        # v3.9.29 新增:設備/英文術語
        "interlock", "bypass", "PMI", "MSDS", "SOP", "PPE", "LOTO", "K3",
        # v3.9.29 新增:管理/操作
        "班長", "副總", "巡視", "操作", "作業", "抓到", "帽扣",
        # v3.9.29 新增:OCR 圖片常見的訊息類用語
        "公告", "通知", "提醒", "今年", "目前",
        # v3.9.30d B22 新增:薪資相關(夜點費誤譯案例)
        "夜點", "夜點費", "日點費", "夜班津貼", "中班津貼",
        "薪資", "薪水", "薪", "減項", "加項", "扣項",
        "獎金", "底薪", "本薪", "勞保", "健保", "夜班費",
    ]
    # 也檢查譯文(if 原文沒命中但譯文有 Yanshui / Dinas K3 等典型錯翻)
    id_trigger_words = ["Yanshui", "Dinas K3", "tilang", "kena tangkap", "kepala shift",
                        "bereaksi", "dicuri", "tertelan", "interlock di-bypass"]
    factory_src = any(k in src for k in factory_keywords)
    factory_id = any(k in result for k in id_trigger_words)
    if not factory_src and not factory_id:
        return result
    # v3.9.29 修補 8/10: idempotent 處理 — 避免「(鹽水廠) (鹽水廠)」重複註解
    # Step 1: 把已經完整的版本暫存,套用詞庫時不會被重複處理
    placeholders = []
    def _stash(match):
        placeholders.append(match.group(0))
        return f"\x00PH{len(placeholders)-1}\x00"
    # 先把已完整的詞藏起來
    result = re.sub(r"Pabrik Yanshui\s*\(鹽水廠\)", _stash, result, flags=re.I)
    result = re.sub(r"Pabrik Taichung\s*\(台中廠\)", _stash, result, flags=re.I)
    result = re.sub(r"Departemen Keselamatan Kerja\s*\(職安署\)", _stash, result, flags=re.I)
    
    # Step 2: 套用詞庫(由長到短,避免短詞先吃掉長詞的部分)
    for wrong, correct in sorted(FACTORY_ZH_ID_POST_FIX.items(), key=lambda x: -len(x[0])):
        result = re.sub(re.escape(wrong), correct, result, flags=re.I)
    
    # Step 3: 處理 standalone Yanshui — 沒中文標註的 Yanshui 都升級成完整名稱
    # 不管前面有沒有 "Pabrik",只要後面沒接 "(" 就升級
    # 「Pabrik Yanshui sudah」→「Pabrik Yanshui (鹽水廠) sudah」(前面 Pabrik 被吃掉,重新加)
    # 「Yanshui sudah」→「Pabrik Yanshui (鹽水廠) sudah」
    # 「Pabrik Yanshui (鹽水廠)」→ 不變(已被 stash)
    result = re.sub(
        r"(?:Pabrik\s+|pabrik\s+)?\bYanshui\b(?!\s*\()",
        "Pabrik Yanshui (鹽水廠)",
        result
    )
    result = re.sub(
        r"(?:Pabrik\s+|pabrik\s+)?\bTaichung\b(?!\s*\()",
        "Pabrik Taichung (台中廠)",
        result
    )
    
    # Step 4: 把 placeholder 還原
    for i, ph_text in enumerate(placeholders):
        result = result.replace(f"\x00PH{i}\x00", ph_text)
    
    # v3.9.30d B22 修補:夜點費條件式翻譯
    # 「夜點費」在台灣勞基法 = 因輪到夜班領的固定津貼,不論加班(uang shift malam)
    # 「夜班加班費」才是 uang lembur malam
    # GPT 常把兩者混淆,這裡只在原文明確含「夜點費」/「日點費」/「夜班津貼」時修
    if "夜點費" in src or "夜班津貼" in src or "夜點" in src:
        # 把譯文裡的 uang lembur malam 改成 uang shift malam
        result = re.sub(r'\buang\s+lembur\s+malam\b', 'uang shift malam', result, flags=re.I)
        # 處理變體:uang lembur shift malam / lembur shift malam(連 uang 一起包以免重複)
        result = re.sub(r'\b(?:uang\s+)?lembur\s+shift\s+malam\b', 'uang shift malam', result, flags=re.I)
        # 處理已經是 shift 但前面誤加 lembur 的:lembur uang shift malam
        result = re.sub(r'\blembur\s+uang\s+shift\s+malam\b', 'uang shift malam', result, flags=re.I)
    if "日點費" in src or "日班津貼" in src:
        result = re.sub(r'\buang\s+lembur\s+siang\b', 'uang shift siang', result, flags=re.I)
        result = re.sub(r'\b(?:uang\s+)?lembur\s+shift\s+siang\b', 'uang shift siang', result, flags=re.I)
    if "中班津貼" in src:
        result = re.sub(r'\buang\s+lembur\s+sore\b', 'uang shift sore', result, flags=re.I)
        result = re.sub(r'\b(?:uang\s+)?lembur\s+shift\s+sore\b', 'uang shift sore', result, flags=re.I)
    
    result = re.sub(r"\s+", " ", result).strip()
    return result


def factory_semantic_translate_zh_id(text):
    """Deterministic Chinese->Indonesian factory translation for high-risk known shapes."""
    src = text or ""
    compact = re.sub(r"\s+", "", src)
    if all(k in compact for k in ["清洗前", "料", "品保"]) and ("偷跑" in compact or "吊去" in compact) and "刮傷" in compact:
        return (
            "Kalau material sebelum dicuci ditemukan sudah dibawa QC duluan, harus segera lapor. "
            "Sebelum dicuci, perhatikan dulu apakah material masuk ada goresan; "
            "kalau tidak, setelah dicuci goresannya bisa hilang atau tertutup."
        )
    if "偷跑" in compact and "品保" in compact:
        return "Kalau material sudah dibawa QC duluan tanpa konfirmasi, harus segera lapor."
    if "吊去" in compact and "品保" in compact:
        return "Material sudah dibawa oleh QC, harus segera dikonfirmasi."
    return None


def build_factory_context_hint_zh_id(text):
    src = text or ""
    if not any(k in src for k in ["料", "品保", "清洗", "研磨", "進料", "刮傷", "吊", "偷跑", "工單", "包裝", "站別"]):
        return ""
    return (
        "【繁中→印尼工廠語義提示】這是台灣不鏽鋼棒材工廠群組訊息，不可逐字翻。"
        "料/進料=material/bahan masuk；品保=QC；清洗=di-cuci/dibersihkan依現場語氣；"
        "吊去=被吊走/移走/帶走，譯為 dibawa/diangkat，不可譯成 dicuri；"
        "偷跑=未照正常流程先拿走/先做/先跑，譯為 dibawa/diproses duluan tanpa konfirmasi，不是偷竊；"
        "反應=回報/通報，譯為 lapor/beri tahu，不是 bereaksi；"
        "被吃掉=痕跡被清洗/加工後消失或被蓋掉，譯為 hilang/tertutup，不是 tertelan。"
    )


def repair_factory_translation_openai_zh_id(src_text, bad_result, reason):
    if not oai:
        return None
    try:
        deterministic = factory_semantic_translate_zh_id(src_text)
        sys_prompt = (
            "你是台灣不鏽鋼棒材工廠的繁體中文→印尼文現場翻譯審核器。"
            "你的任務是修正工廠語義直譯錯誤，輸出印尼員工現場看得懂的自然印尼文。"
            "不要解釋，不要加註解，只輸出修正後譯文。"
            "規則：偷跑不是偷竊；吊去不是被吊車偷；反應是回報/通報；被吃掉是痕跡消失/被蓋掉。"
        )
        user_msg = (
            f"原中文：{src_text}\n"
            f"錯誤印尼文：{bad_result}\n"
            f"錯誤原因：{reason}\n"
            f"語義提示：{build_factory_context_hint_zh_id(src_text)}\n"
        )
        if deterministic:
            user_msg += f"可採用譯文：{deterministic}\n"
        user_msg += "請輸出修正後印尼文："
        # v3.9.8: model_supports() filter
        repair_kwargs = {
            "model": pick_model(src_text),
            "messages": [{"role":"system","content":sys_prompt},{"role":"user","content":user_msg}],
        }
        _rmodel = repair_kwargs["model"]
        if model_supports(_rmodel, "temperature"):
            repair_kwargs["temperature"] = 0.0
        if model_supports(_rmodel, "top_p"):
            repair_kwargs["top_p"] = 1.0
        # v3.9.30b B11 修補: reasoning model 加 reasoning 預算
        if model_supports(_rmodel, "max_completion_tokens"):
            _is_reasoning = not model_supports(_rmodel, "temperature")
            repair_kwargs["max_completion_tokens"] = 500 + (4000 if _is_reasoning else 0)
        elif model_supports(_rmodel, "max_tokens"):
            repair_kwargs["max_tokens"] = 500
        _opt = optimal_reasoning_for_translation(_rmodel)
        if model_supports(_rmodel, "reasoning_effort") and _opt:
            repair_kwargs["reasoning_effort"] = _opt
        if model_supports(_rmodel, "verbosity"):
            repair_kwargs["verbosity"] = "low"
        try:
            if bad_result and model_supports(_rmodel, "stop"):
                repair_kwargs["prediction"] = {
                    "type": "content",
                    "content": bad_result,
                }
            _ck_r = _build_cache_key(getattr(_tl, 'group_id', ''), "zh", "id", "repair")
            if _ck_r and model_supports(_rmodel, "prompt_cache_key"):
                repair_kwargs["prompt_cache_key"] = _ck_r
        except Exception:
            pass
        r = oai.chat.completions.create(**repair_kwargs)
        track_tokens(r)
        return (r.choices[0].message.content or "").strip()
    except Exception as e:
        logger.error("Factory zh-id repair error: %s", e)
        return None

# === Hard replacement tables ===
# These bypass GPT entirely - applied BEFORE sending to GPT (zh->id)
# and AFTER receiving from GPT (id->zh result post-processing)

ZH_TO_ID_HARD = {
    # 製程/站別
    "爐號標籤": "label heat number",
    "爐號": "heat number",
    "無心研磨": "centerless grinding",
    "光輝退火爐": "furnace bright annealing",
    "光輝退火": "bright annealing",
    "退火爐": "tungku annealing",
    "過帳": "input data ke sistem",
    "放行": "release data",
    # 品質/缺陷
    "殺光痕": "bekas grinding mark",
    "車刀痕": "bekas pisau bubut",
    "砂光痕": "bekas sanding mark",
    "軋輥印痕": "bekas roll mark",
    "環狀擦傷": "goresan melingkar",
    "表粗": "surface roughness",
    "直度": "kelurusan",
    "偏小": "under size",
    "偏大": "over size",
    "風險批": "lot berisiko",
    "走ET檢測": "jalankan pengujian ET",
    "開立重工": "buat work order rework",
    "不允收": "pelanggan tidak terima",
    # 設備
    "矯直機": "mesin straightening",
    "壓光機": "mesin press polish",
    "砂光機": "mesin sanding",
    "拋光機": "mesin polishing",
    "眼模": "die/cetakan",
    "引拔座": "drawing bench",
    "皮膜槽": "coating tank",
    "氣壓缸": "silinder pneumatik",
    "安全圍籬": "safety fence",
    "集塵設備": "dust collector",
    "計長器": "length counter",
    "冷水機": "chiller",
    "馬蹄環": "shackle",
    "吊掛物": "beban gantung",
    "護罩": "pelindung mesin",
    "interlock": "pengunci keamanan",
    "標籤機": "mesin label",
    "台車": "troli",
    "天車": "crane",
    # v3.9.30: 移除原本緊接著的 "台車"/"天車" 完全重複定義
    # v3.9.30d B22 修補: 薪資/津貼類術語(夜點費誤譯案例)
    # 台灣勞基法上「夜點費」≠「夜班加班費」:
    #   夜點費 = 因輪到夜班而領的固定津貼,不論有沒有加班
    #   加班費 = 超過正常工時才有
    # GPT 之前誤翻成 "uang lembur malam"(夜班加班費),語意有偏差
    "夜點費": "uang shift malam",
    "夜班津貼": "uang shift malam",
    "日點費": "uang shift siang",
    "日班津貼": "uang shift siang",
    "中班津貼": "uang shift sore",
    "夜班費": "uang shift malam",
    "減項": "potongan",
    "加項": "tunjangan tambahan",
    "扣項": "potongan",
    "獎金": "bonus",
    "全勤獎金": "bonus kehadiran penuh",
    "績效獎金": "bonus kinerja",
    "三節獎金": "bonus hari raya",
    "底薪": "gaji pokok",
    "本薪": "gaji pokok",
    "薪資單": "slip gaji",
    "扣稅": "potong pajak",
    "勞保": "BPJS Ketenagakerjaan",
    "健保": "BPJS Kesehatan",
    # 管理
    "品保": "QC",
    "儲運": "bagian gudang",
    "生計": "production planning",
    "業務": "bagian sales",
    "營業": "bagian sales",
    "人事": "HRD",
    "處長": "kepala divisi",
    "稼動率": "utilization rate",
    "線速": "kecepatan lini",
    "速差": "selisih kecepatan",
    "主機手": "operator utama",
    "印勞": "pekerja Indonesia",
    "加壓": "tambah tekanan",
    "電流": "arus listrik",
    "在製品管制表": "tabel kontrol WIP",
    # 包裝/入庫
    "套紙管": "pasang tabung kertas",
    "太空包": "jumbo bag",
    "噴漆罐": "kaleng spray",
    "木箱": "kotak kayu",
    "櫃子": "kontainer",
    # 訂單
    "允收": "toleransi terima",
    "訂尺": "panjang pesanan",
    "短尺": "ukuran pendek",
    "異型棒": "batang bentuk khusus",
    "遞延單": "order ditunda",
    "急單": "order urgent",
    "不擋非本月": "order bukan bulan ini boleh masuk gudang",
    "不擋": "tidak dibatasi",
    "溢量": "kelebihan produksi",
    "併包": "gabung packing",
    "出貨差": "kekurangan pengiriman",
    # HR/紀律
    "忘卡補": "input lewat sistem lupa kartu",
    "造冊": "buat daftar absensi",
    "班股": "rapat shift",
    "堆高機複訓": "pelatihan ulang forklift",
    "天車複訓": "pelatihan ulang crane",
    "扣績效": "potong penilaian kinerja",
    "劣項": "pelanggaran",
    "納入劣項": "dicatat pelanggaran",
    "提報懲處": "laporkan untuk sanksi",
    "三定": "3 tetap",
    "不要物": "barang tidak terpakai",
    "被釘": "kena tegur",
    "綠卡": "kartu hijau",
    # 環境
    "煙蒂": "puntung rokok",
    "檳榔渣": "sisa pinang",
    "廚餘": "sisa makanan",
    "漏油": "bocor oli",
    "積水": "genangan air",
    "粉塵": "debu",
    # 口語
    "感溫": "terima kasih",
    "有夠": "sangat",
    "母湯": "jangan",
}

# Post-replacement: fix common GPT mistakes in output
ID_POST_FIX = {
    # 爐號 corrections
    "nomor panas": "heat number",
    "label nomor panas": "label heat number",
    "nomor tungku": "heat number",
    "label nomor tungku": "label heat number",
    "nomor oven": "heat number",
    "label nomor oven": "label heat number",
    # 有包到 corrections
    "paket datang ke": "kalau ada packing untuk",
    "saat paket datang ke": "kalau ada packing untuk",
    "Mohon diperhatikan saat paket datang ke": "Nanti kalau ada packing untuk",
    "Mohon diperhatikan saat kalau ada packing untuk": "Nanti kalau ada packing untuk",
    # 三米六米 corrections
    "tiga meter di atas enam meter": "batang 3 meter ditaruh di atas batang 6 meter",
    "Tiga meter di atas enam meter": "Batang 3 meter ditaruh di atas batang 6 meter",
    "3 meter di atas 6 meter": "batang 3 meter ditaruh di atas batang 6 meter",
    # 放=PUT/PLACE asking WHO: must use menaruh/meletakkan, NOT bare taruh (too colloquial)
    # These fix "siapa yang taruh X" → "siapa yang menaruh X"
    "siapa yang taruh": "siapa yang menaruh",
    "Siapa yang taruh": "Siapa yang menaruh",
    "yang taruh batang": "yang menaruh batang",
    "yang taruh material": "yang meletakkan material",
    "yang taruh barang": "yang menaruh barang",
    "yang taruh kotak": "yang menaruh kotak",
    # 品保 corrections
    "jaminan kualitas": "QC",
    "penjaminan mutu": "QC",
    # 點名 corrections (NOT roll call)
    "panggilan nama": "inspeksi pengawas",
    "absen nama": "inspeksi pengawas",
    "roll call": "inspeksi pengawas",
    # 感溫 - should not be translated literally
    "suhu perasaan": "terima kasih",
    "merasakan suhu": "terima kasih",
    # Common GPT errors
    "Polymetal": "寶麗金屬",
    "Bao Li Metal": "寶麗金屬",
    "Bao Li Logam": "寶麗金屬",
    "Changzhou Zhongshan": "常州眾山",
    "Da Shun": "大順",
    "Da Cheng": "大成",
    "Bei Ze": "北澤",
    "Hong Yun": "鴻運",
    "Tian Hua Rong": "田華榕",
    "Jia Dong": "佳東",
    # 營業 common mistranslation
    "bagian operasional": "bagian sales",
    "operasional perlu": "sales perlu",
}

# Customer names - protect from translation by wrapping
# Storage area lookup data (from 儲區查詢.xlsx)
_STORAGE_JSON = '{"6C422209": [["<=3200", "EH28"], [">4200", "EG38"], [">3200<=4200", "EH26"]], "ABE": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "AIK": [[">3200<=4200", "EG14"], [">4200", "EH33"], ["<=3200", "EH28"]], "ALCONIX JP": [["<=3200", "EG14"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "AMERICAN STAINLESS": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "AMS": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "ANCHOR": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "ANIL METALS": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "APEX METAL": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "AWACS": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "B&B": [[">4200", "EH33"], ["<=3200", "EH22"], [">3200<=4200", "EG14"]], "B&J": [["<=3200", "EC40"], [">4200", "EC40"], [">3200<=4200", "EC45"]], "BOBCO": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH34"]], "BOLLINGHAUS": [[">3200<=4200", "EC43"], ["<=3200", "EC43"], [">4200", "EC43"]], "CA-ASD": [[">4200", "EH11"], ["<=3200", "EH12"], [">3200<=4200", "EH12"]], "CA-AUSTRAL": [[">3200<=4200", "EH12"], ["<=3200", "EH12"], [">4200", "EH11"]], "CA-DALSTEEL": [[">4200", "EH11"], ["<=3200", "EH12"], [">3200<=4200", "EH12"]], "CA-FLETCHER": [[">3200<=4200", "EH12"], [">4200", "EH11"], ["<=3200", "EH28"]], "CA-M&S": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-MICO": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-MIDWAY": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-S&T": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-VAN LEEUWEN": [["<=3200", "EH12"], [">4200", "EH11"], [">3200<=4200", "EH12"]], "CA-VES": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-VULCAN": [[">4200", "EH11"], ["<=3200", "EH12"], [">3200<=4200", "EH12"]], "CA-VULCAN NZ": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-WAKEFIELD": [[">4200", "EH11"], [">3200<=4200", "EH12"], ["<=3200", "EH12"]], "CAMELLIA": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "CASTLE": [[">3200<=4200", "EH12"], ["<=3200", "EH28"], [">4200", "EH11"]], "CHANDAN": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "CHANG HSIN": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "CHANGSU": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "COGNE AOSTA": [[">3200<=4200", "EG34"], ["<=3200", "EH28"], [">4200", "EG14"]], "COGNE CELIK": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "COGNE DE": [[">3200<=4200", "EG14"], [">4200", "EH34"], ["<=3200", "EH28"]], "COGNE DG": [[">3200<=4200", "EC47"], ["<=3200", "EC47"], [">4200", "EC41"]], "COGNE FR": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "COGNE KR": [["<=3200", "EH26"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "COGNE UK": [[">3200<=4200", "EG14"], [">4200", "EH34"], ["<=3200", "EH28"]], "COMINOX": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "COMPRINOX": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "CSMU": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "DACAPO": [["<=3200", "EH25"], [">3200<=4200", "EG14"], [">4200", "EH31"]], "DACAPO-K STOCK": [["<=3200", "EH25"], [">3200<=4200", "EG14"], [">4200", "EH31"]], "DAECHANG": [[">4200", "EG34"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "DAMSTAHL": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "DAVER": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "DK METAL": [[">4200", "EG35"], [">3200<=4200", "EG14"], ["<=3200", "EC47"]], "DUFU": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "EGMO": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "EIAM": [[">3200<=4200", "EG14"], ["<=3200", "EG14"], [">4200", "EH33"]], "ESP": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "EURO STEEL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "FASTENAL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "FINE METAL TRADE": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "FSS": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "G HWA": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "GIC": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "GLH": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "GS METAL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "HADCO": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "HAKUDO": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "HAMATECH": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "HANWA": [["<=3200", "EH21"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "HEAP SING HUAT": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "HH": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "HRMETAL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH34"]], "HUA GUAN METAL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "HWA GUAN METAL": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "IM": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "INTEGRITY STAINLESS": [[">4200", "EG34"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "IPE": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "ISE": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "IWATANI": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "JANG ANN": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "JFE SHOJI": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "KANGRUI": [[">3200<=4200", "EG14"], [">4200", "EC45"], ["<=3200", "EH28"]], "KANSAI": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "KDK": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "KIAN": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "KIM ANN": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "KJ": [[">4200", "EG32"], ["<=3200", "EC47"], [">3200<=4200", "EG14"]], "KJ PRECISION": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "KOMINOX AB": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "LAI KING": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "LAURIE": [["<=3200", "EG14"], [">3200<=4200", "EH28"], [">4200", "EH33"]], "LE": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "LEE & STEEL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "LIM MENG SENG": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "LINSTER": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "LOTUS METAL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "LTM": [["<=3200", "EG15"], [">3200<=4200", "EG15"], [">4200", "EG34"]], "M.R. STEEL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "MAINCHAIN": [[">4200", "EG34"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "MAN TAK": [[">4200", "EG34"], ["<=3200", "EG15"], [">3200<=4200", "EG15"]], "MARINE": [["<=3200", "EG14"], [">3200<=4200", "EH28"], [">4200", "EH33"]], "MCB": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "MENAM": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "METAL ESTABLISH": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "METALINOX": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "METALLSERVIS": [[">3200<=4200", "EH14"], ["<=3200", "EH28"], [">4200", "EG35"]], "NAKAYAMA": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "NDE": [["<=3200", "EH28"], [">4000", "EG34"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "NM": [[">3200<=4200", "EG14"], ["<=3200", "EH22"], [">4200", "EG34"]], "NMSK": [[">4200", "EG34"], ["<=3200", "EH28"]], "NOVA TRADING": [["<=3200", "EH27"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "NOXFAP": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "NS METAL": [[">3200<=4200", "EG14"], ["<=3200", "EG14"], [">4200", "EH18"]], "NSC": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "OKAYA": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EH14"]], "OLYMPIC STEEL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "OME": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "PACKER(ISRAEL)": [["<=3200", "EH28"], [">4200", "EG34"], [">3200<=4200", "EH14"]], "PASCAL": [[">3200<=4200", "EH14"], ["<=3200", "EH28"], [">4200", "EG34"]], "PF": [["<=3200", "EH28"], [">3200<=4200", "EH14"], [">4200", "EG34"]], "PLUTUS": [[">3200<=4200", "EI30"], ["<=3200", "EI25"], [">4200", "EI40"]], "PRECISION": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EH14"]], "PRECISION METAL": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EH14"]], "PRECISION METALS": [[">3200<=4200", "EH14"], ["<=3200", "EH28"], [">4200", "EH33"]], "QPLUS": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "RAAJRATNA": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "RHS": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "RINO": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH34"]], "RISEBM": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "SAGAMI": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "SAMWON": [["<=3200", "EC47"], [">3200<=4200", "EG14"], [">4200", "EG32"]], "SCM": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "SCOT": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EI40"]], "SD-BK": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "SD-BKL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "SD-KHS": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "SD-LIM METAL": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "SD-METALPHILE": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "SD-METHA": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "SD-TPS": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "SENG HUAT": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "SENG HUAT METALPLEX": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "SGH": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "SHIMIZU MATERIAL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH34"]], "SHINKO": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH34"]], "SHINKO TH": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH34"]], "SING LEONG-雙馬": [["<=3200", "EH28"], [">4200", "EH34"], [">3200<=4200", "EG14"]], "SLA": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "SMG": [["<=3200", "EH28"], [">4200", "EG33"], [">3200<=4200", "EG14"]], "SPECTROMATRIX": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "STEELINC": [["<=3200", "EH28"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "STEWART": [[">3200<=4200", "EG14"], [">4200", "EH33"], ["<=3200", "EH28"]], "STIRLINGS": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH34"]], "STIRLINGS(5%)": [[">3200<=4200", "EG14"], [">4200", "EH34"], ["<=3200", "EH28"]], "STKSTAINLESS": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "STRONG STEEL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH34"]], "SUNGEUN": [[">4200", "EG33"], ["<=3200", "EG37"], [">3200<=4200", "EG14"]], "SUNGSIL METAL": [[">4200", "EG35"], ["<=3200", "EC47"], [">3200<=4200", "EG14"]], "SUPERFIX": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "SUPREME": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "TAN VIET": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "TCI": [["<=3200", "EH32"], [">3200<=4200", "EH32"], [">4200", "EH32"]], "TEKPOINT": [[">3200<=4200", "EG14"], ["<=3200", "EG14"], [">4200", "EG34"]], "TITAN METALS": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "TK-SCHULTE": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH22"]], "TKMP": [[">3200<=4200", "EG14"], [">4200", "EH34"], ["<=3200", "EH26"]], "TMC": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "TOP SUNNY": [["<=3200", "EH28"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "TOZZHIN THAILAND": [["<=3200", "EH28"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "TSA": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "TSM": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "TUBE SUPPLY": [[">4200", "EG34"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "TUSCO": [[">3200<=4200", "EG15"], [">4200", "EG34"], ["<=3200", "EH28"]], "WESCO": [[">4200", "EG34"], [">3200<=4200", "EG15"], ["<=3200", "EH28"]], "WEST COAST": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "WING KEUNG": [[">3200<=4200", "EG14"], [">4200", "EH33"], ["<=3200", "EH29"]], "WPS": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "YGS": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "YIEH CORP LTD(HK)": [["<=3200", "EH28"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "YONGTA": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "YOSHU": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "YOUCHANG": [[">4200", "EG34"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "YOUNG DONG": [[">3200<=4200", "EG15"], ["<=3200", "EG15"], [">4200", "EH33"]], "？頂": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "？暉": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "力常(觀音)": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "三大興": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "三卯鍛壓": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "三利": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "上晉": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "上海凡斯": [["<=3200", "EC47"], [">4200", "EC40"], [">3200<=4200", "EC45"]], "上海坤成": [["<=3200", "EC47"], [">3200<=4200", "EC40"], [">4200", "EC40"]], "上海億科": [[">3200<=4200", "EC40"], [">4200", "EC40"], ["<=3200", "EC47"]], "上海町芃": [["<=3200", "EH10"], [">4200", "EH10"], [">3200<=4200", "EH10"]], "上銀": [["<=3200", "EH99"], [">4200", "EC40"], [">3200<=4200", "EH99"]], "凡立": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "千里眼": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "大甲永和": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "大成": [[">4200", "EH32"], [">3200<=4200", "EH32"], ["<=3200", "EH32"]], "大連德邁仕": [["<=3200", "EC47"], [">3200<=4200", "EC47"], [">4200", "EC40"]], "大順": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "中國防蝕": [[">4200", "EH35"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "元盈": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "元偉勝": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "升暘": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EG38"]], "天津隆德": [[">4200", "EC40"], ["<=3200", "EC47"], [">3200<=4200", "EC40"]], "方鉦": [[">3200<=4200", "EH72"], [">4200", "EH72"], ["<=3200", "EH79"]], "世廷": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "世華": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "功億": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "北澤": [[">4200", "EG38"], [">3200<=4200", "EG39"], ["<=3200", "EG39"]], "北澤一廠": [["<=3200", "EG39"], [">3200<=4200", "EG39"], [">4200", "EG38"]], "北澤二廠": [[">4200", "EG38"], ["<=3200", "EG39"], [">3200<=4200", "EG38"]], "北澤三廠": [["<=3200", "EG39"], [">3200<=4200", "EG38"], [">4200", "EG38"]], "右勝鋼鐵": [[">3200<=4200", "EH78"], ["<=3200", "EG39"], [">4200", "EH71"]], "台芝": [[">4200", "EH10"], ["<=3200", "EH10"], [">3200<=4200", "EH10"]], "台灣亞錁": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "台灣林吉": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "台灣矽微": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "巨昌": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "巨頻": [[">3200<=4200", "EG14"], [">4200", "EG38"], ["<=3200", "EH79"]], "永川泰": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "永村": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "生計直棒": [[">3200<=4200", "EH99"], ["<=3200", "EH99"], [">4200", "EH99"]], "生計庫存": [[">3200<=4200", "EH99"], [">4200", "EH99"], ["<=3200", "EH99"]], "禾桀": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH38"]], "光翔": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "全利金屬": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "全敏尖端": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "向春": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "名威": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "合順": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "宇隆": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "宇慶": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "有光": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "江陰外庫": [["<=3200", "EC47"], [">3200<=4200", "EC47"], [">4200", "EC40"]], "江陰華新": [[">4200", "EC40"], [">3200<=4200", "EC40"], ["<=3200", "EC47"]], "江蘇迪威": [[">4200", "EC40"], [">3200<=4200", "EC47"], ["<=3200", "EC47"]], "汎新": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "百呈": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "百堅": [[">3200<=4200", "EG37"], [">4200", "EH33"], ["<=3200", "EG14"]], "西邁金屬": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH33"]], "君立": [["<=3200", "EH79"], [">4200", "EH36"], [">3200<=4200", "EH78"]], "壯安": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "宏盈": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "宏荃": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "志典": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH33"]], "志聯": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "甫剛": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "貝加": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "貝克休斯": [[">3200<=4200", "EG38"], [">4200", "EG38"], ["<=3200", "EG38"]], "京碼": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "京鋼": [[">4200", "EC41"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "佳東": [[">3200<=4200", "EH76"], ["<=3200", "EH76"], [">4200", "EH70"]], "佳東-台中": [[">4200", "EH70"], ["<=3200", "EH76"], [">3200<=4200", "EH78"]], "佳東-台北": [[">4200", "EH70"], [">3200<=4200", "EH78"], ["<=3200", "EH76"]], "佳東-高雄": [[">3200<=4200", "EH78"], [">4200", "EH70"], ["<=3200", "EH76"]], "協崎": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "坤泰": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "奇賓": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EG38"]], "孟駿": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "尚智": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "岡山東穎": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "承總": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "易隆": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "昆山金富盈": [[">3200<=4200", "EC40"], ["<=3200", "EC47"], [">4200", "EC40"]], "明石": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "東栗": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "東莞峰作": [["<=3200", "EC47"], [">3200<=4200", "EC40"], [">4200", "EC40"]], "東萊": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "東徽": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH33"]], "武漢機械": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "金大": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "金利山": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "金亞洲": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "金城": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "金耘": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH71"]], "金耘-南營所": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH71"]], "金煜": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH71"]], "長盈": [[">4200", "EG38"], [">3200<=4200", "EG14"], ["<=3200", "EH79"]], "長圓": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "俊來(蘆洲)": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "俊益": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "厚群": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "威孚高科技": [["<=3200", "EC47"], [">3200<=4200", "EC47"], [">4200", "EC40"]], "建新": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "柏緯": [[">3200<=4200", "EC47"], [">4200", "EC42"], ["<=3200", "EC47"]], "津展": [["<=3200", "EH75"], [">4200", "EH71"], [">3200<=4200", "EH71"]], "津展-台中": [[">4200", "EH72"], ["<=3200", "EH75"], [">3200<=4200", "EH72"]], "津展-台北": [[">3200<=4200", "EH72"], [">4200", "EH72"], ["<=3200", "EH75"]], "津展-台南": [[">3200<=4200", "EH72"], ["<=3200", "EH75"], [">4200", "EH72"]], "皇銘": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "研發部": [["<=3200", "EH99"], [">3200<=4200", "EH99"], [">4200", "EH99"]], "研發測試": [["<=3200", "EH99"], [">3200<=4200", "EH99"], [">4200", "EH99"]], "科威聯": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "英鈿": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "重慶九勝": [[">4200", "EC40"], ["<=3200", "EC47"], [">3200<=4200", "EC47"]], "重慶九環": [[">4200", "EC40"], ["<=3200", "EC47"], [">3200<=4200", "EC40"]], "展舵": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "峰作金屬": [["<=3200", "EH74"], [">3200<=4200", "EH78"], [">4200", "EH71"]], "峰勝": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "振家": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "振華興": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "時哲": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "晉易": [[">4200", "EH38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "晉椿": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH38"]], "晉椿(鹿港)": [[">4200", "EH38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "浙江三花": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "益陽": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "退庫重工": [[">4200", "EH99"], [">3200<=4200", "EH99"], ["<=3200", "EH99"]], "高立熱處理": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "高銪": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "商旺": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "域鑫科技": [[">4200", "EC40"]], "常州眾山": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EC43"]], "強淞": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "強實": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "捷流": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "淳康": [[">3200<=4200", "EH10"], ["<=3200", "EH10"], [">4200", "EH10"]], "眾山": [[">3200<=4200", "EH78"], [">4200", "EG35"], ["<=3200", "EH79"]], "祥日達": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "祥英": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "笠源": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "頂翔勝": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "麥億": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "備料庫存": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "凱記": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "勝初": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "勝新": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "勝盟": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "富億鑫": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "尊茂": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "復盛應用": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH78"]], "敦壹": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "朝盟": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "無錫永雋": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "舜欽": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "華友(外)": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH34"]], "華纜": [[">4200", "EH10"], [">3200<=4200", "EH10"], ["<=3200", "EH10"]], "詠勗": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "詠晟": [["<=3200", "EH79"], [">3200<=4200", "EC47"], [">4200", "EC40"]], "進達": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "開滋": [["<=3200", "EG39"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "隆明": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "隆門": [[">4200", "EG38"], [">3200<=4200", "EH28"], ["<=3200", "EH28"]], "隆順發": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "雅信億": [[">3200<=4200", "EH14"], ["<=3200", "EH79"], [">4200", "EH33"]], "廉喬": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "廉錩": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH77"]], "廉錩-台北": [[">3200<=4200", "EH78"], ["<=3200", "EH77"], [">4200", "EG38"]], "廉錩-台南": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH77"]], "慈溪龍華": [[">4200", "EC40"], [">3200<=4200", "EC40"], ["<=3200", "EC47"]], "新創捷": [["<=3200", "EH79"], [">3200<=4200", "EH14"], [">4200", "EH33"]], "新華特聯": [["<=3200", "EH79"], [">3200<=4200", "EH14"], [">4200", "EH35"]], "新萊應材": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "瑞鋼": [[">4200", "EC40"], [">3200<=4200", "EC40"], ["<=3200", "EC47"]], "盟鉦": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "萬揚": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EH33"]], "經捷": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "經貿": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "群鎰": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "聖泰": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "路竹新益": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "鉅泰昇": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH36"]], "鉅銅": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "鉅豐": [["<=3200", "EH79"]], "鼎崴": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "嘉冠": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "嘉碁": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "寧波東葛": [[">3200<=4200", "EC40"], ["<=3200", "EC47"], [">4200", "EC40"]], "慷倫": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "睿緻佳": [["<=3200", "EH78"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "福泉": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "聚祥": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "銓宥": [[">4200", "EH10"], [">3200<=4200", "EH10"], ["<=3200", "EH10"]], "廣泰": [[">4200", "EH10"], [">3200<=4200", "EH10"], ["<=3200", "EH10"]], "慶鋐": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "歐承": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "毅鋼": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "磐石": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "誼山": [["<=3200", "EH29"], [">3200<=4200", "EH78"], [">4200", "EH38"]], "頭份": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "優普洛": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "營三備庫(內)": [[">3200<=4200", "EC40"], ["<=3200", "EC47"], [">4200", "EC40"]], "營三備庫(外)": [[">3200<=4200", "EC40"], [">4200", "EC40"], ["<=3200", "EC47"]], "營業庫存": [[">4200", "EH99"], ["<=3200", "EH99"], [">3200<=4200", "EH99"]], "環友": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "聯岱": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "聯祥": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "邁達斯": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "鴻運": [[">3200<=4200", "EH27"], ["<=3200", "EH27"], [">4200", "EG38"]], "雙和": [[">3200<=4200", "EG14"], ["<=3200", "EH26"], [">4200", "EG34"]], "麒譯": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "町洋": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "晟田": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "畯圓": [["<=3200", "EH19"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "鐿順發": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "鑫誠鐵材": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "恒耀": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "暉": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "頂": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]]}'
STORAGE_LOOKUP = json.loads(_STORAGE_JSON)
# Try loading from storage_data.json (auto-updated via admin panel)
_storage_json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "storage_data.json")
if os.path.exists(_storage_json_path):
    try:
        with open(_storage_json_path, "r", encoding="utf-8") as _f:
            STORAGE_LOOKUP = json.load(_f)
            logger.info("Loaded storage data from storage_data.json: %d customers", len(STORAGE_LOOKUP))
    except Exception as _e:
        logger.warning("Failed to load storage_data.json, using embedded: %s", _e)
# Try loading packaging_data.json
_packaging_json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "packaging_data.json")
if os.path.exists(_packaging_json_path):
    try:
        with open(_packaging_json_path, "r", encoding="utf-8") as _f:
            PACKAGING_LOOKUP = json.load(_f)
            logger.info("Loaded packaging data: %d codes", len(PACKAGING_LOOKUP))
    except Exception as _e:
        logger.warning("Failed to load packaging_data.json: %s", _e)
# Extra customers not in storage Excel but appear in factory chat
# Per-group protected names: {"__all__": [...], "group_id": [...]}
_DEFAULT_NAMES = [
    "寶麗金屬", "田華榕", "蘋果", "賽利金屬", "盛昌遠", "曜麟",
    "LOTUS", "LOTUS METAL", "shinko", "wing keung",
    "高侑", "十元", "小麥", "啊堂", "秋情", "政軒", "碩凱", "汶錡",
    "武駿", "凱銘", "小趙", "阿澤", "法比恩", "山多", "EggEgg", "fang", "Dato潘",
    "阿添", "小叮噹", "多啦A夢", "潘柏良", "大彭",
]
extra_names_by_group = {"__all__": list(_DEFAULT_NAMES)}
EXTRA_CUSTOMERS = []

def rebuild_customer_names():
    """Rebuild EXTRA_CUSTOMERS and CUSTOMER_NAMES from all groups."""
    global EXTRA_CUSTOMERS, CUSTOMER_NAMES
    merged = set()
    for names in extra_names_by_group.values():
        merged.update(names)
    EXTRA_CUSTOMERS = sorted(list(merged), key=lambda x: -len(x))
    CUSTOMER_NAMES = sorted(list(set(list(STORAGE_LOOKUP.keys()) + EXTRA_CUSTOMERS)), key=lambda x: -len(x))

CUSTOMER_NAMES = []
rebuild_customer_names()


def pre_replace_zh(text):
    """Apply hard replacements to Chinese text before GPT translation.
    Returns (modified_text, customer_placeholders_dict)."""
    result = text
    # Protect customer names with placeholders (these survive GPT translation)
    cust_ph = {}
    for i, name in enumerate(CUSTOMER_NAMES):
        if name in result:
            ph = f"__CUST_{i}__"
            cust_ph[ph] = name
            result = result.replace(name, ph)
    # Apply hard replacements (longest first to avoid partial matches)
    for zh, replacement in sorted(ZH_TO_ID_HARD.items(), key=lambda x: -len(x[0])):
        if zh in result:
            result = result.replace(zh, f"[{replacement}]")
    return result, cust_ph


def restore_customers(text, cust_ph):
    """Restore customer name placeholders back to original names."""
    if not text or not cust_ph:
        return text
    result = text
    for ph, name in cust_ph.items():
        # GPT might mangle placeholders, try variants
        idx = ph.replace("__CUST_", "").replace("__", "")
        variants = [
            ph, ph.replace("_", " "), f"CUST_{idx}", f"CUST {idx}",
            f"__CUST {idx}__", f"[CUST_{idx}]",
        ]
        for v in variants:
            if v in result:
                result = result.replace(v, name)
    # Safety: if any customer name placeholder pattern remains, try regex
    result = re.sub(r'__CUST_(\d+)__', lambda m: cust_ph.get(f"__CUST_{m.group(1)}__", m.group(0)), result)
    return result


def post_fix_translation(text):
    """Fix known GPT translation mistakes in output."""
    if not text:
        return text
    result = text
    # Fix specific wrong translations (longest match first)
    for wrong, correct in sorted(ID_POST_FIX.items(), key=lambda x: -len(x[0])):
        result = result.replace(wrong, correct)
    # Remove bracketed hints that leaked through from pre_replace
    result = re.sub(r'\[([a-zA-Z /&]+)\]', r'\1', result)
    # Clean up double spaces (preserve newlines)
    result = re.sub(r'[^\S\n]+', ' ', result)
    result = re.sub(r'\n{3,}', '\n\n', result)
    result = result.strip()
    return result


def translate_openai(text, src, tgt, strict_no_source_script=False, repair_mode=False, bad_result=None):
    if not oai:
        return None
    # v3.9.30 B10 修補: 空字串/純空白不該送到 OpenAI 浪費 token
    if not text or not (text or "").strip():
        return ""
    try:
        src_name = LANG_NAMES.get(src, src)
        tgt_name = LANG_NAMES.get(tgt, tgt)

        # Apply hard replacements before GPT for zh->other
        input_text = text
        cust_placeholders = {}
        if src == "zh":
            input_text, cust_placeholders = pre_replace_zh(text)

        protected, placeholders = protect_mentions(input_text)

        extra_rule = ""
        if strict_no_source_script and src != tgt:
            if src == "zh":
                extra_rule = (
                    " 10. IMPORTANT: Do not leave any Chinese words untranslated unless they are a person's name or __MENTION__ placeholder."
                    " Terms such as 印籍, 印尼籍, 早班, 夜班, 考試, 讀書, 下班後 must be translated into the target language."
                )
            elif src == "ja":
                extra_rule = " 10. IMPORTANT: Do not leave Japanese text untranslated unless it is a person's name or __MENTION__ placeholder."
            elif src == "ko":
                extra_rule = " 10. IMPORTANT: Do not leave Korean text untranslated unless it is a person's name or __MENTION__ placeholder."
            elif src == "th":
                extra_rule = " 10. IMPORTANT: Do not leave Thai text untranslated unless it is a person's name or __MENTION__ placeholder."

        # Get tone from thread-local (set by handler before calling translate)
        _tone = getattr(_tl, 'tone', 'casual')
        _tone_custom = getattr(_tl, 'tone_custom', '')
        # v3.2-0426d: tone_custom is ADDITIVE (appends to preset) instead of REPLACING.
        _preset_text = TONE_PRESETS.get(_tone, TONE_PRESETS['casual'])
        if _tone_custom and _tone_custom.strip():
            tone_instruction = _preset_text + " 【額外語氣指令（補充微調，衝突時以下方為準）】 " + _tone_custom.strip()
        else:
            tone_instruction = _preset_text

        sys_prompt = (
            "You are a professional translator for a stainless steel factory (Walsin Lihwa/華新麗華, Yanshui plant) work group chat. "
            "This factory produces stainless steel bars, wire rods, peeled bars, cold-drawn bars using processes like rolling, annealing, pickling, peeling, cold drawing, and centerless grinding. "
            "This is a group with Taiwanese managers and Indonesian migrant workers operating centerless grinding (無心研磨) equipment. "
            "CRITICAL RULES: "
            "1. NEVER translate @mentions and NEVER translate or romanize person names. Keep all Chinese names in ORIGINAL CHINESE CHARACTERS. "
            "For example: 徐嘉騰 stays as 徐嘉騰, NOT Xu Jiateng. 陳弘林 stays as 陳弘林, NOT Chen Honglin. "
            "Chinese nicknames for people must stay unchanged. Do NOT translate them literally. "
            "2. Any text like __MENTION_0__, __MENTION_1__ etc are placeholders - keep them exactly as is. "
            "3. TRANSLATION TONE/STYLE: " + tone_instruction + " "
            + build_factory_context_hint(text, src, tgt) + " "
            # ★ v3.4 CoD:對 ID→ZH 注入高風險詞字典
            + ((detect_id_zh_risk_terms(text) + " ") if (id_zh_cod_enabled and src == "id" and tgt == "zh") else "")
            # ★ v3.4 CoT:對 ID→ZH 注入二階段思考指令
            + ((build_id_zh_cot_instruction(text) + " ") if (id_zh_cot_enabled and src == "id" and tgt == "zh") else "") +
            "4. Indonesian slang: gak=tidak, udah=sudah, gimana=bagaimana, bgt=banget, org=orang, yg=yang, tdk=tidak, dg=dengan, krn=karena, blm=belum, hrs=harus, bs=bisa, lg=lagi, gw=saya, lu=kamu. "
            "5. TAIWANESE MANDARIN COLLOQUIAL (very important): "
            "乾/干=aduh/astaga, 靠=astaga/waduh, 幹=sial/buset, 傻眼=gak percaya, 扯/誇張=keterlaluan, 笑死=ngakak, 氣死=kesel banget, 累死=capek banget, "
            "啦=lah/dong, 喔/哦=ya/lho, 耶=dong/nih, 嘛=dong/kan, 蛤=hah?/apa?, 厚=ya kan, "
            "醬/降=begitu/gitu(=這樣), 母湯=jangan/gak boleh(=不要), 超/有夠=banget(=非常), 感溫=terima kasih(台語感恩), "
            "CRITICAL: Taiwanese rhetorical questions SUGGEST doing something: 需不需要X=perlu X gak nih(suggesting X should be done), 要不要X=gimana kalau X, 還在X=masih X(often implies criticism). "
            "搞什麼=ngapain sih, 搞定=beres, 人咧=orangnya mana, 怎麼搞的=kenapa bisa begini, 出包=ada masalah, 先這樣=segitu dulu ya, 再說=nanti aja, "
            "X到不行/X得要死/X到爆=X banget, 怎麼這麼X=kok X banget, 有夠X=X banget, "
            "ㄏㄏ=haha, QQ=sedih, 3Q=terima kasih, GG=tamat, XD=haha, @@=bingung. "
            "6. Target Traditional Chinese = Taiwan style, not mainland. "
            "6.5 PARAGRAPH STRUCTURE (CRITICAL — most users complain about this when translating images): "
            "**You MUST preserve the original paragraph structure EXACTLY.** "
            "Rules: "
            "(a) Every blank line (\\n\\n) in source = blank line in translation, in the SAME position. "
            "(b) Every newline (\\n) in source = newline in translation. "
            "(c) DO NOT merge paragraphs. DO NOT remove blank lines. DO NOT add extra blank lines. "
            "(d) If source has timestamps or speaker labels (e.g., '13:48', '潘柏良'), keep them on their own lines. "
            "Example: source '段A\\n\\n段B\\n\\n段C' MUST translate to "
            "'translation_A\\n\\ntranslation_B\\n\\ntranslation_C', NOT 'translation_A translation_B translation_C'. "
            "Test: count the blank lines in source. The translation MUST have the same count. "
            # v3.9.27: 台灣工廠專用術語對照表(印尼工友常用講法)
            "6.6 TAIWAN FACTORY TERMS (use these Indonesian translations, do NOT keep Chinese place names untranslated): "
            "鹽水廠 → Pabrik Yanshui (鹽水廠) [keep Chinese in parens for clarity]; "
            "台中廠 → Pabrik Taichung (台中廠); "
            "冷精棒冷抽課 → Bagian Cold-Drawn Bar; "
            "職安署 → Departemen Keselamatan Kerja (職安署) [NOT 'Dinas K3']; "
            "重大職災 → Kecelakaan kerja parah; "
            "罰單/警告單 → Surat Peringatan (SP); "
            "列管 → Diawasi ketat; "
            "interlock → Interlock (keep English); "
            "bypass interlock → Memintas interlock (NOT 'di-bypass'); "
            "班長 → Ketua shift / Kepala regu; "
            "副總 → Wakil Direktur; "
            "帽扣 → Tali helm; "
            "違規作業 → Pelanggaran prosedur kerja; "
            "違規操作 → Pelanggaran operasi; "
            "工安 → Keselamatan kerja; "
            "巡視設備 → Inspeksi peralatan; "
            "記過 → Catatan pelanggaran; "
            "抓到 → Ketahuan (NOT 'kena tangkap' which means arrested by police); "
            "7. Target Indonesian = simple clear daily language for factory workers. "
            "8. Context: factory work - shifts, overtime, orders, tasks, meals, breaks, meetings, exams. "
            "9. FACTORY VOCABULARY: "
            "【製程/Process】"
            "無心研磨=centerless grinding, 研磨=grinding, 砂輪=batu gerinda, 調整輪=roda pengatur, 刀板=work rest blade, 冷卻液=cairan pendingin, "
            "不鏽鋼=stainless steel, 棒鋼=steel bar, 盤元=wire rod, 削皮棒=peeled bar, 冷精棒=cold-drawn bar, "
            "熱軋=hot rolling, 退火=annealing, 酸洗=pickling, 削皮=peeling, 冷抽=cold drawing, "
            "鋼種=jenis baja, PMI=uji material, 來料=material masuk, 棒材=batang baja, 混料=tercampur material(SERIOUS), 料號=nomor material, "
            "拋光=polishing, 粗拋=rough polishing, 噴漆=spray paint, 洗料=cuci material, "
            "倒角=chamfer, 修磨=repair grinding, 盤元修磨=repair grinding wire rod, 線外修磨=offline repair grinding, "
            "壓光=press polish, 矯直=straightening, 重矯=straightening ulang, 精整=finishing, AP=mesin finishing, "
            "光輝退火=bright annealing, 回爐=kirim kembali ke furnace, "
            "側磨=side grinding(DILARANG/prohibited), 不可側磨=dilarang side grinding, "
            "【站別/Stations - numbers are STATION NUMBERS】"
            "400站=station 400, 401站=station 401, 420站=station 420, "
            "470站=station 470(UT station), UT=mesin UT(di station 470), 480站=station 480, "
            "490站=station 490(秤重站/timbang), 801站=station 801, "
            "OL=sedang produksi/online, 回400=kembalikan ke station 400, "
            "無主=tanpa pemilik/unassigned, 入無主=masukkan ke status tanpa pemilik, "
            "掛單/工單=work order, 重掛單=pasang ulang work order, 無工單資訊=tidak ada info work order, "
            "改制=ubah proses, 去化=ada order baru mau terima, 有單去化=ada order baru untuk serap material, 改制去化=ubah proses produksi, "
            "帳/帳務=data administrasi(ERP), 帳已回400=data sudah dikembalikan ke station 400, "
            "過帳=input data produksi(jumlah&berat)ke sistem tanpa release ke stasiun berikutnya, "
            "放行=release data ke stasiun berikutnya(setelah QC lulus), "
            "退庫=kembalikan ke gudang, 退庫拆包=keluarkan dari gudang bongkar packing untuk dibagi ulang, "
            "發料=issue material, 存檔=simpan data, 暫存=simpan sementara, 短尺=ukuran pendek, "
            "溢量=kelebihan produksi melebihi permintaan, 併包=gabung packing dari lot berbeda dalam order sama, "
            "出貨差=kekurangan pengiriman hari ini, 轉用=dialihkan untuk order lain, 跳無主轉用=pindah ke tanpa pemilik lalu dialihkan, "
            "【班次/出勤】"
            "點名=ada pengawas yang datang(inspection, NOT roll call), 早班=shift pagi, 夜班=shift malam, 中班=shift siang, "
            "加班=lembur, 排班=jadwal shift, 調班=tukar shift, 上班=masuk kerja, 下班=pulang kerja, 打卡=absen, "
            "請假=izin, 病假=izin sakit, 事假=izin pribadi, 特休=cuti tahunan, 代班=gantikan shift, "
            "忘卡補=lupa kartu ID, pakai sistem input waktu, 造冊=buat daftar absensi, "
            "班股=rapat shift, 堆高機複訓=pelatihan ulang forklift, 天車複訓=pelatihan ulang crane, "
            "紅包=angpao, 年終獎金=bonus akhir tahun, 過年不停機=Imlek tidak berhenti produksi, "
            "【產線/設備】"
            "產線=lini produksi, 機台=mesin, 開機=nyalakan mesin, 停機=mesin berhenti, 調機=setting mesin, "
            "上料=isi material, 備料=siapkan material, 產量=jumlah produksi, 目標=target, 達標=capai target, 超產=over production, "
            "訂單=order, 出貨=kirim barang, 交期=deadline, 趕貨=kejar order, 急單=order urgent, 急單備註=catatan order urgent, "
            "下製程=proses selanjutnya, 異常=abnormal/ada masalah, 維修中=sedang diperbaiki, "
            "天車=overhead crane, 台車=trolley, 吊秤=timbangan gantung, 馬蹄環=shackle, 鋼索=sling baja, 吊掛物=beban gantung, "
            "稼動率=utilization rate, 線速=line speed(m/min), 限速=batas kecepatan, 降速=turunkan kecepatan, 提速=naikkan kecepatan, 速差=selisih kecepatan, "
            "撥料=feed material, 線外=offline, 印勞=pekerja Indonesia, "
            "砂光機=sanding machine, 眼模=die/cetakan drawing, 引拔座=drawing bench, 皮膜槽=coating tank, "
            "查修=investigasi&perbaiki, 修護=maintenance, 儀電=instrumen listrik, 備品=spare part, "
            "跳異常=error muncul, 復歸=reset, 復歸無效=reset gagal, 跳機=mesin trip, 恢復生產=kembali produksi, "
            "叫修=panggil teknisi, 進廠查修=teknisi masuk pabrik cek, 電聯儀電=hubungi instrumen listrik, "
            "斷料=material putus, 卡料=material macet, 擠料=material terjepit keluar, "
            "主機手=operator utama, 上料人員=petugas pengisian material, 點檢=cek rutin, 護罩=pelindung mesin/safety guard, "
            "interlock=pengunci keamanan(jangan ditahan pakai benda), "
            "【印尼文機械/設備詞彙 Indonesian Mechanical Terms】"
            "as=軸/軸心(axle/shaft), as roda=輪軸(wheel axle), roda=輪(wheel), roda penarik=拉料輪(pulling wheel), "
            "penarik barang=拉料車/拖料車(material puller/cart), "
            "kopel=萬向接頭(universal joint/coupling), cross joint=十字接頭(cross joint), "
            "as roda penarik barang patah=拉料輪的萬向接頭斷裂(pulling wheel universal joint broken), "
            "patah=斷了/斷裂(snapped/broken off), bengkok=彎了(bent), "
            "retak=裂了(cracked), aus=磨損(worn out), bocor=漏(leak), macet=卡住(jammed), "
            "bearing=軸承(bearing), rantai=鏈條(chain), sabuk=皮帶(belt), engsel=鉸鏈(hinge), "
            "kawat=鋼線/線材(wire), selang=軟管(hose), katup=閥門(valve), baut=螺栓(bolt), mur=螺帽(nut), "
            "tekanan=壓力(pressure), getaran=震動(vibration), gesekan=摩擦(friction), pelumas=潤滑油(lubricant), "
            "gigi/gear=齒輪(gear), kipas=風扇(fan), kipas angin=電風扇(electric fan), motor=馬達(motor), "
            "pompa=泵浦(pump), kompresor=壓縮機(compressor), pipa=管(pipe), tabung=鋼瓶/桶(tank/cylinder), "
            "dongkrak=千斤頂(jack), kunci=扳手/鑰匙(wrench/key), obeng=螺絲起子(screwdriver), tang=鉗子(pliers), "
            "las=焊接(welding), gerinda=砂輪機(grinder), bor=鑽孔機(drill), gergaji=鋸子(saw), "
            "forklift=堆高機(forklift), crane=吊車(crane), conveyor=輸送帶(conveyor), "
            "NOTE: 'As' in Indonesian mechanical context ALWAYS means axle/shaft(軸), never translate as 'as/像'. "
            "NOTE: 'patah' means snapped/broken off(斷了), different from 'rusak'(壞了/故障). "
            "【包裝/入庫】"
            "套紙管=pasang tabung kertas, 入庫=masuk gudang, 優先包裝入庫=prioritas packing masuk gudang, "
            "需求單=formulir permintaan, 可以全收=bisa diterima semua, "
            "櫃子=kontainer(shipping container), 櫃子在路上=kontainer sedang di jalan, "
            "木箱=kotak kayu, 裝箱=masukkan ke kotak kayu, 2700大的木箱=kotak kayu ukuran besar 2700mm, "
            "NOTE: 木箱 context: 3200/2400=box LENGTH mm, 500/1000=weight CAPACITY kg. "
            "把=bundel(bundle), 捆=bundel/ikat, 支/根=batang(piece/rod), 批=lot/batch, "
            "NOTE: X米(三米,六米)=batang X meter(bar LENGTH not distance). 三米上面放六米=batang 3m ditaruh di atas batang 6m. "
            "包(verb)=packing/kemas(NOT wrapping). 秤重=timbang, 貼標=tempel label, 綁鐵=ikat besi, "
            "【訂單管理】"
            "允收=jumlah yang boleh diterima pelanggan, 允收0支=zero tolerance, 不收短尺=tidak terima ukuran pendek, "
            "訂尺=panjang sesuai pesanan, 爐號=heat number/nomor furnace(NEVER translate as 'panas'), 爐號標籤=label heat number, "
            "分捆=pisah bundel, 遞延單=delayed order, 非本月=bukan order bulan ini, "
            "非本月不入庫=order bukan bulan ini jangan masuk gudang, 檔非本月=tahan order bukan bulan ini, "
            "異型棒=batang bentuk khusus, 異型棒不擋=batang khusus tidak dibatasi, "
            "不擋=tidak dibatasi/boleh masuk(exemption), 不擋非本月=order bukan bulan ini BOLEH masuk gudang(exception/exemption, NOT blocked), "
            "入庫目標=target masuk gudang, 壓日期=ada deadline ketat, "
            "管控=kontrol, 不管控=tidak dikontrol(bebas), "
            "【品質/缺陷】"
            "品保=QC, 會驗=joint inspection, 暫留=hold sementara, HOLD=tahan, "
            "客訴=komplain pelanggan, 夾帶樣品=sertakan sampel, 掛檔=simpan ke arsip, 稽核=audit, "
            "螺紋=thread mark, 車刀痕=turning tool mark, 砂光痕=sanding mark, 殺光痕=grinding mark, "
            "剝片=flaking, 軋輥印痕=roll mark, 碰傷=luka benturan, 黑皮=unfinished surface, "
            "偏小=under size, 偏大=over size, 直度=kelurusan(straightness), 表粗=surface roughness, 目視=visual inspection, "
            "開立重工=buat WO rework, 重工研磨至尺寸下限=rework grinding sampai batas bawah ukuran, "
            "不允收=pelanggan tidak terima, 風險批=lot berisiko, 走ET檢測=jalankan pengujian ET, "
            "卡料需關閉電源後再取料=material macet HARUS matikan listrik dulu baru ambil, "
            "【部門/人員】"
            "業務=sales, 營業=sales(=業務), 生計=production planning, 資訊=IT department, 品保=QC, 儲運=gudang&logistik, 人事=HRD, 工安=safety officer, "
            "處長=kepala divisi, 抓資料=ambil data, "
            "【標籤/系統】"
            "TAG=label, 儲區=area penyimpanan di sistem, 轉檔=konversi data, "
            "MES=MES(sistem produksi), 報表=laporan produksi, 條碼=barcode, "
            "標籤機=mesin label, 包裝電腦=komputer packing, "
            "在製品管制表=WIP control sheet, "
            "【安全/環境/紀律】"
            "太空包=jumbo bag/FIBC, 噴漆罐一定要打洞才能丟棄在太空包=kaleng spray HARUS dilubangi sebelum buang ke jumbo bag, "
            "扣績效=potong kinerja(sanksi), 劣項=pelanggaran, 納入劣項=dicatat pelanggaran, "
            "三定=3 tetap(tempat/barang/jumlah tetap), 不要物=barang tidak terpakai, "
            "漏油=bocor oli, 生鏽=berkarat, 掉漆=cat mengelupas, 積水=genangan air, 粉塵=debu, "
            "煙蒂=puntung rokok, 檳榔渣=sisa pinang, 被釘=kena tegur atasan, "
            "提報懲處=laporkan untuk sanksi, 會嚴罰=dihukum berat, "
            "綠卡=kartu hijau(catatan safety), KYT=pelatihan prediksi bahaya, 防火演練=latihan pemadam kebakaran, "
            "調班單=formulir tukar shift, 簽核=tanda tangan persetujuan, "
            "【生活/薪資】"
            "宿舍=asrama, 便當=bekal makan, 餵狗=kasih makan anjing, "
            "薪水=gaji, 加班費=uang lembur, 績效=penilaian kinerja, 匯款=transfer, "
            # v3.9.30d B22 修補:夜點費 ≠ 夜班加班費
            # 夜點費 = 因輪到夜班領的「固定津貼」(不論加班),譯 uang shift malam
            # 加班費 = 工時超出才有,譯 uang lembur
            "夜點費=uang shift malam(NOT uang lembur,夜點費是固定夜班津貼非加班費), "
            "日點費=uang shift siang, 夜班津貼=uang shift malam, 中班津貼=uang shift sore, "
            "底薪=gaji pokok, 本薪=gaji pokok, 全勤獎金=bonus kehadiran penuh, "
            "績效獎金=bonus kinerja, 三節獎金=bonus hari raya, "
            "減項=potongan, 加項=tunjangan tambahan, 扣項=potongan, 獎金=bonus, "
            "薪資單=slip gaji, 扣稅=potong pajak, 勞保=BPJS Ketenagakerjaan, 健保=BPJS Kesehatan, "
            "尾牙=pesta akhir tahun, 春酒=pesta tahun baru, 伴手禮=oleh-oleh, 便當費=biaya makan siang, "
            "量測=mengukur, 尺寸=diameter, 公差=toleransi, 校正=kalibrasi, "
            "【客戶 - NEVER translate】"
            "DACAPO, CASTLE, LOTUS, METALINOX, KANGRUI, SUNGEUN, STEELINC, GLH, shinko, wing keung, "
            "田華榕, 佳東, 蘋果, 常州眾山, 大順, 大成, 巨昌, 北澤, 鴻運, 畯圓, 名威, 右勝, 貝克休斯, 皇銘, "
            "台芝, 百堅, 津展, 曜麟, 廉錩, 盛昌遠, 永吉, 光輝, 寶麗金屬. "
            "NOTE: 蘋果=customer NOT fruit. 光輝=customer OR 光輝退火(bright annealing), context determines. "
            "10. CRITICAL CONTEXT RULES: "
            "a) X米(三米,六米)=bar LENGTH. 三米上面放六米=batang 3m ditaruh di atas batang 6m. "
            "b) 把/捆=BUNDLE counters. 包2把=packing 2 bundel. "
            "c) 包(verb)=packing NOT wrapping. 高侑的今天包2把都這樣=Yang di-packing 高侑 hari ini 2 bundel semuanya kayak gini. "
            "d) Names(" + ",".join(EXTRA_CUSTOMERS) + ")=keep as-is. "
            "e) Customer names=keep as-is, do NOT translate. "
            "f) R+number=round bar diameter(R28.57=bulat 28.57mm). Non-R=hex/special(H26=hex 26mm). "
            "g) S/B=straight bar. E1~E11=cold drawing lines. I1~I21=grinding machines. BF2/3/5=polishing machines. "
            "h) 5F/5L/6S/6T/6U/6W/7E/7F/7G+numbers=work order ID, keep as-is. "
            "i) 課料=section chief designated material. G包=packing method code. AP=finishing equipment. "
            "j) 爐號=heat number(NEVER 'nomor panas'). 有包到X=kalau ada packing untuk X(NOT 'paket datang ke X'). "
            "k) 放=POLYSEMY(multiple meanings, judge by context): "
            "放+把/單/批/工單號/這把/這單/這批=RELEASE data(放行). e.g. 先放這把=release bundel ini dulu, 放了=sudah di-release, 幫放一下=tolong bantu release. "
            "放+地點/方位(地上/旁邊/上面/那邊/架上)=PUT/PLACE. e.g. 放地上=taruh di lantai, 三米上面放六米=batang 3m ditaruh di atas 6m. "
            "放+這些/這批/這個+物體(棒材/料/箱/東西) WITHOUT 地點 + 詢問責任/歸屬(誰放的/誰放的啊/哪個人放的)=PUT/PLACE asking WHO placed it. "
            "CRITICAL: 在這種「誰放的」句型必須用 menaruh 或 meletakkan(formal/written), 不可用 taruh(太口語). "
            "e.g. 這些棒材誰放的=Ada yang tahu siapa yang menaruh batang-batang ini? "
            "這批料誰放這裡的=Siapa yang meletakkan material ini di sini? "
            "這個箱子誰放的=Siapa yang menaruh kotak ini? "
            "放+料/材料(without location)=FEED material. e.g. 放料=isi material. "
            "放假=libur/holiday. "
            "IMPORTANT taruh vs menaruh/meletakkan: taruh(base form) 只適合祈使句「叫某人把東西放哪」(e.g. taruh di sana=放那邊). "
            "當主詞是人問責任(誰放的/誰擺的)→ 必須用 menaruh 或 meletakkan, 不能用裸 taruh. "
            "When ambiguous and context is about work orders or production flow, default to RELEASE(放行). "
            "l) 再=POLYSEMY: "
            "X再Y(condition+action)=hanya X yang Y / X baru Y(=才). e.g. 急單再幫忙安排入庫=hanya order urgent yang tolong bantu atur masuk gudang. "
            "再+verb(without preceding condition)=lagi/sekali lagi(=again). e.g. 再確認一下=confirm sekali lagi. "
            "m) 非本月=bukan order bulan ini(order that is NOT for the current month). 非本月包裝不入庫=yang bukan order bulan ini jangan packing masuk gudang. "
            "n) 不擋=tidak dibatasi/boleh masuk(EXEMPTION, means ALLOWED). 不擋非本月=order bukan bulan ini BOLEH masuk gudang. "
            "CRITICAL: 不擋 means NOT blocked = ALLOWED. Do NOT translate as tidak boleh(=blocked). "
            "e.g. DACAPO不擋非本月=DACAPO order bukan bulan ini boleh masuk gudang. "
            "o) When H、S appear in a list with 異型棒 or customer names, they are SEPARATE product categories(H=hex bar, S=straight bar). "
            "Keep them as individual items with commas. e.g. H、S異型棒=H, S, batang bentuk khusus(three separate types). "
            "p) ELLIPTICAL QUANTITY REPLIES: Chinese speakers often reply with JUST '數字+量詞' as a short answer, omitting the noun. "
            "e.g. Q:『要幾台?』A:『兩台』(=兩台台車=two troli, NOT 'two units'). "
            "Default mapping: 兩台→dua buah(generic, NEVER 'dua unit'); 三把→tiga bundel; 5支→5 batang; 一個→satu buah; 兩件→dua potong. "
            "CRITICAL: NEVER use 'unit' to translate 台/個/件 unless the original Chinese contains '單位'. "
            "q) 台車=troli(hard replacement applied). 削皮=peeling(hard replacement applied). "
            "r) APOLOGY DETECTION (CRITICAL - MISTRANSLATING APOLOGIES AS REQUESTS IS A SEVERE ERROR): "
            "Indonesian 'maaf' family ALWAYS means SORRY/APOLOGY: 對不起/抱歉/請原諒我. "
            "It NEVER means 麻煩你了 (=tolong ya, asking favor) — they are OPPOSITE in meaning. "
            "RECOGNIZE THESE SPELLING VARIANTS as the same apology: "
            "Maaf / Maafkan / Maafkan saya / Mafkan saya / Maf kan saya / Maf kan / Mafin / Mohon maaf / Minta maaf / Mhn maaf — ALL = 對不起. "
            "Workers commonly misspell as 'Maf kan' or 'Mafkan' — these are STILL apologies. "
            "🙏 emoji combined with maaf = sincere apology, NOT a polite request. "
            "When followed by 'kedepan akan lebih...' (以後會更...) or 'gak akan ulang lagi' (不會再犯) — DEFINITELY apology. "
            "EXAMPLES: "
            "Maaf kan saya 🙏 kedepan ya akan lebih teliti dalam bekerja → 對不起 🙏 以後我會更仔細地工作. "
            "Mohon maaf, saya salah → 真的很抱歉，是我的錯. "
            "Mhn maaf telat → 抱歉遲到了. "
            "WRONG: translating any maaf-form as 麻煩你了/麻煩了 — these are REQUEST phrases, OPPOSITE of apology. "
            "p) ELLIPTICAL QUANTITY REPLIES (very important for Taiwan factory chat): "
            "Chinese speakers often reply with JUST '數字+量詞' (number + measure word) as a short answer, OMITTING the noun that was mentioned earlier. "
            "e.g. Q:『需要幾台台車?』 A:『兩台』 (=兩台台車=two troli, NOT 'two units'). "
            "Q:『要幾把?』 A:『3把』 (=3 bundel). "
            "Q:『幾支?』 A:『5支』 (=5 batang). "
            "Because each message is translated independently WITHOUT conversation context, DO NOT invent or guess the noun. "
            "Translate these short replies with a GENERIC Indonesian quantity phrase that stays neutral: "
            "兩台/兩個/兩支/兩把/兩件 (when noun is omitted) → 'dua' or 'dua buah'(generic), NEVER 'dua unit'(too formal and wrong in casual chat). "
            "The word 'unit' in Indonesian suggests abstract units/modules and is WRONG for physical countable items like troli/batang/bundel. "
            "Default mapping for bare quantity replies: 台=buah(generic) or keep context-neutral, 把=bundel, 支=batang, 個=buah, 件=potong(for items/pieces). "
            "Examples: 兩台→dua buah. 三把→tiga bundel. 5支→5 batang. 一個→satu buah. 兩件→dua potong. "
            "CRITICAL: NEVER use 'unit' to translate 台/個/件 unless the original Chinese literally contains '單位' (unit as in department/organizational unit). "
            "q) 台車=troli(hard replacement already applied). 削皮=peeling(hard replacement already applied). These should appear in output as 'troli' and 'peeling' consistently. "
            + _build_custom_examples_prompt() +
            " 11. TRANSLATION EXAMPLES (follow strictly): "
            "【中→印尼】"
            "乾 需不需要提報一下 → Aduh, perlu dilaporkan gak nih? "
            "UT囤一堆料了 → UT udah numpuk banyak material. "
            "品保還在下班 誇張 → QC udah pulang, keterlaluan. "
            "三米上面放六米 → Batang 3 meter ditaruh di atas batang 6 meter. "
            "麻煩他們不要這樣放料 → Tolong bilang ke mereka jangan taruh material kayak gini. "
            "有人知道這些棒材誰放的嗎？ → Ada yang tahu siapa yang menaruh batang-batang ini? "
            "這批料誰放這裡的？ → Siapa yang meletakkan material ini di sini? "
            "這些東西誰放的？ → Siapa yang menaruh barang-barang ini? "
            "高侑的今天包2把都這樣 → Yang di-packing 高侑 hari ini 2 bundel semuanya kayak gini. "
            "來料都短少4-5公斤 → Material masuk semuanya kurang 4-5 kilogram. "
            "已轉達 → Sudah disampaikan. "
            "這批料有問題 → Lot material ini ada masalah. "
            "幫我盯一下 → Tolong awasin ya. "
            "怎麼搞的啦 → Kok bisa kayak gini sih. "
            "人咧 → Orangnya mana? "
            "辛苦了 → Makasih kerja kerasnya. "
            "靠 又壞了 → Astaga, rusak lagi. "
            "先這樣 → Segitu dulu ya. "
            "叫他快點 → Suruh dia cepatan. "
            "砂輪要換了 → Batu gerinda harus diganti. "
            "公差超過了 → Toleransinya udah lewat. "
            "這6把再麻煩今晚入庫 → 6 bundel ini tolong masukin gudang malam ini. "
            "明早業務要抓資料 謝謝 → Besok pagi sales perlu ambil data, makasih. "
            "BF2拋光機維修中 → Mesin polishing BF2 sedang diperbaiki. "
            "44.45前天有跟妳說超產，業務回覆了嗎 → Diameter 44.45 kemarin sudah bilang over produksi, sales udah balas belum? "
            "噴漆後照訂單量拆包 → Setelah spray paint, bagi packing sesuai jumlah order. "
            "品保點錯製程，麻煩退回400-無主 → QC salah pilih proses, tolong kembalikan ke station 400 tanpa pemilik. "
            "帳已回400、料要回去那一個單位？ → Data sudah dikembalikan ke 400, materialnya mau ke unit mana? "
            "去削皮退火 感溫 → Ke proses peeling dan annealing, makasih. "
            "削皮需要台車，再麻煩一下 → Bagian peeling butuh troli, tolong bantu ya. "
            "兩台 → dua buah. "
            "兩台台車 → dua troli. "
            "三把 → tiga bundel. "
            "5支 → 5 batang. "
            "需要兩台 → Butuh dua buah. "
            "再一台 → satu lagi. "
            "削皮需要台車，再麻煩一下 → Bagian peeling butuh troli, tolong bantu ya. "
            "兩台 → dua buah. "
            "兩台台車 → dua troli. "
            "要幾台？ → Perlu berapa? "
            "三把 → tiga bundel. "
            "5支 → 5 batang. "
            "一個 → satu buah. "
            "需要兩台 → Butuh dua buah. "
            "再一台 → satu lagi. "
            "7F414020 請幫放至480轉用收回400，要改制去化，謝謝 → 7F414020 tolong pindahkan ke station 480, lalu kembalikan ke 400, mau ubah proses, makasih. "
            "業務說收～ 請包～ → Sales bilang terima, tolong di-packing. "
            "班長～ 7F656502A 這把溢量請再入無主～ 謝謝! → Kepala shift, 7F656502A bundel ini kelebihan, tolong masukkan ke tanpa pemilik, makasih! "
            "客需求支數7支、不收短 來料只有6支、其中一支短、剔除掉剩5支、能包嘛？ → Pelanggan minta 7 batang, gak terima pendek. Masuk cuma 6, 1 pendek dibuang sisa 5, bisa packing gak? "
            "因為櫃子在路上 9點到 這樣可能可以等一下入庫 → Karena kontainer sedang di jalan, sampai jam 9, mungkin bisa tunggu sebentar baru masuk gudang. "
            "DACAPO都入完了 → DACAPO semuanya sudah masuk gudang. "
            "班長～ 請用2700大的木箱裝，再麻煩幫我抓一下幾點會好，業務下午要出，謝謝 → Kepala shift, tolong pakai kotak kayu 2700, cek jam berapa selesai, sales sore mau kirim, makasih. "
            "那就是帳沒入到 → Berarti datanya belum masuk ke sistem. "
            "資料異常，凱銘在處理了 → Data ada masalah, 凱銘 sedang urus. "
            "研磨排程已更新，急單再麻煩安排洗料拋光 謝謝 → Jadwal grinding diupdate, order urgent tolong atur cuci material dan polishing, makasih. "
            "粗拋完已放行 → Rough polishing selesai, sudah di-release. "
            "麻煩先放這把 → Tolong release bundel ini dulu. "
            "放了 → Sudah di-release. "
            "先放這單 → Release order ini dulu. "
            "幫放一下 → Tolong bantu release. "
            "這批先不要放 → Lot ini jangan di-release dulu. "
            "料放旁邊 → Material taruh di samping. "
            "放地上 → Taruh di lantai. "
            "今日出貨差 DACAPO 7G63837在490 7G687108A在420 OL → Hari ini pengiriman kurang: DACAPO 7G63837 di 490, 7G687108A di 420 sedang produksi. "
            "METALINOX 差2噸等等K4會在出料 可以的在幫包裝 感謝 → METALINOX kurang 2 ton, nanti K4 keluarkan material, kalau bisa tolong packing, makasih. "
            "7G108519D 請幫收回400，有單去化 謝謝 → 7G108519D tolong kembalikan ke 400, ada order baru untuk serap material, makasih. "
            "洗給E7拋了 → Sudah dicuci dan dikasih ke E7 untuk polishing. "
            "包裝遇到常州眾山再注意這個料號，剛接單後續才會投料生產，此訂單不收短尺需將短尺分捆 → Kalau packing ketemu 常州眾山 perhatikan nomor material ini, baru terima order nanti baru produksi, order ini gak terima pendek harus pisah bundel. "
            "剛剛開會決議過年不停機，如果A班D班出勤人數不夠12人，想賺紅包可以代班 → Rapat keputusan Imlek tidak stop, shift A D kurang 12 orang, mau angpao bisa gantikan shift. "
            "人事有通知堆高機複訓課程，1/29 1700-2000三樓會議室。當天來上課就好，加班時數改天用忘卡補 → HRD info pelatihan forklift, 29/1 jam 17-20 ruang rapat lt.3. Datang ikut aja, jam lembur diinput lewat sistem lupa kartu di hari lain. "
            "處長走了 → Kepala divisi sudah pergi. "
            "有壓日期的急單再幫忙處理一下，很多未到站，拋光會一邊產出 → Order urgent deadline tolong diproses, banyak belum sampai, polishing produksi sambil jalan. "
            # ===== v3.9.30 新增:積壓料量公告(配合班長公告反向訓練) =====
            "上個月積壓的30噸料還沒拋,本月又有46噸料到期要拋 → Masih ada 30 ton pemolesan yang jatuh tempo bulan lalu, dan 46 ton yang jatuh tempo bulan ini. "
            "拋光積壓20噸料 → Tunggakan polishing 20 ton. "
            "本月到期的料要優先處理 → Material yang jatuh tempo bulan ini harus diprioritaskan. "
            "拋光機人員一定要注意生產效率 → Personil mesin pemoles harus memperhatikan efisiensi produksi. "
            "研磨積壓的料還很多 → Material grinding yang tertunda masih banyak. "
            "噴漆罐一定要打洞才能丟棄在太空包，本週被查核兩次缺失 → Kaleng spray HARUS dilubangi baru buang ke jumbo bag, minggu ini kena audit 2 kali. "
            "本月入庫目標2950，異型棒不擋，其餘非本月不入庫 → Target gudang 2950, batang khusus bebas, sisanya bukan bulan ini jangan masuk. "
            "本月入庫目標量已達標，目前只入急單、異型棒跟二月以前的遞延單 → Target tercapai, sekarang hanya urgent, batang khusus, dan order ditunda sebelum Feb. "
            "今天沒點名，昨天來過了 → Hari ini gak ada inspeksi, kemarin sudah datang. "
            "應該是上週四D班，傍晚要注意一下小趙跟處長行蹤，免得凱銘被釘 → Harusnya shift D Kamis kemarin, sore perhatikan 小趙 dan kepala divisi, supaya 凱銘 gak kena tegur. "
            "自己稍微看一下設備的料源，有料就是要生產。月底我們不可能是停機的單位 → Cek material di mesin masing-masing, ada material ya produksi. Akhir bulan kita gak boleh mesin berhenti. "
            "之後有包到寶麗金屬注意一下，有一批訂單會備註客戶不要爐號標籤 → Nanti kalau ada packing untuk 寶麗金屬 perhatikan, ada order dicatat pelanggan tidak mau label heat number. "
            "非本月只有異型棒不管控，其他麻煩不要入了，昨天早班沒管控被檢討 → Bukan bulan ini cuma batang khusus bebas, sisanya jangan masuk, shift pagi kemarin gak kontrol kena tegur. "
            "非本月包裝不入庫 → Yang bukan order bulan ini jangan packing masuk gudang. "
            "急單再幫忙安排入庫 → Hanya order urgent yang tolong bantu atur masuk gudang. "
            "大成、SUNGEUN/佳東/麒譯/津展/DACAPO不擋非本月，各班在注意一下 → 大成, SUNGEUN/佳東/麒譯/津展/DACAPO order bukan bulan ini boleh masuk gudang, semua shift tolong perhatikan ya. "
            "H、S異型棒、大成、SUNGEUN、佳東……以上不擋非本月 → H, S, batang bentuk khusus, 大成, SUNGEUN, 佳東... yang di atas order bukan bulan ini boleh masuk gudang. "
            "開天車務必遵守規定目視吊掛物 → Operasi crane WAJIB lihat beban gantung sesuai aturan. "
            "護罩跟外勞宣導一下要蓋好 → Sosialisasi ke pekerja Indonesia pelindung mesin harus ditutup rapat. "
            "印勞打錯系統有提示 可是他們看不懂把他按掉了 → Pekerja Indonesia salah input, sistem ada peringatan tapi mereka gak ngerti jadi ditutup. "
            "拋光機interlock都不要拿東西擋著，上面會查 → Pengunci keamanan polishing jangan ditahan pakai benda, atasan akan periksa. "
            "來料自由端偏小 → Material masuk ujung bebasnya under size. "
            "自由端偏小 → Ujung bebasnya kecil / under size. "
            "自由端直徑偏小 → Diameter ujung bebasnya kecil. "
            "夾頭端偏大自由端偏小 → Ujung jepit besar, ujung bebas kecil. "
            "殺光痕嚴重但表粗有過 → Bekas grinding parah tapi surface roughness lulus. "
            "表粗有過目視沒過 → Surface roughness lulus tapi visual tidak lulus. "
            "涉及軋輥印痕的批次，請協助開立重工研磨至尺寸下限 → Lot kena roll mark, tolong buat WO rework grinding sampai batas bawah ukuran. "
            "護罩要隨時關閉，卡料需關閉電源後再取料 → Pelindung mesin harus ditutup, material macet HARUS matikan listrik dulu baru ambil. "
            "嚴禁運轉中設備直接以手搬動棒材 → DILARANG pindahkan batang baja dengan tangan saat mesin jalan. "
            "矯直機前壓輪故障，卡死無法上昇，已請修護協助處理 → Roda tekan straightening rusak macet, sudah minta maintenance bantu. "
            "氣壓缸更換備品回裝完成，測試OK正常生產 → Silinder pneumatik ganti spare part selesai, tes OK produksi normal. "
            "來料盤元不佳退回線外修磨 → Wire rod masuk kualitas buruk, dikembalikan offline repair grinding. "
            "不可側磨已宣導多次，納入劣項 → Larangan side grinding sudah disosialisasi berkali-kali, dicatat pelanggaran. "
            "E5線速是否過慢，僅2.4～3.6m/min → Kecepatan lini E5 terlalu lambat, cuma 2.4-3.6 m/min? "
            "眼模刮傷整修一次，無法改善，更換眼模 → Die tergores, perbaiki sekali tidak membaik, ganti die. "
            "E11已抽完，要回精整，請放行過帳 → E11 selesai drawing, harus kembali ke finishing, tolong release dan input data. "
            "更換備品後已恢復生產 → Setelah ganti spare part sudah kembali produksi. "
            "報表要記得確實填寫，尤其是雷射校正部分 → Laporan produksi ingat diisi benar, terutama bagian kalibrasi laser. "
            "幫追料 → Tolong kejar materialnya. "
            "幫追帳 → Tolong kejar data administrasinya. "
            "已2900別入帳了噢 → Sudah 2900 jangan masukkan data lagi ya. "
            # v3.9.30d B22 修補:夜點費誤譯案例(歐那實際遇到)
            "這個月薪資的減項應該是上個月夜點費計算錯誤扣回去的 → Pengurangan gaji bulan ini kemungkinan karena koreksi uang shift malam yang salah perhitungan bulan lalu. "
            "核對三月跟前面幾個月份比薪就可以發現夜點費特別高 → Bandingkan gaji bulan Maret dengan beberapa bulan sebelumnya, akan kelihatan uang shift malamnya memang lebih tinggi. "
            "夜點費算錯 → Uang shift malam salah hitung. "
            "這個月夜點費比較多 → Bulan ini uang shift malamnya lebih banyak. "
            "我的薪資減項多一個 → Pengurangan gaji saya ada tambahan satu lagi. "
            "底薪加夜點費加全勤 → Gaji pokok plus uang shift malam plus bonus kehadiran penuh. "
            "【印尼→中文】"
            "Saya mau izin besok → 我明天要請假 "
            "Mesinnya rusak → 機台壞了 "
            "Materialnya udah habis → 料用完了 "
            "Kapan gajinya keluar? → 薪水什麼時候發？ "
            "Saya gak ngerti → 我聽不懂 "
            "Boleh pulang duluan? → 可以先下班嗎？ "
            "Lembur sampai jam berapa? → 加班到幾點？ "
            "Bos, ini udah selesai → 老闆，這個好了 "
            "Ukurannya gak pas → 尺寸不對 "
            "Stoknya masih ada? → 庫存還有嗎？ "
            "Tolong ajarin saya → 請教我一下 "
            "Sudah rusak saat saya disana, saya lupa memberi tahu anda → 我在的時候就已經壞了，忘了跟你說 "
            "Batu gerindanya udah habis, mau ganti → 砂輪用完了，要換新的 "
            "Ini material dari shift sebelumnya, belum selesai → 這是上一班留下來的料，還沒做完 "
            "Saya sudah cek, ukurannya lewat toleransi → 我檢查過了，尺寸超出公差 "
            "Mesin E5 ada masalah, sudah panggil maintenance → E5機台有問題，已經叫修護了 "
            # v3.9.30d B22 修補:反向訓練(夜點費)
            "Pengurangan gaji bulan ini kemungkinan karena koreksi uang shift malam yang salah perhitungan bulan lalu → 這個月薪資的減項應該是上個月夜點費計算錯誤扣回去的 "
            "Uang shift malam bulan ini lebih banyak → 這個月夜點費比較多 "
            "Pengurangan gaji saya ada tambahan satu lagi → 我的薪資減項多一個 "
            "Cek slip gaji, ada tambahan potongan satu lagi → 看薪資單,多了一個減項 "
            "Bandingkan gaji bulan Maret dengan bulan-bulan sebelumnya → 比對三月跟前面幾個月薪資 "
            "Barang ini mau dikirim ke mana? → 這個東西要送去哪裡？ "
            "Yang ini sudah di-packing, tinggal masuk gudang → 這個已經包好了，只剩入庫 "
            "Tolong cek material di line 3 → 麻煩去看一下3號線的料 "
            "Saya sakit, hari ini gak bisa masuk → 我生病了，今天沒辦法上班 "
            "Oli mesinnya bocor → 機台漏油了 "
            "Hasil polishing ada goresan → 拋光後有刮痕 "
            "Permukaan kasar, gak lulus visual → 表面粗糙，目視沒過 "
            "Bos, die-nya udah aus, perlu ganti → 老闆，眼模磨損了，要換 "
            "Siap, saya kerjakan sekarang → 好，我現在做 "
            "Belum sempat, masih proses yang sebelumnya → 還沒來得及，前面的還在做 "
            "Sudah disampaikan ke mereka → 已經跟他們說了 "
            "Saya gak tahu harus taruh di mana → 我不知道要放哪裡 "
            "Ini ordernya urgent gak? → 這張單急不急？ "
            "Mau izin ke toilet sebentar → 我去一下廁所 "
            "Mesin udah jalan normal → 機台已經恢復正常了 "
            "Kemarin sudah bilang, tapi belum diperbaiki → 昨天就說了，但還沒修 "
            "Kita kekurangan material untuk order ini → 這張單的料不夠 "
            "Shift malam ada masalah di mesin I7 → 夜班I7機台出問題了 "
            "Tunggu sebentar, sedang ganti batu gerinda → 等一下，正在換砂輪 "
            "Sudah saya laporkan ke QC → 我已經跟品保回報了 "
            "Packing-nya salah, harus bongkar ulang → 包裝包錯了，要拆掉重包 "
            "Maaf bos, saya telat → 老闆抱歉，我遲到了 "
            "Surface roughness lulus tapi ada bekas grinding → 表粗有過但有殺光痕 "
            "Ini work order yang mana? → 這是哪一張工單？ "
            "Tolong bantu input data ke sistem → 麻煩幫忙過帳 "
            "Materialnya macet di mesin → 料卡在機台裡了 "
            "Pelindung mesin jangan dibuka saat jalan → 機台運轉中不要打開護罩 "
            "Sudah selesai semua, mau pulang → 全部做完了，要下班了 "
            "Gak ada label heat number di material ini → 這批料沒有爐號標籤 "
            "Diameter-nya under size, harus rework → 直徑偏小，要重工 "
            "Saya mau tukar shift sama dia → 我要跟他換班 "
            "Atasan bilang harus pakai pelindung → 上面說要戴護具 "
            "Besok libur gak? → 明天放假嗎？ "
            "Udah dicek, hasilnya OK → 已經檢查了，沒問題 "
            "Belakangan ini atasan sering datang inspeksi → 最近主管常常來巡場 "
            "Tadi pagi ditemukan banyak posisi kerja kosong → 今天早上發現很多工作崗位沒人 "
            "Saya menjalankan mesin pelan dengan kecepatan 500 → 我降速跑機台，速度500 "
            "Menjalankan mesin pelan → 降速跑機台 "
            "Mesin jalan pelan → 機台跑很慢 "
            "Naikkan kecepatan → 加速 "
            "Turunkan kecepatan → 降速 "
            "Kecepatan terlalu cepat → 速度太快了 "
            "Kecepatan terlalu pelan → 速度太慢了 "
            "Ganti batu gerinda → 換砂輪 "
            "Ganti batu polishing → 換拋光輪 "
            "Arus depan dan belakang tambah 0.3 → 前後電流加0.3 "
            "Mesin I16 saya jalankan pelan → 我I16慢慢跑 "
            "UT udah numpuk banyak material → UT囤一堆料了 "
            "QC udah pulang, keterlaluan → 品保已經下班了，太扯了 "
            "Batang 3 meter ditaruh di atas batang 6 meter → 三米放在六米上面 "
            "Tolong bilang ke mereka jangan taruh material kayak gini → 麻煩跟他們說不要這樣放料 "
            "Ada yang tahu siapa yang menaruh batang-batang ini? → 有人知道這些棒材誰放的嗎？ "
            "Siapa yang meletakkan material ini di sini? → 這批料誰放這裡的？ "
            "Yang di-packing hari ini 2 bundel semuanya kayak gini → 今天包的2把都這樣 "
            "Material masuk semuanya kurang 4-5 kilogram → 來料全部短少4-5公斤 "
            "Lot material ini ada masalah → 這批料有問題 "
            "Tolong awasin ya → 幫盯一下 "
            "Kok bisa kayak gini sih → 怎麼搞的啦 "
            "Orangnya mana? → 人咧？ "
            "6 bundel ini tolong masukin gudang malam ini → 這6把麻煩今晚入庫 "
            "Besok pagi sales perlu ambil data, makasih → 明早業務要抓資料，謝謝 "
            "Mesin polishing BF2 sedang diperbaiki → BF2拋光機維修中 "
            "Diameter 44.45 kemarin sudah bilang over produksi, sales udah balas belum? → 44.45前天有說超產，業務回覆了嗎？ "
            "Setelah spray paint, bagi packing sesuai jumlah order → 噴漆後照訂單量拆包 "
            "QC salah pilih proses, tolong kembalikan ke station 400 tanpa pemilik → 品保點錯製程，麻煩退回400無主 "
            "Data sudah dikembalikan ke 400, materialnya mau ke unit mana? → 帳已回400，料要回去哪個單位？ "
            "Ke proses peeling dan annealing, makasih → 去削皮退火，謝謝 "
            "Bagian peeling butuh troli, tolong bantu ya → 削皮需要台車，再麻煩一下 "
            "dua buah → 兩個 "
            "dua troli → 兩台台車 "
            "Maaf kan saya 🙏 kedepan ya akan lebih teliti dalam bekerja → 對不起 🙏 以後我會更仔細地工作 "
            "Maafkan saya 🙏 → 對不起 🙏 "
            "Mafkan saya bos, saya salah → 老闆對不起，是我的錯 "
            "Maf kan saya, gak akan ulang lagi → 對不起，不會再犯了 "
            "Mohon maaf, sudah saya perbaiki → 真的很抱歉，已經修正了 "
            "Minta maaf telat balas → 抱歉太晚回覆 "
            "Bagian peeling butuh troli, tolong bantu ya → 削皮需要台車，再麻煩一下 "
            "dua buah → 兩個 "
            "dua troli → 兩台台車 "
            "tiga bundel → 三把 "
            "5 batang → 5支 "
            "Sales bilang terima, tolong di-packing → 業務說收了，請包 "
            "Bundel ini kelebihan, tolong masukkan ke tanpa pemilik → 這把溢量，請入無主 "
            "Pelanggan minta 7 batang, gak terima pendek. Masuk cuma 6, 1 pendek dibuang sisa 5, bisa packing gak? → 客戶要7支不收短，來料只有6支其中1支短剔掉剩5支，能包嗎？ "
            "Kontainer sedang di jalan, sampai jam 9, mungkin bisa tunggu sebentar baru masuk gudang → 櫃子在路上9點到，可以等一下再入庫 "
            "DACAPO semuanya sudah masuk gudang → DACAPO都入完了 "
            "Tolong pakai kotak kayu 2700, cek jam berapa selesai, sales sore mau kirim → 請用2700大木箱裝，看幾點會好，業務下午要出貨 "
            "Berarti datanya belum masuk ke sistem → 那就是帳沒入到 "
            "Data ada masalah, sedang diurus → 資料異常，正在處理 "
            "Jadwal grinding diupdate, order urgent tolong atur cuci material dan polishing → 研磨排程更新了，急單麻煩安排洗料拋光 "
            "Rough polishing selesai, sudah di-release → 粗拋完已放行 "
            "Tolong release bundel ini dulu → 麻煩先放這把 "
            "Sudah di-release → 放了 "
            "Release order ini dulu → 先放這單 "
            "Tolong bantu release → 幫放一下 "
            "Lot ini jangan di-release dulu → 這批先不要放 "
            "Material taruh di samping → 料放旁邊 "
            "Taruh di lantai → 放地上 "
            "Hari ini pengiriman kurang, di 490 sedang produksi → 今天出貨差，在490 OL中 "
            "Kurang 2 ton, nanti keluarkan material, kalau bisa tolong packing → 差2噸，等等會出料，可以的話幫包裝 "
            "Tolong kembalikan ke 400, ada order baru untuk serap material → 請收回400，有單去化 "
            "Sudah dicuci dan dikasih ke E7 untuk polishing → 洗給E7拋了 "
            "Kalau packing ketemu customer itu perhatikan nomor material ini, order ini gak terima pendek harus pisah bundel → 包裝遇到那個客戶注意這料號，不收短尺要把短的分開捆 "
            "Rapat keputusan Imlek tidak stop mesin, shift kurang orang, mau angpao bisa gantikan shift → 開會決議過年不停機，班別出勤人數不夠，想賺紅包可以代班 "
            "HRD info pelatihan forklift, datang ikut aja, jam lembur diinput lewat sistem → 人事通知堆高機複訓，來上課就好，加班時數用忘卡補 "
            "Kepala divisi sudah pergi → 處長走了 "
            "Order urgent deadline tolong diproses, banyak belum sampai, polishing produksi sambil jalan → 有壓日期的急單幫處理一下，很多未到站，拋光一邊產出 "
            "Kaleng spray HARUS dilubangi baru buang ke jumbo bag → 噴漆罐一定要打洞才能丟到太空包 "
            "Target gudang 2950, batang khusus bebas, sisanya bukan bulan ini jangan masuk → 本月入庫目標2950，異型棒不擋，其餘非本月不入庫 "
            "Target tercapai, sekarang hanya urgent, batang khusus, dan order ditunda → 目標已達標，目前只入急單、異型棒跟遞延單 "
            "Hari ini gak ada inspeksi, kemarin sudah datang → 今天沒點名，昨天來過了 "
            "Cek material di mesin masing-masing, ada material ya produksi → 自己看設備的料源，有料就要生產 "
            "Nanti kalau ada packing untuk customer itu perhatikan, ada order dicatat pelanggan tidak mau label heat number → 之後有包到那個客戶注意，有一批不要爐號標籤 "
            "Bukan bulan ini cuma batang khusus bebas, sisanya jangan masuk, kemarin gak kontrol kena tegur → 非本月只有異型棒不管控，其他不要入了，昨天沒管控被檢討 "
            "Yang bukan order bulan ini jangan packing masuk gudang → 非本月包裝不入庫 "
            "Hanya order urgent yang tolong bantu atur masuk gudang → 急單再幫忙安排入庫 "
            "Order bukan bulan ini boleh masuk gudang, semua shift tolong perhatikan → 不擋非本月，各班注意一下 "
            "Operasi crane WAJIB lihat beban gantung sesuai aturan → 開天車務必遵守規定目視吊掛物 "
            "Sosialisasi ke pekerja Indonesia pelindung mesin harus ditutup rapat → 跟外勞宣導護罩要蓋好 "
            "Pekerja Indonesia salah input, sistem ada peringatan tapi mereka gak ngerti jadi ditutup → 印勞打錯了系統有提示，但他們看不懂就按掉了 "
            "Pengunci keamanan polishing jangan ditahan pakai benda, atasan akan periksa → 拋光機interlock不要拿東西擋著，上面會查 "
            "Material masuk ujung bebasnya under size → 來料自由端偏小 "
            "Barang kecil di ujung → 自由端偏小 "
            "Barang kecil di ujing → 自由端偏小 "
            "Kecil di ujung → 自由端偏小 "
            "Ujung kecil → 自由端偏小 "
            "Ujungnya kecil → 自由端偏小 "
            "Ujung bebas kecil → 自由端偏小 "
            "Diameter ujung kecil → 自由端直徑偏小 "
            "Ujung jepit besar ujung bebas kecil → 夾頭端偏大自由端偏小 "
            "Bekas grinding parah tapi surface roughness lulus → 殺光痕嚴重但表粗有過 "
            "Surface roughness lulus tapi visual tidak lulus → 表粗有過目視沒過 "
            "Lot kena roll mark, tolong buat WO rework grinding sampai batas bawah ukuran → 有軋輥印痕的批次，請開重工研磨至尺寸下限 "
            "Pelindung mesin harus ditutup, material macet HARUS matikan listrik dulu baru ambil → 護罩要隨時關，卡料要關電源再取料 "
            "DILARANG pindahkan batang baja dengan tangan saat mesin jalan → 嚴禁運轉中用手搬棒材 "
            "Roda tekan straightening rusak macet, sudah minta maintenance bantu → 矯直機前壓輪故障卡死，已請修護處理 "
            "Silinder pneumatik ganti spare part selesai, tes OK produksi normal → 氣壓缸換備品完成，測試OK正常生產 "
            "Wire rod masuk kualitas buruk, dikembalikan offline repair grinding → 來料盤元不佳，退回線外修磨 "
            "Larangan side grinding sudah disosialisasi berkali-kali, dicatat pelanggaran → 不可側磨已宣導多次，納入劣項 "
            "Kecepatan lini E5 terlalu lambat, cuma 2.4-3.6 m/min → E5線速太慢了，只有2.4到3.6 "
            "Die tergores, perbaiki sekali tidak membaik, ganti die → 眼模刮傷修一次沒改善，換眼模 "
            "E11 selesai drawing, harus kembali ke finishing, tolong release dan input data → E11抽完了要回精整，請放行過帳 "
            "Setelah ganti spare part sudah kembali produksi → 換完備品已恢復生產 "
            "Laporan produksi ingat diisi benar, terutama bagian kalibrasi laser → 報表要確實填，尤其雷射校正 "
            "Tolong kejar materialnya → 幫追料 "
            "Tolong kejar data administrasinya → 幫追帳 "
            "Sudah 2900 jangan masukkan data lagi ya → 已2900別入帳了 "
            # ===== v3.9.30 新增:工序+噸數=待加工料量 / jatuh tempo=積壓 =====
            # 對應歐那實際遇到的誤譯案例(2026-05-04 班長公告)
            "Masih ada 30 ton mesin pemoles yang jatuh tempo bulan lalu → 還有上個月積壓未拋的30噸料 "
            "46 ton mesin pemoles yang jatuh tempo bulan ini → 本月到期的46噸料要拋 "
            "Masih ada 30 ton mesin pemoles yang jatuh tempo bulan lalu, dan 46 ton mesin pemoles yang jatuh tempo bulan ini → 上個月積壓的30噸料還沒拋,本月又有46噸料到期要拋 "
            "Personil mesin pemoles harus memperhatikan efisiensi produksi → 拋光機人員一定要注意生產效率 "
            "Personil grinding harus memperhatikan efisiensi produksi → 研磨人員一定要注意生產效率 "
            # 通則範例(讓模型抓住模式)
            "20 ton mesin grinding jatuh tempo → 研磨積壓20噸料 "
            "Bahan jatuh tempo bulan lalu masih banyak → 上月積壓的料還很多 "
            "Order yang sudah jatuh tempo harus diprioritaskan → 已積壓的訂單要優先處理 "
            "Belum jatuh tempo, jangan dulu produksi → 還沒到期,先別生產 "
            "Material polishing yang tertunda → 拋光積壓的料 "
            "Tunggakan grinding bulan ini 50 ton → 本月研磨積壓 50 噸 "
            "【印尼日常短句→中文】"
            "Iya → 對 "
            "Iya benar → 對，沒錯 "
            "Ini benar → 這是對的 "
            "Benar → 對 "
            "Betul → 對 "
            "Bukan → 不是 "
            "Salah → 錯了 "
            "Bukan begitu → 不是這樣 "
            "Siap → 收到 "
            "Siap bos → 好的老闆 "
            "Oke → 好 "
            "Oke bos → 好的老闆 "
            "Baik → 好的 "
            "Udah selesai → 做好了 "
            "Udah beres → 搞定了 "
            "Belum → 還沒 "
            "Belum selesai → 還沒好 "
            "Belum sempat → 還沒來得及 "
            "Gak tau → 不知道 "
            "Gak ngerti → 看不懂 "
            "Gak bisa → 沒辦法 "
            "Gak ada → 沒有 "
            "Gak mau → 不要 "
            "Mau → 要 "
            "Bisa → 可以 "
            "Boleh → 可以 "
            "Gak boleh → 不行 "
            "Jangan → 不要 "
            "Tunggu → 等一下 "
            "Tunggu sebentar → 等一下 "
            "Bentar → 等一下 "
            "Nanti → 晚點 "
            "Nanti dulu → 晚點再說 "
            "Maaf → 抱歉 "
            "Maaf bos → 老闆抱歉 "
            "Sori → 不好意思 "
            "Makasih → 謝謝 "
            "Terima kasih → 謝謝 "
            "Sama-sama → 不客氣 "
            "Gak apa-apa → 沒關係 "
            "Gpp → 沒關係 "
            "Santai aja → 慢慢來 "
            "Capek → 累了 "
            "Capek banget → 累死了 "
            "Sakit → 不舒服 "
            "Pusing → 頭暈 "
            "Lagi makan → 在吃飯 "
            "Lagi istirahat → 在休息 "
            "Sudah pulang → 已經下班了 "
            "Mau pulang → 要下班了 "
            "Izin → 請假 "
            "Izin sakit → 病假 "
            "Izin besok → 明天請假 "
            "Telat → 遲到 "
            "Saya telat → 我遲到了 "
            "Masuk → 上班 "
            "Gak masuk → 沒來上班 "
            "Lembur → 加班 "
            "Gak lembur → 不加班 "
            "Mau lembur → 要加班 "
            "Libur → 放假 "
            "Cuti → 休假 "
            "Mau cuti → 要請假 "
            "Di mana? → 在哪裡？ "
            "Yang mana? → 哪一個？ "
            "Kapan? → 什麼時候？ "
            "Kenapa? → 為什麼？ "
            "Gimana? → 怎麼樣？ "
            "Berapa? → 多少？ "
            "Siapa? → 誰？ "
            "Ada apa? → 怎麼了？ "
            "Ngapain? → 在幹嘛？ "
            "Serius? → 真的嗎？ "
            "Masa? → 真的假的？ "
            "Yang bener? → 真的嗎？ "
            "Emang → 確實 "
            "Kayaknya → 好像 "
            "Mungkin → 可能 "
            "Pasti → 一定 "
            "Cepat → 快 "
            "Cepatan → 快點 "
            "Pelan-pelan → 慢慢來 "
            "Hati-hati → 小心 "
            "Awas → 小心 "
            "Bahaya → 危險 "
            "Tolong → 拜託 "
            "Bantu → 幫忙 "
            "Panggil → 叫過來 "
            "Bilang → 跟他說 "
            "Suruh → 叫他去 "
            "Cek dulu → 先檢查 "
            "Lihat dulu → 先看看 "
            "Cobain → 試試看 "
            "Ganti → 換 "
            "Perbaiki → 修理 "
            "Matikan → 關掉 "
            "Nyalakan → 打開 "
            "Tambah → 再加 "
            "Kurang → 不夠 "
            "Lebih → 多了 "
            "Pas → 剛好 "
            "Lewat → 超過了 "
            "Habis → 用完了 "
            "Penuh → 滿了 "
            "Kosong → 沒有 "
            "Rusak → 壞了 "
            "Macet → 卡住了 "
            "Bocor → 漏了 "
            "Mati → 停了"
            + extra_rule +
            # v3.9.8: GOLDEN RULES (added 2026-05) — based on real production failures.
            # These are XML-tagged for GPT-5's "surgical instruction following"
            # (per OpenAI GPT-5 prompting guide best practices).
            " <format_preservation_rules>"
            " RULE 1 — Output format must MIRROR the source format exactly:"
            " - If the source is plain prose without bullets/emoji/⚠️, the translation MUST be plain prose."
            " - DO NOT split a single paragraph into bullet points or numbered lists."
            " - DO NOT add ⚠️, ⚡, ‼️, or any emoji that the source doesn't have."
            " - DO NOT reorganize sentences into a 'cleaner' structure."
            " - DO NOT add introductory phrases like 'Here's the translation:' or '以下為翻譯:'."
            " </format_preservation_rules>"
            " <passive_voice_rules>"
            " RULE 2 — Chinese passive structures with 「被」 must be translated as passive:"
            " - 「將被X列管」 = 'will be supervised by X' (factory is the OBJECT, NOT subject)"
            " - 「被開立罰單」 = 'received a fine' / 'kena denda' (we received it)"
            " - 「interlock 被 bypass」 = 'interlock di-bypass' / 'interlock was bypassed'"
            " - 「被職安署列管」 = 'akan kena pengawasan Dinas K3' (we will be put under supervision)"
            " NEVER translate 「被」 as if the subject does the action to others."
            " </passive_voice_rules>"
            " <factory_terms_rules>"
            " RULE 3 — Keep these factory/safety terms VERBATIM (do NOT translate):"
            " interlock, bypass, MSDS, NG, OK, PPE, LOTO, SOP, OEE, JSA, SDS, EHS, K3, PMI"
            " For 「記過/警告處分」 → use 'SP (Surat Peringatan)' in Indonesian."
            " For 「職安署」 → use 'Dinas K3' in Indonesian."
            " </factory_terms_rules>"
            " <register_rules>"
            " RULE 4 — Match the formality of the source:"
            " - Casual chat (短句, 口語) → casual Indonesian (gak, udah, nih, dong)."
            " - Formal announcement (含「公告/通知/重大職災/列管/停工/警告」) → formal Indonesian"
            "   (use 'tidak', 'sudah', 'ini', avoid slang)."
            " - Keep cause-effect chains intact: 「如果A,將B」 → 'Kalau A, akan B' (single sentence)."
            " - Do NOT soften warnings into suggestions."
            " </register_rules>"
            " IMPORTANT: Preserve the original line breaks and blank lines exactly. If the source has a blank line between paragraphs, keep a blank line in the same position in the translation."
            " Only output the translation. No quotes, no explanation, no prefix."
        )

        if repair_mode and bad_result:
            msg = (
                "Original text (source language): " + protected + "\n\n"
                "Bad translation that leaked source-language words: " + bad_result + "\n\n"
                "Rewrite the bad translation into pure " + tgt_name +
                ". Preserve names and __MENTION__ placeholders exactly. Translate every remaining source-language word."
            )
        else:
            msg = "Translate from " + src_name + " to " + tgt_name + ": " + protected

        _model = pick_model(text)
        # v3.2-0426e: build messages array, optionally using few-shot "messages" format
        # which separates examples into example_user/example_assistant pairs (OpenAI standard).
        if fewshot_mode == "messages":
            _msgs = _build_messages_with_fewshot(sys_prompt, msg, src, tgt)
        else:
            _msgs = [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": msg}
            ]
        # v3.9.5: snapshot what we are about to send to OpenAI so /admin/debug/last-translate can show it.
        # This is the single most useful diagnostic when custom examples aren't taking effect.
        try:
            global last_translate_debug
            _example_count = sum(1 for m in _msgs if m.get("role") in ("user", "assistant")) - 1  # -1 for the actual user msg
            _example_count = max(0, _example_count // 2)  # pairs
            last_translate_debug = {
                "ts": int(time.time()),
                "src_text": text,
                "src_lang": src,
                "tgt_lang": tgt,
                "model_picked": _model,
                "fewshot_mode": fewshot_mode,
                "tone": _tone,
                "example_pairs_in_prompt": _example_count,
                "messages_sent": _msgs,  # full messages array
                "strict_no_source_script": strict_no_source_script,
                "repair_mode": repair_mode,
            }
        except Exception:
            pass
        # v3.2-0426e: build kwargs with all official OpenAI features
        # v3.9.2 (2026-05): GPT-5 series (incl. 5.4, 5.5) are also reasoning models.
        # They reject temperature/top_p/logit_bias/max_tokens and require
        # max_completion_tokens. Without this fix, all GPT-5 calls return 400
        # and fall back to Google Translate — which is why 範例 didn't take effect.
        # v3.9.7 (2026-05): 全面用 model_supports() 集中過濾。
        # 不論使用者勾選什麼進階設定,後端永遠根據實際模型的能力決定要不要送到 API。
        # 任何「誤勾」都會被靜默忽略,不會造成 400 BadRequest。
        _is_reasoning_model = not model_supports(_model, "temperature")  # 保留給下游 code 用
        _kwargs = {
            "model": _model,
            "messages": _msgs,
            # v3.9.22: 加 timeout 30 秒,避免無限等待
            "timeout": 30,
        }
        # Sampling parameters
        if model_supports(_model, "temperature"):
            _kwargs["temperature"] = translation_temperature
        if model_supports(_model, "top_p"):
            _kwargs["top_p"] = translation_top_p
        if model_supports(_model, "seed") and translation_seed and translation_seed != 0:
            _kwargs["seed"] = int(translation_seed)
        # Token limit (different parameter name per family)
        # v3.9.30: 動態 token 預算
        # 之前寫死 2000:對 reasoning model(GPT-5 系列) max_completion_tokens 包含
        # reasoning tokens + 輸出 tokens,長公告(>500 字)做不完,會被截斷或回空。
        # 新邏輯:輸入越長,輸出空間越大;reasoning model 額外給 reasoning 預算。
        _src_len = len(text or "")
        # 估計輸出長度(中文/印尼文 ~1.5x 互譯,給 2x 安全係數)
        _output_budget = max(800, int(_src_len * 2.5))
        # reasoning model 還要額外給 reasoning 預算
        if not model_supports(_model, "temperature"):  # 是 reasoning model
            _output_budget = _output_budget + 4000  # 給 reasoning 4K 緩衝
        # 上限 16K(防止意外炸掉)
        _output_budget = min(_output_budget, 16000)
        if model_supports(_model, "max_completion_tokens"):
            _kwargs["max_completion_tokens"] = _output_budget
        elif model_supports(_model, "max_tokens"):
            # 非 reasoning model 不需要 reasoning 緩衝
            _kwargs["max_tokens"] = min(max(800, int(_src_len * 2.5)), 4000)
        # v3.9.8 (2026-05): Translation-optimal reasoning_effort.
        # OpenAI's official guide: "gpt-4.1: gpt-5.2 with `none` reasoning"
        # Translation is a lightweight, instruction-following task. Higher
        # reasoning_effort HURTS instruction adherence (arxiv 2505.14810).
        # We auto-select the LOWEST viable effort for the model family,
        # ignoring the admin's stored 'reasoning_effort' setting because
        # that was tuned for o-series and is wrong for gpt-5 series.
        if model_supports(_model, "reasoning_effort"):
            _optimal_effort = optimal_reasoning_for_translation(_model)
            if _optimal_effort:
                _kwargs["reasoning_effort"] = _optimal_effort
            elif reasoning_effort in ("low", "medium", "high"):
                # Fallback to admin setting only if no optimal known
                _kwargs["reasoning_effort"] = reasoning_effort
        # v3.9.8: verbosity=low for GPT-5 family.
        # Per OpenAI: "low is often a better starting point for concise responses"
        # This prevents GPT-5 from spontaneously formatting output with bullets,
        # ⚠️ markers, or structural rewrites that aren't in the source.
        # Translation should preserve source format, not reorganize it.
        if model_supports(_model, "verbosity"):
            _kwargs["verbosity"] = "low"
        # Stop sequences
        if model_supports(_model, "stop"):
            _stops = _build_stop_sequences(tgt)
            if _stops:
                _kwargs["stop"] = _stops
        # User ID hash (always supported)
        _user_id = getattr(_tl, 'user_id', '')
        if send_user_id_to_openai and _user_id:
            import hashlib as _h
            _kwargs["user"] = _h.sha256(_user_id.encode()).hexdigest()[:32]
        # logit_bias
        if model_supports(_model, "logit_bias"):
            try:
                _bias = _build_logit_bias(tgt, _model)
                if _bias:
                    _kwargs["logit_bias"] = _bias
            except Exception as _be:
                logger.warning("logit_bias skip: %s", _be)
        # Metadata (requires store=True since 2025)
        if send_metadata_to_openai and model_supports(_model, "metadata"):
            _meta = {}
            _gid = getattr(_tl, 'group_id', '')
            if _gid:
                _meta["group_id"] = str(_gid)[:64]
            _meta["src_lang"] = src or ""
            _meta["tgt_lang"] = tgt or ""
            if _meta:
                _kwargs["metadata"] = _meta
                _kwargs["store"] = True
        # Structured Outputs (json_schema)
        if structured_output_enabled and model_supports(_model, "structured_output"):
            _kwargs["response_format"] = {
                "type": "json_schema",
                "json_schema": {
                    "name": "translation_result",
                    "strict": True,
                    "schema": {
                        "type": "object",
                        "properties": {
                            "translation": {
                                "type": "string",
                                "description": "The final translation, no notes/explanations"
                            },
                            "confidence": {
                                "type": "number",
                                "description": "0-1 confidence score for this translation"
                            },
                            "covered_all_information": {
                                "type": "boolean",
                                "description": "Whether the translation covers ALL information points from the source (key for announcement-type messages)"
                            },
                            "message_type": {
                                "type": "string",
                                "enum": ["announcement", "incident", "general", "question", "command"],
                                "description": "What kind of message this is"
                            },
                            "preserved_terms": {
                                "type": "array",
                                "items": {"type": "string"},
                                "description": "List of terms kept verbatim (machine codes like BF/BF2/CYA, names, numbers)"
                            },
                            "alternatives": {
                                "type": "array",
                                "items": {"type": "string"},
                                "description": "Up to 2 alternative phrasings"
                            }
                        },
                        "required": ["translation", "confidence", "covered_all_information",
                                     "message_type", "preserved_terms", "alternatives"],
                        "additionalProperties": False
                    }
                }
            }
        # Logprobs (信心度)
        _supports_logprobs = model_supports(_model, "logprobs")
        if logprobs_enabled and _supports_logprobs:
            _kwargs["logprobs"] = True
        # Prompt cache key (sticky routing)
        if model_supports(_model, "prompt_cache_key"):
            try:
                _ck = _build_cache_key(getattr(_tl, 'group_id', ''), src, tgt, "trans")
                if _ck:
                    _kwargs["prompt_cache_key"] = _ck
            except Exception:
                pass
        try:
            _event_log_write("translate_call_start", {
                "model": _model,
                "src": src,
                "tgt": tgt,
                "text_len": len(text or ""),
                "kwargs_keys": list(_kwargs.keys()),
            })
        except Exception:
            pass
        try:
            r = oai.chat.completions.create(**_kwargs)
        except Exception as _te:
            try:
                _event_log_write("translate_call_failed", {
                    "model": _model,
                    "error": str(_te)[:300],
                    "error_type": type(_te).__name__,
                })
            except Exception:
                pass
            raise
        try:
            _event_log_write("translate_call_done", {
                "model": _model,
                "finish": r.choices[0].finish_reason if (r and r.choices) else None,
                "content_len": len(r.choices[0].message.content or "") if (r and r.choices) else 0,
            })
        except Exception:
            pass
        # v3.9.5: record raw OpenAI response into debug snapshot
        try:
            _resp_text = r.choices[0].message.content if (r and r.choices) else ""
            last_translate_debug["openai_raw_response"] = _resp_text
            last_translate_debug["openai_finish_reason"] = r.choices[0].finish_reason if (r and r.choices) else None
            last_translate_debug["openai_status"] = "success"
            last_translate_debug["kwargs_keys_sent"] = sorted(list(_kwargs.keys()))
        except Exception:
            pass
        # Track cache hit for stats (if available in response)
        try:
            if hasattr(r, 'usage') and hasattr(r.usage, 'prompt_tokens_details'):
                _cached = getattr(r.usage.prompt_tokens_details, 'cached_tokens', 0)
                if _cached and prompt_caching_enabled:
                    logger.info("Prompt cache hit: %d tokens cached", _cached)
        except Exception:
            pass
        # Batch C: extract confidence
        _confidence = 1.0
        try:
            if logprobs_enabled and _supports_logprobs and r.choices and getattr(r.choices[0], 'logprobs', None):
                _confidence = _calc_confidence_from_logprobs(r.choices[0].logprobs)
        except Exception:
            _confidence = 1.0
        track_tokens(r)
        # v3.2-0426e: parse structured output if used
        # v3.9.30 B7 修補: GPT-5 reasoning model 在 max_completion_tokens 用光 reasoning 預算
        # 後可能回 content=None + finish_reason='length'。要先檢查不要直接 .strip()。
        _msg_content = r.choices[0].message.content if r.choices else None
        if _msg_content is None:
            _finish_reason = r.choices[0].finish_reason if r.choices else "no_choice"
            logger.error(
                "[translate_openai] empty content! model=%s finish=%s — return None to trigger fallback",
                _model, _finish_reason
            )
            # 寫進 last_translate_debug 方便事後 debug
            try:
                last_translate_debug["openai_finish_reason"] = _finish_reason
                last_translate_debug["empty_content"] = True
            except Exception:
                pass
            return None  # 上游會 fallback 到 google translate
        _raw = _msg_content.strip()
        _struct_self_check_failed = None  # v3.5: 記錄 self-check 失敗原因供日誌
        if structured_output_enabled and _raw.startswith("{"):
            try:
                import json as _json_mod
                _parsed = _json_mod.loads(_raw)
                result = str(_parsed.get("translation", _raw)).strip()
                # Override confidence from structured output if available
                _struct_conf = _parsed.get("confidence")
                if isinstance(_struct_conf, (int, float)):
                    _confidence = float(_struct_conf)
                
                # ★ v3.5 Self-check 1:preserved_terms 真的有出現在 translation 嗎?
                # 機台代號、人名、數字保留錯了會嚴重失真,直接擋
                _preserved_terms = _parsed.get("preserved_terms", []) or []
                if isinstance(_preserved_terms, list) and result:
                    _missing_terms = []
                    for _term in _preserved_terms:
                        _t = str(_term).strip()
                        if _t and _t not in result:
                            # 容忍大小寫差異
                            if _t.lower() not in result.lower():
                                _missing_terms.append(_t)
                    if _missing_terms:
                        _struct_self_check_failed = f"missing_terms:{_missing_terms[:3]}"
                        # 降信心,觸發後續警告
                        _confidence = min(_confidence, 0.5)
                        logger.warning(
                            "Structured self-check FAILED: claimed preserved_terms %s "
                            "but they're missing from translation. Original=%s, Translation=%s",
                            _missing_terms, text[:60], result[:60]
                        )
                
                # ★ v3.5 Self-check 2:GPT 自己說沒涵蓋全部資訊?
                _covered = _parsed.get("covered_all_information", True)
                if _covered is False:
                    _struct_self_check_failed = (_struct_self_check_failed or "") + ";claimed_incomplete"
                    _confidence = min(_confidence, 0.6)
                    logger.warning(
                        "Structured self-check: GPT admits incomplete coverage. "
                        "Original=%s, Translation=%s",
                        text[:60], result[:60]
                    )
                
                # ★ v3.5 Self-check 3:訊息類型與我們的分類器交叉驗證(僅 ID→ZH)
                # 如果 GPT 說是 announcement 但我們的分類器卻說 incident,代表有歧見
                # 這時保險起見不要走槽位拼接降級
                if src == "id" and tgt == "zh":
                    _gpt_msg_type = _parsed.get("message_type", "")
                    try:
                        _our_cls = classify_factory_message(text, src="id")
                        if _gpt_msg_type == "announcement" and _our_cls.get("type") != "announcement":
                            # GPT 認為是公告,我們的分類器漏掉了 - 信任 GPT
                            try:
                                _tl.gpt_says_announcement = True
                            except Exception:
                                pass
                    except Exception:
                        pass
                
                # 把 self-check 結果記到 thread-local 供日誌顯示
                if _struct_self_check_failed:
                    try:
                        _tl.struct_self_check = _struct_self_check_failed
                    except Exception:
                        pass
            except Exception as _e:
                logger.error("Structured output parse error: %s", _e)
                result = _raw
        else:
            result = _raw
        result = restore_mentions(result, placeholders)
        if src == "id" and tgt == "zh":
            _raw_factory_result = result
            _bad, _reason, _domains = detect_factory_semantic_error(text, _raw_factory_result, src, tgt)
            if _bad:
                _tl.factory_audit = {
                    "src": text,
                    "type": "factory_semantic_auto_detected",
                    "reason": _reason,
                    "raw_translation": _raw_factory_result,
                    "domain": _domains,
                    "auto_corrected": True,
                }
            result = post_fix_factory_id_to_zh(text, result)
            if _bad:
                _tl.factory_audit["corrected_translation"] = result
        # Fix known GPT translation mistakes and restore customer names
        if src == "zh":
            if tgt == "id":
                _raw_factory_result_zhid = result
                _bad_zhid, _reason_zhid, _domains_zhid = detect_factory_semantic_error_zh_id(text, _raw_factory_result_zhid)
                if _bad_zhid:
                    _tl.factory_audit = {
                        "src": text,
                        "type": "factory_semantic_auto_detected_zh_id",
                        "reason": _reason_zhid,
                        "raw_translation": _raw_factory_result_zhid,
                        "domain": _domains_zhid,
                        "auto_corrected": True,
                    }
                    _fixed_zhid = post_fix_factory_zh_to_id(text, result)
                    _bad_after, _reason_after, _ = detect_factory_semantic_error_zh_id(text, _fixed_zhid)
                    if _bad_after:
                        _repaired_zhid = repair_factory_translation_openai_zh_id(text, _fixed_zhid, _reason_after)
                        if _repaired_zhid:
                            _fixed_zhid = post_fix_factory_zh_to_id(text, _repaired_zhid)
                    result = _fixed_zhid
                    _tl.factory_audit["corrected_translation"] = result
            result = post_fix_translation(result)
            if tgt == "id":
                result = post_fix_factory_zh_to_id(text, result)
            result = restore_customers(result, cust_placeholders)
        # v3.2-0426d Batch B: Round-trip verification
        _did_double_check = False
        _similarity_val = 1.0
        _multi_path_details = None
        try:
            if _should_double_check(text, src):
                _did_double_check = True
                
                # ★ v3.6 多路徑反譯:長句 + 啟用時走兩條路徑
                _orig_char_len = len(re.sub(r"\s+", "", text or ""))
                _use_multi_path = (
                    multi_path_backtrans_enabled 
                    and _orig_char_len >= multi_path_min_chars
                    and src in ("id", "zh") and tgt in ("zh", "id")
                )
                
                if _use_multi_path:
                    passes, _multi_path_details = _multi_path_back_translation(text, result, src, tgt)
                    _similarity_val = _multi_path_details.get("path1_jaccard", 1.0)
                    if not passes:
                        decision = _multi_path_details.get("final_decision", "unknown")
                        logger.warning(
                            "Multi-path back-trans FAILED (%s): orig=%r trans=%r details=%s",
                            decision, text[:80], result[:80], _multi_path_details
                        )
                        result = "⚠️ " + result
                        # 記錄到 thread-local 供監控儀表板用
                        try:
                            _tl.multi_path_fail = _multi_path_details
                        except Exception:
                            pass
                    else:
                        logger.info("Multi-path back-trans passed: %s", _multi_path_details)
                else:
                    passes, _similarity_val, back = _round_trip_check(text, result, src, tgt)
                    if not passes:
                        logger.warning("Round-trip check FAILED: orig=%r trans=%r back=%r sim=%.2f",
                                       text[:80], result[:80], back[:80], _similarity_val)
                        result = "⚠️ " + result
                    else:
                        logger.info("Round-trip check passed: sim=%.2f", _similarity_val)
        except Exception as _e:
            logger.error("Round-trip wrap error: %s", _e)
        # ★ v3.3:低信心警告 - 只在反譯檢查未觸發時才加 ⚠️,避免雙重前綴
        try:
            if (logprobs_enabled and _confidence < confidence_threshold 
                and not result.startswith("⚠️") and not result.startswith("⚠")):
                result = "⚠️ " + result
                logger.warning("Low confidence translation: %.2f < %.2f for %r",
                               _confidence, confidence_threshold, text[:80])
        except Exception:
            pass
        # Batch D: log this translation
        try:
            _tokens_used = 0
            if r and getattr(r, 'usage', None):
                _tokens_used = r.usage.total_tokens
            _log_translation(text, result, src, tgt, _model, _tokens_used,
                             _confidence, _did_double_check, _similarity_val,
                             getattr(_tl, 'group_id', ''))
        except Exception:
            pass
        return result
    except Exception as e:
        logger.error("OpenAI error: %s", e)
        # v3.9.5: record exception in debug snapshot
        try:
            last_translate_debug["openai_status"] = "exception"
            last_translate_debug["openai_error"] = str(e)[:500]
            last_translate_debug["openai_error_type"] = type(e).__name__
        except Exception:
            pass
        return None


def translate_google(text, src, tgt):
    try:
        protected, placeholders = protect_mentions(text)
        lang_map = {
            "zh": "zh-TW", "id": "id", "en": "en",
            "vi": "vi", "th": "th", "ja": "ja",
            "ko": "ko", "ms": "ms", "tl": "tl",
        }
        sl = lang_map.get(src, src)
        tl = lang_map.get(tgt, tgt)
        q = urllib.parse.quote(protected)
        url = "https://translate.googleapis.com/translate_a/single?client=gtx&sl=" + sl + "&tl=" + tl + "&dt=t&q=" + q
        req = urllib.request.Request(url)
        req.add_header("User-Agent", "Mozilla/5.0")
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            parts = []
            for item in data[0]:
                if item[0]:
                    parts.append(item[0])
            result = "".join(parts)
            result = restore_mentions(result, placeholders)
            # Fix known translation mistakes
            result = post_fix_translation(result)
            return result
    except Exception as e:
        logger.error("Google translate error: %s", e)
        return None


def cache_get(text, src, tgt):
    """Get translation from cache if exists and not expired."""
    key = (text.strip(), src, tgt)
    with _cache_lock:
        if key in translation_cache:
            result, ts = translation_cache[key]
            if time.time() - ts < CACHE_TTL:
                logger.info("Cache hit: %s -> %s", src, tgt)
                return result
            else:
                del translation_cache[key]
    return None


def cache_set(text, src, tgt, result):
    """Store translation in cache, evict oldest if full."""
    key = (text.strip(), src, tgt)
    with _cache_lock:
        if len(translation_cache) >= CACHE_MAX_SIZE:
            oldest_key = min(translation_cache, key=lambda k: translation_cache[k][1])
            del translation_cache[oldest_key]
        translation_cache[key] = (result, time.time())


def translate_with_retry(func, text, src, tgt, max_retries=2):
    """Call a translation function with retry on failure."""
    for attempt in range(max_retries + 1):
        result = func(text, src, tgt)
        if result:
            return result
        if attempt < max_retries:
            wait = 1 * (attempt + 1)
            logger.warning("Retry %d/%d after %ds for %s", attempt + 1, max_retries, wait, func.__name__)
            time.sleep(wait)
    return None


def translate(text, src, tgt):
    """Public translate wrapper: guarantees translation_log entry on every successful return.

    Inner function may also call _log_translation along its paths; we mark those entries
    with an internal flag so this wrapper does NOT double-log them.
    """
    # Snapshot log length before to detect whether inner already logged
    _len_before = len(translation_log)
    result = _translate_inner(text, src, tgt)
    try:
        if result is not None and len(translation_log) == _len_before:
            # No inner path logged this turn — write a guaranteed entry now.
            _log_translation(text, result, src, tgt, "wrapper", 0, 1.0, False, 1.0,
                             getattr(_tl, 'group_id', ''))
    except Exception as e:
        import traceback
        logger.error("[TLOG] wrapper log failed: %s\n%s", e, traceback.format_exc())
    return result


# =====================================================================
# v3.7 段落結構保留(分段翻譯)
# =====================================================================
# 問題:GPT 即使在 prompt 中要求保留段落,長訊息仍可能合併段落
# 解法:把訊息以 \n\n 切段,每段獨立翻譯,再用相同分隔符拼接
# 原理:每段獨立翻譯 = GPT 不可能合併不同的 API 呼叫的結果

def _has_paragraph_structure(text):
    """偵測訊息是否含分段結構
    
    判斷標準:
      - 含 \n\n (空行分隔)
      - 或含多個 \n 且總長度 > 閾值
    """
    if not text:
        return False
    # 優先:有空行分隔
    if "\n\n" in text:
        return True
    # 次要:訊息夠長且含多個換行
    char_len = len(re.sub(r"\s+", "", text))
    if char_len >= paragraph_split_threshold and text.count("\n") >= 2:
        return True
    return False


def _split_into_paragraphs(text):
    """把訊息切成段落 list,保留分隔符資訊以便還原
    
    回傳: [(paragraph_text, separator_after), ...]
      separator_after = "\n\n" / "\n" / ""
    """
    # 用空行分段(優先),保留分隔符
    # 把連續空行視為一個分隔符,不加入空段
    parts = []
    current = ""
    i = 0
    while i < len(text):
        # 偵測 \n\n+(連續空行)
        m = re.match(r"\n{2,}", text[i:])
        if m:
            if current.strip():
                parts.append((current, "\n\n"))
            current = ""
            i += m.end()
            continue
        # 偵測單個 \n
        if text[i] == "\n":
            if current.strip():
                parts.append((current, "\n"))
                current = ""
            i += 1
            continue
        current += text[i]
        i += 1
    
    if current.strip():
        parts.append((current, ""))
    
    return parts


def _translate_single_paragraph(text, src, tgt):
    """v3.7 單段翻譯(供分段翻譯使用,跳過分段邏輯避免無窮迴圈)
    
    這函數做最小流程:cache → custom example → translate_openai
    不做反譯、self-check 等(那些是整篇譯文層級的檢查)
    """
    # 1. Check exact custom example
    exact = _check_custom_example_exact(text.strip(), src, tgt)
    if exact:
        return exact
    
    # 2. Check cache
    cached = cache_get(text, src, tgt)
    if cached:
        return cached
    
    # 3. Translate via OpenAI(用 retry,確保穩定)
    result = translate_with_retry(translate_openai, text, src, tgt, max_retries=2)
    if result:
        result = finalize_factory_translation(text, result, src, tgt)
        cache_set(text, src, tgt, result)
    return result


def _translate_paragraphs_separately(text, src, tgt, translate_fn):
    """分段獨立翻譯,然後用原分隔符拼回
    
    Args:
      text: 原文
      src, tgt: 語言對
      translate_fn: 實際翻譯函數(通常是 translate_openai)
    
    Returns:
      拼接後的譯文(段落結構與原文一致),失敗時 None
    """
    parts = _split_into_paragraphs(text)
    if len(parts) <= 1:
        return None  # 沒分段,讓主流程處理
    
    logger.info(
        "Paragraph-split translation: %d paragraphs, lengths=%s",
        len(parts), [len(p[0]) for p in parts]
    )
    
    translated_parts = []
    for idx, (para_text, sep) in enumerate(parts):
        para_clean = para_text.strip()
        if not para_clean:
            translated_parts.append(("", sep))
            continue
        
        try:
            tr = translate_fn(para_clean, src, tgt)
            if not tr:
                # 任一段失敗,放棄分段策略,讓主流程整段重翻
                logger.warning(
                    "Paragraph-split: paragraph %d/%d failed, falling back",
                    idx + 1, len(parts)
                )
                return None
            translated_parts.append((tr.strip(), sep))
        except Exception as e:
            logger.error("Paragraph-split error on paragraph %d: %s", idx + 1, e)
            return None
    
    # 拼接
    result = ""
    for tr_text, sep in translated_parts:
        result += tr_text + sep
    
    return result.strip()


def _translate_inner(text, src, tgt):
    # ★ v3.6 印尼文預處理:在所有後續路徑之前先標準化
    _preprocessing_log = None
    if src == "id" and id_preprocessing_enabled:
        normalized, n_replacements = normalize_indonesian_text(text)
        if n_replacements > 0:
            logger.info("ID preprocessing: %d replacements. %r → %r", 
                       n_replacements, text[:60], normalized[:60])
            _preprocessing_log = {"original": text, "normalized": normalized, "count": n_replacements}
            # 把標準化結果存到 thread-local 供日誌記錄
            try:
                _tl.id_preprocessing = _preprocessing_log
            except Exception:
                pass
            # 用標準化版本後續處理
            text = normalized
        # 進階:nano 模型語意級規範化(成本高)
        if id_preprocessing_nano:
            text_nano = normalize_indonesian_text_with_nano(text)
            if text_nano and text_nano != text:
                text = text_nano
    
    # ★ v3.7 段落結構保留:訊息含分段時走分段翻譯路徑
    # 這個路徑不影響短訊息(沒分段就直接走原本流程)
    # v3.9.27: 來自圖片 OCR 的文字一律走分段路徑(段落保留是首要需求)
    _is_from_image = getattr(_tl, 'from_image_ocr', False)
    if paragraph_split_translate and _has_paragraph_structure(text):
        char_len = len(re.sub(r"\s+", "", text))
        # v3.9.27: 圖片 OCR 不看 threshold,直接分段
        if _is_from_image or char_len >= paragraph_split_threshold:
            try:
                # 用內部 helper 來處理單段(會走 cache + custom_example)
                def _single_para_translate(para, s, t):
                    # 重新進入 _translate_inner 但跳過分段路徑(避免無窮迴圈)
                    return _translate_single_paragraph(para, s, t)
                
                multi_result = _translate_paragraphs_separately(text, src, tgt, _single_para_translate)
                if multi_result:
                    logger.info("Paragraph-split SUCCESS: orig=%d chars, result=%d chars",
                                len(text), len(multi_result))
                    return multi_result
                else:
                    logger.info("Paragraph-split returned None, falling back to single-shot")
            except Exception as _e:
                logger.error("Paragraph-split exception: %s", _e)
    
    # Check custom examples for exact match first (free, no API call)
    exact = _check_custom_example_exact(text.strip(), src, tgt)
    if exact:
        logger.info("Custom example exact match: %s -> %s", src, tgt)
        return exact

    # Deterministic factory semantic engine before cache/OpenAI.
    # This prevents short factory defect/direction reports from being treated as daily language.
    if src == "id" and tgt == "zh":
        semantic = factory_semantic_translate_id_zh(text)
        if semantic:
            logger.info("Factory semantic translation hit: %r -> %r", text[:80], semantic[:80])
            _tl.factory_audit = {
                "src": text,
                "type": "factory_semantic_direct",
                "reason": "deterministic_factory_slot_translation",
                "raw_translation": "",
                "corrected_translation": semantic,
                "domain": detect_factory_domain(text, src, tgt).get("domains", []),
                "auto_corrected": False,
            }
            _log_translation(text, semantic, src, tgt, "factory-semantic", 0, 1.0, False, 1.0, getattr(_tl, 'group_id', ''))
            cache_set(text, src, tgt, semantic)
            return semantic

    if src == "zh" and tgt == "id":
        semantic = factory_semantic_translate_zh_id(text)
        if semantic:
            logger.info("Factory ZH->ID semantic translation hit: %r -> %r", text[:80], semantic[:80])
            _tl.factory_audit = {
                "src": text,
                "type": "factory_semantic_direct_zh_id",
                "reason": "deterministic_factory_zh_id_translation",
                "raw_translation": "",
                "corrected_translation": semantic,
                "domain": detect_factory_domain(text, src, tgt).get("domains", []),
                "auto_corrected": False,
            }
            _log_translation(text, semantic, src, tgt, "factory-semantic-zh-id", 0, 1.0, False, 1.0, getattr(_tl, 'group_id', ''))
            cache_set(text, src, tgt, semantic)
            return semantic

    # Check cache first
    cached = cache_get(text, src, tgt)
    if cached:
        cached = finalize_factory_translation(text, cached, src, tgt)
        _log_translation(text, cached, src, tgt, "cache", 0, 1.0, False, 1.0, getattr(_tl, 'group_id', ''))
        return cached

    # ★ v3.4:長 ID→ZH 訊息嘗試 pivot via English(若啟用)
    pivot_result = None
    if should_use_pivot(text, src, tgt):
        try:
            pivot_result = translate_id_zh_with_pivot(text, src, tgt)
        except Exception as _e:
            logger.warning("Pivot attempt failed: %s", _e)
    
    result = translate_with_retry(translate_openai, text, src, tgt, max_retries=2)
    
    # ★ v3.4:雙翻 ensemble - 比較 pivot 和直譯,選較完整的
    if pivot_result and result:
        result = select_better_translation(result, pivot_result, text)
    if result:
        result = finalize_factory_translation(text, result, src, tgt)

    # If source-language leakage is detected, retry with strict mode.
    if result and not is_translation_valid(result, src, tgt):
        logger.warning("Source-language leakage detected in translation, retrying with stricter prompt")
        strict_result = translate_openai(text, src, tgt, strict_no_source_script=True)
        if strict_result:
            strict_result = finalize_factory_translation(text, strict_result, src, tgt)
        if strict_result and is_translation_valid(strict_result, src, tgt):
            result = strict_result
        else:
            repaired = translate_openai(
                text,
                src,
                tgt,
                strict_no_source_script=True,
                repair_mode=True,
                bad_result=(strict_result or result)
            )
            if repaired:
                repaired = finalize_factory_translation(text, repaired, src, tgt)
            if repaired and is_translation_valid(repaired, src, tgt):
                result = repaired

    if result and is_translation_valid(result, src, tgt):
        result = finalize_factory_translation(text, result, src, tgt)
        cache_set(text, src, tgt, result)
        return result

    # Fallback to Google with retry, then factory post-validation.
    # v3.9.3: DO NOT cache Google fallback results. If the OpenAI call failed
    # (e.g. wrong API params, rate limit), Google will give a literal direct
    # translation that ignores all custom examples. Caching that result poisons
    # subsequent lookups for 1 hour even after the OpenAI bug is fixed.
    # Instead, return the Google result without caching, so a retry can hit OpenAI.
    result = translate_with_retry(translate_google, text, src, tgt, max_retries=1)
    if result:
        result = finalize_factory_translation(text, result, src, tgt)
    if result and is_translation_valid(result, src, tgt):
        # cache_set(text, src, tgt, result)  # ← intentionally disabled
        logger.warning("Used Google fallback (NOT cached); check OpenAI errors above")
        return result

    # Last chance: deterministic semantic fallback before returning None.
    if src == "id" and tgt == "zh":
        semantic = factory_semantic_translate_id_zh(text)
        if semantic:
            cache_set(text, src, tgt, semantic)
            return semantic

    # Last chance: ask OpenAI to repair the latest output instead of returning a leaked translation.
    if result:
        repaired = translate_openai(
            text,
            src,
            tgt,
            strict_no_source_script=True,
            repair_mode=True,
            bad_result=result
        )
        if repaired:
            repaired = finalize_factory_translation(text, repaired, src, tgt)
        if repaired and is_translation_valid(repaired, src, tgt):
            cache_set(text, src, tgt, repaired)
            return repaired

    return None

def detect_work_order(ocr_text):
    """Detect if OCR text is from a factory work order (製造指示書).
    Returns customer name if detected, None otherwise."""
    if not ocr_text:
        return None
    wo_keywords = ["冷精棒製造指示書", "製造指示書", "訂單編號", "客戶名稱", "成品尺寸",
                   "FINAL流程", "FINAL", "MIC_NO", "ID_NO", "HRITABPDIL", "退火代碼",
                   "冷精棒", "收貨人", "短尺", "品保", "特殊", "削皮", "訂單資訊",
                   "成品尺寸MIN", "成品尺寸MAX", "製造指示"]
    keyword_count = sum(1 for kw in wo_keywords if kw in ocr_text)
    logger.info("Work order detection: %d keywords matched in OCR text (%d chars)", keyword_count, len(ocr_text))
    if keyword_count < 2:
        return None
    # Try multiple patterns to extract customer name
    patterns = [
        r'客戶名稱[:\s：]*([^\s\n|,，]+)',
        r'客戶[:\s：]*([^\s\n|,，]+)',
        r'客[户戶]名[称稱][:\s：]*([^\s\n|,，]+)',
    ]
    for pat in patterns:
        m = re.search(pat, ocr_text)
        if m:
            customer = m.group(1).strip()
            if customer and len(customer) >= 2:
                logger.info("Work order customer detected: %s", customer)
                return customer
    # Fallback: try to match any known customer name in the text
    for name in CUSTOMER_NAMES:
        if len(name) >= 2 and name in ocr_text:
            logger.info("Work order customer matched from list: %s", name)
            return name
    logger.info("Work order detected but no customer name found")
    return None


def format_length_zh(code):
    """Convert length code to Chinese."""
    if code == "<=3200":
        return "未滿3200"
    elif code == ">4200":
        return "超過4200"
    elif code == ">3200<=4200":
        return "3200～4200"
    elif code == ">4000":
        return "超過4000"
    else:
        c = code.replace("<=", "未滿").replace(">=", "超過").replace(">", "超過").replace("<", "未滿")
        return c


def format_storage_for_work_order(customer_name):
    """Format storage lookup for work order image detection."""
    entries = STORAGE_LOOKUP.get(customer_name)
    if not entries:
        for key in STORAGE_LOOKUP:
            if key.lower() == customer_name.lower() or customer_name in key or key in customer_name:
                entries = STORAGE_LOOKUP[key]
                customer_name = key
                break
    if not entries:
        return None
    lines = []
    lines.append("\U0001f4cb \u5de5\u55ae\u5075\u6e2c")
    lines.append("\u5ba2\u6236\uff1a" + customer_name)
    lines.append("")
    lines.append("\U0001f4e6 \u5132\u5340\u67e5\u8a62")
    lines.append("=" * 18)
    for length, area in entries:
        zh = format_length_zh(length)
        lines.append(zh + " \u2192 " + area)
    lines.append("=" * 18)
    return "\n".join(lines)


# v3.9.29: 已移除此處的孤兒 OCR 死碼(原 line 5078-5135)
# 該段沒有 def header,實際永遠不會執行,且使用未定義的 image_base64
# 真正的 OCR 函式是 ocr_image_openai() (line 5198 附近)


# v3.8: Vision model upgrade. gpt-4o-mini → gpt-5-mini for OCR.
# gpt-5-mini has noticeably better small-text + handwriting recognition,
# critical for factory work-order photos and shift schedule snapshots.
# Auto-fallback to gpt-4o-mini if primary unavailable.
# v3.9 (2026-05): made into a runtime-mutable global so admin panel can change it
# without redeploy, matching the behaviour of model_default / model_upgrade for translation.
# GPT-5.5 (released 2026-04-23) is also supported here when API access is enabled
# on your tier; it auto-falls back to gpt-5-mini → gpt-4o-mini on failure.
VISION_MODEL = os.environ.get("VISION_MODEL", "gpt-5-mini")
VISION_FALLBACK_MODEL = "gpt-4o-mini"


def _vision_call(messages, max_tokens, cache_key=None):
    """v3.9.28: 完全用 model_supports 過濾參數,符合 GPT-5 系列官方規格。
    
    關鍵差異(對比一般翻譯任務):
    - verbosity = "high"  → OCR 要忠實轉錄,不能壓縮
      (官方 GPT-5 cookbook: "Raise verbosity when you need
       faithful transcription rather than compressed summaries")
    - reasoning_effort = optimal_for_translation (minimal / none)
      → OCR 純抄字,不需要 reasoning
    - timeout=30 雙保險
    
    Tries VISION_MODEL first, falls back to gpt-4o-mini on failure.
    """
    last_err = None
    primary = VISION_MODEL
    for attempt_model in (primary, VISION_FALLBACK_MODEL):
        if attempt_model == VISION_FALLBACK_MODEL and primary == VISION_FALLBACK_MODEL:
            break
        try:
            kwargs = {
                "model": attempt_model,
                "messages": messages,
                "timeout": 30,
            }
            # === 全部用 model_supports 過濾,確保 GPT-5 系列相容 ===
            # Token limit
            # v3.9.30b B11 修補: 之前寫死 3000,reasoning model 會被 reasoning tokens 吃光
            # 改成依 reasoning model 額外加 4K reasoning 緩衝
            if model_supports(attempt_model, "max_completion_tokens"):
                # OCR 預設輸出 3000;若是 reasoning model 加 4K reasoning 預算
                _vision_budget = 3000
                if not model_supports(attempt_model, "temperature"):
                    # 是 reasoning model,額外給 reasoning 預算
                    _vision_budget = 3000 + 4000
                kwargs["max_completion_tokens"] = _vision_budget
            elif model_supports(attempt_model, "max_tokens"):
                kwargs["max_tokens"] = max_tokens
            # Temperature(GPT-5 系列不支援)
            if model_supports(attempt_model, "temperature"):
                kwargs["temperature"] = 0.0
            # Reasoning effort(GPT-5/o-series 才有)
            if model_supports(attempt_model, "reasoning_effort"):
                _opt = optimal_reasoning_for_translation(attempt_model)
                if _opt:
                    kwargs["reasoning_effort"] = _opt
            # Verbosity — OCR 用 HIGH(官方建議:忠實轉錄)
            if model_supports(attempt_model, "verbosity"):
                kwargs["verbosity"] = "high"
            
            try:
                _event_log_write("vision_call_start", {
                    "model": attempt_model,
                    "kwargs_keys": list(kwargs.keys()),
                })
            except Exception:
                pass
            
            logger.info("[Vision] calling %s with kwargs=%s", attempt_model, list(kwargs.keys()))
            r = oai.chat.completions.create(**kwargs)
            
            if attempt_model != primary:
                logger.warning("[Vision] fell back from %s to %s", primary, attempt_model)
            
            try:
                _content = r.choices[0].message.content if r.choices else None
                _finish = r.choices[0].finish_reason if r.choices else None
                _usage_dict = r.usage.model_dump() if hasattr(r, 'usage') and r.usage else None
                _event_log_write("vision_call_done", {
                    "model": attempt_model,
                    "finish": _finish,
                    "content_len": len(_content) if _content else 0,
                    "usage": _usage_dict,
                })
                logger.info("[Vision] model=%s ok, finish=%s, content_len=%d",
                            attempt_model, _finish, len(_content) if _content else 0)
                if _finish == "length":
                    logger.warning("[Vision] WARNING: finish_reason=length, output truncated!")
            except Exception as _le:
                logger.warning("[Vision] logging error: %s", _le)
            return r
        except Exception as e:
            last_err = e
            try:
                _event_log_write("vision_call_failed", {
                    "model": attempt_model,
                    "error": str(e)[:300],
                    "error_type": type(e).__name__,
                })
            except Exception:
                pass
            logger.warning("[Vision] model %s failed: %r", attempt_model, e)
            continue
    raise last_err if last_err else RuntimeError("vision call failed")


def _clean_ocr_status_bar(text):
    """v3.9.29: 改用 token-based 偵測,精準移除手機螢幕狀態列。
    
    判斷邏輯:把每行拆成 tokens,如果**所有 tokens 都是狀態元素**就移除整行。
    只要有一個 token 不是狀態元素就保留(避免誤殺正常內容)。
    
    狀態元素:
    - 時間: "09:40", "12:43", "23:59"  
    - 訊號: "4G", "5G", "Wi-Fi", "wifi", "LTE", "VoLTE"
    - 電量: "99+", "100%", "63%", "<99+", ">99+"
    - LINE 介面: "已讀 1/2/...", "輸入訊息", "換行", "中/英/符"
    - 注音
    """
    if not text:
        return text
    
    import re
    
    # 單一 token 是否為狀態元素
    def is_status_token(tok):
        if not tok:
            return True  # 空 token 視為狀態(允許)
        # 純時間
        if re.match(r'^\d{1,2}:\d{2}$', tok):
            return True
        # 訊號
        if re.match(r'^[345]G$', tok):
            return True
        if re.match(r'^(Wi[-]?Fi|wifi|WIFI|LTE|VoLTE|HSPA|5G\+|5GE)$', tok, re.I):
            return True
        # 純數字 / 99+ / 帶 + 號
        if re.match(r'^\d{1,3}\+?$', tok):
            return True
        # 百分比
        if re.match(r'^\d{1,3}%$', tok):
            return True
        # <99+ / >99+
        if re.match(r'^[<>]\d{1,3}\+?$', tok):
            return True
        return False
    
    # 整行匹配的特殊狀態
    line_status_patterns = [
        r'^已讀\s*\d+$',
        r'^輸入訊息$',
        r'^換行$',
        r'^(中|英|符|空白鍵|語音輸入|注音|倉頡|嘸蝦米)$',
        r'^[\u3105-\u3129]+$',  # 純注音字母
    ]
    
    cleaned_lines = []
    for line in text.split('\n'):
        stripped = line.strip()
        if not stripped:
            cleaned_lines.append(line)  # 保留空行(段落分隔)
            continue
        
        # 整行匹配檢查
        is_status_line = False
        for pat in line_status_patterns:
            if re.match(pat, stripped):
                is_status_line = True
                break
        
        if not is_status_line:
            # Token-based 檢查:用空白切 tokens,看是否每個都是狀態元素
            tokens = re.split(r'\s+', stripped)
            tokens = [t for t in tokens if t]  # 過濾空字串
            if tokens and all(is_status_token(t) for t in tokens):
                is_status_line = True
        
        if not is_status_line:
            cleaned_lines.append(line)
    
    return '\n'.join(cleaned_lines)


def ocr_image_openai(image_base64, mime_type="image/jpeg"):
    """Use OpenAI Vision to extract text from image. v3.8: model upgraded.
    
    v3.9.14: GPT-5 vision OCR 改良:
      - prompt 改為更明確的指令,避免 GPT-5 過度解讀回 NO_TEXT_FOUND
      - 失敗時詳細記錄 raw response,方便偵錯
    v3.9.17: 加 mime_type 參數 — 不能再寫死 jpeg,LINE 可能傳 PNG/HEIC
    """
    if not oai:
        return None
    # v3.9.17: HEIC OpenAI 不支援,改用 jpeg(雖然會失敗但至少 OpenAI 會給明確錯誤)
    if mime_type == "image/heic":
        logger.warning("[OCR] HEIC not supported by OpenAI, sending as jpeg (may fail)")
        mime_type = "image/jpeg"
    try:
        msgs = [
                {
                    "role": "system",
                    "content": (
                        "你是一個 OCR 引擎。任務:把圖片中所有可見的訊息文字,逐字輸出。\n"
                        "規則:\n"
                        "1. **嚴格保留段落結構**:原文中的空行(段落分隔)用空行輸出;原文中的換行用換行輸出。\n"
                        "2. 不要加註解、不要翻譯、不要總結、不要編號\n"
                        "3. 中文、英文、數字、印尼文、日文都原樣輸出\n"
                        "4. **忽略手機螢幕介面元素**:狀態列(時間 / 4G / 5G / WiFi / 電量百分比 / 訊號)、\n"
                        "   未讀數(99+)、輸入框文字(輸入訊息)、鍵盤按鍵(注音、空白鍵、換行等)、\n"
                        "   應用程式名稱、底部 navigation bar — 這些都不要輸出\n"
                        "5. 只輸出訊息內容、貼文、文件、招牌、文章等實質文字\n"
                        "6. 如果圖片中真的完全沒有任何訊息文字,才輸出 NO_TEXT"
                    )
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64," + image_base64,
                                "detail": "high"
                            }
                        },
                        {
                            "type": "text",
                            "text": "請輸出這張圖中的訊息文字內容(忽略手機介面元素),嚴格保留段落結構。"
                        }
                    ]
                }
            ]
        r = _vision_call(msgs, max_tokens=2000,
                         cache_key=_build_cache_key(getattr(_tl, 'group_id', ''),
                                                    "img", "txt", "ocr"))
        track_tokens(r)
        result = (r.choices[0].message.content or "").strip()
        # v3.9.14: 詳細 logging
        logger.info("[OCR] raw result (%d chars): %s", len(result), result[:200])
        # v3.9.14: 多種「沒文字」回應的辨識
        no_text_markers = ("NO_TEXT_FOUND", "NO_TEXT", "沒有文字", "no text",
                           "圖片中沒有", "圖中沒有")
        result_lower = result.lower()
        if not result:
            logger.warning("[OCR] empty result")
            return None
        # 結果過短且看起來像「沒文字」回應
        if len(result) < 30:
            for marker in no_text_markers:
                if marker.lower() in result_lower:
                    logger.info("[OCR] detected no-text marker: %s", marker)
                    return None
        # v3.9.27: 後處理 — 清掉殘留的狀態列文字(雙重保險)
        result = _clean_ocr_status_bar(result)
        # 清完後若空,當沒文字
        if not result.strip():
            logger.info("[OCR] all content cleaned as status bar")
            return None
        return result
    except Exception as e:
        logger.exception("[OCR] OpenAI Vision OCR error: %s", e)
        return None


def ocr_and_translate_image(image_base64, tgt_lang):
    """OCR + translate image text in one API call, preserving layout."""
    if not oai:
        return None, None
    tgt_name = LANG_NAMES.get(tgt_lang, tgt_lang)
    tgt_flag = LANG_FLAGS.get(tgt_lang, "")
    try:
        msgs = [
                {
                    "role": "system",
                    "content": (
                        "You are an OCR + translation assistant for a factory work group chat.\n"
                        "Task: Extract ALL text from the image, then translate each section.\n\n"
                        "OUTPUT FORMAT:\n"
                        "For each distinct section/paragraph in the image, output:\n"
                        "original text...\n"
                        + tgt_flag + " translated text...\n"
                        "(blank line before next section)\n\n"
                        "EXAMPLE:\n"
                        "1.研磨來料前需紀錄來料三點式尺寸\n"
                        + tgt_flag + " 1.Sebelum grinding material masuk, catat dimensi 3 titik\n\n"
                        "2.拋光棒需清洗\n"
                        + tgt_flag + " 2.Batang polishing harus dicuci\n\n"
                        "RULES:\n"
                        "1. Keep the SAME structure, numbering, and line breaks as the original.\n"
                        "2. Each section: original text first, then translation with " + tgt_flag + " flag. Do NOT add section titles or brackets.\n"
                        "3. If there are numbered items (1. 2. 3.), keep the same numbering.\n"
                        "4. Do NOT repeat the original text. Show it only ONCE then show the translation.\n"
                        "5. Translate naturally, casual daily language for factory workers.\n"
                        "6. Target Traditional Chinese = Taiwan style.\n"
                        "7. NEVER translate or romanize person names. Keep Chinese names in original Chinese characters (e.g. 陳弘林 stays as 陳弘林, NOT Chen Honglin). Do NOT convert to pinyin.\n"
                        "7b. NEVER translate customer/company names. Keep them EXACTLY as-is: "
                        "賽利金屬, 寶麗金屬, 田華榕, 佳東, 蘋果, 常州眾山, 大順, 大成, 巨昌, 北澤, 鴻運, 畯圓, 名威, 右勝, "
                        "貝克休斯, 皇銘, 台芝, 百堅, 津展, 曜麟, 廉錩, 盛昌遠, 永吉, 光輝, "
                        "DACAPO, CASTLE, LOTUS, METALINOX, KANGRUI, SUNGEUN, STEELINC, GLH, SHINKO, WING KEUNG, "
                        "BOLLINGHAUS, COGNE, TCI, PLUTUS, SAMWON, DK METAL, KJ. "
                        "If you see ANY company name in the image, keep it unchanged. Do NOT translate 金屬=metal, 鋼鐵=steel etc. when part of a company name.\n"
                        "8. If no text found, output exactly: NO_TEXT_FOUND\n"
                        "9. TABLES/SPREADSHEETS: If the image is a table or spreadsheet, output it as a COMPACT table. "
                        "Only translate column headers and labels. Keep person names as-is in original characters. "
                        "Keep numbers as-is. Use a simple format like:\n"
                        "姓名/Nama | 3/17止/Hingga 3/17\n"
                        "陳弘林 | -600\n"
                        "蔡佳佳 | 200\n"
                        "Do NOT output each cell as a separate translated section. Keep it compact.\n"
                        "10. Factory vocabulary: "
                        "交辦事項=hal yang harus dikerjakan, "
                        "研磨=grinding, 無心研磨=centerless grinding, 拋光=polishing, 來料=material masuk, "
                        "量測=mengukur, 尺寸=diameter/dimensi, 三點式=3 titik, "
                        "雷射=laser, 設備=peralatan, 故障=rusak, "
                        "紀錄=catat, 拋光棒=batang polishing, "
                        "清洗=cuci, 輕調輕放=handle dengan hati-hati, "
                        "環狀擦傷=goresan melingkar, "
                        "重工=rework, 料回削皮=material kembali kupas/peeling, "
                        "補上=lengkapi, C行套環=C-ring, "
                        "廠內=di dalam pabrik, 禁止=dilarang, 餵狗=kasih makan anjing, "
                        "宣導=sosialisasi, "
                        "包裝站=stasiun packing, 啟動=mulai, "
                        "PMI全檢=inspeksi penuh PMI, 抽查機制=sistem sampling, "
                        "每捆=setiap bundel, 鋼種=jenis baja, "
                        "棒材=batang baja, 混料=tercampur material, "
                        "出貨=pengiriman, 依情節=sesuai tingkat pelanggaran, "
                        "增加績效=tambah penilaian kinerja, "
                        "確實=pastikan, 防止=mencegah, "
                        "精整=finishing, AP=mesin finishing, 矯直=straightening, 壓光=press polish, "
                        "退火=annealing, 光輝退火=bright annealing, 酸洗=pickling, 削皮=peeling, 冷抽=cold drawing, "
                        "熱軋=hot rolling, 煉鋼=steelmaking/peleburan baja, 碳廠=pabrik karbon, "
                        "職安署=Dinas K3(inspeksi keselamatan kerja), 查核=audit/inspeksi, "
                        "品保=QC, 儲運=gudang&logistik, 生計=production planning, 業務=sales, 營業=sales, 人事=HRD, "
                        "處長=kepala divisi, 點名=inspeksi pengawas(NOT roll call), "
                        "加班=lembur, 排班=jadwal shift, 早班=shift pagi, 夜班=shift malam, "
                        "砂輪=batu gerinda, 天車=overhead crane, 堆高機=forklift, "
                        "油桶=drum oli, 太空包=jumbo bag, 噴漆罐=kaleng spray, "
                        "入庫=masuk gudang, 退庫=kembalikan ke gudang, 出貨差=kekurangan pengiriman, "
                        "掛單/工單=work order, 重掛單=pasang ulang work order, 取樣=ambil sampel, "
                        "二道門=pintu kedua(gate 2), 捐血=donor darah, "
                        "爐號=heat number(NEVER nomor panas), 過帳=input data ke sistem, 放行=release data\n"
                        "11. Only output the result. No extra explanation."
                    )
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": "data:image/jpeg;base64," + image_base64,
                                "detail": "high"
                            }
                        },
                        {
                            "type": "text",
                            "text": "Extract and translate all text from this image to " + tgt_name + ". Keep the same layout structure."
                        }
                    ]
                }
            ]
        r = _vision_call(msgs, max_tokens=3000,
                         cache_key=_build_cache_key(getattr(_tl, 'group_id', ''),
                                                    "img", tgt_lang, "ocrtrans"))
        track_tokens(r)
        # v3.9.30 B7 修補: vision model 也可能回 content=None
        _content = r.choices[0].message.content if r.choices else None
        if _content is None:
            _finish = r.choices[0].finish_reason if r.choices else "unknown"
            logger.warning(
                "[OCR+translate] empty content, finish_reason=%s", _finish
            )
            return None, f"empty_content:{_finish}"
        result = _content.strip()
        if result == "NO_TEXT_FOUND" or not result:
            return None, None
        return result, None
    except Exception as e:
        logger.error("OpenAI Vision OCR+translate error: %s", e)
        return None, str(e)



def detect_image_mime(raw_bytes):
    """v3.9.29: 從 magic bytes 偵測圖片格式。
    回傳 'image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/heic' 之一。
    無法辨識時回 'image/jpeg'(safest fallback)。
    
    v3.9.29 修正:之前 hard-code 12 bytes 最低長度,導致 8-byte 的 PNG signature
    跟 6-byte 的 GIF 被當作未知。改成依格式分別判斷。
    """
    if not raw_bytes:
        return "image/jpeg"
    # JPEG: 3 bytes
    if len(raw_bytes) >= 3 and raw_bytes[:3] == b'\xff\xd8\xff':
        return "image/jpeg"
    # PNG: 8 bytes
    if len(raw_bytes) >= 8 and raw_bytes[:8] == b'\x89PNG\r\n\x1a\n':
        return "image/png"
    # GIF: 6 bytes
    if len(raw_bytes) >= 6 and raw_bytes[:6] in (b'GIF87a', b'GIF89a'):
        return "image/gif"
    # WebP: 12 bytes (RIFF...WEBP)
    if len(raw_bytes) >= 12 and raw_bytes[:4] == b'RIFF' and raw_bytes[8:12] == b'WEBP':
        return "image/webp"
    # HEIC/HEIF: 12 bytes (offset 4)
    if len(raw_bytes) >= 12 and raw_bytes[4:12] in (
        b'ftypheic', b'ftypheix', b'ftyphevc', b'ftypheim',
        b'ftypheis', b'ftyphevm', b'ftyphevs', b'ftypmif1'
    ):
        logger.warning("[Vision] HEIC/HEIF image detected, OpenAI may reject it")
        return "image/heic"
    logger.warning("[Vision] unknown image format, first bytes: %s",
                   raw_bytes[:16].hex() if raw_bytes else "<empty>")
    return "image/jpeg"  # safest fallback


def download_line_image(message_id):
    """Download image from LINE and return (base64_string, raw_bytes)."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            content = blob_api.get_message_content(message_id)
            # v3.9.17: 確保 content 是 bytes(不是 bytearray 或其他)
            if isinstance(content, bytearray):
                content = bytes(content)
            elif not isinstance(content, bytes):
                # 如果是 file-like object,讀出來
                if hasattr(content, 'read'):
                    content = content.read()
                else:
                    content = bytes(content)
            img_base64 = base64.b64encode(content).decode("utf-8")
            return img_base64, content
    except Exception as e:
        logger.error("LINE image download error: %s", e)
        return None, None


def download_line_audio(message_id):
    """Download audio from LINE and return bytes."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            content = blob_api.get_message_content(message_id)
            return content
    except Exception as e:
        logger.error("LINE audio download error: %s", e)
        return None


# v3.8: STT model upgrade. whisper-1 (2022) → gpt-4o-transcribe (2025).
# Lower WER, better noise robustness, semantic VAD, supports prompt hint.
# Critical for noisy factory-floor recordings with Indonesian-accented speech.
STT_MODEL = os.environ.get("STT_MODEL", "gpt-4o-transcribe")  # or gpt-4o-mini-transcribe (cheaper, slightly less accurate)
STT_FALLBACK_MODEL = "whisper-1"  # auto-fallback if primary unavailable

# Factory vocabulary hint passed to STT to bias recognition toward
# domain-specific terms, station numbers, and known names. Keep <224 tokens.
STT_PROMPT_HINT = (
    "工廠語音記錄。常見詞彙：無心研磨、研磨、拋光、削皮、退火、酸洗、冷抽、矯直、精整、"
    "棒鋼、盤元、不鏽鋼、削皮棒、冷精棒、料號、爐號、鋼種、混料、出貨、來料、"
    "400站、401站、420站、470站、480站、490站、801站、UT、PMI、AP、OL、"
    "工單、掛單、放行、過帳、退庫、入庫、無主、改制、去化、"
    "早班、夜班、中班、加班、點名、班股、堆高機、天車、"
    "Bahasa Indonesia istilah pabrik: as, kopel, patah, bengkok, retak, aus, bocor, "
    "macet, bearing, rantai, sabuk, baut, mur, kawat, motor, pompa, gerinda, las, bor, "
    "shift pagi, shift malam, lembur, work order."
)


def transcribe_audio_openai(audio_bytes):
    """v3.8: Use gpt-4o-transcribe (or fallback to whisper-1) to transcribe audio.

    Improvements over the old whisper-1 path:
      * gpt-4o-transcribe / gpt-4o-mini-transcribe have noticeably lower WER on
        noisy environments and accented speech (factory floor + Indonesian accent).
      * `prompt` parameter biases the model toward our factory vocabulary,
        station numbers, and Indonesian mechanical terms.
      * Auto-fallback to whisper-1 if the new model isn't available on the account.
    """
    if not oai:
        return None
    primary = STT_MODEL
    last_err = None
    for attempt_model in (primary, STT_FALLBACK_MODEL):
        if attempt_model == STT_FALLBACK_MODEL and primary == STT_FALLBACK_MODEL:
            break  # already tried as primary
        try:
            with tempfile.NamedTemporaryFile(suffix=".m4a", delete=True) as tmp:
                tmp.write(audio_bytes)
                tmp.flush()
                tmp.seek(0)
                kwargs = {
                    "model": attempt_model,
                    "file": tmp,
                    # v3.9.30c B16 修補: STT 也加 30 秒 timeout(其他 OpenAI 呼叫都有,只有這個沒)
                    "timeout": 30,
                }
                # Only the new gpt-4o-transcribe family accepts the `prompt`
                # vocabulary-bias parameter via the standard transcriptions endpoint.
                if attempt_model.startswith("gpt-4o"):
                    kwargs["prompt"] = STT_PROMPT_HINT
                r = oai.audio.transcriptions.create(**kwargs)
                # v3.9.30c B16 修補: 防 r 為 None / r.text 為 None
                if r is None:
                    logger.warning("STT model %s returned None", attempt_model)
                    continue
                text = (getattr(r, 'text', None) or "").strip() or None
                if attempt_model != primary:
                    logger.warning("STT fell back from %s to %s", primary, attempt_model)
                return text
        except Exception as e:
            last_err = e
            logger.warning("STT model %s failed: %s", attempt_model, e)
            continue
    logger.error("OpenAI STT error (all models failed): %s", last_err)
    return None


def make_notice(content, target="id"):
    tgt_text = translate(content, "zh", target)
    if not tgt_text:
        tgt_text = "(translation failed)"
    sep = "=" * 18
    lines = []
    lines.append("\U0001f4e2 \u516c\u544a / Pengumuman")
    lines.append(sep)
    lines.append("\U0001f1f9\U0001f1fc " + content)
    lines.append(LANG_FLAGS.get(target, "") + " " + tgt_text)
    lines.append(sep)
    return "\n".join(lines)


def make_notice_from_other(content, src, target="zh"):
    zh_text = translate(content, src, "zh")
    if not zh_text:
        zh_text = "(translation failed)"
    sep = "=" * 18
    lines = []
    lines.append("\U0001f4e2 \u516c\u544a / Pengumuman")
    lines.append(sep)
    lines.append("\U0001f1f9\U0001f1fc " + zh_text)
    lines.append(LANG_FLAGS.get(src, "") + " " + content)
    lines.append(sep)
    return "\n".join(lines)


def get_help_text(group_id):
    sep = "=" * 18
    lines = []
    lines.append("\U0001f310 翻譯機器人")
    lines.append(sep)
    lines.append("【開關】")
    lines.append("/on ・ /off 翻譯")
    lines.append("/img on・off 圖片")
    lines.append("/voice on・off 語音")
    lines.append("/wo on・off 拍工單查儲區")
    lines.append("【個人】")
    lines.append("/skip 不翻譯我")
    lines.append("/unskip 恢復翻譯")
    lines.append("【管理】")
    lines.append("/skipadd 名字 加入白名單")
    lines.append("/skipdel 名字 移出白名單")
    lines.append("/skiplist 查看白名單")
    lines.append("【功能】")
    lines.append("/notice 內容 雙語公告")
    lines.append("/qry 客戶 查儲區")
    lines.append("/pkg 代碼 查包裝碼")
    lines.append("/pw1 班長密碼")
    lines.append("/pw2 儲運密碼")
    lines.append("/scrap 廢料顏色")
    lines.append("/status 查看狀態")
    lines.append("\U0001f4f7 拍工單→自動查儲區")
    lines.append(sep)
    lines.append("中文 ⇄ 🇮🇩 印尼文 即時互譯")
    return "\n".join(lines)


# ======================================================================
# /help Flex Message (v3.2-0420b)
# Dark-themed bilingual help card: Carousel with ZH-TW (blue) + ID (red)
# Switch language via postback: action=help&lang=zh|id
# ======================================================================

HELP_COMMAND_SECTIONS = [
    {
        "num": "I", "zh_tag": "SWITCH", "zh_name": "開關",
        "id_tag": "SAKELAR", "id_name": "TOMBOL",
        "items": [
            ("/on · /off",   "開啟 / 關閉翻譯",       "Aktif / nonaktif"),
            ("/img on·off·ask",  "圖片翻譯(ask=按鈕詢問)", "Terjemahan gambar"),
            ("/voice on·off","語音翻譯",              "Terjemahan suara"),
            ("/wo on·off",   "拍工單查儲區",          "Foto WO cek gudang"),
        ],
    },
    {
        "num": "II", "zh_tag": "PERSONAL", "zh_name": "個人",
        "id_tag": "PRIBADI", "id_name": "PERSONAL",
        "items": [
            ("/skip",   "不翻譯我",   "Jangan terjemahkan saya"),
            ("/unskip", "恢復翻譯",   "Terjemahkan lagi"),
        ],
    },
    {
        "num": "III", "zh_tag": "ADMIN", "zh_name": "管理",
        "id_tag": "ADMIN", "id_name": "PENGELOLA",
        "items": [
            ("/skipadd 名字", "加入白名單",   "Tambah whitelist"),
            ("/skipdel 名字", "移出白名單",   "Hapus whitelist"),
            ("/skiplist",     "查看白名單",   "Lihat whitelist"),
            ("/wrong 譯文",   "標最新翻譯錯", "Tandai salah"),
            ("/wrong N 譯文", "標倒數第N筆",  "Tandai ke-N salah"),
            ("/wrong list",   "看最近10筆",   "10 terjemahan terakhir"),
            ("/export",       "訓練資料統計", "Statistik data"),
            ("/export jsonl", "匯出訓練資料", "Ekspor data training"),
        ],
    },
    {
        "num": "IV", "zh_tag": "FUNCTION", "zh_name": "功能",
        "id_tag": "FUNGSI", "id_name": "FITUR",
        "items": [
            ("/notice 內容",  "雙語公告",       "Pengumuman 2 bahasa"),
            ("/qry 客戶",     "查儲區",         "Cek gudang"),
            ("/pkg 代碼",     "查包裝碼",       "Cek kode kemasan"),
            ("/pw1",          "班長密碼",       "Password mandor"),
            ("/pw2",          "儲運密碼",       "Password gudang"),
            ("/scrap",        "廢料顏色",       "Warna scrap"),
            ("/status",       "查看狀態",       "Cek status"),
        ],
    },
]

# Color palette per language
_HELP_COLORS = {
    "zh": {
        "body_bg":    "#1A1D3A",  # 深夜藍
        "header_bg":  "#12142A",  # 更深藍黑
        "accent":     "#7FB3FF",  # 青藍（分組編號、header meta）
        "cmd_color":  "#FFFFFF",  # 指令代碼白
        "desc_color": "#A8B0C8",  # 說明淺藍灰
        "divider":    "#2A2D4A",  # 分隔線
        "info_bg":    "#252849",  # info strip 稍亮
        "info_text":  "#C5CCE0",
        "btn_bg":     "#7FB3FF",  # 切換鈕青藍底
        "btn_text":   "#0D0F1E",  # 按鈕深底字
        "plate_text": "#5A6080",
    },
    "id": {
        "body_bg":    "#3A1A1D",  # 深酒紅
        "header_bg":  "#2A1213",
        "accent":     "#FF8A95",  # 橘粉紅
        "cmd_color":  "#FFFFFF",
        "desc_color": "#C8A8AE",
        "divider":    "#4A2A2D",
        "info_bg":    "#492528",
        "info_text":  "#E0C5C9",
        "btn_bg":     "#FF8A95",
        "btn_text":   "#1E0D0F",
        "plate_text": "#805A60",
    },
}


def _help_cmd_row(cmd, desc, c):
    """Build a single command row: [cmd_code]  [description]"""
    return {
        "type": "box",
        "layout": "baseline",
        "spacing": "md",
        "paddingTop": "sm",
        "paddingBottom": "sm",
        "contents": [
            {
                "type": "text",
                "text": cmd,
                "size": "xs",
                "color": c["cmd_color"],
                "weight": "bold",
                "flex": 4,
                "wrap": False,
            },
            {
                "type": "text",
                "text": desc,
                "size": "xs",
                "color": c["desc_color"],
                "flex": 5,
                "wrap": True,
            },
        ],
    }


def _help_section(num, tag, name, items, lang):
    """Build one section: header row + divider + command rows."""
    c = _HELP_COLORS[lang]
    rows = []
    # Section header
    rows.append({
        "type": "box",
        "layout": "baseline",
        "spacing": "sm",
        "margin": "lg",
        "contents": [
            {
                "type": "text",
                "text": num + " / " + tag,
                "size": "xxs",
                "color": c["accent"],
                "weight": "bold",
                "flex": 0,
            },
            {
                "type": "text",
                "text": name,
                "size": "xs",
                "color": c["cmd_color"],
                "weight": "bold",
                "flex": 0,
                "margin": "md",
            },
            {"type": "filler"},
        ],
    })
    rows.append({
        "type": "separator",
        "color": c["divider"],
        "margin": "sm",
    })
    for i, (cmd, zh_desc, id_desc) in enumerate(items):
        desc = zh_desc if lang == "zh" else id_desc
        rows.append(_help_cmd_row(cmd, desc, c))
        if i < len(items) - 1:
            rows.append({"type": "separator", "color": c["divider"]})
    return rows


def _build_help_bubble(lang):
    """Build one Flex bubble for the given language ('zh' or 'id')."""
    c = _HELP_COLORS[lang]
    if lang == "zh":
        hdr_meta = "COMMAND REFERENCE · ZH-TW"
        hdr_title = "翻譯機器人"
        hdr_sub = "研磨股C班 · 不鏽鋼棒線部"
        info_line_1 = "📷 拍工單　直接傳照片自動查儲區"
        info_line_2 = "中文 ⇄ 印尼文　即時互譯・免指令"
        switch_btn_label = "BAHASA INDONESIA  ›"
        switch_lang = "id"
    else:
        hdr_meta = "DAFTAR PERINTAH · ID"
        hdr_title = "Bot Penerjemah"
        hdr_sub = "Grup Grinding C · Stainless Steel"
        info_line_1 = "📷 Foto WO　Kirim foto otomatis cek gudang"
        info_line_2 = "Mandarin ⇄ Indonesia　Terjemah langsung"
        switch_btn_label = "中文版 / MANDARIN  ›"
        switch_lang = "zh"

    body_contents = []
    for sect in HELP_COMMAND_SECTIONS:
        tag = sect["zh_tag"] if lang == "zh" else sect["id_tag"]
        name = sect["zh_name"] if lang == "zh" else sect["id_name"]
        body_contents.extend(_help_section(sect["num"], tag, name, sect["items"], lang))

    # Info strip
    body_contents.append({
        "type": "box",
        "layout": "vertical",
        "margin": "xl",
        "paddingAll": "md",
        "backgroundColor": c["info_bg"],
        "cornerRadius": "sm",
        "contents": [
            {
                "type": "text",
                "text": info_line_1,
                "size": "xxs",
                "color": c["info_text"],
                "wrap": True,
            },
            {
                "type": "text",
                "text": info_line_2,
                "size": "xxs",
                "color": c["info_text"],
                "wrap": True,
                "margin": "xs",
            },
        ],
    })

    bubble = {
        "type": "bubble",
        "size": "mega",
        "header": {
            "type": "box",
            "layout": "vertical",
            "backgroundColor": c["header_bg"],
            "paddingAll": "xl",
            "contents": [
                {
                    "type": "text",
                    "text": hdr_meta,
                    "size": "xxs",
                    "color": c["accent"],
                    "weight": "bold",
                },
                {
                    "type": "text",
                    "text": hdr_title,
                    "size": "xl",
                    "color": "#FFFFFF",
                    "weight": "bold",
                    "margin": "xs",
                },
                {
                    "type": "text",
                    "text": hdr_sub,
                    "size": "xxs",
                    "color": c["desc_color"],
                    "margin": "sm",
                },
            ],
        },
        "body": {
            "type": "box",
            "layout": "vertical",
            "backgroundColor": c["body_bg"],
            "paddingAll": "xl",
            "spacing": "none",
            "contents": body_contents,
        },
        "footer": {
            "type": "box",
            "layout": "vertical",
            "backgroundColor": c["body_bg"],
            "paddingTop": "md",
            "paddingBottom": "lg",
            "paddingStart": "xl",
            "paddingEnd": "xl",
            "contents": [
                {
                    "type": "separator",
                    "color": c["divider"],
                },
                {
                    "type": "button",
                    "style": "primary",
                    "color": c["btn_bg"],
                    "height": "sm",
                    "margin": "lg",
                    "action": {
                        "type": "postback",
                        "label": switch_btn_label,
                        "data": "action=help&lang=" + switch_lang,
                        "displayText": switch_btn_label,
                    },
                },
                {
                    "type": "box",
                    "layout": "horizontal",
                    "margin": "md",
                    "contents": [
                        {
                            "type": "text",
                            "text": "MODEL · " + VERSION.upper(),
                            "size": "xxs",
                            "color": c["plate_text"],
                            "align": "start",
                        },
                        {
                            "type": "text",
                            "text": "UNIT · GRINDING-C",
                            "size": "xxs",
                            "color": c["plate_text"],
                            "align": "end",
                        },
                    ],
                },
            ],
        },
        "styles": {
            "header": {"separator": False},
            "body": {"separator": False},
            "footer": {"separator": False},
        },
    }
    return bubble


def build_help_flex(primary_lang="zh"):
    """Build the complete help Flex carousel (2 bubbles)."""
    if primary_lang == "id":
        contents = [_build_help_bubble("id"), _build_help_bubble("zh")]
        alt_text = "🌐 Daftar Perintah Bot Penerjemah / 翻譯機器人指令"
    else:
        contents = [_build_help_bubble("zh"), _build_help_bubble("id")]
        alt_text = "🌐 翻譯機器人指令 / Daftar Perintah"
    carousel = {"type": "carousel", "contents": contents}
    return alt_text, carousel


def send_help_flex(reply_token, primary_lang="zh"):
    """Send help Flex carousel via reply token."""
    try:
        alt_text, carousel = build_help_flex(primary_lang)
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.reply_message(ReplyMessageRequest(
                reply_token=reply_token,
                messages=[FlexMessage(
                    alt_text=alt_text,
                    contents=FlexContainer.from_dict(carousel),
                )]
            ))
        return True
    except Exception as e:
        logger.exception("send_help_flex failed: %s", e)
        return False


def push_help_flex(to_id, primary_lang="zh"):
    """Push help Flex to a group/user (for postback language switch)."""
    try:
        alt_text, carousel = build_help_flex(primary_lang)
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.push_message(PushMessageRequest(
                to=to_id,
                messages=[FlexMessage(
                    alt_text=alt_text,
                    contents=FlexContainer.from_dict(carousel),
                )]
            ))
        return True
    except Exception as e:
        logger.exception("push_help_flex failed: %s", e)
        return False


def build_image_ask_flex(message_id):
    """v3.9.12: Build a Flex bubble asking whether to translate the received image.
    Same visual style as the help bubble (dark navy + cyan-blue accent).
    Single action button: 翻譯這張 (primary). 1 分鐘後自動過期(避免誤觸)。
    """
    c = _HELP_COLORS["zh"]
    bubble = {
        "type": "bubble",
        "size": "mega",
        "header": {
            "type": "box",
            "layout": "vertical",
            "backgroundColor": c["header_bg"],
            "paddingAll": "xl",
            "contents": [
                {
                    "type": "text",
                    "text": "IMAGE RECEIVED · 收到圖片",
                    "size": "xxs",
                    "color": c["accent"],
                    "weight": "bold",
                },
                {
                    "type": "text",
                    "text": "翻譯上面這張圖?",
                    "size": "xl",
                    "color": "#FFFFFF",
                    "weight": "bold",
                    "margin": "xs",
                },
                {
                    "type": "text",
                    "text": "Terjemahkan gambar di atas?",
                    "size": "xs",
                    "color": c["desc_color"],
                    "margin": "xs",
                },
            ],
        },
        "body": {
            "type": "box",
            "layout": "vertical",
            "backgroundColor": c["body_bg"],
            "paddingAll": "xl",
            "spacing": "md",
            "contents": [
                # 說明文字
                {
                    "type": "box",
                    "layout": "vertical",
                    "paddingAll": "md",
                    "backgroundColor": c["info_bg"],
                    "cornerRadius": "sm",
                    "contents": [
                        {
                            "type": "text",
                            "text": "📷 此按鈕僅翻譯上面那張圖片",
                            "size": "xxs",
                            "color": c["info_text"],
                            "wrap": True,
                            "weight": "bold",
                        },
                        {
                            "type": "text",
                            "text": "Tombol ini hanya untuk gambar di atas",
                            "size": "xxs",
                            "color": c["info_text"],
                            "wrap": True,
                            "margin": "xs",
                            "weight": "bold",
                        },
                        {
                            "type": "separator",
                            "margin": "sm",
                            "color": "#3a4a6a",
                        },
                        {
                            "type": "text",
                            "text": "💬 打字 / 語音 → 自動翻譯,不用按按鈕",
                            "size": "xxs",
                            "color": c["info_text"],
                            "wrap": True,
                            "margin": "sm",
                        },
                        {
                            "type": "text",
                            "text": "Chat teks / suara → otomatis, tanpa tombol",
                            "size": "xxs",
                            "color": c["info_text"],
                            "wrap": True,
                            "margin": "xs",
                        },
                    ],
                },
                # 主按鈕:翻譯上面那張圖
                # v3.9.32: 移除 displayText,過期按按鈕時群組不會顯示「翻譯上面那張圖」字樣
                # (LINE 的 displayText 是客戶端本地行為,後端管不到,只能不設)
                {
                    "type": "button",
                    "style": "primary",
                    "color": c["btn_bg"],
                    "height": "md",
                    "margin": "lg",
                    "action": {
                        "type": "postback",
                        "label": "✨ 翻譯圖片 / Terjemahkan",
                        "data": "img_translate=" + message_id,
                    },
                },
                # 提示
                {
                    "type": "text",
                    "text": "不需要就不用按 / Lewati jika tidak perlu",
                    "size": "xxs",
                    "color": c["plate_text"],
                    "align": "center",
                    "margin": "sm",
                },
                {
                    "type": "text",
                    "text": "1 分鐘後過期 / Kedaluwarsa 1 menit",
                    "size": "xxs",
                    "color": c["plate_text"],
                    "align": "center",
                    "margin": "xs",
                },
            ],
        },
        "styles": {
            "header": {"backgroundColor": c["header_bg"]},
            "body": {"backgroundColor": c["body_bg"]},
        },
    }
    return bubble


def handle_lang_command(text, group_id):
    return "ℹ️ 本機器人僅支援 中文 ⇄ 🇮🇩 印尼文 互譯\nBot ini hanya mendukung terjemahan Mandarin ⇄ Indonesia"


def handle_qry_command(text):
    """Handle /qry <customer_name> command to lookup storage area."""
    parts = text.strip().split(None, 1)
    if len(parts) < 2:
        return "⚠️ 請輸入客戶名稱 / Masukkan nama pelanggan\n範例 / Contoh: /qry ABE\n範例 / Contoh: /qry 佳東"
    query = parts[1].strip()
    # Try exact match first
    entries = STORAGE_LOOKUP.get(query)
    # Try case-insensitive match
    if not entries:
        for key in STORAGE_LOOKUP:
            if key.lower() == query.lower():
                entries = STORAGE_LOOKUP[key]
                query = key
                break
    # Try partial match
    if not entries:
        matches = [k for k in STORAGE_LOOKUP if query.lower() in k.lower() or query in k]
        if len(matches) == 1:
            query = matches[0]
            entries = STORAGE_LOOKUP[query]
        elif len(matches) > 1:
            result = "🔍 找到多筆符合 / Beberapa hasil ditemukan:\n"
            for m in matches[:10]:
                result += "  • " + m + "\n"
            if len(matches) > 10:
                result += "  ...(共 / Total " + str(len(matches)) + " 筆)\n"
            result += "\n請輸入完整客戶名稱 / Masukkan nama lengkap"
            return result
    if not entries:
        return "❌ 找不到客戶 / Pelanggan tidak ditemukan: " + query + "\n請確認名稱是否正確 / Mohon periksa nama"
    # Build response
    lines = []
    lines.append("📦 " + query + " 儲區查詢 / Cek gudang")
    lines.append("=" * 18)
    for length, area in entries:
        zh = format_length_zh(length)
        lines.append(zh + " \u2192 " + area)
    lines.append("=" * 18)
    return "\n".join(lines)


def handle_pkg_command(text):
    """Handle /pkg <code> command to lookup packaging info."""
    parts = text.strip().split(None, 1)
    if len(parts) < 2:
        return (
            "⚠️ 請輸入包裝碼 / Masukkan kode kemasan\n"
            "範例 / Contoh: /pkg U\n"
            "範例 / Contoh: /pkg G"
        )
    query = parts[1].strip()
    query_upper = query.upper()
    if not PACKAGING_LOOKUP:
        return "⚠️ 包裝碼資料尚未上傳\nData kode kemasan belum diupload"
    # Try exact match (case-insensitive)
    entry = PACKAGING_LOOKUP.get(query) or PACKAGING_LOOKUP.get(query_upper)
    matched_key = query if PACKAGING_LOOKUP.get(query) else query_upper
    if not entry:
        for k in PACKAGING_LOOKUP:
            if k.upper() == query_upper:
                entry = PACKAGING_LOOKUP[k]
                matched_key = k
                break
    # Try partial match
    if not entry:
        matches = [k for k in PACKAGING_LOOKUP if query_upper in k.upper()]
        if len(matches) == 1:
            matched_key = matches[0]
            entry = PACKAGING_LOOKUP[matched_key]
        elif len(matches) > 1:
            result = "🔍 找到多筆符合 / Beberapa hasil ditemukan:\n"
            for m in matches[:15]:
                result += "  • " + m + "\n"
            return result
    if not entry:
        return "❌ 找不到包裝碼 / Kode kemasan tidak ditemukan: " + query
    # Build response - show specific fields in order
    # Match Excel headers by keyword → display label
    PKG_DISPLAY = [
        ("簡稱",       ["簡稱"]),
        ("詳細包裝方式", ["詳細包裝", "包裝方式說明", "包裝方式"]),
        ("內包裝",     ["內包裝"]),
        ("外包裝",     ["外包裝"]),
        ("固定繩",     ["固定繩", "固定"]),
    ]
    lines = []
    lines.append("📦 包裝碼 / Kode kemasan: " + matched_key)
    lines.append("=" * 20)
    if isinstance(entry, dict):
        for display_label, keywords in PKG_DISPLAY:
            # Find matching field in entry
            for field_name, field_val in entry.items():
                if field_val and any(kw in field_name for kw in keywords):
                    lines.append(display_label + ": " + str(field_val))
                    break
    elif isinstance(entry, str):
        lines.append(entry)
    lines.append("=" * 20)
    return "\n".join(lines)


def get_display_name(group_id, user_id):
    """Get user display name from cache or LINE API. Also caches user language."""
    if group_id in group_user_names and user_id in group_user_names[group_id]:
        return group_user_names[group_id][user_id]
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            profile = api.get_group_member_profile(group_id, user_id)
            name = profile.display_name
            if name:
                if group_id not in group_user_names:
                    group_user_names[group_id] = {}
                group_user_names[group_id][user_id] = name
            # Cache user language and picture from LINE profile
            lang = getattr(profile, 'language', None)
            if lang and user_id not in user_languages:
                user_languages[user_id] = lang
                logger.info("User %s language: %s", name or user_id, lang)
            pic = getattr(profile, 'picture_url', None)
            if pic:
                user_pictures[user_id] = pic
            return name
    except Exception as e:
        logger.warning("Failed to get display name for %s: %s", user_id, e)
    return None


def record_user_name(group_id, user_id):
    """Record user display name in background (best effort)."""
    if not group_id or not user_id:
        return
    if group_id in group_user_names and user_id in group_user_names[group_id]:
        return
    get_display_name(group_id, user_id)


def find_user_by_name(group_id, name_query):
    """Find user_id by display name (partial match). Returns list of (user_id, display_name)."""
    if group_id not in group_user_names:
        return []
    matches = []
    query_lower = name_query.lower().strip()
    for uid, dname in group_user_names[group_id].items():
        if query_lower == dname.lower() or query_lower in dname.lower() or dname.lower() in query_lower:
            matches.append((uid, dname))
    return matches


def mark_translation_wrong(group_id, correct_translation="", add_to_examples=True,
                           offset=1, entry_id=None):
    """v3.4: Mark a translation as wrong with multiple selection modes.

    Selection priority:
      1. If entry_id provided → find that exact entry (works across days)
      2. Else: pick the Nth most recent translation in this group (offset=1 means latest)

    Args:
      group_id: LINE group id (limits search to this group's translations)
      correct_translation: optional human-corrected version, added as training example
      add_to_examples: whether to append to custom_translation_examples
      offset: 1-based index from the most recent (1=latest, 2=2nd latest, ...)
      entry_id: exact translation_log entry id

    Returns: (ok: bool, info_text: str)
    """
    target = None
    if entry_id:
        for entry in translation_log:
            if entry.get("id") == entry_id:
                target = entry
                break
        if not target:
            return False, f"找不到指定 ID 的翻譯記錄：{entry_id}"
    else:
        # Walk most-recent-first, count down to the Nth in this group
        n = max(1, int(offset or 1))
        seen = 0
        for entry in reversed(translation_log):
            if not group_id or entry.get("group_id") == group_id:
                seen += 1
                if seen == n:
                    target = entry
                    break
        if not target:
            return False, f"找不到第 {n} 筆翻譯記錄。可能剛重啟，或翻譯日誌不夠長。"

    target["marked_wrong"] = True
    target["manual_marked_wrong"] = True
    if correct_translation:
        target["correct_translation"] = correct_translation
        if add_to_examples:
            src_lang = target.get("src_lang", "zh")
            direction = "zh2id" if src_lang == "zh" else "id2zh"
            new_ex = {
                "zh": target.get("src", "") if src_lang == "zh" else correct_translation,
                "id": correct_translation if src_lang == "zh" else target.get("src", ""),
                "dir": direction
            }
            # Avoid exact duplicates
            is_dup = any(
                ex.get("zh") == new_ex["zh"] and ex.get("id") == new_ex["id"]
                for ex in custom_translation_examples
            )
            if not is_dup:
                custom_translation_examples.append(new_ex)
                if len(custom_translation_examples) > CUSTOM_EXAMPLES_MAX:
                    custom_translation_examples[:] = custom_translation_examples[-CUSTOM_EXAMPLES_MAX:]
                _save_examples_to_disk()  # v3.9.29: 修補隱性 bug
    _save_translation_log_to_disk()
    save_settings()
    return True, target.get("src", "")[:80]


def mark_latest_translation_wrong(group_id, correct_translation="", add_to_examples=True):
    """Backward-compat wrapper. Old code/tests calling this still work."""
    return mark_translation_wrong(group_id, correct_translation, add_to_examples, offset=1)


def list_recent_translations(group_id, limit=10):
    """Return the N most recent translation log entries for this group, newest first.
    Used by /wrong list command so the user can pick which one to mark."""
    out = []
    for entry in reversed(translation_log):
        if not group_id or entry.get("group_id") == group_id:
            out.append(entry)
            if len(out) >= limit:
                break
    return out


def format_recent_translations_for_line(entries):
    """Format a recent-translations list as a readable LINE message."""
    if not entries:
        return "（沒有翻譯記錄 / Tidak ada riwayat）"
    lines = ["📋 最近翻譯（新→舊）/ Riwayat terjemahan:"]
    for i, e in enumerate(entries, 1):
        src = (e.get("src") or "").strip().replace("\n", " ")[:40]
        tgt = (e.get("tgt") or "").strip().replace("\n", " ")[:40]
        marked = "❌ " if e.get("marked_wrong") else ""
        lines.append(f"{i}. {marked}原: {src}")
        lines.append(f"   譯: {tgt}")
    lines.append("")
    lines.append("標記第 N 筆錯誤：/wrong N 正確翻譯")
    lines.append("Mark salah ke-N: /wrong N terjemahan benar")
    return "\n".join(lines)


def parse_wrong_command(text):
    """Parse the /wrong command body. Returns dict with parsed fields.

    Supported syntaxes:
      /wrong                          → {"mode": "mark_only", "offset": 1}
      /wrong list                     → {"mode": "list"}
      /wrong 正確翻譯                 → {"mode": "correct", "offset": 1, "correct": "..."}
      /wrong 3 正確翻譯               → {"mode": "correct", "offset": 3, "correct": "..."}
      /wrong id=abc123 正確翻譯       → {"mode": "correct", "entry_id": "abc123", "correct": "..."}

    The original `text` is the raw user message starting with the command keyword.
    """
    # Strip the leading command keyword. Accept several aliases.
    body = text.strip()
    for prefix in ("/wrong", "/markwrong", "/錯", "/標錯"):
        if body.lower().startswith(prefix.lower()):
            body = body[len(prefix):].strip()
            break

    # Empty → mark only, no correction
    if not body:
        return {"mode": "mark_only", "offset": 1}

    # /wrong list
    if body.lower() in ("list", "ls", "清單", "列表"):
        return {"mode": "list"}

    # /wrong id=xxx 正確翻譯
    m = re.match(r'^id\s*=\s*(\S+)\s*(.*)$', body, re.IGNORECASE)
    if m:
        eid = m.group(1)
        correct = m.group(2).strip()
        if not correct:
            return {"mode": "mark_only", "entry_id": eid}
        return {"mode": "correct", "entry_id": eid, "correct": correct}

    # /wrong N 正確翻譯  (N is 1-9)
    m = re.match(r'^(\d+)\s+(.+)$', body)
    if m:
        n = int(m.group(1))
        correct = m.group(2).strip()
        if 1 <= n <= 50:
            return {"mode": "correct", "offset": n, "correct": correct}

    # /wrong 正確翻譯
    return {"mode": "correct", "offset": 1, "correct": body}


# v3.4: short-lived JSONL export tokens for /export jsonl LINE command.
# Maps token → (filepath, expiry_ts). Tokens are wiped after use or expiry.
_export_tokens = {}


def export_examples_to_jsonl():
    """Export custom_translation_examples to OpenAI fine-tune JSONL format.
    Returns a public URL good for ~1 hour, or "" on failure.

    OpenAI chat fine-tune format expects:
      {"messages":[
         {"role":"system","content":"..."},
         {"role":"user","content":"中文原文"},
         {"role":"assistant","content":"印尼譯文"}
      ]}
    one JSON per line.
    """
    if not custom_translation_examples:
        return ""
    try:
        token = uuid.uuid4().hex[:12]
        # Use /tmp on Render — survives within container lifetime, which is enough for 1h links.
        path = os.path.join(tempfile.gettempdir(), f"finetune_{token}.jsonl")
        sys_prompt = ("You are a Taiwan-factory ZH↔ID translator. "
                      "Output only the translation, no commentary.")
        with open(path, "w", encoding="utf-8") as f:
            for ex in custom_translation_examples:
                zh = (ex.get("zh") or "").strip()
                idn = (ex.get("id") or "").strip()
                direction = ex.get("dir", "zh2id")
                if not zh or not idn:
                    continue
                if direction == "zh2id":
                    user_msg, assistant_msg = zh, idn
                else:
                    user_msg, assistant_msg = idn, zh
                row = {
                    "messages": [
                        {"role": "system", "content": sys_prompt},
                        {"role": "user", "content": user_msg},
                        {"role": "assistant", "content": assistant_msg},
                    ]
                }
                f.write(json.dumps(row, ensure_ascii=False) + "\n")
        _export_tokens[token] = (path, int(time.time()) + 3600)
        # Clean up expired tokens occasionally
        now = int(time.time())
        expired = [t for t, (_p, exp) in _export_tokens.items() if exp < now]
        for t in expired:
            try:
                os.remove(_export_tokens[t][0])
            except Exception:
                pass
            _export_tokens.pop(t, None)
        # Return a relative URL; LINE will let the user tap it.
        # Caller can prepend host if needed; we'll use the Render hostname through env.
        host = os.environ.get("PUBLIC_HOST", "").rstrip("/")
        if host:
            return f"{host}/export/finetune/{token}"
        # Fallback: absolute path note for the user
        return f"/export/finetune/{token}"
    except Exception as e:
        logger.error("export_examples_to_jsonl failed: %s", e)
        return ""


def handle_command(text, group_id, user_id=None):
    bot_stats["commands"] += 1
    cmd = text.strip().lower()
    if cmd == "/help":
        # Return sentinel; caller detects this and sends Flex instead of Text
        return "__FLEX_HELP__"
    elif cmd.startswith("/wrong") or cmd.startswith("/markwrong") or cmd.startswith("/錯") or cmd.startswith("/標錯"):
        # v3.4: rich /wrong syntax (see parse_wrong_command for supported forms)
        parsed = parse_wrong_command(text)
        mode = parsed.get("mode")

        if mode == "list":
            entries = list_recent_translations(group_id, limit=10)
            return format_recent_translations_for_line(entries)

        if mode == "mark_only":
            ok, info = mark_translation_wrong(
                group_id,
                correct_translation="",
                add_to_examples=False,
                offset=parsed.get("offset", 1),
                entry_id=parsed.get("entry_id"),
            )
            if not ok:
                return "⚠️ " + info
            return ("✅ 已標記翻譯錯誤 / Terjemahan ditandai salah\n"
                    "若要同時加入正確譯文,請輸入 / Untuk menambahkan koreksi:\n"
                    "/wrong 正確翻譯(標最近一筆)\n"
                    "/wrong 2 正確翻譯(標倒數第 2 筆)\n"
                    "/wrong list(看最近 10 筆)")

        # mode == "correct"
        correct = parsed.get("correct", "").strip()
        ok, info = mark_translation_wrong(
            group_id,
            correct_translation=correct,
            add_to_examples=True,
            offset=parsed.get("offset", 1),
            entry_id=parsed.get("entry_id"),
        )
        if not ok:
            return "⚠️ " + info
        offset = parsed.get("offset", 1)
        position_label = "最近一筆" if offset == 1 else f"倒數第 {offset} 筆"
        if parsed.get("entry_id"):
            position_label = f"指定 ID 訊息"
        total = len(custom_translation_examples)
        return (f"✅ 已標記{position_label}錯誤,並加入修正範例 / Ditandai salah & ditambahkan ke contoh koreksi:\n"
                f"{correct}\n"
                f"📚 累積範例 / Total contoh: {total} / {CUSTOM_EXAMPLES_MAX}")

    elif cmd.startswith("/export"):
        # v3.4: export training data
        parts = text.strip().split()
        if len(parts) == 1:
            # /export → status
            total = len(custom_translation_examples)
            zh2id = sum(1 for e in custom_translation_examples if e.get("dir") == "zh2id")
            id2zh = sum(1 for e in custom_translation_examples if e.get("dir") == "id2zh")
            target = 300
            ready = "✅ 達到 fine-tune 建議門檻" if total >= target else f"再 {target-total} 筆達建議門檻"
            return (f"📊 訓練資料統計 / Statistik data\n"
                    f"總範例數 / Total: {total}\n"
                    f"中→印 / ZH→ID: {zh2id}\n"
                    f"印→中 / ID→ZH: {id2zh}\n"
                    f"狀態 / Status: {ready}\n\n"
                    f"匯出 jsonl: /export jsonl\n"
                    f"（產生下載連結 / generate link）")
        elif parts[1].lower() in ("jsonl", "json"):
            url = export_examples_to_jsonl()
            if not url:
                return "⚠️ 匯出失敗 / Export gagal（沒有範例 / no examples）"
            return f"📦 訓練資料已匯出 / Data ekspor:\n{url}\n\n（連結 1 小時內有效 / valid 1 hour）"
        else:
            return "用法 / Usage: /export 或 /export jsonl"

    elif cmd in ("/stats wrong", "/wrongstats", "/標記統計"):
        total = len(custom_translation_examples)
        wrong_count = sum(1 for e in translation_log if e.get("marked_wrong"))
        recent_30d = 0
        cutoff = int(time.time()) - 30*86400
        for e in translation_log:
            if e.get("marked_wrong") and e.get("ts", 0) > cutoff:
                recent_30d += 1
        return (f"📊 翻譯品質統計 / Statistik kualitas terjemahan\n"
                f"累積修正範例 / Contoh koreksi: {total} / {CUSTOM_EXAMPLES_MAX}\n"
                f"日誌中標錯 / Ditandai salah: {wrong_count} 筆\n"
                f"近 30 天標錯 / 30 hari terakhir: {recent_30d} 筆\n"
                f"距 fine-tune 門檻 / Sisa untuk fine-tune: {max(0, 300 - total)} 筆")

    elif cmd == "/on":
        group_settings[group_id] = True
        save_settings()
        return "\u2705 \u7ffb\u8b6f\u5df2\u958b\u555f / Penerjemah aktif"
    elif cmd == "/off":
        group_settings[group_id] = False
        save_settings()
        return "\u274c \u7ffb\u8b6f\u5df2\u95dc\u9589 / Penerjemah nonaktif"
    elif cmd == "/img on":
        group_img_settings[group_id] = True
        group_img_ask_settings.pop(group_id, None)  # 清掉 ask 旗標
        save_settings()
        return "\u2705 \u5716\u7247\u7ffb\u8b6f\u5df2\u958b\u555f(\u81ea\u52d5\u7ffb\u8b6f) / Terjemahan gambar otomatis"
    elif cmd == "/img off":
        group_img_settings[group_id] = False
        group_img_ask_settings.pop(group_id, None)  # 清掉 ask 旗標
        save_settings()
        return "\u274c \u5716\u7247\u7ffb\u8b6f\u5df2\u95dc\u9589 / Terjemahan gambar nonaktif"
    elif cmd == "/img ask":
        # v3.9.10: 詢問模式 — 收到圖片不自動翻,先問使用者按按鈕才翻
        group_img_settings[group_id] = False
        group_img_ask_settings[group_id] = True
        save_settings()
        return "\u2753 \u5716\u7247\u7ffb\u8b6f: \u8a62\u554f\u6a21\u5f0f\n\u6536\u5230\u5716\u7247\u6703\u51fa\u73fe\u300c\u7ffb\u8b6f\u9019\u5f35 / \u4e0d\u7528\u300d\u6309\u9215\nMode tanya: pesan gambar akan menanyakan dulu"
    elif cmd == "/voice on":
        group_audio_settings[group_id] = True
        save_settings()
        return "\u2705 \u8a9e\u97f3\u7ffb\u8b6f\u5df2\u958b\u555f / Terjemahan suara aktif"
    elif cmd == "/voice off":
        group_audio_settings[group_id] = False
        save_settings()
        return "\u274c \u8a9e\u97f3\u7ffb\u8b6f\u5df2\u95dc\u9589 / Terjemahan suara nonaktif"
    elif cmd == "/wo on":
        group_wo_settings[group_id] = True
        save_settings()
        return "✅ 拍工單查儲區已開啟 / Foto WO cek gudang aktif"
    elif cmd == "/wo off":
        group_wo_settings[group_id] = False
        save_settings()
        return "❌ 拍工單查儲區已關閉 / Foto WO cek gudang nonaktif"
    elif cmd == "/skip":
        if not user_id:
            return "⚠️ 無法識別你的身分 / Tidak bisa mengenali identitas Anda"
        if group_id not in group_skip_users:
            group_skip_users[group_id] = set()
        group_skip_users[group_id].add(user_id)
        save_settings()
        return "✅ 已將你加入白名單,你的訊息不會被翻譯\nAnda ditambahkan ke daftar skip, pesan Anda tidak akan diterjemahkan"
    elif cmd == "/unskip":
        if not user_id:
            return "⚠️ 無法識別你的身分 / Tidak bisa mengenali identitas Anda"
        if group_id in group_skip_users:
            group_skip_users[group_id].discard(user_id)
        save_settings()
        return "✅ 已將你移出白名單,你的訊息會被翻譯\nAnda dihapus dari daftar skip, pesan Anda akan diterjemahkan"
    elif text.strip().lower().startswith("/skipadd"):
        name_query = text.strip()[8:].strip()
        if not name_query:
            return "⚠️ 請輸入名字 / Masukkan nama\n範例 / Contoh: /skipadd 秋情"
        matches = find_user_by_name(group_id, name_query)
        if len(matches) == 0:
            return "❌ 找不到「" + name_query + "」 / Tidak ditemukan\n該用戶需先在群組發過訊息才能被認到\nUser harus pernah kirim pesan di grup dulu"
        if len(matches) > 1:
            names = "\n".join(["  • " + m[1] for m in matches])
            return "🔍 找到多人符合 / Beberapa user ditemukan:\n" + names + "\n請輸入更完整的名字 / Masukkan nama yang lebih lengkap"
        uid, dname = matches[0]
        if group_id not in group_skip_users:
            group_skip_users[group_id] = set()
        group_skip_users[group_id].add(uid)
        save_settings()
        return "✅ 已將「" + dname + "」加入白名單,訊息不會被翻譯\n" + dname + " ditambahkan ke daftar skip, pesan tidak akan diterjemahkan"
    elif text.strip().lower().startswith("/skipdel"):
        name_query = text.strip()[8:].strip()
        if not name_query:
            return "⚠️ 請輸入名字 / Masukkan nama\n範例 / Contoh: /skipdel 秋情"
        matches = find_user_by_name(group_id, name_query)
        if len(matches) == 0:
            return "❌ 找不到「" + name_query + "」 / Tidak ditemukan"
        if len(matches) > 1:
            names = "\n".join(["  • " + m[1] for m in matches])
            return "🔍 找到多人符合 / Beberapa user ditemukan:\n" + names + "\n請輸入更完整的名字 / Masukkan nama yang lebih lengkap"
        uid, dname = matches[0]
        if group_id in group_skip_users:
            group_skip_users[group_id].discard(uid)
        save_settings()
        return "✅ 已將「" + dname + "」移出白名單,訊息會被翻譯\n" + dname + " dihapus dari daftar skip, pesan akan diterjemahkan"
    elif cmd == "/skiplist":
        skipped = group_skip_users.get(group_id, set())
        if not skipped:
            return "\u76ee\u524d\u767d\u540d\u55ae\u662f\u7a7a\u7684 / Daftar skip kosong"
        names_cache = group_user_names.get(group_id, {})
        lines = ["\u23ed\ufe0f \u767d\u540d\u55ae / Daftar skip:"]
        for uid in skipped:
            dname = names_cache.get(uid)
            if dname:
                lines.append("  \u2022 " + dname)
            else:
                lines.append("  \u2022 (\u672a\u77e5\u7528\u6236)")
        return "\n".join(lines)
    elif cmd == "/status":
        is_on = group_settings.get(group_id, True)
        if is_on:
            img_on = group_img_settings.get(group_id, True)
            img_status = "✅ 開啟 / Aktif" if img_on else "❌ 關閉 / Nonaktif"
            audio_on = group_audio_settings.get(group_id, True)
            audio_status = "✅ 開啟 / Aktif" if audio_on else "❌ 關閉 / Nonaktif"
            wo_on = group_wo_settings.get(group_id, True)
            wo_status = "✅ 開啟 / Aktif" if wo_on else "❌ 關閉 / Nonaktif"
            return ("✅ 翻譯:開啟中 / Penerjemah aktif\n"
                    "中文 ⇄ 🇮🇩 印尼文 / Mandarin ⇄ Indonesia\n"
                    "🖼️ 圖片翻譯 / Terjemahan gambar:" + img_status + "\n"
                    "🎤 語音翻譯 / Terjemahan suara:" + audio_status + "\n"
                    "📋 拍工單查儲區 / Foto WO cek gudang:" + wo_status)
        else:
            return "❌ 翻譯:已關閉 / Penerjemah nonaktif"
    elif cmd == "/clearcache":
        # v3.9.3: clear translation cache (e.g. after fixing a bad translation
        # that got cached, or after switching to a new model). Anyone in the
        # group can run this — it only affects the in-memory cache, no data loss.
        with _cache_lock:
            n = len(translation_cache)
            translation_cache.clear()
        logger.info("Translation cache cleared (%d entries) by user in group %s", n, group_id)
        return f"🧹 已清除翻譯快取 / Cache terjemahan dibersihkan\n清除筆數 / Jumlah: {n}"
    elif cmd.startswith("/lang"):
        return handle_lang_command(text, group_id)
    elif text.strip().startswith("/notice ") or text.strip().startswith("/notice\u3000"):
        if not is_cmd_enabled(group_id, "notice"):
            return None
        content = text.strip()[8:].strip()
        if not content:
            return "⚠️ 請輸入公告內容 / Masukkan isi pengumuman\n例如 / Contoh: /notice 明天放假一天"
        tgt = group_target_lang.get(group_id, "id")
        if has_chinese(content):
            return make_notice(content, tgt)
        else:
            src = detect_language(content)
            if src and src != "zh":
                return make_notice_from_other(content, src)
            return make_notice(content, tgt)
    elif text.strip().lower().startswith("/qry"):
        if not is_cmd_enabled(group_id, "qry"):
            return None
        return handle_qry_command(text)
    elif cmd == "/pw1":
        if not is_cmd_enabled(group_id, "pw1"):
            return None
        return "🔑 班長密碼 / PW Shift Leader\n" + "=" * 18 + "\n" + pw1_text + "\n" + "=" * 18
    elif cmd == "/pw2":
        if not is_cmd_enabled(group_id, "pw2"):
            return None
        return "🏭 儲運密碼 / PW Gudang\n" + "=" * 18 + "\n" + pw2_text + "\n" + "=" * 18
    elif cmd == "/scrap":
        if not is_cmd_enabled(group_id, "scrap"):
            return None
        return scrap_text
    elif text.strip().lower().startswith("/pkg"):
        if not is_cmd_enabled(group_id, "pkg"):
            return None
        return handle_pkg_command(text)
    elif cmd == "/saran":
        # v3.9.31: 提案系統網址 — 用純文字回傳,使用者可長按複製或用外部瀏覽器開啟
        return ("💡 提案系統 / Sistem Saran\n"
                "長按網址可複製或用外部瀏覽器開啟\n"
                "Tekan lama untuk salin atau buka di browser luar\n\n"
                "https://app-walsin-crm-improvement.azurewebsites.net/improvePropose/personalList")
    elif cmd == "/absen":
        # v3.9.31: 差勤系統網址 — 同上
        return ("📅 差勤系統 / Sistem Absen\n"
                "長按網址可複製或用外部瀏覽器開啟\n"
                "Tekan lama untuk salin atau buka di browser luar\n\n"
                "https://hrm.walsin.com/servlet/jform?file=hrm8w.pkg,hrm8aw.pkg,BPM_JS.pkg,hrm8w_walsin.pkg,hrm8w_walsinhrisp.pkg&locale=US&init_func=%E4%BA%BA%E4%BA%8B_WS")
    return None


@app.route("/callback", methods=["POST"])
def callback():
    sig = request.headers.get("X-Line-Signature", "")
    body = request.get_data(as_text=True)
    # v3.9.19: 記錄所有進來的 webhook(只記事件 type 跟 message type,不記內容)
    try:
        _b = json.loads(body) if body else {}
        _evts = _b.get("events", [])
        _summary = []
        for ev in _evts:
            _t = ev.get("type", "?")
            _msg = ev.get("message", {})
            _src = ev.get("source", {})
            _summary.append({
                "type": _t,
                "message_type": _msg.get("type") if _msg else None,
                "message_id": _msg.get("id") if _msg else None,
                "source_type": _src.get("type"),
                "group_id": _src.get("groupId") or _src.get("roomId"),
                "user_id_tail": (_src.get("userId") or "")[-6:] if _src.get("userId") else None,
            })
        _event_log_write("webhook_in", {"events_count": len(_evts), "events": _summary})
    except Exception as _le:
        _event_log_write("webhook_parse_error", {"error": str(_le)[:200]})
    try:
        handler.handle(body, sig)
    except InvalidSignatureError:
        _event_log_write("webhook_invalid_sig", {})
        abort(400)
    return "OK"


@handler.add(MessageEvent, message=TextMessageContent)
def handle_message(event):
    # v3.9.30c B15 修補: LINE webhook 重發保護
    if _is_redelivery(event) or _is_duplicate_message(getattr(event.message, 'id', None)):
        logger.warning("[handle_message] redelivery or duplicate, skipping msg_id=%s",
                       getattr(event.message, 'id', None))
        return
    
    text = event.message.text.strip()
    if len(text) < 2:
        return

    source = event.source
    is_dm = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
    group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', None)
    user_id = getattr(source, 'user_id', None)

    # v3.9.30c B14 修補: 重置 _tl thread-local 殘留狀態
    # gunicorn worker 會重用 thread,前一次處理可能殘留 from_image_ocr=True 等狀態
    # 不重置會讓 text 訊息誤觸發 OCR 專用邏輯(例如強制分段翻譯)
    try:
        _tl.from_image_ocr = False
        _tl.factory_audit = None
        _tl.rtc_fail_reason = None
        _tl.last_entry_id = None
        _tl.gpt_says_announcement = None
        _tl.struct_self_check = None
        _tl.multi_path_fail = None
    except Exception:
        pass

    # --- DM (private message) mode ---
    if is_dm and user_id:
        # Record DM user for admin panel
        if user_id not in dm_known_users:
            try:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    profile = api.get_profile(user_id)
                    dm_known_users[user_id] = profile.display_name or user_id
            except Exception:
                dm_known_users[user_id] = user_id

        # DM commands
        cmd = text.strip().lower()
        if cmd == "/help":
            # DM /help also uses Flex carousel (ZH + ID bilingual)
            send_help_flex(event.reply_token, primary_lang="zh")
            return
        if cmd.startswith("/to"):
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.reply_message(ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text="ℹ️ 本機器人僅支援 中文 ⇄ 🇮🇩 印尼文 互譯\nBot ini hanya mendukung terjemahan Mandarin ⇄ Indonesia")]
                ))
            return
        # DM: handle /qry command
        if text.strip().lower().startswith("/qry"):
            qry_result = handle_qry_command(text)
            if qry_result:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=_clip_line_text(qry_result))]
                    ))
            return
        # DM: handle /pw1, /pw2, /scrap, /pkg commands
        dm_cmd_result = None
        if cmd == "/pw1":
            dm_cmd_result = "🔑 班長密碼 / PW Shift Leader\n" + "=" * 18 + "\n" + pw1_text + "\n" + "=" * 18
        elif cmd == "/pw2":
            dm_cmd_result = "🏭 儲運密碼 / PW Gudang\n" + "=" * 18 + "\n" + pw2_text + "\n" + "=" * 18
        elif cmd == "/scrap":
            dm_cmd_result = scrap_text
        elif text.strip().lower().startswith("/pkg"):
            dm_cmd_result = handle_pkg_command(text)
        if dm_cmd_result:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.reply_message(ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=_clip_line_text(dm_cmd_result))]
                ))
            return
        # DM: skip other / commands
        if text.startswith("/"):
            return

        # DM master toggle check
        if not dm_master_enabled and user_id not in dm_whitelist:
            return

        # DM translation: strip mentions, detect language, translate
        text_clean = strip_mentions_for_detect(text).strip()
        if not text_clean or len(text_clean) < 2:
            return

        lang = detect_language(text_clean)
        tgt = dm_target_lang.get(user_id, "id")
        if lang is None:
            return
        if lang == tgt:
            return

        # Set translation tone for DM (use global default)
        _tl.tone = translation_tone
        _tl.tone_custom = translation_tone_custom
        _tl.group_id = "__dm__"

        _bp, _bc = bot_stats.get("tokens_prompt", 0), bot_stats.get("tokens_completion", 0)
        result = translate(text_clean, lang, tgt)
        track_group_usage("__dm__", _bp, _bc)
        if not result:
            return
        reply = LANG_FLAGS.get(tgt, "") + " " + result


        bot_stats["text_translations"] += 1
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.reply_message(ReplyMessageRequest(
                reply_token=event.reply_token,
                messages=[TextMessage(text=_clip_line_text(reply))]
            ))
        return

    # --- Group mode (original logic) ---
    # Record user display name for /skipadd lookup
    if group_id and user_id:
        record_user_name(group_id, user_id)
    # Track group for admin panel
    if group_id and not is_dm and group_id not in group_tracking:
        gname = ""
        try:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                summary = api.get_group_summary(group_id)
                gname = summary.group_name or ""
        except Exception:
            pass
        group_tracking[group_id] = {"name": gname, "joined_at": time.time()}
        save_settings()

    if text.startswith("/"):
        # Set tone before commands that may translate (e.g. /notice)
        _tone, _tone_custom = get_group_tone(group_id)
        _tl.tone = _tone
        _tl.tone_custom = _tone_custom
        cmd_result = handle_command(text, group_id, user_id)
        if cmd_result == "__FLEX_HELP__":
            # /help uses Flex Message carousel (ZH + ID bilingual)
            send_help_flex(event.reply_token, primary_lang="zh")
        elif cmd_result:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.reply_message(ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=_clip_line_text(cmd_result))]
                ))
        return

    is_on = group_settings.get(group_id, True)
    if not is_on:
        return

    # Check skip list
    sender_id = getattr(source, 'user_id', None)
    if sender_id and sender_id in group_skip_users.get(group_id, set()):
        return

    if text.startswith("!"):
        return

    # Extract LINE's actual @mention data (blue text = 100% accurate)
    line_mentions = extract_line_mentions(text, event.message)

    # Cache this message for future quote references
    msg_id = getattr(event.message, 'id', None)
    if msg_id:
        message_cache[msg_id] = {"text": text, "ts": time.time()}
        # Trim cache
        if len(message_cache) > MESSAGE_CACHE_MAX:
            oldest = sorted(message_cache.items(), key=lambda x: x[1]["ts"])[:50]
            for k, _ in oldest:
                message_cache.pop(k, None)

    # Check if this is a reply to another message (quoted message)
    quoted_text = None
    quoted_id = getattr(event.message, 'quoted_message_id', None)
    if quoted_id and quoted_id in message_cache:
        quoted_text = message_cache[quoted_id].get("text", "")

    # Strip @mentions for language detection only
    text_for_detect = strip_mentions_for_detect(text, line_mentions).strip()
    if not text_for_detect or len(text_for_detect) < 2:
        return

    lang = detect_language(text_for_detect)
    if lang is None:
        return

    tgt = group_target_lang.get(group_id, "id")

    # Show typing indicator while translating
    show_loading(group_id)
    if get_group_feature(group_id, 'mark_read'):
        mark_as_read(group_id)

    # Protect LINE mentions before translation
    text_to_translate = text
    mention_placeholders = {}
    if line_mentions:
        text_to_translate, mention_placeholders = protect_mentions(text, line_mentions)

    reply = None
    _bp, _bc = bot_stats.get("tokens_prompt", 0), bot_stats.get("tokens_completion", 0)
    # Set translation tone for this group
    _tone, _tone_custom = get_group_tone(group_id)
    _tl.tone = _tone
    _tl.tone_custom = _tone_custom
    _tl.group_id = group_id
    if lang == "zh":
        result = translate(text_to_translate, "zh", tgt)
        if result and mention_placeholders:
            result = restore_mentions(result, mention_placeholders)
        if result:
            reply = LANG_FLAGS.get(tgt, "") + " " + result
    else:
        result = translate(text_to_translate, lang, "zh")
        if result and mention_placeholders:
            result = restore_mentions(result, mention_placeholders)
        if result:
            reply = LANG_FLAGS.get("zh", "") + " " + result
    track_group_usage(group_id, _bp, _bc)

    if reply is None:
        return

    bot_stats["text_translations"] += 1

    # Build reply message based on settings
    sender_display = None
    if sender_id:
        sender_display = (group_user_names.get(group_id, {}).get(sender_id) or
                       get_display_name(group_id, sender_id))

    src_flag = LANG_FLAGS.get(lang, "")
    tgt_flag = LANG_FLAGS.get("zh" if lang != "zh" else "id", "")
    translated_text = reply.split(" ", 1)[1] if " " in reply else reply

    # Flex or plain text based on setting
    flex_msg = None
    if get_group_feature(group_id, 'flex'):
        flex_msg = build_translation_flex(text, translated_text, src_flag, tgt_flag, sender_display, quoted_text)
    qr = build_quick_reply(group_id) if get_group_feature(group_id, 'quick_reply') else None
    custom_sender = get_sender_object()
    # Get quoteToken from original message for reply linking
    qt = getattr(event.message, 'quote_token', None)

    _use_retry = get_group_feature(group_id, 'retry_key')
    _retry_key = generate_retry_key() if _use_retry else None

    with ApiClient(configuration) as api_client:
        api_line = MessagingApi(api_client)
        # v3.8: capture last logged entry id BEFORE reply, so we can map it
        # to the bot's outgoing message_id for reaction-event feedback.
        _entry_id_for_reaction = getattr(_tl, 'last_entry_id', None)
        _reply_resp = None
        if flex_msg:
            if qr:
                flex_msg.quick_reply = qr
            if custom_sender:
                flex_msg.sender = custom_sender
            if qt:
                try: flex_msg.quote_token = qt
                except Exception: pass
            req = ReplyMessageRequest(reply_token=event.reply_token, messages=[flex_msg])
            if get_group_feature(group_id, 'silent'):
                req.notification_disabled = True
            try:
                if _retry_key:
                    _reply_resp = api_line.reply_message(req, x_line_retry_key=_retry_key)
                else:
                    _reply_resp = api_line.reply_message(req)
            except TypeError:
                _reply_resp = api_line.reply_message(req)
        else:
            msg = TextMessage(text=_clip_line_text(reply))
            if qr:
                msg.quick_reply = qr
            if custom_sender:
                msg.sender = custom_sender
            if qt:
                try: msg.quote_token = qt
                except Exception: pass
            req = ReplyMessageRequest(reply_token=event.reply_token, messages=[msg])
            if get_group_feature(group_id, 'silent'):
                req.notification_disabled = True
            try:
                if _retry_key:
                    _reply_resp = api_line.reply_message(req, x_line_retry_key=_retry_key)
                else:
                    _reply_resp = api_line.reply_message(req)
            except TypeError:
                _reply_resp = api_line.reply_message(req)
        # v3.8: bind every sent_message.id back to the translation entry so
        # subsequent reactions can be attributed to the correct entry.
        try:
            if _reply_resp and _entry_id_for_reaction:
                sent_msgs = getattr(_reply_resp, 'sent_messages', None) or []
                for sm in sent_msgs:
                    sm_id = getattr(sm, 'id', None)
                    if sm_id:
                        register_sent_message(sm_id, _entry_id_for_reaction, group_id)
        except Exception as _re:
            logger.debug("register_sent_message skipped: %s", _re)


@handler.add(MessageEvent, message=ImageMessageContent)
def handle_image(event):
    """Handle image messages: OCR + translate with layout-preserving text."""
    # v3.9.30c B15 修補: LINE webhook 重發保護
    if _is_redelivery(event) or _is_duplicate_message(getattr(event.message, 'id', None)):
        logger.warning("[handle_image] redelivery or duplicate, skipping msg_id=%s",
                       getattr(event.message, 'id', None))
        return
    
    source = event.source
    group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', None)
    user_id = getattr(source, 'user_id', None)
    is_dm_img = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
    logger.info("Image received from %s", group_id)
    
    # v3.9.30c B13 修補: 設定 _tl thread-local,避免跨群組污染
    # 之前 handle_image 完全沒設 _tl,OCR 翻譯時會讀到上一次翻譯殘留的 tone / group_id,
    # 導致 A 群組 OCR 圖片用到 B 群組 tone 設定。
    try:
        _tone, _tone_custom = get_group_tone(group_id)
        _tl.tone = _tone
        _tl.tone_custom = _tone_custom
        _tl.group_id = group_id or ""
        _tl.user_id = user_id or ""
        _tl.from_image_ocr = True  # 標記:這次翻譯是 OCR 來的,供下游函式區分
    except Exception:
        pass
    
    # v3.9.19: 寫磁碟 log(跨 worker 共享)
    _event_log_write("image_handler_entered", {
        "msg_id": event.message.id,
        "group_id": group_id or "",
        "user_id_tail": (user_id or "")[-6:],
        "is_dm": is_dm_img,
    })
    
    # v3.9.18: 記錄到 recent images(供 /debug/recent-images 使用)
    try:
        _last_image_received_msgs.append({
            "msg_id": event.message.id,
            "group_id": group_id or "",
            "ts": int(time.time()),
        })
        # 只保留最後 10 張
        if len(_last_image_received_msgs) > 10:
            _last_image_received_msgs.pop(0)
    except Exception:
        pass

    # Record user for whitelist (even if translation is off)
    if group_id and user_id and not is_dm_img:
        record_user_name(group_id, user_id)

    # Check if translation is on
    is_on = group_settings.get(group_id, True)
    if not is_on:
        _event_log_write("image_skipped", {"reason": "translation_off", "group_id": group_id or ""})
        return

    # Check skip list
    sender_id = user_id
    if sender_id and sender_id in group_skip_users.get(group_id, set()):
        _event_log_write("image_skipped", {"reason": "user_in_skip_list", "user_tail": (sender_id or "")[-6:]})
        return

    # DM master toggle check for image
    if is_dm_img and sender_id:
        if not dm_master_enabled and sender_id not in dm_whitelist:
            _event_log_write("image_skipped", {"reason": "dm_disabled"})
            return

    # v3.9.10: 三種圖片處理模式
    # 1. group_img_settings=True → 自動翻譯(原本行為)
    # 2. group_img_settings=False AND group_img_ask_settings=True → 詢問模式
    # 3. group_img_settings=False AND group_img_ask_settings=False → 完全不處理
    img_on = group_img_settings.get(group_id, True)
    img_ask = group_img_ask_settings.get(group_id, False)
    _event_log_write("image_mode_check", {
        "group_id": group_id or "",
        "img_on": img_on,
        "img_ask": img_ask,
        "decision": "auto_translate" if img_on else ("ask_mode" if img_ask else "fully_off")
    })
    if not img_on:
        if img_ask:
            # === 詢問模式 ===
            # v3.9.13: 用磁碟版 pending(_pending_img_set)跨 worker 共享
            _now = int(time.time())
            _pending_img_set(event.message.id, {
                "group_id": group_id,
                "user_id": user_id or "",
                "ts": _now,
            })
            logger.info("[ImgAsk] pending stored: msg=%s gid=%s", event.message.id, group_id)
            # v3.9.12: 用 Flex 卡片詢問,跟 /help 同視覺風格
            try:
                bubble = build_image_ask_flex(event.message.id)
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[FlexMessage(
                            alt_text="📷 收到圖片 - 要翻譯嗎?",
                            contents=FlexContainer.from_dict(bubble),
                        )]
                    ))
                logger.info("[ImgAsk] Flex sent for msg=%s", event.message.id)
            except Exception as _e:
                logger.warning("Image ask-mode Flex failed: %s", _e)
                # Fallback 到純文字 + Quick Reply
                try:
                    from linebot.v3.messaging import (
                        QuickReply, QuickReplyItem, PostbackAction
                    )
                    qr = QuickReply(items=[
                        QuickReplyItem(action=PostbackAction(
                            label="📝 翻譯這張",
                            data="img_translate=" + event.message.id,
                            # v3.9.32: 不設 display_text,過期按按鈕時不會在群組顯示
                        )),
                    ])
                    with ApiClient(configuration) as api_client:
                        api = MessagingApi(api_client)
                        msg = TextMessage(
                            text="📷 收到圖片\n需要翻譯這張圖請按下方按鈕(1 分鐘內有效)\n\n💬 打字/語音會自動翻譯,不用按按鈕\nChat teks/suara otomatis, tanpa tombol",
                            quick_reply=qr
                        )
                        api.reply_message(ReplyMessageRequest(
                            reply_token=event.reply_token,
                            messages=[msg]
                        ))
                except Exception:
                    pass
        return

    # Need OpenAI for image OCR
    if not oai:
        _event_log_write("image_aborted", {"reason": "openai_not_initialized"})
        logger.warning("No OpenAI key, cannot do image OCR")
        return

    # v3.9.25: 把整個 OCR + 翻譯 + reply 移到背景 thread,
    # 因為 LINE webhook 必須在短時間內 return,否則會被 LINE 跟 Render 雙重 timeout 殺掉
    _event_log_write("image_step", {"step": "spawning_background_thread", "msg_id": event.message.id})
    
    # 把需要的東西打包成 dict,thread 內用(不要傳整個 event 物件,event 在 webhook 結束後可能無效)
    _ctx = {
        "message_id": event.message.id,
        "reply_token": event.reply_token,
        "quote_token": getattr(event.message, 'quote_token', None),
        "group_id": group_id,
        "user_id": user_id,
        "tgt": group_target_lang.get(group_id, "id"),
        "tone_info": get_group_tone(group_id),
        "wo_setting": group_wo_settings.get(group_id, True),
        "mark_read_setting": get_group_feature(group_id, 'mark_read'),
        "is_dm_img": is_dm_img,
    }
    
    import threading as _threading
    _bg = _threading.Thread(target=_handle_image_background, args=(_ctx,), daemon=True)
    _bg.start()
    # webhook 立刻 return,LINE 不會 timeout
    return


def _handle_image_background(ctx):
    """v3.9.25: 在背景 thread 跑 OCR + 翻譯 + 回應 LINE。
    webhook 主執行緒已經 return 了,所以可以慢慢做不會 timeout。
    最後用 reply_message(若 token 還活著)或 push_message(fallback)送譯文。
    """
    try:
        # 從 ctx 取出主流程設定
        group_id = ctx["group_id"]
        user_id = ctx["user_id"]
        message_id = ctx["message_id"]
        is_dm_img = ctx["is_dm_img"]
        
        # v3.9.26: 確認 thread 真的開始跑
        _event_log_write("bg_thread_started", {
            "msg_id": message_id,
            "group_id": group_id or "",
        })
        
        # show_loading + mark_read(原本在 webhook 主流程做的,移到背景)
        _event_log_write("image_step", {"step": "before_show_loading"})
        try:
            show_loading(group_id)
        except Exception as _sle:
            _event_log_write("image_step_error", {"step": "show_loading", "err": str(_sle)[:200]})
        if ctx.get("mark_read_setting"):
            try:
                mark_as_read(group_id)
            except Exception:
                pass
        
        # Download image from LINE
        _event_log_write("image_step", {"step": "downloading", "msg_id": message_id})
        try:
            img_base64, img_raw = download_line_image(message_id)
        except Exception as _de:
            _event_log_write("image_step_error", {"step": "download", "err": str(_de)[:300]})
            logger.exception("Image download exception: %s", _de)
            return
        if not img_base64:
            _event_log_write("image_aborted", {"reason": "download_returned_none"})
            logger.warning("Failed to download image %s", message_id)
            return
        img_mime = detect_image_mime(img_raw)
        _event_log_write("image_step", {
            "step": "downloaded",
            "size_bytes": len(img_raw) if img_raw else 0,
            "mime": img_mime,
            "first_bytes_hex": (img_raw[:16].hex() if img_raw else ""),
        })

        # Determine target language
        tgt = group_target_lang.get(group_id, "id")

        # Quick OCR to check if there's text and detect language
        _bp, _bc = bot_stats.get("tokens_prompt", 0), bot_stats.get("tokens_completion", 0)
        _event_log_write("image_step", {"step": "before_ocr"})
        try:
            extracted = ocr_image_openai(img_base64, mime_type=img_mime)
        except Exception as _oe:
            _event_log_write("image_step_error", {"step": "ocr", "err": str(_oe)[:500]})
            logger.exception("OCR exception: %s", _oe)
            return
        _event_log_write("image_step", {
            "step": "ocr_done",
            "extracted_len": len(extracted) if extracted else 0,
            "extracted_preview": (extracted[:100] if extracted else None),
        })
        logger.info("Image OCR result: %s chars, text: %s", len(extracted) if extracted else 0, (extracted[:100] + "...") if extracted and len(extracted) > 100 else extracted)
        if not extracted or len(extracted.strip()) < 2:
            _event_log_write("image_aborted", {"reason": "ocr_empty_or_too_short", "len": len(extracted) if extracted else 0})
            return

        # === Check if this is a work order (製造指示書) ===
        try:
            wo_customer = detect_work_order(extracted)
            if wo_customer:
                _event_log_write("image_step", {"step": "work_order_detected", "customer": str(wo_customer)[:50]})
                # It's a work order — never translate work order content
                wo_on = group_wo_settings.get(group_id, True)
                if wo_on:
                    reply = format_storage_for_work_order(wo_customer)
                    if reply:
                        bot_stats["work_order_detections"] += 1
                        qt_wo = ctx["quote_token"]
                        with ApiClient(configuration) as api_client:
                            api = MessagingApi(api_client)
                            msg_obj = TextMessage(text=reply)
                            if qt_wo:
                                try:
                                    msg_obj.quote_token = qt_wo
                                except Exception:
                                    pass
                            api.reply_message(ReplyMessageRequest(
                                reply_token=ctx["reply_token"],
                                messages=[msg_obj]
                            ))
                        _event_log_write("image_done", {"path": "work_order"})
                # Whether storage found or not, skip translation for work orders
                track_group_usage(group_id, _bp, _bc)
                return
        except Exception as e:
            _event_log_write("image_step_error", {"step": "work_order_detect", "err": str(e)[:200]})
            logger.error("Work order detection error: %s", e)
        # === End work order check ===

        _event_log_write("image_step", {"step": "before_lang_detect"})
        lang = detect_language(extracted)
        _event_log_write("image_step", {"step": "lang_detected", "lang": lang})
        if lang is None:
            _event_log_write("image_aborted", {"reason": "lang_is_none"})
            return

        # Determine actual translation target
        if lang == "zh":
            actual_tgt = tgt
        else:
            actual_tgt = "zh"

        # Translate OCR text using the same translation engine as text messages
        # Set translation tone for this group
        _tone, _tone_custom = get_group_tone(group_id)
        _tl.tone = _tone
        _tl.tone_custom = _tone_custom
        _tl.from_image_ocr = True  # v3.9.27: 標記為圖片來源,觸發強制分段翻譯
        _event_log_write("image_step", {"step": "before_translate", "src": lang, "tgt": (tgt if lang == "zh" else "zh")})

        # v3.9.24: 用 threading 強制限制 translate() 最多 50 秒,
        # 因為 translate() 內部有多個 OpenAI 呼叫(分段、品質檢查、反譯),
        # 即使各別都有 timeout,加總可能超過 LINE webhook 限制
        import threading
        _trans_result = [None]
        _trans_exc = [None]
        def _do_translate():
            try:
                if lang == "zh":
                    _trans_result[0] = translate(extracted, "zh", tgt)
                else:
                    _trans_result[0] = translate(extracted, lang, "zh")
            except Exception as e:
                _trans_exc[0] = e
        _t = threading.Thread(target=_do_translate, daemon=True)
        _t.start()
        _t.join(timeout=50.0)  # 最多等 50 秒

        if _t.is_alive():
            # 超時,放棄等待(thread 會繼續但我們不管它)
            _event_log_write("image_step_error", {"step": "translate", "err": "thread timeout after 50s"})
            logger.error("Translate thread timeout after 50s")
            return
        if _trans_exc[0]:
            _event_log_write("image_step_error", {"step": "translate", "err": str(_trans_exc[0])[:300]})
            logger.exception("Translate exception: %s", _trans_exc[0])
            return
        result = _trans_result[0]
        _event_log_write("image_step", {
            "step": "translate_done",
            "result_len": len(result) if result else 0,
            "result_preview": (result[:100] if result else None),
        })

        if not result:
            _event_log_write("image_aborted", {"reason": "translate_returned_empty"})
            track_group_usage(group_id, _bp, _bc)
            return

        reply = "\U0001f5bc\ufe0f " + LANG_FLAGS.get(actual_tgt, "") + "\n" + result

        # LINE message limit is 5000 chars
        if len(reply) > 5000:
            reply = reply[:4990] + "\n..."

        track_group_usage(group_id, _bp, _bc)
        bot_stats["image_translations"] += 1
        _event_log_write("image_step", {"step": "before_reply"})
        # v3.8: thread quote_token onto image translation reply.
        qt = ctx["quote_token"]
        try:
            try:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    msg_obj = TextMessage(text=reply)
                    if qt:
                        try:
                            msg_obj.quote_token = qt
                        except Exception:
                            pass
                    api.reply_message(ReplyMessageRequest(
                        reply_token=ctx["reply_token"],
                        messages=[msg_obj]
                    ))
                _event_log_write("image_done", {"path": "auto_translate", "reply_len": len(reply), "method": "reply"})
            except Exception as _re_reply:
                # v3.9.23: reply_token 過期或無效 → fallback 用 push
                _event_log_write("image_step_error", {"step": "reply", "err": str(_re_reply)[:300], "fallback": "push"})
                logger.warning("Reply failed, trying push: %s", _re_reply)
                try:
                    with ApiClient(configuration) as api_client:
                        api = MessagingApi(api_client)
                        msg_obj = TextMessage(text=reply)
                        api.push_message(PushMessageRequest(
                            to=group_id,
                            messages=[msg_obj]
                        ))
                    _event_log_write("image_done", {"path": "auto_translate", "reply_len": len(reply), "method": "push"})
                except Exception as _pe:
                    _event_log_write("image_step_error", {"step": "push", "err": str(_pe)[:300]})
                    logger.exception("Push also failed: %s", _pe)
        except Exception as _re:
            _event_log_write("image_step_error", {"step": "reply_outer", "err": str(_re)[:300]})
            logger.exception("Reply exception: %s", _re)



    except Exception as _bg_e:
        _event_log_write("image_step_error", {"step": "bg_uncaught", "err": str(_bg_e)[:300]})
        logger.exception("[BG] handle_image_background uncaught: %s", _bg_e)


def _process_pending_image_translate(event, message_id):
    """v3.9.13: 外層 wrapper — 任何例外都不可吞掉,一定要送錯誤訊息給使用者"""
    try:
        return _process_pending_image_translate_inner(event, message_id)
    except Exception as e:
        logger.exception("[ImgAsk] UNCAUGHT exception: %s", e)
        # 嘗試任何方式回應
        try:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.reply_message(ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text="❌ 處理圖片時發生未預期錯誤 / Error tak terduga saat proses gambar\n" + str(e)[:200])]
                ))
        except Exception:
            # reply 失敗 → 試 push 到事件來源
            try:
                src = getattr(event, 'source', None)
                gid = (getattr(src, 'group_id', None) or
                       getattr(src, 'room_id', None) or
                       getattr(src, 'user_id', None))
                if gid:
                    with ApiClient(configuration) as api_client:
                        api = MessagingApi(api_client)
                        api.push_message(PushMessageRequest(
                            to=gid,
                            messages=[TextMessage(text="❌ 處理圖片時發生未預期錯誤 / Error tak terduga saat proses gambar\n" + str(e)[:200])]
                        ))
            except Exception as e2:
                logger.error("[ImgAsk] emergency push also failed: %s", e2)


def _process_pending_image_translate_inner(event, message_id):
    """v3.9.10: 詢問模式下,使用者按了「翻譯這張」之後執行實際翻譯。
    重用 handle_image 的下載/OCR/翻譯邏輯,但 reply_token 來自 postback event。
    
    v3.9.11: 增加全程 logging + 最終 fallback push,避免任何路徑靜默失敗。
    v3.9.13: 改用磁碟版 pending(跨 worker)+ 外層 wrapper catch-all。
    """
    # v3.9.13: 從磁碟讀(支援多 worker)
    _all_pending = _load_pending_imgs()
    logger.info("[ImgAsk] postback triggered for msg=%s, pending_count=%d",
                message_id, len(_all_pending))
    
    # v3.9.15: 直接 push(reply_token 已被外層 "🔄 正在翻譯圖片..." 用掉)
    # 從 event.source 抓 group_id;若失敗才 fallback 用 info 裡的
    _src = getattr(event, 'source', None)
    _push_to = (getattr(_src, 'group_id', None) or
                getattr(_src, 'room_id', None) or
                getattr(_src, 'user_id', None))
    
    def _reply_or_push(text):
        target = _push_to
        if not target and info:
            target = info.get("group_id")
        if not target:
            logger.error("[ImgAsk] no push target! cannot send: %s", text[:80])
            return False
        try:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.push_message(PushMessageRequest(
                    to=target,
                    messages=[TextMessage(text=text)]
                ))
            logger.info("[ImgAsk] push sent to %s: %s", target[:8], text[:80])
            return True
        except Exception as _pe:
            logger.error("[ImgAsk] push failed: %s | text=%s", _pe, text[:80])
            return False

    info = _pending_img_pop(message_id)
    if not info:
        # v3.9.30: 完全靜默 — 兩種情況:
        # (1) 外層已攔截過(已過期)→ 不應該到這裡,但保險
        # (2) Race condition:另一個 worker 已搶先 pop 並處理中 → 那個 worker 會 push 翻譯結果
        # 兩種情況都不該再發訊息打擾群組
        logger.warning("[ImgAsk] message %s not in pending (expired or other worker fetched) → silent", message_id)
        return

    group_id = info["group_id"]
    if not oai:
        _reply_or_push("❌ 系統未設定 OpenAI,無法 OCR 圖片\nSistem belum diatur, tidak bisa OCR gambar")
        return

    show_loading(group_id)

    # 下載圖片
    logger.info("[ImgAsk] downloading image %s", message_id)
    try:
        img_base64, img_raw = download_line_image(message_id)
    except Exception as _de:
        logger.error("[ImgAsk] download exception: %s", _de)
        _reply_or_push("❌ 下載圖片失敗 / Gagal unduh gambar\n" + str(_de)[:100])
        return
    if not img_base64:
        _reply_or_push("❌ 下載圖片失敗(LINE 端已過期或網路錯誤)\nGagal unduh gambar (kedaluwarsa di LINE atau error jaringan)")
        return
    # v3.9.17: 偵測 MIME 格式
    img_mime = detect_image_mime(img_raw)
    logger.info("[ImgAsk] downloaded %d bytes, mime=%s", len(img_raw) if img_raw else 0, img_mime)

    tgt = group_target_lang.get(group_id, "id")
    _bp, _bc = bot_stats.get("tokens_prompt", 0), bot_stats.get("tokens_completion", 0)
    
    logger.info("[ImgAsk] running OCR")
    try:
        extracted = ocr_image_openai(img_base64, mime_type=img_mime)
    except Exception as _oe:
        logger.error("[ImgAsk] OCR exception: %s", _oe)
        _reply_or_push("❌ OCR 失敗 / OCR gagal\n" + str(_oe)[:100])
        return
    if not extracted or len(extracted.strip()) < 2:
        logger.info("[ImgAsk] OCR returned no text")
        _reply_or_push("ℹ️ 圖片中沒偵測到可翻譯的文字\nTidak ada teks yang bisa diterjemahkan di gambar")
        return
    logger.info("[ImgAsk] OCR ok: %d chars", len(extracted))

    # 工單偵測
    try:
        wo_customer = detect_work_order(extracted)
        if wo_customer:
            wo_on = group_wo_settings.get(group_id, True)
            if wo_on:
                wo_reply = format_storage_for_work_order(wo_customer)
                if wo_reply:
                    bot_stats["work_order_detections"] += 1
                    _reply_or_push(wo_reply)
            track_group_usage(group_id, _bp, _bc)
            return
    except Exception as e:
        logger.error("[ImgAsk] Work order detection error: %s", e)

    lang = detect_language(extracted)
    if lang is None:
        logger.warning("[ImgAsk] lang detection returned None for: %s", extracted[:80])
        _reply_or_push("⚠️ 偵測不到語言,無法翻譯\nTidak bisa mendeteksi bahasa, gagal terjemahkan")
        return
    actual_tgt = tgt if lang == "zh" else "zh"
    logger.info("[ImgAsk] translating %s -> %s", lang, actual_tgt)

    _tone, _tone_custom = get_group_tone(group_id)
    _tl.tone = _tone
    _tl.tone_custom = _tone_custom
    _tl.group_id = group_id
    try:
        if lang == "zh":
            result = translate(extracted, "zh", tgt)
        else:
            result = translate(extracted, lang, "zh")
    except Exception as _te:
        logger.error("[ImgAsk] translate exception: %s", _te)
        _reply_or_push("❌ 翻譯失敗 / Terjemahan gagal\n" + str(_te)[:100])
        return

    if not result:
        logger.warning("[ImgAsk] translate returned empty")
        _reply_or_push("⚠️ 翻譯結果為空,請重試\nHasil terjemahan kosong, silakan coba lagi")
        track_group_usage(group_id, _bp, _bc)
        return

    reply_text = "\U0001f5bc\ufe0f " + LANG_FLAGS.get(actual_tgt, "") + "\n" + result
    if len(reply_text) > 5000:
        reply_text = reply_text[:4990] + "\n..."

    track_group_usage(group_id, _bp, _bc)
    bot_stats["image_translations"] += 1
    _reply_or_push(reply_text)
    logger.info("[ImgAsk] DONE")


@handler.add(MessageEvent, message=AudioMessageContent)
def handle_audio(event):
    """Handle audio/voice messages: Whisper STT + detect language + translate."""
    # v3.9.30c B15 修補: LINE webhook 重發保護
    if _is_redelivery(event) or _is_duplicate_message(getattr(event.message, 'id', None)):
        logger.warning("[handle_audio] redelivery or duplicate, skipping msg_id=%s",
                       getattr(event.message, 'id', None))
        return
    
    source = event.source
    group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', None)
    user_id = getattr(source, 'user_id', None)
    is_dm_aud = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)

    # v3.9.30c B14 修補: 重置 _tl thread-local 殘留(同 handle_message)
    try:
        _tl.from_image_ocr = False
        _tl.factory_audit = None
        _tl.rtc_fail_reason = None
        _tl.last_entry_id = None
        _tl.gpt_says_announcement = None
        _tl.struct_self_check = None
        _tl.multi_path_fail = None
    except Exception:
        pass

    # Record user for whitelist
    if group_id and user_id and not is_dm_aud:
        record_user_name(group_id, user_id)

    # Check if translation is on
    is_on = group_settings.get(group_id, True)
    if not is_on:
        return

    # Check skip list
    sender_id = user_id
    if sender_id and sender_id in group_skip_users.get(group_id, set()):
        return

    # DM master toggle check for audio
    if is_dm_aud and sender_id:
        if not dm_master_enabled and sender_id not in dm_whitelist:
            return

    # Check if audio translation is on
    audio_on = group_audio_settings.get(group_id, True)
    if not audio_on:
        return

    # Need OpenAI for Whisper
    if not oai:
        logger.warning("No OpenAI key, cannot do audio transcription")
        return

    show_loading(group_id)
    if get_group_feature(group_id, 'mark_read'):
        mark_as_read(group_id)

    # Download audio from LINE
    message_id = event.message.id
    audio_bytes = download_line_audio(message_id)
    if not audio_bytes:
        return

    # Transcribe with Whisper
    transcribed = transcribe_audio_openai(audio_bytes)
    if not transcribed or len(transcribed.strip()) < 2:
        return

    # Detect language
    lang = detect_language(transcribed)
    if lang is None:
        return

    tgt = group_target_lang.get(group_id, "id")

    reply = None
    _bp, _bc = bot_stats.get("tokens_prompt", 0), bot_stats.get("tokens_completion", 0)
    # Set translation tone for this group
    _tone, _tone_custom = get_group_tone(group_id)
    _tl.tone = _tone
    _tl.tone_custom = _tone_custom
    if lang == "zh":
        result = translate(transcribed, "zh", tgt)
        if result:
            reply = "\U0001f3a4 " + LANG_FLAGS.get(tgt, "") + "\n\U0001f4ac " + transcribed + "\n\U0001f4dd " + result
    else:
        result = translate(transcribed, lang, "zh")
        if result:
            reply = "\U0001f3a4 " + LANG_FLAGS.get("zh", "") + "\n\U0001f4ac " + transcribed + "\n\U0001f4dd " + result
    track_group_usage(group_id, _bp, _bc)

    if reply is None:
        return

    bot_stats["voice_translations"] += 1
    # v3.8: thread quote_token onto the reply so the translation visually
    # references the original audio bubble. Important in busy group chats.
    qt = getattr(event.message, 'quote_token', None)
    with ApiClient(configuration) as api_client:
        api = MessagingApi(api_client)
        msg_obj = TextMessage(text=_clip_line_text(reply))
        if qt:
            try:
                msg_obj.quote_token = qt
            except Exception:
                pass
        api.reply_message(ReplyMessageRequest(
            reply_token=event.reply_token,
            messages=[msg_obj]
        ))



if StickerMessageContent:
    @handler.add(MessageEvent, message=StickerMessageContent)
    def handle_sticker(event):
        """Record user name when they send a sticker (for whitelist tracking)."""
        source = event.source
        is_dm = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None)
        user_id = getattr(source, 'user_id', None)
        if group_id and user_id and not is_dm:
            record_user_name(group_id, user_id)


if VideoMessageContent:
    @handler.add(MessageEvent, message=VideoMessageContent)
    def handle_video(event):
        """Handle video messages: download thumbnail, OCR, translate."""
        # v3.9.30c B18 修補: video OCR 也會跑 OpenAI(花錢),需要 redelivery 保護
        if _is_redelivery(event) or _is_duplicate_message(getattr(event.message, 'id', None)):
            logger.warning("[handle_video] redelivery or duplicate, skipping msg_id=%s",
                           getattr(event.message, 'id', None))
            return
        source = event.source
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', None)
        user_id = getattr(source, 'user_id', None)
        is_dm = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
        if group_id and user_id and not is_dm:
            record_user_name(group_id, user_id)
        # Video OCR: try to get preview image and OCR it
        if not get_group_feature(group_id, 'video_ocr'):
            return
        if not group_settings.get(group_id, True):
            return
        if not group_img_settings.get(group_id, True) and not is_dm:
            return
        try:
            msg_id = event.message.id
            with ApiClient(configuration) as api_client:
                blob_api = MessagingApiBlob(api_client)
                # Get video content (preview/thumbnail)
                content = blob_api.get_message_content_preview(msg_id)
                if content and len(content) > 100:
                    b64 = base64.b64encode(content).decode()
                    # OCR the preview frame
                    ocr_result = ocr_image_openai(b64)
                    if ocr_result and len(ocr_result.strip()) > 2:
                        lang = detect_language(ocr_result)
                        if lang:
                            # Set translation tone for this group
                            _tone, _tone_custom = get_group_tone(group_id)
                            _tl.tone = _tone
                            _tl.tone_custom = _tone_custom
                            # v3.2-0426e: pass user_id and group_id for OpenAI metadata/user param
                            try:
                                _tl.user_id = getattr(getattr(event.source, 'user_id', ''), '__str__', lambda: '')() if hasattr(event.source, 'user_id') else ''
                            except Exception:
                                _tl.user_id = ''
                            _tl.group_id = group_id or ''
                            if lang == "zh":
                                tgt = group_target_lang.get(group_id, "id")
                                result = translate(ocr_result, "zh", tgt)
                                actual_tgt = tgt
                            else:
                                result = translate(ocr_result, lang, "zh")
                                actual_tgt = "zh"
                            if result:
                                reply = "🎬 " + LANG_FLAGS.get(actual_tgt, "") + " " + result
                                with ApiClient(configuration) as ac2:
                                    api2 = MessagingApi(ac2)
                                    api2.reply_message(ReplyMessageRequest(
                                        reply_token=event.reply_token,
                                        messages=[TextMessage(text=_clip_line_text(reply))]
                                    ))
                                bot_stats["image_translations"] += 1
        except Exception as e:
            logger.warning("Video OCR failed: %s", e)


if FileMessageContent:
    @handler.add(MessageEvent, message=FileMessageContent)
    def handle_file(event):
        """Handle file messages: record user, log file info."""
        source = event.source
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None)
        user_id = getattr(source, 'user_id', None)
        is_dm = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
        if group_id and user_id and not is_dm:
            record_user_name(group_id, user_id)
        fname = getattr(event.message, 'file_name', '未知檔案')
        fsize = getattr(event.message, 'file_size', 0)
        logger.info("File received: %s (%d bytes) from %s", fname, fsize, group_id)


if LocationMessageContent:
    @handler.add(MessageEvent, message=LocationMessageContent)
    def handle_location(event):
        """Handle location messages: translate location info."""
        source = event.source
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', None)
        user_id = getattr(source, 'user_id', None)
        is_dm = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
        if group_id and user_id and not is_dm:
            record_user_name(group_id, user_id)
        if not group_settings.get(group_id, True):
            return
        if not get_group_feature(group_id, 'location'):
            return
        # Translate location title/address if available
        title = getattr(event.message, 'title', '') or ''
        address = getattr(event.message, 'address', '') or ''
        if title or address:
            loc_text = (title + " " + address).strip()
            lang = detect_language(loc_text)
            if lang:
                # Set translation tone
                _tone, _tone_custom = get_group_tone(group_id)
                _tl.tone = _tone
                _tl.tone_custom = _tone_custom
                if lang == "zh":
                    tgt = group_target_lang.get(group_id, "id")
                    result = translate(loc_text, "zh", tgt)
                    actual_tgt = tgt
                else:
                    result = translate(loc_text, lang, "zh")
                    actual_tgt = "zh"
                if result:
                    try:
                        with ApiClient(configuration) as api_client:
                            api = MessagingApi(api_client)
                            api.reply_message(ReplyMessageRequest(
                                reply_token=event.reply_token,
                                messages=[TextMessage(text="📍 " + LANG_FLAGS.get(actual_tgt, "") + " " + result)]
                            ))
                    except Exception:
                        pass


if JoinEvent:
    @handler.add(JoinEvent)
    def handle_join(event):
        """Track when bot joins a group."""
        source = event.source
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None)
        if not group_id:
            return
        gname = ""
        try:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                summary = api.get_group_summary(group_id)
                gname = summary.group_name or ""
        except Exception:
            pass
        group_tracking[group_id] = {"name": gname, "joined_at": time.time()}
        save_settings()

if MemberJoinedEvent:
    @handler.add(MemberJoinedEvent)
    def handle_member_joined(event):
        """Send bilingual welcome when a new member joins the group."""
        # v3.9.30c B18 修補: 避免重發歡迎訊息
        if _is_redelivery(event):
            logger.warning("[handle_member_joined] redelivery, skipping")
            return
        _rtok = getattr(event, 'reply_token', None)
        if _rtok and _is_duplicate_message("mjn:" + _rtok):
            logger.warning("[handle_member_joined] duplicate, skipping")
            return
        source = event.source
        group_id = getattr(source, 'group_id', None)
        if not group_id:
            return
        # Record new members
        members = getattr(event, 'joined', None)
        if members and hasattr(members, 'members'):
            for member in members.members:
                uid = getattr(member, 'user_id', None)
                if uid:
                    record_user_name(group_id, uid)
        # Send welcome if enabled
        ws = get_group_welcome(group_id)
        if not ws.get("enabled", True):
            return
        if not group_settings.get(group_id, True):
            return
        try:
            zh = ws.get("text_zh", "")
            id_text = ws.get("text_id", "")
            welcome = zh + "\n\n" + id_text if zh and id_text else (zh or id_text)
            if welcome:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=_clip_line_text(welcome))]
                    ))
        except Exception as e:
            logger.warning("Failed to send welcome: %s", e)


if MemberLeftEvent:
    @handler.add(MemberLeftEvent)
    def handle_member_left(event):
        """Track when a member leaves the group."""
        source = event.source
        group_id = getattr(source, 'group_id', None)
        if not group_id:
            return
        left = getattr(event, 'left', None)
        if left and hasattr(left, 'members'):
            for member in left.members:
                uid = getattr(member, 'user_id', None)
                if uid:
                    # Remove from skip list
                    if group_id in group_skip_users:
                        group_skip_users[group_id].discard(uid)
                    logger.info("Member %s left group %s", uid, group_id)


if FollowEvent:
    @handler.add(FollowEvent)
    def handle_follow(event):
        """Track when a user adds the bot as friend."""
        user_id = getattr(event.source, 'user_id', None)
        if not user_id:
            return
        try:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                profile = api.get_profile(user_id)
                dm_known_users[user_id] = profile.display_name or user_id
                lang = getattr(profile, 'language', None)
                if lang:
                    user_languages[user_id] = lang
        except Exception:
            dm_known_users[user_id] = user_id
        bot_stats["followers"] = bot_stats.get("followers", 0) + 1
        save_settings()
        logger.info("New follower: %s", dm_known_users.get(user_id, user_id))


if UnfollowEvent:
    @handler.add(UnfollowEvent)
    def handle_unfollow(event):
        """Track when a user blocks/removes the bot."""
        user_id = getattr(event.source, 'user_id', None)
        if user_id:
            bot_stats["unfollowers"] = bot_stats.get("unfollowers", 0) + 1
            logger.info("Unfollowed by: %s", user_id)


if BotLeaveEvent:
    @handler.add(BotLeaveEvent)
    def handle_bot_leave(event):
        """Clean up when bot is removed from a group."""
        group_id = getattr(event.source, 'group_id', None) or getattr(event.source, 'room_id', None)
        if group_id:
            group_tracking.pop(group_id, None)
            group_settings.pop(group_id, None)
            group_target_lang.pop(group_id, None)
            group_img_settings.pop(group_id, None)
            group_audio_settings.pop(group_id, None)
            group_wo_settings.pop(group_id, None)
            group_skip_users.pop(group_id, None)
            group_user_names.pop(group_id, None)
            save_settings()
            logger.info("Bot removed from group %s", group_id)


if PostbackEvent:
    @handler.add(PostbackEvent)
    def handle_postback(event):
        """Handle postback actions from Quick Reply / Flex buttons."""
        # v3.9.30c B18 修補: postback 也會 redelivery,且會跑翻譯花錢
        # postback 沒有 message.id,改用 reply_token 當 dedup key(每個 reply_token 唯一)
        if _is_redelivery(event):
            logger.warning("[handle_postback] redelivery, skipping")
            return
        _rtok = getattr(event, 'reply_token', None)
        if _rtok and _is_duplicate_message("pbk:" + _rtok):
            logger.warning("[handle_postback] duplicate reply_token, skipping")
            return
        data = event.postback.data if hasattr(event.postback, 'data') else ""
        logger.info("Postback: %s", data)

        # Parse postback data (format: action=X&lang=Y&...)
        try:
            params = dict(urllib.parse.parse_qsl(data))
        except Exception:
            params = {}

        action = params.get("action", "")

        # /help language switch: re-send Flex with the other language first
        if action == "help":
            lang = params.get("lang", "zh")
            if lang not in ("zh", "id"):
                lang = "zh"
            send_help_flex(event.reply_token, primary_lang=lang)
            return

        # v3.9.10: 圖片翻譯詢問模式 — 使用者按了「翻譯這張」
        if "img_translate" in params:
            msg_id = params["img_translate"]
            logger.info("[ImgAsk] postback received, msg_id=%s", msg_id)

            # v3.9.30: 過期 / 已處理 → 完全靜默,不發任何訊息(避免騷擾群組)
            # 先檢查 pending 是否還在;不在就直接 return,連 ack 都不發
            _pending_check = _load_pending_imgs()
            if msg_id not in _pending_check:
                logger.info("[ImgAsk] msg_id=%s not in pending (expired or already processed) → silent ignore", msg_id)
                return

            # 只有確認還在 pending 才 ack + 進入翻譯流程
            # 立刻 reply 確認收到(reply_token 必須在 1 分鐘內用掉)
            # 然後用 push 發實際翻譯結果(OCR + translate 可能需要 10+ 秒)
            try:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text="🔄 正在翻譯圖片...\nSedang menerjemahkan gambar...")]
                    ))
                logger.info("[ImgAsk] ack reply sent")
            except Exception as _ake:
                logger.warning("[ImgAsk] ack reply failed (token may be expired): %s", _ake)
            # 然後處理 OCR + 翻譯,用 push 送結果
            _process_pending_image_translate(event, msg_id)
            return
        if "img_skip" in params:
            # v3.9.30: 舊版 Flex 殘留的按鈕(新版已移除「跳過」),保留處理避免 500
            # 過期 → 完全靜默;還在 pending → pop 掉並回覆已跳過
            _msgid = params["img_skip"]
            _pending_check = _load_pending_imgs()
            if _msgid not in _pending_check:
                logger.info("[ImgAsk] img_skip msg_id=%s expired → silent ignore", _msgid)
                return
            _pending_img_pop(_msgid)
            try:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text="✅ 已跳過 / Dilewati")]
                    ))
            except Exception:
                pass
            return


if UnsendEvent:
    @handler.add(UnsendEvent)
    def handle_unsend(event):
        """Clean up cached message when user unsends."""
        msg_id = getattr(event.unsend, 'message_id', None) if hasattr(event, 'unsend') else None
        if msg_id and msg_id in message_cache:
            del message_cache[msg_id]
            logger.info("Unsend: removed message %s from cache", msg_id)


if VideoPlayCompleteEvent:
    @handler.add(VideoPlayCompleteEvent)
    def handle_video_play_complete(event):
        """Handle video viewing complete event — log for analytics."""
        source = event.source
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', '')
        user_id = getattr(source, 'user_id', '')
        tracking_id = getattr(event.video_play_complete, 'tracking_id', '') if hasattr(event, 'video_play_complete') else ''
        logger.info("VideoPlayComplete: group=%s user=%s tracking=%s", group_id, user_id, tracking_id)
        bot_stats["video_play_complete"] = bot_stats.get("video_play_complete", 0) + 1


# v3.8: --- Reaction event feedback pipeline ---

def register_sent_message(message_id, entry_id, group_id):
    """Record that bot sent message_id as the rendering of translation entry_id.
    Allows reaction-event feedback to find which translation a user reacted to.
    Bounded LRU; oldest evicted at SENT_MSG_MAP_MAX.
    """
    if not message_id or not entry_id:
        return
    try:
        sent_message_to_entry[message_id] = {
            "entry_id": entry_id,
            "group_id": group_id or "",
            "ts": int(time.time()),
        }
        # Evict oldest if oversized.
        while len(sent_message_to_entry) > SENT_MSG_MAP_MAX:
            sent_message_to_entry.popitem(last=False)
    except Exception:
        pass


def _apply_reaction_feedback(message_id, reaction_type, user_id, group_id):
    """Look up which translation entry the reacted message corresponds to,
    then apply positive/negative feedback to translation_log.

    Positive (love/like/happy): if entry has a known correct translation,
        promote it to examples; else just record the positive vote.
    Negative (sad/surprise): mark the entry as wrong (operator can later add
        a correction via admin panel or LINE command).
    """
    rec = sent_message_to_entry.get(message_id)
    if not rec:
        return False, "unknown_message"
    entry_id = rec.get("entry_id")
    target = None
    for e in translation_log:
        if e.get("id") == entry_id:
            target = e
            break
    if not target:
        return False, "entry_evicted"

    rt = (reaction_type or "").lower()
    feedback = target.setdefault("reaction_feedback", {"positive": 0, "negative": 0, "voters": {}})
    voters = feedback["voters"]
    prev = voters.get(user_id or "")

    if rt in REACTION_POSITIVE:
        if prev != "positive":
            feedback["positive"] += 1
            if prev == "negative":
                feedback["negative"] = max(0, feedback["negative"] - 1)
            voters[user_id or ""] = "positive"
        # If 2+ positive votes and not yet a training example, promote it.
        try:
            if feedback["positive"] >= 2 and not target.get("promoted_to_examples"):
                src_text = target.get("src", "")
                tgt_text = target.get("tgt", "")
                src_lang = target.get("src_lang", "zh")
                if src_text and tgt_text and not tgt_text.startswith("⚠️"):
                    direction = "zh2id" if src_lang == "zh" else "id2zh"
                    new_ex = {
                        "zh": src_text if src_lang == "zh" else tgt_text,
                        "id": tgt_text if src_lang == "zh" else src_text,
                        "dir": direction,
                        "source": "reaction_positive",
                    }
                    if new_ex not in custom_translation_examples:
                        custom_translation_examples.append(new_ex)
                        # Cap at configured max to prevent unbounded growth.
                        try:
                            cap = int(CUSTOM_EXAMPLES_MAX)
                        except Exception:
                            cap = 1000
                        if len(custom_translation_examples) > cap:
                            custom_translation_examples[:] = custom_translation_examples[-cap:]
                        target["promoted_to_examples"] = True
                        try:
                            _save_examples_to_disk()
                        except Exception:
                            pass
                        try:
                            _save_translation_log_to_disk()
                        except Exception:
                            pass
                        logger.info("Reaction promoted entry %s to examples", entry_id)
        except Exception as ee:
            logger.warning("Reaction promote failed: %s", ee)
        return True, "positive"

    if rt in REACTION_NEGATIVE:
        if prev != "negative":
            feedback["negative"] += 1
            if prev == "positive":
                feedback["positive"] = max(0, feedback["positive"] - 1)
            voters[user_id or ""] = "negative"
        # Mark wrong if 1+ negative vote (single complaint is enough to flag).
        if not target.get("marked_wrong"):
            target["marked_wrong"] = True
            target["marked_wrong_source"] = "reaction"
            try:
                _save_translation_log_to_disk()
            except Exception:
                pass
        return True, "negative"

    return False, "ignored_reaction_type"


if MessageReactionEvent:
    @handler.add(MessageReactionEvent)
    def handle_reaction(event):
        """v3.8: Free quality signal from group members.
        ❤️/👍 → promote the translation to training examples (after 2 positive votes).
        😢/😮 → flag the translation as wrong, surface to admin panel.
        """
        try:
            source = event.source
            user_id = getattr(source, 'user_id', '')
            group_id = (getattr(source, 'group_id', None)
                        or getattr(source, 'room_id', None)
                        or user_id or '')
            # event.reaction is a struct with `type` and `emoji`-related fields
            rx = getattr(event, 'reaction', None)
            rtype = getattr(rx, 'type', '') if rx else ''
            # `messageId` identifies which message was reacted to
            msg_id = getattr(event, 'message_id', None) or (
                getattr(rx, 'message_id', None) if rx else None)
            if not msg_id:
                return
            ok, info = _apply_reaction_feedback(msg_id, rtype, user_id, group_id)
            if ok:
                logger.info("Reaction feedback applied: msg=%s type=%s result=%s",
                            msg_id, rtype, info)
        except Exception as e:
            logger.warning("Reaction handler error: %s", e)


def show_loading(chat_id):
    """Show typing indicator before translation."""
    if not ShowLoadingAnimationRequest:
        return
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.show_loading_animation(ShowLoadingAnimationRequest(chat_id=chat_id))
    except Exception:
        pass


def mark_as_read(chat_id):
    """Mark messages as read in the chat (shows 'read' indicator)."""
    if not MarkMessagesAsReadRequest:
        return
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.mark_messages_as_read(MarkMessagesAsReadRequest(chat_id=chat_id))
    except Exception as e:
        logger.debug("mark_as_read failed: %s", e)


def generate_retry_key():
    """Generate a UUID v4 for X-Line-Retry-Key header to prevent duplicate sends."""
    return str(uuid.uuid4())


def safe_reply(reply_token, messages, retry=True):
    """Reply with optional X-Line-Retry-Key for idempotency."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            req = ReplyMessageRequest(reply_token=reply_token, messages=messages)
            if retry:
                # SDK v3 supports x_line_retry_key param
                try:
                    api.reply_message(req, x_line_retry_key=generate_retry_key())
                except TypeError:
                    api.reply_message(req)
            else:
                api.reply_message(req)
    except Exception as e:
        logger.warning("safe_reply failed: %s", e)


def safe_push(to, messages, retry=True):
    """Push with optional X-Line-Retry-Key for idempotency."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            req = PushMessageRequest(to=to, messages=messages)
            if retry:
                try:
                    api.push_message(req, x_line_retry_key=generate_retry_key())
                except TypeError:
                    api.push_message(req)
            else:
                api.push_message(req)
    except Exception as e:
        logger.warning("safe_push failed: %s", e)


# ---- Webhook Management API ----
def get_webhook_info():
    """Get current webhook endpoint info."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            info = api.get_webhook_endpoint()
            return {
                "endpoint": getattr(info, 'endpoint', ''),
                "active": getattr(info, 'active', None),
            }
    except Exception as e:
        logger.warning("get_webhook_info failed: %s", e)
        return None


def set_webhook_url(url):
    """Set webhook endpoint URL via API."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            from linebot.v3.messaging import SetWebhookEndpointRequest
            api.set_webhook_endpoint(SetWebhookEndpointRequest(endpoint=url))
            return True
    except Exception as e:
        logger.warning("set_webhook_url failed: %s", e)
        return False


def test_webhook(endpoint=None):
    """Test webhook endpoint."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            from linebot.v3.messaging import TestWebhookEndpointRequest
            if endpoint:
                resp = api.test_webhook_endpoint(TestWebhookEndpointRequest(endpoint=endpoint))
            else:
                resp = api.test_webhook_endpoint(TestWebhookEndpointRequest())
            return {
                "success": getattr(resp, 'success', None),
                "timestamp": getattr(resp, 'timestamp', ''),
                "status_code": getattr(resp, 'status_code', None),
                "reason": getattr(resp, 'reason', ''),
                "detail": getattr(resp, 'detail', ''),
            }
    except Exception as e:
        logger.warning("test_webhook failed: %s", e)
        return {"success": False, "reason": str(e)}


# ---- Content Preview & Preparation Status ----
def get_content_preview(message_id):
    """Get a preview image of an image or video message."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            content = blob_api.get_message_content_preview(message_id)
            return content
    except Exception as e:
        logger.warning("get_content_preview failed: %s", e)
        return None


def check_content_preparation(message_id):
    """Check if video/audio content is ready for download."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            resp = blob_api.get_message_content_transcoding_by_message_id(message_id)
            status = getattr(resp, 'status', 'unknown')
            return status  # "processing", "succeeded", "failed"
    except Exception as e:
        logger.debug("check_content_preparation failed: %s", e)
        return "unknown"


# ---- Validate Message Objects ----
def validate_message_objects(messages, msg_type="reply"):
    """Validate message objects before sending. Returns True/error dict."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            if msg_type == "push":
                api.validate_push({"messages": messages})
            elif msg_type == "broadcast":
                api.validate_broadcast({"messages": messages})
            elif msg_type == "multicast":
                api.validate_multicast({"messages": messages})
            else:
                api.validate_reply({"messages": messages})
            return {"valid": True}
    except Exception as e:
        return {"valid": False, "error": str(e)}


def get_line_quota():
    """Get LINE monthly message quota info."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            quota = api.get_message_quota()
            consumption = api.get_message_quota_consumption()
            return {
                "quota": getattr(quota, 'value', None),
                "type": getattr(quota, 'type', None),
                "used": getattr(consumption, 'total_usage', None),
            }
    except Exception:
        return None


def get_follower_count():
    """Get bot follower count."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_number_of_followers(var_date=time.strftime("%Y%m%d"))
            return getattr(resp, 'followers', None)
    except Exception:
        return None


def get_bot_info():
    """Get bot's own profile info."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            info = api.get_bot_info()
            return {
                "name": getattr(info, 'display_name', ''),
                "picture": getattr(info, 'picture_url', ''),
                "status": getattr(info, 'chat_mode', ''),
            }
    except Exception:
        return None


def get_group_member_count(group_id):
    """Get group member count from LINE API."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            count = api.get_group_members_count(group_id)
            return count
    except Exception:
        return None


def fetch_all_group_members(group_id):
    """Fetch all member IDs in a group using LINE API."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            member_ids = []
            token = None
            while True:
                try:
                    if token:
                        resp = api.get_group_members_ids(group_id, start=token)
                    else:
                        resp = api.get_group_members_ids(group_id)
                except AttributeError:
                    # Try alternative method name
                    if token:
                        resp = api.get_group_member_ids(group_id, start=token)
                    else:
                        resp = api.get_group_member_ids(group_id)
                # Extract member IDs from response
                ids = getattr(resp, 'member_user_ids', None) or getattr(resp, 'member_ids', None) or []
                member_ids.extend(ids)
                token = getattr(resp, 'next', None) or getattr(resp, 'next_token', None)
                if not token:
                    break
            # Record names for all members
            for uid in member_ids:
                record_user_name(group_id, uid)
            logger.info("Fetched %d members from group %s", len(member_ids), group_id)
            return member_ids
    except Exception as e:
        logger.warning("Failed to fetch group members: %s (type: %s)", e, type(e).__name__)
        return []


def push_message_to_group(group_id, text):
    """Push a message to a group (not a reply)."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.push_message(PushMessageRequest(
                to=group_id,
                messages=[TextMessage(text=text)]
            ))
            return True
    except Exception as e:
        logger.warning("Push message failed: %s", e)
        return False


def setup_rich_menu():
    """Create a rich menu with common bot actions."""
    if not RichMenuRequest:
        logger.warning("RichMenuRequest not available, skipping rich menu setup")
        return None
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            # Delete ALL existing rich menus first
            try:
                api.cancel_default_rich_menu()
            except Exception:
                pass
            try:
                existing = api.get_rich_menu_list()
                for rm_old in (existing.richmenus or []):
                    try:
                        api.delete_rich_menu(rm_old.rich_menu_id)
                    except Exception:
                        pass
            except Exception:
                pass
            # Create new rich menu (2 rows x 3 columns)
            rm = RichMenuRequest(
                size=RichMenuSize(width=2500, height=1686),
                selected=True,
                name="翻譯Bot選單",
                chat_bar_text="📋 選單",
                areas=[
                    RichMenuArea(
                        bounds=RichMenuBounds(x=0, y=0, width=833, height=843),
                        action=MessageAction(label="說明", text="/help")
                    ),
                    RichMenuArea(
                        bounds=RichMenuBounds(x=833, y=0, width=834, height=843),
                        action=MessageAction(label="狀態", text="/status")
                    ),
                    RichMenuArea(
                        bounds=RichMenuBounds(x=1667, y=0, width=833, height=843),
                        action=MessageAction(label="查儲區", text="/qry ")
                    ),
                    RichMenuArea(
                        bounds=RichMenuBounds(x=0, y=843, width=833, height=843),
                        action=MessageAction(label="翻譯開", text="/on")
                    ),
                    RichMenuArea(
                        bounds=RichMenuBounds(x=833, y=843, width=834, height=843),
                        action=MessageAction(label="不翻我", text="/skip")
                    ),
                    RichMenuArea(
                        bounds=RichMenuBounds(x=1667, y=843, width=833, height=843),
                        action=MessageAction(label="公告", text="/notice ")
                    ),
                ]
            )
            result = api.create_rich_menu(rm)
            rid = result.rich_menu_id
            # Generate simple image for rich menu
            _upload_rich_menu_image(api_client, rid)
            # Set as default
            api.set_default_rich_menu(rid)
            logger.info("Rich menu created: %s", rid)
            return rid
    except Exception as e:
        logger.warning("Rich menu setup failed: %s", e)
        return None


# Rich Menu image (2500x843 JPEG, base64 encoded)
_RICH_MENU_IMG_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAA0JCgsKCA0LCgsODg0PEyAVExISEyccHhcgLikxMC4pLSwzOko+MzZGNywtQFdBRkxOUlNSMj5aYVpQYEpRUk//2wBDAQ4ODhMREyYVFSZPNS01T09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT0//wAARCAaWCcQDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwDgKKKWthCUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUABooopgFFFFAC0lAooAWiiikAUUUtACUUUUwCiiigAooooAKKKKAFooooAKKKKACiiigAooooAKKKKACiiigBaKKKBBSUtFABRRRQAUUUUAFFFFMApKWkoAWiiikMO1FFFMAooooAKKKKACiiigApaSigAooopgFFFFAC0UlLSAKKKKACjFFJTAWiiigAooooAKKKKACiiigApKWigAooooAWikpaACiiigBKKWigBKKKKACiiloASlopKACiiigApaKSgApaSlzQAUUUUAFFFFABRRRQAUUUUwCiiikAUUUUAFFFFAAKWiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFACiigUUwCkpTRQAUUUUgCiiimAUdaKSgBaKKKQBRRRTAKKKKACiiigYUUtFAhKWkpaACiiigAooooAKKKKBhRRRQIKWkooAWikpaACiiigAooooAKKKKACiiigAooooAKKKKAA0lLSU0AtFFFABRRRQAUUUUAFFFFABRRS0AJRRRTAKU0lFABRRRQAUUUUgFpKKKAFooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLRRQAUUUUAFFFFAwooooAKKKKBBRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKAFooooAKKKKACiiigAopKWgAooooAKKKKACiiigAooooAKKKWgYlFFFAgooooGFFFFABRRRQAUvaiimAlFLRQAUUUUCDiiiimMp0UUVykhSUUUDCiiigAooooAKKKKAClpKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgFFFFABRRRQAUUUd6AFoopaQCUUUUAFFFBpgFFFJQAtFFFABS0lLQAUUUUAFFFFABRRRQAUUUUAFFFFABS0lLQAUUUUCCiiigAooooAKKKKACiiimAUlFLQMKKKKACiiigAooooAKKKKAFpKWkoAKO9FFABRRRTAKKKWgBKWkpaQBRRRQAUlFFMBaKKTvQAtFFFABRRSUALRRRQAUUUUAFLSd6KAFoopKACiiigAooooAKKKKACiiigApaSigAooooAKKKKAFoo7UUAFFJS0wCkopaACiiigAooooAKKKKQBRRRQAUopKUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoo70UAFLSUooAKKKSmAtFJS0AFFFFABRRRQAlFLRQAUUUUAFFFFABRRRQAUUUtABRSUUALRSUtAwooooEFFFFACUtFFABRRQKAFpKWkoAWiiigAooooAKKKKACiiigAooooAKKKKACiiigBKWkNFMBaKKKACiiigAooooAKKKKACiiigApaSlpgJRS0lAAaKKKQB1ooooAKKWkpgLRRRQMO1FFFAgooooAKKKKACiiigAooooAKKKKAFooooAKKKKBhRRRQAUUUUAFFFFAg7UUUUAFFFFAAKKBS0AJRRRTAKKKKACiiigAooooAKKKKAFooooGFFFFAgooooAKKKKACiiigAooooAKKKKACiiigYUtJRQAtJS0lABRRRQAUUUUAFFFLTAKKKKACikpaBBRRRQAUUUUDKdFJ3ormJCiiigYUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABS0lFABRRRQAUUUUAFFFFABRRRQAUUUUAFFLSUAFFFFMAooooAKUUlKKAFpKKKQC0neiimAUUUUAFFFFABSUtFABS0gpaACiiigAooooAKKKKACiiigAooooAKWkpaACiiigQUUUUAFFFFACUtFFACUUUvamAgpaKKQwooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUtMBKKKKACloopAFFFFCASiilpgFFFFABRRSUALRSUUAFGaKKAFpaQUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUCgBaKSloASloopgJS0UUAFFJS0AFFFFABRRRQAUUUUgFooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAd6WkpaYBQaKKACiiigAooooAKSiloAKKSloAKKKKACiiigAooooAKWkpaACiiigAooooAKKKKACiiigAooooAKKKKAF5pKKKAClpKWmAUUUUgCiiigAooooAKKKKACiiigApKU0lMApaSloAKKKKACiiigAooooAKKKKACiiigAooopgLSUtJQAUUUUAFFFFAwpRSUtAgooooAKKKKACiiigAooooAKKKKACiiigAopaSgApaSigBaKKKBhRRRQAUUUUAFFFFABRRRQIKKKKACiiimAUUUUAFFFFABRRRQAUUUtACUUUtABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUDCiiigAooooAKKKKACjNFFABS0gzmlpgFFFFABRRRQIKKKKQBRRRVAU6KSiuUAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAClpKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKWgQUUUUAFFJS0AJRS0lMYtFAopAFFFFABRRRTAKKM0UAFFFFABRRR3oAWiiigAooooAKKKKACiiigAooooAKKKKACloFFAgooooAKKKKACiiigAooooASiilpjCiiikAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKYBRRRSAKWkpRQAUlLSUwClpKKAFopKKAFpKKKAFpKKWgAFFFJQAvaiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACl7UCigAoopKAFopKKAFopKWgAopKWmAUUUUAFFFFABRRRSAWiiigAooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFAoAKWiimAUUUUAFFFFABSUtJQAUUUUAFFFGaAClopKAFopKWgAooooAKWkpaACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgFLSUtABR3pKKAFooopAFFFFABRRRQAlFFFMApaSloAKKKKACiiigAooooAKKKKACiiimAtFFJQAUUUUAFFFFABRRS0AJS0CikAUUUUwCiiigAooooGFFFFABRRRQIKKKKACiiigAoFFFAxaKKSgBaBSUtABRRRQAUUUUAFFFFAgooooAKKKKYBRRRQMKKKKBBRRRQAUUUUAHelpKWgAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKBhRRRQAUUUUAFFFFABRRRQAoopBS0wCiiigQUUUUAFFFFIAoooqgKVFFFcoBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLRRQAlFFFABRRRQAtFFFAgooooAKKKKACiiimMBS0UUgCkoooAKKKKYBRRRQAUUUUAFLSUCgBaKKKACiiigAooooAKKKKACiiigAooooAWiiigQUUUUAFFFFACUtFFABRRSUwCloooGFFFFABRRRQAUUUUAFFFFAAKKKKACiiigAooo7UwCiiigAooopALRQKKACkoopgFFFFABS0lFAC0lFFAC0UlGaAClFJ1paACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAFooopgJRS0UgEooooAKKKKAFooFFMAooooAKKKKQBS0lLQAUUUUAFFFFMBKWiigAooooAKKKKACiikoAWkpaKACiiigAooooAKKKKACiiigApaKKYBRRRQAUUUUAFFFJQAtFFJQAUUUUAFFFFABRRRQAUUUtABRRRQAtFJS0AJS0UlAC0UUUAFFFFABRRRQAUUUUAFFFFAwooooEFFFFMYUUUUAFFFFAgooooAKKKKAFoooNIBKKKWmAUlLRQAUUUUAFFFFABRSUtABRRRQAUUUUwCiiigAooooAKKKKAClpKKBi0UCigQUUUUAFFFFABRRRQMKKKKACiiigQUUUUAFFLSUAFFFLQMKSlooAKKKKACiiigAooooEFFFFAwooooEFFFFMAooooGFFFFAgooooAKKKKAFFFHaikAUUUUxhRRRQIKKKKACiiigAooooAKKKKACiiigAooooGFFFFABRRRQAUUUUAFFFFAC0UUlMBaKKKBBRRRSAKKKKBhRRRTApUUUVzCCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigApaSloAKKKKBBRRRQAUUUlABS0UCgYtJRRQAUUUUAFFFFMAooooAKKKKAClpKWgAooooAKKKKACiiigAooooAKKKKACiiloEFFFFABRRRQAUUUUAFJS0UAFJS0lMYtFFFIAooopgFFFFABRRRQAUUUUAFFFFABRRRmgANFGaKYBRRRQAUtJRSAWigUUAJRRRTAKWkooAKKKKACiiigApaSigApaO1FABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLSUGgBaSiigBaKSigAooooAWkopaACiiimAUUUUAFFFFIBaKKKACiiigAooopgFFFFABRRRQAUUUUAFFJS0AFFFJQAtFFFABRRRQAUUUUAFFFHegBaKKKYBRRRQAUUUlAC0lFFABRRRQAUUUUAFFFFABRRRQAUtJS0AFFFFABRRRQAUtJRQAtFFFABRRRQMKKKKACiiigQUUUUAFFFFMYUUUUAFFFFABRRRQIKKKKACiigUALSUtJQAUUUtACUtFFABRRRQAUUUUAFFFFABRRSUALRRRTAKKKKACiiigAooo7UALSUUUALRRRQMKKKKBBRRRQMKKKKACiiigQUUUUAFFFFAwooooELRRRQMKKKKACiiigAooooAKKKKACiiigQUUUUAFFFFABRS0lMAooooAKKKKACjvRS9qACiiikAUUUUAFFFFMAooooAKKKKACiiigAooooAKKKKBhRRRQAUUUUAFFFFABS0lFABS0lLTAKKKKBBSUtFABRRRSAKKKKBhRRRTApUUUVzCCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoopaBBRRRQAUUUUAFFJS0AFApBS0DCiiloASiiimAUUUUAFFFFABRRRQAUtFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLSUtABRRRQIKKKKACiiigAopKWmAUUUUDCiiigBaSiigAooooAKKKKACiiigAooooAKKKKYBRRRQAUUUtIBKKWigApKWkoAKKKKYBRRRQAUUUUAFFFFABRRS0AFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAKKSlpKACiiigBaKSigBaKSloASloooAKKWkoAKKKKACiiloAKKKKACiiimAUUUUAFFFFACUtFFABRRRQAUUUUAFFFFABRRRQAfSiiigAooooAKWkooAKKKKACiiigAooopgFFFFABRRRQAUUUUAFFFFABS0UUAFGKKKACiiigAooooAWiiigAooooAKKKKACiiigYUtJRQIKKKKACiiimAUUUUDCiiigQUUUUAFFFFABRRRQAtJRRQAtJS0UAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMAooooAKKKKACiiikAtFFFABRRRTAKKKKACiiikAUUlLTAKKKKACiiigAooooAWiiigYUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUCCiiigAooopgFFFFACiiiikAUUUUAFFFFAwpKWkpiFopKWgAooooGFFFFAgooooGFFFFABRRRQAUUUUAFFFFABRRRQAUtFFMApKWigQUUUUAFFFFABRRRQAUUUUDKVFFFcwgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKWkoAKKWigBKKKKACiiloASiiigAoopaAEpaSigBaKSigBaKSimAUUUUALRRR2pALSUUUAFFFFMAooooAKKKKACilpKAFooooAKKKKACiiigAooooAKKKKACiiigApaSloEFFFJQAtFFFABRSUUALSUUUwFooopDCiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACilooASlFJS0gCkpaKYCUtJRQAUtFJQAUUUUAFFLRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUtJRQAtFJRQAUUUtACUtFFABRRRQAUUUUAGaKKWgBKKKKAClpKWgAooopgFFFFABRRRQAlLRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAAooooAKKKKACiiigBaSiigAooopgFFFFABRRS0AFJS0UAJS0UUAFFFFABRRRQAUUUUAFLSUUAFLRRQAUUlLQAUUUUAFFFFABRS0UAFJS0UwEooooAKKKKACiiigAooooAKKKKBhRRRQIWikpaACiiigApKWigAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFAooFAC0UUUAFFFFACUtFFABRRRSAKKKKYBRRRQAUUtJQAUtFFABRRRQMKKKKACiiigAooooAKKKKBBRRRQAUUUUAFFFFABRRRQAUUUUwCjvRQKAFopKWkAUUlFAC0UlLQMKSiimIKWkpaACiiigAooooAKKKDQAUUUUDCiiigAooooAKKKKACiiigBaKKKYBRSUtAgooooGFFFFAgooopAFFFFUMp0lGeaK5RBRS0lABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFBIHWrtnpN7en93HsU/wAT0AUqNwHUiuptvCkSgG6mYn/Z6VfTRdKiGGCH/epXA4bcv94Ubl9RXeDT9IH8EFH2HSP7kFFwOE3L/eFJuX+8K7z7DpH9yCj7BpH9yCi4HB7l/vCjcv8AeFd59h0j+5BR9h0j+5BRcDg9y/3hRuX+8K7z7BpH9yCk+w6R/cgouBwm5fUUu5fUV3X2HSP7kFH2LSP7kFFwOE3L6il3L6iu6+xaR/cgpPsOkf3IKLgcKWX1FG5fUV3f2HSP7kFJ9h0n+5BRcDhty+opNy+oruvsOkf3IKPsOkf3IKLgcNuX1FJuX1Fd19h0j+5BR9i0n+5BRcDhty/3hRuX+8K7n7DpP9yCj7DpP9yCi4HDbh6ijcvqK7j7FpP9yCl+xaT/AHIKLgcNuX+8KNy+oruPsWk/3IKPsWk/3IKdwOHyvqKMr6iu4+xaT/cgo+xaT/cgouBw+4eoo3L6iu3+xaV/cgo+xaT/AHIKLgcRuX1FG5fUV232LSv7kNH2LSv7kNFwOK3L6ijcvqK7X7Hpf9yGk+x6V/choA4vcvqKNy+ortDZ6V/chpPsel/3IaLgcZuX1FG5fUV2f2PS/wC5DSfY9L/uQ0XA43K+oo3D1Fdl9k0z+5DR9k0z+5DRcDjdy+oo3D1Fdl9k0z+5DSfZNM/uQ0XA47cvqKNw9RXY/ZNM/uQ0fZNM/uQ0XA47cPUUbh6iuw+yaZ/cho+yab/chouByG5fWk3D1Fdf9l03+5DR9l03+5DRcRyG5fUUbh6iuu+yab/chpPsmm/3IqLgcluHqKNw9RXW/ZdO/uRUn2XTv7kVO4HJ7h6il3D1FdX9l07+7FR9l07+7FSuM5TcPUUZHqK6r7Np392Kk+zaf/dip3A5bcPUUbh6iup+zaf/AHYqPs2n/wB2Ki4HLbh6ijI9RXUG20/+5FSfZtP7JFRcDmMj1FG4etdP9n0/+5FSfZ9P/uxUXA5nI9aNw9a6b7PYf3YqPs9h/diouBzO4etGR610v2ew/uxUfZ7H+7HRcDmtw9RRuHqK6U29j/djpPs9h/djouBzeR60ZHrXSeRY/wB2Ok8iy/ux0XA5zI9aMj1ro/Isj/DHTGsrN+gUfSi4GBRWxLo6MMxOc+9UJ7G4g+8u4eq07gV6KQUtACUUUUwClpKKAFopKWgBKKKKAFooooAKKBRQAUUUUAFFFFABRS0lABRRToopJ22Qxs59hQA0UEgdTitu08M3UwDXDqinsOtasXhywhH712Y/7VK4HHbl/vCjcvrXcDTNJT+GL8aX7BpX9yGlcDhty+oo3L6iu6+w6T/cgpPsOk/3IKLgcNuX1FG5fUV3P2DSf7kNH2DSf7kFO4HDbl9RRuX1Fdz9g0n+5DR9h0n+5DRcDh9y+oo3L6iu4+w6T/cgo+w6T/chouBw+5fUUm5fUV3H2HSf7kNH2HSf7kNK4HD7l9RS7l9RXb/YtJ/uQ0fYdJ/uQUXA4fcvqKXcvqK7f7DpP9yGj7DpP9yGi4HEbl9RRuX1Fdv9i0r+5DR9h0r+5DTuBxG5fUUbl9RXb/YtK/uQ0n2LSv7kNFwOJ3L6ijcvqK7b7FpX9yGj7DpX92Gi4HEhl9RS7h/eFdr9i0r+7DSfYtK/uQ0XA4vcvqKTcvrXa/Y9L/uQ0n2LS/7kNFwOM3L6ijcvqK7P7Fpf9yGj7Fpf9yGi4HF7l9RS7l/vCuy+x6X/AHIaPsel/wByGi4HGbl9RRuX1Fdl9j0v+5DR9j0v+5DTuBxu4eoo3D1Fdl9j0z+5DR9j0z+5DSuBxu4eopdw9RXY/Y9M/uQ0n2PTP7kNFwOP3D1FG5fUV2H2TTP7kNH2PTP7kNO4HH7l9RRuHqK6/wCx6Z/choNppn9yGi4HIZHrRuHqK6/7Jpv9yGj7Hpv92Gi4HIbh6ijcvqK677Jpv92Gj7Jpv92Gi4HI7h6ijcPUV132XTf7kVIbXTf7sVFwOS3L60bh6iut+y6b/dipPsunf3YqLgcpuHqKMj1FdX9l07+7FSfZdO/uxUXA5XcPUUZHrXVfZdO/uxUn2bT/AO7FRcDldw9RS5HrXU/ZtP8A7sVH2XT/AO7FRcDlsj1FJuHrXU/ZtP8A7sVH2bT/AO7FRcDltw9RS7h6iun+zaf/AHYqT7NYf3YqLgczuHqKMj1rpvs9h/djo+z2H92Ki4HM7h6ijI9a6X7PYf3Y6Ps9j/djp3A5rI9aMj1rpfs9h/djpPs9h/djouBzeR60ZHrXSfZ7D+7HSfZ7H+7HRcDnNw9aMj1ro/s9j/djpPs9j/djouBzuR60ZHrXRfZ7L+7HR5Fl/djouBz2R60ZHrXQeTZf3Y6PIsz/AAx0XA57I9aWt42do/AC/hUEukRsMxOw+tFwMiirM9jPByV3D/Zqt/OmAtFA6UUAFFFFAC0lFLQAUlLRQAlFLSUwCiiigAooooAKKKKACiiigAooooAWik6naAST0ArSs9EvbvB2iNP9rg0mxmceOtJuX+8K6qDwvbIAZ5XJ/Sri6RpcY5EZ+tLmEcTuX1FG5fUV3H2DSh/BDS/YdJ/uQ0cwHDbl9RRuX1Fdz9g0n+5DSfYdJ/uQ0cwHD7l9RRuX1Fdx9h0n+5DR9h0n+5DRzAcPuX1FG5fUV3H2HSf7kNJ9h0n+5DRzAcRuX1FG5f7wrt/sOlf3IaPsOk/3IaOYDiNy+oo3D1Fdv9h0n+5DR9h0r+5DRzAcRuHqKNy+ort/sOlf3YaT7DpX92GjmA4jcvqKXcvqK7Y2Ok/3IaPsOk/3YaOYDidy+ooDL612v2HSv7sNL9h0r+7DT5gOJ3L6il3L6iu0+xaV/dho+w6V/dhpcwHFbl9RRuX1Fdr9i0v+5DR9i0r+5DRzAcVuX1o3L6iu1+xaX/dho+xaX/dho5gOL3L6ik3L6iu0+xaX/dho+xaX/dhouBxm4etG4eors/sel/3IaT7Hpf8Adhp8wHG7h60bl9a7L7Hpn92Gk+x6Z/dho5gOO3L60bl9a7H7Hpn9yGk+x6Z/cho5gOP3D1FG4etdh9k0z+7DR9k03+7FRcDj9w9RS7hjqK677Jpn92Kj7Jpv92KjmGchuHqKXcPUV1v2TTf7kVH2TTf7kVFwOR3D1FLuHqK6z7Lpv9yKj7Lp39yKjmA5PcPUUbh611f2XTv7sVH2XTv7sVO4jlNy+tG4eorq/sunf3YqPsunf3IqVwOU3D1o3D1FdV9l07+7FR9l07+7FRzAcruHqKNw9RXU/ZdP/uxUfZdP/uxU7gctuHqKNw9a6k2un/3YqT7Np/8AdiouBy+4etGR611H2aw/uxUn2aw/uxUcwHMZHqKMj1rp/s1h/dipPs1h/djouBzOR60uR6iul+z2H92Ok+z2H92OjmA5vI9RRketdJ9nsf7sdJ9nsf7sdFwOcyPWjI9a6P7PY/3Y6Q29j/djouBzuR60ZHrXReRY/wB2Oj7PY/3Y6Lgc5ketLketdD5Fl/djo8iy/ux0XA57I9aMj1rofIsv7sdJ5Fl/djp3A5/I9aMj1roPIsv7sdJ5Fn/djouBgZHrRketb/2ezP8AClNawtXHGB9KLgYfaitSXSB1hc/8CqhNbzQnDoceo6UXAiooopjCiiigAooooAKKKKACiiigBaKKKYgooooAKKKKACiiigAooooAKKKKBlOkpTSVzCCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKkt4JbmURQIWc+naltbeW7uFggGWb9K7SytLbRbPLEb8fM/c0gKmmaBBZqJrzEknUg/dFT3WtQQfurddxHYdKyr/U5rtiqkpF2A71SGB0oAuTapdzfx7B6LVZpJW+/Ix+pplLQAc+po59TRRQAc+po59TRRQAfiaOfU0UUAHPqaOfU0UUAH4mj8TRRQAnfqaX8TRRQAn4ml/E0UUAH4mj8TRRQAnPqaOfU0tJQAc+po/E0UUAHPqaOfU0UUAGPc0n4mlooATn1NH4mlooATHuaPxNLSUwD8TR+JoooAPxNH4miigAx7mjHuaKKADn1NJ+JpaSgA59TR+JpaKAE/E0c+ppaKAE/E0fiaWkoAPxNH4miigA59TR+JopKAF/E0n4miigA/E0fiaKKAD8TSfiaWigBPxNH4miimAn40fjRRQAc+tH40UUAH40fjRRQAfjR+NFFAB+Jo59TRRQAfjR+dFFAB+JpPxNLRQAnPqaUMw6MRSUUATx3c8f8AFu+tXINQST5JRtJ/Ksyj60AaV1p0Vwu+HCt2x0NYssbwyFJFww/WtC2u3gOCdyenpV+eCK/gyMbv4WoA56inzRPDKY5Bhh+tMqgCiiigApaSigAoopaACkpaKACiiigAooooAKKKKACjoOaXtW/oGi+fi7u1/dj7invSbsBW0rQpr7Es+Y4P1NdHix0mLCqqkDt941FqWqrbjyLcAuOOOi1gSs8rl5GLMe5qQNK51uWQkQIAvqetZ8l1cyHLzMfbNR0lAC5Y9WNHPqaSimAvPqaOfU0UUAJz6mjn+8aWkoAOf7xo59TRRQAc+po59TRRQAfiaPxNFFAB+Jo/E0UUAHPqaOfU0UUAHPqaOfU0UUwDn1NJz6miikAc+po/E0UUAHPqaPxNFFMA59TR+JoooAOfU0fiaKKAD8TR+JoooAPxNJ+JpaKAD8TSfiaWigBPxNH4miigA59TRz6miimAc+ppOfU0tFABz6mjn1NFJQAvPqaPxNJRQAfiaPxNFFAB+Jo/E0UlAB+Jo/E0UUAH4mjHuaKKADn1NHPqaKKAD8aPxNFFAB+NFFFACfjR+JoooAKOfWiigA/E0fjRSUwF/Gk/GiigA/E0UUUAH4mk/GlFJQAv40c+tFFAwDMOjkVKl3PH/FuHvUNFAjSh1BH+WUbSfyoutPiuF3w4Vu2OhrOqe3uXgPXK+lIClJE8T7JFwf50yt+WKG+gyMZ7HuKxJonhlMcg5H600wI6WikpgFLSUUALRSUtABRRRQAlFLSUwCiiikMKKKKACiiimIKsWNlPfzeXbrx3Y9BTtN0+TUbgRpxGPvNXXFrXSLMIgAwOAOrGpbGQ2Wk2enR75Nrv3Z/6Uy61xVJS3XcR3bpWVd3k12+ZDhOyjpUFSBZlv7qYndKV9hUBd2PzOxptFMBcn1NAz6mkpaAEyf7xpcn1NFFACc+po59TS0lABz6mjn1NFFABz6mjn1NFFABz6mj8TRRQAfiaOfU0UUwD8TSfiaWkoAX8TSc+ppaKBB+Jo59TRRQAnPqaPxNFFAw59TRz6miloATn1NHPqaWigBPxNH4miigQfiaPxNHaigA/E0c+popKYBz6mj8TRRQMPxNH4miigA/E0fiaKKAD8TRz6miigA59aOfU0UUAH4mjn1NFFABz60n40UtACc+tHPrRRQAn4mjn1paKAD8aOfWiigBPxo/E0tJQAfjR+NFFMA/Oj8aKKAEpfxpKKAFopKKBB+NH40UUAH40UUUxhRRRSAOfU0odx0cikooAsR3kqdTuHvVuK6inG18fQ9KzKTr0osBcu9NDAyW/B/u1lsCGIYYI6itO1vGjIWQ5XsamvbRblPNiwJAO3ei4GLRQQQSGGCOooqgCiiigAooooAKWkpaACikpaYgooooAKKKKACiiigAooooAKKKKBlOkoormEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFLQAlGCSABkngCitjw1Yi6v8AzXGUh5+tIDc0TT49MsjNNjzXGWJ7CsnUb1r2c8kRKeBWlr95jFsh6/ex6Vh9qQCUUtFMAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKSigAooooAKKKKAFpKKKACiiigApKWkoAWikopgFFFFABRRRQAUUUUAFFFFABSUtFACUtJRQAtJRRQAUUUUAFGaKKAEpaO9JQAUUUUAFFJS0AFJRRQAUlLRTAKKKKACkoooAKKWigBKKWkoAKWkooAKKKKACiikoAWkoooAKKKKACprW4NvIO6HqKhooA1NQtluoPMj++oyD6isHnv1rb0yfOYW7ciqOqW/kXO4fdfmmgKVFFFMApaSigApaKKACgdaKKAA9aKKKACiiigAooowSQo6scCgDR0PTjqF4N4/cxnLH3rpdVvxawiCDAcjAA/hFFhAml6SNwAbGWPqawJpWnmaVzksajcBvXJJyT1NJQKKYBRRRQAUCiigApKKKACiiigAooooAKKKKACiiigAooooAOlFFFMApKKKAFpKKKACiiigAooooAKKKKACiiigAooooAKKKSgAooopgFFFFABRRRQAUUlBoAKKKKACiiigAoopKAClpKKACiiigAooooAKKKKACiiigAoopM0ALRSUZpgFBpKKAClpKWgApKWkoAKKKKACiiigAoopKAClpKWgAFFFGaAJracwSA/wAJ6ir19brdW+9PvgZU+tZdX9Nn6wsfpQwMfnkHqKKuapB5NxvUfK9U6YBRRRQAUUUUwFooopAFJS0lABRRQKYwooooAKdFE88yQxjLOcCm10Xhax3F7yQf7K5pNiNW3hg0jTgOMgZY+prAubh7qYyyH6D0q5rN2Z7jylPyJ1+tZ1QMKKSlFMApaKKACiiigQUUmaKACiiimMKKKKBBRRRQMKSlpKAFpKWigBKWiigAooooEFFFFABSdaWigAFFGaSgBaTNFFABRRRQAUUUUxhRRRQIKSjNFAw4ooooAKKKKACiiigAooooAKSiigBaSiigAooooAKKKSmAtJmiigAooooAKKKSgQvakoooAKKKKACilpM0AGaKKKYwooooAKKKKACikooAKKKKACrdjc+W/luflPQ+lVKKQFvVbXI+0Rj/AHgKy63bSUXFuUbkgYNY9zCYJ2j7Z4poCKiiimAUUUUAApaKKYBRRRQIKKKKACiiigAooooAKKKKACiiigClRRRXMAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAB4Ga7Xw/biz0cSMMFuT9K4xV3yIn94iu6viLfRtq8fuxSYHN3EpmuZJDzk4qOgdM0UAFFFFABSUtFABRRRQAUlLRQAUUUUAFFFFABRSZooAKKKKYC0lFFABRRRSAKKKKYBRRRSAKKSigAooopgFFFFABRRRQAUZpKKACiiigAooooAKKKKACiiigAooooAKKKSgAopaTNABmiiigApKKWgApM0UUAFJS0UwCijNJQAZpaSloAKKKSgAooooAKKKKBBRSUUDFpKKKACiiigAooooAKKKKACiikpgPhcxzIw9ea0tTjE1lvHUc1lHoa2YP3unkf7OKQHO9RmilI2sV9DikqgCiiigApe9JRQApooooAKKKKACiiigAq/oVr9q1WNWGVXk1Qro/B8QLTzEcg8UmBc8QT4RIFP3utYdXdYk8zUZB2XpVOpAKKKKYBRRRQAUUlFIAoo70UwCiiigAooooAKKKKACiikoAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUlABRRRTAKKKKACiiigAozRSUAGaKKKACiiigApKKKACiiigAooooAKKKKACiikoAWkpaKACiikzTAKM0fWk7UAFFFFABRRS0AFJRRQAUUUUAFFJRQAZooooAKKKKACiiimAUUUUAFPjcxyq47GmUh6UAauox+dZbx1UZFYfat+3Pm2GD6VgHhmHoaSAKKKKYBRRRTGGaWkpaQhKKKKBhRRRTEFFFFAAFLkIOrHAruEC2GiqBwdn61yOmR+bqUC+jZrqfEEmy1SMd2xUSAwCSxLHq3JpKX2ooGJS0UUAFFFJQAUUUUwCiiigAooooAKKKSgQuaSiloGFFJRQAtFFFABRSUtAgo7UZpKBi0UUUCCk+lFFABRRRQAUUUUAFFFFMYUUUlAB9KKKKACiiigAooooAKKKKACikooAKMUUUAFFFFMAooooAKKSigAzRRRQAUUUUCCkzRRQAtFJS0AFFHekoAKKKKBhRRRTAKKKSgBaSiigAoFFFABRRRQAYooozQAUUUlAFmxk8u4APQ0/WYuElHbg1UU4dT6GtW9XzbAn8aXUDBopAeKWqAKKKBQAtJS0UxBRRRQAUUUUAFFFFABRRRQAUUCigAooooApUUUVzAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAE1mM38A9Wrs9eONNQeorjbD/kI2/+9XY+If8AkHx0mBzY6CigdBRQAUUUUAFFFFABRRRQAUUUlAC0UlFABmiiigAooooAKKTNLQAZopKKAFopKWgAopKKACiiimAUUUUAFFFFABRRSUALRSUUAFFFFABRRRQAUUUUAFFFFABRSUtABRSUUAFFFFABRRRQAUUlFABmiiigAoopKYC5pKKKACiiigAooooAKKKKACiiigApKWkoAKKKKACiiigAooooAKKKKYB2pOaKKACiiikAdjWvpnNoR71kVraT/wAezfWgDEnGLmUf7VR1Lc/8fUv+9UVUgCiiigAoopaAAUUUUAFFFFABRRRQAdjXW+EVAsJm9a5Lsa67wn/yDZaTAy75t1/MfeoKlu/+Pyb61FSAKKKKAFooopAFJRRQAUUUZpgFFGaM0AFFFJQAUUUUAFFFFMAooooAKKKKACiiigAooooASloopgFFFFIAooooASiiigAooopgFFFFABRRRQAUlFFABRRRQAUZopKACiiloASiiigAooooAKKKKACiiimAUUUUAFJRRQAZNJS0lAC0lLSUALRSUUAFFFHagAoopKAFpKKKACiiigAooopgFFFFABRRQKAFpKWkNABQaKKANXTDm1x9axZxtuJB71s6V/x7n8ayLr/j7k+tJARUUUUwCiiigYUUUUAFLSUtAhKKKKYBRRRQBpeHl3auvsM1teI2/eovbrWP4b/5C4/3a1vEf/H0n0qHuBk96KO9FAwoozSUAFFFFMAooooAKKKSgAooooAKKKKACiiigAooooAWkoozQAtJRS5oASlpKKACiiigQUUUUAFFFFMAoopM0DFpKKKACiiigAooooAKKSigBaKSigAooooAKKKKACiiimAUUUlAC0lFFABRRRQAUlFFAgooooAKKKKAFpKKKBiUtFFABRRRmmIKSiigYUUUUAFFFFABRR2ooAKKSigBc0lFFABRRRQAVsJ82n8/3TWOa2IP+PD/AICaTA54UtJ3P1paoApaQUtABRRRTEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUDKVFFFcwgooooAKKKKACiiigAooooAKKKKACiiigAooooAnsP8AkI2/+9XZeIf+QfHXG2P/ACEbf/ersfEP/IPjpMDmuwpaOwooAKKKKACiiigApKKKACiiigAooopgFFJmikAUUUUAFFFFABRRRQAUZoopgFFFFABRRRQAc0UUUAJRS0UAJRRRQAUUUUAFFFFABRRRQAUUUUAFJRRQAUUUUAFHeiigAopM0UALmkoooAKKKKACkopaYCZopaKAEoopaAEoooNABRRRQAUUUUAJRRRQAUZoooAM0UYooAKKKKYBQaKSgAoopaAEoo70UAFFLSUAHY1raV/x6n/erJrW0n/j1P1pMDFuf+PqX/eqKpbn/j6l/wB6oqoAooooAKKKKAFoo7UUAFFFFABRRRQAdjXXeEv+QdLXInoa67wl/wAg6WkwMm7/AOPyb61FU13/AMfk31qGpAKKKKAClpKKACiiimAlLRSUAFLSUUAFFHFFABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKKSgAooopgFFFFIAooopgFFJRQAUUUUAFGaDSUAGaKKWgBKKWkoAKKKKACiiigAopKKAFooooAKKSigAzRRRQAUUUUwDNJRRQAUUUUAFFFFABSZoooAKKKKACiiigAooopgFFFFABRSUUALRSUtAC0neiigAoNFJQBq6T/x7n8ayLr/j6k+ta2k/8e5/Gsm6/wCPqT60luBFRRRTAKKKKBhRRRQAtFHekoEFFFFMYUUUUCNTw5/yGB9K1vEn/H0n0rI8Of8AIYX6Vr+I/wDj6T6VD3AyaDRmkoGFFFFMAooooAKKKSgAooooAKKKKACiiigAooooAKKKKACiiigQUUUUAFFFFABRRRQAUGiimMSjNFJQAuaSiigApaSlzQAUUUUAFFFJQAUUUtAhKKWigYlLxSUUAFFFJQAtJRRTAKKKKACkoooEFFFFABRRRQAUUUUAFFFFABRRRQAUlFFABRRRTGFFFFABRRRQAUUUlABmiiigAooooEFFFFAwooooADWxB/x4f8ANYxrYg/48P+AmkwMDufrRSdz9aWqAKKWimAUUUUCCiiigAooooAKKKKACiiigAooooASilooApUUUVzAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAE9h/yEbf/ersvEP/ACDo642w/wCQjb/71dj4h/5B8dJgc32FFIOlLQAUUhNGaADNFFFABRRRQAUUmaKAFpKKWgBKWkooAOKKKKACiiimAlLRRQAUUUlAhaKSkz3NAx3Wius8MeDf7Vt1vNQZkgb7qDgmuk/4QHRD/DN/31U8yCx5filxXpzfD/RT084f8CqCT4daaf8AVyyD6mjmQWPNyKK72T4cr/yxu9v1FU5vh5qK5MN7E3tijmQHHUlb9z4O1yDOIDN/u1k3Gnajak/aLGWMDuRTugK1FM3rnBJB9MU7tTAKWkooAWikpaACkpaSgAooooAKSlooASiikNABRRR70ALRTQ4LYXLH0Aq7b6Vqd1/x76fK4PcCgCpS4roLbwZrlx96LyR/tCtOH4c37f6++iHsBS5kBxmDRg16BF8N4P8AltdFvpVpPhzpQ+/JKfxpcyA81xRjNeoD4e6IOvnH/gVKfh/ohHSYf8Co50B5dikrq/E3gyXSYGu7FjLbqfmT+Ie9clkEZHQ1SdwFzRSUUwFpKKKAFpKKKACiiigAooopgFFFFABRRRQAlFFFIAooopgFFGaKACiiigArW0n/AI9W+tZNa2k/8ezfWkwMW5/4+pf96oqluf8Aj6l/3qiqgCiiloASiilAoAKKKKACiiigAooooAOxrr/Cf/INlrkPWuu8J/8AINlpMDKu/wDj8m+tQ1Ld/wDH5N9aiqQClpKM0AFFJRTAWikpaACiikoAKKKKYBRRRQAUUUUAFFFFABRRRQAUUlFAC1e0zRtQ1Yn7HCfLBwZCOBWe3QD1OK9l8N28dtodrHGoUbMnHepk7AeY6p4b1TS4vOuI/MiHV1HArJHrXuN7Ek1nNHIoZWQ8H6V4jNGIbmaIHhGIFKMrgNoopKsBaKSigBaSiigApKKKYBS0UlABRRRQAZooooAKKKSgBc0lFFABRRRQAUUUlAAaKM0UAFFFFABRRRQAUUUUAFFFJTAM0UUUAFFFFABRSUUAFFFFABRRRQAUUUUAFGaM0UwCiikoAWiiikAlLRRQAUUUlMAooooAKKKKANXSf+Pc1kXX/H3JWvpP/HuayLr/AI+5PrSQEVFFFMAopaSgAooooAWkpaSgAooopgFFFFAGn4c/5DA/3a1/Ef8Ax9J9KyPDv/IXH0rX8R/8fSfSoe4GRRR3opjCiiigQZpKWigBKKKWgYlFFFABRRRQAUUUUhBRRRTAKKKM0AFFFFABRRRQAUUUmaAFpM0UUxhRRRQAlFOpp6ZPSgAzR9KsWNhe6i+2xtnl9WA4FdVp3gCWTa+p3I2nqicEVLkkBxRkUHBJJ9MVat7DULr/AI9rGWQHuBXqmn+GtKsB+5tlc+sgzWrHFHEMRxqg/wBkYqHU7AeUweE9cmxmDyv94Vfj8B6uwy1zCv4V6UCfWlzUe0YHnP8Awr/Uf+f2H8qjk8A6qvK3ULfhXpdFHOwPJ5/CGtw5xF5v+6Kzp9L1K2/4+LCVAO5Fe00140kGHRWHuM01UYHhm5QcEkN6EUpFewX2gaZfIVmtUGe6DBrl9S8AY3PpdxtHXY/OatVEBw1FW7/Tr3TXK3tu8Y7ORwaqdsitE0wCkozRTAKKKSgBaSiigQUUUUAFFFFABRRRQAUUUUAFFJRQAUUUUDCiijNMAooooASlopKAFpKKKACiiigQUUUUwCiiigYUUlFIAooooADWxB/x4f8AATWP2rYg/wCPD/gJpMDA7n60Udz9aBVALRRRTEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAJRRRQBTooormAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAJ7D/AJCNv/vV2PiH/kHx1x1j/wAhG3/3q7HxD/yDo6TA5rsKKOwooAKKKKACikozQAtJRRQAUUtJQAUUUUAFFFFABRRRTAKKSloAKKKKAEooooAShRuljTsXH86U0Rf8fEX++P50Ae52UaxWUCIAAI14H0qwKitv+PaH/rmv8qj1KR4tPnkjOGVCQaxGWqK8di8W6+oY/akOGI5FWE8ba+nWWJvqtVysLnrXHpRXmEXxA1VP9bDG/wBFq9B8SAMC40+Q+60uVgeg0ySKOQfvI0f/AHhmuVtfH2kzECVXh/3q2LXxBpN3jyL6Nie2aLAJe+H9KvQRNaIM90GK56++HdhJlrCV4W/2zkV2auHGUII9QadwaLsDyXUfBes2JJjT7Uo/uCufmSSB/LuImicdiK97xVO90mxv4ylzbRsD3C8/nVKQrHiGO9Fd7q/w9UbpdIm2d/LfnNcVfWN3p03lX1u8LdiR1qk0wK9JSnim5pgLRSZozQAtFAPGe1W9O0y+1STZYW7yDu+OBRcCpjjNLDHLcSeXbRPK/oBXoOkfD2FMS6tMZm67U4xXX2emWVlGEtraNQO+0Z/OpcgPLtO8E6zekGVRaKf74rptP+HenxYa+leZ/wDZOBXbYo4qeZjM2z0LTLNQsNpHx3ZQTWhHGkYwiKv+6MUkkiRqWkZVUdyay7rxJo9rnzb+IEdhU6sDYorjLr4h6XASIYZJvdazJ/iU7f8AHvp7r/vU+Vgei8UteWP8QdYf/VxxJ9Vqu3jfxA3SaFfotPlYrnrdFePHxfr7yRA3KYLgHA7V6zZSNJYwu5yzKCaTVhkepqH0y7UjIMTfyrwoDaWUdFJAr3bUf+Qddf8AXJv5V4Sf9ZJ/vGrgDFooFFWIKKKKACiiigAooopgFFJRQAtFJS0gCikopgLRSUUABooooEFFFFABRRRQAVraT/x6t/vVk1raT/x7N9aTGYtz/wAfUv8AvVFUtz/x9S/71RVQBRS0UAJS0UUAFFFFABRRRQAUUtJQAdq67wl/yDZa5Guv8J/8g2WpYGRd/wDH5N9aizUt3/x+y/WoaQC0lFFMApetJRQAUUUUAFFFFMAooooAKKKKACiiigAooooAKSlpKACiiiiwCH+D/fFe2aP/AMgm2/3BXiZ/g/3xXtejf8gm2/3BWcwLU/8AqJP9w/yrxG8/5CN1/v17fP8A6iT/AHT/ACrxC+/5CV1/v0QGRUlFJWghaKKSgBaKKKYBRSUUAFFFFABRRSUALRRSUALSUUUAFFFFABRRSUAFFFJQAUZAGTSMdqk+ld34S8H29xZx6hqYLmQbo0HYe9JuwHB+bH/eP5UCWP8AvH8q9oXw/pQ/5c4/++ad/YGlf8+cf/fNR7QdjxYSR/3v0pd8f94/lXtP9gaV/wA+cf8A3zR/YGlf8+cf/fNHtBWPFd8f94/lR5if3v0r2r+wNK/584/++aR/D2lOpU2ceD6Cj2gWPF+1Fdj4y8KRabCL7TgRFnEiHnHvXHZHUdDVp3GFFFJVCFpKKKACiiigAooooAKKKKACikpaYBRSUUgClFFFABRRSUALRSUtACUUUUwCiiigAoxRRQBq6T/x7msi6/4+5K1tK/49zWTdf8fUn1pICKiiimAtJRRQAtFFJQAtFHWigBKKKKYBRRRQM0/Dv/IXH0rW8R/8fSfSsnw7/wAhcfStbxH/AMfSfSoe4jJPWig9aKYBRSUUDCiiigAooooAKKKKQBRRRQIKKKM0wCiikoAWikpaACiikoAWkoopjCiiigAooooAKWgdau6TpV1rF19ns1OB9+QjhRSbsBViilnlWGCNpJW6KK7TQ/A24Lcau2e4iHGPrXSaF4es9GhAjQPMfvSHr+FbFYynfYCC3tILWMR28SRqP7oxUuKJJEiQvIwVV5JPauU1nxxY2ZaKyU3MnTcvQGoSbA6rNUrvV9OtM/abyOMjsTXl2peJdX1Jj5tx5cfYR8GshtznMjvIf9s5rRUwPULnxvo8JISQzf7tUX+Idip+SymP4155tUdFX8qWq9mgO/HxFtv+fCb86lT4h2J+/ZTD8a89BNGaPZoD1C28b6RMQJHMP+9Wxa6xp15j7NeROT2BrxfCnqoP4Uq5Q5jd4z/sHFL2YHuuaK8f0/xJq2nEeTceYndZOa7DSPHVndFYr9DbydN56GocGgOqubeC5jKTxJIpGPmGa4vXfAqPun0h9j9TG3OfpXaxSxzRiSJw6HkMD1p9Sm0B4dcQTWs7QXUTRSr1Vqjr1/XNBs9ZtzHcIFk/hkHUH3ry7WdIu9FujBdKShPySAcGt4zuBQpaTHNLWggooooAKKKKACiiigAopKM0ALRSUUwClpKKQBRSUtMYUUUUAGaSiigAooooAKKKKBBRRRQMKTNFFAhaKSigAopaSgBaKSigAPStiD/jw/4Caxz0rYg/48P+AmkxmB3P1paQ9T9aWqAKKKKYgooooASloooAKKKKACiiigAooooAKKKKACiiigClRS0lcwBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBPY/8hC3/AN6ux8Q/8g+OuOsf+Qjb/wC9XY+If+QfHSYHNdhRR2pKAFopKKACjFFFABRS0lABRRSUALRRSUwFpKKWgAopKKQBRS0UwCiilxQAlGKuabpt5qlx5FjCXYfebstdVB8OrlkBn1BAx7BelK6QHEkURj/SIv8AfH867z/hXC974flQPhwFkVxfD5SD0pcyCx3Nt/x7Q/8AXNf5VDqv/ILuf9w1YjTy40T+6oX8qrar/wAgy4/3DWYzxBTw3++aXNMX+L/fNLmthDs0mTSUZoACAeqg/WkAwcoSn+7xS0UAW7XVNRs2DW95Jx0DNkV0Wn+P9SgIF7Gs6j+4MGuSxS5pWQHrGl+NNJv8LJJ9mkP8L10UciSoHjYMp5BBrwUgHnAz6960NN1rUtLkDWly2O6yHIqXEZ7ZVW+sLW/haK6hWRWGORyK5XRPHtrclYdTX7PJ08w/dNdhHLHNGJInDowyCD1qbWA828Q+BZ7MPc6QTLD1MP8AEK4xgyuUdSrqcFT2r301zHinwlb6xEZ7VVhvFHBAwGqlIR5TSqGeRY40LyMcKo71bh0jUZtTOmpbN9pBw3HAHrXpvhnwnaaLGJZQs14w+ZyOB9KpsDn/AA54DaYJda0SFPKwDgj616Ba2tvaRCK2iSNAMfKMVJ3pk9xFbwtLPIqRqMlielZt3GTUyWWOFC8rhFHUk1xGt/ECCEtDpUfnP083+GuK1HWNR1OQveXTnP8AChwKai2B6TqnjbSbHKxP9qcdVjrk9Q8fapclhZIluh/vDmuTwByAM+tKST1qlFCJ7rUL+7Ytc3kpJ7K2BVQqD975j6tzTqDVWAQADooH0pdxHekopgLmjNJRQA9T+9h/66Cvc9O/5B1v/uCvCV/10P8A10Fe66b/AMg63/3BWcxoNR/5B11/1yb+VeFf8tJP94171cxefbSw5x5iFc+ma4QfDZdzE3w5OelKLsBwOKMV6B/wrZMf8fv6VWufhxcpGWtr9Cw/hK9avnQrHEUVa1DT7vTLk299C0T9s/xCqvSqAKQ0tJQAUUlFABS0UUAFJRRmgQUUUUxhRRRQAUUUUCCiiigYUUUUAHatbSf+PVvrWTWtpP8Ax6n60mBi3X/H1L/vVFUt1/x9S/71RVQBS0lFAC0UCigAooooAKKKWgAoopKADtXX+Ev+QbLXIGuv8J/8g2WkwMi8/wCP2b61DUt3/wAfs31qKkAUUUUAFFFFABRRRQAUYpKKYC0daTNFAC0UlFAC0UlFAC0lFFABS0lFABS0lFAA/wDD/vivatF/5BFr/uCvFG/h/wB8V7Von/IItf8AcFZzAuTf6mT/AHTXiF9/yE7v/fr2+b/Uyf7prxG//wCQnd/79EAIKKKStQFpKKKACjNFFABRRSUAFFFFABS0lFABRRRQAUUUUAFBopKBC0lFFAwoopKAGv8Awf7wr23RABotmB0EYrxNhyn+8K9u0Uf8Se0/65is6g0XaZ50Q/5aJ/30KS4yLeUj+4a8Ouprj+0Lv/Spx+9P8ZqFG4Hufnxf89E/76FHnR/89E/76FeEedcf8/U//fdL59x/z93H/fZqvZge8CRCcB1/OnV4poM9wdfsAbmcgyjIL8Gvau5qGrAY/i0A+Gr3P/POvG0/1a/SvZPFv/ItXv8A1zrxuP8A1a/StaYhaSlpK0AKKKKACiiigAopKKAClpKKACilpKACj6UUUwCiiigAoopKQC0UUUAFFFFMAooooAKKKKANTSv+Pc1k3X/H1JWtpX/Huaybr/j7k+tJARUtJRTAWkoooAWikpaACiiigBKKWkpjCiiigDT8O/8AIXH0rW8R/wDH0n0rJ8O/8hcf7ta3iP8A4+k+lQ9xGTSUtFAxKKKKACiiimAUUUUAFFFJQIWjNJRQAtFJRQAtJRRQAUUUUAFFFFMAooopDCkoozTAM0ZzSVc0jTJ9Yv1s7cEDP7x+yik3YCxoOjXGuXnlQgrAp/eSdsV6tpmm22mWi29rGFVRye5NN0vTrfS7JLW1QKqjk92NXQeK55SbAWsfXfENlo0R85t82PliU8mszxV4sj0wNaWJEl2RyeoT615xNNLcTNNcSGSVjkknNVGFwNLWNf1DWJCZ5THD/CiHH51lgAdBikzRWySQARTcVYtLW5vpRHZwPM/oBXTaf4CvrhQ95cLCv9zHNJySA5A4HU03zEHc/lXqFr4E0eEAyo8j+u7itODw7pUH3LRf+BDNT7RAeOh1/wBr8qfXs02kafNEYntItpGDhcGvMfFWkLouq+VFnyJRujB7CnGdwMiim0tWAGmnnqM0tFAGlo+vaho8u63lLxfxRuc5HtXpOgeI7LWogI2EdwB80THmvI/pToZJIJlmt5DHKhyGHH51EoJge6daqalpttqdo1tdxhlYcHuPpXP+FPFcepqtpfER3ajAJ6PXW1g7pgeOa/oVxoV35coLW7H93J2+lZle1anp1vqdk9rcoGRhwe4PrXkesaXPo9+1pODt6xv6it4TvowKNJS0VoISijNJQAUUUUDCiiigQUUlFABS0lFMYUUUUAFFFFACUtFJQIWikooAWikooGGaKKKBBRRRQAUUUUDCg0UUAFFFFMANbEH/AB4f8BNY56Vrwf8AHj/wE0mBg9z9aWk7n60tMAoopKYhaKKKAEpaKKAEopaKACiiigAooooAKWkooAKKKKAKdJS0lcwBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBPY/8hG3/wB6uw8Qn/iXxVx9h/yEbf8A3q7DxD/yD46TA5r0ooooAKKKKACiiigAoopKYC0UUUAJS0lFABRRRQAUUUUAFLSUUALS4ZiFTlmOBSVNZf8AIRtc9N4pAeu+GNJh0nSYo41G91DO3ck1sVHb/wDHvF/uD+VQarcvaaZcXEYy8aZFZDLfFHFeWjx/qxz+6Trj7tKvj7VjIi+WnzMB92nysD1Cqeq/8gy5/wBw/wAqswuZII3PVlBP5VW1X/kGXP8AuGkB4an8f++adTU6P/vmlrYQUtFFABRRRQAUUUUALSUUlAA2D1Ga1tD8R6hocoMMjSwE/NGxz+VZJpKGgPZ9C16y1u38y1cCQD54ieVrVrwmzuriwukubKQxyoc8dD9a9a8KeIE1/T/NK7J4/lkX1PrWbVhmyIYxMZREgkIwXA5NP6UtYXirxBHoNh5mN1xJ8sS+/rSAl8QeIbLQrfdcMHlI+SIHk15ZrWv6hrc5e5kKRfwxKcYHvVG7uri+unuryQySuc89B9Kiq1EQDgYAx9KcDTaWqAWikozQAUUUUwA0lFFACUUtJQAL/rof+ugr3XTf+Qdb/wC4K8LX/Ww/9dBXuum/8g63/wBwVEhosilqK7lMFpNKoyY0LD8BXmv/AAsHVCzfuVwDjpUJXA9P4pcV5cfiDqmR+6Tk/wB2vSNPna5sLedxhpEDGhqwGX4t0mHVNFnDoDLEpeNu+RXjozjDfeXhvrXvF7zZzf7hrwmf/j6ucf8APU1cBMZSUUVYC0lLSZoAWkoopiCiiigYUUUUAFFFFABRRRQAUUUUAFFFFABWtpP/AB7N9aya1tJ/49W+tIDFuv8Aj7l/3qiqW6/4+5f96oqoBaKSloABRRRQAUUUCgAopaSkAtFFFMBD3rr/AAn/AMg2WuQ7Guu8J/8AINlpMDIu/wDj9m+tRVLd/wDH7N9aipAFFFJQAtJS0lABRRRTAKKKKQBRRRTAKM0UUAFFFFMAooooAKWkopAFFFFACN/D/vivatE/5A9r/uCvFW/h/wB4V7Von/IHtf8AcFZ1ARcm/wBTJ/umvEL/AP5Cd3/v17fN/qZP9014hf8A/ITu/wDfogBBRRScVqAUUUUAFLSUUAFFFFABRRRmgAozRRQAUUUUAFFJRQAUZopKAFzRRSUALRSUtMAPVP8AeFe26L/yB7T/AK5ivEW/g/3hXt2if8ga0/65isagyzc/8e8v+6a8LvRjUbv/AK6mvdZ/9RJ/umvDL3/kI3n/AF1NFMCGilpK1EX9A/5GGw/66ivbe5rxPQP+RhsP+uor2z1rKpuMxvF3/ItXv+5XjUf+rX6V7N4u/wCRavf+udeMp/q1+lOmIdRRRWoBRmkooAKKKKACiiigApaSimAtJRRmgAopKKAFopKWkAlLRRQAUUUUwCiiigAooooGFFFFAjU0r/UGsm6/4+5K1tK/49zWTdf8fclJAR0ZpKKYC0lFFABRRRTAKWkpaAEooooGFFFFAGn4d/5C4/3a1vEf/H0n0rJ8O/8AIXH+7Wt4i/4+k+lQ9xGTSUppKBhRRRTAKKKKBBRRSUALRSUUAFFFFABRRRQAUUUUDCiiimAUUUlAC0lLRQISg0tLjNAxqq8jpHEpaSQ4UDvXq/hbQ49G01VwDcSDdI39K5bwBowurptUuEzHGcRZ/vV6NisKkugDelcr4x8SjS4TZ2bA3cg5I/gHrWt4j1iPRtMe4JBlIxEvqa8hmnluZ3uJ2LSyHJJ7e1EI3ARmZnZ3Ys7HLMeppM02nRpJLKkUSF5HOFUd622AcDlgACWPAA7112geCp7sLcarmKHqIu5ra8KeE4tORbu/USXbDIB6JXW4rKU+wFWysLWwhWK1hVFUdcc/nVqkZgqksQAOpNcxq3jXTrFjHbA3Uo6hO1ZpNgdRTWYDqQPqa8uvvGms3JIhZIIz2xzWJPf307FpbuUk+jYq1TYHs01zBEheWZFQdTury7xnq8Orasn2Y7ooBtDf3qwmkmYYe4mYehemYx0rSMLagLRRRVgLSUtFACUtFLQAqlldXRirqcqw6ivS/B3iUanCLO7IW7jHf+IV5nUsE8trcR3NuxWaM5UjvUzjdAe41ieKNFTWdNZAv+kR/NE3vU3h3WYtZ01LhDiQfLIvoa1a59UwPCpEaOR4pAQ6Haw96aTXYfEHR/s10mp26/JL8sgA6e9cb9K6Yu6ELRSUZqhi0ZpKKACiiimAUUUUAGaM0lFAC0UlLQAUlFFABRRRQIKKKKBhRRRQIKKKKACiiigYUUlFAhaSiigBaKSlpjEPStiD/jx/4CayK14P+PD/AICaTAwe5+tFHc0tUAUUlLQAUUUUCCiiigAooooAKKKSgBaKKKACiiigAooopgU6SlpK5QCiiigAooooAKKKKACiiigAooooAKKKKACiiigCew/5CNv/AL1dh4i/5B8VcfYf8hG3/wB6uw8Rf8g+KkwOaoo5wKKACiiigAooooAKKKTNAC0lFFABRRRTAKKSl7UAFFFFABRRRQAVLaf8hC1/66CoqltP+P8Atf8AfFJge5W3/HtF/uD+VU9f/wCQHef9czVy2/49ov8AcH8qqa9/yBLz/rmayGeJqeG/3qch/fxf74/nUa/xf7xp6f6+L/fH8616CPdrX/j2h/65r/KodV/5Btx/uH+VTWv/AB6w/wDXNf5VDqn/ACDbj/cP8qyGeGD+P/falpF/j/32pa2EFFFFABRRRQAUUnWigBaSiigAoopM0AKOorvfhd/q7r61wQ6iu9+F33Lr61MthnoHavOPih/x9Wlej9q85+J//HzaVEdwOHI5NFKeppK1EFGaKSgBaSiimAUUUUAFFFFABRRRQAq/62H/AK6CvddN/wCQdb/7grwpf9bD/wBdBXuumf8AINt/9wVnMaDUv+Qbdf8AXJv5GvC8nL8/xGvdNS/5Bt1/1yb+Rrwr+J/940QAVifl+te4aL/yBrP/AK5CvDT2+te46J/yBrP/AK5CnMRYvP8Ajzm/3DXhE/8Ax93X/XU17vef8ec3+4a8In/4/Lr/AK6mlAGMoopK0AWikopgLRRRmgAopKWgAopKKAF7UlLSUAFLSUUAFLSUUAFLSUtACVr6T/x7N9aye1aulf8AHs31pMDGuv8Aj6l/3qiqW5/4+pf96oqoApRSUtABRRRQAUUUUALRSUtABRSUUABrr/Cf/INlrkDXXeEv+QdLSYGRdj/TZvrUVS3n/H7N9ahpAFFFFABRRRQAUUUUwCiiigAopKWgAooooAKKKKACikpaACiiigApKWkoARui/wC8K9r0P/kD2v8AuCvFG6L/ALwr2vQ/+QNa/wC4KzqAXJv9S/8AumvENQ/5Cl3/AL9e3zf6mT/dNeH6h/yFLz/fopgQUUZpK1AWigUUAFFFFABRRRQAUUUUAFFFJQAUUUUAFFJRQAUtJRTAKKKWgApKKKQA38H+8K9u0X/kDWn/AFzFeIt/B/vCvbtF/wCQPaf9cxWVQZan/wBRJ/umvDL7/kJXn/XU17nP/qJP9014Xff8hK8/66mimBH2oo7UVsI0NA/5GGw/66ivbPWvE9A/5GKw/wCuor2zuaxqbgjF8Xf8ize/9c68aj/1a/SvZfF3/Is3v/XOvGU/1a/Sqp7AOooorQAooooAKKKKYBxRRSZoAWkoooAKKKKACiiikAUtJS0AFFFFABRSUUwFo7UlFAC0UUUAFGaKKBmppX+oNZN1/wAfcla2lf6g1k3X/H3JSQiKiiimAUUUUDCiiimAUUUUAFFFFABRRRQBp+Hf+QuPpWt4i/4+k+lZPh3/AJC4+laviI/6UlQ9xGSaKKKYwoopKBC0lLSUAFFFFABRRRQAUUUUDCiikzQAtFJRQIWkzRRTGFFFFABS0lGaAHU+GF7meK2j5eZtoqMcnFdP4C08Xesvcuv7uBcr9amTsgPQdJsU0/TobaNdu1Ru+verbEAEk4A70tc9401T+ztEcIcSz/Ih9K592Bwfi7V21bWGCE/Z4DtQf7Q71h+9OAOOep5P1pCK6UrIBCcDOMk8AV6R4K8NCwgGoXyA3Uoyqn+AVzvgbRBqWom9uFzb2/Kg9GNeoAYHt2rKcugC1T1PVLXS7Rri7kCKOg7mk1TUYNLsZLu5YKiDgepryTWdXudavGubkkJn93H2AqYxuBoa94nvdYkKKzQWuflQHBP41iDA6D8e9NzS1ukkApNJRRTAKSlpKACkpaSmAUUUtABRRmkoAWjNJRQBteFdYbR9XQlsW85CyDsPevW0cOiupyrDIPtXhDDcpX1r1HwJqx1HRVilbM0B2kf7PasKkeoG5q1jHqWnTWkmMSLjPpXi08L2tzLbSAho2KgH0r3TtXmXxD077Lq0d8i4S4G049aKctbAcpRRS10AFFFGaACikooAKKWigBKKWkoAKKM0UAFFFFABRRRQAUUUUCCiiigAoopKACiiigAopaKBiUtFJQAUtJS0xCGtiD/jx/4CaxzWxB/x4/8AATSYzBP3jRR/EfrS0xCUtJRTAWiiigYUlLRQIKKKSgBaKSloASloooAKKKKACijFFAyl3opaSuYQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAT2H/IRt/wDerr/EX/IPirkLH/kI2/8AvV1/iL/kHxUmBzQpaBRQAUUUmaACiiimAUUUUAFFFGaACikooAKKKKAClpKUUAFFFFABUtp/x/23++KhqW0/4/7X/fFJge523/HtF/uD+VVNe/5Al3/1zNW7b/j2i/3B/Kqmu/8AIFu/+uZrIZ4iv8X+8acn+ui/3x/Omr/F/vGnR/6+L/fH8616CPdrX/j1h/65r/KotU/5Blx/uGprX/j1h/65r/KodU/5Blx/uGshnhi/xf77UtIv8X++1LWwgopKKACiiigAooozQAUlFFABS0UUAKOoru/hd0uvrXCDqK7v4XfduvrUy2A9C7V5x8UP+Pm0+tejdq85+KH/AB8Wn1qY7jOHPU0UHqaStRBRRRQAUUUUAFFFJQAtFJS0AGaKSloAVf8AWw/9dBXuumf8g23/ANwV4Sv+th/66CvdtM/5Btv/ALgrOYw1L/kG3X/XJv5GvCf43/3jXu+pf8g66/65N/I14R/G/wDvGiACHt9a9y0P/kC2X/XIV4a3b617lof/ACBbL/rkKcxFi9/485v9w14Rcf8AH3df9dTXu95/x5zf7hrwi4/4/Ln/AK6mlAGR0UUVoAUUUUAFLSUUwCiiigAooooAKKKKACiiigAooooAKKKKADtWtpP/AB7N9ayq1dK/492+tJgY1z/x9S/71RVLc/8AH1L/AL1RVQBS0lLQAUUUUAFFFLQAlFFFABRS0lAAa63wn/yDpa5LtXW+Ez/xLpaTAyLv/j8m+tRVLd/8fkv1qKkAUUUZoAKKKKYBRRRQAUUUUAFFFFABRSUtABSUtJQAtFJRQAtJRmigAzRSUUwB+i/7wr2rQv8AkDWv+4K8Ubov+8K9r0L/AJA1r/uCsqg0XZv9S/8AumvD9Q/5Cl3/AL9e4Tf6mT/dNeH6h/yFLv8A36KYEFFFJWohaKSigAopaKACikooAWkoooAKKKMUAJRS4oI9x+dACUUhI9R+dJkeo/OgB1Gabkf3h+dGR/eH50AOzRTdw/vD86AR6j86AHUUmaMigAb+D/eFe36J/wAge0/65ivD2P3P98V7hon/ACB7T/rmKyqDLU/+ok/3TXhd9/yErz/rqa90n/1En+6a8Lv/APkJ3n/XU0UwI6KO1JWwjR8P/wDIxWH/AF1Fe2eteJeH/wDkYrD/AK6ivbe5rGpuMxfF/wDyLN7/ANc68Zj/ANWv0r2bxf8A8ize/wDXOvGY/wDVr9KqmIdRRRWgBRRRQAUlFFMAooooAKKKKACiiigAooooAKWkpaQBSUGimAUUUUALSUUUAFLSUtABRSUtAGppX+oNZN1/x9yVq6X/AKg1lXX/AB9SUkBFRRRTAKKKKBhRRRTAKKKKACiiigAooooA0/Dv/IXH0rV8Rf8AH0n0rK8Pf8hcfStTxF/x9L9Kh7iMo0lKaSmAtJRRQAUUUUAFFFFAwooozQIKKSigYtJRRQIKKWigYlFFFMAopKKACikpaAAttUse1eoeA7H7JoKuw+eVt2favL1QyyRwj/lowFe16bEINOtogMbYwKyqMC3Xl/j+/N1rYtlbMcCg/jXps0ixRPI3RBk14jfzm51G5mY53SHH0qaa1AhzzSqjyukUYy7kBRTa6DwTY/bfEKOVytt8xraTsgPRtB02PTNJgtYx0Xcfqa0GwASeg5p1YPjDUzpuhytG2JpBtSubdgcJ4y1ptW1Q28TZtbc4XHRjWBQBge55P1pa6YqyASloopgFLSZooAWkoopgFFFJQAUUUUAFFFFABSUtIaAAmug8DX5svEQiJwl0NmK56n285tryC5XgxODSkroD3Wud8cWIvPD0zgZeAblrdtJPOtIZeu9Af0ou4VuLSWFhw6kVzLRgeGKcqPpzS06VDFczxHqsjCm11LYAooozTAKKSigAooNFABRRRQAUUUUAFFFFABRRRQAUZopKACiiigAooooEFFFFABS0lFAwpaKSgAooozTEB6VsQf8AHj/wE1j9q2IP+PH/AICaTGYJ+8aKO5opiCiilpjCiiigQUUUUAFFFFABRSUtABRRRQMKKKKBBRSUUDKnekpaSuYQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAT2P/IQt/wDerr/EP/IPirkLH/kIW/8AvV13iH/kHxUmBzfaij0ooAKKKKACijNFMAopM0tABSUUUAFFFFABRRRQAtJRRQAUUUUAFS2n/H/a/wC+KiqW0/4/7X/roKTA9ztv+PaL/cH8qqa7/wAgW7/65mrdt/x7Rf7g/lVTXf8AkC3f/XM1kM8RH8X+8adH/r4v98fzpi9X/wB406P/AF8X++P51r0Ee8Wv/HtD/wBc1/lUOqf8gy4/3DUtr/x6w/8AXNf5VFqv/IMuP9w1kM8MB+//AL7UUi9H/wB80tbCCiijNABQTSUUAFFFFABS0lLQAUUlFACg8iu8+F33br61wa/eFd58Lul19aUtgPQe1ec/FD/j4tPrXo3avOfij/x8Wn1qI7gcOeppKD1NFagFFFFABRRmkoAPpRRRQAUtJRQAtJRRQIVf9bD/ANdBXu2mf8g23/3BXhK/62H/AK6CvdtM/wCQbbf7gqJjQal/yDbr/rk38jXhP8T/AO8a921L/kG3X/XJv5GvCf4n/wB40oDA9vrXuOif8gaz/wCuQrw49vrXuOif8gaz/wCuQpzEWLz/AI85v9w14Pcf8flz/wBdTXvF5/x5zf7hrwi4/wCPu6/66mlACOiiitACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUALWrpP/AB7N9aya1tK/49m+tAGNc/8AH1L/AL1RVLc/8fUv+9UVMBaKKKACiiigAooooAKKKKACg0UUAHY11nhT/kHS1yeetdZ4U/5B8tJgZN3/AMfkv1qKpbv/AI/JfrUVIApKWkpgFLRRQAUUUd6ACiiigApKWigBKKWkoAKKTNFAC0lFFAC0lFFABRRRQAjdF/3hXtehf8ga1/3BXijdF/3hXtehf8ga1/3BWdQaLs3+pf8A3TXh+of8hW8/369wm/1L/wC6a8P1H/kK3n+/RTAr0UtJWogooooAM0UUUAFFFJQAuaKSigBaWkooAlt4Jbm4jt4F3SyHCivQtN+H9gkCnUWaaUjJwcY9q5jwQgfxPCSPudK9brKcncDmP+ED0D/n2b/vqkPgPQP+fZv++q6iio5mM5b/AIQLQf8An3b/AL6pP+EC0H/n3b/vquqoo5mByv8AwgWgj/l3b/vqqt/8PdMkhIsWaCXsxOa7PFJjmnzMDwm/s59PvpbO5XEsZwfeoK6v4kRqmvW7AYLqc+9cnW0XdAKf4P8AfFe46L/yB7T/AK5ivDT1T/fFe46J/wAge0/65is6gFuf/USf7prwu/8A+Qnef9dTXuk/+ok/3TXhd/8A8hO8/wCupopgRdqKKWthF/w//wAjFYf9dRXtvrXiWgf8jFp//XUV7b61jU3AxfF//Is3v/XOvGU/1a/SvZvF/wDyLN7/ANc68ZT/AFa/Sqp7AOoopK0AKKKKACiiimAUUUUAFFFFABRRRQAUUUCgBaKKKQBSUtJTAKKKKACiiigAoopaAEooooA1NL/1BrKuv+PqT61q6V/qDWVc/wDH1J9aSAiooopjCiiigAooopgFFFFABRRRQAUUUUCNPw9/yFh9K1PEX/H0v0rL8Pf8hYfStTxF/wAfSfSoe4GUaSiimMKKKKACiiigAopKKBBRRRQAUUUtACUUUUAFFJS0wCkooxQMKMUUUAJRS0lAi5osXn69YR4480Zr2gLjgdBXkXhJN/ia3/2SDXrx6msKj1GZfiaf7P4evZM4Pl8V40pJXJ7816p48lKeHJF/v8V5Uv3FHtVU9gHg16B8NLbFrcXZH+sO38q88zgE16t4Cj8rwxD6sxNOpsB01ecfEa887UbezB4h+Yj616NmvIfFk/n+JrpvQACs6a1AyDRSZoroAKSlpKACiiimAUUUUgCiiigAooopgFFFFACUUUUAJTJRmJhT6Q9DQwPYvClx9p8P2zk5wNv5VsVy3w8k3+F4wf4ZDXU1yS3A8Y8RwfZvEd3Fjj71Z1b3jpdniyYj+JBWBmumOwC0lGaCasAopKKAFopKWgAooooAKKKKACg0lFAC0lFFABRRRQIKKKKACiiigAooooGFLSUtABSUUUALSUUUxB2rYg/48f8AgJrHNa8H/Hj/AMBNJgYXc0Udz9aBTGLRRRTEFFFFABRRRQAUUUUAFFFFAwooooAKKKKBCUUtFAylRRRXMIKKKKACiiigAooooAKKKKACiiigAooooAKKKKBE9j/yELf/AHq67xD/AMg+KuRsf+Qhb/71dd4g/wCQfFSGc32ooooAKKKSmAppKKKACiiigBaSiigAzRRRQAUUUUAFFFFABSUUUALUtp/x/wBr/wBdBUVS2f8AyELX/rpSA9ztv+PaL/cH8qqa7/yBbv8A65mrdt/x7xf7g/lVTXf+QLd/9czWQzxBf4v9405P9fF/vj+dNX+P/eNOj/18X++P51r0Ee8Wv/HrD/1zX+VQ6p/yDLj/AHDUtr/x6w/9c1/lUWqf8gy4/wBw/wAqyGeFr/H/AL7UuaaP4/8AfalrYQUUUUAFFFKKAEpaSigAooooAKKDSUAKOorvPhb0uq4Qda7v4W/8vVTLYEehdq85+KP/AB8Wn1r0btXnHxR/4+LT61MdxnDnqaKD940laiFozSUtACUUUdqACiiigAopaSgQUUUUDFX/AFsX/XQV7rpf/INtv9wV4Uv+si/66CvddL/5Blt/uCs5ghdR/wCQddf9cm/ka8J/if8A3jXu+pf8g66/65N/I14R/E/+8aIDEPb617lof/IFsv8ArkK8Nbt9a9y0P/kC2X/XIU5iLF5/x6Tf7hrwi4/4/Lr/AK6mvd7z/jzm/wBw14Rcf8fd1/11NKAMjooorQAooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAB2rW0r/AI9m+tZPatXSv+PZvrSYGPc/8fUv+9UVS3P/AB9S/wC9UdUAUUUUAFFLSUAFFFFABS0lFABRQaKADtXWeFP+QfLXJ+tdZ4U/5B8tJgZN3/x+S/Woqlu/+PyX61DQAtJS0UAJS0lFAC0lFFABRRR2oAKKKKACiiigAopKKACiiigAooooAKDSUUwEbov+8K9s0L/kDWv+4K8Tb+H/AHhXtmhf8ga1/wBwVlUAuzf6l/8AdNeH6h/yFbz/AH69wl/1L/7prw/UP+Qref79FMGV6KKK1ABRRRQAUlLSUAFFFFABRRRQAUoptKKAOi8C/wDIzx161XkvgX/kZ469aFYz3Gc345v7vT9Caeym8qQH71ee/wDCT69tGdQ6gdq7v4i/8i4/1ry7+EfQVUErAa3/AAk2vf8AQQ/Sj/hJte/6CH6Vk5oFXyoD0n4e6lfajDetf3HnFGAU46V2Zrgvhf8A8e9//viu8NYS3A8v+Jn/ACG7T/cNchXYfEz/AJDVp/uGuPraOwCHqn++K9y0T/kDWf8A1zFeGt/B/vivctE/5A1n/wBcxUVALc/+ok/3TXhd/wD8hS8/66mvdJ/9RJ/umvCr/wD5Cl7/ANdTRTAjooorURf0D/kYtP8A+uor271rxHQP+Ri0/wD66ivbh1NZVNwRi+L/APkWb3/crxmP/Vr9K9m8X/8AIs3v/XOvGU/1a/SqpgLRRRWgBRRRQAUUUUwCiiigAooooAKKKKAClopKAFooooAKKKKAEpaKSgAooooAKKKKACiiloA09K/1BrKuf+PqStTS/wDUGsu5/wCPqSkgIqKKKYBRRRQAUUUUxhRRRQAUUUUCCiiigDT8Pf8AIXH0rU8Rf8fSVl+Hv+QuPpWp4i/4+k+lS9wMk9aKKKBhRRRQAUlFFAgooooAKKKKACjNFFABRSA0UALSUUUwDNJS0lAC0UUUALSUZooA2/Bo/wCKmi+lesnqa8i8KPs8S2uf4mAr1/1rCpuM5T4gg/8ACP8A/Aq8uA+UfSvWvHMPmeG5yB9wZrydR8i/Srp7ANIypFeqeBLqK48ORIjDfGx3L3FeW4q3pupXuk3BmsJShb7y9jVSjdAe1swVSW4ABJJrxbV51uNYupkOVLEZq9qHizWdQgaB5fKiYYYDqaxAMcVMI23Admlpopa0AKKKKACiiigAooooAKSilpgJRRRmgQtFJmigYGiiimIMUmKWkPAJpDPTPhwMeGgf+mhrrK5vwBH5fhiL3cmukrlluB5R8QD/AMVU/wD1zFc7mt3xy4fxbPj+FBWFXRDYApaSiqAWikpaACiiimIKKSigYppKKKBBRRRQAUUUUAFFFFABRRRQMKKKKACiiigQUUUtAxKKWimAUUUUCENa8H/Hj/wE1kmtaD/jx/4CaTGYXc/WijuaKYhaKKKYCUtFFABRRRQAUUUUAFFFFAwooooAKKKKACiiigClRRRXMIKKKKACiiigAooooAKKKKBBRRRQAUUUUAFFFFAE9j/yELf/AHq67xD/AMg+KuRsf+Qhb/71db4h/wCQfFSYznOwpKPSimAUUUUAFFLSUAFFFFABRSUtABRSUUAFFFFAgooooGFFFFABUtn/AMhC1/66CoqltP8Aj/tf+ugpAe6W/wDx7Rf7g/lVTXf+QJd/9czVu2/49ov9wfyqpr3/ACBLz/rmayGeID+L/eNOj/18X++P50xf4v8AeNPj/wBfF/vj+da9BHvFr/x6w/8AXNf5VDqn/IMuP9w1Na/8esP/AFzX+VQ6p/yDbj/cP8qyGeFr/H/vt/Oihf4/99v50VsIKKKKACiiigAooNFABSUUUAFFFLQADqK7z4W/8vVcGOorvPhb0uqmWwHoXavOPij/AMfNp9a9H7V5x8Uf+Pm0+tTHcZw7dTSUHqaK1EFFFBoAKKKKBBiiiigYUUUUAFFFFACr/rIv+ugr3XSv+QZbf7grwpf9ZF/10Fe66V/yDLb/AHBWcwQ7Uv8AkHXX/XJv5GvCP4n/AN417vqX/IOuv+uTfyNeEH7z/wC8aIAxG6D617lof/IFsv8ArkK8NboPrXuWhf8AIEsv+uIpzAsXn/HnN/uGvB5/+Py6/wCupr3i8/485v8AcNeD3H/H5df9dTSgDGUUUVoAUUlLTAKKKKACiiigAooooEFFFFAwooooAKKKKACiiigArV0r/j2b61lVq6V/x7H60mBj3P8Ax9S/71R1Jc/8fMv+9UdUAUUUUAFFFFAC0UUlIApaSigAooopgHY11nhX/kHS1yddX4V/5B8tJgZV3/x+S/Woqku/+Pyb61FSQBS5pKKYgooooAKKKKBhRRSUALRSUUAFFHeigAooooAKOKM0lABRRRTAKKKKAEbov+8K9s0L/kDWv+4K8Tbov+8K9s0P/kDWv+4KyqAi7L/qn/3TXh2o/wDIWvP+ule4y/6p/wDdNeHaj/yFrz/rpRACvmiiitQAUtJRQAtJS0UAGKSlpKADNFFFACYpaKSgDofAp/4qiKvXBXkPgY/8VTDXrhPNYz3Gct8Rf+Rbf615aD8o+gr1D4in/imn+ory4fdX6CrhsAtFJS1Yj0D4X/8AHvf/AO+K7w1wfwu/497/AP3xXe1hLcZ5h8Tf+Q1af7hrjq7H4mf8hq0/3DXH1rHYBp/g/wB8V7lon/IGtP8ArmK8OPVP98V7lon/ACBrT/rmKioBan/1En+6a8Kv/wDkKXn/AF1Ne6z/AOok/wB014Xf/wDIUvP+upop7iIu1FFFbAX9A/5GLT/+uor271rxHQP+Rh0//rqK9t7msam4IxvF/wDyLN7/ANc68Yj/ANWv0r2fxf8A8ize/wDXOvGE/wBWv0qqewDqKKK0AKKKKACiiigBaSlopgJRS0UAFJS0UAJS0UUAFFFJSAKKWimAlFFFABRRRQAUUUUAFLSUtAGlpf8AqDWXc/8AH1JWppf+oNZdz/x9SUkBFRRRTAKKWigBKKKKYBRRRQAUUUUAFFFFAGn4e/5C4/3a0/EX/H0lZnh7/kLD6Vp+Iv8Aj5Soe4zKJpKWkpgFFFFAgoo70UwCiiikAUUUlAC0maKKAEopaKYBRRSUAFFFFABRRRQAUtJRQMu6RL5GtWMvYSjNe0IwdAw6MM14WWKASDqhyK9p0adbnSbWVTnMYz9cVjVQEfiCH7RoV5EBktGcV4wBtBU9VOK92dQ6FWHDDBrxXVbZrTVrqBhg+YWH0opvoBTxSYpaK2ASilooAKKKKACiijNMAoopKAFpKKKBBRRSUALmkoooGLSUUUCFopKKBi02TIiP4U6p7G3N3qNtaqMmV8UPYD13wzbfZtBtY8dVDfnWoe/0plvH5VvFH/cQL+QqO+nW2spp26IpNcnUDxvxDN9p8RXcuc87aoU6RjJPNIf45CabXUlZAFLSUtUAUUUlAgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAFLSUGmMWigUUgCiiimAHpWtB/x4/8AATWSela0H/Hj/wABNJiMLufrS0dz9aKYCUtFFMAooooAKKKKACiiigYUUUUAFFFFABRRRQAUUYopgUqKKK5RBRRRQAUUUUAFFFFAgooooAKKKKACiiigAooooAnsf+Qhb/71db4h/wCQfFXJWP8AyELf/errfEP/ACD4qTGc32FFFFMAooooAM0UUlIApaKSmAUUUZoAKKKKACikooAWikopALSUZopgFTWn/H/a/wDXQVDUtp/x/wBr/wBdBQwPdbb/AI9ov9wfyqpr3/IEvP8Armat23/HtF/uD+VVNe/5Al5/1zNYjPD1/i/3jTo/9fF/vj+dNH8X+8acn+vi/wB8fzrXoI95tf8Aj1h/65r/ACqHVP8AkG3H+4f5VLa/8esP/XNf5VFqn/INuP8AcP8AKsuozwpf4/8Afb+dLSD+P/fb+dLWogoopKYBRmlpKACiiimAUUUUALSUlFIBw6iu7+Fv/L1XBjqK7z4W/wDL1SlsB6H2rzj4o/8AHxafWvR+1ecfFH/j5tPrUR3GcMepopT1NJWogopKKAFpKKKAFooFFAgooooGFFFFACr/AK2L/roK910v/kGW3+4K8KX/AFsX/XQV7rpf/IMtv9wVEwHal/yDrr/rk38jXhB++/8AvGvd9S/5B11/1yb+Rrwg/ff/AHjSgDEboPrXuOhf8gSy/wCuIrw1jxXuWhf8gSy/64inMCzef8ec3+4a8Hn/AOPy6/66mveL3/jzm/3DXg8//H5df9dTSgMZRRRWggooopgFFFFABRRRQAUUUUAFFFFAgooooAKKKKBhRRmjNABWrpf/AB7n61lVqaX/AMezfWkwMi5/4+Zf96o6kuP+PmX/AHqjqgCiiigAopaKAEopaKQBSUtJTAKKKKACur8K/wDIPlrlPWur8K/8g+WkwMm7/wCPyX61FUt3/wAfkv1qKkIKKKSmMWikooAWkoooAKKKKACiiigAopKKACiiigAooopgHeiikoAKWkzS0ANb+H/eFe26F/yBrX/cFeJN0X/eFe26F/yBbT/cFZVARdm/1Mn+6a8N1H/kK3n+/XuU3+pf/dNeGaj/AMha8/36KYENApKUVqAoFBAHcVo6HpFxrd99mt/lUcvJ/dFdtF8PdPVAJLl3PripckgPOOP7w/Ojj+8Pzr0v/hX2l/8APR6X/hX+l/32pe0QHmfH94fnRx/eH516X/wr/Sv77Uv/AAr/AEv++1HtEFjzPj+8PzpMV6Yfh/pf/PR6wvEfgmTTbVrzT5WmjTl0PYe1CmmBx9FGcjNJVgdB4G/5GmGvWz1NeR+Bv+Rqhr1w9TWM9wOU+Iv/ACLT/UV5cPur9BXqPxF/5Fl/qK8uH3V+gq4bDClFJS1Yj0D4Xf8AHvf/AO+K76uB+Fv/AB73/wDviu+rCW4zzD4mf8hq0/3DXH12PxM/5DVp/uGuPrWOwhrdU/3xXuOif8ga0/65ivDm/g/3xXuOif8AIGtP+uYqKgy3P/qJP9014Vf/APITvf8Arqa91n/1En+6a8Kvv+Qnef8AXU0UwI6KSithGhoH/Ixaf/11Fe3V4joH/Iw6f/11Fe3d6xqbgjF8X/8AIs3v/XOvGE/1afSvZ/F//Is3v/XOvGE/1afSqp7AOooorQAooooAKKKKAFpKWimIKKKKACkpaKACkpaM0DEooooAWkopaACkpaSgAooooAWkoooAKWkpaANLTP8AUGsu5/4+nrU0z/UGsu5/4+npICKilpKYBRRRQAUUUUwCiiigAooooAKKKKANPw9/yFh9K0/EX/HylZnh/wD5Co+laXiL/j5Soe4GUaKDRTAKKKKACiikoAKKKKACikpaACikopiCiiigYUUUUAFFJRQAUtJRQAtFJRQMdgHg9DXo/wAO77z9Ha2dv3kTn8q83BxW/wCCtS+weIVRjiO5AT8aiaugPWK81+I1gYNSi1BV+WYbDj1r0j+VZPifSxqmizQBcyqN0f1rGLswPHu9LQEKko4+ZDtP1FLXSgCiikoADRRSUxBRRRQAUUUnvQAtJS0UDEopaKACkoopgFFJQaAClzSUUCHd8V1Hw+0/7Xrb3TD5LcZU+9coWIHyjLHgD1r13wbpX9laHEjD97L87H69qyqSshm/XNeO74Wnh+WMHD3HyrXS15j8Q9R+1avHZI2Utxu49axgrsDkhwqj2paD1orrAKKKSgQUUUUAFFLSUAFFFFABRRRQAUUUUAFFFFABRRRQMKKKKBBRRRQAUUUUwClpKUUDCiiikAUUUUwA1rQf8eP/AAGsntWtB/x5f8BNJiMPufrRR3P1opgFFFFMAooooAKKKKBhRRRQAUUUUAFFFFABRRRQIKKKKYylRRRXKIKKKKACiiigBaSiigQUUUUAFFFFABRRRQAUUUUAT2P/ACELf/errPEB/wCJfFXJ2P8AyELf/erq/EH/AB4RUmM53sKKOwooAKKKSmAUUtJQAUUUUAFFJS0AFJRRQAUUUUAFFFFABRSUUALUtp/yELX/AHxUVS2n/H/a/wDXQUmB7rbf8e0X+4P5VU17/kCXn/XM1btv+PaL/cH8qqa9/wAgS8/65mshnh6/xf7xpyf6+L/fH86av8X+8aVP9fF/vj+da9BHvVr/AMesP/XNf5VDqn/IMuP9w/yqW1/49Yf+ua/yqLVP+Qbcf7h/lWQzwpf4/wDfaihf4/8AfalrZCEooooAKKKKACiiigBKKDRTAKKSigBw6iu9+FvS6+tcCvUV33ws6XX1qZbAehdq84+KX/HxafWvR+1ecfFL/X2n1qI7jOGPU0lKeppK1EFFFFABRRRQAoooooEFFFFAwooooAVf9bF/10Fe66X/AMgy2/3BXhS/62H/AK6CvddL/wCQZbf7gqJgh2pf8g66/wCuTfyNeEH77/7xr3fUv+Qdc/8AXJv5GvCD/rJP940oDGt0/GvctC/5All/1xFeGt0/GvctC/5All/1xFOYizef8ec3+4a8HuP+Py6/66mveL3/AI85v9w14Pcf8fd1/wBdTSgDGUUUVoAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiimAVqaX/x7N9ay61NM/492+tJgZFx/wAfMn+9UdSXH/HzJ/vVHTAKWkooAKKKKAFooooASiiigAooooAPWur8LH/iXyVynauq8Lf8g+WkwMq7/wCPyX61DUt3/wAfkv1qKkIKKKKYwpKKKAFopKKAFopKKACiiigAoopKAFo5oopgFJRmigAooooAKKKWkA1ui/7wr23Qv+QLaf7grxJ+i/7wr23Qv+QNa/7grOYIuzf6l/8AdNeGal/yFrz/AH69zl/1T/7prwzUv+Qtef8AXSimMr0ucc0lIehrUR6V8MolXR7iXHzNL1rtM1x3w1/5AMv/AF1rscVzy3GGfalrzbxd4k1jT/EBtrO4KRBM4rIHjHX/APn6P5U1BgewfhSZryH/AITHX/8An6P5Uj+MdfVCRdHj2p8jA9gpkqLJGyOMqwIIqvpM0lxpVrNKcu8YZj71ZJqAPCLxBHqF2g4AmbFQ1Y1H/kKXn/XZqr10LYRv+Bf+Rqhr171ryLwN/wAjVBXruOaynuByfxF/5Fp/qK8uH3V+gr1L4jf8i0/1FeWj7q/QVcNhhQKKUVYjv/hb/qNQ/wB8V39cD8Lv9RqH++K76sJbjPMfiZ/yGrT/AHDXHV2HxM/5DVp/uGuPrWOwCHqn+8K9x0T/AJA1p/1zFeHH+D/fFe46J/yBrP8A65ioqAW5/wDUSf7prwq+/wCQnef9dTXutx/x7yf7prwm9/5Cd5/11NFMBlJRR2rYRoaB/wAjDp//AF1Fe3eteIaB/wAjDp//AF1Fe3d6xqbgjG8X/wDIs3v/AFzrxhP9Wn0r2fxf/wAize/9c68YT/Vp9KqmA6ijvRWgBRRRQAUUUUALSUUtMQUlLRQAnNHailoGJRS0lABRRRQAUUUUALSUUUAFFFFABRS0UAJS0UUAaWmf6g1l3P8Ax9SVqaZ/qTWXc/8AH09JARUUtJTAKKKKYBRRRQAUUUUAFFFFABS0lFAGl4f/AOQqPpWn4i/4+lrM0D/kKj6VpeIf+Plal7gZZooNJQAtJmiigAzRRRQAlLRSZoAKM0UUwEpaKKQBRRRmgAopKKYC0lFFABRSUUDCiiloAKAzoyyRnEiHKmiigD2Lw3qaaro8Nwp+YDaw9xWtXlfgjWf7M1QWszYt7g4Gein1r1MHI4rmkrMDzLx1ohsL8X9umLec4YD+FvWuWIr23UrGHUbGW0uBlJBj6V49q+mz6RfvZ3KkYOY27MO1a05X0ApUUE03NagLSE0ZooEGaKKWgYnelopM0ALRSUUAFFFFAgooopjEopfpRikAlJS4qexsp9RvY7K1UmSQ4JH8I9aG7CNrwTop1XVluZF/0a2O7PZj6V6uo2gADAHQVQ0TTIdI02O0hUDaMufVu9aBrlk7sZT1bUI9N02e7kIxGuQPWvFp53uriW4lJLSMWB9q6n4ga2Ly8XTbZ/3UPLkdz6VyNbU421AXNLTRS1qIWikooAKKWkoAKKKKACiiigAooooAKKKKACiiigYUUUUAFFFFAgooooAKKKKYBS0lLSGFFLSUAFFFFMQHpWrB/wAeX/AayjWrB/x5f8BpMZh9z9aWk7n60tUIKKKKACiiigYUUUUAFFFFABRRRQAUUUUAFFHeigQUUlFMZTooorlEFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigCex/wCQhb/71dXr/wDx4R1yll/yEIP96uq8Qf8AHhFSYznR2oo7UUwCiikoAWikooAWkopaAEooooAKKKKACikooAM0UUUAFFFFIBaltP8AkIWv/XQVFUtr/wAf9r/10FAHutt/x7Rf7g/lVTXv+QLd/wDXM1btv+PaL/cH8qqa7/yBbv8A65mshnh6/wAf+8acn+vi/wB8fzpq/wAf+8acn+vi/wB8fzrXoI95tf8Aj1h/65r/ACqHVP8AkGXH+4f5VNa/8e0P/XNf5VDqn/IMuP8AcNZDPCl/j/32paRf4/8AfalNbCCkopKAFpKKDQAUGkpaYCUUUtABRRRQADqK734W9LquDHWu8+Fv/L1Uy2A9E7V5v8Uv9fafWvR+1ecfFL/X2n1qI7jOGPU0lB6mitRBRRRQAtJRRQAUtFJQAtFJRQAtFFJQAq/62H/roK920v8A5Blt/uCvCV/1sP8A10Fe7aX/AMgy2/3BUTBDtR/5B11/1yb+Rrwg/ff/AHjXu2pf8g26/wCuTfyNeE/xv/vGlAY1ug+te5aF/wAgSy/64ivDX6V7loP/ACBLL/riKcxFm8/485v9w14Pcf8AH5df9dTXvF5/x5zf7hrwe4/4/Lr/AK6mlAGR0tFFagFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAK09M/492+tZlaWm/wCob60mBlXH/HzJ/vVHT7j/AI+ZP96mUwCiiloASlopKAFpKWigApKKKACiiigArqvC/wDyD5K5Wuq8L/8AIPkpMDIu/wDj8l+tRVLd/wDH5L9aioAKKKKACiiigAooooAKKKSmAueaSiigApaSigAooooAKKKKACiiigAoopKABjwv+8K9t0L/AJAtp/uCvEW6L/vCvbtC/wCQLa/7grKoCLsv+qf/AHTXhmpf8he8/wCule5y/wCqf/dNeG6l/wAhe8/66UQGVqD0NLSHoa0Eem/DX/kAy/8AXWuxrj/ht/yAZf8ArrXYd6wluM8i8d/8jS3+5XPV0Pjz/kaT/uVztbR2EOBpsv8Aqm/CjNNlP7o02B7joR/4kdn/ANclq6ev4VR0D/kB2f8A1yFXz/SsOozwrUP+Qpef9dmqvVjUP+Qpef8AXZqgrdCN/wADf8jVBXrwryLwMP8AiqYK9eFZT3Gcn8Rv+Raf6ivLR90fQV6n8Rf+Raf615aOi/QVUNgEoFFFaCO/+Fv+o1D/AHxXf1wHwt/1Gof74rvqwluM8x+Jv/Ias/8AcNcdXY/E3/kNWf8AuGuOrWGwAeqf74r3HRP+QNaf9cxXhp6p/vivctE/5A1n/wBcxUVALc/+ok/3TXhN9/yE7z/rqa91uP8Aj3k/3TXhV9/yErz/AK6mimBH2ooorYRoaD/yMOn/APXUV7d614joH/Iw6f8A9dhXt3rWNTcEYvi//kWb3/rnXjCf6tfpXs/i/wD5Fm9/6514wn+rX6VVPYB1FFFaAFFFFABRRRQAUUtJTAKKKKAClFJS0AJRQaKACiiigAooooAKKKKACiiigBaSiigBaKSigDS03/UGsy5/4+nrT03/AFJrMuf+Pp6SAiooopgFFFFMYUUUUAFFFFAgooooAKKKKANLQP8AkKj6VpeIf+PlKzdA/wCQqPpWl4h/4+VqXuBld6KDRQAlLRSUAFFFFMAopKWkAUUUlABRRRTAKKKM0AFFJRQMWkoooAKKKKACiiigApelFJQAvUcHB7H0r03wT4hGpWf2O6YC7hGOf4l7V5jmpbW6nsruO6tWKzRnI96mcboD3PNY/iPQ4NbsTFIAsy8xv3BpvhvXoNcsVkQhZ1GJE759a2M1z7MDw69tLiwu3tbtCkqHv0YVBXr3iPw9ba5bYcbLhfuSDrXlepadd6VdNbX0ZRh0f+FvxreE7gVqMUClrQQlFLSUxhRRRSEFFJRQMWikooAXNFJS0wCloFT2Nnc6hcrbWURkkY4z2H1NJuwEcEEtzOlvbIXmc4AH9a9V8J+G4tEtN8gD3cg+dz29qXwx4Yt9EhEj4lu2HzyHt7CugrnnO4CEVzvjDXl0bTSsbD7VMNsYHUe9bl7M8FpLLFH5jouQnrXi2r391qWpS3F7uEuduw/wilCN2BSJZmZ3OXc7mPvRRS10gFLTadTEFLSUUhhRS0lMAooooEFFFFAwooooAKKKKACiiigAooooEFFFFABRRRTAKKKWgA+tFFFIYUUUUAFFFFMQGtWH/jy/4DWVWpD/AMef/AaTGYnc/WlpO5+tLVCCiiigYUUUZoAKKKKACiiigAooooAKKKKACijvRQIM0UlFMZTooorlEFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigCex/4/7f8A3q6rX/8AjwjrlbL/AI/4P96up1//AI8IqQznu1FJniimAZooooAKKKSgBaKKKACikooEGaKKKBhRRRQAUUUUAFFFFABUtp/yELX/AK6CoqltP+Qha/8AXQUgPdrb/j2i/wBwfyqprv8AyBbv/rmat23/AB7Rf7g/lVXXP+QLd/8AXM1kM8OXq/8AvGnx/wCvh/3x/Omr/H/vGnR/6+H/AHx/Otegj3i1/wCPaH/rmv8AKodU/wCQZcf7hqa1/wCPaL/rmv8AKodU/wCQZcf7hrIZ4UvR/wDfb+dFNU/f/wB9v50E1sIdSZpKKYBmiiigAooooAKWkxS4oAKKBS0AA6iu8+Fv/L1XBjqK7z4W/wDL1Uy2A9D7V5x8Uv8AX2n1r0ftXnPxS/19p9aiO4zhD1NFKeppK1EFFFJQAtFFGaAClpKWgApKWigBDRRiigAU/vYf+ugr3fS/+Qbbf7grwgf62H/roK920z/kGW3+4KiY0O1L/kG3X/XJv5GvCf4n/wB417rqR/4ll1/1yb+Rrwkfef8A3jSgDEf7te5aD/yA7L/riK8NbpXuWg/8gSy/64inMRZvP+POb/cNeD3H/H5df9dTXvF5/wAek3+4a8IuP+Py6/66mlAGR0UUVqAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMArS03/UN9aza0tN/wBQ31pMDKuP+PiT/eqOpLj/AI+JP96o6YBS0lLQAUlLSUALSUUUAFFFFABRRRQAHoa6nwv/AMg+SuWrqfDH/HhLSYGTd/8AH5L9aizUl1/x9y/WoqAFzSUUUAFFFFAC0lFFMAooooAKKKKQBRRRQAUUUUAFFJRQAtJRRTAKKKKAGv0X/eFe36D/AMgW0/3BXiL9F/3hXtug/wDIFtP9wVlUGi9L/qn/AN014bqX/IXvP9+vcpf9U/8AumvDdT/5C97/ANdKKYFfNIehopD0NaiPTvhr/wAgGX/rrXYd64/4a/8AIBl/6612HeueW4zyLx4f+KpP+5XO10Pj3/kaT/uVzwraOwgpsp/dH8KdTJf9U34UwPcfD/8AyA7P/rkK0D/Ss7w//wAgKz/65CtE/wBKwe4zwrUB/wATO8/67NUIqfUP+Qnef9dmqCt1sI3/AAP/AMjTBXrwryHwP/yNMFevCsp7jOU+Iv8AyLUn1FeWD7o+gr1P4i/8i2/1ryz+EfQVdPYAoooqxHf/AAt/1Oof74rvq4D4W/6nUP8AfFd/WEtxnmHxN/5DVp/uGuOrsfib/wAhq0/3DXHVrHYAP8H++K9y0T/kDWf/AFzFeGnqn++K9y0T/kDWf/XMVFQC1cf8e8v+6a8Kvf8AkJXn/XU17rcf8e8n+6a8Kvf+Qjef9dTRTAjope1JWwjQ0D/kYdP/AOuor27ua8R0D/kYdP8A+uor271rGpuCMXxf/wAize/9c68YT/Vr9K9n8X/8ize/9c68ZT/Vp9KqmAtFFFaAFFFFABRRRTAKKKKACiiigApaSigAooooAKKKKACiiigAopaSgAooooAKKKKAFopKKANHTv8AUmsy5/4+XrT07/Ums25/4+XpICKiiimAUUUUwCiiigYUUUUCCiiigYUUUUCNHQP+Qqv0rT8Q/wDHytZmg/8AIVX6VpeID/pK1L3GZfeig9aSgQUUUUwCikpaAEooooAKKKKACikpaBiUUUUAFFFLQAUUlFABRRS0wEooopABooooASkpaKYFjTr+60y8W7s2IcH5l7MK9X8Pa/a65ah4mCzqPnjPXNeQVNZ3VxY3a3NpIY5V7jvWco3A9yAqpqel2mqWxgvIldexI5WsTw34vtdUVYLsrBdjqCflP411APFYO6YHleu+DL7TC01lm5tuv+0PwrmT8rFXBRh/C3Br3qsXVvDGmaoGaWBY5T/y0Uc1pGpbcDx/pQa6zU/AWo2pZ9PkW4jH9881zNxZXtqxW4s5gR3C8VqppgQZpaZuUfeO3/epQyH+NfzqrgLS0ZX+8PzpCyd3X86ACjijIP3ct/u81Yt7C+umC29nMSe5XildAVqActtUF2P8K8mur07wDqVyVa/kW3jP9w812WkeFdL0sKyQiWUf8tHHNQ6iQHEaF4Ov9U2zXWba29f4j+Fej6TpFlpNuIrOFV4+ZsctV0cUuaxlJsBaTIzjPPpWB4g8VWOjoUDCa4I+VFORn3rgB4q1UauNRaUnB5g/h2+lCg2B6/XGeM/Ci36Nf6eoW6QZdR/GK6LRtXttYsUubZgcj5l7qavn2pJuLA8Ew4cxmNvMBwUA5zV230fVLkZhs5Mf7S4r18aNpovGu/scfnt1fFXgMDAxj6Vr7UDxWbQ9WgUtLZuR/sjNUSGVtrqyN/dYYNe9AH2/KsPXvC9hrETExiK47SqOaFV7geRUlXNT0650q8a1vE2sPut2YVUNbJ3AKKSimAUUUUAFFFFABRRRQAUUUUABooooAKKKKACiiigAooFLTASloopAFFFFABRRRQAUUUUwA1qQ/wDHn/wGss9K04f+PP8A4DSYGL3P1oo7n60VQC0UUUCCiiigYUUlLQIKKKKACiiigAooooGFFFFAgooopgUqKKK5QFpKKKBBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAE1l/x/wAH+9XU6/8A8eMdctZf8f8AB/vV1Ovf8eMdJjOe7CijtSUALRRSUwFpM0UUAFFBooAKKKKACiiigAozRSUALRRRQAlLRRmgA6VLZ/8AIQtf+ugqLmpbP/kI2n/XQUgPdrb/AI9ov9wfyqprv/IFu/8ArmauW3/HtF/uD+VU9e/5Al3/ANczWQzxBT9//eNKn/HxD/vj+dMB+/8A7xpN+ySNz0Dj+da9BHvlsf8ARof+ua/ypl/G09jNEn3nUgUllKsljBIhBVo1wR9KnrMZ5AfBPiJC37iE5YnrUL+EfECdbVD9K9mApfxp8wjw6XQNYh+/ZOf90VUktbyI/vLK4H/AK9759aa8SSDEiKw9xT5x2PACdp+dWT/eGKAVPRgfoa9xn0LSrjPnWELZ9RWReeBNGuM+Whg/3KfOI8nwfSiu7vfhvImWsLwv6CQ4rnr/AMLa1YZ8618xR3iGaakgMWinSK0TbZo3iPo4xTeoyORTAKKM0lMBwPIru/hZ1u64IfeFd78LP+XqplsB6J2rzn4o/wCvtPrXooPFedfFE/vrP61EdxnDN940mKU9TSVqISiloIwMnge9ACUUqBpG2wxPKfRBmtew8La3qGDFbeWp/wCegxSuBj80hZR1ZR9TXd2Xw1dsHULxk9RHzW7Z+A9GtsGRDPj+/S5kB5Sp3fcVnP8As81PHaXkpxFZXDf8Ar2i30LSrfHk2EK474q+kaRjEaKo9hU84WPFIvD+sTfcsmH+8Ksr4Q19+lsg+tey0UudjPHk8E+IDJGfIiAVwTz2r1eyiaGyhif7yKAas0hobuBV1P8A5Bl3/wBcm/ka8IB5f/eNe66u6ppN4zEACFv5V4Qh3At6k1UAY5jxXueg/wDIDsv+uIrwtj8te56B/wAgOx/64iiYi3d/8ek3+4a8Huf+Py6/66mveLv/AI9Jv9w14Nc/8ft1/wBdTSgAyiiitQCiiigAooooAKKKKAEpaKKACiiigAooooAKKKKACiiimAVpab/qG+tZtaOnf6g/WkwMu4/4+JP96o6kn/4+JP8AeqOgAooopgLSUUtABSGlpO9ABRRRQAUUUUAHauo8MH/QJa5ftXT+GP8AjwlpMDKuv+PyX61FUt1/x9y/WoqACjNFFABRRRTAKKKKQBRRRQAUUUUAFGaSigAzRRRTAKKKKACiikoAKWkpaAGt0X/eFe3aD/yBLT/cFeIt0X/eFe3aD/yBLT/cFZVAL0v+qf8A3TXhmp/8he9/66V7nL/qn/3TXhmp/wDIXvf+ulEBlYmg9DRQe9aiPTvhr/yAZf8ArrXYd64/4af8gGX/AK612Nc8txnkHjz/AJGk/wC5XPV0Xj0f8VSf9yudxW0dhBTZf9Ufwp9Nl/1Z/CmwPb9A/wCQFZ/9chWgaoaB/wAgOz/65CtA9Kwe4zwrUP8AkKXn/XZqgqfUP+Qnef8AXZqgrdbCN/wP/wAjTBXrwryHwP8A8jVBXrwrKe4zlPiL/wAi2/1ryz+FfoK9S+I3/Itv9a8szwv0FXT2AKKKK0A7/wCFv+p1D/fFd/Xn/wALP9TqH++K9ArnluB5f8Tf+Q1af7hrj67D4m/8hq0/3DXH1rHYAP8AB/vivctE/wCQNZ/9cxXhp/g/3hXuWh/8ga0/65ioqAW7j/j3k/3TXhV8P+Jlef8AXU17rP8A6iT/AHTXhV9/yErz/rqaKYEdFFFbCL+g/wDIw2H/AF1Fe3eteI6B/wAjFYf9dRXt3rWNTcEYvi//AJFi9/6514wn+rT6V7P4v/5Fm9/6514wn+rX6VVMB1FFFaAFFFFABRRRTAKKKKACiiigAFFFFIAooopgFFFLQAlFFLQAlFLRQAlFLRQAUlFFABRRRQBo6d/qTWZc/wDHy9aWnf6k1nXP/Hy9JARUUUUwCiiimAUUUUAFFFFABRRRQAUtJS0DNDQv+Qov0rR8Qf8AHytZuhf8hRfpWl4g/wCPlfpUvcRl96KDSUwCiiigAooooAKKKKACkopaBhRRRQAUUUUAFFJRQAUUUUAFFFFABRRRQAUlLSUALSUUUwFxRSZooAXPIIJDDoR2rqNC8a3mnBYdQzcW/wDfP3lFctRk0nFMR7Tpmt6fqkQe1uFOf4WOG/KtDOOteDRvJDJ5tvI0Ug/iU10ml+N9VscJcgXUf95zzWLp9hnqtRyRJKu2VFdfQiub03xvpN7tV3aGTvuGBW/De2k4BiuYnz6MKizQFOfw5o9wcy6fESe+KoyeCtEc/LbhPpXRj2pfwpXYHLf8ILo/o1TReC9FjOWtw/1ro6KOZgZkHh/SLfBisIQR3xWgkaRrtjVVA7AU+oJb21hB825iTH95gKNWBLijr0rm9S8baRZblWRpZB0CDIrkNU8c6pegpaIttH2dTzVKDYHoepazYaZE0l3cIMfwg/N+VcDrvjm7vg0OmAwQn/lr/Ea5SZ5LiUy3MrTSH+JjSda1jTS3AUszOXdizt1Y96M0lJitANXw7qd3pmqxNZZbzW2tF2avZIizxIzDaWUEj0rynwJYfbfEQkYfLbDeK9ZrnqbgGKz9Q1zTtOO25uUDf3QckVR8X64dF0rdF/x8TfLH7GvJJWeeVprljJK5yzE0oQ5gPa9P1iw1Ef6Lco7f3c8/lV+vB7aWW0nWe0kaKVTkMDXqvhPxLHrVr5UxCXkYw6f3vcUShYC9r+h22t2RhnGJB/q5B1U15Jqdhc6XetaXabXX7rdmHtXuFY/iPQbfXLIxSALMozHIOoNOE+UDxyipr+yudNvHtLxCsing9mHtUFdCdwFooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUUwCloooAKKKKQBRRRQAUUUUAFFFFMAPStOH/j0/wCA1mGtKH/jz/4DSYGP3P1opO5+tLVAFJS0lAC0UUUCCiiigAooooAKKKKAEpaKKACiiigAopD1opjKdFFFcogooooAKKKKACiiigQUUUUAFFFFABRRRQAUUUUATWX/AB/Qf71dRr3/AB4x1y9n/wAf0H+9XT69/wAeEdIZz/pRmj0opgGaKKKACiiigAooooAKM0lLQAlLSUUALSUUUAFFFFABRRRQAVNZ/wDIRtP+ugqGpbM/8TG0/wCugpMD3e3/AOPeL/cH8qp6+f8AiSXn/XM1bt/+PaL/AHB/KqWv/wDIDvP+uZrMZ4cD97/eNB5GDSL/ABf7xpa1EdN4d8aXmjW4tZ4xPbj7pJ5FdAvxLsv4reT8q85pcmlygelL8StLP3oZh+FWI/iLoj/eE6/8Bry3J9aKXKgPX4PG2hzfdnZf94YrTg1zS5wDHfQ5PYuK8M2Keq5oCIOQuD9aXIFz3+O4hl/1U0b/AO6wNS/hXgsN7eW5zb3ksf0Natp4t1y1IP2ppwOzmlyjPZaK84sviRMpA1CzUDuY+a6bTvGWi35AW48ontJxSswNG+0bTtQBF3aRyE/xEc1yGrfDqNgZNLuGV+0b/dru4Z4p13QyJIPVTmnE9qLsDw3UtG1PS5Cl5auQP40GVqgrBh8pB+le+zRRzxmOZFkQ9VI4rkNd8BWV7um04/ZpuoQfdNWpCPMh1Fd58LP+XquP1HTL3Srkw38JQg8OB8p/Guw+FnW6py2A9Dxha85+KZ/f2n1r0XPFecfFQgT2n1qI7jOIzzQWC9Tj0FXNK0fUNZuBFYQkqTzIw+X869G0HwJYaeFlvv8ASp+6t90H2q3KwjgNM0LVNVcLaWzKp/jkGBXZaV8OoECyancM7941+7XdRRpFGEjUIg6KB0p9Q5MZn2Ojadp4H2SzijI7gc1fpk08UC7ppEjHqxxWDqXjPRtPJVpjKw/5580tWB0HApcn0rze++I9y5IsLRNvYv1rn7vxXrt0Tm7aEHshp8rYHscl1BF/rJ40/wB44qnPr2l24/eX0P0DV4pNd3k5/wBIvJZM+pqv5aH7wyfrVcgj2GbxvocJwZ2b/dGaqP8AETQ1+757f8BryoKo6Ling46UciC56W/xJ0sfdhmP4VE3xJsv4beT8q863GjJp8qA6XxH4yutatzawR+Rbn7zA8n2rl8ADA6CnE5pKaVgGN0r3TQf+QHY/wDXEV4Y44/Gvc9B/wCQHY/9cRUzAtXf/HpN/uGvB7j/AI/br/rqa94u/wDj0m/3DXg9x/x+3X/XU0QGMopKK0ELRRRQAUUUUAFFFFABRRRQAUUUUAFJS0UAFFFFABRRRQAVo6b/AKk/Ws6tHTv9SfrQwMu4/wCPiT/eqOpLj/j4k/3qjoAKKKWmAlFLSUAFFFFABRRRQAUUUUAFdN4Z/wCPCSuZNdN4a/48ZKTAyrr/AI+5frUdSXP/AB9y/Wo6ACiikoAWiiigAooooAKKKSgAooopgFFFFABRmikoAWkoooAKKKKADNFFJQAN0X/eFe3aB/yBLT/cFeIN0X/eFe36B/yBLT/cFZVARel/1T/7prwvU/8AkL3v/XSvdJf9U/8AumvC9T/5C97/ANdKKYytSkcUUE1qI9N+Gn/ICm/6612Vcb8M/wDkBTf9da7KueW4zyPx7/yNJ/3K52uh8enHio5/551zu4e9bReghcU2X/VH8KUMPeklYeU3Wm2B7hoH/IDs/wDrkKvnpWf4f50Kz/65CtBuh+lYPcZ4Vf8A/ITvP+uzVXzU1+f+Jnef9dmqDNbrYR0Hgc/8VVb169nFeP8Agf8A5Gq3r1/GSaynuM5P4jn/AIpp/qK8tB+VfoK9S+JHHhlz6EV5WGG1eD0FXDYB2aM03cPejcM9DV3EegfCz/Vaj/vCvQcV598KzmLUf98V6FWEtxnl/wATP+Q1af7hrj8V2XxMH/E6tP8AcNcdWsNhAf4f94V7hof/ACBrP/rmK8PP8P8AvCvcND/5Atn/ANcxUVBlu4/1En+6a8Lv/wDkJXn/AF1Ne6XH/HvJ/umvCr4/8TK74P8ArTRT3AiooyPQ03cB61rcRo6B/wAjDYf9dRXt/rXh2gsP+Eh0/r/rRXuGeTWVTcDG8X/8ize/9c68ZT/Vp9K9l8Xn/imb3/rnXjMf+rT6VVPYB1FFFaAFFFFABRRRTAKKKKACiiigAooopAFFFFMAooooAKKWigBKKKKACiiigAooooAKKKWgC/p3+qNZtz/x8vWlp/8AqTWbcf8AHy9CAjooooAKKKKYBRRRQAUUUUAFFFFAwpaSloAv6H/yE1+laOv/APHyv0rO0P8A5CY+laGv/wDHytS9xGZQaDSGgBaKSimAUUUUDDNFFGaAClptLQAtFFJQAtJRRnigAooooAKKKKACiiigAopKKAFpKKKADNFFJTAKKKKYBRRRQISilopAIcN94Zp8UssBzBK8R9jTaKLIZpweINZg4TUZSB2Jq0njHXE/5bbvqawqSp5EB0f/AAm+t4/g/Oo38Za4/wDy0CfQ1gUUciA05vEOtT8PqMqg9gaoyzTTnNxO8p/2jUVLTUUAABfujFFFFMBKMUtFMAxQBS0ooEdp8Miovbxf4tg/nXovavHvCmqDStdjkc4imOyQ+gr19GV1DKcqwyD6iuWotRnCfE23mMVpcAFolfDf7NcCev1r3S9tIb61ktrlA8bjBBryTxL4fn0G6PBe0c/JJ/d9jV05dAMapLa5ntLlLm1kMc0ZyCKiNJmttwPXPC/iaDW7YI5Ed2g+eP19xW/mvCLa4ntLlLm1kMcyHIIr1bwt4lh1y22OQl5GPnT19xXPOFgJvEvh+31yzKsAlwgzHIOorya9s7jT7t7W7TZKp/BvpXufWsXxL4dt9ctCGAS4Qfu5B1ohO24HkFGanvLO40+7e1u0KSqcezfSocV0J3AKKKKYBRRRQAUUUUAFFFFABRRRTAKBRRQAtFFFABRRRSAKKKKACiiigAooopgB6VpQ/wDHp/wGs09K0ov+PT/gNJgY3c/Wlo7n60UwCiiimIKKKKACiiigAooooAKKKKACiiigAooopDCijFFMClRRRXMIKKKKACiiigAooooEFFFFABRRRQAUUUUAFFFFAE1n/wAf0H+9XTa7/wAeMdczZ/8AH9B/vV02u/8AHjHSGYHYUUdqKYBRRRSAKKSlpgFJS0lABRRQaACiikoAWkpaKACiikoEGaKKQmgYtS2f/IRtP+ugqHNS2Z/4mNof+mgoYHu9v/x7Rf7gqlr5/wCJFef9c6u2/wDx7Q/7gqtrEElxpF1DEMu6YUVkM8KX+L/eNLWp/wAI3rilh9ib7x7Uf8I3rf8Az5P/AN81pdCMuitX/hG9b/58n/75pR4a1v8A58n/AO+ad0BlUVrf8I1rf/Pk35Uf8I1rf/Pk/wCVHMgMqgVrf8I1rf8Az5P+VH/CNa3/AM+bflSugMqg1q/8I1rf/Pm/5UHw3rf/AD5P+VF0BlBsUxlRuWXJrWPhvW/+fJ/ypP8AhG9b/wCfJ/yougK1lqmo6c4azvZUA/gzwa6/SPiM6FYtYt+Om+Pk1zB8Oa3/AM+T/wDfNMbQtZXrYS/gtJ2YHsmnapZanCJLOdJAR90HkVd614faWWv6dMJrO1uoXBz8oOD9a9D8M+KL27ZLTWNOmhn6Bwvyn61DQHR6hp9rqNuYLyFZEPTI6Vi+GfDR8P31z5Um+2l5XPUe1dNijFK4xm3iub8ReFxr+p28ly+22hwSB1b2rpqCKdwK9nZ29jAILSJYoh/Co61P7UjNtUsegriPEfinVcva6Hps7Ho0rL0+lLcDqtS1iw0uIveXCIQPuZ+Y1w+rfEaSQtHpFuNvTfJwa5SfS9fvpvNu7W5mkPdweKVdA1o8Cxk/75q0kIZe6rqGoMTeXkjg/wAGeBVNQq/dGK0h4d1v/nxf/vmnDw5rf/Pk/wD3zVXQGXnNJWt/wjet/wDPk35Uf8I3rf8Az5P+VO6AyaStf/hG9b/58n/Kk/4RvW/+fJ/yougMmlrV/wCEa1v/AJ8n/Kj/AIRrW/8Anyb8qOZAZVLWp/wjet/8+T/lS/8ACN63/wA+T/lRdAZNLitX/hG9b/58n/KlHhvW/wDnyb8qLoDIYfKPrXuWhf8AIDsv+uIryQ+G9bI/482/KvXtHieHSbSKQYdIgGHoaibBE13/AMek3+4a8HuP+P26/wCupr3i84s5v9w14Pcf8fl0f+mppwGR0tJS1oIKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigArR07/UH61nVo6d/qT9aGBl3H/HxJ/vVHUlx/x8Sf71R0wCilooASilpKACiiigAooooAKKKKAEPSun8Nf8eElcyehrpfDX/HjJSYGXdf8AH3L9aiqW5/4+pfrUVIBaSl7UUAJS0maKYC0lFFABRRQaYBRRRQAlLSUUAFFFFABRmkooAWikooAKKKB0oARui/7wr2/QP+QJaf7grxBui/7wr2/QP+QJaf7grKoBel/1T/7prwvU/wDkL3n/AF0r3Sb/AFL/AO6a8K1I/wDE2vP+ulEAK9IelFFagem/DI/8SGb/AK612gryvwDr8Wl3L2F422CdtyuegavT0uIXUMkyMD3BrCS1GQXGl2FzL5txaRSP03MOai/sPSv+gfB/3zV7zY/76/nR5if31/OpAo/2HpX/AED4P++aDoWlEYOnwf8AfNXvNj/vr+dHnR/89F/OnqAsUaRRrHGoVFGAB2obofpSedF/z0X86xPEniG00vTpWWZHnYFY0U5OfehAeSX3OpXh/wCmzVBTiS7vI33pGLH6mm4roSEb3gY/8VXb17AOprw3SL86Vq9vfAbhE3zD1FeyWGrWWoWyT29xGQwyVzyDWM1qMsXVrBdxGK5iWWM/wt0qp/Yek4/5B0H5Vd8+L/nqn50nnR/89U/OpAp/2HpX/QOg/wC+aBoelf8AQOg/KrvnR/8APVfzpwlj/wCeifnRqBDZ2FpZbvslvHDv+9sHWrVM82P/AJ6L+dQ3V/a2sTST3EaKozyaQHnXxM/5DVp/uGuPrW8T6uNZ1l7mMHyU4jz1xWTW8VZCEP8AD/vV7foRzolmR/zzFeIEblI7kcV6d4I8R21xpkdhcyLFPbjaNxxuFTUQzsSAQQRkGqDaNpjMWaxgLMck7etWxPCRxKn50edF/wA9E/OsgKf9iaX/AM+EH/fNJ/YWlf8AQPg/75q750X/AD0T86Xzo/8Anon509QKSaLpkUiyR2MKupyrBeQavfWk86P/AJ6J+dNeeFVJaVAB3JoAyPGJx4Zvf9yvGY/9Wn0r0Lx74jt5LX+zLGQSux/eMpyMV58BgADoOK1gtBC0tJS1oAUUUUAFFFFMAooooAKKWkoAKKKKQBS0lFMBaKKKACiikoAWikpaAEooooAKKKKACiiigC/Yf6k1nXH/AB8PWjYf6o1nXH/Hw9JAR0UUUwCiiimAUUUUAFFFFABRRRQMKWiigC/of/ITH0rQ1/8A4+U+lZ+h/wDITH0rQ17/AI+V+lT1EZdFFFMAooooGFFJRQAUtJRQAUtFGaACikpaBBRRRQAYoopKAFpKKKBhRRRQAUUUUAFFFFMBKKWigAoopDQIKM0UUAGaKSigAoopaYBRRRSAKKKKACiiigAooopjCiiikAtFFJQIU4Iweldx4K8VGMppepScdIZWP6GuGppGe+McgjtUyjdDPfAwNV76zgv7V7a6jDxuMEHtXE+DPFu/ZpmqSYccRSk9fY13ma52mmB494l8PXGg3R4L2bn5JP7vsaxq9yvbOC+tXtrqMPG4wQRXkviXw9PoN4RgvZufkk/u+xraE+jAx6ltrmezuUubWQxzIcgjvUeKBWlrget+FvEkGuWoVyEu0GHT19xXQ14Ta3M9ncpc2shjmQ5BHf2r1jwv4jh1y0+bCXSDEkfr7iuecLAP8SeHrbXLQqwCXCjMcg6/SvJr20nsLt7W6TbLGcex+le6VyXj3RVvdNN9Cn+kweg6jvRCdnYDzKkpQQRkdKK6QEooooAKKKKACiiigAooopgFLSUUALRRRQAUUUUgCiiigAooooAKKKKYgPStGH/j0/4DWcelaMP/AB6/8BpMZj9z9aKM/MfrS0wCiiimIKSlooAKKKKACiiigAooooAKKSloGFFFFABRRRQBSooormEFFFFABRRRQAUUUUCCiiigAooooAKKKKACiiigCaz/AOP6D/erptd/48Y65mz/AOP6D/erpdc/48Y6QzB7Cik9KKAFpKKKYC0UmaKAFpKPpRQAUUUUAFFJRQIWkoooGFBopKBBRS0UANo3MhV0+8hBFBooA9u8O6lDqmj29xCwOECsPQitSvEtB1690G58y1O6JvvxN0rvrP4haRNGDMsscncEcVm0M7Dn1o59a5j/AITrRP8Anq/5Uv8AwnWif89X/KlZjOn/ABo/GuY/4TrRP+er/lSf8J1on/PR/wAqLMDqPxo/GuX/AOE70T/no/5Un/Cd6J/z0f8AKizA6n8aPxrlv+E80T/no/5Uf8J5of8Az0f8qLMDqfxpOfWuX/4TzQ/+er/lR/wnmh/89H/KizA6g59aTJ9a5f8A4TzQ/wDno/5Uf8J5of8AfeizA6nn1o59a5ceO9EP/LR/yrR0rxJpeqzGG1nHmddrdTRZiNcA9zQVU9QDS0UhhRRRQAUUUUAIRkc0mMDAAAp1FADefWl59aWsnV/EOm6Q6peTgO38APIosBqc560o+tcwfHOh/wDPVqT/AITrQ/8Anq1OzA6n8aPxrlv+E70T/nq9L/wneh/89X/KizA6j8aPxrl/+E70P/nq/wCVH/Cd6J/z1f8AKizA6j8aPxrl/wDhO9E/56v+VH/Cd6J/z0f8qLMDqPxo/GuW/wCE70T/AJ6P+VL/AMJ3on/PR/yoswOo/Gj8a5f/AITvRP8Ano/5Uf8ACd6J/wA9H/KizA6j8aK5j/hOtE/56P8AlVW7+IekxRnyElkkx8oxxmizA2vE+oxabolxLI2GZCqe5rxTJbLt95zuP1rV13Xb3XbkSXZ2xL9yJeg96yz1rSKsISig0VYBRRRQAUUUUAJS0UUAFFFFABRRRQAUUUUCCiiigAooooGFaOn/AOpP1rOrQ0//AFB+tDAzJ/8Aj4k+tR1JP/x8SfWo6YBRRRQAtJRRQAZooooAKKKKACiiigA9a6Xw3/x4yVzVdJ4b/wCPKSkwMy5/4+5frUdSXP8Ax9S/WoqQgooopjCiiigAopKKYC0UUUAJRRRmgAoopKACjNFFABRRRQAUUUUAApaSigAbov8AvCvbtBP/ABJLT/cFeIOeF/3hXtuhH/iSWn+5WcwRfmP7mT/dNeFajzq15/10r3OX/Uyf7prw3UB/xNrzj/lpSgBWpRS7T6UuPatQEwCMEVOl3eRrtjvZlUdADUIB9KXB9KALH2/UP+ghP+dIdQ1D/oIT/nVfn0pDn0pWQFg3+of9BCf86T7ff/8AQQn/ADqDB9KMH0osgJxf3/8Az/z/AJ1EztI++V2d/wC8x5puD6UYPpRZALSGl59KMH0qgEpY5JYf9RM8Wf7powfSjB9KTAl+2X3/AD/z/nS/bL7/AJ/5/wA6h2n0NGD6UrICb7bf/wDP/P8AnSi9v/8An/n/ADqDB9KMH0osgLP2+/8A+ghP+dRy3FxONtxcySj0Y1Fg+lHPpTsgFpKMH0owfSgAo6MGBIYdCO1GD6UnPoaAJ/tt6BgX0w/Gj7dff8/8/wCdQYPpRg+lFkBP9tvv+f8An/OnC+vv+f8An/Oq+D6Uc+lFkBZ+233/AD/z/nTXu7xgQ97Mw7gmocH0owfSiyAQADOO/WkxTsH0pMH0pgJRS4PpRg+lACUUc0UAFFFFABRRRTAWikopAFFFFAC0UlLTASlpKKACg0UUAFFFFABRRRQAUUUUAFFFFAF+w/1RrOuP+Ph60LD/AFRrPuP+Ph6SAjoooqgCiiigAooooAKKKKACilooGHaiiigRf0T/AJCY+lX9e/4+V+lUNE/5CQ+lXtd/4+V+lS9wM0nmkoNFMYUUUUAFFFFABRRRQAUUUUCClpKKAFozSUUAFFJRQAtJRRQMWiiigAopKKAFpKWkzTAKKKKACkpaSgQUtJS0AFFFFABRSUtMAooopAFFFFABSUtGKYwooopAFFFFMAooopAJRS0lAhCOnJBHII7V3/g3xb5mzTdVkxIOIpSevsa4GjHTnBHII7VMo3Ge9DmoL6xt7+0e2uow8bjGCOlcb4M8WeZs03VHxIOI5Sevsa7sGudppgeOeI/D8+g3ZVgXtXP7uT09jWPivcdQsbfUbR7a5QPG4xz2ryPxBodxoV4Y5AWt2P7uTt9DWsJ30YGVViwvZ9OvYru2Yq6MAfcd6gxQVLFEUEs7BQPrWj2A9x0+7S9sYbmP7si5qS6QSW0qEZDIR+lVNBtHstGtbeT7yJzV2UhYnJ7KTXL1A8MuYvs95PB/zzc1HU+pOJdXu5F6M/FV6647AFFFFMAooooAKKKKACiiimAUtFFABRRRQAUUUUAFFFFIAooopgFFFFAgrQi/49f+A1n1oRf8ev8AwGkxmQep+tLSdz9aKYC0UUUwCikpaBBRSUtABRRRQAUlLRQMKKKKACiiigAooooApUUUVzCCiiigAooooEFFFFABRRRQAUUUUAFFFFABRRRQBNaf8fsH+9XS65/x4x1zVn/x+w/71dLrn/HjHSGYFFHajPFABRSUUwFooooAKKKSgBaSg0UAFFFFAhKKWigApKWkoAWkoooAKKKSgYuaU4J5GaSigBdq/wB0UYX+6KKKBCbV/uijav8AdFFFABhf7oo2r/dFFFACbV/uijav90UtFMBu1f7gpdqf3BRRQMTav90Uu1f7oopaQAFX+6Kkgke2nSe2YxzIcqwNR0tFgPVfCnjCDVIltr5lhvFGOTgP9K63NfPwJDBgxVl6MOorqtE8c6hpqLDej7TAP4jy9Q49gPV6K5ay8d6NdAbneEntJxWmniLRpB8uowD6tU2Ga2aKzP7c0r/oI2//AH1UUviTRohzqEJ+jUWA2KK5S88e6NbA7DJMR02c1yWseO9S1BWislFvCf4hw1HK2B2Hinxda6NA0NsyzXjcKoOQp968pup5r24e5vHMsznJLHp7UxizOZJGLyN1ZuppK0UbCG7E/uCk2p/cFPNJVAJtT+4KNqf3BS0UgE2p/dFG1P7gpaKADav9wUbV/uiiimAbU/uijav90UUtACbV/uil2r/dFFFABhf7opRx0GKSigBc0UlFMAooooAKKKSgBaKKKACiiigAooooEFFFFABRRRQAUUUUDCiiimAVoaf/AKk/Ws+r9h/qT9aTAzZ/+PiT61HUk/8Ar5PrUdMBaSlpKACiiloASijtRQAUUUUAFFFFABXSeHP+PKSubrpPDn/HlJSYGXc/8fUn1qOpLn/j6k+tR0CCkoooGFFFFMAooooAWkopKACjNFFABRRRQAUUUdqACjtRRSAKKKKBBRSUUxjXPC/7wr2/QP8AkCWn+4K8QcfKPY5r2vwzNHP4ftHjYEbOcdqzmCNRl3Iy+oxXn118O5p76ecXYAlbIFeiUnFZptDPOf8AhW8//P4KP+FcT/8AP2K9G4o4p8zA85/4VxP/AM/go/4VvP8A8/gr0bijijnYWPOD8N5/+fwfnR/wraf/AJ/B+dej8UcUc7Cx5x/wraf/AJ/BR/wraf8A5/BXo/FHFHOwPOP+FbT/APP4KP8AhW03/P4K9H4o4o52B5z/AMK3n/5+xR/wreb/AJ/BXo3FHFHOwPOv+FcTf8/go/4VxN/z+CvReKOKOdgedf8ACuJ/+fwUf8K4m/5/BXovFHFHOwPOf+Fbzf8AP4KP+Fbz/wDP4K9G4o4o52B5z/wref8A5/BSf8K3n/5+xXo/FHFHOwPOP+FcXH/P2tH/AAri4/5/Fr0fijijnYHnH/CuLj/n8FH/AAre4/5/BXo/FHFHOwPOP+Fb3H/P4KP+Fb3H/P4K9H4o4o52B5x/wre4/wCfsUf8K3uP+fsV6PgUcUc7A84/4VvP/wA/gpf+FcT/APP4K9G4o4o52B5z/wAK4n/5/BR/wrif/n7FejcUcUc7A85/4VxP/wA/a0f8K3n/AOfwV6NxRxRzsDzSf4c3qRk291GzDs3euRurae0uXt7mMxyoeQa95OK8v+JIiGtRFMeaVG/6VcJNsRyNFB60VqAUUUUwFpKKKQBS0lFMBaKSigBaSiigAooooAKKMUUAFLSUUAFFFFABRRS0AXbHiI1n3H/Hw9aFj/qjWfcf8fD0kBHRRRVAFFFFABRRRQAUUUUALRRRQAUUUUAXtE/5CY+lX9d/4+V+lUNE/wCQkPpV/Xf+PlfpU9QM2koopjCiiigAooooEFFFFAC0lFBoAM0UlFAC0lFFAC0lFFABRRRQMWikopiCiiigAooooGFFFFAgpKWkoAKWikoAWkFLRTAKKSloAKKSigBaSlopAFFFFMYCigUUAFFFFABRRRQAUUUUAFGKKKBCjtgkEcgjtXoHg3xZ5uzTdTkxIOI5SfvfWvP6PQgkEcgjtUSjzDPegaq6hYW2o2rW93EHRh3HT6VxvhHxgrqmn6s4WQcJKeh+td0rhlDKQVPQ+tczTTA89vfh3Msx/s66Xyz2l6itTw54Ig024F3fOJ5x91f4RXXcUoIp8zaAWsXxVqiaVok0pI3sNijuc1e1LUrXTLZp7uVUUDgE8t9K8n8Ra5Nr1/5rZS3j4iT296cI3YGQoOPm5JJJNLRRXVYApKKKACiiigAooooAKKKKYC0UUUAFFFFABRRRQAUUUUgCiijNMAooooAO1aEX/Hr/AMBrP5q/F/x6/wDAaTAye5oo7n60tMAooopgFFFFAgooooAKKKKACiiigYUUUUAFFFFABRRRTApUUUVyiCiiigQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBNZ/8AH7D/AL1dJrn/AB4x1zdp/wAfsP8AvV0et/8AHlHSGYVFJRTAWkoooAXNFJRQAUUUUAFFFFAgooooASjNFFABSUuKKBhSUtFABRRSUALRRRQAUUUUCCiiigAooopgFFFFAwooooAKKKKAClpKKAClyR0NJRQAjAN95QfrTfLT+4BT6KQDdi+lHlx/3BTqKLACgL91QPpS5z3pKKYC0UUUAFFFFABRSUtACUtFFABRSUUALRSUtABRSUtMAooooAKKKKACikpaAEoopaACiiigAooooAKKKKACiiigAooooAKKKKACiiigAq/Yf6k/Ws+r9j/qT9aAM6f/AF7/AFqOnz/6+T60ymAUUUUALSUUUAFFFFABRRRQAUUUUAHauk8O/wDHjJXN9jXR+Hf+PKSkwMu5P+lSfWo6kuf+PqX61HQAUUUUAFFJRQAGijNFMAooooAKKKKACiiikAZooopgFFFJQAtFJRQAUtFFABWxoPiO/wBCZhbkSQN1jboPpWPRSauB3I+JEnezpf8AhZD/APPn+lcLk+tJuPrS5EB3X/CyZP8Anz/Sk/4WTL/z5/pXDZPrRk+po5EB3H/CyZf+fP8ASj/hZMv/AD5/pXD5PqaMn1NHIgO4/wCFkzf8+f6Uf8LJm/58/wBK4fJ9TS5PqaORAdx/wsib/nz/AEo/4WRN/wA+f6Vw+T6mjJ9aORAdx/wsib/nz/Sj/hZE3/Pn+lcPk+tGT60ciA7j/hZM3/Pn+lH/AAsmb/nzrh8n1oyfU0ciA7j/AIWRN/z5/pR/wsmb/nzrh8n1oyfWjkQHcf8ACyZv+fP9KP8AhZM3/PnXEZPrSZPrRyIDuP8AhZE3/Pn+lH/CyZf+fP8ASuIyfWjJ9aORAdv/AMLJl/58/wBKP+Fky/8APn+lcPk+tGT60ciA7j/hZEv/AD5/pR/wsmX/AJ8/0rh8n1oyfWjkQHc/8LJl/wCfP9KP+FkS/wDPn+lcNk+poyfWjkQHc/8ACyJf+fOj/hZEv/Pn+lcNk+poyfWjkQHc/wDCyJf+fP8ASj/hZEv/AD5/pXDZPqaMn1o5EB3P/CyJf+fP9KP+FkS/8+f6Vw2T60ZPqaORAdz/AMLIk/58/wBKP+FkS/8APn+lcNk+poyfWjkQHc/8LIl/586P+FkSf8+dcNk+tGT6mjkQHazfEa8dCLe1QN2LVyN3dT3t09zdOXlfqT2qDJ9aKaikAGiiiqAKKKKYBRRRSAKKKKYC0UlFABRRRQAUUUUAFFFFABRRRQAUUUUALRRSUAXrH/VGs+f/AI+Hq/Y/6o1QuP8Aj4ekgI6KKKoAooooAKKKKACiiigApaKKACiiigZe0X/kJD6Ve1z/AI+Vqjo3/IRH0q9rf/Hwv0qXuIzqSjvRTGFFFJQIKKKKACiiigAooooAKKKKBhRRRQIKM0lLQAUUUlMAooooGFFFFABS0lLQAUUlLQAUUUUCCiiigAooopgFFJS0AFFJRQAtFJRQAUtJS0DCiiigAooooAKKKKBBRRRQAUUUUAFFFFAAQCOf/wBVbOleKNW0rCxTefH/AHZTnFY1FS4pjO4i+I7hf39oS3+yKr3fxDvpVK2duiZ7v2rj8n1oNT7OIE97fXeoS+ZezvKeyk8D6VBmkpatKwC0lFFMAooooAKKKKACiiimAUUUUALRSUtABRRRQAUUUUAFFFFABRRRQAUUUUABq/F/x6/hVA1fi/49v+A0mBknqaKO5+tLTAKKSlpgFJS0UCCiiigAooooAKKKSgBaKKKBhRRRQIWiiigZRooormEFFFFABRRRQAUUUUCCiiigAooooAKKKKACiiigCa0/4/Yf96uj1v8A48o65y0/4/Yf96ui1r/jyjpDMPsKSjtRTAKKSlzQAUUUUAFFFFABSZoooEFFFFABRRRQAUUUUAFBoooASlopKBi0UUUCCiiigYUUUUCCiiimMKKKKACiiigAooooAKKKKACiiigAFFFLQAlFFLQAlFFFAC0UUUAFFFJQAtFFFABRRRQAlFFLQAlLRSUAFFFLTAKKKKACiiigBKWikoAWikpaACiiigAooooAKKKKACiiigAooooAKKKKACkopaYCVfsf9SfrVCr9j/qT9aTAzZ/9e/1plPm/17/WmUwCiiigAooooAKKKKACiiigAooooAO1dF4d/wCPKSud7Gui8Pf8eT0mBmXH/H1J9ajp9z/x9SfWo6ACikzRTAKKKKACiiigBaKSigBaKKSkAtFJRmmAtFFJQAtJS0lAC0UUUAFFFFABSUUUAFFFFABRRRTAKKKKACiiikAtFFFMAooooAKKKKQwooooEFFFFABRRRTGFFFFAgooooASloooAKKKKAEpaKKACiiigAooooASilooAKKSloAKKKKACiiimAUUUUgCiiimAUUUUAFFFBoAKKKO9AC0lFFAC0UlLQAUUlLQAlLSUUAXbL/VmqE//Hw9X7PPlGqE/wDr3oQEdFFFMAooooAKKKKACiiloGFJS0UAJS0UUAXdG/5CI+lXtb/4+Vqjo3/IRH0q9rf/AB8LUvcDO70Ud6KYBSUtJQIKKKKACiiimAUUlLSGFJRRQAUCiigAoopaAEooopgFFFFABRRRQIKKKKBi0lFFAC0UlLQIKSlpKYwpaKKBCUUUUDCiiigAooooAKKKWgAooNFABRRRQAUUUUAFFFFAgooooAKKKKACiiigYUUUUAFFFFABRRRQAUUUUAFFFFMAooooAKO9FFAC0UlFAC0lFFABRRRQAUtJRQAtFJS0AFFFFACdq0I/+PX/AIDWeavxf8ev/AaTAyu5paTuaKoBaKKKACikpaACiiigQUUUUAFFFFABRRRSAKKKWmAlFLRTGUaKKK5RBRRRQAUUUUCCiiigAooooAKKKKACiiigAooooAmtP+P2H/erotb/AOPKOuctP+P2H/erotb/AOPKOkxmH2opO1FMApaSigAooooAKKKKACiiigQUUUUAFFFFABRRRQAUUUUDEpaKSgBaKKKACiiigAooopiCiiigYUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUALSUUUALRRRQAUUUlAC0UUUAFJRRQAUtFFACUtJS0wCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooASilopgJV6y/wBUfrVKrtl/qj9aQGdP/r3+tMp83+vf60ymAUUUUAFFFFABRRRQAUUUUAFFFFAAeldD4f8A+PKSuerofD//AB5PSYGZc/8AH1J9aiqS5/4+pPrUdABiiiimAUUUUAFFFFIAooopgFFJRQAUtFFABRRRQAUtFFACUtFFACUUUUAFFFFABRRRQAUUUUAFFFFABS0lLQAUUUUwCiiigAooopDCiiimAUUUUCCiiigYUUUUAFJRS0CCiiigAooooASlpKWgAopKWgAooooAKKKKACiikoAWiiigAoo7UUAFFFFMAooooAKKKKACiiigAooooAKKKKAFooooAKSiigAooooAuWZ/dGqE/wDr2q/Z/wCrNUJ/9e1JAMoooqgCiiigAooooAKWiigYUlLSUAFLRSUCL2j/APIQH0q7rf8Ax8LVLR/+QgPpVzW/+PhanqBQJpKKKYBRRRQAUUlFAxaKKSgBaKQ0tACUUUUAFFFFMAooooAKKKKACiiigAooooAKKKKBC0lFFAxaSiimIWikooGLSUUUAFFFFABRS0lABRRRQAUtFFABRRRQAUUUUAFFFFAgooooAKKKKACiiigAooooGFFFFABRRRQAUUUUAFFFFABRRRTAKKKKBBRR2ooGFLSUUAFFFFAgooooAKKKKBhS0lLQAlFFLQAlX4/+Pb8KoGr8f/Ht+FJgZXc/Wlo7miqAKKKKACiiigAooooEFFFFABRRRQMKKKKAClpKWgQUUUUxlGiiiuUQUUUUCCiiigAooooAKKKKACiiigAooooAKKKKAJbT/j8h/wB6ui1r/jzjrnbT/j8h/wB6uh1r/jzSkxmHRR2pKYC0UUUAFFFFABRRRQAUUUUCCiiigAooooAKKKKACiiigYUUUUAFFFFABRRRQIKKKKYCUtFFAwooooAKKKKACgUUUAFFFFABRRRQAUUtJQAUUUUAFFFFABS0lFABS0lFAC0UUUAFFFFABRRRTASloooAKKKMUAFFFFABRRRQAUUUUAFFFFABRRRQAUUZooAKKKKACiiigAooooAKKKKACiiimAVcs/8AVH61Tq5Z/wCqP1pAZ83+vf60ynTf65/rTaYBRRRQAUUUUAFFFFABRRRQAUUUUAHauh8P/wDHm9c9XQaB/wAeb0mBmXP/AB8yfWo6kuf+PqT61HQAUUUUAFFJRQAUUtJTAKWijFABRRRQAlLRRQAUtJS0AFFFFABRRRQAUlFFABRRRQAUlLSUALRSUUALRRRQAtFFFABRRRTAKKKKACiiigAooooGFFFFAgooooAKKKKAEpaKKACiiigAopKWgAopKKAFopKKAClpKWgBKWiigAooooAKKKKYBRS0lIBaSiimAUUUUAFFFFABRRRQAUUUtACUUUtACUtFJQAUUtJQAUUtJmgC3af6s1Rn/wBe9XrT/VmqM/8Ar3oQDKKSlpgFFGaKACiiigBaKKKACiiigAooooAu6R/yEB9Kua1/x8LVPSP+QgPpVzWv9etT1Azz1pKDRTAKM0UUABoooFAxaSiigAopaKACkpaSgAooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUtACUUtJTEFFLSUDCilooASilpKACiiigApaKKACiiloASiiigAoopKAFoopKAFooooAKKKKACiiigAooooEFFFFAwooooAKKKKACiiigAooopgFFFFABRRRQAUUUUCCiiigAooooGFFFFABRRRQAUtJS0AFXY/+Pf8ACqVXY/8Aj3/CkwMvufrS0dzRVCCiiigAooooASloooAKKKKBiUtJS0AFFFFAgpaSloAKKKKYyjRRRXKIKWkooEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUATWn/H7D/vV0Gtf8eaVz1p/x+Q/71dBrP8Ax5pSGYnaiiimAUUUUAFFFFABRRRQAUUdKO9AgooooAKKKKACiijFABRRRQMKKKKACiiigApKDT7e2ubuURWkDTOey0CGUBgTgHmu00X4eXd0Fl1aXyoz1iH3q2fEXgqyXQz/AGXEVngGV9WpcyGeZ0tHbBGCDg/WiqAKSig0AFFJySqqMszBQPeuvk+H2onT4ri3uF81ly0RHNJuwHI5pKtX+l6jpzEXtnJEB/ERnNU1YN0NFwHUopq/M4RQWduigda6vQvA+o6kVlvh9lt+pVhyw9qG7AcvRiuu8WeDDpEQvNMDSWyj94h5I965HIIyDkGhO4BRRRTASilpCQOScUAFFIXT+9+lJ5if3v0oAdRTPMT+9+lHmJ/e/SgB1FN8xP736UeYn979KAH0U0Mp6Gl70AOop3kXP/PtJ+VOEFz/AM+sn5UXAjxS1J9nuf8An1k/Kj7Pc/8APrJ+VFwIqKl+z3P/AD6yflR9nuf+fWT8qLgR0VJ9nuf+faT8qX7Pc/8APrJ+VFwIaKe8csf+tiZM+oplABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgFXLP/AFR+tU6t2n+qP1pAUJf9c/1plOl/1z/Wm0wCiiigAooooAKKKKACiiigAooopgFdDoH/AB5PXPV0Gg/8eb1LAzLn/j6k+tR1Jc/8fMn1qOhAFJS0UwCiiigAooooAKKKKACiiigAooooAWikooAWikooAKKKKACiiigAooooAKKTegOCTn6UGSP+8fyoAKKbvQnAPP0pJCRGxHUCgBS6jgnml3r/AHv0r0zwz4Y0i80G1uZ4C0siZY1rDwfon/PsfzqHMDx8Mp6GnV6T4n8N6VZaHPcW8BWRRkGvNEOY1J6kU4yuAtFFITVABIFG4etb3gzSrTWdWmt75WaNE3AKcc123/CBaF/zyl/76qXKwHleR60o5r1T/hAdB/55S/8AfVcJ4t0220nXWtLJWWIIGwTk5oUrhYx8UUEhRknip2sb1bT7W1rILb/noRxVXAr0UtFMBKKWigBKSnVe0TS5dY1WKzi4XO529AKTdgM8MpOAeaU169ceDNEuIVje3I2jAKnBrnr/AOHGCW068EY7BxmoVRAcFSZrU1vw9e6GiveSqwY8Ad6ySQBzVp3Admik5BwQQe2RijNO4BRketXdEtor3Wre1uATFIwDAV6UfAeg5OIZf++qhzsB5RketGa9VPgLQf8AnlL/AN9Vw3i/SbbRtbW0sgREY92GPehTuBiUtAU0uDVgJRS4NGDQAlFFLQAneilpKYBRRRSAKKKKYBRRRQAUUUUAFFFFABS0UUAJRRRQAUtJS0AJS0lFABRRRQBbtP8AVmqM/wDr3q7af6s1Sn/17UICOlpKWmAUUUUDCiijFABSikpaBBRRRQAUUUUAXdI/5CA+lXNa/wBetUtI/wCP8fSrms/69al7gZ560UUUwCiiigAooooGFLSUUALRRRQAUUUUAFFFFABRRRQAlFLSUwEJpCwUZY4FBre8EWsF74jMN1GJI/KyFPrSbsBz/mx/3v0o81P736V7R/wjmkHrZp+VNbw1pBH/AB6L+VZ+0A8bDBhlTmnCuh8eWNtp2tRw2cfloY8kD1rnAauLugHUtJRmqAWkoooAWiiimAUUUUAJRmg0xunHUnApAOyCcA80ua9F0vwZpl7oNs13G63DLksDjmsy/wDh7dxZaxukZR0Qjmo9ogOOpat6jpOoaVzf27RoejnoaLDTNQ1FwtnavIP7w6Cq5luBUxmiu40z4eu+2TU7gFTyY14IrH8WeHv7CukaDJtJOhP8J9KlTTdgOfpKcRSVYDTSZpX4Qkda7nwn4W0vVNFW5vEdpS2CQcCplKwHC596WvUj4E0L/nlL/wB9VxnjTSLPRb+KGxVlRlyQxzSVRN2A5/NLmm55qeCzu7mJpbW3aVE+8V7VdwIqM00nDbWBDehFLQAtLTCau2ulaneQCa0spJYj0cdKG7AVc0Zq/wD2Brf/AEDJaX+wda/6BktLmQGfRWj/AGBrX/QMlo/sDW/+gZLRzIDPorR/sDWv+gZLR/YGtf8AQMlo5kBm0taP9ga3/wBAyWj+wNa/6BktHMgM6itH+wNb/wCgZLR/YGt/9AyWjmQGdSZrROga3/0DJaT+wdb/AOgZLRzoDPozWgdA1r/oGS0f2BrX/QMlp8yAz6K0P7A1r/oGy0j6FrEcbSSafIqIMsT2FHMgKFFJnNLVAFFFFABRRRQAUUUUAFFFLQAgpc0UUAIavR/8e/4VRq7H/wAe/wCFAGb3NFHc0UCCiiimAUUUlABS0UUAFFJS0DCiiigAooooAKWkpaBBRSUUxlKiiiuUQUUUUCCiiigAooooAKKKKACiiigAooooAKKKKBktr/x+Q/71dBrP/Hmlc/a/8fkP1rf1j/jzSkBiUtFJTAWiiigAooooAKKKSgBc0UlLQIKKKKACiiigAooooAKKKKBhRRRQAUUUtACV1fw81SOy1k2cyptueEYjlTXKUCR4pElibbIhBU+lJ6oD6CFBGQQeh4rK8NaqmsaLBdofmI2sPQitXNZDPH/G+kf2TrZkjXbb3PKex71zpNeo/EZ9OfRWiuJkW7HMIzzXlaklRnrWkXoIcTSUU+GGW6njtoFLSynAA9PWqEb/AIF0c6trYmkXNtbct/vdq9gA/CsrwzokWiaTHbKAZCMyN/eNa9ZN3GRTQxTLiWJJB6MM15X8RLfTLS+ig0+MJcuMuV6LXper6hDpenTXkzALGDjPc9hXht7eS6jfz305O+Zs4PanEDvfhqulT28g8lDfIed4zkeor0IYxx0rwTTNQn0nUYr62YhkPzgfxL6V7bo+qW+r6fHeWzAq4+Yf3T6USQy7LGksbRyKGRhgg9xXk3jLw0+i3bXVspaylbPH8B9K9bFQ3tpDfWkltcoHikGCDSTsB4LSVreJNDm0HUTA4JtnOYn9vSsnvWqdxBW94KsbfUPEDQXab4wmQKwK6f4dn/iqH/3KUtgO/wD+ES0T/n0FH/CI6J/z6Ct2isrjMA+END/59KT/AIQ/Q/8An0roKSncDA/4Q/Q/+fSkPg/Q/wDn1P510FIaLgeYfEHRdP0m1tXsYTGzkg1xsJ/0u3/66L/OvQfisf8ARLH/AHjXnsH/AB923/XRf51a2A96it4DFGTbxcoP4RUn2aD/AJ4Rf98ilh/1Mf8AuCpKzYEf2eD/AJ4Rf98ij7NB/wA8Iv8AvkVJnFGaAI/s0H/PCL/vkUfZoP8AnhF/3yKlooAi+zQf88Iv++RR9mg/54Rf98ipaKAPP/ijHHHYW3lxonz9VGK8+PWvQ/ip/wAg+2/3xXnh6mtYbCYlFFFUAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgFW7X/Vn61Uq1a/6s/WkBQl/1z/Wm06X/XP9abTAKKKKACiiigAooooAKKKKACiiimAdq6DQf+PN65+t/Qv+PJ6lgZtx/wAfMn1qOpLj/j5k+tR0wCiiigBKWiigBKWkpaACiiigAooooAKBRRQAUUUUAFFFFABRRRQAUUUUALTW4RvoaWmt/q2/3TQB6h4M0jT7zw3by3FsjyEcsRW7/wAI7pP/AD5x/lWf4A/5Fa3+ldLWDeozjPGGh6ZaeH5p4LZUkU8ECvLGybds/wB017H48OPC8/8AvCvHm/492/3TVQ2Ee0eDh/xS9l/1zrbrE8GHPhez9krcqHuMwfGf/Is3X+6a8ZiP7lPpXs3jP/kWbr/dNeMRf6lPpWkBEtIaUUhrQDrPhn/yMNz/ANcq9TAryz4af8jDcf8AXKvVBWM9xhivLvFunXOreOXtLQgSeWvJ7V6lXlHjK+m0zx6L63J3xIpIH8QpRA6nQfA1jp5We9Iubkd/4fyrprizt7i0a1kiTyWGCoHFM0u/i1LT4buBgyyLzjse9W6TbA8T8QaRJomqyWr5MTHdG/Y+1ZteueM9EGr6QxjTNzCN0Z/nXkYzyCMFTtP1raEroAooxU1rZ3d7KIrK3aaQ9h2qmxESJJLKkMCl5ZDhVFeueEPD6aJpw8wA3U3zSt6e1UvCHhFNJUXl9iS8YcZ6J7V11YylcYVS1XUrbSrGS7u3Cog4z3PpVtyVRmCliBnA715j4n03xR4hvyWsnjtIziOPP6mpQGTI+qeNtbZocqoPyAniNfcV2uh+A7CwKzXzfabgc5/h/KuY0Xwt4m0zUoru3BjIIDjsV716om4opYYJHI9KpvsI5jxb4Xg1axMlpGsV3CPk2jAYeleUujxyvFMpSSM4dT2Ne/471wnj7wx58batp8f71B+9QfxD1pxkM43wv/yM1n/vivbe5rxHwqwbxLZkdN4r27+I0p7gGK5nxD4Ottd1Jb2aVlYJtwK6iiouBww+G+n/APPd/wA6d/wrjT/+ez/nXb0VXMwOI/4Vxp3/AD2f86bL8O9OjhdxK+VBI5ruKjuf+PaX/dP8qFJgeDuuyWRB0Ryv5UlPmP8ApNx/11amV0IQUlLSUAFFFFMBaSiigAooooAKKWigApKWigQUUUUDCkpaSgAopaSgAooooAKKKKALVr9w1Sn/ANc1XbX7hqlN/rmoQEdLSUtMAooooAWiiigAooooAKKKKACiiigC5pH/AB/j6Vc1n/XrVPSP+P8AH0q5rP8Ar1qeoGf3ooopgFFFFABS0lFAwoopaBCUUtJQMWikooEBZV+8aN6f3v0rsvh/pVjqVveNeQ+YyOAvsK68eFtG/wCfQVm6lnYZ49vT+9+lG9P736V7F/wi2j/8+gpf+EX0f/n0FL2oHje+P+9+lG+P+9+leyf8Ixo//PotL/wjGj/8+i0e1A8YMkf94/lXSfDtlbxWQp/5Zeleh/8ACMaP/wA+i1PZaJp1jcefa24STGMj0pOpdAX6Q9KfimkVkB5X8THVfEcQJ/5ZDtXKiSP+9+le2ajoenalOJryDzJAMA+1Vv8AhFNF/wCfStYzsgPHfMT+9+lHmR/3v0r2L/hFdF/59KP+EV0X/n0FV7UDx0SR/wB79Kd5kf8Ae/SvYf8AhFtG/wCfQU4eF9G/59BR7UDx3en979KAyk4B/SvY/wDhGNH/AOfRa5/xrounWGhPNa24SQHgihVLsDzw0UDkD6UGtQENWdKtGv8AV7a1UZLMG/Kqxrsfhrp/n6hPfuOIflQ+uaibsgPSI0CRqijAVQMU7FLimyuI4nkboilj+FcwHmvxMvvP1C3sVbKIp3j3rR+G2rLLbSabIFEsfKEdStcRrt7/AGhrd3dZyrv8vNN0fUX0rVoLyM42kK/P8PetuX3QPdazPEGmR6rpM1s65bG5P97tV20uY7u2juIWBjkXcpFOmmigjMk0ioi9STWWwHhckckE0kEoxJEdrU2tvxnLp8muefps6SrIMybf71YRPTnv610p6AI7qEbk9PSvVfh8QfDaEf3ql0vw7pU2l2sklqpZowSfWty0tILKEQ20YRBzgVjOdwJm6V5d8T5FXWoFJOfLB6V6lVC+0mxv5BJdQLIwGMkVEXZ3A8M8xcEg9K9U+HtiLfw6HkT5pWJOe4rM8b+HLUpYR2EPlyTzbGIHau3sbYWtlBboMCNAp+taSndAUdS8P6XqS7bm1Ue6DBrgvFnhW00K2W5t58K5wsbHk16l0ry74kX/ANp1qOyVspbgNx60oN3A5Mn5c+1esfD7/kU7f/eNeTk8GvWPh7n/AIRS3/3m71dTYDpxRRS1zgFFFFMAooooAKKKKACij8qTP0oAWikz9KM/SgBaMUn5UtACYqlrQ/4k93/1yNXqo61zo930/wBUaa3A8Q/if/ep1NB+Z/8Aep1diAKKKKYBRRRQAUUUUCFooooGFFFFAhO1XY/+Pf8ACqVXY/8AUfhSYGb3NFHc0UwCkpaKYBRRRQAUUlFAC0UlLQMKKKKACiiigQUtFFMBKKWigZRooorlEFFFFAgooooAKKKKACiiigAooooAKKKKACiiigZLa/8AH5D9a39Y/wCPNKwLX/j7h+tb2r/8eaUgMWlpKWmAUUUUAFFFFACUUUUCCiiigApaSigBc0lFFAC0UUUDCiiigApaSloAStLQdK/trVBY+aI8/wAVZtdF4C/5GuOk9gN7/hWa/wDP/wDpTT8Mh21Af9816LRWfMxnL+FfC9x4dllC3olgk6x46Vr6taXd7b+VaXQt88MSOTWhijFAHAT/AA3FxJ5k2pPI57uSai/4VjGOl/j8K9DwKMCi7Cx494o8Jnw/bRTLc+e0jbVjA5JrrPAfhQ6dGNS1FP8AS5BlEP8AyzFddPaW9xIjzxK7R/d3DOKnp3YDqaTxRXK+OvEi6Lpxt4GBvLgYUZ6DuakDkviFr/8AaWof2bbP/o8B/eEfxNXJCmA9SzZYnJJPWnAj1FarQQvSt/wd4gfQdTVJCTZTnDr/AHT61ggj1FKQrAg4IND1A+gYZEliWSJgyMMqw7in15x8PfEuxho9/KMf8sHJ/wDHa9HrNjM3XdHtta097W4UZI+Ru6n1rxnU9OuNJv5LK7Uh0+6395fWveK57xf4ej1zTzsUC6jGY29famnYDx2um+HX/I0P/uVzUiSQyvDMMSxttYe9dL8Ov+Rpf/cFXLYR67RRRWQzjdW8f2WmalLZSWcrvH1YHiqf/CzrEsq/YJfmIHWuO8Wj/iqrvjtWOR88fA++P51fLoB9BRSCWJJB0dQw/GnmobIf6Fb/APXJf5VP0qAPPfiv/wAeth/vGvPYP+Pu3/66L/OvQviv/wAelh/vGvPYP+Pq3/66L/OtFsB9AQ/6mP8A3B/Kn1HCf3Mf+4P5VJ2qAOc8a69caBpi3NsgZmYDkVyC/EvUP4rP9K2vip/yAY/98V5kM+vanFXA7YfEq+/58v0q1YfEO7ub6C3aywJXCk+ma4IZq7pBP9s2X/XVarlQHu9FHeiswOB+Kv8AyDrb/roK87PU16J8Vv8AkHW3/XQV52etaw2EFFFFUAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgAq1a/6s/WqtWbb/Vn60gKMv+uf602nS/61/rTaYBRRRQAUUUUAFFFFABRRRQAUUUUAHat7Qv8AjzesGt7Qv+PN6TAzbn/j5k+tMqS5/wCPmT61HTQBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQIKRvuN/umlPSkb/Vt/umkB654A/5Fa3+ldLXNeAP+RXt/pXS1g9yjmvH/wDyK83+8K8gb/j3b/dNeu/EE/8AFLTf74rx9j/o7cj7vrVw2A9p8Ff8ivZ/7lbprA8FH/imLP8A3K381m9wMLxl/wAizd/7hrxiH/UJ9K9o8Yj/AIpm7/3DXjEI/cp9K1gBIOlGKUCitBHVfDXjxFP/ANc69Vryr4b/APIxz/8AXOvVawnuMK8i+IP/ACNsn/XIV67XkXxC/wCRtk/65CiG4Gl8NtYNvdSaTM3yPzDn1716XXgUFxJZ3UV3ASJYWBFe12Gr2tzo8eoNKqxFMsxPQ96c1qBokZHNeReNtK/svXWeNQsFzyg9+9dHrnxCtod8OkJ58nQS/wAIrg9Q1K81S4M99L5j9sdBTgmBBXoPwxuVe2urYqm+JhtOOcV51muh8B34s/EsaO4WOYHcT0zVz1Qj2Gis2fXNLt2KzXsSkdRmqj+LdAj+9qUVYWGbtJ+JrmpPHXh2PpfKx9hVOX4jaGn3GL/SnYDsfxNFcHJ8TtOX7lnK/wBDVWX4pRAHy9MkOPU0WYHo1McKylWGVIwR61leH9ci1jSEviohz95SelZ+u+NtL0oNHC/2q5H/ACzTtRYDm7vQv7H8fWssI/0a4fcP9k+lemY5NeRWfiO+1nxXZPfMoh8wCNAMYr13uaGAvauA8b+KNW0fXEtbCSNYjEGO4Z5rv657xF4R0/X7hLi53LMg2hge1CA4MeOvEH/PaD/vkUv/AAnXiD/nrB/3yK6P/hW2nf8APZ/zo/4Vrp3/AD3f86u8QOc/4TrxB/z1g/75FI/jjxAyMpmgwwwflFdH/wAK10//AJ7v+dIfhrYdrh/zovEDztmLMzN1c7j9aAa3fF3h+Dw/LbpbyFxKec1gVqncQ6lptLTAKWkopgFLRRQAlLRRQAUUUlABRRRQAUtJRQAUUUUAFFFFABRRRQAUUtJQBZtfuGqc3+uarlt9w1Tm/wBc1AEdFLRTAKKKKAFooooAKKKKACiiigAooooAuaT/AMf4+lXNZ/161T0n/j/H0q3rH+vWp6gUKKKKYBRRRQAClpKWgBKWikoADRS0lAwptONNNAjv/hd/x73/APviu9Brx7w34mbw9HOi2xm84547Vt/8LKk/6BzfnWEou4HpFFebf8LMl/6BzfnSH4mTD/mHt+dTyMZ6VRXHeGfFOp+ILrEenNFar96ZumfSuxqQCiijNABRRRQAmKXFFFACUUtc/wCItR1fTAZ7GyN3CBllXqtAG9mlzXmQ+J0+4q2nMCOoPGKkHxMl/wCgc351XKwPSc1y/wAQ/wDkWpPrWB/wsuX/AKBzfnWbr3jKXXNOaza0MQbqxqlB3A51fuL9BSmmjgAegp1bgNIzhR1Y7R+Nex+E9MGl6FBCVxIRlz61514N0r+1deQuuYLfl/r2r14DAwO1Y1H0AWsLxffiy0YqGw87iIfQ9a3a8w+Il7I/iTTrUn91HIOPU5rNAaUfw20/y1zOxJ5zn1p3/CtdO6GZ+feu1hOYYz/sj+VSCq5mMytA0caJZm1jmaSLPyhjnbVbU/C1pq0xkvZ5z/so+BW/ilqbiOP/AOFdaGOiyD/gVR3PgDRoraSRRJuRcr83eu0qC8/485v9w0czAh0kbNLtl/uxgVdzXic/ijXLee4ihuwEjchRjoK9K8F3tzf6Ak94++UtyabiB0OaM009K4Lx94l1XRtTht9NkVQ6g4Izk1NrgdzLBHM6NIoJjOV9jUoNZ+hG8bSYJNQkDzyKHJA6Z7VfNAEV3cJa20txIcJGpJrwy+unvdRuLqQ5LuQD7V6R8RdV+x6OLKNv3l0drDuBXmCDaoHoMVtTQC9q2PDep31vrNjbRTkW5blM1kVc0P8A5GGx/wB6rlsB7kvQH2p1NT7q/QU6uYDiPiLqd9pyWRsZvLLk7veuN/4SjXv+fwV1HxT/ANXYfU1wXatoRTQGqfE+vf8AP6KkTW/FUq7oXZ17MB1rEb7jfQ1694IRG8LWZZEJ2nkiidkI89/tbxd6P+Rprat4u/2/yNex+VH/AM80/wC+aPKi/wCeaf8AfIqOYZ4ydX8W/wDTT8jTf7Y8Wesn5GvZ/Ji/55R/98ijyIv+eUf/AHyKOZAeM/2z4s9ZPyNOGs+LPV/yNeyeRF/zyj/75FHkxf8APKP/AL5FHMB44db8VKCXdwB1ODUY8Ta73vRXr2owxf2fcfuo/wDVn+EeleHOMTzAf3zVwtIDU/4STXP+f2mS6/rEsbJJd5Rhhh61nClrVRQAv6mlFApaoAooopiCiiigAooooAWiiigAooooADVuP/UfhVSraf6j8KQGd3NFHc0VQwooooEFFFFACUUtFAxKWkpaACiiigAoopaQgooopgFFFFMCjRRRXKAUUUUCCiiigAooooAKKKKACiiigAooooAKKKKBktr/AMfcX1re1f8A480rBtv+PuL61u6v/wAeiUgMeijtRTAKKKKACkpaSgAooooEFFFFABRRRQAUUtJQMWikpaACiiigApaSigBav6Dqq6LrCXzwmVV/hHWs80UWA9BPxQtAedNl/Oj/AIWlZf8AQOm/Osv4faNYatJeC/h8zywNvPSu0/4QrQD/AMuX61m7IZz/APwtKy/6B0350f8AC0rL/oHTfnXQ/wDCFaB/z5frR/whegf8+X60tAOePxSsv+gfL+dJ/wALRsv+gdL+ddD/AMIVoH/Pl+tNk8FaBsY/YugPejQCHw/4203Wrj7MAbec/dRz96uoFeAahH9j1i4W2JjML5jI6ivTfDXja0vLa2tb18XzDaR/eptAdliuf1Hwdo+p3r3d5HJJK3ffwPpXQ9qSpTA5X/hX/h7/AJ95P++qUfD/AMP/APPvJ/31XU0tF2Byw8AeH/8An3f/AL6pf+EA0D/ng/8A33XU0U7sDlV8A6EkiSJFIrodykP0NdPEnlxKmSdoxk0+kJ9KQATXB+NfGYtA+m6VIGuDxJKOie1UfF/jaZnn0zTFMTKSksv9BXBY5JJJJ5JPU1SiApZmYs7FnY5Zj3NdR8Oj/wAVS3+4K5fFdP8ADv8A5Gk/7lW9hHsFFFFZDPD/ABc4XxVd53dP7tY5kQsn3vvj+H3r3ubSbCeUyy2sbOepIqP+xNM/584v++arm0AtWP8Ax42//XJf5VMelCqFUKowAMClqQPPPiv/AMelj/vGvPIf+Pq3/wCui/zr0P4sf8elj/vGvPIf+Pq3/wCui/zq1sB9AQf6iL/cH8qkpkA/cR/7g/lUmKgDn/F/h+TxDpy20U4hIYHcRmuU/wCFZ3X/AEEk/wC+a9MxRTvYDzX/AIVpdf8AQTT/AL5qay+Hd3bX0FwdRRlicMRt64r0SijmYBQaKQ0gOB+K3/IPtv8AroK87PWvQ/iv/wAg62/66CvO+9aw2ELS0lFUAtFFFMAooooAKKKKACiiigAooooEFFFFAwooooAKKKKACiiimAYqzbf6s/Wq1Wbb7h+tICjL/rX+tNp0v+tb602mAUUUUAFFFFABRRRQAUUUUAFFFLQAnrW9of8Ax5tWD2Nbuh/8ebUmBnXH/HzJ9aZUlx/x8yfWo6YBRRRQAUUlFAC0lLSCgBaKKKACiiigAooooAKKKKACiiigAooooAKKKWgBDTXP7tvoadTX/wBW30NDA9b+H5/4peCumJrmPh//AMivBXTVzvcZQ1nS4NY057K5JEbHJxXLn4a6Vt2iR8fWu3pcUXsBS0nTotL0+KzgJKRjAzV0CjFLSAxPGP8AyLN57Ia810XwlqOq6XFdW0gCHjFeleMf+RYvf+uZqh8OufCsP1qk7IDk/wDhAtZx/rVpD4C1r/nopr1eijnYHCeDPC+paPrMt1eMDGybRiu7oopN3AK888X+EtU1XX3vLRgIygUfWvQ6Q/WhOwHkf/CBa7/eWnt4K8SG2W2M7CBTkIDxXrP40n4mnzMDyNfAGtqOCo+grN1rQb3Q/J+2kHzvu4r278687+Kv39N+ppxk7iOCzTWJAypIOeo606mSfc/GtWB6XofgzQ77Rra6uYp3llXLsZOprRHgHw5/z5ufq1aHhM/8U1Zf7lbFYMZzP/CBeG/+fE/nSjwJ4cHSx/WugluYIXVJpURm6BjjNOWeFhlZYz9GFAHPjwR4fHSyP/fVOHgnQP8AnzP/AH1W8Z4VHzSxj/gQp0cscq7o3DD1BzQBip4U0mNNkcUqL/dWQgVzfjvQNM0/QftNrBtmEgG8nJr0GuR+JX/IsH/rqKFuB5ZE7wyxzxf6yJty/WvZ/DGuQa3pkcqOPOUBZU7g141BHLPKkMCb5XOFX1NbNroPiqxuBPZ2UkUo9G4NaSSYHs1FedQav47hAE2miXHcCorzx3renTCC/wBN8qUjO32rOwHpVJXl5+JWof8APjSf8LK1D/nxp8rA9RJ9qT8K8u/4WVqH/PjR/wALJ1H/AJ8qOVgWPil/x9WX1ria1Nf8QT+IJInuIfK8roKy+tbQVkIUUtJS1QBRmlopgJS0UUAJS0UlABRRS0AJS0lFABRRRQAUUUUAFFFFABRS0UAJRS0UAWLb7hqnN/rmq5b/AHDVOb/XNQBHS0UUwAGlpKWgAopKWgAooooAKKKKACiiigC5pX/H+PpVvWP9etVNK/4/h9Kt6x/r1qeoFCiiimAUUUUDCloooEFJS0lABRS0UDEpMUtFAhMUwg09mVevPsOa1tI8MarrDKY4jBbn/lq3+FS2kBh8lwiAs7cBRzmuy8OeBLi8ZLnWMxQdRD3auv0Dwlp2iqHVBNckfNIwyK6EDFZSn2GQ2ltBZ26QW0YjjQYAAqbNBqveXcFlbPcXUgjiQZJNZrUCS4uIraB5p3CRoMsx7CuF0jxJJr/jopbsy2UMZCjP3z61zfi7xXNrsxtrYmOwQ9B1k96f8N8DxWQOP3VXy6AewUGkFKelQBw/ibxRdaB4qiQjzLJ4hvjHUe9dZpmp2mq2q3FlKsiHrjqPavNviYM+I4v+uIrn9K1O80i5FxYylT/Eh6MPpWnJdXA91pCM8Gud8N+LbLWkWJyILvvGx610dZtWA5jxD4M0/WN0sai3uuzrwPyrzfWfD+paI5F3Ezw54mUcGvb6jmhjmjKTIroeoYZqlJoDwJTkZBpwr0XXvAME5a40l/Ik6mM8hq4K/sLzTJjFfW7RN24yDW0ZJiIM0FsL7k4H1pMjFdJ4H0I6vqYu51/0S3OeejmnJ2Qzt/BGj/2VoqNKuLif5pP6V0lNHAwBwOBTq53qAGvI/iKxTxTC4/gIb8q9bJ6V5J8Rf+RmT6VUNwL8fxNeOJEOmsdoxnPWnf8AC0v+oW/51k+BNKstY1GaK/iMiKPlAOK7seCNA/59G/76puyYHNj4pj/oFP8AnTx8UlPXSpP++q6MeCdA/wCfRv8Avqnf8IToI/5dD/31U6Ac4Pigh/5hcn/fVKfiP9pR4U0mQs6kAA10X/CFaD/z6H/vquYl0y00v4h2lvZxbYWjBKtzzT0A42ew1GWSaQWUmZG3Yx0r1bwJBNB4cjSeMxvu6Gui8qP/AJ5p/wB804AAYAA+lJyuAhHFcjrnh+bVfGVncun+jW6hiT3PpXYUmKkBoAHAGAOlK2AMnoKXFLigDyHxPBrWsa5Lc/2fJ5K/Ig+nesa4sby0AN3btEG6E96922j+6PyriPiaANNs8AD94egrWE+gHnOKuaIP+Kgsf96qtXNE/wCRgsf96tZbAe3p91fpS0i/dH0pTXKB558UzhdP+prg88Cu8+Kn3dP+prhB0rop7AMc/I30Nev+Bj/xStn9DXkEn3G+hr1/wN/yKtp9DU1AOh7Vn6zqsGjae17c5ManBxV+ub8e20934YlitovNkLghayQFD/hZGkkcRuaP+Fj6V/zyevOk0PVQOdOb8qcdE1XB/wCJc35VpyxA9m0LWbbW7I3VqCEB2kGtKuT+HdpcWmhPHcwmJzJnaa6ys3uBW1H/AJB9x/1zP8q8LlH+kzf75r3TUP8AjwuP+uZ/lXhcv/HzP/vmtKQCCnUClroASloopgFFFFAgooooGFFFAoELRSUtACUtHeigAq1H/qPwqr2q0h/c/hQBn9zRQepopjCiikoAWikpaAEooooELRRRQMKKKKACiijNAC0UUUCCikopgUqKKK5QCiiigQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQMltv+PuL61u6t/x6JWFbf8AH3F9a3NW/wCPRKQGPS0UUwCiiigApKKKACiiigQUUUUAFFFFABRRRQMWiigUAFFFFABRRRQIKBRSCgDvfhV/rb/6CvR68P0B9fWSb/hHwScfvK2PN8f/AN0/lWbWpR6xRXk/nfED+6fyo874gf3T+VKwHrFMk+43+6a8qM/j/wBD+VN+0ePhnhjxzxRYDnNY/wCQ7ef79SeHQP8AhJbM4HWqc5mN3Kbricn5/rV3w9/yMdn9a06CPdfT6UtAHT6UuORWQzktd8cWmi6mbKa2Z2H8QNUR8TNOP/LpJ+dcv8QVH/CVtnHTvXO4X/Zq1FAemf8ACy9P/wCfWT861/Dfi218QXEkMELRsgycmvHOP9muz+FhH9sXAGPuHpQ42A9TzTe4ozxR3qAPBtb/AORgv/8Arsapirmt/wDIfv8A/rsaqdq1Qgrpvh5/yNJ/3a5k10vw8P8AxVR/3RQ9gPYaKM0VkMKK8617x5f6brc9lFbBkj6GqI+JWogrmyGCQKdmB6nRUVtKZ7aKUjG9A2PqKlpAee/FcZtLH/eNefW65u7b/rov869D+Kv/AB6WP+8a8/tRm+th/wBNF/nVrYD3yAfuY/8AcFSUyL/Vp/uin1AEF3eW1nHvuZVjU9zVYa3ph/5fI/zrlPip/wAgeHkj5uxrzQoB/HJ/31VKNwPdf7b0z/n8j/OlTWdOdwiXcZYnAGa8HKf7cn/fVWdKXGs2Xzyf65f4qfKB79mkPSkoqQOC+K//ACDrb/roK87716J8V/8AkHW3/XQV53/Ea0hsIWiiiqAWiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMAqzbfcP1qtVi3+4frSApS/wCtb602nS/65vrTaYBRRRQAUUUUAFFFFABRRRTAKKKKADsa3dE/482rC7Vu6J/x5tUsDOuP+PmT60yn3H/HzJ9ajpgLRSUtACUUUUALRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUtJRQAE0xslWHqMU4000Ad54V8XaVpOiRWl05WVOvFbX/Cf6F/z2P5V5OQO6j8qNo/uL+VZuIz1n/hP9B/57/pS/wDCfaABk3HT2ryTav8AcX8qbIo8p/lXp6UcoHv+n3sGoWcd3bNuikGVNWawfBXHhaywMfJW9WTAxfGH/IsX3/XM1n/Dr/kVYPrWj4uGfDV6P+mZryfTPEes6fYR21ldLHCBnBFUldAe40V4yPGPiIf8vy/9804eMfEX/P8AL/3zRyMD2SiuO+H+ranq1tPNqM4lCnC4GK7GpasAUlBri/iHqeq6TDaT6bcCJXcq/GaAO0NJXi//AAmPiP8A5/l/75pf+Ex8Rf8AP8v/AHzVcjC57PmvO/ip/rNN+prnf+Ex8Rf8/wAv/fNUdT1fUNXaM6jMJTH9zAxiqjGzEUqZJ938afTZPuj6itGB7X4U/wCRasv92tgCsnwuu3w7Zj/ZrXFYMZw3jLw/ea94isIoJHigWM+ZKp6VrQeCtJhsxADcE45fzTkn1qHxN4qHh7VrSGSDzIJ1JcjqvvW3pmr2Oq24msrhJFPbOCKAPPvEHgPUbeOSfTL2e4Tk+TvOQK6P4dJJHoDRzbxIr4Ic5IrpL++t7C0kubmVEjjXJyetZPhPVI9Xtrq8hh8qJ5fl46j1ouBv1yHxL/5Fj/tqK6+uQ+Jf/Isf9tRRHcDz3wz/AMjHY/8AXQV7h614d4Z/5GSx/wCugr3L1qp7gFec+PPD+sapr6T2FsZIhEF3Z6GvRqb+NSnYDxseCfEZ62+Pxpf+EH8Rf88P1r2Pv1P50Y9zVc7A8d/4QjxD/wA8P1psngzxBEhc2/CjJ5r2P8TUN6D9jm5P3D3o5mB4NyCyt1U4P1pwok/4+J/+urUorZCCiloqgCkoooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACloooAKKKKALFt9w1Tm/1zVbt/uGqk3+uahAMooopgApaSloAKKKKACiiigAooooAKKKKALelf8AH8PpVzV/9etU9K/4/h9Kt6t/r1qeoFHvRQetHWmAUUUUALRRRQAUlLRQAUlFLQMSnxKkk8ccsvlRuQC/pTDQQCMHoaGI9V0HwhpFlHHcKouZSM+aeQfwrpVRVGFAUDsBgV5x4E8Sm1lGlX8n7pv9S5/h9q9JBrmle+oxMUUkkiRoXkYKo5JJxTYZo7iJZYXDI3Rh3qQM/W9csNFtTPfTBePlQdWPpXkfiPxRdeILg+YWitVPyRD+tesX/hrStQuDPeW5lf3bj8qrf8IfoI/5cf1qotIDxcOv+RW54O1W00rxB9qvHKRFNu7Fem/8IhoP/Pl+tL/wiOg/8+IP1qnMCt/wnnh0Hi8z/wABpf8AhPPDx/5fD/3zVj/hEdB/58Fpw8JaEOlglToB51441ax1XWY7ixlMsYj2k471z/mL7/lXtA8LaKOligpf+EY0b/nyWqU7AeLeaFcOjOjqchl4Ndr4a+IDW+y01rLR9BOB0+tdr/wjOj/8+SU1vC2isMNYoRSckwNS2uYbuBZreRZI2GQQamrBubODw7pN1daZGymNd2wtkGofDXi+w1yMISILr+KJjUWA6PFUdXjsPsEr6kqGBVyxYc1eZgqliQABkmvJ/GniJ9YvTaWzkWcJwSP4zVRTbA57UXtJdRkaxjMFmzgKpOePWvZfD1raWui28VgQYdoIYdz3rxUqCMY4xivQPhrqxaGXSpn5i5iz3rSadgO+xRRRWICHtXkvxF/5GaP/AHa9a7ivJ/iOP+Kki/3TVw3An+GH/IYuOR0NeoZ9x+deE6UdVFw39jZ83+LFau7xr6SU5RuwPYsj1H50bvcfnXjufG3pJQD429JKnlA9jyPUfnXCayQvxJstxAzGOprmAfG3o9ZepNqi3anViyXOPlOecVSiB7n50QPMqA/7woE8OcebHk/7QrwRriYDLXM3/fZrtvAvhu5uZV1TUXnESnMUbOeT60nGwHpNFFB4qAEZgoyxAHqTTfNT++n/AH1Xmvj7xMbq4/svT5sRxn97Ipxk+lciLi5AAF3LgdPnq1C4HvHmJ/fX/vquK+JjA6ZaYZT+8PQ158Lm6/5+5f8Avs0jyzSYEsryAdNzZxVxhZgJVzRP+Rgsf96qdXNE/wCRgsf96tJbAe3L90fSlNIv3R9KDXKB578U/u6f9TXCV3fxT+7p/wBTXCA+4rop7ANcfI30Nev+Bx/xStn9DXkTYKkZHI9a7vw74z0vS9Ft7Kc/vIxhqVRXA9DxSEAjkCuR/wCFhaL/AHzSH4h6L/fNY8rA67Yv91fyo2L/AHV/KuQ/4WJo3qaQ/EXRvU0crA7EAAcAD6Utcb/wsXRf7xpw+Imif3zRysDqL/8A48bj/rma8MlH+lT/APXQ16Rc+P8ARZbaWMOcspArzd2DzyuCNrtkc1rSVgAUUceopa3ASilpKYgooooAKKKKAFpKWkoAKWkpaBiUUUtACVaT/U/hVU1ZT/VfhQBQ7mijuaKYBRRRQAtFJRQAUUUUCFooooAKKKKBhRRRQAtFFFAgooopjKNFFFcogooooAKKKKBBRRRQAUUUUAFFFFABRRRQMKKKKAJbb/j7i+tbmq/8eqVh23/H3F9a29V/49VpAZHaiilpgJRRRQAUUUUCCiiigAooooGFFFFAgooooGLRRRQAUUUUAFFFFACUClNNzQB3nwpP+kagP9kV6RXlHw71Wx0q4vWv7gQiQDbnvXc/8JfoP/QQX8qza1Gb1FYP/CX6D/0EF/Kj/hL9A/6CC/lSsBvYFI4+VuB0NYY8YaD/ANBBfypG8YaBgj+0F6elIDyXXB/xUN9/vVJ4eH/FR2f1qLVpo7jWrueFt0TtlW9am8O/8jJZ/Wtegj3MdPwpaQdKWshnLa54KsdZ1E3k8jK57Cs//hWumf8APV67mindgcL/AMK10z/ns1avhzwjaeH7x7i2kLF12kGukxRii4CdqO9Lik7ikB4Nrf8AyMF//wBdjVQVc1r/AJGDUP8Arsap1qhAa6T4ef8AI1H/AHRXNmuj+Hv/ACNX/AaGB7D3NLSdzRWQzxLxj/yNt39Kxmz8n++v863fF1tdP4qunS3kZccEKax2tLzC/wCiy/eH8J9a0Wwj3rTv+Qfbf9cl/lVmqum5/s22yMHy14/CrVZjPP8A4q/8edl/vGuC09d+qWi+riu9+Kv/AB52P+8a4vw9F53iOxjx1OatbAe5R/cX/dFOpAMAfSlqAOF+KSu2kwiNGc7uijNebGKf/n3l/wC+DX0C6I4w6q31Gaj8iH/njF/3wKpSsB4B5U//AD7y/wDfBqzpUcv9s2WYJR++X+A17t9nh/54Rf8AfAoFvCDkQxAj/YFNyEPAp3agUdqgZwPxX/5Btt/10Fedd69F+K//ACDLb/roK87/AIjWsNhBS0lLVAFFFFMAooooAKKKKACiiigAooooAKKKKACiiigAoFFFABRRRTAKsW/3D9ar1Yt/uUAUpf8AWt9abTpP9a31ptABRRRQAUUUUAFFFFMAooooAKKKKADsa3dE/wCPNqwq3dF/482qWBnXH/HzJ9ajqS4/4+ZPrUdMAooooAKKKKAFooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKQ0AIa1ND8P3evCY2bhPJOGrL713fwrP72/H+0KmTsgM7/hXusH/luKP+Feax/wA/Ar1fH1ox7msuZgeUf8K81j/n4FNf4d6wyMPtC8jFes49zRj60czGZnhywl0zRbeznbdJGuCa1KSlqQMbxb/yLd7/ANczXikX+pT6V7X4u/5Fu8/3DXikX+pT6VpAB+KOnNLTZAShVfvHgVoI9T+GtuYfDe4jl5Ca6+svw5aiz0K0iAwTGGP1IrUrB7jErmfH1kLvwzO4GWhG4V09QXlul1aS28gysikGhMDwBOUX6c07FT3ds9pfXFvIMFJDge1R4rdCG4p2KdjijFMBmKa4zsHq4FSYqzpdm1/q9raIMlnDflSewHs2iRGHR7WMjkIK0BTY0CRqgHCqBT652M8+8f6FqWr6zZ/YLfzEVDubPAqvpHw91CB1uJNTa3cfwL0r0KG7gnnmhikDSQnDjuKmp3A5mbwfbX+3+1riW5C9FDbQa37Gyt9PtUtrSIRxIMBRS3d3b2cDzXMqxxoMkk1BpWq2mrWxuLJ98QOA3rSAvVx/xM/5Fn/tqK7CuO+Jx/4pgf8AXUU1uB574a/5GSx/66Cvcu5rwzwz/wAjLY/9dBXufc1U9wA9K5/XfFmm6HeC1vGPmMu7AHaugPSvJvifx4liJjYjyRyFzUxQHSn4jaN/tflR/wALH0j+635V5XvX/nm3/fFLuX/nm3/fFacqA9RPxI0kdI2P0FRTfEbTHieP7PICykDcMVyfgfTm1DxHE32fdDAdz7kwMV6fqfhnSdUTbc2aZ/hZRjFS7JgeME75ZH7O5YfjTsV2mq+ARY2k11Ff/JENwQjt6VxYOc/WtotMQUUUVQBRRRQMSiiigQtFJS0AJS0lFABRRS0AJRRRQAUUUUAFFLRQAlFLRQAUUlLQBPbnCGqkv+uarcH3DVSb/XNQgGUUUUwFoopKAFopKWgAooooAKKKKACiiigC3pX/AB/fhVvVv9cv0qppX/H7+FW9V/1y1PUCjRRRTAKKKWgYUUUlAgpaSigYUtJS0AFJS0lAhOeCDhlOVPoa7rTvH62ujpHdQtNdxjHB+971wtJjmplFMDS13xJqetbhPMY4OcRLxivVPCS7fDGnj/pkK8Wk+4fpXtXhQ/8AFNWH/XIVlNWGa9NJAGSQB6mlNZuvwT3Gi3UNsCZnQhQDjmoQF3zov+esf/fQo86L/nrH/wB9CvF08KeLAgHkTD6yGpB4U8V/88Jv+/pqrAey+dH/AM9Y/wDvoUedH/z1j/76FeN/8Ip4r/54Tf8Af00f8Ip4q/54Tf8Af00coHsvnxf89Y/++hR50X/PSP8A76FeN/8ACK+Kv+eM3/f00j+FfFmw4hm/7+mlyge0g5GRyKKpaMksWk2sdwCJUjAYH1q7UgZfiUZ8PX3/AFzNeIwgr86MUcNwynBr2/xH/wAgC9/65mvEo/uH/erWmgN0+LNWfSX06WXcGG3zB1ArEAwMfnS4pcVqopAJV3Q746Zrlrdg/KG2t75qnSEZx7HIoaugPXPEviKPSre2SIg3NyyiNfY9a3kJZFJ6kA14YLi4u9UtpbqUyMrqFz25r3OP/Vr/ALornlGwC4ryr4jr/wAVHD/umvV68r+I3/Iwwf7tOG4DvhiP+JxcDAPBr1AKP7q/lXk3gPUbPTNTmlvZhErAgGu//wCEt0P/AJ/1/Kia1A2to9B+VGB6D8qxD4t0L/n/AF/KkPi7Qf8An/X8qkDbIH90flXlfxHV5fFNvFFGXdosKqjrXcf8JfoP/P8Aj8qm05tH1S6Op2ZSeZRs3nkgfSmnYDlPCngZgyX2tLyOUg9Pc16EiqiBUUKo4AHagUuaTbYC1W1G3kurKWCKYws4xvHarGaWkBwY+G1rks14zMxyxx1NOHw4tP8An7b8q7qiq5mBwrfDq1AOLwj8K4CVBFcSxA5EblQfXFe7yfcb6GvCrs/6fdf9dW/nWlNtgR1b0T/kYLH/AHqp5q1op/4qGw/3q0lsB7iv3R9KDSL91fpS1yged/FY4SwPpk1S8MeD7fW9HjvZbgoz9VHarnxX/wBXY/Rq2/hwP+KVgP1rS9ogUB8OLL/n5b8qP+Fb2B6zk/hXdUVPMwOF/wCFa6d/z2P5Uf8ACtdO/wCe36V3VIaOZgeV+KvBtpomkm9hlDsHC7cetUvCHhi28QpO07iMxNgDHWu0+JH/ACKzf9dRWR8K/wDV3f8Av1V3YCf/AIVnY/8APz+lH/CtLH/n5/8AHa7sU6p5mBwg+GtkOlz/AOO08fDmzH/L035V3NFHMwOGf4eWixswujkDPSuBmTyrmaIHIjcrn1r3Sb/Uv/umvDLw/wDExu/+uxrWnJsCI0UUVuIKKKKBhS0lLQAUhpaSgAooooAKWkooAKsp/qvwqtVhP9V+FAFH+I0UfxGimAUUUUAFFFFABS0UUAFFFFAgooooGFFFFAgpaSloAKKKKYFGiiiuUAooooEFFFFABRRRQAUUUUAFFFFABRRRQMKKKWgCS2/4+4vrW3qn/HqtYlt/x9xfWtvVP+PVaAMiiiigAooooAKKKWgQlFLRQAlFLRQAlFFFAxaKSloAKKKKACiiigAooooAQ00040lAGz4b8KS+JIpJY7ryfL7Y61s/8Kvuf+gh+laPwp/487oe9egAcVm3qM8v/wCFX3P/AEEf0pR8L7j/AKCP6V6fS0rgeYj4X3H/AEEf0p3/AArC4/6CP6V6YKDRdgeEazpbaNqklg8nmlP4/Wm6XdJY6rb3cgykZ5Fa3j7/AJG24+grnq0WqEeqwfELSJp44VDbpGCjjua7AHNeD6LF5+u2caqM7wele7is5Kwxaw/EniW18PpG1ypbf0xW5XmXxakzNaw/9Myf1pID0LS7+LU9PhvIP9XKuRVqsLwT/wAirZf7lb1AFDWNUg0iwa8uP9WvWuYX4k6QwB2sPwrQ8fjPheb6149Go8pOB09KqKuBYv51u9TurpBhJpCy/SoaXFGK0Qhpro/h7/yNX/Aa5w10fw9/5Gr/AIDSYHsPc0UdzS1kMYY0JyY0J91o8qP/AJ5p/wB80+loABxRRmigDz/4q/8AHpY/7xrmvAcHneLbZsZEakmuk+Kp/wBFsf8AeNVfhZZ77q8vGHCYCmrWwHpdFFIagAJpPwNeceIvHl9Ya5cWVpAGjhON3rWePiLq/wDz7CnYD1jPsaPzryf/AIWLq/8Az7ip7L4g6pPf28D24CySBSfrRygeo5o7Ugp3akBwHxX/AOQZbf8AXQV52eteifFf/kGW3/XQV52eprWGwgpaSlqgCiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMA7VPB9yoKnh+5+NAFOX/AFrfWm06T/Wt9abQAUUUUAFFFFABRRRTAKKKKACiiigArc0X/jzasPtW5ov/AB5tSYGfcf8AHzJ9ajqS4/4+ZPrUdABRRRQAUtFAoAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACkpaKAG4rtvhlPDBcXomkVNx43HGa4qjBzwzD/dOKUldAe9Q3dtM+yGdHb0BzU9eV/DKMtrk0m5yFUjls16oKwasxhmqrahZoxDXUQI6gsOKnmOIXPoprwHWSW1XUH3yffJ+9QlcD35JUkQNGwZT0Ip2awfBQx4Wsskn5O5zW9SegGL4vP/FNXn+4a8VhP7lPpXtHi/8A5Fq8/wBw14tD/qU+lawAkzWh4fs21DxBaW4GVVwz/Ss769K774YaWQs+rSry/wC7XNOT0EehxqqIqKMBRgU6mA04ViMWio4Zo5lJicMAcHHrUlAHl/xI0s2uqR6jGv7u4+RsdiK5Cva/EWlJq+kTWjAb2X5G9DXi0kTwSvBKpWSJipB9q1gxCUUUVoAnbNd58ONEfe+r3CYHSDPp3rgzzxXT+HPGd1pOy2vF86zHAwMeWKmd7aAesVHPIsMDyscBFLflUGn6jaalbLPZzLIjeh5/KuX+IWuCzsBp1u/+kT9SD90VildjOV0rxL9k8X3OpyufstwTvFamtfESZi0elQbUHSY85/CuIAAGMU1hlSPUYrXkQHoMfhW48TafFfX2tSOZl3YUYA9q6nw3oaaDpxtI5TKCc7sYrC+GuofaNGltmcboH2qCe1dkWGPvL+dZMBc1xnxQfHhpB6zCusmuYII2eaeNVUZJLCuB8TteeMpo7PRIy9nE2ZJm4G729aaA43Q5hb69YyscDzQCfSvd1YMNynIPINeTa74GvtM05bu1lNwyDMigcp71veDvGttPaR2OpyeVPGNqMf4h705a7Ad5UFxZ211j7RCkmO5FOjnikUMksbA+jVKCD0IqAKH9jab/AM+kf/fNL/Y2m/8APnH/AN81foouBDb2sFsu2CFIx/sjFTUhIHUgfjUFze21rC0txPGiKMk7qAMHx7ei18OSx5w0/wAq15Mv3R9K6Hxp4hTXL5IrUk2sByrf3jXPVvBWQC0UlFaAFFFFAhaSiigAooooAKKKKACiiigBaSlpKAClpKWgAooooAM0lLSUALRRRQMnt/u1Ul/1rVag+6aqy/61qBDKWkpaYBRRRQAUUUUAJS0UUAFFFFABRRRQBb0v/j9/Creq/wCuWqml/wDH6PpVrVf9ctLqBSooooGFLSUUCCilpKBhS0UUAFJS0lAC0lFFACUhpaQ0ARyH5D9K9o8KH/imtP8A+udeLyj5D9K9o8Kf8izYf9c6yqAbIoxQKqatdPZ6Vc3MYy8UZYD3rEC3+dH515InxH1cqCYDz/s07/hYurf88D/3zVcrA9Zoryf/AIWLqv8AzwP/AHzR/wALG1X/AJ9z/wB80crA9X/Ok5968oPxH1b/AJ9z+VMb4j6uFJ+zn8qOVgeuUtUdIuXvdLt7mUYeRAxFXSakDN8Rn/iQXv8A1zNeIxH5D/vV7Z4j/wCQBe/9czXiUX3D/vVtTAlp1NFLWwBRRSUAS2v/AB/W/wD10X+de6x/cX/dFeFWn/H9b/8AXRf517rH/q1/3RWFTcB9eV/Eb/kYIf8Adr1SvK/iN/yMMP8Au1MNwM7QPC9x4gillhn8sRNtPvWqfhren/l9H5Vq/C7/AI8b7/roK7unKTuB5cfhne/8/wAPypp+Gd9/z+j8q9ToqeZgeVf8Kzv+16Pyp1vpuoeA7yG/ln820mcRSj0z3r1OuL+KC7vDSg/89RQncDqft1otvHcNOixSAbCT1qwrB1DKcg8g14Gb28uGs7eed2hicbFBxivdbD/kH22f+eYokrAWa86v/iJd2uo3FqmnBxE5UNnrXoteX6l4B1S51W5uY5gElkLAUo26gaugeOrrVdYhspLARLIcbs13ledeG/BOo6ZrcN5PMGjQ5Ir0TNErdAEk/wBW30NeEXR/0+6/67N/OvdpD+7b/dNeDXJ/0+6/67N/OtKQDc1b0TnxDY/71U6uaH/yMNj/AL1ay2A9xX7q/QUuKRfuj6CnVygedfFf7lj9Grc+HP8AyKkH41h/Ff8A1dj9Grc+HX/Ip2/41b+EDqa4P4i61qWl3Vkun3Pkh1O/jrXeV558S9Pvr68sTZ25lCqd2O1KO4Evw61rUtUuLpdQufNCH5eOld5Xn3w102/sbm6a8tzEGPGe9ehAUS3A5L4kf8iu3/XUVkfCr/V3f+/Wz8SB/wAUw3/XVayPhWP3d5/v0/sgeggcVyHxF1S/0vToJNPn8l2fBOO1dj2rivibaXN1pcC2sRkYSAkAVMdwMPwl4i1m88RR293eeZER93Feo15F4Q0+/h8TxSTWzogHUivXKc9wGzf6l/8AdNeFXh/4mN5/12Ne5zf6l/8AdNeFXf8AyEbz/rsa0pANFLTRS10ALRRRQIKKKKACiiigYUUUUAFLTaWgAqwh/dfhVep1/wBX+FAFLuaKO5opgFFFFABQKKWgQUUUUAFFFFAwooooAKKKKBBS0lLTASilooGUaKKK5RBRRRQIKKKKACiiigAooooAKKKKBhRRRQIKWkooGS23/H1F9a2tU/49lrFtv+PuL61tap/x7LQBk0UUUAFFLSUAFFFFAgpaSigBaSiloASiiloASlooFAwooooAKKKKACiiigAptONNoA9C+FB/0W7r0EHivPvhQP8ARruvQQOKyluMyvEOtxaFYfa5oy656Vyx+J1hxi1Y5IFX/iaP+KYb615IQNkXT7y00rgfQlpcC5tY5wMCRQQKmz0+tU9IH/Eptv8ArmP5VbPUfWkB454/P/FWz/QVzwNdB8QP+Rtn/CudJwCfStFsI6j4e2v2rxSkpGUhQ5+tew1xPw10hrLSnvZlxJdHcoPUCu1rNvUYteRfEu6Fxr4jB/1MZU+1esyyiKF5G6IpY/hXg2uXv2/VdQus5WRiVpxA9d8DnPhWz/3K6Cud8CHPhW0/3a6KpYHN+Pf+RYnryXTrSe+litbVd0zDgetetePP+RYnrzjwVz4ktPoKuLsgM26tbiymaG7heORTg8cfnUPXoQfpXvF9ptnqEZju7dJF9xXF678P7cRyXOmTGHaN3ldqakKx50RXR/D4f8VV/wABrnmBV3Q9Ubaa3/A00Nt4k82dwiBepNN7AexY60nes/8At3S8/wDH5H+dH9u6XnH2yP8AOsrDOH13x3qen61PZw26skfSqI+I2rn/AJdV/OsHxJKlx4iupYm3I3QjvWcAa0UUB2P/AAsbVv8An1Wk/wCFi6tj/j1WuQwaWnyoRreIPEl5r0cSXUQQRn5cdya9L8CaadO8OQiRcSyjc1eaeFtIbWdcihwfIjO6U+npXtsahEVV6KMCol2GOqG6kENtLKTgIhP6VNXMeP8AUv7P8OTBGxLL8qj1HepQHkV5Obu/ubo9ZXNNHSmqu1QPxp/atkIbVnTP+QxZf9dlqvVjTP8AkMWX/XZaTA98XoPpTu1NXt9Kd2rIZwPxX/5Blt/10Fedd69E+K//ACDrb/roK88PWtYbCEpaKKsAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAVPD92oKnh+7SApyf6xvrTadJ/rG+tNpgFFFFABRRRQAUUUUwCiiigAooooAK29GP+iNWJ2rb0b/j0akwKFx/x8SfWmU+f/j4k+tR0AFLSUtACUtFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUtABSUtNY4Rj7UAd98LIPlvpyP48A16HXKfDuyNt4dWVhg3B3V1dc8txkN422znb0jJ/SvAb1vMnu5P7zE17l4guVtdDu5GOP3ZA/KvCSSbdiepBNVAD2rwWP+KXsv9ytysXwZ/wAitZf7lbZqWBieL/8AkWrz/cNeKw/6lPpXtXjD/kWbz/cNeKRMFt0J7D86uAFm0tJdQvYbKEEtMwU47D1r3DS7GPTtOhtIwAI1AOO59a5L4eeHWtYTqt4mJphiNT2X1rugKUncBMVXv7pLGxmu5D8sSljVquF+Jmr/AGbTU06Jv3k5xIB2WpWoFL4ca6Z9SvrKdziRzJFnvk9K9GHSvAdMu30rULa9i6wMM+4r3Wxu47yziuYmDJIoOR605KwFnNee/ELw8wb+2bNM44nUenrXoVNljSWNo5FDIwwQe9JOwHgYIIBHSlroPF3huTQ7wzwKWspTkEfwGufrdO4CUUuKRuFJ/KmBteDodQm11YdNuXtwVPmP1A/CtDxJ4R1iC6lvhIb5G5Zz1H4V0vw80c2Okm8mXE1182COgrsCM1i5ageAE4JVgykdQwwav6ZoOpaxFJLp0QkWM4bJxzXp/iLw1pmpW8s80axTKpbzV4xisH4XMPI1JEYsizfK3rVc+gHP2nhbxbYymSyiELnqQ3WprjS/HSwvJLL8iDLHd2r1fHuagvBus5x6xkfpUXA8g8L6c/iPV/J1O/cRR/MyF8b/AGr1+1tobS3SC2jEUaDCqBivB5S8GozNDI0ckcnysDjFel+C/F66kq6fqLBLxB8rH/loKckB2RUEEEZB6g968z8deEPsrvqmmxkwscyxr2PrXpuap6tf2Wn2Ek1+6rFjlT1b2xUpgeHW11d24BtbuRB2yxNX01/XYxhdSbH0qpfTQXGozTWkXlQyPlE9q0b/AMN6rp9tHcvAZIJFDbl5I/CtdAGf8JLr4/5iR/Knwa/4nupRFa3skjnphePzqjpz2i6pbrqaMLUviQMMYFe1aVp+nWdqn9nQRpE4DAgZyKmTSA5DSfDvia7HmazqzQr/AM8l5zXHeJLC+0/Vntb+aSQfejbccEV7fWB4t0FNc0wqqj7TF80Te9TF6geOgY6U6lZGjkaKQEOh2sD60ldCEFFLSUwCiiigAooooAKKKKACiiigAopaSgBaKKKAEpaKKACijrRQAZopKKAFFFFFAyaH7pqrL/rWq1D92qsv+tahCGUtFFMAooooAKKKKACiiigAooooAKKKKALWl/8AH6PpVzVf9ctU9L/4/R9Kt6p/rlqeoFKilpKYBRRRQMKWkooAKKKKACiiigAopaTrQAlBpcUYpiIpf9WfpXs/hMf8UzYf9c68akHyH6V7D4Xu7VfDlipuIgRHggsARWNQZvCmyxpLE0cihkYYIPeoft1p/wA/MP8A32KT7faf8/UP/fYrGwFT/hHtJxj7FH+VH/CP6T/z5R/lVv7faf8AP1D/AN9ij7faf8/MP/fYp6gVP+Ef0n/nyj/Kj/hH9J/58o/yq59utP8An5h/77FH220/5+Yf++xRqBT/AOEe0n/nyj/Kmnw7pB62SflV/wC22v8Az8w/99ik+22v/PzD/wB9ijUCSKJIo1jjXaijAHpT8VB9ttf+fmH/AL7FH260/wCfmH/vsUrAU/EQ/wCJBe/9czXiMf3D/vGvateuraTQ7xUuIiTGeA4rxiMfKf8AeNbUwHCloorYApDTqQjigB9r/wAf9v8A9dF/nXu0X+rT/dFeEW3/AB/23/XRf517vF/q0/3RWFXcB9eV/EY/8VDD/u16nXlHxHP/ABUcX+7Uw3A3Phcf9Dvv+ugrvDXAfCs5tb8f9NBXf9qUtwMHxL4mg8P+V58ZfzOmKj8OeK7bX7h4oIihQZOa5z4q4/0Q8f5NVfhcR/aVyBj7tFtLgen1x/xLGfDQ/wCuorsB0rkviR/yLX/bUUo7geUxr/pNt/10Fe9acP8AiX2//XMV4VEP39t/10Fe7WH/AB4W/wD1zH8quYE9J+dL3Feb6j8Q7201O4tVsVZYnKhs9ahK4Ho/50tef6B48u9U1mGxlsxGshxuzXf55NDVgGyf6t/9014Rc/8AH/df9dW/nXu8n+rf/dNeEXI/0+6/66t/OtKQDO1XNFO3xBYk9mqoKltJPJ1C2l/uuK1ewHuyfdH0FOqOE7oY29UBqSuUDzr4r/6ux+jVt/Dlh/wikA9M1T+Juny3WjpdRKW+znkAc81z3hDxjbaJpv2S8XKj7rDrV7oD1XNJwewP1rjP+Fj6R6N+VKPiNo/o35UuVgdoAB2Apa40fEXR/wDb/KlHxE0f/a/KlysCb4j/APIsN/11Wsj4Wf6u8/36r+K/Fum61ozWdsW8wuG5HpWf4N8QWegLOLwnMhyMVai+UD1qkIB6gH61x3/CxdIHXd+VIfiNo4/vflU8rA7HaAeg/Klri/8AhY+j+rflR/wsfRvVvypcrA7Cb/VP9DXhV2P+Jlef9djXorfETRmRly2SMdK85mkWW7uJlPyySFh9K2ppoBoFOoGPWlrcBKKWkoEFFFFAwooooAKKKKACiiigAqdf9V+FQVOv+q/CgRS7mijuaKYwooooEFLSUtABRRRQAUUUUDCiiigAooooAWiiigQUUlFMZSooorlEFFFFAgooooAKKKKACiiigAooooAKKKKBhRRRQBJbf8fcX1rb1P8A49lrEtv+PuL61tan/wAey0gMukoFFMApaSloASiiigQUUtJQAUUUUALSUUtAwopKWgAooooAKKKKACiiigAoxRRQB6D8Kf8Aj3u/rXoXavPfhT/x73f4V6FWT3GZfiHRYtc05rOaQop7iuW/4VlY4UfbZOCD09K72ii4ENrALa1jgU5EagA089R9afSHGaQHjPxAOPFs+eBgdaf4O8MS65eLc3ClLGE5JI++a6u+8KWWt+Kri7ubvcsZG63HU12Vtbw2sCQW8YjiQYVR2qr6APijWKNUjUKijAA7U40U2aVIomkkYKiAsxPpUgcz4+1f+y9AkRG/fT/IF9QeteOsNsDD0Brd8W622u600qsfs8OVi9xWG4zGw9qtIR7N4COfC1t/u10lcp8OphJ4ZRQfuHBrq6l7jOf8cQzT+GrhYELt1wOtcJ8NtPmn17z5IZEihj6suOa9bIBGO1IkaJ9xFXPoMUXAdVDXJPK0W8kzjbETV+ud8dXotPDVwCcGYbBSQHju7cWf+8c0EAjnP4GkQYjUdwKWtkIb5af7f/fVN8tfV/8AvqpMUUWAaBjgUtLRTASljjkmmSCBC8shwqilVHkkWOMAuxwAa9Q8F+El0lBfX4V71xlR1EY9qluwGn4Q0BNC0tUYA3MozK1dBSCjNZbjFNeRfEHVxqWuC1ibMVpxx0Ymu68Z6+miaQ5RgbmUbYl/rXjgLElnbc7HJJqooB1Xk0bVZI1kjs2ZG5Ugdao9VI/KvUvhzq5v9Ha0lIMlodo46irk7CPOTourj/lxf8qtaLomqya3ag2bKqSBmYjoK9t2j0H5UBR6D8qjmGN6GlzkUuKaeAakDgvisf8AQbYf7YrzzvXb/FS5Vrq1tActt3muHzzWsNhDqKQUtWAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAWpYfu1DU0P3aQFST/AFjfWm06T/WN9abTAKKKKACiiimAUUUUAFFFFABRRRQAdq29H/49GrE7VtaP/wAerUmBRn/4+JPrTKfP/wAfEn1plABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABS0lFABUtrbveXsFrECWdxke3eoq7L4b2FtLfSX80qedH8scZPX3qZOyA9GsLZLKyhto/uxqAKsU2g1gM4/wCJl8LfQBbqf3krjj2ryqXAicf7NdP8QdT/ALQ8Q+QhzHaDYfrXLSn90/0rWK0Ee2+DP+RWsv8ArnW5WF4N48L2P/XOt2s3uMw/GHHhm8J/uGuB8BeFX1NotQv0K2kfKKf4zXpurratp8n2/wD49gP3n0qWw+zGyh+x7fs+weXt6YoTAmRAqhVACgYAHanUUGkBDd3MVpbSXE7BY413Ma8P1zUpNY1ia9kOQTtT/d7V1XxC8SC5l/siyf5FOZnHf/ZriMAAAdBWkI9QExkYIru/htrvlltGun4BzCT3PpXDY4pYpJYJ47iBissZypFXJXEe/ZpwrD8K69DrmmpICBcIMSJ3+tborFjILy1gvbZ7e5jDxOMMDXlHibwvc6JO0sKtNZMflYDJX2r16uL8eeJo7G2bTbba9zMMN32CnFu4Hmta/hbRn1vWEjIP2eE7pG9+wrFBAABONx5PpmvYfB1lp9lo0a2MqTFxmRwfmY+9aSdkI3o0WNFRAAqjAAp1JmmSSLHGzuQFUZJPpWIzl/iFq/8AZ+hmCJsT3Bwo9u9YnwnbEV/H/t1zHivWDrWuSTKSbeE7YvpW58LZwmp3NuTy4LAVpayA9QpkyF4ZEHVlIp4pazA8F1a2uLbWLyOW3lyJDghDg1UVpo5FkjjnSRDlWCHg19APbwynMkSN9VFM/s+z/wCfaP8A75FXzgcTpvj0L4e8y7tZ3vo/kCbD8/vXE6vqmp63deffR3BAPyRhDhRXt62lsvS3jH/ARS/ZrcdII/8AvkUkwPE9C0q81LVreKK3kCBwZGZcACvbYoljgSIYIRQozTljRPuIq/QYp1Ju4GLq/hfSdWRhc2qh2HDrwRWb4Z1GGx1GTw3JP5skA3RyN/EPSupkkWKNpHOFUZJrwy+1KYeIptSt2IlimJQjuM00rge7UVmeH9Wi1jSYbuJgSww49G71pZqQPOPiHoPkTDVrVPkbiZQPu+9cR15Fe069faXBp8sWpzIIpBtKg5P5V4w+wTSCI5i3HYfbtW1N6ANpKdSVqIKKKKACilpKACilpKACilpKACilooAKSiigApaKKACkopaACiiigAooooAmh+6aqy/601ah+6aqy/600IBtFFFMAooooAKKKKACiiigAooooAKKKKALWmf8fv4Vb1T/AFq1U0z/AI/Pwq3qn+uWp6gU6SlopgFJRRQMWkoooAKKKKBC0lFFAwpaSgGgRp6RoV/rKSPYKrCI4bccc1fPgfX/APnlH/33W78L+be//wB8V3uKxlNpjPJD4H1//njF/wB90w+BfEHaID6S169ijFRzsDx4+BPEP/PL/wAi03/hBPEP/PL/AMi17FijH1/OjmA8d/4QTxF/zy/8i0f8IJ4i/wCeX/kWvYce5pce5p8wHjv/AAgniL/nn/5Fpw8C+Iv+ef8A5Fr2DHuaMUuYDyD/AIQTxD/zz/8AItH/AAgviH+5/wCRa9fxRijmYHkB8C+Iv7n/AJFo/wCEF8Rf88//ACLXr9GPrRzMDyEeBfEP9wf9/akXwPr4GPJj/wC+69bxRinzsDyb/hCNe/55R/8AfVU9T8Narpdqbm9iRYh1IbNey4rlviF/yLUv1FNTdwPLe2aDSD7q/QUvatwH2o/0+2/66L/OvdY/9Wv+6K8LtP8AkIW3/XQfzr3WP7i/7orCruAp6ivJviP/AMjJF/u16weoryb4jf8AIyxf7tKG4Gz8Kj/o9/8A9dBXoNeffCr/AI99Q/66CvQaUtwMLxL4ag8QiITzNH5fTFReG/CdtoFxJNDO0hcY5FdDSg1NwFrkviR/yLX/AG1FdbXJfEj/AJFr/tqKcdwPMIv+Pi3/AN8V7pYf8eNv/wBcx/KvC4v+Pi3/AN8V7pYf8eFv/wBcx/KrqAWO9eV6v4D1i51e5uYNhjlcsMtXqlGKhOwHmfhvwVq2m67b3lxs8uNsnDV6VjnNLilobuAyQfu2/wB014TdD/T7r/rq38693f7jfQ14Td/8hC6/66t/OtKQEdNcfLx1BzTqQ1swPY/CepLqeg2824F1G1h6Yrarxzwn4hbQdQImJNpMcOP7teuW1xFdQLNbyCSNhkEHNc0o2YEjoroVdQykYIPeuW1DwFo93IXhTyGJycc11dLilewHAt8M7Q/dvnH/AAEU3/hWdv8A9BF/++RXoFFPmYHAj4aWw/5iD/8AfIp3/CtbX/n/AH/75Fd5RRzMDg/+FbWv/P8Ayf8AfIpP+FbW3/QQk/75rvaKOZgcEfhra/8AP+//AHyKT/hWlr/z/v8A98iu9paOZgcD/wAK0tf+f9/++RSj4aWn/P8AP/3yK72ijmYHCL8NrMf8vr/98ipF+HViOt3Ifwrt6KOZgeVeLvDdtoNtBLBKXaR9pBrmicV3/wAUeLGz/wCulefZ5ropu6AdSUmaWtACiiigAooooAKKKKBBRRRQAVOv+r/CoKnX/V/hQMpHqaKO5ooAKKKKBAKWgUUwCiiigAooooAKKKKBhRRRQAUtIKWgQUUUUxlGiiiuUQUUUUCCiiigAooooAKKKKACiiigAooooGFFFFAElt/x9RfWtrU/+PZaxIOLiP61ual/x6rQBk9qKO1LQAUlLRQAUUUUCCkoooAKKKKAFopKWgAooooGFFFFABRRRQAUUUUAFAopKAPQvhV/qLz8K9CrwzR9dv8ARA62LYEn3q0v+E61/wDvis2ncZ6P4n19PD9itzJEZAxxiuUPxOT+HT8/jXKat4j1PWIBBfMDGDmsvHpTUe4jvj8TvTTf/Hqafie3/QMH/fVcHikxT5UFzWn8UXX/AAkja3aRmFmxviByGFet6HrFtrWnR3dsw5HzL/dNeGYrX8Na/ceHrxpYx5kEg/eRdiaHELntckixoXdgqKMlicACvL/G/i/+0S2maW5FuDiWUfxGsvxD4t1HXMxAm2tP+ean71YCqAMAcUlEBAoAwOBTgKXFLirA7D4ca7Fp91Jpl24SOdtyMegPpXqSkEAg5B7jpXz2VB+o6Edq2dP8V67p0Yjju2ljXordqhxC57bS5ryNfiJrQHMCH8aU/ETWSOLdB+NTysZ60zBVJJwB3PSvKviFr0ep3qafaPvggOWYdC1ZGo+LNc1KNopboxRN95F71jqABgVUYiHGgdKKWtAEopaSgBKKWigBuSCCpwR0PpXZeFfHMtkyWWssZICcLP3WuONNIzwalq4Hv8E8VxCssEivGwyGU5qDU9Qt9LsZLu7kCRoM+5PtXkfhjxNd6BcbGYy2TdYyfu1H4m8Q3XiG7y+Y7RP9XEO9TygVNc1efXNTe8nyEziJOyiqQoxSgVaQCit/wRqH9neJotxxFMNrfU1g06JzFPFMvBjcN+VDVwPoGivOl+Je1ADYAsByc0x/iZKR+705c+7VnysZ6RWfrGrWmk2b3N3IqhRwueWP0rzW8+IOs3GRBEsHuDmucvr281KbztQuWnftntTUWA/V9Sl1fVJb6bgMf3a/3RVQUuKUCtEhAKWiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQIKKKKBhRRRQAUUUUwFqWL7tQ1NF92kBUk/1jU2nSf6xqbTAKKKKYBRRRQAUUUUAFFFFIAooopgHatrR/8Aj1asXtWzo5/0ZqTEU7j/AI+H+tR1Jcf8fMn1qOgYUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUtJRQAGkjklt5lmt5WjlXowPSlppFDA7zw74/wAFLXWxjsJx3+tdN4i8QWum6FJeJKjmRdsYU5JJ7142QDwelKS7KqNIzIv3VJ4FZuADWd5ZHlmO6SQ5Y+tRyj9030qXFI65Uj1qrAe0+EOPDFj/ANcxW5Xl+lePxp2m29mbEN5Kbd2etXP+Fmr/AM+H61k4sZ1nizB8NXwPIMZrgPAXiltMaLTr9ybWTiNz/CfSrOqeP11HTZ7T7EF81duc9K4gIPLCnsKqMNNQPoRWDKGUggjII71x3jfxWumwtp9i4a8kGCR/yzHr9a5TTvGmpWOitp+PMkAxHMTygrnHZ5ZGlmcySucsx7mhQ11AZhiSzsWdjlmPc0uKUClrWwgFGKWigC3pGqXOjagt5aE8H94nZxXsWh6zaa1Yrc2rg5HzJ3U14jVrStTu9IvRc2UhU/xp2aolG4HrPizxBFoWmNJkNcyDESeprxyaaa5ne4uHLyyHLMf5Vb1rVrnW9RN3c/LxhY+y1SFEY2ASrOmanfaTcCbT52Q/xLnIaoMU0iqauB6j4e8dWepFbe+Atro8Afwt+NUfiB4mWOD+ybCQGWTmV1Odo9K87Kg9f0oC88kknuTmp5NQEAAGBVvSdRn0jU4r63+8nDD1FVsUYqrAez6H4k07WoA9vMqSY+aNzjB9q2x6jmvn5MxyCSN2Rx0IOK1rfxPr9soWPUnKjoCKzcGM9szRmvHV8aa+o5uc/hTh4317/nrU8jA9gzRmvH/+E417/nrSf8Jxr3/PWnyMD2DPsaRmAGScD1NeOv428QEcXG38KpXXiTXbxCk+ovsPVQKORgdv478VQ21m+mWMgkuJRiQqchR9a8yUYGM5Pr60oXkkkknqSc07bWkY2A6TwR4ij0K6mivGYWkg4xztNaGtfEO6uN0OkxeUmcCbuR9K4vaPSlxRyK4h88811MZrqZpZD1JNIDikApwq0rAOpKKKYC0UlLQAUUUUCEpaKKACiiigAoopKBhRRS0AJSikpaACkoooAWiiigAopcUlAEsP3TVaX/WGrMP3TVaU/vDQgG0UUUwCiiigAooooAKKKKACiiigAooooAtaZ/x+fhVvU/8AWrVTTTi8H0q3qg/eKaXUCnSUtFACUUtJQAUUUUAFFLRQAlLSUtACUUUhoA774Xsq29+WZR846mu882P/AJ6J/wB9V4NHPcwZFvcNFu6470pvtRH/AC/yfnWMoXYz3jzY/wDnon/fVHmx/wDPRPzrwf7fqP8Az/yfnS/b9R/6CEn50vZsD3fzY/8Anon50ebH/wA9E/76rwj7fqP/AD/yfnR9v1H/AJ/5Pzo9mxHu/mx/89E/76o82P8A56J/31XhH2/Uf+f+T86Pt+o/8/8AJ+dHsxnu/mx/89E/76o82P8A56J/31XhH2/Uf+f6T86Pt+o/8/0n50ezYHu/mx/89E/76o82P/non/fVeEfb9R/5/pPzo+36j/z/AEn50ezA9382P/non/fVHmp/z0T868I/tDUf+f8Ak/Ol/tDUv+f+Sj2YHu3mJ/fT86PMT++n514V/aWpf8/8lL/aWpf8/wDJR7Nge6eYn99fzrmfiAynwzLhlPI6GvMhqepf9BCSkkvb2aMxz3byRnqppqm0wIh91foKWkFLW4Etn/yELb/roP517pH9xf8AdFeEQyeTcxTYzsYNj1ruB8SEUAfYe3rWNSLb0A9APavJviN/yMkf0rZPxKT/AJ8f1rk/EerjXNSW7EXl7RjbShFpgdZ8KwRb35x/y0Fd9+BrwzT9V1HSw40+6MIc5YDvVv8A4SrxD/0En/KhwbYHf+LvFcvh26t4o7MTiVS2ScYrAHxMuf8AoGJ/31XJajqN9qkiSahcGZkGFJHSqoWqUO4Hcj4l3HfTV/76rO8Q+MX13TfsbWYh+YNuBzXMhadiqUEA6Hm6t/8AfFe6WH/HjB/1zH8q8KQlJEcdUORXZQ/ES5igjjGnodigZ3damcW9gPS81yt5480ezvJbWYyeZGcNhawf+FkXWf8AkHIf+BVx99cG9vpbpkCmVixX0qFTfUD1TS/GukapfJZ2zP5r8AFa6SvCtMvZNM1CO9hQM8ZyF9a6j/hY2o5/48UH40Om+gHpb/cb6GvCbs/8TC6/67N/OuqPxF1Egj7CnIx1rkJHMs0kpGDIxYj0zV04tALSGgUuK1AjYZrS0bX9S0RsWkpaHPMRPBqhikIpNXA9DsfiRZuANQtzA3fbzWxB420Gb7t2R9VxXkRUU1o1PUVm6aA9qTxNo79LxPzqQeINKP8Ay+R/nXh3koex/OjyU/2v++qXsgPcv7f0r/n9j/76o/t7Sv8An8j/AO+q8N8hP9r/AL6o8hP9r/vqj2QHuX9vaV/z+x/99Uf29pX/AD+R/wDfVeG+Qn+1/wB9UvkJ/tf99UezA9w/t7Sv+fyP/vqj+39K/wCfyP8AOvD/ACE/2v8AvqjyE/2v++qPZAe4/wDCQaV/z+R/nSf8JBpX/P5H+deIeSno3/fVHkp6N/31R7ID2/8A4SDSv+fyP86P+Eh0n/n8j/OvEPKT/a/76o8pf9r/AL6o9kB3vxF1GzvrO0W1mWRlkyQD0FcRTFQL0z+Jp4rWMbIBaKXtSVYC0UUUCCijNFABRRRQAUUUUAFTr/q/wqCph/q/woGU+5oo7migQUUUUwClpKWgAooooGFFFFABRRRQIKKKKBi0UUUxBRSUUDKVFFFcogooooEFFFFABRRRQAUUUUDCiiigAooooAKKKKAFU7XVvQ1v3Y32Of8AZBrnzyK3rNxcWAB9MUgMrtRSsCrFT2NApgFFFJQIWiikoAKKKKACiiigYtFJS0AFFFFABRRRQAUUUUAFFFFABSUtFADaWlpKACloooAKMUUUAIRSYp1FMBmKUClopAFFFFACYpMU7FFMBmKXFOooAbilFLS0AFFFFABRRRQAUUtJQAUhFLRQA3FGKdRQAlLRRQIDRRRQMQ0lOoxQISgUuKKBhS0UUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigQUUUUDCiiigAooopgFTJwlRDk1JKdkRpAVG5YmkoopgFFFFABRRRTAKKKKACiiikAUUUUAFaujNmORfQ1lVc0qTy7vaejUMCS9XbdN71BV7U0wVkH41RpIAooooAKKKKYBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUCCkpaSmAYpKdSUhhikIpaKAG4oxTqKLANpRS4ooAKKWigApaKKAFpKKKAEpKdSUANxSgUuKKACjFFLQA3FLiiloAQCjFLRQAmKTFOooAbijFOxSUwG4oxTqKAG4o206jFFgGgUuKXFLQA3FLiilpgIKWkpaACijNFAC0UlFAhaKKKACikooAKWkooGFFFFAC0lFLQAlFFFABRRS0AFLSUUAFLSUUATJwlU2OWJq3IdkRqpQgCiiimAUUUUAFFFFABRRRQAUUUUAFFFFAE1k228Q+pxWlqi5RW96yEbZIr/AN05rcnUTWe4c/LmpYGXSUClpgJS0lFAxaKSigQUUUtABRRRQAU3FLRQA0ik206igBmBS4p1GKLDG4oxT6KLCGYoxTqXFFhjMUYp+KMUAMxRg0/FGKdgGYoxT8UYpWAZijFPxRigBoFKKWimAYpTSUUCENJTqTFAxuKMU7FGKLCG4oxT8UmKLDGhaUCnUUAIBS0UUAJRilooAbilxS0YoATFJinYoosA3FLilxRTAKWiigAxSEUtFADcUhFPooAZilxTqKLAMxS4p1JigBMUYpaKLAJijFLiiiwCYoxTsUmKLANxRT6MUwG4pRRRQAUtFFAhKXtRRQAUUlLQMKKKKYgooopDCpm4iP0qJRlhS3DYTHrQBWHSloopiCiiigYUtFFAgooooAKKKKACiiigAooooAWiiimAUUUUDKNFFFcogooooAKKKKBBRRRQMKKKKACiiigAooooAKKKKAFrQ0ibbI0Td+lZ1ORijq69VNAGnqMOyXzAPlbg1UrVjdLy1+owfY1mSxNDIUb8KQDaKSlpgFFJRQAUUUUAFLSUUALRSUtABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKBRRQAUlLRQAUUUUAFJS0UAFFFFABRRRQAUtJRQAtFFFABRRSUAFFFFABRRRQAUUUUAFLSUUwFooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKUDJxTAdEMnNR3LZbaKmYiOOqZJJyepoASloooAKKKKYBRRRQAUUUUgCiiigAooooAKcrFHVh2NNooA6Btt1aZH8QzWSQVJB6ip9KudjGFzw3T2qa/tyD5yD/AHhSAo0UUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgFFFFIAooooAKKKKACiiigApKWkpgFLSUUALRSUtABS0lFIBaKSloAKSlpKAClopKAFopKWgAooopgFFFFABRRRQAUUlFAC0lAooAWikFLTAKKKKADFFFFABRRRQAUUUUALmikooAWikooAWikooAKKKWgBKKKKACiiigAooooAKWkooAWiiigAp8Yy2aaAScCpGYRpQBFcPkhR2qGgkk5PeimAUUUUAFFFFABRRRQAUUUUAFFFFABSUtFACda2NMmEkBjbqv8AKsipbaYwThx070mBYuYzFOR2PIqKtS4jW6hDJ16issggkHgjrQMKWkpaBBSUUtAwooooEFFFJQAtJRRQMKKKKYBRRRQAUUUUAFFFFABRQKKACiiigQUtJRQAUUUtABRRSUALRSUUAFLSUUDFopKM0xBS0UUDCkpaKAEooooAKKKKAFopKKBC0UUUDCiikoAWikozQAtFFFABRRRQAUUUUAFFBopgFFFFABRRRQAUUUUCCg0UUAFFJS0AFFFFABRRSUALRSUUALRRRTAKKKcg3H2oGOjXAzUEzbn9hU0z7FwOpqrSAWiiimAUUUtABRRRQIKKSloAKKKKACiiigAooooAKWkpaYwooooAo0UUVyiCiiigAooooAKKKKACiiigApaSloASilooASiiigApRSUtAE9nctbS56ofvCteWOO7hDKc+h9KwasWl29s3qh6igB8sbRPtcY96bWqrwXkXY/zFVJ7J0yY/mX9aAKlFBBB+YEfWjrQAUUUUAAooooEFLRRQMKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAtFJS96ACkoooAKKKKACiiigAooooAKKKKYC0lLRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUlLRQAlFLRQAUUY9KcqE9aYCAEnipQBGuTSFljHNVpJDIfagAkkMje3amUUUAFFFFMAooooAKKKKACiiigAooooAKKKKACiiigAyQQQcEdK2rG7W4i2PjeBgg96xaVWZGDIcEUmBo3VqYyXjGU7j0qrV+0v0mGyXCv+hp1xZK/zRfKfTtSAzqWnPFJGcMp+opmRTAWiiigAoopaAEooooAKKKKACiiigAopKKBBRRRQMKWkopgLRRRQAUUUUAFJS0UAFFFFABRRRQAUlLRQAUUUUAFFFFIBaKKKYBRRRQAUUUUAFFJS0AFFJRQAUUUUAFFFFABRRS0AJS0UUxhRRRQIKKKKACiiigAooooASloooAKM0UUAFFFFABRRRQAUtJRQAUUtJQAUUtJQAUtJS/SgYUoBJ4pQhPWnkrGOaAAAIpJqtI+9vbtRJIXPt6UygQtFFFMAooooAKKKKACiiigAooooAKSlooAKKKKACiiigC7p935TeVIfkPQ+lXbq1Ew3x/e/nWL1q7Z3xixHKcp2PpSaAjIKnDAgjsaK1JIorpNwIz/AHhVGW1li7bh7UAQ0lB46jFFAwoFFFABRS0lMQUUUUDCkpaKACkpcUUAFJS0UAFFFFABRRRQAUUUtAhKM0UUDCiiimAUUlLSAKKKKYBRRRQAUUUUAFFFFABRS0UAJS0lFAC0lFFABRRRQAUtFFACUUUUxC0lFLQAlFLRQMKO9FFABRSUtABRmikoAWkpaSgAooooAKKKKAFopKKBC0UlLQAUlFFABS0lLQAUlLRQAlFLjPSnCP1oGIqlvpT2ZY1zSM6oPeq7MXOTQAMxZsmkoopgFL0pKWgAooooAKKKDQAUUUUCCiiigAooooAKBRRQAtFJS0xhRRRQBRooorlEFFFFABRRRQAUUUUAFFLSUALRRRQAUnelooAKSlooASloooAKKKKAFR3jbdGxU+1aEGqEcTr+IrOpKAN0S2twOqn60hsoH6HH0rDH5VIJpV+7IwoA1/7Oj7OaT+zU/vtWX9quP+erUv2u4/56mgDT/s1P77Uf2an99qzPtdx/z1NH2u4/56mgDU/s1P77Un9nJ/fasz7Xcf8APQ0v2u4/56GgDS/s5P7zUv8AZyf3mrL+13H/AD0NL9ruP+ehoA0v7OT+81H9nJ/easz7Xcf89DS/a7j/AJ6mgDS/s5P7zUf2cn95qzftdx/z1NJ9ruP+ehoA0/7OT+81H9np/easz7Xcf89DR9quP+ehoA0/7PT+81H9np/eas37Xcf89DSfarj/AJ6GgDT/ALPT+81H2BP7xrM+1XH/AD0NH2q4/wCehosBpf2en95qPsCf3mrN+1XH/PQ0farj/noaANL7An940n2BP7xrO+1T/wDPQ0fap/8AnoaYGj9gT+81H2BP7xrO+1T/APPQ0fap/wDnoaANH7An940n2BP7xrP+0z/89DR9qn/56GgDQ+wJ/eNH2FP7xrP+1T/89DR9qn/56GgDQ+wp/eNH2FP7xrP+1T/89DR9pn/56GgDQ+wr/eNH2Ff7xrP+1T/3zR9qn/56GgC/9iX+8aX7En941n/aZv75o+0zf3zQBf8AsS/3jR9jX+8aofaZv75o+0zf3zRYC/8AY1/vGj7Gv941n/aZv75o+0zf3zRYC/8AZF/vGj7Iv941R+0Tf3zR9om/vmgC99kX+8aPsi/3jVH7RL/fNH2ib++aAL32Vf7xpPsq/wB41S+0S/3zR9ol/vmgC79lX+8aPsq+pql583980efN/fNMC79lX1NH2VfU1S+0S/3zR9ol/vmgC59mX1NH2ZfU1TE8v980faJf7xoAufZh6mj7MPU1T8+X+8aPPl/vGgC39nHqaPs49TVTz5f7xo8+X+9QBb8gepo8gepqp58v96jzpP7xoAt+QPU0eQPU1U86T+9R50n96gC35A9TSeSPU1V86T+8aTzpP7xoAt+SPU0eSPU1U82T+9S+dJ/eoAteSPU0eSPU1V86T+9SedJ/eoAteUPU0eSPU1V86T+9R50n96gC15Q9TR5Q9area/8Aeo81/WgCz5a0bFFVfMf+8aQux7miwFoui9x+FRPcdkH41B9aWmAEknLHJooooABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiikAVat76WEAN86+9VaKANqK+t5hhjj2NSG3tpem38KwsA0oZl+6xH0pWA2f7PiPRjR/Zyf32rKFxOOkrUv2u4/56mgDU/s1P77Uf2an99qzPtdx/z1NH2u4/56miwGn/AGan99qP7NT++1Zn2u4/56Gj7Xcf89TQBp/2an99qT+zU/vtWZ9ruP8AnqaPtdx/z1NMDT/s5P77Un9nJ/fas37Xcf8APQ0fa7j/AJ6GgDS/s5P7zUf2cn95qzftVx/z1NH2q4/56GiwGl/Zyf3jR/Zyf3mrN+1T/wDPU0farj/noaANL+z0/vGj+z0/vGs37VP/AM9DR9qn/wCehosBo/2en940f2en941nfap/+eho+1T/APPQ0AaP9np/fNH9np/eNZ32qf8A56Gj7Vcf89DQBo/2en940fYE/vGs77VP/wA9DR9qn/56GgDR+wJ/eNH2BP7xrO+1T/8APQ0fap/+ehoA0PsKf3jR9hT+8az/ALTP/wA9DR9pn/56GgDQ+wp/eNH2FP7xrP8AtM//AD0NH2mf/noaAL/2Ff7xo+xJ/eNUPtM3980faZ/75oAv/Yl/vGj7Ev8AeNZ/2mf++aX7TN/fNMC/9jX+8aPsa/3jVD7RN/fNH2ib++aAL/2Nf7xo+xr/AHjVD7RN/fNH2ibP3zQBe+xr/eNH2Nf7xqj9om/vmj7RL/fNAF77Iv8AeNJ9kX+8apfaJf75o+0S/wB80WAu/ZF/vGj7Iv8AeNUvtEv980nny/3zQBe+yL6mj7Kvqao/aJv75o+0S/3zQBe+yr/eNJ9lX1NU/Pl/vmjz5f75oAufZV9TR9mHqap+fL/fNHny/wB80AXPsw9TR9mHqap+fL/eNHny/wB40AXPsy+po+zj1NU/Pl/vGjz5f75oAt/Zl9TR9nHqap+dL/eNL50v940wLf2cepo+zj1NVPOk/vGk86X+8aALn2ceppPs4/vGqnnSf3jR50n940AW/s49TR9nHqaq+dJ/eNHnSf3jQBa8geppPIHqaq+dJ/eNHnSf3jQMteQPU0eQPU1V82T+8aXzZP7xoAs+SPU0eSPU1W82T+9SebJ/eNAFryR6mjyR6mq3myf3qPNk/vUCLPlD1NHlD1NVvNk/vUebJ60AWfKHqaPKFVvNk/vUnmOf4qALWxR6UFkTuKqFmPVjSUATPcZ4QVESSck5pKKYXCloooAKKKKACiikoAWikpaACkpaKACiiigBKKWigAooooAKKKKACiiigCSGeWE5jY49Kvw6mh4lG0+1ZdLSsBuA20wyNh+tNNlC3IOPpWKODxxTxNMvSVqLAa39nx/3mpP7OT++1Zn2q4/56tS/a7j/AJ6GiwGn/Zyf32o/s5P77Vmfa7j/AJ6Gj7Xcf89DRYZpf2cn99qP7OT++1Zv2q4/56Gj7Xcf89DRYDS/s5P77Uf2cn99qzftVx/z0NH2q4/56GnYDS/s9P75o/s9P75rN+1XH/PQ0farj/noaLAaX9np/fNJ9gT+8azvtVx/z0NH2qf/AJ6GiwGj/Z6f3jR/Z6f3jWd9qn/56Gj7Vcf89DSsI0fsCf3mo+wJ/eNZ32qf/noaPtM//PQ07DNH7An940fYV/vGs77TP/z0NH2mf/noaLCND7Cv940fYE/vGs/7TP8A89DR9pn/AOehoA0PsK/3jR9hT+8az/tM/wDz0NH2mf8A56GgZofYU/vGj7En941n/aZ/+eho+0z/APPQ0WAv/Yk/vGj7En941Q+0Tf3zR9om/vmgC/8AYl/vGk+xp/eNUPtE/wDfNL9om/vmgC99jT+8aPsa/wB41R+0Tf3zR9om/vmmBe+xr/eNH2NP7xqj9om/vmj7RN/fNAF77Iv940n2Rf7xql9ol/vmj7RL/fNIC79kX+8aPsi/3jVL7RL/AHzR583980AXfsi/3jR9lX+8apefL/fNHny/3zTsBd+yr/eNJ9lX1NU/Pl/vmjz5f75osBc+yr/eNH2ZfU1T8+X++aTz5f7xoEXfsy+po+zL6mqXny/3jS+fL/eNAFv7OvqaPsy+pqn58v8AeNHny/3zQMufZ19TR9nX1NU/Ol/vGjzpP7xpgXPs6+ppPs6+pqp50n940edJ/eNIC39nX1NHkL6mqnnSf3qPOk/vUAW/IX1NHkD1NVPOk/vGjzpP7xpgWvIHqaPIHqaq+bJ/eo82T+9QBa8gepo8kepqt5sn96jzZP71AFnyR6mjyR6mq3myf3qTzZP71AFryR6mjyR6mqvmyf3qPNk9aALPlD1NHkj1NVvNk/vUvmv60AWPKHqaPKHrVbzX/vUeY/8AeoAseUPWjyx6mq/mP60eY/rQBY8setG1R6VX3v8A3jTSSepoAsmRFqJ5mPA4FR0UwDvzRRRQAUUUtACUUtFABRRRQAUUUUAFFFFMQUUUUgCiiigAoopaACiiigAopKKYylRRRXKIKKKKACiiigAooooAKKKKAFopKWgA70UlLQAUUGigAooooAKKKKACkpaKADFJS0UwEoFLQKACiiigQUUUUDCiiigAooooAKKKKACiiloAKKSigAopaKAEooooAKKKWgBKKKKYBRRRSAKKKKACilooASlpKWmAlLRRQAUUUlAC0lFLQAUUtFABSUtFABRRRQAUUUUAFJS0UwEpaKKAAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABS0UUAFFFFABRRRQAUUUUAFFFFMAooooAKKKKACiiigAooooAKKKKACiiigAooooAKWkopAFFFLQAUUCimAUUUUgCiiigAooopgFFFFIAooopgFFFLQAlLRRQAUUlLQAUUUUAJRRS0AFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKWkoAWikpaACikooAWikooAWiiigAooooAKKKKBhRRRmmAUUUtAgooopDCkpaKAEpaKSmAtFJSigQUUUUAFFFFABRRRQAUtJRQAtFFFABRRSUALRRRQAlFLRQAUUUUAFFFFABRRRQAUUUUAFFFFABSUtFACUtFFABRRRTAKKKKACiiigYUUtJQAtJS0lAC0lFFAC0UlFAC0UUUAFFFFABRRRQIKKKKBhRRRQAUUUUAFFFFMAooozQAUUUUALRRRQAUUUlAC0UUlAhaKKKACiiigAooooAKKKKACkoooGLRRRQAUUUUAFFLRQAlFLRQAUUUUAFFFFAgooooAKKKKBhRRRQAUUUUAFFFFMAooooAKKKKACiiigApaSloAKKKKACiiigAooooEFFFFABRRRQAUUUUDClpKWgAooopgFFFFAFGiiiuUQUUUUAFFFFABRRRQAUUUUALSUUUAFLSUUALRRRQAUUUUAFFFFABRRRQAUUUlMBaKKSgBaO9FFABS0lLSASilooASiiimAUUUUAFFFFABRRRQAUUUUAFLSUUAFFFFAC0lLSUAFFLSUAFLRRQAlLSUtACUtJS0wEopaSgAoopaACiiigBaKKKACiiigAooooAKKKKACkpaKADFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigAoxRRQAUUUtABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMAooooAKKKKACiiigAooooAKKKKACiiigApaKKQBRRRQAUUUUAFFFFABRRRQAUUUUwCiiigAopaKACiiigAooooAKKKKACkpaKQCUUtFMAooooAKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFFABRRRQAUUtFABRRRQAUUUUAJRS0UAJRS0UAFFFFABRRRQAUUtFACGiiigAooooGFAopaBBRSUtAwooooEFFFFMYlLRRQIKKKKACiiigYUUUUCCloooAKKKKAEpaKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFAwooooAKKKKACilooASilNFAAKKKKBBRRRQAUUUUAFFFFAwooooAKKKKYBRRRQIKKKKBhS0lLQAUUUUAFFFFABRRRQIKKKKACiiigAooooAKKKSgYUUtFABRRRQAUUUtABRRRQIKKKKACiiigAooooAKKKKACiiigYUUUUAFFFFMAooooAKKKKACiiigQUUUUDCloooAKKKKACiiigAooooEFFFFABRRRQAUUUUDClpKWmAUUUUAFFJRQBSooorlEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUALRRRQAUUUUAFFFFABRRRQAUUGimAUUUUAFLSUtIBKKKWgAopKWgApKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFLRQAUUUUAJS0UUAFFFFABRRRQAUlLSUwFpKWkoAKWiigAooooAWkopRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKWgUUAFFFFABRRRQAUUUUAFFFFABRRRQIKKKKYwooooAKKKKACiiigAooooAKKKKAClpKKQC0UUUAFFFFABRRRQAUUUUAFFFFMAooopAFFFFMBaKSloAKKKKACiiigAooooAKSlooGFFFFAgooooAKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFFABRS0lAC0UUUAFFFFABRRxRQAUlLRigQlLRRQMKKKKAFooozQAUUUUAFJS0UAJRRRTAKWkpaQBRRRQAUUcUUAFFFFMAooooAKKKKACiiigYUUUtAhKWiigAooooASloooAKKKKACiikoAWiikoAWiiigAooooAKKKKACiikoAKWiigAooopgFFFFABRRRQMKKKWkISloopgJS0lLQMKKSloASloooEFFFFABRRRQAUUUUAFFFFAwooooAKKKKYBRRRQIKKKWgYUUUUAFFFFABRRRQIKKKKACiiigAooooAKSlooAKKKKACiiigYUUUtACUtAooEFFFFABRRRQMKKKKBBRRRQAUUUUAFFFFAwooooAKKKKYBRRRQIKKKKACiigUALSUtFAwooooAKKKKACiiigAooooEFFFFABRRRQAUUUUwClpKKAFooooAOaKKKBlGiilrlEJRRRQAUUUUAFFFFAC0lFFABRRRQAUUUUAFLSUtABRRRQAUUUUAFFFFABRRRTAKKKKBBS0UUgEpaKKBhSUtJQAUUUUwCiiigAooooAKKKKACiiigAooooAKWiigAooooAKKKKBBRRRQMSilpKAClpKUUwEpaKKACiiikAUtJRTAWiiigAooooAKKKKACiiigAooooAKKKKAFpKWkpgFFFFABRRRQAUUUUAFFFFABRRRQAUUtJQAtFFFABRRRQIKKKKACiiigAooooGFFFFABRRRTAKKKKACiiigAooooAKKKKQBRRRQAUtFFAgooooAKKKKACiiigYUUUUAFFFFABRRRQIKKKBTGFLRRSAKKKKACiiimAUUUUgCiiimAUUUUAFFFFABRRRQAUUUUAFJS0lAC0UUUwCgUUUAFLSUtABRSUtABSUtFABRRRQIKKKKBhRRRQAUUUUAFFFFABRRR2oAWikpaACiiigAooooAKKKKAAdaKBRQAUUUUAFFFFABRRSUwFooooAKKKKACiiigYUUCigQtFJS0AFFFFABRRRQAUlLSUAFLSUUCCiiloGFFFFABRRRQAUUUUAFFFFABRRRQMKKKKYgo7UUtACUUtFIYUUUUCCg0UUxhSUUtIBKWgUUwCiigUCCiiigAooooAKKKKACiiigYUUUUCCiiimAUUUUgCiiimMKWiikAUUUUwCiiigQUUUUAFFFFABRRSUALRRRQAUUUUAFFFFABRRS0AFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQMKKKKYhaSiigYUUUUCCiiigBaKKKACiiigYUUUUAFFFFAgooopgFFFFIAooopgFFFFABRRS0AFFFFABRQRRQBRpaSiuUAooooAKKKKACiiigAooooAKKKKACiiigApaSloAKKKKACiikoAWiiigApKKKYC0CiikAUUtFABRRRQAlFLSUAFFFFMAooooAKKKWgBKKWkoAKKWigBKKKWgAooooASloooEFFFFACUtFJQMWkoooAWikooAWiiimAUUUUgCjtRRQAtFFFMAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUtACUUUZoAKKKWgAooooAKKKKBBRRRQMKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFFABRRRSAKKKO1AC0UUUAFFFFABRRRTASloopAFFFFABRRRQAUUUUAFLSUUwFooooAKKKKQBRRRQAUUUlMBaKKKACiiigAooooAKKSloAKKSimAtFFFIAooopgFFFLQAUUUUAFFJS0AFJS0UAFFHFJQIKKKWgYlLRRQAUUUUAFFFFABS0lFABS0lLQAUlFLQAUUUUAFJRRTAWiiikAUUUUAFFFFMBKWiigApKWigAooooAKKKKAClpKKYBS0lLSAO1FFFABRRSUAFFLRQAlFFFMAooopAFFFFAC0UlFAC0lLSUALRRRQAUUUUwCiiigAoopaBhSUtFIAooopiEopaKAAUUUUDCiiigAooooAKKKKBBRRRQAUUUUAFFFFABRRRQAUUUUxhRRRSEFFFLQAUUUUxhRRRQAUUUUCCiikoAWikooAWiiigYUUUUCCiiloASiiloAKKSloAKKKKACiiigAooooAKKKKACiiigAooooGFFFFAgooooAKKKKYBRRRQAUUUUAFLRRQMKKKKACiiigAoooNABRRRQIKO9FFAwooooEFFFFMAooooGFLSUtAgooooAKKKKBlKkoorlEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLSUtABRRRQAUlFFMAzS0lFIAoopCwHemAtLTdy+v6Uu4e/5UgFopNw9/wAqNy+p/KgBaWm7h6/pRuH+RQA6kpNw/wAijcPX9KAFopNw9f0o3D/IpgLRSbh/kUu4f5FAhaSk3D/Io3D/ACKAHUUm4ev6Um4ev6UALRSbh6/pRuHr+lAxaKTcP8il3D1/SgAoo3D/ACKNw/yKAFpKTcP8ijcP8igBaWm5Hv8AlRuH+RQA6kpMj/IoyP8AIoAWikyP8ilyP8igBaSjI9/ypMj/ACKAFpaTI/yKMj/IpgLmikyP8ijI/wAikAtFJkf5FG4f5FADqKTcP8ijI/yKYC0UmR/kUZFAC0UmRRke9AC0UmRRkUALRSZFGRQAtFJkUuRQAUUZFJkUALRSZFGaAFooyKM0AFFGRRkUAFFGaM0wClpMijNABS0maM0ALmikzRmgBaKTIozQAtFJkUZHrQAtFFFABRRRTAKKKKQBRRRTAKKKKACiiigAooopAFFFFABRRRQAtJRRTAKWkooAKWiikAUUooOB1oASgUm5fWl3D1/SgBaSjI9f0oyPX9KACik3D/Io3D/IoBC0Um4f5FG4f5FADqKbuHr+lLuHr+lAC0Um5fX9KNw9f0oAWjNJuHr+lG5f8igBe9J3o3L6/pRuHr+lMBaKTI9f0o3D/IoAWim7h6/pS7h6/pQAtJSZHr+lLuH+RQAtFJuHr+lG4ev6UALRSZH+RRuH+RRcBaKTcP8AIo3D/IoAWik3D/IoyP8AIoAXtRRuHr+lGR/kUAFLSbh/kUmR/kUwHUlG4f5FJuH+RQAtLSbh/kUZFABRSZHr+lGR/kUALRSZH+RS5FAC0UmRRkf5FAC0Um4f5FG4f5FAC0UmRRkUALRSZFLkUAFLSZFGaACijNGRQAUUZFGaACgUZoyKAFopMijNABRRn60ZFMAoozRQMKKM0UCClpKXNABRRmigAoooyKACijI9aKACiiigAoooFAC0UUUAJRS0lABRRRTAKKKKAFopKWkAlFLRQAUUUUAFFFFABRRSUALRSUtMApaSjcvrSGFLTdy+v6Ubh/kUCHUUm4f5FJuX1/SncB1FJuH+RRuHr+lAwoo3D1/Sk3D1/SgBaWm7h/kUu4f5FAC0Um4ev6Ubh6/pQAtFJuH+RRuHr+lAC0Um4ev6Ubh6/pRcBaKTcP8AIo3D1/SgQtFJuH+RRuH+RQAtFJuH+RS7h6/pQMKKNw/yKNw9f0oEFFGR6/pRuX1/SgBaKTcP8ijI/wAigYtFJuH+RRkf5FAC0Um4f5FG4f5FAC0UmRRuH+RQIKWk3CjcKLgFLSZHrRkUwFopMj1oyPWkAtFJkUZH+RTAWikyKXIoAKKMijIoAKKMijIoAWikyKMigBaKTIoyKAFopMijIoGLRSZFGRQAtFJmjNAC0UmaM0ALRSZozQAtFJkUuaYgopM0uRSAKKM0ZFAwooyKMj1oELSUtJTGLRSUtABRRRQAUUUUAFFFFABRRSUALRRRTEFFFFABRRRQAUtJS0AFFFFAAaKKKBlGiiiuUQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUtABRRRQAlFFSQQS3Moit0Luew7UARkgcmr1lpN7ekGOMoh/jPSuh0vw7DaL596RJIOfYVPea1Bb/u7ZdxH93pSuBUtfCcKgG6lLn/ZNX00rSLcchcj+9WHcaleXBJaTaPReKrFnb70jH6mgDqfL0df+eH5UY0j/AKYflXK4FGBSA6rbpH/TD8qNuj/9MPyrlcUYFAHVY0f/AKYflSbdI/6Y/lXLYowKYHU7dI/6YflRt0j/AKY/lXLYowKQHU40j/ph+VGNI/6Y/lXLYFJimB1WNI/6Y/lSY0j/AKY/lXLYoxQB1ONI/wCmP5UY0j/pj+VctijFAHU7dI/6Y/lRt0j/AKY/lXLYowKAOo26R/0w/KjGk/8ATH8q5fFGBQB1GNJ/6Y/lSY0n/pj+VcxgUmKAOpxpP/TH8qQjSf8Apj+VcvgUYFAHUY0n/pj+VJjSf+mP5VzGBRgUAdPjSf8Apj+VGNK/6Y/lXMYFJgUAdPjSf+mP5UY0kf8APH8q5jFGKAOnxpP/AEx/KkxpX/TH8q5nAowKAOmxpX/TH8qT/iV/9MfyrmsCjAoA6X/iV/8ATH8qT/iV/wDTH8q5vAowKAOk/wCJX/0x/Kk/4ln/AEx/KubxS0wOj/4ln/TH8qMaZ/0x/KucwKMUgOixpn/TL8qP+Jb/ANMvyrnMUtMDof8AiW/9Mvyo/wCJb/0y/KudxRigDof+Jb/0y/Kj/iW/9MvyrnsCjAoA6DOm/wDTL8qP+Jd/0y/KufwKMCgDf/4l3/TL8qTOnf8ATL8qwcUYFAG8Tp//AEy/Kkzp/wD0y/KsHFFAG9mw/wCmX5UmdP8A+mX5Vg4FGKAN4mw/6ZflSZsP+mX5Vh4FJTA3c2H/AEy/KjNh/wBMvyrCxRigDdzYf9MvypM2P/TL8qw8UYoA3M2P/TP8qT/Qf+mf5ViUUAbebH/pn+VJmx/6Z/lWJRigDbzY/wDTP8qAbH/pn+VYnFFAG1my/wCmf5UZsv8Apn+VYtJQBt7rL/pn+VJusv8Apn+VY2BRgUAbH+hH/nn+VIbazl6BfwrIxSgkfdYigC/Lo8TAmJip96z57GeDkruX1FTxXU8XRs/Wr0OoRyfLKMH36UXYGDxRW3d6fFON8OFf9DWNJG8UhSQEGquA2iiigAooopgFFFFABRRRQAUUUUAFFFLSAKSlpKYBRRRSAKKSgmgBafDHJPJshjZ29BV/SdFm1AiST93B6n+L6V1CJY6TCAoVcev3jSuBh2Xhm4mw1xII1/u9614vDumwAGXcT6k1TuddmkJW3Xavq3Ws+S4mkOXlf86V2B0YtNHjGP3f4ineXo4/54/lXLZz1JNFIDqduj/9Mfyo26R/0x/KuWxSYoA6jbpH/TH8qNukf9Mfyrl8UYoA6jbpH/TH8qTbpP8A0x/KuXxRimB1G3SP+mP5UbdJ/wCmP5Vy2KXFIDqNuk+kP5UY0j/pj+VcvijApgdRt0j/AKY/lSbdJ/6Y/lXL0YoA6jbpP/TH8qNuk/8ATH8q5fAoxQB1G3Sf+mP5Um3Sf+mP5VzGBS4oA6bbpP8A0x/Kk26T/wBMfyrmcCjAoA6bbpX/AEx/KjbpX/TH8q5nAoxQB0u3Sv8Apj+VG3Sv+mP5VzWKMUWGdLjSv+mP5UbdK/6Y/lXNYFGBTsI6TGlf9Mfyoxpf/TH8q5vAowKLAdJjS/8Apj+VGNL/AOmP5VzdJiiwHSY0v/pj+VGNL/6Y/lXOUYFAHRY0z/pj+VGNM/6Y/lXOYoxQB0eNM/6Y/lSY0z/pl+Vc7QAM0AdFt0z/AKZflRjTP+mX5Vz2KTFAHRY03/pl+VJjTf8Apl+Vc9ijigZ0ONN/6ZflSY03/pl+Vc/ikxQB0ONN/wCmX5Un/Et/6ZflXP4oxTA6D/iXf9MvypCNO/6ZflXP0tIRv/8AEu/6ZflSf8S7/pl+VYGKXFMZvf8AEv8A+mX5Uf8AEv8A+mX5Vg0UCN7/AIl//TL8qQ/YP+mX5VhUUAbv+gf9MvypP9A/6ZflWHSUAbn+gf8ATL8qP9A/6ZflWHRQBt/6D/0z/Kj/AEH/AKZ/lWJRigZtZsf+mf5Uf6D/ANM/yrExRgUAbX+hf9M/yo/0L/pn+VYuKMU7AbX+hf8ATP8AKkzZf9M/yrGxRgUCNnNl/wBM/wAqP9C/6Z/lWPxSUAbP+hf9M/yozZf9M/yrGoxRYDZzZf8ATP8AKkzZ/wDTP8qx6KANjNn/ANM/ypCbI/8APP8AKsiigDVa2s5OgX8Kry6TEeYmKn3qjkjoxFSx3U8XRsj3oAhuLKeA5K7l9RVbrW3DqCSfLKMH36U26sI5xviwr+3Q07jMeilkR4nKSAgikpiDvS0lLQAUlLRTGJRRRQIKWkpaBhRRRQAUUUlAC0UUlIQtFJRQAtFFWLGynv5vLgXju3YUMCt3AHJPQVp2Wh3t1hmHkoe7d66Cx0m002PzJNrP3dulRXWthMpbLn3PSp5hjbfwzaRjNwzOfUHira6fpMPBCf8AAqwpr25mOXkI9lqElj1dj9TS1A6by9JH/PH8qNmk/wDTH8q5j86MUAdPt0n/AKY/lSbNJ/6Y/lXM4oxQB02zSf8Apj+VBXSfSH8q5nFGBQB023Sf+mP5UbdJ/wCmP5VzOBRiiwHTbdJ/6Y/lRt0n/pj+VczijFFgOm26T/0x/Kk26V/0x/KuawKMCmB0u3Sv+mP5Um3Sv+mP5VzeBRgUAdJt0r/pj+VG3Sv+mP5VzWBRgUgOl26V/wBMfyo26V/0x/KuaxRimB0u3Sv+mP5Um3S/+mP5VzfFFFgOk26X/wBMfyo26X/0x/KubxRgUWA6Xbpf/TH8qTbpf/TH8q5vFGKLAdHt0v8A6Y/lRt0v/pj+Vc3ijFFgOkxpn/TL8qTGmf8ATL8q5zFGKLAdGRpnpD+VGNM/6Y/lXOYoosB0WNM/6ZflRjTP+mX5VzuKMUWA6HGm/wDTL8qMab/0y/KuexRinYDoD/Zv/TL8qTGm/wDTL8q5/AoxRYDoP+Jd6RflRjTvSL8q5/FFFgN/Gnf9MvyoI0//AKZflWBRRYDe/wCJf/0y/Kj/AIl/pF+VYNFFgN7/AIl//TL8qT/iX/8ATL8qwqMUWA3f9A/6ZflR/wAS/wD6ZflWFikosBu/6B6RflR/oHpF+VYVFFgNz/QP+mX5Un+gekX5ViYoosBtn7D/ANM/ypP9B/6Z/lWLijFOwG1/oP8A0z/Kj/Qf+mf5ViUYosBtZsv+mf5UE2X/AEz/ACrGo4osBs5sv+mf5UmbL/pn+VY1FFgNnNl/0z/KkzZ/9M/yrHoosBsZs/8Apn+VGbP/AKZ/lWPijAosBr7rP0j/ACozZ/8ATP8AKsjFFFgNfNn/ANM/yozZ/wDTP8qyKTFFgNfNn/0z/KlzZ/8ATP8AKsfFFOwGvizP/PP8qRrS0k6AfhWTSgsOjMKQFybSUPMLbT71QntZoPvqSP7wqzHdzR98j3q7FexTDbIMH3p3YGGMUVq3emq+ZLfhvT1rLZSrFWBDDsaaYCUUUUwCiiigQUUUlAxaSlpKACloooAKKKKYgooooAWiiigAooooAKKKKYFGiiiuQAooooAKKKKACiiigAooooAKKKKACiiigAooooAKWkoz6ck9qAJbW3lu7hYIAS7H8q7ixsrXRbPcxG/Hzuev4VW8P6cmnWXnz4EzjLE9hWbqd+17OQpIiU8D1pAO1HU5bxiqEpEOgHU1n4A6UtJQAUUUUAFFFFABRS0lAC0lFFABRRRQAUUUUAFFFFACUUGigAzRRRQAUUUUAFFFJTAKKMUUAFFFFABRRRQAUUUUAGaKKKAEopaSgAooooAKKKKACiiigAooooAKSiigAooooAKKKKACkpaSgAooooAKKSlpgJRS0UAJRRRQAUc0UUAFFFFABSUUUwCiiigAooopAFFFFMApaSigAoNGaKAJ7a7eBgCSydxV+4givoMjGexrIqxaXBt5BnlD1FAFGSNopDG4ww/Wm1uajarcwebH99RkH1FYf1600AUUUUwCiiikAUUUUwCiiigAopaKACiiigBKKWigBpNbOg6Ob1xcXIIgU8D+9VTSdPbUbwRjiNeXNdVqN3Hp9qsMAAbGFHpUsBuo6jHZIIYAN44AHRa5+WR5nLysWY0jEsxZjlj1NJUgFFFFMBaKSloAKKKKACiiigAooooAKSlpKACig0UAFFFFABRSUtABRSUtMApKKWgBKWiigAoopKACiiigAooooAKKKKACiiigApKWkoGKKKKKBBRRSUxi0lFFABRSUUALmkoooEFFFFABRRRQMKKKKACikooELRSUUAFFFFMAooooAKSlxRQACkpaSgAoooNABRRRTAKKKKBhRRSUCFopKM0hhSUtJTATFT2108BwfmQ1DRQBqTwRXsORjP8AC1YskbRSGNxhhV21nMEmDyh6irl/bLcweYn31GQR3o2AxKWjHY9RRTEFFFFMYUUUUAFFFFIAooooAKSlpKAFpKKWgQlFFOijeaZIYxlnOBQBY06xl1C5EUfCj7zegrrx9l0mzCoAAPTqxqO2gh0nT8cAgZY9yfSsK5uXupjI54/hHoKh6jJby9lu3Jc4Tsoqt+lGaBQAlFLiigBKWiigAooooAKKKSmAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAlAoopgLRSUUAHNFFFACUtJRQAtFJRQAtFFJQAUUUUAFFFFABRRSUAFFFFMQUUUUAFFFFAwooooAKKKSgBaSiimAUUUUAFFFFABRRRQAUUUUAFFJRQAUUUUAFFFFABmiiigAooooAKKKKAEpaKSgAooooAt2t20RCSHKHv6VYvLRLqPzI8CQDII71mVbsbkxsI3PyHp7UAZpBVirDDDqKStXVbUMvnxjkfeA71k00wFopKWmAUUUUAFFFFABRRRQAUUUUwCiigUCFooooAKKKKACiiimMo0UUVyCCiiigAooooAKKKKACiiigAooooAKKKKACgUGigArV8O2P2zUQ7jMcPJ96yjwM12fhu2FrpAlcYZ8k/SkwG6/ebEFtGcE/e+lYQ4qW6lM91JITnnAqKgAooooAWkoooAKWkzRQAUZoooAKKKKACiiigAoopKACiiigAooooAKDRSUALSUdKKYBRRRQAUUUUgCiiigA6UUUUwCiijNABRR1pKACiiigAooooAKKKKACiikoAWkoooAKKKKACgUlLQAUUlFABmgmiimAlFFFABRRRQAUUUUAFFFFABRRSUAFFFFMAooooAKKDRSAKKKKYBmiiigAopKKACloooAKXqKSjNAGnps+VMLHkdPpWfqdv5FyWH3X5ohkMc6uPXmtDU4vOs946jmkBh0UgORmlqgCiiimAUUUUAFFFFABS0lLSAKKKKYBQeeB1PAoq9o9t9r1SKMj5RyaTA6nR7VNN0ne4w7Dc5rBupmubh5m7ngelbfiCfZAkCnG+sA+1QAyilNJTAKKKKACiiigAoopaACikooAKKKKACiiigAooooAKKKSgBaTNFLQAlLSUtMBKWiigAopKKACiiigAooooAKKKM0AFFJRQMWkpaKAEpaKM8UCCkoopjCikpaACikooAKKKKBBRRRQAUUUUAFFFJQAUUUUAFFFFAwooopgFFFFAgooooAKKKSgBaSiigAooopgFFFJQAtJRRSAWkoopjEpaSloAKSiloEFFFFAwrR02bIMLHkdKzqfC5imVx2NDELqdv5NxuUfI/SqdbmoxiayLDkqMisMdKEAUUcUVQwopKWgAoopKQC0UUUCCikpaACiiigBK6HwvYg77yQc/dXNc+FLsqDqxwK7dVWx0hQOCE/WpkwMrWbozT+Sp+ROv1rOpSdxLHqxzRSGFKKSimIWiiikMKKKSmAtJRRQAUtJS0AFJRRQAUUUUAFFFFABRRRQIKKKKACkpaSmMKKKKACiiigAoopKAFpKKKACiiigAzRRRQAUUUlAC0UlFMAooooAKKKKBBRRRQAUUUUDCkpaSgAooopgFFFFIAooopgHFFFFABRmkzS0AJRRS0AJRRRQAUUUUAFFFJQAtFJRTAWiiigBKWikoAWkoooAKKKKACg9KKSgDWsZRPAUbkjg1kXcP2e4ZP4f4as2Mnl3IB6NU+sxZjWUdV4pdQMmiigVQgooooAKKKKACiiimMKKKKACiiloEFFFFIAooopgFFFFAFGloorlAKSlpKACiiigBe1JRRQAUUUUAFFFFABS0lLQAUlLRQABd7qn94iu8vMW2i7V4/djFcRZjdfwD1au117jTEHsBSYHMDp9eaKOwpaAEooooAKKKKACiiigAooooAKKKKACikooAKKKKACiiigAoopKYC0lHWigAooopAHaiijNABRmjrQKYBRRQTQIKKSigYUUUUAFFFFABRRRQAUUlLQAGkoooAKKSigBaKSigBaSlpKAFpKKSgBaKSimAUtJR3oAKKKKADvRRRQAUUUUAFJRS0AJS0UUAFJRRQAUUUUwCiiigAooooAKKKKACiijNABRSUUAFFFFAAelbMJ82wIP93FY1a+mHNoR70gMEjazL6HFJT5xi5lH+1TKoAooooAKKKKYBRRRQAUUUUgClopKYC10Xg+INPNMR9w4Fc7XW+EFxZTv6mpYFbWpfM1JlzwnSs81Y1A51Cc+9V6kBKKKKYCUtFFABRSUtACUUUUAFFFFABRRRQAUUUUAFFFFMApKWigBKKWikAlLSUUwCiiigAooooAKKKKACikooAKKKWgYlFFLQIKKM0lMYUUGigApKWkoAKKKKBBRSUtABRRRQAUUUUAFFFJQAUUUZoAKKKKBhRRRTAKKKSgBaKSigQtFJRQAUUUUAFFFFMAoopM0gCiiiiwwooopgFFFFAgoopM0DFpKKKAClpKKBC0HpSUtAGvanzbIA9xWCw2uw9Ca29MObbH1rGuRi5ce9CGMopKWqEFFFFIYUUUUxBRRRSAKKSloAKKKKALelR+bqkC44DZrpvEEmy2WMd2rA8PDOrr7CtnxIf3yL+NS9wMfpSUp60lAwooooAM0UUUAFFFFABRRRQAUZoooAKKKKBBRRRQMKKKKBBRRSUwCiiigAooooGFFFFACUtJRQAUUUUALSUUUAFFGaKADikoopgFFFFAgooooGFFFFABRRRQAUUUlABS0UlMAooopAFFFFABRRRTAKKSigAoopaAEpaKSgApaSigAooooAKKKSgQUUUUxhRRRQAtJRRQAUUUUAFFGaTNAC0lFFABS0lFACqcOp9DWvdr51gfcZrHNbUXzWHP9w0mBzo6UtJ3I96WqEFFFFABRRRQMKKKKYBRRRQIKWiikAUUUUAFFFFABRRRTGUqKKK5RBRRRQAlFLSUAFFFFABRRRQAUUUUALRRRQAlLRSUAT2HOo2/+9XaeIf8AkGx1xdj/AMhG3/3q7PxD/wAg6OkwOZ7CijsKKACiiigAoopKAFopKKAFopKKACiiigAooooAKKKSmAtJRRQAUUUUgCiiigAooooAKKKM0wCijNJQIKKKKBhRRRQAUUUUAFFFFABRSZooAM0UUlAC0UUUAJRRS0AJRQaKACiiigBKKWkpgFFFFABRRRQAUUUlAC0UlFABS0UUAFFJS0AFFFJQAUUUUAFFFFMAoopKAClpBS5oAKKQ0UAFFFFABRRRQAUUUUAFa2lf8ep+tZNa2lf8ep+tIDFuf+PqX/eqOpLn/j6l/wB6o8VQBRRRQAUUUUwCilopAJS0UUAFFFFAB2Ndf4R/5Bstch2Ndd4R/wCQdNSYGVef8f031qGpbz/j+m+tRVIBRRRTAKKSigAooooAKKKKACiiigAozRSUAGaUUlLTAKKKKACiiigBKKKKACiiigAooooAKKSigBaSiigAooooAWiikoAWkoopgFFFJ3oAKKKKBhRRRQIKMUUUAJS4oooAKKKSgBaKSigBaSiigAooooAKKKKYBQaSigAooooAKKKKBhmkpaKYgoopKQC0lFFABRRRTGFFFFABRRSUALRmkooAO9FFFABRRRQAUUUUAFFFFAGppZ/cGsi7/wCPuStfSv8AUGsi7/4+3pLcRHRRRVAFFFFABRRRTAMUUUUgCiiigAooooA0/Df/ACFx/u1reJP+PpPpWV4c/wCQuP8AdrV8Sf8AH2n0qXuBk0maO9FAwooooEFFFFAwooopCCiiimAdqKKKACiiimMKKKKQgopKKYBRRRQAUUUUAFFFBoGJRRRQAUUUUAFFFBoAKKSimAUUUUCCiiigYUYopaACjFFJQAUUUUAFJS0UAFJRmimAUUUUAFFFFABRRRQAlLSUtACUtJRQAUUUUAFFFFABRSUUCCiiimAUUUUAFGaKSgYvWiiigQUUUUDCiikoAKKKKACiiigAooooEB6VtW//AB4f8ANYp6VtW/8Ax4f8ANJjOeP3j9aKD1P1oqkIKKKKACiiigAooopjCiiloEFFFFIAooooAKKKKACiiimMpUUUVyiCiiigBKWg0lAC0lFFABRRRQAUtFJQAtFFFAgpKWigZNY/8hG3/wB6uz8Q/wDINjrjLH/kI2/+9XZ+Iv8AkGx0mBzPYUUDoKKACiiigAopOlFABRRRQAUUUUAFFFJQAtJRRTAKKKKQC0lFFABRRRQAUUUUwCikJooAM0UUUAFFFFABRRRQAUUlFAC0lGaSgBc0UlGaAFooo/EfnQAlFBKjvTS6/wC1+VADqKb5i+j/APfNHmL/ALX5UAOzRTQynv8AnS59x+dAC0UmaWgApKKKACiiimAUUUlAC0lFFABRRRQAUtJS0AFFFJQAUUUUAFFFFABRRRQAUUUlMBaKSigApaSigQUUZooGFFFFABRRRQAUUmaKAFrW0r/j1P1rJrW0n/j2P1oYGLc/8fUv+9UdSXP/AB9S/wC9UdMAooopgFFFLSAKKKKYBRRRSASlpKKAF7Guu8I/8g6WuR7Guv8ACP8AyDZqTAyb3/j+m+tQVPe/8f031qCpAWikopgFFFFABRRRQAUUlLQAUUUlAC0lFFABRRRQAUUUUwCiiigAooooAKKKKACkoooGFFFFAgoopaACkoopgFFFFABSUtGKBiUlLmkzQIWkzSUZA7j86AFpabvQd/yo8xP9r8qAHUU3zE9G/wC+aPMT/a/KkA40lN3A9xRn3H50wHUUmaM0ALSUUtABRRiigApKWkpgFFFFABRRRQAUUUUAFJRS0AFJRmigYUUUlAhaKSimMWiikNAC0maKKAA0UUUCCiiigYUUUUAFFFFAgoopKAFooooGaml/6g1kXf8Ax9yVraX/AKg1k3X/AB9yUluIj7UUUlUAtFFFABRRRTAKKKKQBRRRQAUUUUAafhz/AJC4+la3iT/j7T6Vk+HP+QuPpWt4k/4+0+lS9wMiig0UDCiiigQUUUUDCiiigQUUUUAFFFHFABRSZooAKKKKYBRRRSAKWkooAKKKKYwopKKAFpKKWgBKKKKACkoooAKKKXimAlFFFAgooooGLSUGkoAWiiigAopKKACiiigAooopgFFFJQAtJRRQAUtJS0AFFJRQAuKSlpKACikzRQIKKKKYBRRRQMKKKKBBSUUUDCiiloAKKSigBaSiigAooooAKKO9FABRRRQIWkozSUDFNbVv/wAeH/ADWIelbVv/AMeP/ATSYHPnqfrRQfvH60VSEFFFFABRRRTGFFFFAC0lFLSEFFFFABRRRQAUUUUwCiiigClRRRXKAUUUUAFJS0UAHakpaKACkpaKAEopaKACiiigQUnelooGTWP/ACEbf/ersvEP/INjrjbH/kI2/wDvV2fiH/kGx0mBzPYUUg6CigBaSig0AFFFFABRRRQAUfSiigBKWkopgFFFFIAooooAM0UUUAFFJRTAKKKKACiiikAUUUUwCiikoAWkoooAKKKX3oASkqa3t57uYQ2kLSyHoAOPzrrNK+H13cASanN5C94hzmk2kBxZYZwAWP8AsjNadhoGsajg21k2w/xNxXqmmeGdJ0wA21qu8fxNzmtZUCjCgKPYYqXIDzaz+Hd7KA13fCL1UDNbVr8PNLTH2l2mP1xXY0cVN2MwIvBmgxY22Wfq1Wk8OaRGPls0rWzRRdgZ39h6Z/z5x/lTH8PaS4+azT8q1KKVwOfm8HaFN9+y/I4rMuvh3pMnNszQn65rs6Kd2B5nefDm8iBa0vhJ/skYrCvfDWtWALTWRKD+JTmvacUhUEYIBHvT5mB4EeG2srKw6hhijFe1aloGmamMXdqrH1XiuN1b4fSxBpNLn3jtCf8AGqUxHD0lWL2zurCYw3sDROO2OPzqvVAFJRRTAKKKKACiiigAopKWgAooooAKKKKACiikoAKWkopgL2opKKACiiigAooooEFFFFAwooooAKKKKACkpaKACtbSf+PZvrWTWtpX/Hsf96kBjXP/AB9y/wC9UVS3X/H1L/vVFVAFFAooAWiigUAFFFFABRRRQAUUUUAHY11/hH/kHTVyHY113hH/AJB01JgZV7/x/TfWoKmvf+P6b61DSAKKKKACkopaACikpaACikooAKKKKACiiimAUUUUAFFFFABRRRQAUUlFAC0UlFABRRS5oASg0UlAC0UlFMBaKSikAtFGaOAOTTAKMepFXtN0nUNVfbZW7Mvdm4xXZaX8P4E2vqcxmPUoOMVLkkB5+gaRtsSO7H0Wtaz8K63egMtp5cZ/iJr1Wy0qxsY9lrbIg9xk1cAx7fSocxnndr8OZmAN3qGPVQK17bwBo8YHnq0p+uK63FFTzMDDi8I6HF9yzH4mrKaBpSDizT8q080ZpXYGf/Yemf8APpH+VRSeHdJkHzWaVq5ozRdgc9N4L0GUHNoQfUNWVdfDuwfJtZzD9ea7akxT5mB5he+ANTgybW4FwOw6VgXulalYMReWbpjuBmvb8U1o0cYdFYf7QzTU2B4MpB9R9eKfXrep+EtJ1AMzW4jkPR14xXG6t4H1CyBksm+1RjnHQirU0wOVop0iPFIY5UZHHUMMUlWISiikpjCiikoELSUUUAFFFFAC0nFBpKBhS0UUwCiiigAopKKBBR2opO1AxaSlooAKKKKACiikoAWikooAKKKKACiiigQtFFFAGnpf+oNZN1/x9yVraX/qDWVdf8fclJbgRUUlFUAtFFFABRRRTAKKKKQBRRRQAUUUUwNPw7/yFx9K1vEn/H2n0rJ8O/8AIXH0rW8Sf8fafSoe4GT3pKKKBhmiiigQUUUUAFFFFABRRRQAUlFFAC0lFFAC0lFFABRRSU7DFpKWkoAKKKKACijNFABR1opKACiiigAooooAKWkopgLSUUUgCiikpgLSUUUwCiiigAooooAKKKKACikooAKKKKACilooAKKDgDJrpNA8I3eqoLi5Y29ufu5HL0nJIDm/xH50GvRpfAGntFiKVkfH3q43XdCu9DnCT/vIXPySjvUxqJgZVFFJVgFFFFMQUUUUDENLSUUAFFFFABRRSUALSZoooAWkoooAKUUUUAFFFFAgooooAKKSigYtJRS0AJW1b/8AHj/wE1inpW1b/wDHj/wE0mBz5+8frRQfvH60VQBSUtJQAtFJS0AFFFLTEFFFFIAooooAKKKKACiiimAUUUUxlKiiiuQQUUUUAFFFFABRRRQAUUUUAFJS0UCCiiigYUlLSUwJ7L/kIW/+9XZeIf8AkGxVxtl/yELf/ersvEP/ACDYqlgcz2FFHYUUAFBooNAB2pM0UUAFFGKKYC0lFLSASig0UAGaM0UUAJS0UmaYC0lFFABRRRQAUUlLQAlLSUUALRSUUAFFFFABRRirml6XeavdC3sYix/ic9F/Gi4FVQWYIil2PAVRk11+g+Bbm82z6qfJi6iIdWFdV4c8J2eioJHAnuiOZGHSui6Vm5DKWnaXZabCIrOBY1Htk/nV2imSSxxIWkdUUdSxxUgPppOOvFcrq/jrTbLMdp/pUw6qOBXGan4x1m/LLHL9mib+ACmotgep3Wo2dom65uI0H+9WFd+OtCtshbkysOwFeUylpnLzSyOx9WNNCqOij8qrlEehT/Em3UnyLFpPfNVJPiVdE/u9Nx+NcTj6UU+VAdl/wse//wCfAfnU0fxJuR/rNNz/AMCrh6UU+VAehw/Ei3JHn2LR++a1rTx1oVxhWuTGx7EV5P8AgKCFPVR+VLlC57pa6hZ3abre4jcf71Ws14FGzROGikkRvZjW7p3jDWrAqGn+0RD+AipcRnsFFcfpPj/TbsiO9U2sp6A8iurhninjEkMiup6FTmlYCG+sLS/gaG7gWRG65HP51wOveAZYN9xozb16+Qe3416PRQnYDwF0eKVopkMci8FWGKSvY/EPhiw1yImRBFcAfJKo6V5brWiX2h3PlXkZMZ+5KOjVopXEZ9GaSiqAKWkpaAEpaKKACiik+lAC0UmaKYBRRRQAtJRRQAUUUUAFFFFABRRRQAUUUUAFFFFABSUtJTAKKKKAAdK19J/49j9aya1tK/49j/vUmBjXX/H1L/vVFUt1/wAfcv8AvVFTAKKKWgAooooAKKKKACg0UUAFFFJQAvY11/hH/kGy1yHY11/hD/kGzUmBkXv/AB/TfWoamvf+P6b61BSAWikooAKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFJQAUUUUDCiig0CFpKWkpgFFFFIApKWkpgFFBpM0AFFNJwMk8VveHvC97rbCRgYLQHlyOW9qTdgMm0tri+uBb2cJllPYdPzrvNB8CQxBZ9XbzZOvldlrp9J0ey0m3ENnCq+rHqTWgKylO4xkMEUEYjhjVFAwABipKKr3d7bWcRluZljUep5qALFISAMkgD3ri9T8f2sTFNOhNwf7x4xXJ6h4l1fUMrJdFIz/AAKMVSi2B6neavp9kM3N1Gv0OaxLrx3okORHMZW9AK8vI3NudnY+7ZpMD+6Pyq1TA72b4jRA/udPL++aqP8AEa5P3NOx+NcaaSq5EB2A+Il73sB+dTR/EeUf6zTP/Hq4mijkQHolv8RLBiBcW7RfrWxaeL9DuyBHdgMexGK8jwO4H5UbVPUY+nFL2aA92huIZlDRSo4Po1S14Xa3NxZuHtLiSNh6tmuk03x1qVqQt4gul9emKhwYHqFFYGkeLdL1PaizCKY9Ubj9a3gwYZBBHqKi1gM3VtB0/Vo9t3ApbHDDgivPde8HX2lhprXNzbjknuor1WmnBGOPoapSaA8GyDnrn0PWkzXpniXwZb6juutPxBddSB0f2rze7triyuWt7yIxSqcEHoa2jK4EdFJmjNUIWiikNAxaKKKYgxRRR1oGFFFJQIWkoooGFFJRQAtFJS0AJRRRQAtFFJQAtJRRQAUUUUCCiiigAooooAWikpaANPS/9Qayrr/j7krW0r/UGsm7/wCPuSkgIqKKKoYUUUUAFFFJTELRRSUgFooooAKKKKANPw7/AMhdfpWt4k/4+0+lZPh3/kLr9K1vEn/H2n0qXuMye9JQetFABRRRQIKKSigBaKSigAooooAKKKKACiiimMKKSigAooooAKKKSgBc0maKKACiiigAooooAKWkooAWkoopgFFFFIBKWkopgFFFFABRRRTAKKKKACiiigApKKKQBRRRTAKKTNJmgB2aUkAc0zPIABLHgAdTXc+E/B7Epf6uuO8cJ7e9TKSQEfhDwo1yyahqceIhzHEe/vXoaqFUKoAA4AFCgKAFAAHAApSa5ZSbYBWT4nsUv9DuYnAOFyD6VrVV1P8A5Btx/uGhbgeIKcg57Ej8qWmr/F/vmnV2LYApKWkpgLSUUUALSUUUAFJS0lABRRRQAUUUUDClpKXNAgoopKBC0lFFAwooooAM0UUUCFo60lFAwNbVv/x4/wDATWKa2bf/AI8R/umkwMA/eP1ooP3j9aKoAoxRRTAKKKKQgooooAWiiigAooooAKKKKACiiimAUUUUwKVFFFcgBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFACUUtFMCay/5CFv8A71dl4h/5BsVcbZf8hC3/AN6ux8Qn/iWxVLA5rtRR2FJQAtJRRTAKKWigBKKKKQBRmiigAoopKAFopKKYBRRRQAUUUlAC0lFFABRRRQAUUUlAC0UlLQAUv1pM1u+F/Ds2vXW58pZRn53/AL3sKG7AM8OeHLrXrjjMVop+eQ9/YV6tpemWmlWi29nEEUdT3NT2lrBZ2yW9tGI40GABU1ZN3GFITgZJx9ahu7qCzt2nuZBHGoySTXmniXxnc6kzW2mkw2vQyd3oSuB1PiDxpY6WzQW+Lm5H8APA/GvPNV1zUdXcm8uGEf8ADGpxiswDHqT3J5NLVqNhCjA6CkozSVQAaSlooASilooAKKKM0ALSdaSigBc0lFFACEA9RV3TNY1HSZQ9jcsq90Y5zVOkxQB6h4e8c2eolbe/H2a5PAz0auuUhgCCCD3HSvASobr+Y6103hvxheaO6wXjG4s+nPVKhxGetYqvf2NtqFq9tdxCSNxyDSaff2uo2q3NpKJI2HbqKs5qAPIPFPhW40KQzwBprFjwQOU9q53jqORXvs0Uc8TRSoHjYYZSM5rynxh4WfRZTeWal7Fzyo/5ZmtIyEczQKQe3IpasAooooAKKKKACiiimAUtJRQAUUUZoAKQ0UUAFLSUtABRRRQAUUUUAFBoNJQAUUUUwCiiigArX0n/AI9j/vVkVr6T/wAezf71JgY11/x9S/71RVJdf8fcv+9UdMApaSloAKKKKYBRRRQAUUlFABRRRSAXsa67wh/yDpq5E9K67wj/AMg6akwMm9/4/wCf61BU97/x/wA31qCkAUUUUAFFFFABRRRQAUUUUwCiiigAoopKAFopKKACiiigYUUUUCCiiimAUUmaWgAooooAKSjNFACYpCdo5/TvTiQBk/l613Pg3wlvKanqie8UR/maluwFbwn4Oa72X+rKVi6xwnv716LHGkUapGoVFGABxinbcDjiisW7jDNRyzJDG0krhEUZJY4rP1vXLLRbYy3Ug3n7kY6sa8w1zxDf63KfOcxW+fliU04xuB1Wu+PI4y1vpCea/Qyn+E1w97fXeoTGW+naVz74FVx6AYpK1UUgH59hRTe1LmqELSGiigYUlLSUwDFFFFAgoopaQwopM0ZpgBwfUH1HFbmj+LNT0kqhc3FuP+WbdvxrCopNJgev6H4m0/WUxDKI5gPmjY4xW1Xgylo5BJG5R1OQynGK7fw144ZGS01k5B4Wf+lZShYD0LFZOv6BZ65amO4ULKB8koHKmtSKRJY1kjYMjDIYHNPqL2A8R1fSrvRrw214px/BJ2YVSr2zWNJtdYsmtrtAQfuv3U+teR61o9zol8ba5BKHmOTswraMrgZ9LRS1oAlFFGaBBSUUUDCiiigQUlLSUDCilpKAFpKWigBKKKKACiiigAooooAKKKKBBRRRQAUUUUAFLSCloA09K/1BrKu/+PyStXSv9Qayrv8A4+5KS3AipKWiqGFFFFMAooooEFJS0UgEooooAWiiigDT8O/8hdfpWt4k/wCPpPpWT4d/5C6/StbxJ/x9p9Kl7gZBooNFABSUUUAFFFFABRRRQAUUUZoAKKQ0UxhmiiigAoopKAFpM0UUAFFFFABRRRQAUUUtACUUUUAFFFFABSUtFACUUUdKACiiimAUUUUwCiiigBKKKKQBRRRTAMUtJRSAKSnYpDjOBkn0HJpgIafa21xfXK29nEZZWOOOg/Gug0Pwffaptluc21seQe7fhXomk6LZaRAIrOEKf4m7mspVEtgMPwx4Pg0zbdX2JrvqM9ErqqXpVe8u4LK3ee6kEcaDJJrC7YEzOqKWZgqgZJPQVwOveL2u9Vg0/S3Ig34kl/ve1ZXibxZPrDNbWZMNmDgkdZKwbABdTtAOm8VooaXYHuUX+qT/AHRUGp/8g64/3D/Kpoj+6T/dFQ6n/wAg64/3D/Ks1uB4gvRv99v50tIv8X++386WuxbAJRRRTAKKSloASloooATtRS0lABRRRQAUUUUAAooooAKKKKBBRRRQMKKKKACiiigAooooADWzb/8AHiP901jHpWzb/wDHiP8AdNJgYB+8frRQfvH60VSEFFFFABRRRQAtJRRQAtFFFABRRRQAUUUUwCiiigAooopjKVFFFcggooooAKKKKACiiigAooooAKKKKACiiigBKWiigCay/wCQhb/71dj4i/5BsVcdZf8AIQt/96ux8Q/8g2KkwOZ7ClpB0FLQAlLSUUAFBozRQAUUUUwCiikoAKKKKACiiigApKKKACiiigAooooAKKSloASiijNABSUGnwQy3VxHbW6lpZThQKAL+gaPNrmpLbRZESnMr+gr2SwsoNPs47W2QJHGMACs7w1okOh6YkCAGVhmR+7GtgGs27jHVT1PUrXS7N7q7kCRr+ZNO1G/t9Os5Lq6cJGg/OvIPEGuXGvXpmmJW3Q4ii7CklcB/iHxBda9clpCY7VT8kQNY9Gc0VolYQlFFFMAooooAKKKKACiiigBKKKKACiiigAooooAWkopaACiikpgaGi61eaHdie0cmMn95ETwwr1vQ9atNbsVuLVxnHzp3U14metXdI1S50a/W7tGI/56J2YVMo3A9yqO4t47mB4ZkDxuMMDVPRNXttZsEurZhyPmXuprSrMZ4z4q8PSaDqB2AmzlOY2/u+1YmK9z1jTINW06S0uEDK4+U+h7GvF9S0+fS9QlsrkHfGflb+8PWri7iKmKKWirEJiilpKYwooooAKSiigAooopAFFFFMAooooAWiiigAooooAKSiimAUUUUAFGKKKACtbSv8Aj3P1rJrW0r/j3b60mBjXP/H1L/vVHUlz/wAfUv8AvVHTAKWkpaYBRRRQAUUUUAFJS0UgEpaSlpgB6Guu8I/8g6auR7V13hH/AJB81SwMm+/4/pvrUFT33/H9N9agpAFFBooAKKKKYBRRRQAUUUlABRRRSAKKKKACjtRmimAZoopKYBS0hopAFFFFABRR0pM0wFopKWgAo6Ak9BR1ra8L6I+t6koYEWsRzI3qfSk3YDW8FeGTeyrqV+n7hDmND/EfWvSAMDAGBTYYo4IliiUKiDCqOwp9YN3GFc/4o8SwaJb7VxJduPkQdven+KPEEWiWRIIa5cYjT+teUXNxNd3L3N05eaQ5JPaqjG4Be3dzf3TXN5IZJW9ei+1QUppK2SEFGaKKACiiigAozRmigAooooGFFFFABRRRQAUlFFMAooooEFIQCMEcGloxQB0PhbxTPosy290xlsWOOeqe9eq2s8N1Ak8Dh43GVYV4Vj1ro/CXiSTRblbe4YtZSHnP8B9aynDqhnrFZuuaPbazYNbXCjPVH7qfWr8MqTRLLEwZHGQR3p9ZbAeHajp8+mX0lndLh0+6f7w9aq9K9X8Y+H11jTzLCMXUI3IR/F7V5QwZWZXXa6nDD0NdEJXQCUlLSVYBRRSUAFFFFABRS0ZoAKSiigAooooAKKKKACiiigQUUUUAFFFFABRRRQAUUUUAFLSUtAGppf8AqDWRd/8AH3JWtpf/AB7msm6/4+5KSAjopKWrGFFFFAhKWiigAooopAFFFFABRRRQBp+Hf+Quv0rV8Sf8fSfSsrw7/wAhdfpWr4k/4+0+lS9wMmkpTSUAFFFFABRRRQAUUZpM0AGaKKKYxKWjNFABSUUUAFFFFABRRRQAUUUUAFFFFABRRSUALRRRQAUUUUwCkoooADRRSUALRSUZoAWikoyKACijIozQAUUtGKAEpM0EqP4gfYGpba0u7xwlrbSOx9VwKLpARUoI6csfReTXUab4D1G5Ia/lFsnXC812GkeFNL0za6wiWYf8tGqHUSA4HSPDGqaqVZYvJgPV24P5V3mi+EdO0vbIyie4H/LRq6AAAYAAHtS1jKbYCAYHFLUF3eW9nC0tzKsaqMnJ5rhtd8dNJug0dSB0889R+FSotgdNr3iKx0aEmZw8x+5Gpzk+9eX61rV9rc++7crED8kQPAqpK7zStNM5klY5LE1Ga6IU7ANqay/5CVr/AL4qLFTWX/IStf8AfFW9gPb4v9VH/uCodS/5Btx/uH+VTxf6qP8A3RVfUv8AkG3H+4f5VyLcDxFOjf77fzpaah+9/vt/OnV2LYAooopgJS0UlAC0UUUAFFJRQAUUUUAFFFFABRRRQIKKKKBhRRRQAUUUUAFFFFABRRRQAHpWxb/8eP8AwE1jmtm3/wCPH/gJpMDAPU/Wig/eP1oqkIKKKKACiiigAooooAKKKKAFooooAKBRRTGFLSUUCCiiimMpUUUVyCCiiigAooooAKKKKACiiigAooooASlpKKYC0UUUgJrL/kIW/wDvV2PiH/kHRVx1l/yELf8A3q7DxD/yDoqTA5rsKKM8CigAooopgFFFJQAuaSiigAooooAKSiigAooooAKKKKACiiigBKKWkoAKKKKACkpaSgBCQASa9D+HugeVD/a92n7yT/VA/wANcf4c0ltZ1mK2wfJU5lPpXtEUaxRLGgwqDaAPaokwHnimSSJFG0jsFVRkk0+uB+ImvmNBpFo/zv8A64j+GpWoznPF3iGTXL8xxMRZQnCD+/WEDTQMDFOFaJCFopKWmAUUUUAFFFJQAtFJRQAUUUUAFFFFABRRRQAUUUUAFFFJmgBaKTNFMBaBRRQBreG9bl0LUVmQk27nEqdj717JaXMN5bJcW7h45BlTXg3sehrtPh9rxtbr+ybp/wB1JzESfu+1RJAel1ynjvQP7U0/7XbqPtVuMg+q966ukYBlIYZB6is07DPAAcjOMHuPSiui8b6P/ZOtNLGuLe5+Yein0rnTWydxCUUZpKYBRRRQAUUUUAFFFFMAooooAKKKKAFpKKKAFpKWkoAKKKKYBRRRQAUUtFACVraV/wAe7fWsqtXSv+Pc/WkwMa5/4+pf96o6kuf+PqX/AHqjpgFLSUUwFoopKAFooooAKKKKQBRRRTAOxrrvCP8AyD5q5HtXXeEf+QfNUsDJvf8Aj+m+tQVPe/8AH9N9agpAFFFFMAooooAKKSigAooopAFFFFMAzRRRQMKSlpKBBRRRTAKKKKACkoooAKKKKACjNJR0yT0FAEtvDJdXEdtCCZJW2jHb3r2TQdKi0jTI7WNRvxmRv7xrjvh1o255NWuF5HyxA9x616EKxm7jFqhrOqQaTp8l1OR8o+Vf7x9KuyOsaM7nCqMk+gryXxbrbazqjKh/0WA7VHZj60oq7AzNS1CfVL57y6JLMflX+6PSqtBNJW6VhC9qSiigAopQCSAqszH+FRk1JJbXMa7pLWZV9dhpXGQmkp3bjkUUwEooooAKKKKBBRRSUwCilpKBhRiloFACYop6KzuEjRnY/wAKjJp81tcQDM1tKg9SvFK4iKil69KDTASl4IwRkGkooA7PwN4jNrOul30mYXP7pz2PpXpAORXgnPG04Ycg+lep+CNeGq6cLedv9JgGCPVfWsZxtqM6ivNPiBoYs7pdTtkxFKdsijsfWvS6qanZRajYTWkw+WVcZ9KiLswPDTiipr60ksL6azmBDRsdue69qr5rpQC0UUUwClpKKACiiigAooooAKKKKACiiigQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBqaX/AKg1k3X/AB9yVraX/qDWTdf8fclCAjopKWqAKKKKACiiikAUUUUAFFFFABRRRQBp+Hf+Quv0rU8Sf8fafSsvw7/yFx9K1PEn/H2n0qXuBk0UHrRQMKKKKBBRRSUDCiikpgLSUUUAFFFFAgooooGFFFFABRRRQIKKSloGFFFJQAtJRRTAKWk705EeWQRxI0jnoqjNIBKQlfUH2zXWaT4GvboCTUZPs6HkKvJNdXZeENGtVGbZZXH8TVDqJAeVKkj8JBK30WpPsl2Rxazf98GvaYbO2gXbDAiD2FS7R6L+VR7UDxH7Jd/8+k3/AHwaPsl3/wA+k3/fBr27aPQflRtHov5Ue1YHh5tLv/n0m/74NJ9kvP8An0m/74Ne47F/ur+VGxf7q/lR7VgeHfZLz/n0m/74NKLK9PS0m/74Ne37B/dX8qNo9F/Kj2rA8UXTNRf7tnJ+K1ag8Oa3P/q7MfjxXsOPYflS/lR7VgeXW/gfWpcecI4h7NWta/DteGur9z/sgV3gpal1GBz9l4O0a1AJthK4/iatuG3igQJDGqKOwFS1HLPDCCZZY0A/vNipu2A/FJXP6j4y0ex3L5xkkHQKMiuV1Lx5f3AKWUAgXtJnmqUGwPQ7q9trOMyXEyIoHOTzXH6v4/hTdFpUXnNjHmNxiuFvLq6vZDJeXDzN7nFQ/StI0u4Fu/1C81KXzb64aU9hnAFVs0gorVKwBRRRTAMVLZD/AImVr/vioqnsf+Qna/74pS2A9uiH7pP90VX1Mf8AEtuP9w/yqxF/qk/3RUGp/wDIOuP9w/yrkW4Hhqfxf75/nT6Yn8X++1OrsWwC0UlLTAKKKSgBaSiikAUUUGgAoopKYC0UUUAFFFFABRRRQIKKKKBhRRRQAUUUUCCiiigYdq2Lf/jx/wCAmsc1sW//AB4/8BNJgYJ+8frRQfvH60VQgooooAKKKKACiiigAooooAKWkpaACiiigYUUUUwEOaKWimBSooorkEFFFFABRRRQAUUUUAFFFFABRSUUAFLSUtMAooopATWX/IQt/wDersPEX/IOirj7L/j/ALf/AHq6/wAQ/wDIOipAc2OlFA6Cg0wCkoooAKKKKACig0lAC0UlFABRQaKACiiigAoopKBC0lFFAwooooAKKKKACkPAzS1f0SwOp61bWePlZsufTFAHoXw/0j7BpH2qVcT3PLZ9O1dZ2psUYiiSNeAihR+FONZNjM/W9Sj0nS57yU8IMAeprxO4nlvLmW6uGLSytkk+nauw+JOqme9j0yJvli5lHrXF4q4oQUtFFUAUtJRQAtJRRQAUtJRQAUUUUAFFFFABRRRQAUUUUAFJS0maADNJSVreHNCuNfvxDHlLdD+9k/wouBQtbO7vCRaW7y47gcUXFrc2j7bqB4j6kYFe36bp1rplottZxKiKOeOtRa5pltqumy29ygPykqccg1HOB4jRQ6GKSSJuqMRSVoA6lVmRlkjOHjO5SPUU2jNAHs3hXWF1nR4py2ZkG2UehraryTwFq39m659nkbFvc9f97tXrWaxkrDMHxlpK6toM0eP3kQ8xT9O1eNgsV+YYYcEelfQLKGBDcg9RXjHi7TTpfiKdAMR3BMiewqoMTMbNHakpa0AKKKM0wCiiigAooooAKKKKACiiigAooooABRRRTAKKKKACiiloASlpKWgArV0r/j3P1rJrW0r/AI9z9aTAxrn/AI+pf96o6luf+PqX/eqKmAtFJS0wEpaSigAopaSkAtFFFMAooooACOK63wl/yD5q5Lsa6zwl/wAeE1SwMq9/4/pvrUNTXv8Ax/TfWoaQBRRSUAFFFFMQUUUUAGaKKKBhRRSZoAWkoooAKKKKYBRRSUgCiiimAUlFGKACiiigAqS1tnvLyG1jBLSMAfpUVdh8N9N+0ahNqMi/LD8i5pSdkB6Dp1mlhYQ2kf3YlAqzRTJ5VgheVzhY1LGucZyfj/WzY2AsLd8T3A5x/d715quFAA6Crutag2q6xcXbtlNxEXstUq3irIQtFJRVALRnHJpKQjIwe9AHonw90aD7D/aVxGHndsLkdBXaSwxTRmOVFZGGCCOtch8PdXhn006fIwW4iOdp7iuzFYS3Geb+L/CRs92oaWhMPWSIfw+4rjeCMjpXvLKGUqwBB4INeb+MvCjWTvqOmoTA3MkQ/h9xVRn0YHG0UmQRkHijNagLSUUUwCiiikIWikooGLVixsrjULtLW0QvI35AU2ztbi/u0tbRC8rnt/CPWvWfDPh2DQ7MAAPcuMySH1qZSsIZ4b8M2uiwBiokumHzSEdPati6tILuBoLiNXjYYIIqemSyJDG0kjBUUZYntWN7jPGdf08aXrU9qn+rzmP6VnVp+JNQTU9dnuIuYlO1D6isuuiOwBRRSE1QC1d0bU5NI1WG8jJCAgSj1WqGTSHB4boaT1QHvFrcJdW0c8RBSRQwqQ81xHw21Uz2UmmzNmSE7hn+7Xbg5rmkrMDz74laTjytVhX5vuy47CuD4IBHQ17jq9iuo6ZcWjD/AFq4HtXiEkTQTywOMNE5X8BWtN3QCUlFFaiFooooGFFFFAgooooAKKKKACiiigAooooAKKKKACiiigAooooGFFFFABRRRQBqaX/qDWTdf8fUla2l/wCoNZN1/wAfclCERUtFFMAooopgFFFFIAooooAKKKKACiiigDT8Pf8AIXH0rU8Sf8fafSsvw9/yFx9K1PEn/H2n0qXuBknrRR3ooAKSiigAoopKYBRRRQMKKKKAClpKKACiiigAooooAKSiigBaTNFFABRRSUwClpKs6Zp8+qX6WdsDuY/M390etJuwE2kaVdazeC3tFO0ffk7KK9P0Pw7ZaNCPKQPPj5pWHJqxo2lW+kWKW1uoGB8zd2NaGa55TuAUZrN1fWrHSIPMvJQD2UfeP4VwmqeOdQumK2CC3j7OOppKLYHpbyxp9+RF/wB5sVH9sth1uYf++xXjFxqF9dsTc3ckhPviq/Pd3/76NX7Jge3fbbX/AJ+Yf++xR9ttf+fmH/vsV4jj/af/AL6NGP8Aab/vo0/ZAe3fbbX/AJ+Yf++xS/bbX/n5h/77FeIf8Cf/AL6NH/An/wC+jR7ID243tp/z8w/99ipUdZFDIwZT0IOQa8LcfIfmbp/eNeveEhjw1ZDJPyd6iUOUDXrKu/EWlWcrRT3QDpwwHatXHIrxfxPFH/wk+oHbyZPWiEbgegT+O9Dhztmkc+y1l3PxGTkWdn5noWOK4IBR0FOrVUkB0N5421q5J8oi2H+zzWLc3t5eNuvLp5SffFQYoq1FIBQAvT9eaDRR1qgDFJS0ZoAKM0maKAFopKUUwCp7H/kJ2v8AvioM1NZH/iZWv++KmWwHt8X+qT/dFQ6iM6fcAf3D/Kpov9Un+6KccEYIyDXGB4GMqXDqyfO33hjvTuvI5Fez3ug6XeqRcWcbE98VyWrfD4hWl0qclv8Ank3ArojVQHDYoqa6tbiynMF3E0Ug9RwfpUNap3AWikopgFFFFABRRRQIKSlooAKKKKACiiigAooooAKKKKBhRRRQAUUUUCCiiigYGti3/wCPH/gJrHNbFv8A8eP/AAE0mBgn7x+tFB+8frRVCCiiigAooooAKKKKBhS0UUAFFFFAgooooAKKKKYwooooEUqKKK5QCiiigAooooAKKKQ0ALSUUUwCiiigBaKKKQBRRRQBNZf8f9v/AL1df4h/5B0VchZ/8f8Ab/71dd4h/wCQdFSYHN9hRR2FFMAoopKACiiigBaSg0UAFFFFABRSUUAFFFFABRRRQAUUUUAJS0UlAC0UUUALXc/DOw3S3OoOuQcBD6VwjHapP4V7D4MsvsPhy3jIwW+Y/jUyegI3qgvrlLSzluJDhUUnNT1yXxEv/s2hfZlOHuTgVmhnmV5cve3s91KcvI559qhpSOAPakrYQUtJRQIKKKKBhRRRQAUUUUAFFFFABRRRQAUUUUAFFFFACGkNOqaysrjULxLS0QvLIcf7v1oAfpGl3Os6glnaKeT879lFey6NpVto+npaWqgBR8zd2NQeG9At9C08QxANMwzLJ3Y1r4xWbdwDNZHiTWYNF0ySeVgZGUqidzmrGsapbaRYSXd24CqPlHdj7V45rOrXWtX7Xd0SBn93H2UUJXGUWZpHeR/vOxY0UUVqIKKKKADe0bJKpw0TBx+Fe2+HdRGp6LbXWcuyfP7GvEvb1r0D4X3+YrqwduQ25B7VEloB6B14rhfifp/madDqCDLxNtP0ru6zfENot7ol3CwyfLJH1qIgeH57+tFIqlB5bfeQ4NLW4BS0UUAFFFFABRRRQAtJRmigAooooAKKKKACiiimAUUUUAFFFFAC0UlLQAla2l/8e5+tZNaul/8AHufrSYGPc/8AH1L/AL1R1Jc/8fUv+9UdMApaSigApaSloASlpKUUAFFFHSgAooopgHY11nhLiwmrk+xrrPCX/HhNUsDKvP8Aj+m+tQ1Lef8AH9N9ahpAFFFFMAooooAKSiigBaSiigAooooAKKKDQAUUlFABRRRTAKSlooAKKKKACm0tGKAGP9w46ngfWvX/AAbYix8O24K4eRdz/WvKtPtjd6na26jkyBj9K9wjjWONY1GFUAAVnNgOrlvH+pmw0IxRtiWY7ce1dT7V5V8Qb77Xr6QKfkt12ke9RFXYzmVG1Qo7UtAorcQUUtFAwooooAltria0uUuLZyk0ZyCK9W8LeJIdbtQjkJdoMOh7+4ryWpbW5ns7lLm1kKTIcgjvUSjcD3WmuqupVgCpGCD3rB8L+JYNctQrEJdoMPGe/uK3s5rHYDzPxl4UawkbUdNQtbMcyRj+H3FceCCMg8V726K6FHUMrDBB715j4x8KNp0rahpyFrVjl4x/B9K1hIDk80tIMEZHIpRWggopaKBjc1Ja2897dJa2iF5XOMDt70ttbT3lyltaoXmc4AHb616v4W8NQaHahmAe7cZdz29hUylYQ7wv4cg0O05Ae6cZkkPb2rfpOBTJJFjRndgqKMkntWF7jFllSGNpJGCooySe1eYeLvFUmqytZ2LlLRThmH8dJ4w8VPqczWNi5W0Q4dh/Ga5UcDA4ArWEOrAXgDA4FFGaK1EBpKKSmMDRRRQI0vDd+2m6/bTA4SRtsn0r2lGV1DqchhkV4E+dh28MORXs3hW/GoaBbTZyVUIfwrGohmwa8m8faeLLxEZ0XEU6jH1r1oVxvxJsfP0aK6Uc27bifapg7MDzPFLS9QD6jNFdACUUUUCCiiigAooooAKKKKACiiigAooooAKKKKACig0UAFFFFABRRRQAUtJS0Aael/6g1k3X/H09aumf6g1lXP8Ax9PSQEdFFFWAUUUUgCiiigAooooAKKKKACiiigDT8Pf8hZfpWp4j/wCPpPpWX4e/5Cw+laniP/j6T6VL3Aye9JQetFABRRSUAGaKKKYwooooEFFFJQMWikoFAC0lLSUALSUUUAFFFFABRRRQAlFLSGmAhJ/hGWPQetepeDNEXStNWWRQbmcbmbuB6Vw/hHTf7S1+MEZjt/3h969aAA4HAHQVjUl0Ad0rA8VeIotEtdqYe6k4RPT3rU1K+j0+wmu5iNsS5x6143qF/Nqd/Le3DFmc/LnsvaohG7AZd3VxfXLXF5K0krHqT09qjpKWuhKwC0UUtMApKKKAFopKSgBW+6a9f8J/8i3Zf7lePt9w16/4S/5Fqy/3Kyq7AbNeNeKB/wAVPf8A+/XsteN+J/8AkZr/AP36mluBk0tFLiugBM0UYooAWjpSUUALmkopaACkoooAWikooAKlsf8AkJ2v++KiqWx/5Clr/vilLYD3CL/VJ/uinZA5JwKbF/qk/wB0VX1TI0y4wcHYcGuMC2rBhlSCPUUteM6T4m1fTXLJctPGHOY36Yr0vw74ks9cg/dtsuFHzxn+lU4NAT65olprNqYrlAHA+SQDlTXkmr6Zc6PfNaXQPB+R+zCvbic1h+KtEj1rS3TAE8Y3Rv3HtVQnZgeQ0UMjxu0ci7ZEOGHpRXSAtFFFMQUUUUAFFFFAwooooEFFFFAwooooEFFFFABRRRQAUUUUAFFFFMYHpWvB/wAeX/ATWQela9v/AMeX/ATUsDCP3j9aKD94/WiqEFFFFABRRRQAUUUUDCloooAKKKKBBRRRQAUUUUxhRRiimIpUUUVyAFFFFABRRRQAUlLRQAlFFFMAoopaACiiikAUUUUAS2f/AB/wf71dd4hP/EvirkrP/j/g/wB6ut8Q/wDIPipMDnOwpKPSimAUGiigAooooAKKKSgBaSiigAooooAKKKSgBaSiigAooooAKKKKAFzRmm5ooAmt4jcXUMAHMjj+de6WqCO1hQDAVAP0rxvwtCLjxTYIegbJr2ofyrOQIK8v+I955+tQ2ueLcZx9a9PJwK8W8T3P2vxPezZyOFH4UR3BmXSGlpK0ASilooASijFFABRRRQAlLRRQAUUUlAC0lGaTNAC0UlLTAWiinIjySLHEheRzhVHU0AOt4Jrq4S3tkLzSHCqK9b8J+GodDtA7gPdyDLue3tVbwb4WTR7cXV0A97IMk/3B7V1VZSlcYVXvryCwtHublwkaDJJqWWVIYmklcKiDLMTwK8l8YeJH1y7MFuStjEcD/bPvSSuBQ8Sa7Pr+oGVyVtoziKPt9aycU44orVKwhtFLijimAlGaKKAENbvgi8+xeKYWJwki7D9TWFU1pKYdQtJQcFZlpPYD3z2prKHUq3Rhg0yCQTQpKOjqDUnesRnhmuW/2XX76HGB5px9Ko10nxAtvs/inIHEqbq5ut1sIKWkpaYBRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigBa1NM/wCPdvrWVWppn+oP1pAZFz/x8yf71R1Jcf8AHzJ/vVHTAKWkpaACkpaKACiiigAooooASloooAD3rq/Cf/HhNXKdjXV+E/8AjwmpMDKvP+P2b61DU15/x+zfWoKQC0UUUwCkoooAKKKKACiikoAWiikoAKKKM0AFFJRQAtJRS0wEooooAKWikoAWiigdaAOi8B23n+KUkIysUZ/OvV68/wDhlbgte3JHIbaK9BrCe4yKZxHE8h/gUtXh+qTm71e8uCc+ZJkV7Lr04t9Fu5On7sgflXiKZMYJ6nNVTAUUuaKK1AKWkooAWiikoAWkozSUATWt1PZXSXVo5SZDkEd69Y8L+I4NctByEukGJI/6ivIams7u4sbtLq0cpKhzx3qJRuI91zmkeNJEZHUMrDBB71i+GfENvrloCCEuUGJI/wCordrHYZ5j4v8ACTac73+nKWtmOZIx/B9K5IEEZHQ17zIiyIyOoZWGCD3ry3xh4Xk0mZr6yQvZucsoHKVrCXRgcxmpbS2nvbpLa1QvK5wAO31plrDNe3EdvaL5kshwAOcfWvWPCvhqHQ7UM4D3bjLue3sKqUrCF8L+GoNDtgzAPduPnkPb2FdBRSMwVSzEADkk9qweoxsjKqlmIVVGST2rzLxl4ra/kfTtOcrbqcSSD+KpvGnitrt303TZCIVOJZR39hXFgADAGBWsIdwEGAMAcUtGKK2EKKKSigYtJRmigAooooAB1r0H4YXebS4sSf8AVnf+defV1Hw8uPI8RumeJkC1E1oB6sOlZniO2+1aBeQ4zuTitOmSoJInQ9GUiudAeCjgFf7p2/lRU15H5N9cxntK1Q11LYBKKKKYgooooAKKKKACiiigAooooAKKKKACiiigYUUUUCCiiigAooooAKWkpaANLTP9Qayrn/j6krV03/Umsq5/4+pKSAjoooqgCiijimAUUUUgCiiigAooooAKKKKANLw9/wAhZfpWr4k/4+0+lZXh7/kLr9K1fEn/AB9p9Kl7gZBpKD1ooAKKKKYwooooAKSlpKBBmiiigYUUcUUAFFJRQAtFFJQAtFFFABRSCloAKO+aKRziNj7UAehfDazCadLeMPndyo+ldnjisfwlbi38P24A++N1bVc0nqBwHxJ1AhYNOjbBzuf3FcLjHFbvjG4+1eJpnzwi7RWJit4KyAbS0UVYBS0lFAC0maKKACiikzQAN9017B4S/wCRasv9yvHm+4a9h8Jf8i1Zf7lZVdgNmvG/E/8AyM9//v17JXjXif8A5Ge//wB+ppbgZlFGaK6ACigUUAFFFJQAtFJRmgAozSUtABRSUUwCprH/AJClp/vioamsP+Qpaf74qZbAe4xf6pP90VX1T/kGXH+4f5VYi/1Sf7oqDUx/xLbj/cP8q5FuB4Wn8X++f51PaXU9jdx3dq5SWM547ioVH3v99v507FdVroD2jQtUj1jS4ryPALD5l9DWlivPfhpeFJ7qzY8Mdyj0r0OuaSswPKvHmmix14ToMJdDcfY1zXSvRfiXbg6VFdY5jbH51523Wuim7oBKKKK0EFFJRQMWigUUAFFFFABRRRQAUUUUCCiiigAooooAKKKKBhRRRTEBrXg/48f+A1kHpWtb/wDHl/wGpYzDP3j9aKD95vrRVCCiiimAUUUUAFFFFAxaKKKQBRRRQIKKKKYwFLSUtACUUUUwKVFFFcggooooAKKKKACkpaSgAoopRQAYooooAKKKKACiiigCaz/4/wCD/errPEH/ACDoq5Kz/wCP+D/errfEH/IPipMDnOwoo7CimAUUlLQAUlFFABRRRQAUUlLQAUlFLQAlFLSUAFFFFABRRSUAFFFFACUZopDQB0ngCPzPE6t/zzGa9d9a8q+Gwz4guD6KK9VrKQxkzbYZG9FJrweeQyXlxIf4pG/nXuWoHbp85H9w/wAq8HU/NIf+mjfzqoCH0UlLVgFFFFABRjPagnAzXVaF4Gu9UtkubuXyIXGUA+9ik3YDlSDRXWa14EvdPhaexk+0RqMsp+9+FcnzkggqRwVPUUJ3AKSlpKYBSGig0wEooooAKKKQkAZPSgAzjGASTwAOpr03wN4W+xRrqWoIDdOMop/gHr9azfAnhUyMuranHgdYYyP1r0UVnJjH0jMqKWYgAckntRmvO/HPisyO2labJwOJpFP6CpSuBS8aeKW1KZtOsHItUOJHH8Z9K5LoMDpSDAGBRWqVhBSZoJopgFFFFACUUtFADaQnbhv7pzTsU1x+7f6UAe5eH5PN0KykPO6IVoN3rH8Ind4YsPaOtgjisXuM8y+KMeNWtJcf8s8VxZ613fxUX97aN+FcKeprSOwhKWkoqwFooooAKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFFABRRRQAVqaZ/qG+tZdammf8AHu31pAZFx/x8yf71R1Jcf8fMv+9UdMApaSigBaSlopgFFFFIAooooAKKKKYB2NdX4T/48Jq5Tsa6rwn/AMeE1SwMq8/4/ZvrUNTXn/H7N9ahpAFFFFMAopKKAFopKKACiig0AFFJRTAKKKKACijNFABRRQKACijNFABRRRQAUvY/SkoP3G+lAHpfw3j26NM/99812Fcr8PRjw8p9TXVVzy3Gc746lMXhmYg9WAryUDCgV6h8SGx4aI9ZRXmBrSnsADpSUtJWgBRRSUALRSUtACUtJS0CCiiigCxZXlxYXaXVo5SVD2/ir1nw14ht9ctAwIW4UYeOvH6sWV5cafdpdWjlJUP4MPeplG4z3OmTRJNG0cihkYYIPesjw34gt9csw6kJOoxJGeorarDYDG0fw3p2j3E09pH88pzkj7v0rYAxS0hIUEk4A5JoAGYKpZiABySe1ec+MvFrXLvpumSYiHEso7+wp3jLxabgvpulyYj6SzL39hXEYAXArWEOrATAAwOlFFFbAFIaWkoAKSiigApaSloAKKKKAFFa3haXyvE9if7z4NZIq9oh269YH/ppUy2Ee3+tFIOVBpa5hniviOPyfEN0nvmsytzxou3xXcgf3RWFXTHYBaSlpKoQUUUUAFFFFAwooooEFFFFAwooooAKKKKACiiigQUtJS0AJRRRQAUUUUAaWmf6k1l3P/H1JWrpn+oNZVz/AMfUlCAjooopgFFFFABRRRQAUUUUAFFJS0AFFFGaANLw/wD8hZfpWp4j/wCPtPpWV4f/AOQsv0rU8R/8fS/Spe4zJNFHekpgLRmkooAKKKKBBRRRQAUUUlAwooooAKKKKACiiimAUUUUALRSUZoAWkf/AFTUUj8xt9KTA9p0LA0Szx/zyFX+9ZHhaYTaBakHO1Ata9cr3A8X1wk63dk9fMNUDWz4ttza+JJ4yMbhvrGrpjsAUlFFUAUUlFABRRRQAUUUUwEb7hr2Hwl/yLVl/uV4833TXsPhL/kWrL/crGrsBs1414n/AORmv/8Afr2WvGvE/wDyM1//AL9TS3Ay6KKK6AFpKKKYBmjNJRSAKKKKACiiigAooooASp7D/kKWn++Kgqaw/wCQpaf74pS2A9yi/wBUn+6Kg1P/AJBtx/uH+VTxf6pP90VBqf8AyDrj/cP8q5FuB4co+9/vt/OlpF6N/vt/Ol712LYDpfh+SPEqgdChzXqorzf4bWjSahcXZHEXyg/WvSK5anxAcr8Rsf8ACLyZ/wCei15ea9F+JlwF0iK2zzI2fyrzo1tS2AKKSitQFpKKUUAFFFFABRRQaAEpaSloAKKKKBBRRRQAUUUUAFFFFMAooooAQ1rW/wDx5f8AAayTWtB/x5f8BpMZidz9aKP4j9aKYgooopgFFFLQAlLSUtAxKWiikAUUUUCCiiimAUtJRQAUUUUxlKiiiuQQUUUUAFFFFACUUUUwClpKWkAtJS0lABRRRQAUUUlMCaz/AOP+D/errPEH/IPirlLL/j/t/wDerq/EP/IPiqWBznYUUdqKYBRRRQAUUlFAC0lFFABS0lFABRRRQAUUUlAC0lFFABRRRQAUUUUAJSGlpDQB1fw2bHiGceqivVe5ryT4fPs8T7T/ABjFet9z9azluBX1EZ0+4H+wf5V4Mv3pP+ujfzr324Xfbyr6of5V4Ky7J51PUSt/OnEAoooqwClpKWgCW1CtfWyv90yrn8693iVViVUACgDGK8C9MHBByD6GvTfBXiqO+hTTr5wl1GMIxP3x/jUSQHZ4rhfGnhETq+paWgWdeZIx0f3+td1SGpTsM8AJ5IIKkHBB6ikzXoHjfwiZN+qaWn7wcyxAfe9688Bznggjgg9RWqdxD6KSlFMAxSUtFADTwCScAV13gnwq2pzrqOoIVtEOY0I++ff2qt4P8MSa3dC5uVK2URzz/GfSvWYoo4IkiiUKiDCgdqzlIBVVVUKoAUDAA7Clzig1yvjPxOmjWxtbVg17KMDH8HvUpXGU/HPiv7FG2madIDcuMSOD9we3vXmo4zySSck+poZ2kkaSVi0jnLMe5orVKwhc0UlLTAWikooAWikooAWiiigAof8A1bfSlFIwyuPXigD2jwiu3wxYj/pnWwRxWb4cTy/D9inpEK0u1YsZ5z8VD89otcIeprtvik2dRtI/9jNcSetaQ2EJRRS1YBRRRQAUUUUAFFFFABRRRQAUUUUwCiiigAooooAKKKKACiiigArU03/j3P1rLrU0z/j3P1pMDIuP+PmT/eqOpLj/AI+ZfrUdMBaKKKACikpaYBRRRQAlLRRQAUUUUAHY11XhT/jwmrlT0NdV4U/48JqlgZN7/wAfs31qKpbz/j9m+tQ0IAooooAKKKKACiikoAWkpaSmAUUUUAGaKKKACkpaSgBaQUUUALRSUUALRSUUABpD9xvpRQeQfpQB6n8PDnw6vsa6quO+G0m7RJV/uPiuxrnluM5L4kDPhvPpKK8xPWvVfiBHv8MSY7ODXlPUZrWnsAUUUVYAaSijtQAUUUUAGaWkooELRSUooAWnRRyTTLDAhklc4VQOaI45JpkhgQySucKoFeo+EfC0ejwi5ugJL1xkn+59KmUrAO8H+GV0WFri4O68mHz+i+wrpqKQkCsHqMWmuodGVhlWGCPaq1pqdlezSxWtwkjxHDhT0q3QB5V4u8LyaRK13ZqXsnOSo6oa5jtkHIr3ieGOeJopUDo4wynvXlfi3wxJo07XNqpeyc54/grWE+jA5qij36ikNagLSUdqKYBRRRQAUUUUAApaKKBC1d0UZ16wH/TSqNafhpPM8S6evpJUy2Ge0r90fSloormA8e8and4ruf8AcFYVa3imTzPEt034Vk10x2AKKKKoAooooAKKKKACiiigAooooAKKKKACiiigANFFLQISlpKWgBKKKKACloooA0dN/wBSay7n/j6etTTv9Say7n/j6ehAR0UUUwCiiigAopKKAFoopKAFpKWkoAWiiimBpeH/APkLL9K0/Ef/AB9J9Ky9A/5Cy/StPxF/x9J9Kh7gZR60lKaKYxKKKKACiiigQmaKWkoGFFGKKACiiigAooooAKKKDTAOKM0lFABS0lFAC0o549abS5oA9I+HN4JtGe3Y/PHIePauv7V5N4K1Madr6o5xHcjZ7A16uG9Olc01Zgef/EmwKS2+ooM7v3be1cQTzXtOtacmq6ZNaOBl1wpPY14zc20tldSWlwpWSI4we49a0pvSwDM0Z4puaWtQFooopgFFFFABRRRigBG+6fpXsHhH/kWrL/crx9h8p+leweEf+Rasv9ysauwG1XjXif8A5Ga//wB+vZa8b8T/APIzX/8Av1NLcDKoooroAKKKSgBaSiigAooooAKKKKYBRRRQAlTWH/IVtP8AfFQ45qfT/wDkK2n+/Uy2A9xi/wBUn+6Kg1P/AJB1x/uH+VTxf6tP90VBqf8AyDrj/cP8q5FuB4enRv8Afb+dPSOSaVIYVLyyHCqKtabo+pajIY7W1kGXPzuPl616P4Y8J2+jgXE+Jrsjlj0X6V0SmkgL3hjSBo+kR27AGZhmQ+prYNFZPiLV4tH0qW4dhvI2ovcmufVsDgvH+oC811LdDlbUbW+tcxTpJJJpnnmbMkhyxptdcVZAFFFFUAlLRRQAUUUUAFFFFACUtFFAgooooAKKKKACiiigAooopjCiiigBD0rWg/48v+A1knpWrB/x5/8AAaTAxT94/Wij+I/WiqEFFFFAwooooAWkoooELSUUUhhRRRQAtFFFMAooooAKKKKYFKiiiuQQUUUUAFFFIaACiiimAUtJS0gFpKKKACiiigAooopgTWX/AB/2/wDvV1fiH/kHxVyll/x/2/8AvV1fiH/jwiqWBznaikopgLSUUUAFFFFABRRRQAUUUlAC5pKKKAFNJRRQAUUUUAFFFFABzSGlooASiloxQBseD5fJ8WWRPAdsGvZscmvCtOm+y6paXGf9W9e6Qtvhjf8AvKDWcgQ5hlSPUV4drdubXX72AjGGz+de5V5N8QLU2/iVp8YW4UYojuM5iilNJWggoopKAClV2R1kjYpIpyrDqKbRQB6b4N8XrqKLYai4S7UYVyeH/wDr12dfPwLK6vGxV1OVYdQa9M8F+L11BF0/UnCXaDCOTw4/xqJRGdoRXn/jXwfu36ppSYccyxAdfevQKOvapTsB8/jnPGCDgg9RTq77xr4PJL6ppKfMOZYgOvuK4BTuzwQRwQeoNaJ3EOra8L+HptfvhwVs4zmR/X2FVtA0W413UFtoARCOZZOwHp9a9k0zT7fTLKO1tUCxoPTqfU0pSAls7WGztkt7dAkaDAAqU0tZPiHXLfQ7BriYgyEYjTuxrMZV8V+I4dBsSQQ11IMRR+/qa8fuLia7uZLm5cvNIcsT29qm1K/udUvnvLxy0jngH+EelVq1irCEpRRRVAFLSUtABRRRQAUUUUALS0lFACinwqXubeMfxygUytTw3am98Q2sIGdrB/yoewHstlF5FlDF/cQCp6KDWAzyj4jzCXxJEgPEcWK5Q1q+J7n7V4kvXzkJIVFZZraOwhtLRRVAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKACtPTP9QfrWZWppn+oP1pMDIuP+PmT/eqOpLj/j5k/wB6mUwEpaKKACiiigAoopKAFooopgFFFFAB2NdV4V/48Jq5XtXU+Ff+PGapYGVef8fs31qGpbz/AI/ZvrUVABRSUtABSUUUAFFFFMAzRRRQAUUUZoASlzSZooAKKSigBaKSigBaKSigBaSiigYtA60lKOtAjv8A4YTDyb2AnneCK72vLPh3ceT4ha3J4lQtXqYrCe4zJ8UW/wBp0C6TGcKW/KvF4z+6X8a94u4/OtJo+u9Cv6V4XcxG3vLi3PBicirpgNzRTaWtACiiigBKKWigBKBRSZoEOp8ccksqQwIXlc4VQOTTEDySLFEheVzhVHU16j4O8KppMIvLwB72QZ5/gHt71MpWAf4Q8Kx6RELq6Ae9cZJ7J9K6mijpWLdxjScV5/4y8YHL6ZpMnPSWZT09hTvGvi3BfS9Lk+bpLKp6ewrz/GP6+9XCPVgWtM1G50m9W8s3IcHLjPEnrmvX/D+vW2uWKzwMBIB+8jPVTXi1W9K1K50i+W7s2IYH507OKqUbge6A1HcQRXMLwzIHjcYKms3QNctdcsVuLdsOB88Z6qa1c1lawHkvi3wzLodwbi3UvZSHg/3K5zP4iveLq3iuoHguIw8bjDKRXkvivwzNoVyZoQ0ljIcq39z2Nawn0YGDmlpo/SlrUBaKBS0AJS0UUAFFFFAgrofAluZ/EytjiEbq5/Ga7j4Y2uZ7q8I4K7Qaib0GeiU1jhST0Ap1UNauPsmj3U+cbENc4HjOpS+fqt1L6yEVWo3Fmd/77lvzorqWwBRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAFpKKKBC0lFLQAUUUUDNDTv8AVGsy5/4+XrT07/Umsy4/4+XoQiOiiimMKKSloEFFJS0AFFFFABRRRQAUUUUwNHQP+Qqv0rT8Rf8AH0n0rM0H/kKr9K0/EP8Ax8p9Kh7gZXekJo70UxhSUUUALSUUUAFFFFABRRRQAUUUUwCikooAKWkooAKKKKACiikoAWkopKADLAhkOHU5U+hr1nwjrSaxpKEkefCNjr34715NV7RdWm0XUVu4SSh4lQdxUTjdAe09RXM+L/DC6zD9otsJeRjg/wB72Nbmm6hb6lZpdWrhkcc4PQ+lWq51dMDwmaKW3naC5jMcqnBVh1+lJXsGueHbDWo8XCbJe0qj5q4HVPBmq6eS0IFxD228tW8aie4HPCinyW9xAcTWssZH94UzPsa0ugFpKMj3oyvc0XAKKNy+tG5fWgAb7p+leweEv+Rasv8Acrx5mXaee1ew+ETnwzZH/YrKrsBs1414oP8AxU9//v17J3rxjxTIg8UX4LgHfU0twM2jNNBDdDml59K6AFzRSUUALSUtIaYC5opKWgAooooAKKBS0AFT6eP+Jtaf79Q1Pp//ACFbT/fqZbAe3Rf6tP8AdFOYBgQwyD2psX+rT/dFPPvXGAyOJIl2xoqL6AU7IFVJ9SsYATNdwpj1auZ1fx7Y2oaOwUzzdj/DTUWwOl1PUrbTLRri7kCqBwM8t9K8k1/W59cvjPLlYV4iT0FQanql7q85mvpS392MH5VqnXRCnbVgLRSUtagFFJRQAtFFFABRRRQAlLRRQAUUUUCCiiigAoopKAFooopjCiiigAooooAQ1qwf8ef/AAGso9K1IP8Ajz/4DSYGMfvH60UH7x+tFNCCiiimMKKKKBBRRRQMKKKKQBRRRTAKWk70tABRRRQIKKKKBlKiiiuUQUUUGgApKKKYBRRRQAtFFFIAopaSgAooopgFFFFAE1l/x/2/+9XVeIf+PCKuVsv+P+3/AN6up8Qf8eEVSwOd7UUdhRTAKKKKACig0UAFJmlpKACiiigAooooAKKKKACiiigAooooAKKKKAClpKKAB87OOoINe2eG70X+hWs4OflC/lXilejfDK/8zT5tPY/NAdw/GplsCO5rhvibZGSwt71R/qG+Y/Wu5rO1+wGpaNc2h/jXI/CoW4zxE88+tJQQUZkYYZGK4+lJWogopDRQAUlLRQAlALK6ujFXU5Vh1BpaKAPTPBfi5dQRdP1Jwl2gwjno4/xrtBXgClldXRijqcqw6g16b4M8XLqCLYai4S7QYVyeHH+NRKIHZ4z1rhvE/gT7deC60lkheRsSqeBj1Fd1RUXGZmg6NbaLp6W1uoz1du7GtOiq97eQWNrJc3LhI0GSTQBFq2p22k2L3d04CKOB3Y+grxvWtXudbv2u7kkLn93H2Uf41Y8S6/Pr9+ZGJW1jOIo+31rHPrWkY2EBpMUtJVgJRS0lAAKWkooAWikpaACiiigApc0lFADh1Fdp8MrEzajPfkcRAxg1xDHCnHU8D617D4H03+z/AA9CWGJJxvf61M3oCOiqpqk/2bTLmYnGyMkflVuuU+Il/wDY/D5jRv3kzhce1ZrcZ5U8pnlknbrKxY0lGNoCjtRW4hKKKKACiiigAooooAKKKKACiiigAooopgFFFFABRRRQAUUUUAFFFFAC1pab/qD9aza0tN/49z9aQGTcf8fMn+9TKfcf8fMn+9TKYCUtFFMBKWkpaAEpaKKACkpaKACiiigAPQ11Phb/AI8Za5btXUeFv+PGWpYGVef8fs31qKpbz/j9m+tRUAFFFJTAKKKKACiiikAUUlFABSUUUwCiiimAUUUUAFFGKKQBRRRQAUUUUALRSUZoAv6JeGx120uQcfMEP417aGBAI5BFeAsxADDqh3D8K9o8M339oaDaT7ssUw31rKaGavevH/GtkbLxLKcYW5+cV7B2riPiVppm0+PUI1zJCdpx6UoPUDzmlpMg8joaWthBS0lLQMKKKKBCGlRHkkWKJC8rnCqOpp0ccksqRQoZJXOFUdTXp/g/wnHpUa3l6BJeuM89E+nvUylYYzwd4STSoxe3yh71xkA9E+nvXXUU01i3cBc1wnjTxf5e/TNLkzIeJZVP3fYUvjTxb9nV9M0uQGZhiWVT90e3vXnYzkkkkk5JPc1cIdQHevOSeSfU0lApa2AT60nenYpMUCLek6nc6PfLd2jEEH507MK9f0LWrbWrFbm2YbsfOh6qa8Uq9o+rXOi3y3Vqxxn94nZhUSjcZ7fUV1bQ3du9vcRh4nGCpFVNF1e11mxW5tW7fOndT71pCsXoB5D4p8MTaFcGWEGSyc/K39z2NYOPyr3e7tYbu2eC4QPG4wQRXk3ijw3NoVyZIwZLJz8rf3fY1rCfRgYNFLSVqAZopKKAFpM0UlAhWbapb0r1rwJYmy8ORblw0rb/AMDXlmnWrX2p21moz5rgN7CvcbWIW9tFAvSNQv5VjUfQZNXK/EK8Ft4deHOGuDsFdVXmPxKvvtGqw2KnIgAc496iKuwOOAwij0FLQetFdIBRRRQAUUUUAFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigApaKKYwooopAaGn/6o1mXH/Hy9aWn/AOqNZtx/x8vQgI6KSlpiEpaKSgBaKKKACiiigAooooAKKKWmBoaD/wAhVfpWl4h/4+V+lZug/wDIVX6VpeIf+PlKh7gZXeijvRTAKSiigYUUUUAFFFFABRmkopgLRSUUALSUlLQAlLRSUALRSUUAFFFFABRRRQAlJS0UAaWg65daDdeZAS1ux/eRHp9RXquj6zZ6xbLNaSAkj5oyfmWvGKltLm4sZxNZytE45+U/e+tRKFwPchzTgPSuD0fx+hCw6tEVfoHTp+NdlYahaahF5lpOko77T0rBxaAkmtLefmaCOQ/7QqD+x9NP/LjB/wB81eoqbgUf7H03/nxg/wC+aQ6Nph62EH/fNX6T8Kd2BQ/sXS/+fCD/AL5o/sTS/wDnwg/75q/Rmi7Aof2Jpf8Az4Qf981chhjgiWKFAiL0UdBTiaM0agKayLzw3pV7M809qhkc5Zscmtfml/ChOwHJ3HgLR5s7TLH/ALtZN18OSuTZXZPp5hr0LikIFNTaA8jvfB2t2eTsSZf9jk1jT2t1bEi5tZYsd2HFe6DI6VDcWtvcjFxAko/2hVqq+oHhasG+6QfpS16lqfgjSb4lo1a3c9PL4FclqfgbVbIF7VlniHQD71aKomBzVFOnjltn2XULwMOzikx+RrRO4BS0lLTAO9FFFIBan0//AJCtr/vioKm0/wD5Clr/AL4pS2A9vi/1af7oqLUCV0+4KnBCHB/CpYv9Un+6Kg1PjTbj/cP8q5FuB4cS828zSvL87fePvQAFGFGB7U1Ojf77fzp1daSsAUUUtUAUUUUAFJS0lAC0UUUCCiiigYUUUUAFFFFABRRQKACiikoAWiiimAUUUUAFFFFACGtSD/jz/wCA1lnpWpB/x6f8BpMRjH7x+tFB+8frRTAKKKKYBRRRSAKKKKBhRRRQAUUUUwClpKWgAooooEFFFFMZSooorkEFFFJQAUUUUwCiiloAKKKKQBRRRQAUUUUwCiiigCay/wCP+D/erqPEH/HhFXL2f/H/AAf71dRr/wDx4xVLA57tRRRmmAUZpKKAFooooAKKKSgAooozQAUUUlAC0UUUCCikooGLRSUUAFFFFAC0UmaKAFzW14R1L+zfEUDs22KY7ZDWLSHOODgg5B+lJq4H0ApDAEdDyKDyKw/CGrDVtChlJHmoNrjuMVuVkM8f8c6YdN8QPIi4gueU+veudzXr3jbRv7W0R/KXNxD80Z/nXkAJPUYIOD9RWkXoIdRSUtUAUUUUAFFFLQAClVmV1dGKupyrDqDSUUwPT/Bni1dRjWw1Bwl2gwrE8SD/ABrss18/qzK6ujFHU5Vh1FdZp3xA1G0gWG6iSYKMB+5+tZuPYD1GWVIomkkcKijLMegFeTeMfEr63dm2tmK2URx/vn/CoNd8W6lrUfkMRBb91Tgt9awRgAAdBTjEAoopKsBaTNFFABSUtJQAtFJS0AJS0UUAFFFFABSUUhO0E/l70AanhvTG1fXre2A+RD5jHtxXtqKqKEQAKBgCuQ+Heimx0o3s64mufmGeqiuwrKTuAteT/EHUxfa8LeNsx2w2sP8Aar0bX9RTS9InunbBCkJ/vdq8ReV55XuJTmSVtzU4IAzmkoorUAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUtABWlpv+oP1rNrR03/UH60mBl3H/HxJ/vVHUlx/x8yf71R0wCiiimAUUlLSAKKKKYCUUtJQAtFFFAAeldP4X/48Zq5jtXT+F/8AjylpMDLu/wDj9l+tQ1Ld/wDH5L9aipAFFFFABRSUtABSUUUwCiikoAKKKKACiiigBaSiigAooooAKKKKACjNJRQAGiikoAOvXvXefDPUcfaNNkbkndGPauDq9ol+2l6zb3anA3BG+hqZK6A9vHIqvf2iXtlNbOARIhXntUsMiTRLLGco4yCKkxWIHg95ZvYX09nIMGFyoz3FRCu++I2ik7NWt0yV+WRQOvvXBcdR0raLugEpaKKoAoJwCTRmmt8yketAHo/w80OKKz/tS4QNcSHCZ/hHqK7iuR+H+qw3ejLa7gJ4Tgp3xXW5PpWEtximuF8Z+LhbB9N0uQNOwxJIDwv096d4y8XrbK2naXIGnYYkkU8IPb3rzo5JJYlmJySe5qoR7gM5JJYksTkk9zRTqSthBS0UUAFFFFACUoFFLQBe0XVrrRL0XNqx25/eR9mFev6Lq9rrFitzasOR8yd1PvXiVaGi6vdaJei5tWJX/lpH2Yf41E43Ge21BeWkN7avb3CB45Bggj9aq6NrNprNmtxauCSPmTPKn3q9JIsSGSQhUUZLE8AVgB4nrVgdK1aezBJRG+QnuKoGtXxNfJqOv3E8JzGrbVPqKy66Y7AJRS0VQCUgpcVJb20l5cxWkIJkmbaMdqTdgOv+G2l+fdzapKv7tfkjz616QKo6Np0el6ZDaRgDao3Y7t3q/XPJ3YEdxOlvbyTyHCRrkmvENSvGv9Tubtzks5APtXonxD1b7FpAs42/eXJ2nHUCvMAMAL6VpTXUB1FFFagFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigAooooGFFFFMQUtJRQMWiiikBf0//AFJrNuP+Ph60tP8A9Uazbj/j4ehCI6KKKYBRRRQAUUUUAFFFFABRRRQAtFFFMDQ0L/kKj6Vo+If+PlKzdC/5Cg+laXiD/j5T6VD3AyyeaSg9aKYwooooEFFJS0DCkoo7UAFFFFMAooooAKKSigAooooAKKKM0AFFJS0AFFFFMAooooAMUUdqKAF7YPSun+Ht4LTXJLbolwMAe9cuKs6fcmy1O2uwceU9RJXQHuFc/wCNUuD4emmtJpIpYuQUPJrcgkEsMcg6OoNMvIFubSWFxlXUiuZbgeMrq2qbR/xMbjp/epf7Y1Uf8xGf/vqqk8TW9zNA4wyOePamZrqSTQFz+2NV/wCgjcf99V2Pw71e4uZLq0vLh5WBzGXOTXA1r+E7z7D4ktpGOEYFW/GpnFWA9iri/iMLyC1t7y1uZYUj+VghxnNdqOeayvE9iNQ0K4gxkgbx+FYJ6geUrq+qp93UZ/xarcHinW4D8tzv/wB81jL935uoJFBrp5UwOrg+IGrR4E8ELAdx1rWtPiLZvhbu2lU+qjivPc0ZJ71LpoD2Cy8T6ReAbLuOMn+FzzWtHIkqhomDr6g14PsXOQoB9RV2z1XUbJg1teS4H8JbiodLsB7bxSgelec6b8QbmIqmpwB06Zj612Gl+JNL1JB5NwqOf4HODWbi0Bbv9KsdQQrd20chI+8RyK4vWPh+ybpdIm9ykn9K9BzRQpNAeGXdpc2MhjvYHhYd2HBqHtXt99p1pqERju4EkBGMkciuE13wJLbhp9IYunUxtyfwraNW+4HF0UsivFI0UqGOReqMORTc1re4BmprA/8AE1tP98VBU2n/APIVtP8AfFKWwHuMZ/dp/uiq+qH/AIltx/uH+VTx/wCqT/dFQan/AMg24/3D/KuRbgeHJ0b/AH2/nTu9MTo3++386fXYtgClpKWmAUUUUCCiiigYUUUUAFFFFAgooooGFFJS0AFFFFABRRRQAUUUUwCiiigQUUUUAIelacP/AB6f8BrMPStOD/j0/wCA0mBjn7x+tFB+8frRTAKKKKACiiigAooooAKKKKBhRRRTAKWkpaACiiigQc0UUUxlKiiiuQQUlLSUAFLSUtMAoopaQCUUUtABSUUUAFFFFMAooooAls/+P+D/AHq6jX/+PGKuYs/+P6D/AHq6fX/+PGKpYHPUtFFMBKWkooAWikooEFFFFABRRRQAUUUhoGLSUUUAFFFFABQaKSgBc0lFFMApc0lFIBc0maSigDpfAetf2ZrX2aVsQXfBJ6Ka9cByOOlfPhB6qSGByCK9d8Ea+usaUI5W/wBKtxiQe3aokgOmIGCD34NeSeOtCOlaobuFP9FuTngcIa9bqlq+mwatp8tncrlHHB9D2qYuwHhdKKtarplxpGoyWV0p3Kfkbsw9qqitQDFLSUUALSUUtMBKKKKACjOKKSgApaSloAKKKKACjNFFABRRRQAUUUUAFFFFABRRRQIStrwloj65rCKyn7NAd0jdsjtWXZ2c9/eR2dqheWQ447D1r2bw7osOh6YlrEAXIzI/941MnYZpxoqKEQBVUYAHal9qXpWP4m1iLRtIluHP7wjbGvfJ71nuM4j4j6wLu+TS4HzFFzLj+9XG0skkk80k8x3Sync596StkrIQUtJS0wCiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAUtJRQAtFFJQIWkpaSgYtaOnf6g/Ws2tLTv8AUn60gMqf/j4k/wB6mU+f/j4k+tMpgFFJRTAKWkopALRRRTAKKKKACiiikAHoa6bwv/x4y1zJ6V03hj/jxloYGXd/8fkv1qKpbv8A4/JfrUNJALSUUUwCiikoAKWkooEFFFFAwooooAKSlpKACloooAKKSigAooooAKKKKBBRRRQMSgruBHrRSgUAem/D3Wftmm/YJm/fW3Cg9WX1rsa8O0nUZdK1OK9hJG04ceq969psLyG/s4rq3bMci5HtWM1ZjHXdvHd20kEwykilTXi+t6XJo+qS2cgITJMRPda9urnfGOgLrWnExAC5hG5D3PtSjKzA8jopzq6O0cqlZEOHU9jTM1uIXNJRSUwJLe4ntZhNazNFIO6nGa0p/E2tXEHkvdFV6bkPNZVJSsgFHUkkknkk9TS0lAoGLRRRTAKKKKBBRRS0AFFJRQAtJRRQBNZXt1YTebZTvE3cA8H61bvde1a/j8u4umCdwh61m0tKyABgcClpKWqAKBRSgUDDgAk9BXffDzQWRW1e7jw78RKw6D1rm/C2hSa5qSggi0hOZG7N7V6/DEkEKRRjCIMKPasakugDsUyV0ijaSRtqKMsT2qQ1xPxC137Laf2ZbN+/mH7zHZazSuwOJ8Saq2sa3Nc5/dofLQfTvWXQFwAB2pcc10pWQC5ooopgFFFFAgooozQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKWgBKWkpaBhRRRSAvWH+pNZtx/x8PWlYf6o1m3H/Hw9CEMooopgFFFFABRRRQAUUUUAFGOaKKYhaKKKBl/Qv8AkKD6Vo+IP+PlKz9C/wCQoPpWh4g/4+UqHuBlmig9aKYBRRRQMSiiigAo70UUAFFJRQAZooopgFFFHegAopKKACiiimACloopAFFJRTAKWkooAWikzRQAuaRhuUj8aKM0gPXPBl/9v8OwSMcsuVP4Vuk1558NL3bNd2LNgAAoK9BrlkrMDybxzZ/ZPEskgGI5wNv1rAr0L4mWJl0+3vVH/HufmP1rz0dAfUZrem7oBcUqsY3SRTgo4bP0pKXqCPUYq2B7bpN2t9plvcr0kQVakUPGyHowIrkfh1fGfR5LUnm2bArrya5GrMDxTXLU2OuXltjCq/y1Qrr/AIkWQh1S3u0HEinefeuPrpg7oApaSl7VYBRS0UAHNKvytuQlG/vL1pKKLAb+keLtU0whHf7RB33nLfhXe6L4q03VlCrJ5Mx/5ZyHBNeSA4NAOH3qSrjow6is5U09gPeaK8s0LxreaaVhv83Ft69XFeiaZqtnqluJrOZXGOVB5X61hKLQFHxB4ZsdaiJZBFcDlZF4JPvXl+saReaNcmG9T5M/LKB8rV7XuzVTUNPttStWt7uMOjDuOV+lVCbiB4jU1h/yFLT/AHxWv4j8L3OhymSMGayJ4YdU+tZFkf8AiZWhB43jmtr3QHuEf+rT/dFQap/yDbj/AHD/ACqxEP3Sf7oqDVP+Qbcf7h/lXMtwPDE6N/vt/On0iDhv98/zp1di2AKSlopiCiiigAooooGFFFFAgooooGFFFFABRRRQAUUUUCCiiigAooopgFFFFABRRRQAh6Vpw/8AHp/wGs01pQ/8en/AaTAxz94/Wig/eP1opgFFFFABRRRQAUUUUAFFFFAwooopiClpKWgYUUUUCCiiimMpUUUVyCCkpaSmAUtJS0gCigUUAFFFFABRRRQAlFLRTAKKSloAls/+P6D/AHq6fXv+PGOuYs/+P6D/AHq6bXf+PKKpYGD2oo7CkpgFFFFABRRRQAUUlFAhaKSigAooooGFFFFABSUUUAFFFFABRRRQAUUUUAFJS0lABV3RdVm0XVYr2EnaDiRf7wqlSUAe9WF5Df2cd1buGjkGQR61ZxXkvgjxIdHuxZXbk2Ux+Un+A16yjK6hlIKkZBHesmrDMTxT4dg16xKEBblBmOSvIbu1nsrp7W7jKTIcEHvXvdc/4p8M2+u224AR3SD5JB3+tNOwHjtFWL6zubC7a1vIykqnuPvfSq9aCCiiimAUUmaKACiiigAooooAKWkpaAEpaKKACiiigAooooAMUd6KXNABinRxyTSpDAheVzhVFEMcs8yQW6GSZzhVFepeD/CUekxi7vAJLxxnnolJysBL4O8MJolp584DXsoyzf3fYV0+KWkPArK4yOaRIYnlkbaiAlifSvG/FmuNrurFkP8AosPyxj+971vePvE/2l20nT5MRqf3zqep9K4cAAYAwKuERBSikpa0AKKKKACiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAtFJRQAtJRRQAtJRRQAVpad/qD9aza0dP/wBQfrSYGXcf8fEn1plPuP8Aj4k+tR0wFpKKKYBS0UUgCiiimAUUUUgFopKKACul8M/8eUtc0e9dL4Z4spaTAzLv/j8l+tQ1Lef8fkv1qKhAFFFFMAoozSUAFFFFABRRSUALSUUUAFFFFABRRRTAKKKKQBRRRQAUUUmaAFpKKKAFopKWmIX69K63wL4h/s66/s67f/R5j8jH+FvSuRpDzjnBHQ+hqZK6Ge/A5FBrivA3icX0K6bfOBcxDCMf4xXaA1g1YZw/jnwsbpW1TTkxOo/eIB94ev1rznOc5BBBwQeoNe/EcVwPjLwcZC+paSmJOskQ/i+lXCQHn9LTeckEFWU4ZT1BpRWohaKKWmMSloooAKKKKBBS0lLQAUUlFAC9aSiigAooooASlopKACiijmmMXNXdJ0y51i/WztVPJ/eP2UUzSdMu9YvFtbFCST88mOFFeu+H9DtdDsVgt1BkPMkndj/hWc5WAsaNpVvpFglrbqAFHzH+8fWr9AqK5uIrW3eedwkaDLE1gBT1zVYdH02S7mYZAwi/3j6V4ze3c1/eSXlwxMkhzz2HpWn4n12TXdQLAkWsZxGvr71jE5reEbagJRRRWggooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAUUtFAwooooAKKKKAL1j/qjWbP/wAfD1pWP+qNZs//AB8PSQhlFFFMAooooAKKKKACiiigQUUd6WmMKKSigDQ0P/kJj6Vo6/8A8fK1naH/AMhMfStDXv8Aj4Woe4GX3ozRRTGFFFFABRRRQAUlFFABRRRmmAUUlFABRRRQAUUUUAFFFFABRRRTAKKKSgApaSigBaKKKAEooooA1PDF59h8R2kpOEY4evZVIIBHQjNeCklcOp+ZWBFe2aJdrfaTb3C90A/KsKq1uA3X7Eahotzan+Ncj8K8WQEAqwwVYr+Ve9MAQQe/FeN+JrMaf4guoAMITuX8aVJ62Ay6TOKM0ldAHS+AL77L4g+zs2EuFJP1r1SvCrSdrW9t7hDgq4/KvcLWZbi3jmQ5V1BFc1RWdwOd8f2P2vw87quZImBH0rysHcoIr3W+gFxZTQsMh0I/SvEJ7drW6mtmGDExBq6T6AR0tFFbAFFFFMBaSiigApKKKAEqxYX13ptwJ7GUxuDnbn5W+tQUlJq4Hqfhrxba6uiwXJEN2OCD0f6V04PY14MMhg6MUdejDqK7zwn4xLFLDV3AbpHN2/GuedO2wHdTQxzRNFKivGwwVI4Neea74Pl0/Uob3TFMltvBaPuvNejqQQCDkHkGlIB4IrNNoBkP+qT/AHRVfVP+Qbcf7hq3VTVP+QZcf7hoW4Hh69G/3z/OlpF6N/vml712LYAooopiCiiigYUUUUAFFFFAg70Ud6KBhRRRQAUUUUCCiiigAooopgFFFFABRRRQAUUUUgA9K0of+PT/AIDWaelaMP8Ax6f8BoAyD94/Wig/eP1opgFFFFAwooooEFFFFABRRRTGFFFFABS0lLQIKKKKACiiimMpUUUVyCEopaKYBRRRSEFFFFAwopaSgANFFFABRRRTAKKKKAJbT/j+g/3q6bXf+PKOuZtP+P6D/erpdd/48o6lgYPaijtRTAKKKKACikooAKKKKACiiigAozRRQAlFFFABRRRQAUUUUAFFFFABRRRQAUUUlAC0lFFAAQCMHpXd+B/FnkFdK1ST5OkMrdvY1wtB5/ofShq4H0CCCOKWvNvB/jM2xTTtXfMfSKY9vY16OjrIgdGDKwyCO9ZNWGZPiHw9Z67amOdQsoHySDqDXk2taNe6JcmG+QlM/LMB8rV7jVa/sba/t2gu4lkjYY5HT6U1KwHgxpK7DxD4CurItcaQTNB18nqwrjnDRymKVGjkHVGHStE7iClpKKYC0UUtACUtFFACUtJRQAtFFFABRRS0AFGKMUo5cIoLOeigdaAEqxp2n3eq3QtrCJpHJwzY4X610eg+Br3Uts2o5trf/nmfvNXpGl6TZ6VbLBZQhABy2OT9TUOQWMrwv4UtdChEjgS3jD55D2+ldHSUE4HXpWYxa4bxv4tW1R9M02QNcsMSODwg9Ki8YeNVhD6fpDhpjw8o6LXneSWLMxZ2OWY9TVxiIMYySSSeST1NFFFagFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKAClpKKACiiigAooooAK0dP8A9SfrWdWjp/8AqT9aQGXP/wAfEn1qOpJ/+PiT60ymAUlLRQAUUlLQAUUUUwCiiikAUtJRQAdjXS+Gv+POSua7V0nhv/jyloYGZd/8fkv1qKpbz/j8l+tQ0gCiiimAUUUmaAFpKKSgBc0ZpKKAFooooAKKKSgBaKKSgBaKSjNAgooooAKKKKBhRRRTAKKKKAAmkpaSgB0ckkMyTQuUlQ5VhXq3hHxPHrVsIJ2CXsYwyn+P3FeT1JbzTW1wlxbSGOZDlWB/nUSjcD3oGjGa5fwp4rg1iJbe6IivVGCD/H7iupFYvQZyHinwZBqe67sAIbwDkD7r/wD1681urWezuGt7uJopl7Edfeveqy9b0Gx1q3Md3EN3aReGH41UZ2A8Vorf1zwnqWjsWVDc2w5DoPuj3rBBDZ2nOOtbJpgFJS0lMQUUUUAFLSUUAFFFFAC4ooooAKKKKACkNO5PSnW8M11MIbSFppT/AAqKNhkJ4BLHArb8PeGb3XZQQphtB96Rh94e1dJ4e8A/Ml1rTbiOVhXt9a7yKGOCIRwoqIvRVGBWcp9gKWkaRZ6PaC3s4wo/ibu1aHSiop5ooImmmkCRqMsx7VkA6SVIo2kkcIijLMegFeWeMfFDaxcGzs2K2UZ5I/5aH/Cl8XeLJNWdrKwYx2Sn5mHWT/61csAAMAYFawh1YDs9hwBRSUVqAtFFFAgooooAKWkooAKKWkoAKKKKACiiigAooooAKKKKACiiimAUUUUALRRRQMSloooAKKSloAu2X+qNZ0//AB8PWjZf6o1nT/8AHw9JCGUUUUwCiiigAooopgFFFFABRRRQAUUUtAF7RP8AkJj6Voa9/wAfC1n6J/yEh9Kv67/x8LUvcZm0Ud6SgBaKSloAKKSigAooooAKSiimAUUUUAFFFFMAooopAFFJRTAWkoNFABRRRQAUUUUAFLSUUALRRRQAoHUe1ej/AA4vfN0l7InLW5z+decCuk8BXv2TxCIScLcjH5VnUV0B6rXnfxLswk1peoOuQ5r0Sud8b2X23w5cBVzImCtYRdmB5NRTVOVB9OKdXWAjDKn6V6z4GvRd+HYVJy8I2tXlArsvhtemK/uLFjxL84/Csqq0A9IryfxzY/Y/ETSAYW5+evWK4j4mWoawhvB1jO386ypu0gPPKKU0ldQBSUUUwCiiigA5ooooAKKKXpQAUuARz/8AqpKWgDs/CHixrZk0/VJMxHiKU/w/WvRVYMoZSCCMgjvXg5wRg9K6vw34yl0yMWmoK00A+4w6r9awnT6oD06sjxRdrZaDdTOcfLge9UpPG2jJB5gl3HH3B1rhfE3iSbXpVRVMVoh+VD1P1rOMG2BhLwv1JNLSUV1ALRRRTAKKKKACiiigAooooAPpRRRQAUUUUCCiiigYUUUUCCiiimAUUUUAFFFFABRRRQMDWjD/AMev/AazjWhD/wAev/AaQGSfvH60UH7x+tFMQUUUUDCiiigAo5oooAKKKKYgooooAKWkpaACiiigAooopjKVFFFcghKWkpaYBRRRSELSUUUDFpKKKACiiigAooopgFFFFAEtp/x/Qf71dLrn/HlHXNWn/H7B/vV0uuf8eUdSwMGik7CimAtJRRQAUUUUAFFFJQAtFJRQAUUUUAFFFFABRRRQAUUUUAFFFGaACikooAKKKKYBRRRQAtFJRQAvBGCMium8MeMLnRnFvdlp7Mn6stcxSUmrge9WGoWuo2y3FnMssZ7g9KsZrwrStXvtHuBNYTFR3jP3T+Fel+H/ABtp+qBYbo/ZbnoQ/Rj7Vm42GdVWRrHhvTNXjK3Vuqt/fQYNawIIBBBB6EU6lewHler/AA/1Cz3SadILiLtGB81crc21zZuUvLd4GHZhXv2PSoLmxtbpStzbxyZ7soJqlIDwVfm6HNLivVb/AMA6NdFniV4nPoeKwbr4cXiZNpfRkdlIqlJCOHoropvBOuxE7YfN/wB2qj+GddTg6ZKad0Bj0VpHw/rYP/ILmpyeGtefppkoouBl0V0EHgrXpiN0Plf71atr8N758G6vowO4A5pcyA4o8ck4qS3ilunCWsLzP6KK9PsPh9o1thphJK49W4rpbTT7K0ULb20SY7hRmk5hY8y0nwJqt9h7wi0i7qw5Nd5ovhbS9HX9zAJJO7yc/lW5RUOTYwxRRmub13xjpukqURxcXA48tD0PvSSA3ru6gs4GnuZVjiUcsxrzLxT42m1DfZaUTFbdGk7t9Kw9a1zUNbmL3cpEY+7GvA/EVmYAGAMCtIx7iGgYz3J6k9TS0tFaAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFBooAKKKKYBRRRQAUUUUAFFFFABRRRQAUUUUAKK0NP/wBSfrWcK0LD/Un60gMyf/j4k+tR1JP/AMfEn1qOmAUtJS0AJS0UUAFFFFABRRRTAWkoooAOxrpPDf8Ax5yVzfauj8Of8eclJgZt3/x+S/WoalvP+PyX61FQAUlFFABRRRQAUUUUAFFJmjNAC0lFFABS0lFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMAooooAKMUUUAORmjkWSN2SRTlWU4Ir0Dwx44V9lnrBCuOFm7H6159RweDyKmUUwPfEdZEV0YMrDII706vHdB8U6horhd5nts8xscn8K9K0TxHp+sxL9nlCy94mPzCsXFoZrsqupVgCp6gjg1zOteCtM1MtLEn2efsU4X8q6eiknYDyLVPB2s6cSUj+1xjvGOlc9JmJ9kytG46qRXv2Kz77R9PvkK3NrGc9woBq1UYHiHXpRXpV98O9Mmy1o8kTn1ORWJc/DvUo8m3vI3HpirU0I5CjNbk3g/XoulsZPcVWbw5rqHnTJTVcyAzM0ZrRHh7XCf+QZLUyeFtff/AJh8i/WjmQGTmlHPSujt/AutzEb2WH/eFbFp8OGyDfXocdwgxS50BwZIX7xxVqy06/v2AsrOSUeoFep6f4N0WxYMkBkYf3zkVuxQQwrthiSMf7IxUOp2GedaV8PbmfbJqs4WM9Y1+9Xc6Xo1hpUIjs4FXH8RGW/OtDFFQ5NgFIabLLHDGZJXCIoyWJ6VxniDx5Bb7rfSQJpenm9VFJJsDpdX1ey0i2M15KF4+VM8t9K8r8ReJrzXZSuWhtAfljHBP1rNvbu5v52nvZWlkJzgngfSoK2jC24DcdgMCloorQAooooAUUUUUCClpKWgBKKKKACiiigAooooAKKKKACiiigYUUUUAFFFFMQUUUUDFooooASloooAKKKKBF2y/wBUazp/+Ph60LL/AFVZ8/8Ar3pIBlFFFMAooooAKKKKACiiimAUUUUALRSUtAF3Rf8AkJD6Voa7/wAfC1Q0X/kJD6Ve1z/j4Woe4zOopKKYBS0lFABmkpaSmAtFJRQAtJRRQAUUUUAFFFJQIKKKKBhRRS0AJRS0lABRS0lMAooooAKKKKAFopKWgQVLa3LWl5BcqfmjcVDSHkYpNaDPdbWZbi2ilU5DIDmlmjEsLxsMh1Irlfh/q63mlfYpG/f2/XPcV1orlaswPD9WszpWq3FlPldrEqcdc1WDx/3v0r2+90qxvyDdW0bsP4iOapnwvo5/5dFrRVbIDx3en979K6n4e2stxr/2uMHyYUKs2OM12/8Awiujf8+orTtLO3sYfKtYljT0UdaUql1YCxXI/EiVV8O7CfmaQYFdYTXl3xA1Vb7VUs4WzHbjD+5qYK7A5c0UlFdQC0UUUwCiiigAooooAKKKKAFopKWgAzRnikooAML/AHV/KjNFFIAooopgFLSd6KAFooooAKKKKACikpaAEpaSloAKKKKBBRRRQAUUUUAFFFFMAooooAKKKKACjpRRQMD0rQi/49f+A1nnpWhD/wAe3/AaTAyT94/Wig/eP1opgFFFFABRRRQAUUUUIAooopgFFFFAC0UUUAFFFFAgooopjKVFFFcogooopAFFFFABRRRQAUUUUwCiiigAooooAKKKKAJbT/j9g/3q6PXP+PKOuctP+P2H/ero9b/48o6kDCoo7UUwCiiigAoopDQAtJRRQAUUUUAJS0UUAFFFFABRRRQAUUlFABRRRQAUUUUAFFFFMAooooAKKKKACiiigBMUhGTnkHsw6inUlAG5ovizVtHwiy+fbjqj8tXf6N430rUgEmf7LMf4JO9eSgUYBOcc+vepcbhc+gUdZEDIwZT0INOFeHadrmqaY4azu2wO0hyK6nT/AIjzR4TUbMyHu0fFQ4sZ6RRXNWXjfRLoAPciBj/C9bVvqdjcjMF1G+fQ0rAW6KQMD0IP40Z+lIBaTn1paKAEoxQTgckfnVa41Gzthme5jQD1NFgLOKTOK5298a6Hag7btZmH8K1zl/8AEksCunWTKezvyKaiwPRC4VSzEADqSa57V/Gmk6apVZRcyj/lnGea8z1LxDq+qE/a7oqvpGcVlgDOcZPqetWoCOj1rxjqmrbo0f7Pbnoq8N+dc/33ElmP8R60maKtJIBc0lFFMAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFFABRRRQAUUtFACUUUtACCr9j/qT9ao1esf9UfrQBmT/wDHw/1plST/APHw/wBajoAWikpaACiiimAUGiigAooooAKKKKADtXR+HP8AjzkrnO1dF4d/485KTAzbv/j8l+tQ1Nd/8fcv1qGhAFFJRQAtFJRQAtIaKKACiijpQAUUUUAFFFFABRRRQAUUUZoAKKKSgBaKKKYBRRRQAUUUUAFFFFABRRRQAUUUUALnFOjd4nEkLtG4/iQ4plLRYDrdG8eXtkFi1GP7TEOAV6iu50vxJpeqKv2e6QSH/lmx5FeMZoHynKFkPqhwazdNDPfs8Uma8a07xRrOm4ENyJIx/DJya6ew+I0ZAS/snDd3XpUODA7+jHpWDZ+LtEuwMXqIx/hateC8tZ1zDOjj2NTYCfB9aMH1NGfp+dLSAbz60YPqadSH8KADn1oxSFgBkkD8ap3Oq2FqD9ou40x6miwF2lrlr7xzo1rnyZftDDslc5f/ABEvJsrp1t5A9ZBmqUGwPSJpooIzJNIqIOpJrldZ8eadZZjsgbuXplOgNedX2qahqLl726dieynAqoMD7oAq1T7gaer69qesSZu5ysfZIzjj3rMGAPlAAoorVJIAooopgFJRRQAUtJSigAooooAKKKKACiiigQUUUUDCiiigAooooAKKKKACiiigAoFFFMQtJS0UDCiiigQUUUUAFFFFAFyz/wBUaz5/9e1X7P8A1RqhP/r3pIBlFFFMBKWiigAooopgFFFFABRSUtAC0UlFAF7Rv+QkPpV/XP8AXrVDRv8AkIj6Vf1z/XrUvcDNzRQaKBhRSUtACUtJRTAKKKKACiiigApKWigApKWkoAKKKKACiiimAUUmaKAFopKBQAtFHaigAopaSgAooooAKSlooAsaffXGmXqXdo22RDyOzD3r1DQfFun6rEqySCC46GNz1NeUUDg5BKn+8OtRKCkB7yDkZBGPrS14lBq+p2y7YbyTH+02an/4SPWu15WXsmB7LTZHVELOwVR1JPSvGz4i1o/8vlVrnU9Quxi4u5CPRWxR7Jgdx4o8ZwwQvZ6S4lncYaUdEFedncWLOxZmOWJ6mlAA6D6n1oraMFEBKWiirAKKKWgBKKKKACiiigAooooAKKKKACiiigAooooAKKKWgBKKKKACiiloASilooASiiigBaKSloAKKKKBBRRRQAUUUUwCiiigAooooAKKKKBhRRRQAGr8X/Ht/wABqhV6L/j3/CkwMs/eNFHc/WimAUUUUAFLSUtACUUUUIAooopgFFFFAhaKKKBhRRRQAUUUUwKVFFFcogooopAFFFAoAKKKKACiiimAUUUUAFFFFABRRRQBJa/8fkP+9XSa2f8AQo65y1/4/Yf96ui1v/jzjqWBh9qKO1FMAopKKACkpaKAEpaKKACiiigBKWikoAWikooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRk+tFFACYU/wAK5+lKrOpyk0qf7rYpKKQFqLUdQh/1V7MMermraeItbT7t6fxrLoosBsjxRr3/AD+Cmt4k1tx817+VZNFFkBcl1PUZv9bey/g2KqO0j8vPM/8AvNmkzRTsA3ao/hX64pfpS0lABRRS0wEpaKKACiiigAooooAKKKKACiiigAooooAKKKKAClpKKACiiigAooopgFFFFABRRS0AJRRS0AFFFFACUUtFACUooooAKvWP+qP1qjV2y/1R+tIDOn/17/WmU+f/AF7/AFplMAoFFFABRRRTAKKKKACiiigAooooAO1dD4d/485K57tXQ+Hv+POSkwM27/4+5frUNS3f/H3L9aioAKM0UUAFFFFAwopKKBBRRRQAtFFFABRRRQAUUlFAC0lFFMAopaKACiiigAooooAKKKKACiiigAoopaAEooooAKKKKBhRRRQIKPrzRRQABV/ugfQVIkkifcuJl+j0yiiwFyPU9Ri/1d7L+LVZTxHrSfdvPzrKopcqA2D4o13H/H4tRP4h1mQfNen8KzKKOVATy39/N/rL2b8HNV2LMfnlkf8A3mzRRjinZDECr2VR+FL+NL1pKACjNFFMApaSigQtJS0lAC0UlFAC0UUUDCiiigQUUUUDCigUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUtFMQUUUUDCiiigQUUlFABS0UlAy5Z/wCrNUJv9e1X7P8A1RqhN/r2oER0tFFABRRRQAUUUGgAopaKYCUUUtACUUUUDL2j/wDIRH0q/rn+vWqGj/8AIRH0q9rn+uWpe4Gd3pKXvSUAGaKKSgBaKSimAUtJS0AJRS0lABRS0lABRRRQAUlLRTAKSlpKQBRS0UAFFFAoAKKKKACiiimAUUCigAopaKAEpaKSmAUUtFACYopaKAEooooAKBS0UAJS0UUAFJS0UCEopaKBhSUtJQAUUUtACUUUtABSUtFACUUUtACUtJS0AFJS0UAJS0UUAFFJS0AFFJRQAtFFJQAtFFFAgooopjCiiigQUUUUDCiiigAoopO9ABV+L/j2/CqBq/F/x7fhSYGX3NFB6mimAUUUUALSUUtACUtJRTAKKKKBBRRRQMWikpaACiiigAooopgUqKKK5BBRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAktf+P2H/AHq6LW/+POOuetf+P2H/AHq6HW/+POOpYGHRR2pKYBS0lLQAUUUUAFFJRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUwCiiigAooooAKKKKACiiigAooooABRRRQAUUUUAFFFFAAKWkpaACiiigAooooAKKKKAEpaKKYBRRSUALRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFLRQAmKWiigAooopgFFFFABRRRQAUUUUAFFFFABV2y/wBUfrVKrtn/AKo/WkBnTf69/rUdPm/17/Wm1QBRRRSAKKKKYBRRRQAUUUUAFFFFAB2roPD/APx5yVz9dD4f/wCPOSkwMy7/AOPuT61FUt3/AMfcv1qKhAJRS0lAC5pKKKACiiigApaSigBaKM0lAC0lFFABRRRQAZooopgFLRRQAUUUdqACiiigAooooAKKKKAFopKKACiiigYUCiloEFFJS0AJS0lLQAlLSUtABRRRQAUUUUwEpaSloGFFFFIQUUUUDE+lFLRTASloooEFFFFAwo7UUCgApaSigAooooEFFFFAwooooAKKKWgApKWigBKKKKACiiloAKKKKYgoopKAFooooAKSlpKACiiloGW7T/V1nzf69qv2h/d1Rm/1zUkIjpaKKYBRRRQAUUUUwFooooAKKKKBid6WkooAvaP/AMhEfSr2uf65KoaP/wAhEfSr+t/69al7iM00UGigYlFLSUwCiijFABRRRQAUUUUAFJS0UAFFFFABRRRQAUUlLQAUUUUAFFFFABRRRTAKKKKACilooASlpKWmISilooGJRS0UCCiiigYlLRRQAUUUUCCikooAWiiigAooooAKKKKBiUUUUAFFFLQAlFFLQAlFFLQAlLRRQAlLSUtABRRRQAUUUUAFFFJQAtFFFABRRRQAUUUUwCiiigQUUUUDCiiigAooooAQ1ei/49/wqlV2P/j3/ChgZh6mig9TRQAUUUUAFFApaYCUUUUAFFFFABRS0lAC0UUUAFFFFAgooopjKVFFFcggooooAKKKKYBRRRQAUUUUAFFFFABRRRQAUUUUAS2v/H7D/vV0Ot/8ecdc9a/8fkP+9XQa1/x5x1IGHS0nalpgJRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigBaKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUCgBaKSigANLSUtABSUtJQAUUUUAFAoooAWiiigAooooAKKKKYBRRRQAUUUUAFFFFABRRRQAVctP8AVH61Tq5af6s/WkBnTf65/rTKfN/rn+tMqgFopKWgAooooAKKKKACiiigAooooAO1dB4f/wCPR65+ug0D/jzkpMDNu/8Aj8l+tQ1Ndf8AH3L9ahoAKKKKACiiigAooooAWkoooAWkoooAKKKKACiiigAoFFFAC0UUUwCiijFABRRRSAKKKKAClpKKACiig0wCiiigYUtJS0CEooooAKWiigBKWiigAoopKBhRS4opiEpaKKACiiikAUUUUAFJS0UwEopaKACikpaAE6UtFFAwopaSgQUUtFAxKKWk70AFFGKWgAooooEFFFFACUUtJQAtFFFABRRRQAnejvS0UAFFFFMAoopKBhS0UUAWrX/V1Qm/1zVftf8AV1Qm/wBc1IQ2iijvTAKKKKAA0UUUALRRRTAKKKSgBaKKSgZe0j/kID6Vd1r/AFy1R0j/AJCA+lXta/1y1L3EZx60UUUDEopaSgAoozSUwFoopKYAKWiikAUUUUAFFFFABRRRQAUUUUAFFFJQAtJRRTAKUUlLQAUCiloASloooAKKKKYBRRRQAlLRRQIKKKKACiiigAooooAKKKKACiiigAooooASjtRRQMKKWigBKKKKACilpKAAUUvakoAWiiigAooooASloooAKKKSgBaKKKACiiigBKWkpaACiiimAUUUUCCiiigAooooAKKKKACrsf8AqPwqlVyP/UfhQwM09TRQepooGFFFFAC0lFFMAooooAKKWigAooooAKKKKACiiigAopaKYFGiiiuQQUUUUAFFFFMAooooAKKKKACiiigAooooAKKKKAJbX/j8h/3q6DWv+POOuftf+PyH/erf1r/jzjqWBi0lHYUUwCiiigAooooEFFFFABRRRQAUUUUAFFFFAwooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRS0AFFFFABRRRQAUUUUAFFJRQAtFFFMAooooAKKAKMUAFFGKKACiiigQUtJRQAUUUUDCiiigAooooAKKKWgAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFW7T/AFZ+tVKt2v8Aqz9aQFCb/XP9aZTpv9c/1plMBaSlopgFFFJQAtFFFAB2ooopABooooAK39A/49JKwDW/oH/Hm9DAzbr/AI+5frUVTXX/AB9y/WoaACiiigAooooAKKKKACiiigAooooAKKKKACiiigApaSigBaKKKACiiigQUUUUAFFFFMYUvakFLSASiijvTAKKWigBKKKKAFopM0ZoAWkozRmgAoo7UUAFLSUCgBaKKKACilooASloxRQAlFLSUwCiikoAKKKKBhRRRQAtFFFABRiiloAKKKKBBRRRigAooooAKKKKACiiigAooooAKKKKACiiigAoopKYxaSlpKAFpKWigAooooAs23+rqjN/rmq/bf6uqE3+uakhDaKKWmAUUUYpgJS0UUAFFFFABQaKKACikpaALmkH/iYD6Vd1r/XLVLSf+QgPpV3Wv9ctSwM+ijvSGgYUUUUwCkoooAKKWigAooooAKKKKACiiigAooooAKKKSgAopaSmAUUUUAFLRRQAUtJS0AFFFFABRRRQAUUlLTEFFJRQAtJRS0AFFFFABRRRQAUUUUAFFFFABRSUUDClpKWgBKWiigAooooASiiigBaKSigBaSlooASlpKWgAooooAKKKKACiiigBKWiigAooopgFFFFAgooooAKKKKACiiigAooooAKtx/6j8KqVbj/ANR+FJgZx6mig/eNFMYUUUUwCiiigApaSigAxRRRQAtFFFABRRRQAUUUUAFFFFMClRRRXIIKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFAEtr/AMfkP+9W/rX/AB6R1gWv/H5D/vVvaz/x6R0mBi9qKO1FABRRRQAUUUUAFFFFABRRRQAUUUUAFFJu5xsc/Rc0oDn7sMx/4AaADNJmpks7yXiO0lP1U1ch8O65cY8rTzz6nFFwM6iujt/AuvSkebEsQ/3q1rX4azNg3WolfVQKXMgOFLAdx+dIDnpXq9j4B0W2wZ4zOw7k1Q8YeDIpbT7XpEXlyxD5ox/EKXMB5xmgU0kgkOCrKcMD2NTR21zIu6O2kZfUKaq4DKKmFlff8+kv/fJpwsL4/wDLpL/3yaLgV6Ks/YL7/n0k/wC+TUckE8LBZoWRm6AjrRcCOkqz9ivD0tJf++TR9gvf+fSX/vk0XArUVZ+wX3/PpL/3yaPsF9/z6Sf98mi4FakpzKVJVhgjqKSmAlGaKsWFhealcCCwhMsh6+gpAV8jucUoIPQ13+kfDpcLLq1xuPeEdvxq34o8HWKaG8mlweVNAN2c9QKXMB5tRSK25QcY9qWqAKMUoyWVR1Y4FbieENfkRXSzBVhkHd2pXsBhUVvf8Ib4i/58R/31Sf8ACG+Iv+fEf99UcyAwaK3T4O8Q/wDPiP8Avqk/4Q7xF/z5D/vqjmQGHRW5/wAIb4ix/wAeI/76o/4Q3xF/z4j/AL6o5kBh0Vt/8Id4i/58h/31S/8ACHeIf+fEf99U+ZAYlFbg8G+IT/y5D/vqsq8tJ7C7e1u02TJ95fShMCGiipba2nvJxDax+ZKei0xEVJWn/wAI/rX/AD4Gj/hH9a/58DSuhmZRWn/wj+tf8+Bo/wCEf1v/AKB5ougMykrTPh/W/wDoHmk/4R/W/wDoHmi6AzqStE+H9a/58DVS8sruwdUvYfKZuQPWi4EWaKltLO7vnZLKAysvJHpVsaDrJ/5cGouBn0taI0DWv+fA0v8Awj+tf9A80XQGbRWl/wAI/rX/AEDzR/wj+tf9A80XQGbRmtE6BrX/AD4Gk/sHWf8Anwai6Az80VNdWd1YyCO8h8pz0HrUFNALRRRQAUUUUwCiiigAooooAKKKKACrVr/q/wAaq1atfuH60AUJf9c/1ptOm/1z/Wm0AFJRS0wCkopaACiijFABRRRSAKKWkoAK39B/49HrA7Vv6D/x6PSYGbdH/S5PrUVS3X/H1J9aipoAoopKAFoFFFABRRRQAUUUUAFFFFABRRRQAUUHHcgfjSZH95fzoAWikyP7y/nS5H95fzouAtFJkf3l/OjI/vL+dAC0UZH95fzo4/vL+dABRRx/eX86Tj+8v50ALRRx/eX86OP7y/nQAUUmR/eX86XI/vL+dABS0ZH95fzo4/vL+dMAoo+X+8v50v0OaAEpKWkJH91z9FoAKSnBZGPyQzH/AIAamjsL+X/V2ch+qmldAV6K04vDmuzH93p/5mr8HgfXpf8AWQrF/wACpcyA52kyB/EPzruLX4bzvg3WoFfVQK3LLwDo1vhp4zOw7k4pOaA8sDA9KUV6F4v8HRNa/bNIi2SRD54x/GK87HfPBBwQexpxlcB4oqRLW5dQy28hU9CFNO+yXf8Az6y/98mndAQ0tTfY7v8A59Zf++TS/Yrz/n1l/wC+TRdAQUVIYJhIIzC4kPRcc1J9gvv+fST/AL5NF0BXpKnNleDraS/98mmG1ux/y6y/98mndARUVIbe6AybaUAdflNRZp3AWikooAWijmkzjqQPqaBi0tM3qO+fpzUsEFxcsFt7aVyfVSKVxDRR9SBW/Y+DNcu8F4BAh/iJrptO+HdrEytqFybj1XGKlzSGedjB6HNLiuw8Y+El02MX2lxn7OOJIxzt9644kFcg01K4Bkf3gPxoyv8AeH516D4Z8KaRqWhQXVzAWlfOWzWmfAehn/liR+NT7RAeV5H94fnRXqZ8B6GoJMJOAe9eZahGlvqVxBGMIjYUU4zTEQ0lNLAdfyHWtEaHqx07+0PsjfZx37/lVXQFClpo56U4UxhRRRQIKKKKACiiigYlLRRTEJS0UUDCiiigQUUUUDLNv/q6pTf65qu2/wDq6oy/65qSENopaKYCUtFFMAooooAKKKSgAopaDQAUlFFAy7pH/IQH0q7rP+uWqWkf8hAD2q5rP+uWpe4GeTSUtJTAKKKKAClpKKAFooo60AFFFFACUtFFABRRRQAUUUUAJSmkpRQAlFFFMAooooAWlpKKAFooooAKK1/D+gT6886wy+UIhndjr7VLfeE9asgWNsJIx/EDzU86vYDDNJSyK8TlZY3Qj1XFMyOxBp3AdRWpovh691uF5bRgqp1zWgfAmtDoV/OlzpAc1S10n/CCa16r+dKPAmtf3l/Ol7RAc2KMV0w8B6x3dfzrF1TTLjSbz7Ndff7H1pqaewFPFGKWkqgCiijNABRSUtMAoooyByegoAaWA65oBHqPzrvfAmgW13pElxqEAk8xiFB9K0b3wDpU3Nrut29uaxdVXsB5kKWuuvfh/fw5NpcicdgeK5m9s7jT7pra7QJMvJANXGaewFek3Y/hf/vmkkPyfiK9j0vSrCXS7Z2tUJMYycUpz5QPHdw/uv8A980m4f3X/wC+a9s/sXTf+fSP8qin0bTVgkItE4Q9vao9qB4zmikfiaYDoJGA/OgVsmAtFFFMAoopaACiiigAooooAKKKKACiikoAKWkpaACiiimIKKKKACiiigAooooAKKKKACiiigAq1H/qfwqrVpP9T+FDGZ/c0UHqaKACiiimAUUUUAFFFFABRRS0AJS0UUAFFFAoAKWikoAWikooApUUUVyiCig0UwCiiigAooooAKKKKACiiigAooooAKSlooAltf8Aj8h/3q3dZ/49I6wrX/j8h/3q3tZ/49Y6QGL2ooooAKKKKACiiigAooooAKKKKACgjII6ZoooA9I+HrWOoaY0UttGZ4G25I+8PWu1W1tlHywRD/gAryHwVq8Wka6HuX228i7SfevVodVsp2VYZw+7pjms5bjLYiQdEQf8Bp2AOgA+lZmra/pujso1CcxbhkcZzVbRPFWna5fSW1gzMY13EkYqQNvHuaKd1FU9VumsdMubtU3mFCwX1oAtUh+oH1rzO08eazrF2llp9gqSyHG4HO33rZ1rQPEstqJbPWWecD5oiMA/jTsBnfEDwxE0T6rYGNWXmdA3UetdB4EdZ/Cts5VDyRyteTahLqqSvb6jJMJAcMGzg16t8OgR4Sg/32pvYR0wRT/An/fNL5a/3F/Kl/A0uakYmxf7q/lXAePAB4m0YBV+/wBhXoNef+Ov+Rr0Uf7VNAd4iLsX5E6D+GneWv8AcX8qE+6v0FOpAN8tf7q/lTJUXyn+RfuntUtMl/1T/wC6aAPCdT/5C13/AL9Vat6p/wAhi8/36qfU4rZCENdJ4C1YaZroglIEFxxk/wB6ubOPUfnTS5jZZVYbom3jnuKGroD6DpHVXQo4yrDBHtWb4c1Aanodrc7gzlBvwehrTPTrishnifijTX0nxBPARiOUl4/pWVmu8+KP2OSK3mjnja7Vtu1Wz8tcFkHuPzrSL0ESRnE0P++K9303/kHWx/6Zj+VeCKf30XIPzCvedMP/ABLLX/rkP5VMwLlJTScAn0FcfY+PLe710aWLUqxlMe7NQM7KikrH8Sa/F4fs47iaMuHbbigDZoxWN4Z8QReILNriGIxhTjBraoATHvRXI6346tNI1aSwlgLOmOfrXVwSiaCOUDAdQw/GgB1eN+OD/wAVdd/hXsprxrxz/wAjdefhVw3EzCBrpfh8f+KsiH/TNq5muj+Hx/4q6H/carlsB7Bgeg/KjA9B+VApaxGJgeg/KjA9BWNrPijStFuUgv5mSRxlQFzxWf8A8LB8Pf8APy3/AHzTsB1OB6D8qMD0H5Vzdp440K7uo7eG4YySHCjb3rpaQDSBg8D8q8x+KPGrWmP+edennpXmXxSH/E1tP+udVHcBvwt51S6/3K9OCjA4H5V5n8LR/wATS7/3K9OHSiW4CYHoPyowPQflRXL3XjvQrW6kt5p3EkbFWG3vUoDqMD0H5UYHoPyrk/8AhYXh/wD5+H/75q3pnjHR9UvVtLSZmlboCuKdgOhwPQflRgeg/KgUUgPLficf+J1bj2rkK674nf8AIct/oK5LvW0dhBRRRVgFFFFABRRRQAUUUUAFFFFABVq2/wBX+NVas233D9aQFGX/AFz/AFptOl/1r/Wm0wCiiimAUUCikAUUUUAFFFFABRRRQAdq3tB/49HrCNbuhf8AHo9DAzbr/j7k+tR1Ldf8fUn1qGhALSUUUwFooopAFFFFABRRRQAUUUUAFFFJ3oA7DwV4b0/XNOee9QmQNjOa6T/hX2if3D+dU/hd/wAgeX/fNdwKxk3cZyX/AAr7RP7jfnR/wr3Rf7rfnXXUVPMwOR/4V9ov91vzo/4V9ov91vzrrqKOZgcj/wAK+0X+4350v/CvtF/uN+ddbRRzMDkv+Ff6L/cb86P+Ff6J/wA82/Outoo5mByX/Cv9E/55t+dH/Cv9E/55t+ddbRRzMDkv+Ff6J/zzb86X/hANE/55n866yijmYHKf8IBon/PJvzo/4QHRP+eTfnXV0UczA5T/AIQDRM/6s/nXAeJLGDTNbezthiIDIFe014943P8AxVUn+7VwbuBhkbhtzjNeleBGsdT0kebax/aITtOR1HrXmma6LwRrMGj6rKbxykEqYz6GrnsI9XS1t0HyQRj/AICKfsUdET8qo6frVjqLlLSRnI7lSBUWq+ItM0l9l/OY2xx8vWsBmoB6AD6UY9z+dc/oHiyz17Up7WzQ7Yl3bz3rosUANoqlrd4+naRcXcab2iXIX1rg9K8Z63rOpRWVtbLGZD8zZ+6KaVwPSPw/OvOPHfhlI2bVNPMYXcDNHu6e9bevaZ4p8ktpmqeYccxlcV5pqB1SOdk1FpxJuG7Odpqogey+HVVtDtTtT7n92tLy0/uL+VZvhsFdBtAQc7K1c/Wpe4DfLT+4v5Unlr/cX8qfSUgOB1NgnxLs/lX7hGMcV3mxf7q/lXnuusV+JNjjuteh9+hqmAmxf7i/lSeWn/PNP++adS1IFLU0UaZdEIn+qb+H2rwv+J/9417xqf8AyDLr/rk38jXhGPnf/eNaUwJIIJrmXyrdPMk/u1sWvhHXrkArZhVPctis7TL2TTtSgvIzjYw3+617fZ3CXVpFcR/dlUMMVU5NAeb2vw71CXH2m7EPqBzWxafDmwjIN1cNP+ldvRWfOwMG08JaHaEGKyG4dyc1sRW8MKhYoo1A9FqakpXYCYp1NpGdVGWZV+pxSAJY0mjaORQyMMEHvXkXi/QJNDvS8YJs5iSjf3T6V6jd6tYWaFri6iUD0YGuU8Q+MfD13YS2bZuQ6kY29DVRugNjwGd3hW1Pua6KuX+HjZ8KQY6B2x9K6ek9wGyfcb/dNeIXttc3niK6gtIjLKz8ACvb3+63+6a8Y/tW70TxTdXVoQRv/eIR94VUAO08MeB4LEJd6rie56hO0ddnsXbt2rtxjGOKz9C1i11qwS6tWHI+Ze6mtOpbYHn3jHwft36jpMeCOZYh3964PqPT1B7V74QCCCMivOPHPhj7M7app8f7sn97GOx9a0hPowOKopoORkUtbALRRRQAUUUlAC0UUUAFFFFMAopKWgAooooAsW/3Kpy/65quW/3Kpy/61qQhlFLSVQBS0lLQAUd6SloASloFFABRRRQAUUlFAy7pP/IQH0q5rX+uWqWk/wDH+PpV3Wv9ctT1EZ560lKaKYxKKKWgQlLSd6WgYUUUUAFJS0UAFFFFABSUtJQAUUUUAFFFFMAooooAKKKKAClopKAFprNtXP4UGtfwppbavrsaFcwwHdJ7ipk7IR6L4K0v+zdBiDriWX52PseldBikRQqhVGFUYH0p1crd2MpX2lWN+my6tkcfTFebeN9G0zRXhFj8skxOY85xXqjsEQsxwFGSa8W8SakdW1u4uR/qx8iD0xVwvcDsPhhzp9z9a7nHufzrh/hgMafc8dxXc5qZbgGPc/nS4pM0tSAYrhPiXYM1vBqKL/qflb8a7uqWr2Kajps1rIMh1OPr2pxdmB4lmilnhktLmW1mGJIWINMzXYncBaKKKAClpKUUAFSW1s97dw2cYJaZgv0qPjqeldv8OtGLyPq1wnH3IwR+tTOVkB3Om2i2NhBbIABGgB+tWqKK5AIbq4jtbaS4lOEjXcTXiuqXr6hqU905J3MQv0r1vxFpU2sacbSK6NuGPzkDqPSuS/4Vu4AA1M4HH3a0pyS3A4OU/u/xH869w0U50e1/65iuIf4bSsuBqZ6jtXe2Fv8AZLKG3LbjGoXPrTqSTAsVDcf8e8v+4f5VNUcq743TONykVkB4O/8Ar5/+ujfzoFdy3w2YyOw1RvmYt931pP8AhWz/APQUP/fNdKqIDh6X8K7j/hW0n/QUP/fNKPhu/fU2/wC+aftYgcPiiu6/4Vuf+gm3/fNYviXwx/YFvHN9rM+84IIximqiegHP0UGirASlopKAFpKWkoAWiikoAWiiimAUUUUCCiiigAooooAKKKKACiiigAooooAKtR/6n8Kq1ZT/AFP4UDKB6mig9TRQAUUUUwCiiigQUtJS0DCiiigBKWikoAWigUUALRQKKBBRQcUUDKVJRRXMIKKKKACiiigAooooAKKKKACiiigAooooAKKKKAJLX/j8h+tb2sf8eiVg2v8Ax+Q/71b2sf8AHpHUgYwooopgFFFFABRRRQAUUUUAFJS0lABSGlpp45oAkt2jW8tzOu6LzBvX2r2/SbDToLWOSxgCI4DLzmvF10XVbiEPDaMyuPlOOteueDzeL4fghv4vLlhGwD1FRIB3i3RodZ0aVHTMsal4z3z6Vw/wtDR+ILuKRcSJFhhjvmvT55PKhaTbu2jOPWuG8H2N6njLUL6a08i3nUlanoM78dKy/Ev/ACL1/wD9cTWoOlZniSOWXw9fRwrukaIhR6mkBxXwm04eRcaky/M/yAmvR8cVgeCNNk0zw3BBMmyQ/Mw9M1uzMyRMyLuYDIX1psDhfiXf2CWa2LQJJeSdCOqD1ribDxBrWm2i2llemOFei471e1zRfEVzfXGo31pwWOCDnC9qzrPRtVvZBHbWbkn++MVatYRYk8X+Iwmf7RYfhXq3hm6mvNBt57h98rD5m9a8l8ReHb7QoIHvMfvjgYOcGvVfB/8AyLVr9KUrdBm3mvM/ifLJFrGnyxNtkQ5VvSvTMV5j8UUkfVrBIkZ3boqjNStwMP8A4S3xH/0Ejx/s1oeHvE+u3fiG0t7m/Z4n+8uOtGm+AdWvI/Nu3FspGVAOSazvD9tLZ+Nra2nG2SMkEVegj2wU2X/Vv/umnDpTZf8AVP8A7prMZ4Vqn/IZvf8Afq94V0mHWtaFncsVj2FuPaqGqf8AIZvf9+t34c/8jWP+uTVq9hHUf8K60v8A56vQfhzpR6yPXZ0tZ3YzJ0DQ4NDtHt7d2aNm3YPart7btc2zwrKY9wwWFWaKQHES/DjTpZDI9zKzt1JOaYfhtpva5cV3NIelPmYHjnjDw/b+H7q1W2lL+YwzmvWNM/5Bdp/1yH8q4D4qf8flj/vCu/0z/kF2n/XIfypsCy/KNj0NeS6JoWqx+NlupLQrCLgsW9s164KXA9qm9gE7muT+IemXup6VbxWEIldXJIPpiusp1O4HI/DvSr3StIeK/hEUhYkD2rrqKKQHlnjHwxrOo+J5rq0tFkhO3DZ9K9MsUaOygRxhljAI98VPRTuAV4345H/FW3f4V7JXjnjj/kbLv6CnDcDniK6H4ff8jfD/ALhrnyK6DwB/yN8H+4a0lsI9hFFFLWIzkvFXgxfEV/DdG8MBjXbjGc1ij4WoP+Ymf++a9FLKOrAfU0b0/vr+dO7A4TTvhwtlqEF1/aJfymDbdvWu+pgZT0ZT+NOFK9wFPSvMvij/AMhS0/6516YeleZfFH/kK2n/AFzpx3AX4Xf8hO7/ANyvTR0rzL4Xf8hO6/3K9NHSiW4CHrXAaj8N473UZrv+0SnmsW27eld/TSy5+8v50kB53/wq6P8A6Cbf981paD4DTR9UjvhfGUp/DtxXY71/vr+dKCpPDKfoadwHClooNIDyv4m/8hy3+grku9db8Tv+Q3b/AErke9bR2EL2oooqwCikpaACiiigAooooAKKKKACrNt9w/Wq1WLf7n40AUpP9a/1ptOl/wBa31plABS0UUwCiiigAooopAFFFLTAKKKKAA9DW5oX/Ho9YfatzQ/+PR6TAzrr/j7l+tRVLdf8fcn1qKhAFLSUtMAooFFIAooopgFFFFABS0lLQAlJ3p1GKQHonwwZU0eXc6r854JxXb+dF/z1j/76FeCpLPEMQzNGPQGlN3e/8/cn5msnC7A9486L/nrH/wB9Cl86L/nrH/30K8E+13v/AD9yfmaPtd7/AM/cn5mlyDPe/Oi/56x/99Cjzo/+esf/AH0K8E+13v8Az9yfmacLy9/5+5P++jRyAe8+dF/z1j/76FKsiOcK6k+xzXgpvbwDJu5PzNek/D3SbqCybUb6R2e45RGP3RScbAdnRS0VICZphljU4aRAfQtVbVr+LTNOmu5jgIpx7ntXiV3ql/f3kt3LcurStnaD92mo3A928+L/AJ6x/wDfQo86L/nrH/30K8E+03n/AD9Sf99Gl+13n/P1J/30arkA9686L/nrH/30KPOi/wCesf8A30K8F+13n/P3J/30aPtd7/z9yfmaOQD3nzos/wCtj/76FeP+NiD4pkIYEbeoOaxhd3v/AD9yfmajZndt0jF29TVRjYBc1NZyRJqFs1wu+EON6+oquTgZNXV0XVpo1eKzZlbBU461bA9s0+K2S0iNrEscbKCAB2rO8V6JFrWkSxGMGdAWib0NS+GZbmXRYBeQeTLGoQr9O9akr+XGz4ztGcVh1A8v+F0bReINQikGHRNrcdxXqVcF4Nsb6Lxdqd3c2vkxTD5ffmu9obuBj+LP+Ravv+udcn8LdN/d3GpOufM+VD6V13imGa48PXkNum+V0wq+tR+EbCTTfD1tbSpskHLD3ovoBtVwfxH1eCO2XTI4kkmkOXbHMfpXcTuY4HdRuKg4Hqa8f1fSPEN7fXOo3VmBvJyc9AOlOO4EMPivXraBIYb3EaDCjb0ofxp4jAH+n4yQPu1U07RdU1V9tlasVzgs4wKj13R7vRLmO3vB8zkFSOhq9APaNDuJbnR7aadt0jrlj61oelZXhvnw/Z/7lamMYrJgeU/EGaW38V289u+yVV+VvSss+KvER5OpN/3zWn8RIZZvFNvHBG8jsvAUZplj4B1a6gaW6YW/y7lUHJb2rVWsBb8Ea/rF94mjtr29MsJjJK4716jXj3gOGS38ceRMu2SNCrCvYKiW4FbU/wDkG3X/AFyb+Rrwn+N/9417tqX/ACDbr/rk38jXhJ/1j/7xqqYC4DAqeh4r074d6qLnRmtppAJIGwoJ/hrzDvUkcs8JJglaMng471co3A91e7tkGXuIlHu4rNu/E+jWgJmvV467ea8ZfzJOZZZG/wCBGmGJAhwG/E5qPZge7aZqVtqloLq0YtETgEjrVyuZ+H//ACLEX++a6as2BFOSkEjL1CkivEdS1zWL26nW4v3KLIVCjjAr265/49pf90/yrwOb/j8uf+uzVpBXAjK7jl3dj7saMDBwB09KfikI4P0rQR6p8NX3eFo1/uua64VxHwvkzoskf91q7cVhLcYjD5W+hrw7XV2+Ib1fRq9yNeJeJ12eKb4e4qoAJoOs3GhagtzAS0LH99H2I9a9l06/t9Sso7q1cPHIMj2rwofSt7wl4gk0G+EcpLWMxww/umqnHqgPYaiufKMDicqIypDbumKpajrdhp2nfbridfJIyuDktXlviHxTf67I0YZoLPPEY/i96iMWwKOsQ2kGr3Edg++23fL7VTFIF2jA6UtdC0AWiiimAUUUUAFFFFMApKWigBKWiigAooooAnt/uVUm/wBc1W7f7lU5f9a1CENpaSimAUUUUDFoooxQAUUUUCCiiigBKKKKALmk/wDH+PpV3Wf9ctUtJ/4/x9Ku6x/rlqXuBn0lL3pKYwpaSigQUtJS0DCiiigApKDS0AFJRRQAtJRRQAtJRRQAUUUUwCiiigApaSlFABSGlo9ycCgBh3ZAQbnY4UeteteDNEGj6SnmL/pEw3Oe+D2rlvAXh43l0NVu0xBGf3QPc16ZiuepK+gAKWkqK4njtoHmmYLGgySe1ZAcz4/1n+z9J+yRNie5+UEHlRXlo4GO/f61f17VZNZ1ea7cnYDtRfQDvVCuiEbIDT0rX9Q0eJ4rJ8K3Jq9/wm2u/wDPYflXP0Yp8iYHQHxtr2DiYdPSu98G642taQHnI+0x8SivIjW/4H1b+y9dEUjbbe54b69qicFbQD16kNJnNHWsRnnfxF0Ro5F1e2T5ekwHc+tcQORkcg17rdW8d1bvBMoZHBBBrx3xDos2hak0DAm3c5iftit6cugjNpaQU6tgCiipbW2nvblLW1QvK5wAO3vSbsBa0TSpta1SO0iHyZzK3oteyWdrHZ2kdvCAEjUKKy/DGgxaHYCMYad+ZH9/StuuacuZgFFITUEl5bRPsluIkb0ZgDUAWKKq/wBoWf8Az9w/99ikOoWX/P3D/wB9inYC1RVX+0bEdbuH/vsVZVldQykFT0IpWAWiig4AyeAKADFGKr/b7P8A5+of++xR9vs/+fqH/vsUAWaKrfbrT/n6h/77FL9utP8An6h/77FFgLHauH+Jv/IPg+tdh9ttf+fmH/vsVxXxKmilsIBFKjnPRWzVwXvAcFRSUtdYBRSUUAFLSUUALRSUtABRRRTEFFFFABRRRQAUUUUAFFFFABRRRQMKKSlpgFWU/wBT+FVqsp/qvwpMCh3NFHc0UwCiiigAoopRQAUUUUCCiiigYUUUUAFFFFAC0UUUCCiiigCjRRSVzALRRRQAUUUUAFFFFABRRRQAUUUUAFFFJQAtFFFAElr/AMfcP+9W7rH/AB6pWFa/8fcP1rd1j/j0SpAx+1FHaimAUUUUAFFFFABRRRQAUUUUAJTJPuH6VJimOMoQO9AHtvhY7vDdgSB/qgOla9cRoPjHQ7LRbW1nuHWWJNrDb3rQPjvw73u3H/AayaA6jFA/Cs7R9Zs9YhaaxZ2jU43MuM1o5pDFooooAKKjnk8qF5NpbaM4Fcs/xC0SGVorgzxSKcFWTFAHTXswt7SWZgCEGcGltJUnt0ljAwwzwK5S58eeHrm0li8+TLKQBsq94G1AajoAkB6SMMe3aiwDfHumjUPDU+1cyw/Mn1qx4PGPDVoD1AwfrW3JGssbI4yrDBFQWFnHY2wgi+6CT+dAFisZ/sj+IgJkUzqP3RPatgiuI8RXf2Lx7pDE4R8hqYHbYwfxrgvEWmi28faZqEY4uMhz9K7339eao6jpyX0sDtwYjkGi4F8Hp9KSX/VP/umgcAfSkk/1b/7ppAeE6p/yGr3/AH6s+H9YbQtU+3LAJjsK7Scdaqaof+J1ej/bqtmtegjuj8T5wf8AkFL/AN9Uf8LRl/6BQ/76qb4d6Vp+o6VcPeWqSusmAW9K63/hGdE/6B0X5VDsM4v/AIWlL/0Ch/31S/8AC0pP+gUP++q7T/hGdE/6B0X5Uf8ACM6J/wBA6L8qV0BxR+KUn/QLH/fVJ/wtJz/zCx/31Xbf8Ixoh/5h0X5Vna94c0eLRLyWKwjWRIiVYdjRoB554n8Tf8JFJbs1qIPJbJ5zmuqtfiPpcFrDC0bZjQKeK85iGY1z1rr/AAV4Wsdf06W5u2IdX2gAdqtpWA3f+Fm6T/zzf8qP+Fm6V/zzb8qnHw40fuSfwpf+Fc6MPX8qnQBlr8R9LuLmOBY23SNtHFdojbkVh3Ga5O28A6TbXUVxHndEcjiusUbVAHQDFJgOrnPEHi+x0G7W3u1bcwyMCuirB17wrY67cpPdscoOABSAx/8AhZmj/wB1/wAqP+Fl6N6P+VSf8K50jszflR/wrnSf77flT0AZ/wALK0Y9n/KuG8Q6lFq2tTX0AIjkAABrsdT8BaZaadNOkjbo1yOK87TkH6kVcUhCkVv+AR/xV8H+4awDXQeAv+Rug/3DVS2A9goNLRWIzzP4kXWqQazarYPOsZjOfLUkZrkPt/iH/ntd/wDfBr3lkVuWRT9Rmm+VH/zzT/vkU0wPFtEv9e/tq0WSa6MbSAMChxiva+5poiQHIjQf8BFOobAUnivMfiicata/9c69NPSvMfin/wAhW1/65047gL8LTnVbof7FenA8V5d8K/8AkLXX/XOvUF6US3AXuK8X8QXuurr92kEt0Iw527UOMV7TTTGhOSin8KSYHhP2/wAQf89rv/vg1v8Agi71iTxLEl1JcGI9Q6kCvVvKj/55p/3zSiNAchFB9hTuA6iiipA8r+J3/Ibt/pXJd6634nf8hu2+lclW0dhBRRRVgFFFFABRRRQAUUUUAFFFFABVi3+5VerFv9ykBSl/1rfWm06X/WN9abTAKKKKYBRRRQAUUUUALRRRQAUUUUAHatzQ/wDj0esPtW3of/Ho9JgULr/j7k+tRVLdf8fcn1qKgAoopKYC0UlLSAKWkpaAEopaKACiiigAooooASkxS0tADcUmKfQRQAzFJ0p+M1oaDolxruoLbQAiFT+9k7AUm7AX/Bnh19b1ATzqRZQHJJ/jNevoqogVQAoGABVXTrCDTbKO0tUCRxjHHc1brFu4xaKKKkDzf4m6q0ksWkxN+7PzSfWuFArpfiDhfExJ7jp61Lofge91Swa7uJDbZGYV7t9a1i0kBy2OKMVc1PTL3SbgwahCY2HRh90/jVOrWoCYpQKUUYpgJijFLilxQBDN/qmr3Lw/zoNjkD/Ur2rxCRdyEDrXp2leN9DtdLtreaaRZIowrDb3rOaA7QUVy48feHu9zIPqlbGkavaaxbmeyLtEDjcy4zWVgNCiiigAooqC8uPsts8xUsEGSB1oAnqOZUeJlkA2nrmuaXx7oDMUknlikBwUdMGquueM9JfR5/sd0zTfwjFOwHWwwxQxCOFFRB0AFcj8SdJ+26It3GCZrVht+lb3hu+/tHQra6zkuOa0J4UuIWikUFWGCKNgM7wz/wAi9Z/7lauelQWVqlnapbx/cTpU/cUmBiW8kH/CSzRPGpnIyjEcgVtdDXF3919l+JVkrHCSxMD9a7XFMDhf7P8AsXxRimUYW6iZz9a7yqc+nxTahDet/rIlKirlDYFXU/8AkG3X/XJv5GvCP43/AN417vqf/INuv+uTfyNeEH77/wC8aumBNbxGe6ht1ODM4QH0rsP+FcXZ/wCX41yWmH/ib2P/AF2Fe7J90fSnOTQHm/8Awre6/wCf80h+G10QR9vPNel0VHOwMnw1pLaNpKWbyeYVYndWtRRUgMlTfEyf3gRXmkvw4vmuJXW84dy3516dSYpp2A8yHw3vMf8AH6aQ/De+xxeV6dikxT52BzXg3w9PoFvNFPL5m/pXTUmKWpeoCE1wWu+BLnU9auL6O72LL/D6V32KTFNOwHmg+G973vad/wAK4vCMG+r0oClp8zA83f4e6hLFHFNqbvHEMIpPSmf8K3ux0via9LpKOdgeL+INCm0C5ignl8zzV3A1lV2fxQ41Sx/65muLBreDugHUUlLVAFFFFMAooooEFFFFAwopKWgAopKWgCeD7lVJf9a1WoD8lVJf9a1CAbRRRTAKBRRQAtFJS0AFFFFAgpDS0UAJRRRQBc0r/j/H0q7rP+uWqWlf8f4+lXdZ/wBctS9wM/vSUtFMBKKKWgAooooGFFFJQAtJRRQAUUUUAFFFFABRRRTAKKKKACiiigApaSjNAC1r+GtCl16/C4K2kZzI/r7Uzw9oVzr12I4gUtlP7yT29q9c0zT7bTLNLW0QJGo7dz61jUnbRATWtvFa26QQqFjQYAFS0UVgAhrz34ja/gDR7V/mbmYj0rq/EutRaJpb3DEGVhiJfU141LLJcTvcTsWkkO4k/wAq0hG7uBGowAOwp4pOgrpPCvhQ69azXM8rQRdImA6mtm0gOdorotT8F6vYbnhUXEI/izz+Vc9IrxOVmieNh/fGKakmAhpjZ4ZThlO4fUU7IPQg0Gm9QPW/ButrrGjp5jf6TCNsoroRXifh/V5ND1aO6TJiY7ZV9fevZ7S4iu7ZLi3YPHIMqRXNJWYE2Kz9a0i21mwe1uVBz91u6mtGiouB4lrGk3Wi3htrtTt/5ZydmFUq9t1XS7TVbNre8jDqeh7qfWvLvEHhi90SQuA09pn5ZAMkfWuiFS+jAxK6jwVrel6RIyX0WyaQ8T4zx6Vy2cinw21xd5FtC0mOuBmrkk0B7lb3ENzEJIJFkUjIKnNS143pLeJNHmEljBLs/iiIJDV6XoesSajp7TXVs9vKn30x/KuZxsBPrmqw6Pp0l5MR8o+Rf7x9K8Tv7ufU72S8uZH3yHIAbG0eldL4sl1rXdQOLKRbSI4jTB5PrWH/AGPq3/Pi/wCVXBIDO2H+/J/30aXZ/tv/AN9GtA6Rqneyk/75pP7J1P8A58ZP++avQDOkT5PvP1H8Rr3XQuNFtOc/uxXjMmk6oUx9ik6j+GvadGRo9JtkcYYRjI9KidgLwqG9/wCPKf8A3D/KpxUN2pa0mVRklCB+VZAeCFcyS/M/+sb+I+tG3/af/vo1oHRtUWSXNk/MjHp70n9k6n/z4yf9810q1gKAX/bf/vo0oH+3J/30avf2Tqf/AD5Sf980f2Vqf/PlJ/3zTugKeD/z0f8A76NGD3Zj9Tmrv9lan/z5Sf8AfNJ/Zepf8+Un/fNCaAqg0tTS2N9Chea1dEHVitQZz0q07gLS0lFMBaKSigBaKKKYBRRRQIKKKKACiiigYUUUUAFFFFABSUtFABRSUooAKsJ/qvwqvVhP9V+FDAo9zRR3NFMAooooAKWkpaACiiigAooooAKKKKACiiigQUtJS0DCiiigCjRQaK5hBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBJa/wDH5D9a3dX/AOPVKwrX/j8h+tbur/8AHqlJgY/aiiigAooooAKKKKACiiigAoopaACkpadHFLPII7eJpJG6KozQBCwAGTiuk8KeDZ9blW5vFaGyU9xgv7fSt/wv4DwUvNbAJ6rAOn1NegxxrGipGoVVGAB2qHICG0tILK2S3tYljiQYCgVOKDxXPW3iOO+8VPpFph44Y90kg9fSoGdFSZpKzvEE81tol3cWzbZYk3KfegDSzXP+JfCtjr0JZlEV0B8syjn6VF4Q8UQeILFQxCXkYxIh7+4rpAaewHg2r6Re6LeG2v4tv92QD5WH1rtfhVdYa8sifuAMBXa61Y2F/p0kepIhhUElm/h968j0PVovD/iGW4t2aW1DFc9yKrdAe2UteWap8R7+aFksbVYQSMSZ5r0LQbmS70e3nmbdI65Y+tS1YDRry/4nuYdb06ZTgo4r1CvL/iqu7UbMUID0eymWeyhlU5DIP5VPXlWmeP59N0mGyWzWaWIY3E1Z0Hxjquq+KLWC4Iht3zmMdDT5QPTKa/8Aq3/3TT8UyT/Vv/umpA8G1b/kOXw/26q1a1j/AJDt9/v1UrVbCPTPhYwGkXmWH+tFd0HX+8v514Rpi64Y3OkPKsefm2DvV7Z4x/563P5VLQz2rcP7w/OjcPUV4sF8Y/8APW5/KnBfGX/Pa4/KlYD2fcPUVna/htCvgCP9S1eV7fGX/Pa4/KorhfFggc3E05hx84I4xRYDFhHyCvS/hYQNFnyQP3p4rgNN0u+1QsNPiEm3rmta28P+LLVSLQGFT1CnrVPYD2Hcv94fnS7l/vD868j/ALK8bD/ls/507+y/G/8Az3f86iwHrWQe4ory7TdO8ZJqlq1zM5hD/Pz2r1AZ2jPXFIBaTIHUgUVwXji38Rzaoh0d3WELztPegDvNw/vCl3D+8K8eFj44/wCe0v508WHjg/8ALeX86qwHpuvsDol1gj7hrw5D8p/3jXRSaX41ljMcksjI3VSetZt5oWqabame9txHEDyRVR0EUq6DwD/yN8H+4a52ui8A/wDI3wf7hqnsB7DRRRWIzO1DW9P02ZYry4EbsMgHvVT/AISzRP8An8WuG+KSBtetM/8APM96w9F8KX+uQPPYhPLjbadz45qraAeqf8Jbon/P4KP+Eu0T/n7Fee/8K41r0h/7+Uf8K41sdof++6LID0A+LtDx/wAfdcF8QtTs9Tv7eSyl8xVTBPpTP+Fc63/0y/77pR8OtcHTyf8AvumrIQ74eajaabqVxJey+WrJgGvQF8V6Lj/j8Feff8K61s9TD/33S/8ACuda/wCmP/fyh2Yz0H/hK9F/5/Fpf+Eq0b/n8WvPf+Fc6z/0x/7+Uf8ACudZ9If+/lKyA9B/4SrRf+fxansdf0y/uPItbkPJ6V5Hrnhi80GKOW/VNsrbV2tnmrvw9RV8VrgEcetHLpcD2OiiipA8r+J//Iat/pXI96674n/8hq2+lcj3raOwhaKSirAWiiigAooooAKKKKACiiigAqxB9yq9TwfcNICnJ/rG+tNp0n+sb60ymAtFFFMAooooAKWkpaACiiigAooooAT1rc0T/j0esM9DW5of/Ho9JgZ91/x9yfWo6kuv+PqT61FQAtFFJTAWiiikAClpKWgAooopgFJS0UAFFJRSGFLmm0tAhaKbWnoOiXmu3QitlKwg/vJT0A9qTdgGaPpN1rV6traKdufnk7KK9g0TR7XRrFba1UcD5m7saNF0e00ayW3tUA/vN3Y1o1jKVxhRRRUgFFFFAHOSeFoLvxG2q3x8wJ/qoz0+tdEAAMDge1MEqGQoHUuOq55p+aAKmpadaanatb3sKyIfXt715b4l8I3eis1xbbriyJ6gfMv4V67TXRXUqyhlYYIPenGVgPAQQRkf/qp1d/4p8Dhg97owCuOXg7H3FcCVZHMcilJFOGVhgitoyuISlooqgEpDgDJwKfHHJNII4UZ3bgBRmu98M+BACl5rQDN1WDsPrUykkBheFvCM+tSrcXaGKxU55GC/tXq9rbQ2lukFvGEjQYAAqSNFjRURQqqMADtQTWLdxi5oJrnrPxEt94ouNKtsNFAgZpB6+ldBSAKKyPFV1PZ+H7q5tW2zRDKmqvhLxNb+ILFeQl3GAJIz/MU7AHiHwjpuuIztGILo9JkHNeV61oV5odyYb2M7Twko5B+te6iqOtw2E2mTDVERrYLlt3amnYDlfhffebpc1iTzbnj8a7mvF/D2twaFr0tzFuNkxIx39q2NV+It7dYh02AQqzgebnnFDTA9RoqrprO+nwPI+9mQEn1q1UgeX+PZjbeMdOnU4KsM/SvTLeVZ4ElQ5V1BFeW/EwZ1+3x12HFWdM8ff2fosNp9m864iXbycA1fLdAem0V5/wCFvFeo6v4nWC6xHC8ZIjB4zXf1LVgKuqH/AIll1/1yb+RrwbPzv/vGveNU/wCQZd/9cm/lXgw++/8AvGrpgW9NP/E2sSSAPOHJ7V7mlxCEX9/F0/vivA8dwcHsaf591/z9SfnVSjcD3r7TB/z3i/77FH2qDvcRf99ivBfPuv8An6k/M0ySW48tibiT/vo1PswPoFXV13IwYeoOaWub8BM7eGIS7Fjnqa6Q1nYBM0wzRdPNj/76pLjP2eXHB2n+VeFzXd8bu4/02UYlYdapRuB7sJo/+esf/fVHmx/89U/76rwkXmoDpfS0v23UP+f6Wq9mB7sroxwrqT6A0/FeN+FL+9XxRYiW7d42J3KT1r2OoasAtNZ1X7zKv1NOrgvic1zFDZz207xbSQ23vQtQO582P/nqn/fVHnR/89U/76rwc3+ohf8Aj/l6Z61t6XoXiPVtMa+tL18fwKxxuFU4WA9c86L/AJ6x/wDfVBnix/ro/wDvqvDLqTWbKYxXk9xE47kHb+dRC8vWGftshHrmmoAdf8T2V9SsSjq37s/dOa40UO8spBmlaQjoT2oFaxVkA6lpKWqAKKKKACiiimIKKKSgBaSlpKBhS0lLQBND9yqkn+tarcP3aqS/61qEA00UUUxBS0lLQMKKSloEFFFFABRRRQMSiiigRb0v/j/H0q7rH+uWqWl/8f4+lXdY/wBctSwKBpKWkpgFL2oFFABRRRQMKKMUlABS0lFABS0lFAC0UlFABRRRTAKKKKACiiigBDWl4f0qLWNSW3nuRBGOSScbvas0005DBlYqynKsD0qXsI900+xt9OtEtrSMJGo7d/erdeeeEfGvMen6w/zHiOY9/Y16ArBgCCCD0I6GuZprcY+obu5htLZ7i4cJHGMsTT2cKpZmAA5JPQV5X428TnVrk2Fk5FnEcOw/jPpSSuBmeJNbk13U3nJIt0OIk7fWswUxRgcdBTwQASeldKVkBLZ2cuoX0NlbjLyt+Qr2zS7GLTtPhtIVAWNf171yHw70MxQtqtynzy8Rg/w+9d1isJyuwExVG/0mw1BcXlqkv4Vfqrf3kNhZy3U7BY4xnPvUq4zzPxlommaM6fZJiJZekPYVzNWNU1CXVtTmvZicucKvoKr10xvbUQhAIOehrrvAniP+zrgaZev/AKNIf3bH+A+lcnikYZHoRyD6GiUboD3wEEZByKWuD8D+KxKE0vUpMSqMRSMfvCu8rmasAVm67qNppmmS3F7taPGAh/iPpV25nitoHmmcJGgyxPpXjninXpNf1EspItIjiNf73vTjG7AzbicXN1JOkQiWRsiMdFrtvhaTi6HGN3pXBjrXd/C3/l6/3q2n8IHoQ6dB+VLj6flQOlL+Fc4CY9h+VGPYflS0tADcew/KjHsPyp1FADcew/KlApaKACiiigBMfT8qMfT8qKKADH0/KjH0/KjNGfagAx9PyoI9h+VGfakP0oAwvGX/ACLVz0/KvH4/9WteweM/+Raua8fi/wBWtdFECQUtIKK2AWiiigAooopgFFFFAgooooAKKKKBhRSUtABRRRQAUUUUAJS0UUAFWE/1X4VXqwn+r/ChgUT1NFB6mimAUUtFABRRRQIKKKKBhRRRQAUUUUAFFFFAhaKSloAKKKKYFE0UUVygFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAElr/wAfkP1rc1f/AI9UrDtf+PuH61uav/x6pSYGR2oo7UUAFFFFABRRRQAUUUUAFLSUtABVrS9SuNIv0vbU/Mn3hj7w9Kq0UAe5aJqsGsadHeQEfMPmX+6fStAnAryb4e6rNZ659gyTbzjcR6NXpuqWcl9ZPbxXUlsW4Lp1xWTVhnHeOfGC2yPpmlvuuGGJJB0UVh/C4Y8R3DM2SYslj3NbL/DC2Ylv7RmJY5JPUmnW/wAODayGS11eeFzxuUdqelhHd71x94fnWZ4kw3h2/AYf6k1z58Eah/0MV3TJPAt/JG0b+IbpkYYZT3pKwHmVhdXNhcx3dnIY5ozwR/F7V7B4W8VW2u2eHZYbuMfvEY4H1rnf+FXxjj+0ZKcvwz8tt0eqTIxGCRVOzAzvHnis6jM2l6c5FtGcSuP4z6Vxy4AAHQV0nibwa3h+yW7W5aYO2CDWHpen3Wr3qWlihZ2+83ZRTVkBJpmmXOs3qWVmmWYjc3ZRXuOl2YsNOhtQ27y1AJ96oeGvD9roFiIYQGmYfvJD1Y1tVDdxhXmPxSH/ABMrL8a9OrzL4qHbfWB5Oc4A6miO4HEEgD8eB616H4D8JtE8esagCsmP3MZ7CofBXg4ysmqatH8vWKE/zNejAADAqpSAWmyfcb/dNOpsn3G+hqAPBdY/5D1//v1UPer2sr/xPr//AH6pkda1Wwj0v4W/8gW6/wCuvpXcfl+VecfD3WdP03S7iK8mEbtJkD1FdZ/wlei/8/YqHuM3Pyo/KsT/AISvRf8An7Wl/wCEq0X/AJ/F/OpswNus3xDxoN90/wBS1Vv+Eq0X/n8WqWteJdIm0e7iS6DM8RCgdzTsBzvwnPNyO2P616UOleP+BtfsdAaY35cbxgbRmuvHxH0Pv53/AHzTaYHY0VhaT4pstXINnHMyE43leM1u1IBRWZrmtQaJai5uo5GiJwSi5x9awh8R9CIyDN/3zRYDrjRXIH4i6F/02/75rW0LxJZa80gsVkKx9WYYFOwGzS/lSVi6v4kttJuRBNbXMjHnMaZFIDaP4VyfxIP/ABTEn+8Kjb4i6IGIYTAjqCvSsPxd4u0zWtFazs/M81iD8wwBVJAcSOg+grofAB/4q+D/AHDXO9h9K6HwB/yOEH+4at7CPZKKQUtZDPLPif8A8h+z/wCuZrX+FZzpt6P+mtZHxQ/5D1n/ANczS+BPEWm6JZ3MWoSMjySbl2jPFX0A9SoxXO2vjTQ7u5jghuHMkhwoK966KoATbRilrE1PxVpGl3n2W7mZZcZwFzQBtYpcVjaV4n0rV7s21lKzyhdxBXHFbNABiimu4RGZuijJrm28daAsjxm4fcjbT8vegDF+K/FlYn/prXOfD8/8VYv0rQ8f69putWNqthKzyRyZYEY4rN+H5/4q1fpWi2A9lpKTuaWswPLPih/yGbb6VyHeuv8Aif8A8hq2+lcgOtbR2ELS0lLVgFFFFABRRRQAUUUUAFFFFIAqeD7lQVND9ygCpJ/rG+tMp8n+sb602mAUUUUAFFFFMBaKKSgBaKSigBaKKKAENbmif8er1hmtzRP+PVqTAoXX/H1J9aiqS6/4+pPrUVAC0UUUAFFFHSgApaSloAKKKKACiiimAUUUHA5JA9zSAQ00nHXqeg7mr2naZfarOIbGAsT/ABkfL+deieHvA1np+2fUMXNx1w3RD7VMpJAcn4b8G3msMtxehrezznphm/CvUtPsbbTrVbaziWONR0HerAXAAAAA6ClxWLlcYtFJVPVNTtdKs3uryQJGo6d2+lIB2p6hb6ZZSXV1IEjRSeT1+lVPDWrPrWlLfMmwOTtHtXk/iXxFc+ILlnkJS1TPlxA8fWvR/h5/yKdv+NNqyA6ekPWlpD1pAeW+NNSvNL8YxXFjKyMF5XPDD0rrvDHiu016IRsRDdr9+MnqfauI+Iv/ACMyf7prmIpJYJknt5DHKhyritOW6A+gQaWuM8H+Mo9TC2WpMIrxRhWJ4k/+vXZis2rALXLeKvCMGrxtc2oEV4oyCP4/Y11NFCdgPBbmCa0uXtrqMxzIcFT3+lR16/4o8M22u2xYAR3aD93IO/sa8kvrS5068e0vYzHKhx7N9K2jK4hbK+n029ivbY/PEckf3h6V7Roeqwaxpsd5AQdw+cejdxXhxauq+HGpy2mtPYgkwSrkL6H1pTVxnq5PpXDeOvFwsYm0zTnDXbjDuvSMf411+oW0t1ZyQQXDQOwwJF6iuHk+GEbuztqczuxyzHqTUKwGV8LQf7fvCzFmKZJPUmvVa4Sz+HclhKZbPWJ4ZCMFlHWrh8Jar/0Ml1+VDswNHxudvhS9zx8teO2N1c6fdRXllIY5kwcj+KvSrrwRf3cJhudfuZYz1Ujg1Q/4VguMf2jJVKwHUeFvEtvr2n+bkR3EY/exnt71w3jrxM2q3jadZuRZxH5yP+WhrWtvh1PaOz2mrTRM4wxXuKwPE3hA+HrRblblpg5+YHsaStcDnP6Vo6Fo11ruoJb2ykIjBpJey4pNC0W7169W3tVIj/5aS9lFeyaLo9ro1itraIAAPmbuxpykBctYBb20cKnIRQM1LRRWYHlXxK/5GG3/AN01ygySFVSzMcBQOSa6v4lAnxHaqqlmZSFUdSa3PBfhAWwTUtUQNOwzHGeiD/GtVKyAf4H8KPYFdUvsi5dcIn90V2ppaKzbuBU1P/kGXX/XJv5GvBgP3kn+8a951L/kG3X/AFyb+VeDE/vZf941dMB9JT4YZrhtlvC8reijNadv4Z1y6x5VpjP9/itboDJpsmfLbjtXVwfD/W5cecYox7NWlB8M9y4udQkXPULUuSA6HwF/yK8P1rpaoaNpkWkaclnC5dU/iPU1fFYvcBk4zBIB/dP8q8ai8Na1e3lwYLTAMzYL8V7TRQnYDy22+HmqyY+1TRx/7rZqTWfA6aTos999sklkiGdhHFenVk+J4vO8PXsfqlPnYHj+mTeTqlpP02sP1r3OI7o0PqoNeAAlVBHVXH6GvdtIuBdaXbzKc7kA/SqmBcrkviTb+Z4ZklHWJhXW1meIrBtS0W4tEALSDjPrULcDx7RdMm1nU4bOEHBwXbsB3r2yytIrKzitoFCxxjAArE8HeHF0HT/3oBu5eZG9K6OnKVwKd/p9pqEXlXtukyehFcD4j8CQ2kMl7p9zsjXkxucKPpXpJrzL4ia59ruRpFs58qM/viD1NON7gcYpyKcKQUo610IQtLSUtAwooooEFFFFMYlLRSUCFpKKKACloooGTQ/cqpL/AK1qtQ/dqrL/AK1qEIbRRS0AJS0lLTASilooAKKKKACiikoAWiikoAt6X/x/D6Vd1f8A1y1S0v8A4/h9Kuav/rlpPcZRooooEFFFFAwooooAKKKKAEpaKSgBaKKKAEopaSgAooopgFFLRQAlLSUtACGmmn0hoAiZQRg9K7Hwb4vkspE07U3LwNxHKeq+1ciRTdtRKNwO18beLftAbS9KkOwj97KO49K4dRgACnhQBgUYpKNgFFa3hzSH1rVo7cA+ShzM3oKykR5JFjiUtI5wqjvXsHhLQ00XSkRgDcSDdI3f6UpysgNqCJIIUijGEQAAVJRRXOAhrzH4g699suv7KtX/AHMR/fEHhq6vxn4gXRtNKRMDdzDbGvpXkg3MS7nLscsT61rTjfUAApwoOAMk4FW9P02+1OYRWNuzk/xEfL+dbXSArY49KSu8034eApv1O5YOf4E5ArmPEOjS6JqJt3y0LcxOe496Smm7AZBByCpKspyrDsa9A8H+M/MCadq77ZBwkx/i+tcERTWAIwaJQTA67x74jN/OdMspP9HQ/vWU/eNccF7AcU7GOlKKIxsA3FaWh6/eaA0htIlkEnUN2qgRTSKbjcDp/wDhYesDpaR/nR/wsXWf+fOL865YitDQNGbXNTNkshi2ru3CocUgNofETWf+fOP86d/wsTWP+fOP86uj4aP/AM/70v8AwrR/+f8Aeo90Cl/wsTWP+fOP86P+Fiax/wA+Uf51d/4Vo/8Az/vR/wAK1f8A5/3o90Cn/wALE1j/AJ8o/wA6P+Fiav8A8+Uf51c/4Vq//P8AtR/wrWT/AJ/3o90Cn/wsTV/+fKP86Q/EXV/+fKP86un4ayf8/wC1NPw1l/5/2o90CkfiLrH/AD5R/nSf8LE1n/nzj/Orv/CtJf8An/ej/hWkn/P+9HugUv8AhYes/wDPnH+dH/Cw9Z/59I/zq7/wrST/AJ/3o/4VpJ/z/vR7oFH/AIWHrH/PpH+dL/wsPWP+fWP86tT/AA5eGCSX7c52KWx64rilGc+xIqlGLA6LUfGmpanYvZz2yIj9WBrAUYUD0oUU/FaxikACiloqgCiiigQUUUUwCiiigAooooAKKKKBhRRRQIKKKKBhRRRQISlpKWgAqwn+q/Cq9Tp/qvwoYyn/ABGijuaSgApaKKYgooopDCiiimAUUUUAFFFFABRS0UAFFFFABRRRQBRooormEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAEtp/x+RfWtvVv+PVKxLX/j8i+tberf8eyUmBkUUUUAFFFFABRRRQAUUUUAFFFFABRRSGgDZ8HnHiq2+le19zXifhA/8VXbfSvbO5rOW40FFZ+vXEtpol5cQttkjjLKfevOvBfirWdT16K2urlnjIyQakD1Wkpa5fx7qt3pOipcWT7JDJgn2oA6bFAFcJ8Odf1HWZ7kX8xkVFyue1d5QBzPjnSbrWNNgs7QDe0nJPYVd8OaBaaBYLBbqDKR+8kI5Y1B4v1aTRrO2vF5RZMSD1FbNrMl1axXEZBWVAwx70wJqWsTxPr9v4f0xriUgzMMRR92NO8K3k9/oMN1dHMkhJPt7UgNqse/0Sz1LV4Lu7w7W3KRnpmtauT1zVW0vxjpqM+2C5yH/pQB1oGBxS00MPz5/CuJ1nxY0vii10bTZMqG/fyD+QoA7imv9xvoaVeg+lI/3G+hoA8K1n/kP3/+/VM1c1n/AJD9/wD79Uz0P0rZbCNzQ/CFzr9q1zDIqhW28nFaP/Csb0/8vC/99V0Xwx/5Ac//AF0rtKzbdxnlP/CsL3/n4X/vqj/hWF5/z8L/AN9V6tSYpXA8q/4Vjff891/76qO4+HN9bQSTmZCI1Lfer1jGap6qP+JVd/8AXI00wPB84LD+6cVPDY31xH5tvaySJ/eVciqzfel/3q9Z+GzE+Fxz0kNW3YRzfgS/1bSL8WNzZz/Y5zx8n3G9a9TpozS5rNu4yC+tUvbKa2kGVlUqa8Rn0nUILueEWM5CSEKQnGK92zSYoTsB4OdN1LHFhPnp9yvWvBmknSdAhhkXErfM3HPNb2KWhu4CYrlPHWrX2n2SW+mWzyz3GQXVM7BXWUEUgPAWsNRBaSazuMnlmKVD2zXumuHbo90ePuHtXhSj5W/3zWsXcQV0PgD/AJG+3/3DXPV0PgH/AJHC3/3DRLYD2IdaWk9aKyGeWfFE/wDE+tP+uZrkBHKRn7LMffbXW/E//kYrHPpXoekWls+lWrG3jJMY7Vd7IDxzRIpRrtkfss4HmryV4Fe79zUK2tuhBWBAR3AqapbuAV4/8QYpT4qLLbSuNnVVyK9gqN7eGRtzxIx9SKE7AeWfDVJR4lkL20sY8nqy4r1eo0hijOUjVT6gVJQ2BDcjdbyADqprwS7tbhdQus2NwczNj5D619AEVGYYieY0P4ChOwHgHkTrk/Ypx77K3/h9/wAjYv0r1e+ghFlOfKT7h7V5T4E48Zyf9dD/ADqr3QHsfc0Unc0tQB5Z8T/+Q1b/AErkB1rr/if/AMhq2+lcgOtbR2EOopKWqAKKKKYBRRRQAUUlFAC0UlLQAVND92oamh+4aAKkn+sb602lk/1jfWkpgFFFFABRRRQAGiiikAUUUtABRRRQAlbmif8AHq9YZrc0T/j1ehgZ91/x9SfWoqluf+PmT61HQAUUUUCCiiigApaQUtAwooooATvS0UUwCrekvYpqcR1RS1qT8wFU6OvXpUsD3XTYLOCzjGnxxpAwyuzoat15l4D8SmznXS76TMEh/cuT0PpXpgORkGsGrMYtFIWABJIAHUntXHeKPHFvp+6003E910JH3V/GklcDY8Q+IrLQrUyXDhpT9yMdSfevI9b1q91278+8chAf3cQPC1Wu7i4vrlrm8maWZv4mPSosVqo2AYfut9DXr3w7P/FKW/415Gw+Vvoa9c+Hf/IqW/40TA6ik7ilo71kB5J8RP8AkZk/3a5jFdT8RB/xUqf7tcwBxW8dhDQCGVlYqynKsOoNekeDPGQuQmnas4WcDEcp/jHv7151ikI6EEgg5BHY+tEo3A+gQc0V574L8ZFnTS9Wf5+kUp/i9jXoOeKxasMDXP8AizRbDVtNdrx1hkjGUmPGPatPU9TtdLtGubyUIijOM8n6V5J4m8T3fiGcqCYrJT8sYP3vc04pgYrp5crx7gwQ4Df3vet3wN/yNkX0FYKqAMAYFdB4HGPFcX0FaPYD2PuaWisnxRdz2Ph28ubZissaZUjtWIGrSGvMfBHifVtT16O2u7lnjKglTXp5p7AJij8K474i6zfaPY2slhKY2kchiKh+HGt3+sJeG/mMvl4257UAdvxXP+LdEl160htI32Luy7egrf7Vz3i3V5NFFndA/u9+1x65pIDU0fSrTR7JLWzjCqo+Y92NaFQxSrLCkiEFXUMCPesjxR4ig8P2HmuQ07nEcfc0wN2iqWj3EtzpcE8/+skXJq7SAzJtFs7nV49SnQSTQjCA9BWnXNz6w1n4zg02RsRXMZYE9jXRMyqpZiAFGST2FADqK4u08UvqvjdNPs3/ANDiQ7iP4zXaUAVNT/5Bt1/1yb+VeDN/rZP94171qX/INuv+uTfyrwVh+9lz/eNaUwOl+H139l8T7CcLLHt/GvXh6V4PpE5tdbsZs4HmgH6V7qjq6h1PDDIpTWoD8CjijPtTXkVBl2CD1Y4qAHUtRxSxzJvikV16ZU5FSUAFFI7bUZvQZryvUfHmsy3E0UMUcKxuUBU88U0rgeqFgoyxAHqay9X1DTk0+dJ7uIblIwGGa8ludd1m6z5+pS7T/DWe48xt0hLt6k1apgIQP3g7b2I+les/D+7Fz4YgUnLxkg15P7V3HwvvALi9smb0KiqmtAPR6KKKxAKKM1na5rFtounvd3TYAHyr3Y0AZ/jDxAmi6YwQg3Uo2xr/AFryElnZpJGLSOcsx71oXNxqHiTVnmILzEFkj/ur7VQA5KkEMvBU9RW8FYQYoApaK0AKKKKACkoopjFopKWgQUUUUDEpaSloEJS0UUDJYvu1Vl/1jVai+7VWT/WNQgG0UUUxC0lLRQAUUUUDCiiigQUUUUAJRS0UDLWl/wDH8PpVzV/9ctU9L/4/h9Kuat/rVqXuBRooopgFFFFAgooooGFFFFABRRRQAUUUUAJRRRQAUUUUwCiiigApaBS0AJRiiloAbikxTqKAGYpCQASTxTsVteFNAk13UQXUizhOXbs3tUydkBvfD7w8Wb+2L1MD/lipHT3r0QVHDEkMSRRqFRBhQPSn1zSdwHVU1PUINNsZbu5cKkYz9ass4VSzEAAZJPavJPG3iM6xqH2W2Y/Y4DjI/jPeiMbsDI1fU59Y1KS9uCfmOEXso7VWSohUimuhKwG/4Q02w1TVvJ1CQgjlI+z16xaWlvZwiK1iWKMfwqK8NgnltriO5gYrJGcgj0r2Xw9q8es6VFdxkBiMOv8AdNZVE7galY/iXRYta0x4GGJVGY37g+lbFBrJaAeDzwy208lvcLtljOGFRGvQPiFoJkjGrWifOnEqgfe968+BBGRyK6oSugFxS4pKdVgFIRS0tADCK6b4d/8AIzv/ANcq5siuk+Hn/I0P/wBcqiewHq9FFFcoBRUU1xBBjzpo489N7AZqP+0LL/n7g/77FAFmiq32+z/5+4P++xR9vs/+fuD/AL7H+NAFmk4qv9vs/wDn7g/77FJ9vsv+fuD/AL7FAFjFGKr/ANoWX/P5B/32KT+0LL/n8g/77FPUC1ik20KwZQykEHoRS0rgVtQH/Evuf+uTfyrwlRy3+8f517vf82Fx/wBc2/lXhKj5n/3j/OtqQDwKdikFLXQAUUUUAFFFFMQUUUUAFFFFAwooooEFFFFABRRRQAUUUUAFJS0UAFFFFMAqdP8AV/hUFTr/AKr8KQyn3NFHc0UwCiiikIKKKKYwooooAKKKKACiiigBaKKKBBRRRQAUUUUxlGiiiuUQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAS2v/H3F9a2tW/49krFtf+PuL61tat/x7JSAyaKKKACiiigAooooAKKKKACiiigApKWmmgDX8If8jXbfSvbe5rxLwj/yNdt9K9sz8xrOW40JNFHPE0UyB0YYZT0NUrXRdMs5hNa2MMUgGNyjmrN1dRWlu887bY0GSazLHxRpOoXS29rcb5G6CpA2qr3lla30Qiu4EmQHO1hxViqGratZ6RbrPeuURm2gj1oAfY6XYaeWNlaxwFuuwdauVl6Pr+n6yXFjIXKdcitSgDjficM+HAD/AH6xPBfjCHTtBlttSYlrZd0fq3tW58TP+ReT/fryoKCq8dBVpXQFnxBq11rl695dsQM/u4+yivXfBX/Ir2leLyj5DXtPgr/kV7WiSsBu15j8VSy6hp7IcMrZBr06vMvip/x/2H40kAaj4/Mnh+K2sQy30i7JWP8AAPauZ8KqT4rsixJYkkk9zWbWt4SGfFtiPrVWsI9wHQfSkf7jfQ0Ch/uN9DWYzwrWf+Q/f/79Uz0NXNZ/5D9//v1TPQ1sthHp3wx/5Ac//XSu1rivhh/yBLj/AK612p6Vk9xmN4g8SWHh9Ea+34fptGazLHx9o1/fR2kAm8yQ4GV4rG+K3+rta5LwqP8AiqLTj+IU0tAPcap6t/yCrv8A65GrnrVPVf8AkF3f/XI1IHgbsqyShmA+Y9a7rwd4w0fRtE+yXkjiXeT8oyMVU8AaNY6xPerfwiTaxxntXZjwNoOP+PVatsCr/wALH8Pf89Jf++aP+Fj+Hv8AnpL/AN81a/4QXQf+fUUf8INoP/PsPyqdAIbb4g6Dc3UVvE8peVtq/LxmusUhlBHQjNc5B4M0S3uI5o7cB4zlTiujUYAA6ChgLWNrXibTNElSO/lZGfpgZrZrK1bw/p2sSI99CHZOhpAZH/Cw/Dv/AD8Sf980f8LD8Pf895P++alPgXQj/wAu9H/CC6F/z709AM3VvHehXWmTwQzSGR1woIrzSMqVIDAkknivXD4E0M/8sawPGfhnTdI0NrmzTbKCAKqLQjgyK3/AI/4rC3/3DWCOg+ldD4BH/FXwf7hqnsB6/R3paMVkM8o+J/8AyMVj9K9J0UEaPaZB/wBWK82+J4/4qC0PomRXPrr2tKiompzKi8KAelXa4HvFFea/Di81XUdVlmu72Wa3jUqVbpmvSM8VLQDqKjZwilm6KMmvF9W8Saw+s3vkajKkSykIoPAFCVwPa80ZrzD4e6tqd74jkhvL2WaMRZ2seM16bQ1YB2aTn0pr/cbHXBrxO+17W11K7jXU5gqykAZ6ChK4Hs98CbKcY/gNeS+BB/xWcn/XU/zrNOu62ylW1Ocg9RnrWl8PwT4sVj1PJPqaq1kB7F60UtFQB5Z8UP8AkM230rjx1rsPih/yGbb6Vx/eto7CHUlLRVgFFFFABRRRQISilpKBhRRQaAFqaH7pqCpofuGgCo/32+tJTpP9Y31ptMAooooAKKKWgBKKKKAFooopAFFFFAhDW3on/Hq9Ypra0T/j1ehjKFz/AMfUn1qOpLn/AI+ZPrUdABRRSUALRRRQAUtJRQAtFFFMBKWikoAKKWkpANOT0OCOQR2r0Xw544totFI1Z2E8AxxyWHavOyKQgZzjkVLjcDpNf8aahq+YbUm1tT3U8sPeubAA6d+vvRSihRsAYoxS0VQEbj5G+hr1v4dj/ilLf8a8mb7jfQ1618PP+RUt/wAaznsB09FLRWQzyf4iD/ipY/8AdrmBXUfET/kZI/8AdrlxW8NhARSGnU2qAawzjBwQcgjsfWu88PeO0t9GeLVNzXMAxER/y0+tcIaQjPWk43AuazrF7rt2bi9chM5SIHhapYpcUuKEhiVv+CP+Rri+grANb3gg/wDFVxfQUpbCPZKjnhiuIWhnjEkbjDKehp+eTUF5dw2VpJc3DbYoxlj6VgMr2mi6ZZz+fa2MUUnTco5rQxWJpninStUuhb2kxaQ9sVt0AVb7TrPUEVL23SZVOQGHSm2Ol2OnBhZWyQB/vbB1pNT1Wz0uJJL2URq5wCabpmsWOqh/sUwk2fex2oAu4riPiiu7RIh713NcT8Tv+QNF/vU1uBleGvGkOneGTBe7nuLYYjH9/wBK4vV9Ru9Xv/tl65Ls42p2Qe1MKjg45Apkg+5/viteUD3TQh/xJbX/AHBWhVDQxjRrX/cFX6xe4HlnxKkkh8S2M0TFXQbgR7VW1jxxfappq2UK+QSMTSDqfap/id/yHrT/AHDXI1qkmhHRfD7C+LYQOnlmvYq8c8Af8jfF/wBczXsQqZ7jK+o/8g+5/wCuTfyrwhx++l/3jXvGof8AHhcf9c2/lXhMv+vm/wB806YEZBBVh1U5FdpH8RZreyhggs1d40Ckt61x2KCK0cbgdBeeO9duc7FSAHuhrFu9X1e7VvtGpTMD/Dniq+KRh8hpcqA9Z+HZJ8LxliSfMPJrqa5X4df8itH/ANdDXVVi9wGyfcb6GvB7vjULsf8ATZv517y/3T9K8Hvv+Qnd/wDXZv51dMCGikLAMF5LHoo6mul0DwZf6uyzXYa2te+eHNaNpCMGwsrrU7pbaxiMjscFgPlX616r4V8LQaDCZGPmXcg+dz2+lamk6NZaRbCGyhVOPmfHLfWr+KylK4wopaKgCve3UVnaSXMxOyMZOK8X8Ra9N4h1EzuSttGSIkr2m7t0ubaSCQBlkUjBrwrU7B9L1W4sZAQY24PqDVwsBvfD0FvFsZHaJq7TxH4Ms9V3T2uLe665XgOfeuV+GMBk1qefHEYK5r1MUSbTA8R1TR9Q0mYx3sB4/jQfL+dUa9t10QnRrozoroIyQD64rw+NtyZ9Sa1hK4D6KKK0AKKKSgQtJS0UDCkpaSgApaKKACiiigRLF92qsn+sNWYvu1Wk/wBYaEA2iiimAtFFFABRRRQAUUUUDCiiigQUUUUDLWl/8fw+lXNW/wBatU9L/wCP4fSrmrf65aT3Ao0UUUAFFFFABRRRQAUUUUAFFFFABRRSUAFLSUtACUUUUwCiiigBaKKKAClpKKAFopK19F8N6jrUoEcZhgz80jjHHtSbsBV0jSrjWr5bW2B25/eSdlFew6TplvpVhHaWygKo5PqfWo9F0a00azEFqgz/ABP3Y1o1zSlcBKSnVWvoppbSSO3fZIwwG9KgDiviD4m+zxHSbGT99IP3rD+AV5wuFGM//Xr0j/hW1rJK01zqM8kjnLMe9WI/h1pKfeuJm+taxkkB5iGX+8KeHT++v516tH4D0VOqs31q3H4O0NP+XNG+tP2iA8f8xOzA10PgrXG0rVxE4c2twcPjse1ekJ4a0VOmnQfiKnj0bTYiDHYwqR6LUyncC8DkZHfmloAAGBRWYEc0ayxNG4yrgqw9q8c8UaM2iau8QGLaY7oj6CvZ6w/FeiprOkSRY/fINyMOvHaqhKzA8eFOFJseN2jlXbIhwyntTq60AUUUUAIa6P4e/wDI0N/1yrnDXR/D3/kaW/651E9gPWKQ0tIa5QPOviqTtssEj5+xrgvmz/rH/M13vxVztsuCfn7VweRn7jflW8LWAAW/vt+ZpQW/vv8AmaNw/ut+VLkf3W/KrsgEy399/wAzRhv77fmadkf3W/KnDH91vyosgIyG/vv+ZprK3y/O33x3NTcf3W/Kmtj5flb747UmkB7lpA/4lVt/uCrtUtI/5BVt/uCrtczAgvR/oU//AFzb+VeED77/AO+f517vef8AHnP/ANc2/lXhOPnf/fP862pAKOtOpop1dABRRRQIKKKKYBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFTr/AKv8KgqZf9X+FAyp3NFHc0UxBRRRSAKKKKYwooooAKKKKACiiloEFFJS0AFFFFABmiiigZRooormEFFBooAKKKKACiiigAooooAKKKKACiiigAooooAltf8Aj7i+tbWrf8eyVi2v/H3F9a2tW/490pAZPalpKKACiiigAo70UUAFFFFABRRRQAU006mmgDX8I/8AI1W1e2H7xrxPwj/yNVt9K9s/iNZy3GjN8QWMupaLcWcGPMkQhd3SuH8KeBdW0fW4ry5MJjQc7Wya9KoAFSAveuc8b6Dc6/pcVtaMgdJNx3nAxXSUcUAcb4E8L33h6W4a8aMiQYGw5rsqKKAON+Jhx4fT/fry1FcoNsbsMdQK9S+Jv/IAj/36wvhte2kjPpV5bxuw+aN2HJPpVp2QjimguHXC2szEnoFr2rwlby23h21imQo4GSD1FaSWVqhyltGCPRasVLdxjT1rzL4qAnUNPAGSc4Ar0415Z8ULnGt2QjPzwkMfahAch9nu+hsrj/vmt/wRpt5L4ptpzayxxRA7mdcCvT9Glt9S0u3uxCh3oB09K0kRUGEUKPQCm5CAUOfkb/dNLUdywjt5XJwFQn9Km4zwvWGzr18f9uqueDTruX7RqFzOOjucfnUWeDWq2EeofC7/AJAlz/11rt64b4WH/iTXX/XWu5zWb3Gc/wCKPC8PiNYhNO8Xl9NvesnTPh7badqcV6l5KxjOdp7121FFwCqmq/8AILuv+uR/lVvNU9WP/Equ/wDrkf5UgOC+FikXV+ccZP8AOvSO3SvCtM1m/wBHlm+wMBvY5zWh/wAJr4gH/LUfnVOLEeyc+lHPpXjR8a+IP+ew/Omnxn4gP/LYfnRysZ7Nz6UoryPRfFuuXOuWVvPNmOSTDDPUV65SasAuaM0lcB4/8QanpOoQR2Em1W60kB6D+FH4V4t/wmviD/nsPzpR408Qf89h+dVysD2iuS+JA/4piQ44DCuHHjTX/wDnsPzqtqPiPVNUtTbXrgxHqM01F3Ayx0H0FdF4B/5G6H/cNc77V0PgH/kboP8AcNXLYR7DRRRWIzyn4oD/AIn1r/uGuOCu7LHGCZHO1QO5rs/iapk8Q2aLjcyYGfWtrwb4MSxKajqW2S4YZjQcqo9a0TsgNfwXon9i6IiOMTzfPL7GuhxxQOvNBPYVm2Bj+Kb9NN0C6mY4LIUX614ehZhvf7z8mu3+JetC8vE0q3fMcXzSY7N6VxeM1pFaCOr+Gf8AyNMv/XGvWa8n+Ggx4ol/6416zipluMRvun6GvG7Xw3f67fX8+nvFhbhgyuea9kf7rfQ15V4N1MWHja9gc4jupSgHvmhARj4f696wfnW/4P8ABt/pWrG+1CSPgfKqGu8HpS0nJgFFFBpAeV/E4g63bj0Arkcc10Hj27W78VSLG25IlAyPWsCto7CCiiirAKSiigBaKSigAooooAKKKO9AC1LD92oqlh+5QBVk/wBY31ptOf77fWm0AFFFFMAooooAKWiigAooopAFFFFAARW1ov8Ax6PWJW1ov/Hq9DAo3P8Ax8yfWo6kuP8Aj5k+tR0AFFFFABRRRQAUopKKYC0UUUAFFFFABRRRSATFGKWigBMUYpaKACikozQAjfcb6GvWvh5/yKlt+NeSN9xvoa9b+Hn/ACKlv+NZz2BHT0UUVkM8o+In/IyR/wC6a5bNdP8AEU/8VJH/ALprlxW8NhDqSiirAMUmKWigAooooAa1bvgf/ka4voKwjW74H/5GuP6VEtgPY+5rM8R2M2paDd2Vvt82VMLu6Vp9zRWAzzrwZ4K1TRdaW7vPJ8sDHytk16PSCloYHL+OfD91r9hBDZ+Xujck7zioPAnhm88Prc/bDHmXGNhzXX0UXAK4n4nHGjw/71dtXEfE/H9jwg92px3A84CSFQywyMpH3lHFM8m4kdFS1mY7xwFrtvhvdwSNNpl1Gj9DFuHX1r0KOytYm3R28an1Aq3LoBFpMbRaXbI4wwQZBq6aSis2B5d8TI5Jdfs0hRpHKH5VGTXKfY73/nwuf++K6zxlqhtfHFnPEebchH+hr0mAw3EKTIqlXUEHFac1kB5p8O9Jvf7f+3TW8kMMaFfnGCTXqYpqjAwBgU+obuBW1DjT7n/rm38q8JkP7+U/7Zr3HWZVg0i7kYgARN/KvDFbcN/945q6YD60dO0LUdVt2nsFR0U4IPWs8etdH4E1M6frwhdsQ3A2gf7VaSdkBTPhTXQf+PYflSHwnrzAhbZcn1Fez0VlzsDF8KaVLpGhxWs5BkzubHY1tUUhNQA2Q4jY+gNeJ2Wmz63r11bW0kaMZm5c4zz2r17W7tbLR7q4c4EaE14bHNKlwbmBykvmGRWHUVpAD1zQfBmm6UBJKv2mf+9IM7T7V0wArmPBviePXLMRTEJewjDr/eHqK6cGod+oC0UyWWOGMySuqIOrMcAVwHifx6BvstF+Z+jzHoPpQk2B3A1GzN79jWdDcdSgPIq1Xg1hqNzYaxFqnmNJOrfOzH7wPWvb9NvoNRsYru3YNHIMg+9OUbAWSK5fxd4Rh16MTQt5V5GPlbs31rqaCM1KdgOW8EeGpvD9pP8Aa2Vp5mydvQV0/SlxikNO9wOd8d3n2TwxOwOC7BfzryJPlUL6V3XxRvwxttORs7hvYDtXCA81tTWgElLSClrUAooooAKKSloASloooAKKKSgBaKKKAJIvu1Wk/wBYasxfdqtJ/rDQgG0UUCmIWiiigAooooAKSiloGFFFFABRRRQBa0z/AI/R9Kuat/rVqnpv/H6PpVzVf9ctJgUaKKKACiiigAooooAKKKKACiiigAopM0tABSUUUwClpKKAFpKKBQAtFFFABRRRQA1huUr616t4D1T+0dCRGwJIDsI9q8pPWum+H2ofYvEDW7n5LhQqj3rKoroD1eiiiucAoxRketFACYoxS/hR+FACUtRtLGv3nVfqaha+tE+9dQj6sKLAWqKz31nTI/v6hbj/AIGKqTeKtGh63kbf7pp2YG1RXKzePdDiJy8jf7oqz4f8Wadr1zLBah0kToH43CizA6KikFLSA8z+IOi/Y70apAv7qY/vcf3q5GvbdXsItS02a1mAKspwT2PrXis8DWtzLbOwcxNt3Doa6KUr6AMooFGa2Aaa6P4fH/iqW/651zjGuh+Hv/I0t/1zrOewHrRpKO9LXKBXurG1vMfardJdvTcOlV/7F0v/AJ8IP++a0KM0XAz/AOxdL/6B8H5Uf2Jpf/PhB+VaFHFFwKH9i6X/AM+EH/fNL/Yumf8APjB/3zV+ii7Aof2Npv8Az4w/980f2Npn/PjD/wB81foouAiIqIFQAKOgFLRRQBDef8ec/wD1zb+VeEH77/75/nXu95/x5z/9c2/lXhH8cn++f51tS3AUdaWkFOroAKKKKYgooooGFFFFABRRRQIKKKKBhRRRQIKSlooAKKKKACiiigAooooGFTL/AKv8KhqVf9X+FAFXuaKO5opgFFFFABRRRQAUUUUCCiiigBaKSloAKKKKBhRRRQIKKWimMoUUUVyiCiiigAooooAKKKKACiiigAooooAKKKKACiiigCW1/wCPuL61tar/AMeyVhwHFzGfetzU+bVTSYGTRRRQAUUUUAFFFFABRRRQAUUUUAFIRS0UAa3hAZ8VW2PSvbMcmvBLG8m0+9S7tseanTPSt8+Pdd67YqmSbA9O1rURpOmy3jRlxGMkCuKHxQgIBFm/PtXO3/jHWNQtHtZ/LEbjBxWAMgAA8Cko9wPQv+Fn2/8Az5yflSj4n2vezk/KvPNx9aTcfWnyoD2Lwt4tg8RyzRwwPGYhk7q6T8K+f7a9vLNy1ncyQM3BKHrVg65rP/QTuP8AvqpcRnonxPP/ABIIv+uleb6JenT9ZsrsHAjYbqZc6jqF5GI7u8lmQchXPAqsVyMVSWgj6FhcSwpIOjqG/On147B4512CCOGMRFYwFBPXFJL431+Vcb0T3U1PKxnquq6raaVatcXkqoFGQueW+leKazqMms6rcX8ox5nCr6DtUV5eXd/J5l9cyTsOgY8CoapRsI9R+Gd95+gG1Jy1u2D+NdnXhmj63f6I8rWBX9794N0rSbx1r7DrGPoaTi7jPXydoJPA9TXEeOfFdva2L6fYSiS6l4JU5CjvXD3niPXL1Sst9JGp6hD1rJ2/MWJJY9WPehRENQbVwPrTj0NKBQRVgemfCwH+xbo4/wCWtdxz6V4RYaxqWmxNFY3DRoxyQDVv/hKte/5/X/OocWM9F8XeLf8AhG5oIzbmTzVzmsAfE9e9k1cXqGp32pur30xkZBhcnpVXLZ601ER6CPidH3s2pl18R4bm0mg+yODIhUHHrXBAn1pwJ9afKgEOSzH1OaSn0lVYBuKMU6ilYC7oAx4j08/9NK90GMCvBbK5NlfwXaruaFtwB712P/Cyrn/nzSpkmM9JOK8u+KIzqltVj/hZd1/z5pXOeI9ek8QXEU0kQjMfYUlF3EY2OaWnY5oxWlgEpwopRQAV0PgH/kboP9w1z1XtG1SXRtSS+hRXdQRtbpSewHueR60ZHrXmP/CydR/584fzo/4WTqH/AD5Q/maz5WMr/E4k6/bYOCEPI7VF4X8a3OkulrqTNPaE4Dn7yf8A1qydf1mbXr5Lq4jWNkGAF6VmEdqtR0Ee+Wl3Be2yXFrIskTjgg1g+MfEsWh2DJEwa8lGEUdV96840DxHf6B5i2p8yKQYEbHhT6is26uZ726e5u5Gklc9SensKSjqBEzSSyvLMxaSQ7nY9zSgUAUtWB1Xw3KJ4llZ3CjyepNep/aYP+e0f514IkkkbbopGjb1XrTjd3v/AD+zfnUONwPeGuICD+/j6eteE3cxtdenuoz80NyWBH1phu73/n9m/OoHG8Hcck9T601GwHvem3K3enW84bO+ME/WrYIrx/TPHOqaZYR2cNvC6x9GbrVl/iNrbDC21uKnlYz1j8K5vxT4ptNGs3WKRZbtwQiKc4PvXnF54u128yGuTCD/AM8zWOzNJIZJXaSQ9WY8mmoADvJLK8spy8jFifrTqSlrQQUUUUwEooooAKKKKACiiigAooooAWpYvumoqli+7QBVf77fWm0rH52+tJQAUUUUwClpKWkAUUUUwCiiikAUUUUwCtnRv+PVqxa2dGP+juKTApXP/HzJ9ajqW6GLqT61FQAUUUUAFFFFABRRS0AFFFFABSUtJQAtFJS0CCiiimMKKKKQBSEUtBoAY/3G+hr1r4d/8ipb/jXkxGQR61s6X4q1XSbJLO0WMxJ03daiSuB7RmkryP8A4T3Xv7kNKPHuvZ+5DUcjGO+Iv/IyR/7tcwKuarqd1rF4Lq9CiQDA21UArWKshBRS0VQCUUtFACUUtJQA01u+Bv8Aka4/pWGRVrStRn0nUFvbZFeRezdKmSA91zyapaxqSaVpc966llhXJA715z/wsXWc/wDHtBVXU/G2qanp8tlPBCscowxXrWXIxm8vxOtWAP2STn2p3/CzLb/n0k/KvO1BVQAeAMU7J9avkQj0P/hZlr/z6SflW94Y8UQ+IvP8mF08nGdw6149k+tT21/e2Zb7HdSQbvvbD1pOAz3r8K4b4o5/sq34/iNcL/b2tf8AQTn/ADqC71HUL5FS9u5JlXoGPSkoMCbw/fHTdes7rPyg7T+Ne4qdyhh3Ga+fvQ91IIroF8a68kaorR4UYBzTlG4HsXPpVDV9WtNJtHnupVXaMqmeWNeUyeMNflGDOE91NZN1dXV9J5l7cyTt23npSUGAaneSanqF1fS8NMcgelet+Brz7Z4atgTl4htavIK1dJ8Q6lo0Lw2JUo5yd3aqlHQD2yms4RSzkKo6k9BXkTeN9fYY3Rj3BrOvNe1i9UrPfShD1VTwaj2bA6nx74mjuY/7J0+QMM5kkU8fSuGxjgdKQADp3606toxsgF6UqTNbzR3CfehYMKSmnoQabQHuekXi32l29yGB8xAT9au5HrXj+i+Mb/RbH7JDDHKgOQX6irp+I+rdrSCsHBgepZ9KjmljhjLzOsaDqzHivKZ/iBrkqkLFAme4rEvtX1TUci7vZCh6xg8U1BgdD468ULqjjTdPY/Z4z+8f+/7VyG3AAHSlVQBgDAp2K0UbCH2d3cafeR3lo5WaM5/3vrXpDfELTotKjnZHa6YYMYHQ15pikwAc459aHFMZq634l1PXXPnyGGA9IkPBHvWSqhRhRgU4CnCmlYBuK6Lwh4mfQbnyLjL2Mp5H/PM+1YBpCO1DVwPebW6gu4FmtpFkjboVOamzXhumaxqOktmxuGCdfKJ+Wulg+JF+oAubOIn1WsXBgem5HrVLVdRt9LsnurpwqqMgE/ePpXAz/Ei8ZcW1nHu/2q5bVNWv9Xl8y/mZh1EYPyimoMCLVdQl1bVJr6XI8xvkU/wiqw60oFKBW6VgFFLRiimAtJRS0AJS0lLQAUUUUCCiiigYUUUUgJIulVpP9Yasx/dqq/3zTQCUUUUwFpKWigBKWiigBKKKWgBM0tFFABRRRQBa03/j9H0q5qv+tWqWnHF6tXtVHzKaT3AoUUUUAFFFFABRRRQAUUUUAFFFJQAtJRRTAKKKKACiiigAoFFLQACiiigAooooAQ0+0na1v7a6TrC+6mUlJq4HqEnj/R4IkLmV3KgkKM81nT/Eyzzi3tJT/vCvPtoHQUvPrWXs0B2cvxJvD/qLKP8A4FVOX4g65JwkECfSuX5oxT5EBty+Mdfl/wCWwT/dNU5df1yX72pTD6GqOKMU+RASPf6lIfn1GdvxqBnnf79xI31NPxRinygQmJT97J+poEMf9wVNilxT5QIdijotW9KvpdL1SC9hJBRgrf7veocUFQQR6jFDiB7vZ3Md3aRXERyki5Wqmsa5YaPAZLyYAjpGD8x/CvNLDxhf6doy6dbIpKjCu38IrCuZ57uczXczzSH+Jj0rFU3cDode8aX+qlorQm2tj0K/eYe9c5k9zk9ye9J1pRW0YpALRRSVYCGui+H3/I0t/wBc6501f0LVn0TUjexxiQlduDUTV0B7bQTXnI+JE/8Az5rR/wALIn/581rn5JAdR4k8UWvh4RfaY3fzDgbKw/8AhZWm/wDPvP8AlXK+JvET+IRCHhEflHOR3rEwc1pGnpqB6L/wsrTv+fef8qUfErTf+fef8q855peafs0B6OPiVpn/AD7z/lTh8SdL/wCfef8AKvN+aOaPZID0j/hZGl/88J/yoPxK0oY/cT8kDpXmxJpDk49jmj2SA96tbhLm2jnThXGRmpdwryq28f6pa20cCW0JWMYBPepP+Fjat/z6wVl7Ngel3hH2Ofn/AJZt/KvCM/O/++f511cnxD1SSJ42tYcOpU/jXKD7xb1JNa04tbgOFLSClrYBaKKKYBRRRQAUUUUAFFFFAgooooAKKKKACkpaKACiiigAoopKAFooopjCpV/1f4VDUy/6v8KQFXuaKSlpgFFFFABRRRQAUUUUCCilooAKKKKACiiigAooooAKKKKYFGiiiuUAooooAKKKKACiiigAooooAKKKKACiiigAooooAAdrK3oa6G6Hm2GR/dFc6elb2myCex2nqODSYGWOlFLKhjlZT2NNoAWkxS0UwCkpaMUAJS0UYoAKKKWkAlFFFMANJilopANopaKAEopaKYDcUYp2KMUgG4pcUuKKYCUuKWigBMUYpaWgBuKKdSUANpcUtFACUYpcUUANxSYp9JigBuKXFLS0AJilFFFAC0d6SloAKKKKAEoyaKMUAGaKKKYBilooFABRRRQAUlLRQA2jmlxRigBKKdRQAmKMUtLQA3FGKdSUAJikxT8UUAMxRinUUAMxS4pcUUAJilFGKXpQAClpKKYBRRRQAUtJRQAUUUtACUUtJQAUUUtABUo+WMn2qIDJFOnbbHj1oArE8k+tFFFABRRS0wCiiigAooooAKKKKACiiigArS0V+ZE9azasafL5N2pPQ8GkBZ1Bdt0T/eqrWnqceUWQfw1m0IBKKKKACg0UUAFLSUtABRRRTASilzSUAFKKSigApaKKACiiigAooooAbSEU+kxQA3FGKdijFIBBSilxRTAKKKKACiiigBKWiigBMUmKdRigBmKWlxRSATFLS0UwExSYp1FADcUmKfRigBmKMU/FGKAGbaUCnYooGJikxT6SmIbilxS4paAG4paWigYUhFLRQIZikxT8UYpAMxS4p2KAKYxMUtFLQAmKMUtFAgo7UUUDCkNLRQA00c0uKMUAN57mlxS4oxQAYopaBTAKWikoAKKKKAFoo70UAFFFFABRRSUALRRQBk4pASLwlVCckmrMx2x49arU0AUCigUwCiiloAKKKKAEpaSloAKKKKACiiigCW2bZdRt71r6ku+AMOxzWHnHI6it5CLixGOflx+NSwMmilIIJU9RxSUwCiiigApKWigAooooASig0UwCiiigAooooAKKKKAClFJS0AFFFFABRRRQAUhFOpKAExSYp1FACYoxS0UAJijFLRQAUmKWigAooooAKTFLRQA3FGKdRQAgFFLRTAKKKKAExSYp1FIBvNHNLilosA3HrS4FLS0AJikp1FMBKSlooATFIRTqKAGYpMU/FGKLANApQKXFKBQAlLS0UwCiiigAooooEFFFFAwooooEFFFFABRRRQAUUlLQAUUUUAFJRS0DEpaSlpgFSOdsRpiDLUXDcBaQEFFLRQAUUUUwCiiigQUUUUAFLSUtAxKWkpaACiiigApaSimIWikooAo0UUVygFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABVzSp/JuNjHCvxVOjpyOooA2NShz++UfWs+tWxuFu7fa/3wMMKoXVubeTH8B6GkBDS0lLTAKKKKAClpKWkAlLRSUALSUUUAFFFFMApKWigBKKWigAooooAKKKKACiigUAFLRRQAlFFFABRRRQAUUtJQAUUUUAFFFFABRRRQAtFFFACUtFFABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAWkopaACkoopgFLSUUALRSUUgCiiigAooopgFFLRQAlFFFABRS0UAJS0UUAFFFFACUtJTkG4+1AD417moJ23SewqaZ9i4HU1V+tABRS0lMApaKKACiiigAooooAKKKKACiiigApOmCOo5paSgDetpFu7PnrjBrLkjMcjI3UUWFybefk/I3WtO9txMgkj+8B+dIDKoo9qKACiiigApaSloAKKKKAEpaKKACiiigAooopgFFFFABRRRQAUUUUAFFFFAwooooEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLSUtABRRRQAUYoooGFJS0gpgFLQaSgBaKSloEFFFFABRS0lABRRRQAUUUUAFFFFAwooooAKKKKACiiigAooooAKKKKACiiimAUUUUAL2oopKAFooooAKKKKQBRRSUALT4x3pqjcaWV9i4HU0wIZm3PjsKZRRTAMUUtFAgooooGJS0UUAFFFFABRRRQAUUUUALWjpE+C0DH3FZtOR2Rw69VOaANC/g8uTeB8rdfrVWtaN0vbX68H2NZksbQyFG/A1IDKKKKYBRRRQAUUUcUAJRS0lMAooooAKKKKACiiigBaKBRQAUUUUCCiiigYtJS0UAJRS0UAAooooASloooAKSlooASilooASilooASilopgFFFFABRRRQISilooATFLRRQAUUUUDCiiigBKKKKACiiigAopaKAExRS0UwEpaKKACiiigAooooAKKKKACiiigAooooAKKKKAEpaKKACkpaSmIKWkpaBiUtFORc8mkA5BtXJqu7bnJqWZ8fIPxqCgBaKSl7cUwHpDLIjPHGzKvUjtTK63RrXy9KKdPPHNcvdReRdSxY+42KlSuwIqKKKoQUUUtACUtJRQMWkpaSmAUtFFIQUUUUDCilopgUKKKK5RBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABSUtFABRRRQA+GV4JRIh5H61uRSxX0Hv3HcVgU+KV4XDxnB/nQBcuLV4Gz1Tsah+laVtfRXK7ZMK3cHvST2AJ3QnHsaQGdRT3ikjPzqR70zIoAKKKKACiiigAooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFLSUAFFFFABRRRQAUUUtACUUUUAFFFFABRRS0AFFFFABRRRTAKKKKACkpaKACiiigAooooAKKKKACiiigBaKKSgAopaKACkopaAEooopgLSUUUAFFFFABRRRQAUUUYoAKKKKACilooAKKKKACigAnoKesfc0ANVd30p7sIl9+wpHlVBheTVZmLHLUADMWbJopKWmAUUUUAFFFFABRRRQAUUUUAFFFFACUUtFABRRRQAlaWm3uzEEx4/hNZ1JQBtXlnvPmRYDenrWcQQcEYI7VPZaiYwI58lezelXpYYbpNwIPuKkDJoqeWzlj5A3D2qA5B+YYpgLRSUtABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooGFFFFAgoopKAFooooAKKKKACiiigAooooAKKWkoAKKO1FAwpaSloEFFFFABRRRTGJS0UUAFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigYUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKQC0UUUwCiiikAUUlKAT0oAKUKW+lOCdzSPKqcLyaAFZhGvvVZiWbJNDEsck0gqkAUUUtAhKKKWgApKKKBhRRRQIWiiigYUUUUAFFFFAhe1JRRQBNa3LWsu4cqeorYdIryAEH6H0rBqa2uZLZsryvdaTQyWaJ4W2uOOxqOtWKaG6jxwc/wnqKrzWBGTEePQ0gKdFKyOhwykU3imAtJS0UAJRS0UwEopaKAEopaSgAooooAXFFFFABRRRQAUtJRQAUUUUAFLRSUAFLRRQAlFLSUwClpKKACiiloASlopKBC0UUUDCikpaACikpaBBRRRQAUUUUAFFFJQMWiiigBKWiigAooopgFGKKKACiiigBaSiigAooooAKKKKACiiigAooooAKKKKYBRRRSEHekpaKBiUtFJQAUtKFJp4QDk0ANVM8miSQIMDrSSTADCVB15PWgA6nmiiigAq1p1s11epEvrk/SqwBJAUZJ6Cut0Sw+x2od1/eycnPUe1KTsgNFVCKFT7qjArnvEloVlW5RfkPDfWujqK5gW5t2hfow49qyTswOFpKnu7Z7S4aFwRjofUVBW4gooooAKKKKACiiimAtFFFIAooooAWikzRTGUaKKK5RC0lFFABRRRQAd6KKKACiiigAooooAKKKKACiiigAooooAKKKKADvnofWrlvqMsI2t86/rVOloA247+3lHzYB/2qeY7WXoV/CsHigFh91iPpSsBuGwgPTNJ/Z8Pv8AnWOJpR0lb86Xz5v+erfnRYDX/s+H3/Ol/s+H3/Osbz5v+erfnR583/PVvzoA2f7Ph96P7Ph9/wA6xvPm/wCerfnR583/AD1b86LAbP8AZ8PvR/Z8PqfzrG8+b/nq350efN/z1b86LAbP9nw+9H9nw+prH8+b/nq350efN/z0b86dgNj+z4fej+z4fU/nWP58v/PRvzo86X/no350AbH9nw+9J/Z8PqayPOl/56N+dHnS/wDPRvzosBr/ANnw+9H9nxep/Osjzpf+ejfnR50v/PRvzosBr/2fD6mk/s+H1NZPnS/89G/Ojzpf+ejfnQBrfYIfU0fYIfU1k+dL/wA9G/Ojzpf+ejfnRYDW+wxeppPsMXqayvOl/wCejfnR50v/AD0b86LAav2GL1P50fYYvU1lebL/AM9G/OjzZf8Ano350AapsYvU0fYYvU1lebL/AM9G/OjzZf8Ano350WA1TZReppPsUXqayvNk/wCejfnS+bJ/z0b86LAan2KL1NH2OP3rL82X/no350edL/z0b86dgNT7HF70n2OP3rM82X++350ebJ/z0b86LAaf2SP3o+yR+9Znmyf32/OjzZP77fnRYDS+yx+po+yx+prN82T++aPNk/vn86LAaX2WP3pPssfvWd5kn99vzo8yT++aLAaP2WP3o+zR+9Z3mSf3zR5kn980WA0fs0fvSfZkrP8AMk/vmjzH/vmiwGh9mj9TR9mj96z/ADH/AL5o8x/75osBofZ4/ej7Onqaz/Mf++aN7/3zRYC/9nT3o+zp71Q8x/75o3yf3zQBf+zp70eQlUN7/wB80b3/AL5oAv8AkJ60nkJ71R3v/eNG9/7xoAveQvvR5K+9Ud7/AN80b3/vGiwF3yVo8lapb3/vGje/940wLnkr70eSvvVPe/8AeNG9/wC8aALnkrR5K1S3v/eNLvf+8aALnlL70nlLVTe/940b2/vGgC35S0eWue9VN7f3jRvb+8aALflL70eUtVN7/wB40b2/vGgC15Yo8sVV3t/eNG9v7xoAtbFpcIvXFVNzeppDk9zQBaMyL05qF5mbgcCo8UUAFFLRQAlLRRTAKKKKACiiigApKWigAooooAKKKKACiiigAooooAKKSigA+tSQzSQnMbH6dqjpaANSHVFPEyke4qyJLWYdUrCoHHTilYDdNnbt0/Sk/s+H3rEEkg6SN+dO8+b/AJ6t+dFgNn+z4fej+z4fU1j+fN/z1b86PPm/56N+dFgNf+z4fU/nR/Z8PqayPPl/56N+dJ503/PRvzosBsf2fD6n86T+z4fU1kefN/z0b86Xzpf+ejfnRYDW+wRepo+wRep/Osnzpf8Ano350edL/wA9G/OiwGt9gh9T+dH2CL1NZPnS/wDPRvzo86X/AJ6t+dFgNb7BD6n86PsEPqfzrI86X/no350edL/z0b86LAa/2CH1P50fYIvU/nWR50v/AD0b86POl/56N+dFgNf7DF6mk+wxe9ZPnS/89G/Ojzpf+ejfnTsBrfYYvU0fYYvesnzZf+ejfnSiaX/no350WA1fsMXvR9ii96yvOl/56N+dJ5sv/PRvzosBrfYovej7FF71k+bL/wA9G/OjzZf+ejfnRYDV+xRe9H2OL3rK82X/AJ6N+dHmy/8APRvzoA1fsUXqaPsUXvWV5sv/AD0b86PNk/56N+dAGp9ji96PscXqazPNl/56N+dJ5sn99vzpgan2OP3o+xx+9Znmyf32/OjzZP8Ano350gNP7JH70n2SP3rN82T++aPNk/vt+dFgNL7JH70fZI/es3zZP77fnR5sn99vzpgaX2WP3pPssfvWd5kn98/nR5kn99vzoA0fsiepo+yx+9Z3mSf3z+dHmSf3z+dAGj9lj96PssfqazvMk/vn86TzJP75oA0fsye9H2ZPes/zJP75/OjzJP75osBofZk96Psye9Z3mSf3zR5kn980WA0Ps6e9H2dPes/zJP75pfMf++aLAXvs6epo8hPeqPmP/eNJ5j/3zRYC/wCQnvR5C+9UPMf++aPMf++aLAX/ACE96TyF96peY/8AfNJ5j/3jTsBe8lfejyF96peY/wDeNHmP/eNKwF3yVpPJWqe9/wC8aN7/AN40AXfJWk8lap73/vGje/8AeNOwFzyVo8lap73/ALxo3v8A3jSAueSvvSeUtVN7/wB40bn/ALxoAt+UtHlLVTe/djRuf+8aALflCjyxVTc3940bm/vGnYC35YpPLFVdzf3jRub1NFgLXlijYKq7m/vGjc3qaLAWvkHUimmZF6c/Sqx+tHSiwx7ys3sKZRRTEFFFBoGFFFFAgpaKKACiiigYUUlLQIKKKKBh3ooooAKKKKBBRRRQAUUUUDFBKnKkg+1XINSkTCyjcPUdapUUAbaXltKPmwP96l8i1k5BH4Vh4FKGYdGYUrAbX2GA9CaP7Ph9TWP5kn/PRvzpfOl/56t+dFgNf+z4fU0n9nxeprJ86b/no350edN/z0b86LAa/wDZ8Pqfzo/s+H1P51kedN/z0b86POl/56N+dFgNf+z4fU0n2CH1P51k+dL/AM9G/Ojzpf8Ano350WEa/wBgh9TSfYIfU/nWT50v/PRvzo86X/no350WA1vsEXqaPsMXqfzrJ86X/no350ebL/z0b86dhmr9hi9TR9hi9TWV5sv/AD0b86PNl/56N+dFgNX7DF6mj7DF6msrzZf+ejfnSebL/wA9G/OiwGt9ii9TSfYovU1l+bJ/z0b86PNk/wCejfnRYDU+xReppPsUXvWZ5sv/AD0b86TzZP8Ano350WA1fsUXqaPsUXqay/NkP/LRvzo82T/no350WA0/sUXqaPscXqazPNk/56N+dHmyf89G/OiwGn9ji9TR9jj9TWZ5sn99vzo8yT++350WEaf2OPvmj7HH71meZL/z0b86PMk/56N+dFgNL7JH70v2OP3rM8yT/no350ebJ/fb86LAaX2SP3o+yJ71m+ZJ/fb86PMk/wCejfnRYDS+yR+9H2SP1NZvmSf32/OjzJP75/OnYDR+yx+9H2WP3rO8yT++350eZJ/fNIDR+yx+9J9mj96z/Mk/vmjzJP75p2A0Psyepo+zJ71n+ZJ/fNHmSf3zQMv/AGaP3o+zR+9UPMf++aN7/wB80AX/ALOnvR9nT3qh5j/3zS+Y/wDeNAF37OnqaPs6e9Ud7/3zS+Y/940WAu/Z09TR5CetUt7/AN40b2/vGiwFzyE96PIWqW9/7xpd7/3jTAueQvvSeQtVN7/3jSb3/vGgRc8laPJWqe9/7xo3v/eNFhlzyVo8lap73/vGje/940AW/KWjylqpvb+8aNzf3jQBb8paPKWqm5v7xo3N/eNFgLXlrR5a1V3N/eNG5v7xoAteWtHlrVXc3940bm/vGmBa8taPLWqu5v7xo3N/eNICz5Yo8sVW3N/eNG5v7xoAs7Foyg7iq2Se5opgTtMAPlFQs7N1NJSUAFFFLQAlLznGMk9BUttazXT7YELe/pXSaZosdriWfDy/oKltICvoekmNhdXK8/wqe1b9FFZN3AKKKKQFHVNOS/gx0kX7rVyM8EtvKY5kKsP1rvKqX1hBfR7ZR8w6MOtVGVgOKoq9f6Xc2bnK74+u4dBVHrWqdxBRRS0wEpaSimAtFJS0gCiiigAzRRRTGUqSiiuUQUUUUAFFFFAB3ooooAKKKKACiiigAooooAKKKKACiiigAooooAKBRQKAFoopKBC0lFFAwopaKACkopaAEpaSloAKKKKYC0UUUCCiiigAooooASloooAKKKKBhRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAOtLSCloASilpKACloooASiiigApaKKACijFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUtJRTAWkpaKACkoooAWiiikAUUUUCCiiimMWiiikAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigBKWiigAooooAKKKKACiiigAooFFABRRRQIKKKKBhRRRQAUUUUwCiiigAooooAKKWk70ALRSUtACUtJS0AJRRRQAtJRRQAUUUtABRRRQAUUUd6AFooooAKKKKACiiigAooooEFJRRTGFFFLQAlFFFAC0UlFABRRS0AJS0UUwCiiigAooooAKKKKQBRRRQAUUUUALSUUtABRSUtABRRRTAKKKKACiiigAooooGFFFFAgoNFFAwpaSloASloooEFFFFABRRRQAUUUUDCiiigAooooEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAC0UUUwCiiigAooooAKKKKBhRRRQIKKSloAKKKKACiiimAUtFFABRRRQAUUUUAFFFFABRRRSAKKKKACiiigAooooGFFFFABS0UUwEopaKBBSUUtACUtFFABSUtFABRRSUALRRSUwFopKKQC0UlLQAUUUUwCiiigAooooGFFFFAgooooAKWkpaBhRRSUAT2Vq95dLbxsFZucmuhtvD1vGQ1wxdh6HiuYVmRtyMVI7itG11q8t8DcGXvnrUyT6AdZFFHEAIkVPoKfWZZa1bXOFkPlP0+bvWlWTTW4C0UUUgCiiigAooooAQgMMMAR6HpWddaLZ3BLbSjf7PSrF5f29mpMrgt/dHWsO68QzyEi2UIh/vdaqKfQCtqmlPpyrI0isjnAA6is+nzTyztmWRm9ieKjrZX6iClpKKYC0UUUgCiiimAtFJmigCjRRRXKAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUCiigAoooFABRRRQAUCigUALRRRQIKKKKBhSUtFACUUUUAFFFLQAUtIKKYhaKKKACiiigAooooAKKKKBhRRRQAUUUUCClpKKBhRRRQAUUUUwCiiigAooooAKKKKADtS0UUAFFFFABRRRQAUUUUAFFFFABRRRQIKKO1FAwooooAKKKKACiiigAooooAKKKKACiiigAFLSUtMAooxRSAMUUUUxhRRRQIKKMUtACUUtFABRRRQAUUUUAFFFFABRRRQAlLRRQAUUUUAFFFFIAooopgFFFFABRRRQAUUUUAFFFFABRRRQACiiigAooooAKKKKACiiigAooooAKKKKYBRS0UAJRS0UAFFFJQIWkpaSgYUUUUAFFFFABSikpaACiiigAoFFLQAUUUUAFJS0UAFBoooAKKKKACiikpgFFLRQAlLRRQAlFFLQAlLSUtABSUUtABRRRQAUd6BRQAUUUUAFFLRQISloooGFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigBaSiigYtFFFABRRRQIKKKKACiiigAopaKBiUUd6WgBKKKKBBRRRQAUUUUAFFFFABRRRQAUUUUAFFFLQAUUlLTAKKKKACikpaACiikoAWiikoGLSUUtABRRRQIKKKKACloooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKBhRRRQAUtJS0wCiiigAooooASloooEFFFJQAtJRRQAtFJRQAtFJS0AFFFFMBKKWigAooooAKKKKACiiigAooooAKKKKBhS0UUAFFFFABRRRQAfzrV0zWZbZhHOS8PcnqKyqKTVwO9hljnjEkTBlPcU+uP0nUXsZwrEmFjgj0rrkdZEDoQVYZBrKUbAOoooqQDp1rD1XWxGTBaEF/4n7Cm67qm0G1tm+b+Nh29q52rjHqwHO7SOWkYsT60lJRWogpaSimAUUtJQAtFFFIYUUUUCCiiimMo0UUVyiCiiigAooooAKKKKACkpaKACiiigAooooAKKKKACiiigAooooAKUUlLQAZoopKACloooAKSlooAKKKKACiiigBaKKKYgopaSgAooopAFFFFABRRRQMKKKKYBRRRQAUUUUAFFFFABRRRTAKKKKACiiigApaSigBaKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAUUUUgCiiigApaSloAKKKKYBRRRQMKKKWgTCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKQBRRRTASloooAKKKKACiiigAooooAKKKBQAUUUUAFFFFABRRRQAUUUUAFFFLTASlpKKAFopKWgAooxRQAUlFFABRRRQAUUUtACUvaiigAFFFFABRRRQAUtJ3paACiiigAooooAKKKKYBSUtJQAUUUUALRRSUAFLSUUALRRRQAUUUGgAooopgFFFFIAopaSgApaKKACiiigAooooAKKKKACiiigAooopgFFFFABRRRQAUUUUAFFFLQMSl70lLQIKKKKBhSUtJQIWikpaACiiigYUUUUAFLSUtACUUUtACUUUtACUUUUAFFFFABRS0UAJRRRQAUUtFACUtJS0xBSUtJQAtFJS0AFFFFAwooooABRRRQAUUUUCClpKWgAooooAKKKKACiiigAooooAKKKKACiiigAooozQAUUUUDCiiloAKKKKYBRRRSAKKKKYBRRRQIKKKKAEpaSloASlopKAFopKWgYUUUUAFFFFMAoopaAEopaSgQUUUUDCiiigQUUUtABRRRQAUUUUAFFFFABRRRQAV0Hh2+OTaSt7qa5+nwyGGZJQcbTk0pK6Gd7VLVb1bKzZ+rtwo9PerNvKLiCOYcBxmuV126NxflQflj+XFZRV2BnMzMxZzlmOSaSiitgCilpKAClpKKBBRRS0AFFFFAwooooEFFFFMCjRRRXKAUUUUAFFFFABRRRQAUUd6KACiiigAooooAKKWigBKKKKACiiigAoopaACiiigAooooAKKKKACiiigAoFFFAC0UUUCFpKKKACiiigAooooAKKKKYwooooAKKO9FABRRRQAUUUUwCiiigAooooAKKKKAClpKWgAooooAKKKKACiiigAooooAKKKKACiiigAooopgFFFFABRRRQAUUUUgCiiigAooooAKKWkoAKWkpaYBRRRQAUUUUDClpKWgTCiiigAooooAKKKKACkpaKACiiigAooooAKKKKACiiikAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiloASiiigAoopaAEooooAKWkopgFFLSUAFLRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUtACUtFFABRRRQAUUUUAFFJRTGFFFFAhaSlooAKSiigAoopaACiiimAUUUUDCgUUUgCloooEFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACikpaACiiigAooooAKKKKQBS0UUwEpaSloGFFFFAhKKKKACloooGFFFFABRS0UAJS0lFAC0lLRQAUlLRQAUUUUAFFFFABSUtFABRRRQAlLRRQAlFLRTEFFFFABRRRQAUUUUDCiiigAooooEFFFFACiiiigAooooAKKKKACiiigAooooGFFFFAgooooAKKSigYtFFFABS0UUAFFFFMAooooAKKKKBBRRRQMSilpKBBS0lFAxaKSimAtFFFIAooooAKKKKYC0UlLQAUUUUhBSUtFABSUtJTGLRRRQAUUUUCCiiigAooooGFFFFAgo60UUAb+lanDDpksc0m2RRhB61hO5kdnPVjk03juKKSVgCiiimAUUUUxhRS0UAJS0lLQIKKSlpDCiiigQUUUUxlGiiiuUQUUUUAFFFFABRRRQAUUUUAFFFFABRRS0AFJRRQAtJRRQAUUUUAFLSUtABRRRQAUUUUAFFFFABRRRQAUtAooEFFFLQAlLSUUAFFFFABRRRQAUUUUxhRRRQAUUUUAFFFFAC0UUUwEopaKAEopaKAEooooAKWkpaACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKWkooAKKKKAFpKKKAClpKKAFpKKWgAooooAKKKKYBRRRQMKXpSCg0CFopKWkAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAAUUUUAFFFFABRRRQAUUUUAFFFFMAooooAWkoooAWiiigAooooAKKKKACiiigAooooAKKKKACiiigBaSlpO9AC0UCigAooooAKKKKACiiigYlFLSUxC0UlFABS0UlAC0UUUAFFJS0wCiiigAooooAWiiikAUUUUAFFFFABRRRQAUUUUwCiiigAooooAKKKKACiiigAoopaAEopaSgAooooAWiiigYUUUUAFFJS0CCgUUUDCilpKAFooooAKKSloASloooAKKKKACiiigAooooAKKKKACiikoAWiiimAUYoooEFFFFABRRRQMKKKKBBRRRQMKKKWgQlLRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAwoFLRQAUUUUAFFFFMAooooAKKKKACiiigApKWkpgLSUUUgCilopgFFFFIQUUUUDFopKWmAUUUUgCiiigQUUUUAFJS0lAC0UUUxhRRRQIKKKBQMKKKKACiiigQUUUUgCiiigAooopgFFLRQAlFFLTASilopDCiiigQUUUUAFFFFMCjRRRXKAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRS0AFFFFABRRRQAUUUUAFFFFABRRRQAtFFFAgooooAKKWkoAKKKKYBRRRQAd6KKKBhRRRQAUUUUAFLSUtABRRRTAKKKKACiiigAooooAO1FFAoAKKKKACiiigAooooAKKKKACiiigApaSigAooooAKKWkoAKWiigApKWigBKWiigAooooASlopKAFooopgFFFFAwooooAWiiigQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRS0UAJRS0lAC0lLRigBKKWimAlFLRQAlFLRQAlLRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUALSUtFABRRRQAUUUUAFFFFABRRRQAUlLSUwCloooASloooAKKKKACiiigAooopgFKKSlpAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUwCiiigAopaSgYUUUUCCloopDEopaKYhKWiigYUUUUwCg0UnegApaSigQtFFLQMKKKKQBRRRQAUUlLQAUUUUAFFFFABRRRQAUUlLQAUlLRQIKKKKACiiimAUUUUAFFFFABRRRQMKKWjFAhKKWigAooooGFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigAopaKAEooooGLRRRQAUUUUCCiiimAUUUUAFFFFAwooooAKKKKACjFFFABRRRQIKKKKACiiloGFFFFAgooooAKKKKBhRRRQAUUUUxBRRRQMKKKKBBRSUtAwooooEFFFFABRRRQAUUUUDCloooASilpKBBS0lLQAUUUUAFFFFAwpaKKBCUUtFMChRRRXKAUUUUAFLSUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUtFACUtAooEFFFFAwooooAKKKKACiiigAoopaACiiigQUUUUAFFFFABRRRQAtJRS0AJRS0UxhSUUUAFFFFAC0UUUAFFFFMAooooAKKKKACiiigAooooAKKKWgBKKWigBKKKWgBKKWkoAKKKWgBKKWigBKWiigApKWigAooooAKKKKAEpaKKACiiigAoopKAFooopgFFFFAwpRSUtABRRRQIKKKKACiiigAoopaAEooo70DCiiigQUUUUAFFFFABRRRQAUUUUAFFFFAB2oxRRQAUUUUAFFLSUAFAopaACiiigAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUtJS0AFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiimAUUUUAFFFFIAooopgFFFFABS0CikAUUUUAFFFFABRRRQMKKKKBBRRS0DEoopaYBRRRQISilooASloooASilooASloooAKKKKBhRRRTAKKDRQIKKKKAFooopDCiiigAooooAKKKKACiiigAooooEFFFFABRRRQMKKKSgBaKKKYBRRRQIKKWkoAKKWigAooooAKKKKACiiigAooooAKKKKBgKKKKACiiigQUUUUALSUtJQAtFJS0AFFFFABSUtFACUUtFAwooooAKKKKACiiimAUUUUCCilooAQ0UUUDCiiigQUUUUDFooooEJS0UUAFFFFMYUUUUAFFFFAgooooAKKKKBhRRRQIKKKKBhRRRQAUUUtACUUtFAhKKWigYlLRRQAUUUUCCijFFAwooooEFFFFABRRRQAUUUUwFooopAAopaKYFAUYoorlAMUYoooAMUYoooAMUYoooAMUYoooAMUYoooAMUYoooAMUYoooAMcUYoooAMUuKKKBBijFFFAwxRiiigAC+9G2iimAYoxRRSAMUAUUUALijFFFAgxRiiigYYoxRRQIMUYoooGBFGKKKADFAGaKKYBijFFFABijFFFAC4oxRRQAmKMUUUwDGaMUUUALijFFFACYpcUUUAGKMUUUAGKMUUUAGKMUUUAGKMUUUAGKMUUUAGKMUUUAGKXFFFAxMUYoooELijFFFAARRiiigBMUuKKKYxMUYoopCDHFGKKKAF20mKKKBi4pMUUUALijFFFABijFFFABijFFFMQYooooAMUuKKKADFGKKKADFGKKKAExS496KKAExS4oooAMUY5oooAMUmKKKBi4oxRRQAYoxRRQIMUbfeiigYYoxRRQIMUYoooAMUuMCiigAxRiiigAxQRiiigAxSYoopjFxRiiigQYoxRRQAYoxRRQAYpMUUUDFxRiiigQYoxRRQAYoxRRQAY4oxRRQMMUYoooEGKMUUUAG2giiigYu2jbRRQIMUYoooAMUYoooAMUYoooAMUYoopgGKMUUUDDFGKKKYBjNGKKKADFGKKKADFGKKKBBijFFFACgUYoopAG2jFFFMBMUuOKKKQwxRiiigAxRiiigQYoxRRTAMUYoooGGKMUUUCFxRiiigBMUYoooAXFGKKKBhijFFFAgxRiiimMMUUUUAGOKMUUUgDFGKKKYgxRiiigYuKMUUUAGOKTFFFAC4oxRRQAYoxRRQAYoxRRQAYoxRRQAYoxRRQAYoxRRQIMUYoooGGKMUUUAGKMc0UUCFxSYoooGLikxRRQAuKTFFFABijFFFABilxRRQIMUYoooAMUYoooGAFGKKKBBijFFFABijHNFFMYEUYoooAMUYoooEGKMUUUDFxRiiigAxRiiigAxQBmiigAxRiiigAxRiiigAxRiiimIMUYoopAGKMUUUwAijFFFABijFFFIAxRiiimMKXFFFABikxRRQAYpcUUUAJijFFFABijFFFABilxRRQAYoxRRQAYoxRRQAYoxRRTAMUYoopAGKXFFFACYoxRRQAYpcUUUCDFGKKKBhijFFFABijFFFAgxRiiigAxRiiimAuKMUUUhj1XI60UUUwP/2Q=="

def _upload_rich_menu_image(api_client, rich_menu_id):
    """Upload the rich menu image."""
    try:
        img_data = base64.b64decode(_RICH_MENU_IMG_B64)
        blob_api = MessagingApiBlob(api_client)
        blob_api.set_rich_menu_image(rich_menu_id, body=img_data, _headers={'Content-Type': 'image/jpeg'})
    except Exception as e:
        logger.warning("Rich menu image upload failed: %s", e)


def delete_rich_menu():
    """Delete ALL rich menus."""
    deleted = 0
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            try:
                api.cancel_default_rich_menu()
            except Exception:
                pass
            try:
                existing = api.get_rich_menu_list()
                # Try multiple attribute names
                menus = getattr(existing, 'richmenus', None) or getattr(existing, 'rich_menus', None) or []
                if not menus and hasattr(existing, '__iter__'):
                    menus = list(existing)
                logger.info("Found %d rich menus to delete (type: %s)", len(menus), type(menus).__name__)
                for rm in menus:
                    rid = getattr(rm, 'rich_menu_id', None) or getattr(rm, 'richMenuId', None)
                    if rid:
                        try:
                            api.delete_rich_menu(rid)
                            deleted += 1
                            logger.info("Deleted rich menu: %s", rid)
                        except Exception as e:
                            logger.warning("Failed to delete rich menu %s: %s", rid, e)
            except Exception as e:
                logger.warning("Failed to list rich menus: %s", e)
    except Exception as e:
        logger.warning("Delete rich menu error: %s", e)
    logger.info("Deleted %d rich menus", deleted)
    return deleted


def get_sender_object():
    """Build Sender object for customized bot display name/icon."""
    if not MessageSender or not sender_name:
        return None
    try:
        kwargs = {"name": sender_name}
        if sender_icon and sender_icon.startswith("http"):
            kwargs["icon_url"] = sender_icon
        return MessageSender(**kwargs)
    except Exception:
        return None


def get_insight_followers():
    """Get follower demographics from Insight API."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            # Get follower demographics
            demo = api.get_follower_demographics()
            return {
                "ages": getattr(demo, 'ages', None),
                "genders": getattr(demo, 'genders', None),
                "areas": getattr(demo, 'areas', None),
                "available": getattr(demo, 'available', False),
            }
    except Exception:
        return None


def get_message_delivery_stats(date_str=None):
    """Get message delivery stats for a specific date (reply/push/multicast/broadcast)."""
    try:
        if not date_str:
            date_str = time.strftime("%Y%m%d", time.gmtime(time.time() - 86400))
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            result = {"date": date_str}
            try:
                s = api.get_number_of_sent_reply_messages(var_date=date_str)
                result["reply"] = getattr(s, 'success', 0)
            except Exception:
                result["reply"] = None
            try:
                s = api.get_number_of_sent_push_messages(var_date=date_str)
                result["push"] = getattr(s, 'success', 0)
            except Exception:
                result["push"] = None
            try:
                s = api.get_number_of_sent_multicast_messages(var_date=date_str)
                result["multicast"] = getattr(s, 'success', 0)
            except Exception:
                result["multicast"] = None
            try:
                s = api.get_number_of_sent_broadcast_messages(var_date=date_str)
                result["broadcast"] = getattr(s, 'success', 0)
            except Exception:
                result["broadcast"] = None
            return result
    except Exception:
        return None


def get_message_interaction_stats(request_id):
    """Get user interaction statistics (opens, clicks) for a sent message."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_message_event(request_id=request_id)
            return {
                "overview": getattr(resp, 'overview', None),
                "messages": getattr(resp, 'messages', None),
                "clicks": getattr(resp, 'clicks', None),
            }
    except Exception as e:
        logger.debug("get_message_interaction_stats failed: %s", e)
        return None


def get_statistics_per_unit(date_str=None, num_days=7):
    """Get daily follower statistics (new followers, blocks, unblocks) for recent days.
    Uses LINE Insight API: /v2/bot/insight/followers?date=YYYYMMDD"""
    results = []
    try:
        for i in range(num_days):
            d = time.strftime("%Y%m%d", time.gmtime(time.time() - 86400 * (i + 1)))
            try:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    resp = api.get_number_of_followers(var_date=d)
                    results.append({
                        "date": d,
                        "followers": getattr(resp, 'followers', None),
                        "targeted_reaches": getattr(resp, 'targeted_reaches', None),
                        "blocks": getattr(resp, 'blocks', None),
                    })
            except Exception:
                results.append({"date": d, "followers": None, "targeted_reaches": None, "blocks": None})
    except Exception as e:
        logger.debug("get_statistics_per_unit failed: %s", e)
    return results


def upload_rich_menu_image_custom(rich_menu_id, image_bytes, content_type="image/png"):
    """Upload a custom image to a rich menu (from admin panel upload)."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            blob_api.set_rich_menu_image(rich_menu_id, body=image_bytes, _headers={'Content-Type': content_type})
            return True
    except Exception as e:
        logger.warning("upload_rich_menu_image_custom failed: %s", e)
        return False


def send_imagemap_message(to, base_url, alt_text, width, height, actions):
    """Send an Imagemap message to a user or group.
    actions: list of {"type": "message"|"uri", "text"|"uri": str, "x": int, "y": int, "w": int, "h": int}
    """
    if not ImagemapMessage:
        return False
    try:
        imap_actions = []
        for a in actions:
            area = ImagemapArea(x=a["x"], y=a["y"], width=a["w"], height=a["h"])
            if a.get("type") == "uri":
                imap_actions.append(URIImagemapAction(link_uri=a["uri"], area=area))
            else:
                imap_actions.append(MessageImagemapAction(text=a.get("text", ""), area=area))
        msg = ImagemapMessage(
            base_url=base_url,
            alt_text=alt_text or "圖片選單",
            base_size=ImagemapBaseSize(width=width, height=height),
            actions=imap_actions,
        )
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.push_message(PushMessageRequest(to=to, messages=[msg]))
        return True
    except Exception as e:
        logger.warning("send_imagemap_message failed: %s", e)
        return False


def get_all_follower_ids():
    """Get all follower user IDs (paginated)."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            follower_ids = []
            token = None
            for _ in range(50):  # safety limit
                try:
                    if token:
                        resp = api.get_followers(start=token)
                    else:
                        resp = api.get_followers()
                except AttributeError:
                    if token:
                        resp = api.get_follower_ids(start=token)
                    else:
                        resp = api.get_follower_ids()
                ids = getattr(resp, 'user_ids', None) or getattr(resp, 'follower_ids', None) or []
                follower_ids.extend(ids)
                token = getattr(resp, 'next', None) or getattr(resp, 'next_token', None)
                if not token:
                    break
            return follower_ids
    except Exception as e:
        logger.warning("get_all_follower_ids failed: %s", e)
        return []


# ---- Room (multi-person chat) support ----
def get_room_member_count(room_id):
    """Get number of users in a multi-person chat."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            return api.get_room_members_count(room_id)
    except Exception:
        return None


def fetch_all_room_members(room_id):
    """Fetch all member IDs in a multi-person chat."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            member_ids = []
            token = None
            while True:
                try:
                    if token:
                        resp = api.get_room_member_ids(room_id, start=token)
                    else:
                        resp = api.get_room_member_ids(room_id)
                except AttributeError:
                    break
                ids = getattr(resp, 'member_user_ids', None) or getattr(resp, 'member_ids', None) or []
                member_ids.extend(ids)
                token = getattr(resp, 'next', None) or getattr(resp, 'next_token', None)
                if not token:
                    break
            return member_ids
    except Exception as e:
        logger.warning("fetch_all_room_members failed: %s", e)
        return []


def get_room_member_profile(room_id, user_id):
    """Get profile of a member in a multi-person chat."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            profile = api.get_room_member_profile(room_id, user_id)
            return {
                "display_name": getattr(profile, 'display_name', ''),
                "user_id": getattr(profile, 'user_id', ''),
                "picture_url": getattr(profile, 'picture_url', ''),
            }
    except Exception:
        return None


# ---- Rich Menu enhanced ----
def get_rich_menu_by_id(rich_menu_id):
    """Get a single rich menu by ID."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            rm = api.get_rich_menu(rich_menu_id)
            return {
                "id": getattr(rm, 'rich_menu_id', ''),
                "name": getattr(rm, 'name', ''),
                "size": {"width": getattr(getattr(rm, 'size', None), 'width', 0), "height": getattr(getattr(rm, 'size', None), 'height', 0)} if getattr(rm, 'size', None) else None,
                "chat_bar_text": getattr(rm, 'chat_bar_text', ''),
                "selected": getattr(rm, 'selected', False),
                "areas_count": len(getattr(rm, 'areas', []) or []),
            }
    except Exception as e:
        logger.warning("get_rich_menu_by_id failed: %s", e)
        return None


def get_default_rich_menu_id():
    """Get the ID of the current default rich menu."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_default_rich_menu_id()
            return getattr(resp, 'rich_menu_id', None)
    except Exception:
        return None


def get_user_rich_menu_id(user_id):
    """Get the rich menu ID linked to a specific user."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_rich_menu_id_of_user(user_id)
            return getattr(resp, 'rich_menu_id', None)
    except Exception:
        return None


def validate_rich_menu_obj(rich_menu_dict):
    """Validate a rich menu object before creating it."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.validate_rich_menu_object(rich_menu_dict)
            return {"valid": True}
    except Exception as e:
        return {"valid": False, "error": str(e)}


def download_rich_menu_image(rich_menu_id):
    """Download the image of a rich menu."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            content = blob_api.get_rich_menu_image(rich_menu_id)
            return content
    except Exception as e:
        logger.warning("download_rich_menu_image failed: %s", e)
        return None


# ---- Rich Menu Alias enhanced ----
def get_rich_menu_alias(alias_id):
    """Get rich menu alias info by ID."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_rich_menu_alias(alias_id)
            return {
                "alias_id": getattr(resp, 'rich_menu_alias_id', ''),
                "rich_menu_id": getattr(resp, 'rich_menu_id', ''),
            }
    except Exception:
        return None


def list_rich_menu_aliases():
    """Get list of all rich menu aliases."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_rich_menu_alias_list()
            aliases_raw = getattr(resp, 'aliases', []) or []
            return [{"alias_id": getattr(a, 'rich_menu_alias_id', ''), "rich_menu_id": getattr(a, 'rich_menu_id', '')} for a in aliases_raw]
    except Exception:
        return []


def update_rich_menu_alias(alias_id, new_rich_menu_id):
    """Update an existing rich menu alias to point to a different menu."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            from linebot.v3.messaging import UpdateRichMenuAliasRequest
            api.update_rich_menu_alias(alias_id, UpdateRichMenuAliasRequest(rich_menu_id=new_rich_menu_id))
            return True
    except Exception as e:
        logger.warning("update_rich_menu_alias failed: %s", e)
        return False


def list_rich_menus():
    """List all existing rich menus."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            result = api.get_rich_menu_list()
            raw = getattr(result, 'richmenus', None) or getattr(result, 'rich_menus', None) or []
            menus = []
            for rm in raw:
                menus.append({
                    "id": getattr(rm, 'rich_menu_id', '') or getattr(rm, 'richMenuId', ''),
                    "name": getattr(rm, 'name', ''),
                    "selected": getattr(rm, 'selected', False),
                    "chat_bar_text": getattr(rm, 'chat_bar_text', ''),
                })
            return menus
    except Exception as e:
        logger.warning("list_rich_menus failed: %s", e)
        return []


def link_rich_menu_to_user(user_id, rich_menu_id):
    """Link a specific rich menu to a user."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.link_rich_menu_id_to_user(user_id, rich_menu_id)
            return True
    except Exception:
        return False


def unlink_rich_menu_from_user(user_id):
    """Unlink rich menu from a user."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.unlink_rich_menu_id_from_user(user_id)
            return True
    except Exception:
        return False


def multicast_message(user_ids, text):
    """Send a message to multiple users at once."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.multicast(MulticastRequest(
                to=user_ids,
                messages=[TextMessage(text=text)]
            ))
            return True
    except Exception as e:
        logger.warning("Multicast failed: %s", e)
        return False


def build_confirm_template(text, yes_label, yes_data, no_label, no_data):
    """Build a ConfirmTemplate message for yes/no interactions."""
    if not TemplateMessage:
        return None
    try:
        return TemplateMessage(
            alt_text=text,
            template=ConfirmTemplate(
                text=text,
                actions=[
                    PostbackAction(label=yes_label, data=yes_data),
                    PostbackAction(label=no_label, data=no_data),
                ]
            )
        )
    except Exception:
        return None


def build_carousel(columns):
    """Build a CarouselTemplate message.
    columns: list of {"title": str, "text": str, "actions": [{"label": str, "text": str}]}
    """
    if not TemplateMessage or not CarouselTemplate:
        return None
    try:
        cols = []
        for c in columns[:10]:  # max 10 columns
            actions = [MessageAction(label=a["label"], text=a["text"]) for a in c.get("actions", [])]
            cols.append(CarouselColumn(
                title=c.get("title", "")[:40],
                text=c.get("text", "")[:60],
                actions=actions[:3]  # max 3 actions
            ))
        return TemplateMessage(
            alt_text=columns[0].get("title", "選單"),
            template=CarouselTemplate(columns=cols)
        )
    except Exception:
        return None


def broadcast_message(text):
    """Broadcast a message to all bot followers."""
    if not BroadcastRequest:
        return False
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.broadcast(BroadcastRequest(
                messages=[TextMessage(text=text)]
            ))
            return True
    except Exception as e:
        logger.warning("Broadcast failed: %s", e)
        return False


def manage_rich_menu_alias(alias_id, rich_menu_id, action="create"):
    """Create or delete a Rich Menu alias."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            if action == "create" and CreateRichMenuAliasRequest:
                api.create_rich_menu_alias(CreateRichMenuAliasRequest(
                    rich_menu_alias_id=alias_id,
                    rich_menu_id=rich_menu_id
                ))
                return True
            elif action == "delete":
                api.delete_rich_menu_alias(alias_id)
                return True
    except Exception as e:
        logger.warning("Rich menu alias %s failed: %s", action, e)
    return False


def batch_link_rich_menu(user_ids, rich_menu_id):
    """Link a rich menu to multiple users at once."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.link_rich_menu_id_to_users(
                rich_menu_id=rich_menu_id,
                user_ids=user_ids
            )
            return True
    except Exception as e:
        logger.warning("Batch rich menu link failed: %s", e)
        return False


def batch_unlink_rich_menu(user_ids):
    """Unlink rich menu from multiple users at once."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.unlink_rich_menu_id_from_users(user_ids=user_ids)
            return True
    except Exception as e:
        logger.warning("Batch rich menu unlink failed: %s", e)
        return False


def add_line_emoji_to_text(text, emoji_product_id="5ac1bfd5040ab15980c9b435", emoji_ids=None):
    """Build a TextMessage with LINE emojis embedded.
    This is a helper - LINE emojis use $placeholder in text + emojis array."""
    # LINE emoji format: text has $ placeholders, emojis array maps each
    # For simplicity, return a TextMessage with emoji support
    try:
        if not emoji_ids:
            return TextMessage(text=text)
        emojis = []
        for i, eid in enumerate(emoji_ids):
            emojis.append({
                "index": text.index("$", sum(1 for c in text[:text.index("$")] if True)),
                "productId": emoji_product_id,
                "emojiId": eid
            })
        return TextMessage(text=text, emojis=emojis)
    except Exception:
        return TextMessage(text=text)


def build_translation_flex(original, translated, src_flag, tgt_flag, sender_name_display=None, quoted_text=None):
    """Build a Flex Message for translation with original + translated text."""
    try:
        body_contents = []
        # Quoted message context (if replying to another message)
        if quoted_text:
            qt = quoted_text[:50] + "..." if len(quoted_text) > 50 else quoted_text
            body_contents.append({
                "type": "text", "text": "↩ " + qt,
                "size": "xxs", "color": "#6a6a7a", "wrap": True, "margin": "none",
                "style": "italic"
            })
            body_contents.append({"type": "separator", "margin": "sm"})
        # Sender name
        if sender_name_display:
            body_contents.append({
                "type": "text", "text": sender_name_display,
                "size": "xs", "color": "#8a8a9a", "margin": "none"
            })
        # Original text
        body_contents.append({
            "type": "text", "text": src_flag + " " + original,
            "size": "sm", "color": "#b0b0b0", "wrap": True, "margin": "sm"
        })
        # Separator
        body_contents.append({"type": "separator", "margin": "md"})
        # Translated text
        body_contents.append({
            "type": "text", "text": tgt_flag + " " + translated,
            "size": "md", "color": "#ffffff", "wrap": True, "margin": "md", "weight": "bold"
        })

        flex_obj = {
            "type": "bubble",
            "size": "kilo",
            "body": {
                "type": "box", "layout": "vertical",
                "contents": body_contents,
                "backgroundColor": "#1a1a2e",
                "paddingAll": "16px",
                "cornerRadius": "12px"
            }
        }
        return FlexMessage(
            alt_text=tgt_flag + " " + translated,
            contents=FlexContainer.from_dict(flex_obj)
        )
    except Exception as e:
        logger.warning("Flex message build failed, falling back to text: %s", e)
        return None


def build_quick_reply(group_id=None):
    """Build Quick Reply buttons for translation messages, respecting per-group command toggles."""
    try:
        # Core buttons always shown
        items = [
            QuickReplyItem(action=MessageAction(label="📖 說明/Info", text="/help")),
        ]
        # Command-linked buttons: only show if that command is enabled for the group
        if is_cmd_enabled(group_id, 'qry'):
            items.append(QuickReplyItem(action=MessageAction(label="🔍 儲區/Gudang", text="/qry ")))
        items.append(QuickReplyItem(action=MessageAction(label="❌ 標錯/Wrong", text="/wrong ")))
        items.append(QuickReplyItem(action=MessageAction(label="❌ 不翻我/Skip", text="/skip")))
        items.append(QuickReplyItem(action=MessageAction(label="✅ 翻譯我/Unskip", text="/unskip")))
        if is_cmd_enabled(group_id, 'pw1'):
            items.append(QuickReplyItem(action=MessageAction(label="🔑 班長密碼/PW1", text="/pw1")))
        if is_cmd_enabled(group_id, 'pw2'):
            items.append(QuickReplyItem(action=MessageAction(label="🏭 儲運密碼/PW2", text="/pw2")))
        if is_cmd_enabled(group_id, 'pkg'):
            items.append(QuickReplyItem(action=MessageAction(label="📦 包裝碼/Kemas", text="/pkg ")))
        if is_cmd_enabled(group_id, 'scrap'):
            items.append(QuickReplyItem(action=MessageAction(label="🎨 廢料色/Warna", text="/scrap")))
        # Camera quick reply button (opens camera directly)
        if MsgCameraAction and get_group_feature(group_id, 'camera_qr'):
            try:
                items.append(QuickReplyItem(action=MsgCameraAction(label="📷 拍照/Foto")))
            except Exception:
                pass
        # Clipboard quick reply button (copy useful text)
        if MsgClipboardAction and get_group_feature(group_id, 'clipboard_qr'):
            try:
                items.append(QuickReplyItem(action=MsgClipboardAction(
                    label="📋 複製儲區指令",
                    clipboard_text="/qry "
                )))
            except Exception:
                pass
        # Camera Roll quick reply button (opens photo album)
        if MsgCameraRollAction and get_group_feature(group_id, 'camera_roll_qr'):
            try:
                items.append(QuickReplyItem(action=MsgCameraRollAction(label="🖼️ 相簿/Album")))
            except Exception:
                pass
        # Location quick reply button (share location)
        if MsgLocationAction and get_group_feature(group_id, 'location_qr'):
            try:
                items.append(QuickReplyItem(action=MsgLocationAction(label="📍 位置/Lokasi")))
            except Exception:
                pass
        # URI-based buttons → 改成 MessageAction,讓使用者能複製網址 / 用外部瀏覽器開啟
        # (LINE 的 URIAction 強制用內建瀏覽器,無法複製網址)
        try:
            items.append(QuickReplyItem(action=MessageAction(
                label="💡 提案/Saran",
                text="/saran"
            )))
            items.append(QuickReplyItem(action=MessageAction(
                label="📅 差勤/Absen",
                text="/absen"
            )))
        except Exception:
            pass
        return QuickReply(items=items)
    except Exception:
        return None


# ─── Admin Panel ────────────────────────────────────────

ADMIN_HTML = '''<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="UTF-8">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<meta name="viewport" content="width=device-width,initial-scale=1,user-scalable=no">
<title>翻譯Bot 管理後台</title>
<script src="https://accounts.google.com/gsi/client" async defer></script>
<link rel="manifest" href="/manifest.json">
<meta name="theme-color" content="#7c6fef">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="Bot管理">
<link rel="apple-touch-icon" href="/icon-192.png">
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#0a0a0a;color:#e0e0e0;min-height:100vh;padding-bottom:env(safe-area-inset-bottom)}
.header{background:linear-gradient(135deg,#5b6abf,#8b5fbf 50%,#b07cc3);padding:18px 16px;position:sticky;top:0;z-index:100;box-shadow:0 2px 12px rgba(91,106,191,.3)}
.header h1{font-size:18px;color:#fff;display:flex;align-items:center;gap:8px}
.header .platform{font-size:13px;font-weight:400;opacity:.8}
.login-wrap{display:flex;justify-content:center;align-items:center;min-height:80vh;padding:20px}
.login-box{background:#1a1a2e;border-radius:16px;padding:32px 24px;width:100%;max-width:360px;box-shadow:0 4px 24px rgba(0,0,0,.4);border:1px solid #2a2a3e}
.login-box h2{text-align:center;margin-bottom:24px;color:#7c6fef}
.input-field{width:100%;padding:14px 16px;border-radius:12px;border:1px solid #2a2a3e;background:#0d0d1a;color:#fff;font-size:16px;margin-bottom:16px;outline:none;transition:border .2s}
.input-field:focus{border-color:#7c6fef}
.btn{display:block;width:100%;padding:14px;border:none;border-radius:12px;font-size:16px;font-weight:600;cursor:pointer;transition:all .2s}
.btn-primary{background:#7c6fef;color:#fff}
.btn-primary:active{background:#6358d4;transform:scale(.98)}
.btn-red{background:rgba(240,71,71,.15);color:#f04747;border:1px solid rgba(240,71,71,.3)}
.btn-red:active{background:rgba(240,71,71,.25);transform:scale(.98)}
.btn-sm{padding:8px 14px;font-size:13px;width:auto;border-radius:8px;display:inline-block}
.btn-dark{background:#2a2a3e;color:#e0e0e0;border:1px solid #3a3a4e}
.tabs{display:flex;background:#0a0a0a;border-bottom:1px solid #2a2a3e;position:sticky;top:56px;z-index:99;overflow-x:auto;-webkit-overflow-scrolling:touch;scrollbar-width:none}
.tabs::-webkit-scrollbar{display:none}
.tab{flex:none;padding:12px 16px;text-align:center;font-size:13px;font-weight:400;color:#8a8a9a;cursor:pointer;border-bottom:2px solid transparent;transition:all .2s;white-space:nowrap}
.tab.active{color:#7c6fef;font-weight:700;border-bottom-color:#7c6fef}
.panel{display:none;padding:16px}
.panel.active{display:block}

/* Stats grid */
.stats-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}
.stat-card{background:#1a1a2e;border:1px solid #2a2a3e;border-radius:12px;padding:16px 12px;display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:80px}
.stat-value{font-size:28px;font-weight:700;color:#e0e0e0;font-family:ui-monospace,SFMono-Regular,"SF Mono",monospace;letter-spacing:-.5px}
.stat-value.highlight{color:#7c6fef}
.stat-label{font-size:12px;color:#8a8a9a;margin-top:4px;display:flex;align-items:center;gap:4px}

/* Cards */
.card{background:#1a1a2e;border-radius:12px;padding:16px;margin-bottom:12px;border:1px solid #2a2a3e}
.card-title{font-size:15px;font-weight:600;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center}
.card-sub{font-size:12px;color:#8a8a9a}
.badge{display:inline-block;padding:3px 10px;border-radius:6px;font-size:12px;font-weight:600}
.badge-on{background:rgba(67,181,129,.15);color:#43b581}
.badge-off{background:rgba(240,71,71,.15);color:#f04747}
.badge-yellow{background:rgba(250,166,26,.15);color:#faa61a}

/* Feature badges */
.feat-badges{display:flex;gap:8px;flex-wrap:wrap;margin:10px 0}
.feat-badge{display:inline-flex;align-items:center;gap:4px;padding:4px 10px;border-radius:6px;font-size:12px;font-weight:500}
.feat-badge.on{background:rgba(67,181,129,.12);color:#43b581;border:1px solid rgba(67,181,129,.25)}
.feat-badge.off{background:rgba(100,100,120,.12);color:#8a8a9a;border:1px solid rgba(100,100,120,.2)}
.feat-badge.ask{background:rgba(124,111,239,.15);color:#a899ff;border:1px solid rgba(124,111,239,.35)}

/* Select */
.sel{appearance:none;background:#2a2a3e;color:#e0e0e0;border:1px solid #3a3a4e;border-radius:8px;padding:6px 28px 6px 12px;font-size:13px;cursor:pointer;outline:none}
.sel-wrap{position:relative;display:inline-block}
.sel-wrap::after{content:"▼";position:absolute;right:8px;top:50%;transform:translateY(-50%);pointer-events:none;color:#8a8a9a;font-size:10px}

/* Channel dropdown (full width) */
.ch-select{width:100%;appearance:none;background:#1a1a2e;color:#e0e0e0;border:1px solid #3a3a4e;border-radius:12px;padding:14px 40px 14px 16px;font-size:15px;font-weight:500;cursor:pointer;outline:none}
.ch-select-wrap{position:relative}
.ch-select-wrap::after{content:"▼";position:absolute;right:16px;top:50%;transform:translateY(-50%);pointer-events:none;color:#8a8a9a;font-size:12px}

/* Toggle */
.toggle{position:relative;width:48px;height:26px;cursor:pointer;display:inline-block;flex-shrink:0}
.toggle input{display:none}
.toggle .slider{position:absolute;inset:0;background:#3a3a4a;border-radius:13px;transition:.2s}
.toggle .slider:before{content:"";position:absolute;height:20px;width:20px;left:3px;bottom:3px;background:#fff;border-radius:50%;transition:.2s;box-shadow:0 1px 3px rgba(0,0,0,.3)}
.toggle input:checked+.slider{background:#7c6fef}
.toggle input:checked+.slider:before{transform:translateX(22px)}

/* Whitelist / user items */
.wl-item{display:flex;align-items:center;justify-content:space-between;padding:14px 16px;border-bottom:1px solid #2a2a3e}
.wl-item:last-child{border-bottom:none}

/* User card */
.user-card{background:#1a1a2e;border:1px solid #2a2a3e;border-radius:12px;padding:16px;margin-bottom:12px}
.user-name{font-size:16px;font-weight:700;margin-bottom:6px}
.user-id{font-size:12px;color:#8a8a9a}
.user-admin-row{display:flex;align-items:center;justify-content:flex-end;gap:8px;margin-top:10px}
.admin-label{font-size:13px;color:#faa61a}

.empty{text-align:center;color:#5a5a6a;padding:32px 16px;font-size:14px}
.toast{position:fixed;bottom:80px;left:50%;transform:translateX(-50%);background:rgba(30,30,50,.95);color:#fff;padding:10px 24px;border-radius:10px;font-size:14px;z-index:200;opacity:0;transition:all .25s;pointer-events:none;box-shadow:0 4px 20px rgba(0,0,0,.4);border:1px solid rgba(124,111,239,.3)}
.toast.show{opacity:1}

/* DM section in groups panel */
.dm-section{background:#1a1a2e;border:1px solid #2a2a3e;border-radius:12px;padding:16px;margin-bottom:12px}
.dm-toggle-row{display:flex;align-items:center;justify-content:space-between;margin-bottom:12px}
</style>
</head>
<body>
<div id="app">

<!-- Login -->
<div id="loginPage">
<div class="header"><h1>🤖 翻譯Bot 管理後台 <span class="platform">LINE</span></h1></div>
<div class="login-wrap">
<div class="login-box">
<h2>🔒 管理員登入</h2>
<div style="font-size:11px;color:#666;margin-bottom:8px">v3.0</div>
<div style="display:flex;gap:6px;margin-bottom:10px">
<button id="modeSuper" class="btn btn-primary btn-sm" style="flex:1;font-size:12px" onclick="switchLoginMode(&apos;super&apos;)">超級管理員</button>
<button id="modeManager" class="btn btn-sm" style="flex:1;font-size:12px;background:#2a2a3e;color:#aaa;border:1px solid #3a3a4e;border-radius:6px" onclick="switchLoginMode(&apos;manager&apos;)">管理員</button>
</div>
<div id="superLoginFields">
<input class="input-field" id="pwInput" type="password" placeholder="輸入管理密碼" autocomplete="off" onkeydown="if(event.key===&apos;Enter&apos;)document.getElementById(&apos;loginBtn&apos;).click()">
</div>
<div id="managerLoginFields" style="display:none">
<div id="googleBtnWrap"></div>
<div style="font-size:11px;color:#666;margin-top:6px">使用超級管理員指定的 Google 帳號登入</div>
</div>
<div id="loginMsg" style="color:#f04747;font-size:12px;min-height:18px;margin-top:4px"></div>
<button class="btn btn-primary" id="loginBtn" type="button">登入</button>
</div>
</div>
</div>
<script>
var _loginMode='super';
var GCID='__GOOGLE_CLIENT_ID__';
function switchLoginMode(mode){
  _loginMode=mode;
  document.getElementById('superLoginFields').style.display=mode==='super'?'block':'none';
  document.getElementById('managerLoginFields').style.display=mode==='manager'?'block':'none';
  document.getElementById('modeSuper').style.background=mode==='super'?'#7c6fef':'#2a2a3e';
  document.getElementById('modeSuper').style.color=mode==='super'?'#fff':'#aaa';
  document.getElementById('modeManager').style.background=mode==='manager'?'#7c6fef':'#2a2a3e';
  document.getElementById('modeManager').style.color=mode==='manager'?'#fff':'#aaa';
  document.getElementById('loginMsg').textContent='';
  document.getElementById('loginBtn').style.display=mode==='super'?'block':'none';
  if(mode==='manager') initGoogleBtn();
}
var _gsiInited=false;
function initGoogleBtn(){
  if(_gsiInited)return;
  _gsiInited=true;
  if(typeof google==='undefined'||!google.accounts){
    setTimeout(initGoogleBtn,500);_gsiInited=false;return;
  }
  google.accounts.id.initialize({client_id:GCID,callback:handleGoogleLogin});
  google.accounts.id.renderButton(document.getElementById('googleBtnWrap'),{theme:'outline',size:'large',width:280,text:'signin_with',locale:'zh-TW'});
}
function handleGoogleLogin(response){
  var m=document.getElementById('loginMsg');
  m.textContent='Google 驗證中...';m.style.color='#aaa';
  fetch(window.location.origin+'/api/admin/manager-login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({credential:response.credential})})
  .then(function(r){return r.json()})
  .then(function(d){
    if(d&&d.ok){
      m.textContent='';
      document.getElementById('loginPage').style.display='none';
      document.getElementById('mainPage').style.display='block';
      window._ADMIN_KEY='';
      window._MANAGER_TABS=d.tabs;
      window._MANAGER_ID=d.user_id||'';
      applyTabFilter(d.tabs);
      if(typeof loadAll==='function') loadAll();
    }else{m.style.color='#f04747';m.textContent=d.error||'登入失敗'}
  }).catch(function(e){m.style.color='#f04747';m.textContent='連線錯誤: '+e.message});
}
function applyTabFilter(tabs){
  var allTabs=document.querySelectorAll('.tabs .tab');
  allTabs.forEach(function(t,i){
    var key=TAB_KEYS[i];
    if(tabs&&tabs.indexOf(key)<0){t.style.display='none'}
    else{t.style.display=''}
  });
  if(tabs&&tabs.length>0){switchTab(tabs[0])}
  else{switchTab('overview')}
}
document.getElementById('loginBtn').addEventListener('click',function(){
  var m=document.getElementById('loginMsg');
  m.textContent='登入中...';m.style.color='#aaa';
  var k=document.getElementById('pwInput').value.trim();
  if(!k){m.textContent='請輸入密碼';m.style.color='#f04747';return}
  fetch(window.location.origin+'/api/admin/status',{headers:{'X-Admin-Key':k}})
  .then(function(r){if(!r.ok)throw new Error('HTTP '+r.status);return r.json()})
  .then(function(d){
    if(d&&d.ok){
      m.textContent='';
      document.getElementById('loginPage').style.display='none';
      document.getElementById('mainPage').style.display='block';
      window._ADMIN_KEY=k;
      try{localStorage.setItem('bot_admin_key',k)}catch(e){}
      if(typeof KEY!=='undefined') KEY=k;
      applyTabFilter(null);
      if(typeof loadAll==='function') loadAll();
    }else{m.style.color='#f04747';m.textContent='密碼錯誤'}
  }).catch(function(e){m.style.color='#f04747';m.textContent='連線錯誤: '+e.message});
});
document.getElementById('pwInput').addEventListener('keydown',function(e){
  if(e.key==='Enter') document.getElementById('loginBtn').click();
});
</script>

<!-- Main -->
<div id="mainPage" style="display:none">
<div class="header"><h1>🤖 翻譯Bot 管理後台 <span class="platform">LINE</span></h1></div>
<div class="tabs">
<div class="tab active" onclick="switchTab('overview')">總覽</div>
<div class="tab" onclick="switchTab('groups')">群組</div>
<div class="tab" onclick="switchTab('skip')">白名單</div>
<div class="tab" onclick="switchTab('users')">使用者</div>
<div class="tab" onclick="switchTab('names')">保護名單</div>
<div class="tab" onclick="switchTab('storage')">儲區</div>
<div class="tab" onclick="switchTab('packaging')">包裝碼</div>
<div class="tab" onclick="switchTab('passwords')">密碼</div>
<div class="tab" onclick="switchTab('scrap')">廢料色</div>
<div class="tab" onclick="switchTab('insight')">數據</div>
<div class="tab" onclick="switchTab('examples')">翻譯範例</div>
<div class="tab" onclick="switchTab('forms')">表單</div>
<div class="tab" onclick="switchTab('settings')">設定</div>
</div>

<!-- Overview Panel -->
<div class="panel active" id="panel-overview">
<div class="stats-grid" id="statsGrid">
<div class="stat-card"><div class="stat-value" id="st-uptime">0h 0m</div><div class="stat-label">⏱ 運行時間</div></div>
<div class="stat-card"><div class="stat-value" id="st-text">0</div><div class="stat-label">💬 文字翻譯</div></div>
<div class="stat-card"><div class="stat-value" id="st-image">0</div><div class="stat-label">🖼️ 圖片翻譯</div></div>
<div class="stat-card"><div class="stat-value" id="st-voice">0</div><div class="stat-label">🎤 語音翻譯</div></div>
<div class="stat-card"><div class="stat-value" id="st-wo">0</div><div class="stat-label">📋 工單偵測</div></div>
<div class="stat-card"><div class="stat-value" id="st-cmd">0</div><div class="stat-label">⌨️ 指令</div></div>
<div class="stat-card"><div class="stat-value highlight" id="st-cust">0</div><div class="stat-label">👥 客戶</div></div>
<div class="stat-card"><div class="stat-value highlight" id="st-groups">0</div><div class="stat-label">💬 群組</div></div>
<div class="stat-card"><div class="stat-value" id="st-dm-users">0</div><div class="stat-label">👤 DM使用者</div></div>
</div>
<!-- API Usage Card -->
<div class="card" style="margin:16px 16px 0" id="apiUsageCard">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🔑 OpenAI API 用量</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">
<div style="font-size:13px;color:#8a8a9a">Tokens（本次啟動）</div>
<div style="font-size:13px;text-align:right" id="st-tokens">0</div>
<div style="font-size:13px;color:#8a8a9a">預估花費</div>
<div style="font-size:13px;text-align:right" id="st-cost">$0.00</div>
</div>
<a href="https://platform.openai.com/settings/organization/billing/overview" target="_blank" style="display:block;margin-top:12px;padding:10px;text-align:center;background:#2a2a3e;border:1px solid #3a3a4e;border-radius:8px;color:#7c6fef;font-size:13px;font-weight:600;text-decoration:none">💳 查看 API 餘額</a>
</div>
</div>

<!-- Groups Panel -->
<div class="panel" id="panel-groups">
<!-- DM Section -->
<div class="dm-section">
<div class="dm-toggle-row">
<span style="font-weight:600;font-size:15px">📨 私訊 DM 翻譯</span>
<label class="toggle"><input type="checkbox" id="dmToggle" onchange="toggleDM()"><span class="slider"></span></label>
</div>
<div class="card-sub">總開關關閉時，只有白名單內的人可以私訊翻譯</div>
<div id="dmWlList" style="margin-top:10px"></div>
</div>
<div id="groupList"><div class="empty">載入中...</div></div>
</div>

<!-- Whitelist/Skip Panel -->
<div class="panel" id="panel-skip">
<div class="ch-select-wrap">
<select class="ch-select" id="skipGroupSelect" onchange="loadSkipList()">
<option value="">選擇群組...</option>
</select>
</div>
<div class="card-sub" style="padding:8px 4px;font-size:12px">開啟 = 不翻譯該成員訊息</div>
<div class="card" style="padding:0;overflow:hidden">
<div id="skipListContent"><div class="empty">請先選擇群組</div></div>
</div>
</div>

<!-- Users Panel -->
<div class="panel" id="panel-users">
<div class="ch-select-wrap">
<select class="ch-select" id="usersGroupSelect" onchange="loadUsers()">
<option value="">全部使用者</option>
</select>
</div>
<div id="usersList"><div class="empty">載入中...</div></div>
</div>

<!-- Protected Names Panel -->
<div class="panel" id="panel-names">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:4px">🛡️ 翻譯保護名單</div>
<div class="card-sub" style="margin-bottom:12px">名單內的名字翻譯時會保持原樣不翻（人名、公司名皆可）</div>
<div style="display:flex;gap:8px;margin-bottom:12px">
<input id="newNameInput" type="text" placeholder="輸入名字..." onkeydown="if(event.key==='Enter')addName()" style="flex:1;padding:10px 12px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:14px;outline:none">
<button class="btn btn-primary btn-sm" onclick="addName()">新增</button>
</div>
<div id="namesList"></div>
<div id="namesCount" style="font-size:12px;color:#8a8a9a;margin-top:8px"></div>
</div>
</div>

<!-- Storage Panel -->
<div class="panel" id="panel-storage">
<div class="card">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;font-weight:700;font-size:15px">📦 儲區資料更新</div>
<div class="card-sub" style="margin-bottom:14px">上傳 Excel 檔案自動更新儲區查詢資料</div>
<input type="file" id="storageFile" accept=".xlsx,.xls" style="display:none" onchange="previewStorage()">
<button class="btn btn-primary btn-sm" onclick="document.getElementById('storageFile').click()">選擇 Excel 檔案</button>
<div id="storageFileName" style="margin-top:8px;font-size:13px;color:#8a8a9a"></div>
</div>
<div id="storagePreview"></div>
<div id="storageActions" style="display:none;margin-top:12px">
<button class="btn btn-primary btn-sm" onclick="uploadStorage()">確認更新</button>
</div>
<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:6px">目前資料</div>
<div id="storageStats" style="font-size:14px;margin-bottom:14px">載入中...</div>
<button class="btn btn-dark btn-sm" onclick="downloadJson()">下載 JSON</button>
</div>
</div>

<!-- Packaging Panel -->
<div class="panel" id="panel-packaging">
<div class="card">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;font-weight:700;font-size:15px">📦 包裝碼資料更新</div>
<div class="card-sub" style="margin-bottom:14px">上傳 Excel 檔案（第一列標題列，含代碼/Code欄位）</div>
<input type="file" id="packagingFile" accept=".xlsx,.xls" style="display:none" onchange="previewPackaging()">
<button class="btn btn-primary btn-sm" onclick="document.getElementById('packagingFile').click()">選擇 Excel 檔案</button>
<div id="packagingFileName" style="margin-top:8px;font-size:13px;color:#8a8a9a"></div>
</div>
<div id="packagingPreview"></div>
<div id="packagingActions" style="display:none;margin-top:12px">
<button class="btn btn-primary btn-sm" onclick="uploadPackaging()">確認更新</button>
</div>
<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:6px">目前資料</div>
<div id="packagingStats" style="font-size:14px;margin-bottom:14px">載入中...</div>
<button class="btn btn-dark btn-sm" onclick="downloadPackagingJson()">下載 JSON</button>
</div>
</div>

<!-- Passwords Panel -->
<div class="panel" id="panel-passwords">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">🔑 密碼設定</div>
<div style="margin-bottom:16px">
<div style="font-size:13px;color:#8a8a9a;margin-bottom:6px">班長工號密碼 (使用者傳 /pw1 時顯示)</div>
<textarea id="pw1Input" rows="3" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical"></textarea>
</div>
<div style="margin-bottom:16px">
<div style="font-size:13px;color:#8a8a9a;margin-bottom:6px">儲運工號密碼 (使用者傳 /pw2 時顯示)</div>
<textarea id="pw2Input" rows="3" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical"></textarea>
</div>
<button class="btn btn-primary btn-sm" onclick="savePasswords()">儲存密碼</button>
<div id="pwSaveResult" style="margin-top:8px;font-size:13px"></div>
</div>
</div>

<!-- Scrap Panel -->
<div class="panel" id="panel-scrap">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">🎨 廢料鋼種顏色</div>
<div class="card-sub" style="margin-bottom:14px">使用者傳 /scrap 時顯示的內容</div>
<textarea id="scrapInput" rows="16" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical;font-family:monospace"></textarea>
<div style="margin-top:12px">
<button class="btn btn-primary btn-sm" onclick="saveScrap()">儲存</button>
<div id="scrapSaveResult" style="margin-top:8px;font-size:13px"></div>
</div>
</div>
</div>

<!-- Insight Panel -->
<div class="panel" id="panel-insight">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">📊 好友趨勢（近7日）</div>
<div id="insightTrendChart" style="min-height:160px;position:relative">
<canvas id="trendCanvas" width="600" height="180" style="width:100%;height:180px"></canvas>
</div>
<div id="insightTrendData" style="font-size:12px;color:#8a8a9a;margin-top:8px"></div>
</div>
<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">👥 好友人口統計</div>
<div id="insightDemoData" style="font-size:13px;color:#8a8a9a">載入中...</div>
</div>
<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">📈 昨日發送統計</div>
<div id="insightDelivery" style="font-size:13px;color:#8a8a9a">載入中...</div>
</div>
</div>

<!-- Examples Panel -->
<div class="panel" id="panel-examples">

<!-- ① 主要操作區:最近翻譯日誌(一進來就能標錯) -->
<div class="card">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
<div>
<div style="font-weight:700;font-size:15px">📜 最近翻譯日誌</div>
<div style="font-size:11px;color:#8a8a9a;margin-top:2px">點「標錯+修正」即時加入範例 · 下次翻譯立即生效</div>
</div>
<div style="display:flex;gap:6px;align-items:center">
<label style="font-size:12px;color:#8a8a9a;display:flex;align-items:center;gap:4px"><input type="checkbox" id="logOnlyWrong" onchange="loadTranslationLog()"> 只看錯誤</label>
<button class="btn btn-sm" style="background:#2a2a3e;color:#aaa;border:1px solid #3a3a4e;padding:4px 10px;border-radius:6px;font-size:12px" onclick="loadTranslationLog()">↻</button>
</div>
</div>
<div id="tlogList"><div class="empty">點 ↻ 載入</div></div>
</div>

<!-- ② 累積進度 + 工具 -->
<div class="card" style="margin-top:12px">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
<div style="font-weight:700;font-size:15px">📚 範例累積</div>
<span class="badge" style="background:rgba(67,181,129,.15);color:#43b581;font-size:11px;border:1px solid rgba(67,181,129,.3)">✓ 即時生效</span>
</div>

<div style="display:flex;gap:8px;margin-bottom:10px">
  <div style="flex:1;background:#0d0d1a;padding:10px 6px;border-radius:6px;text-align:center;border:1px solid #2a2a3e">
    <div style="color:#7c6fef;font-weight:700;font-size:20px" id="trainTotal">—</div>
    <div style="color:#8a8a9a;font-size:10px;margin-top:2px">總範例</div>
  </div>
  <div style="flex:1;background:#0d0d1a;padding:10px 6px;border-radius:6px;text-align:center;border:1px solid #2a2a3e">
    <div style="color:#43b581;font-weight:700;font-size:20px" id="trainZh2id">—</div>
    <div style="color:#8a8a9a;font-size:10px;margin-top:2px">中→印</div>
  </div>
  <div style="flex:1;background:#0d0d1a;padding:10px 6px;border-radius:6px;text-align:center;border:1px solid #2a2a3e">
    <div style="color:#faa61a;font-weight:700;font-size:20px" id="trainId2zh">—</div>
    <div style="color:#8a8a9a;font-size:10px;margin-top:2px">印→中</div>
  </div>
  <div style="flex:1;background:#0d0d1a;padding:10px 6px;border-radius:6px;text-align:center;border:1px solid #2a2a3e">
    <div style="color:#f88;font-weight:700;font-size:20px" id="trainWrong30">—</div>
    <div style="color:#8a8a9a;font-size:10px;margin-top:2px">30天標錯</div>
  </div>
</div>

<div style="display:flex;gap:6px">
  <button class="btn btn-sm" style="flex:1;background:#2a4a7c;color:#fff;border:none;padding:8px;border-radius:6px;font-size:12px" onclick="showImportDialog()">📥 批次匯入</button>
  <button class="btn btn-sm" style="flex:1;background:#3a5a3a;color:#fff;border:none;padding:8px;border-radius:6px;font-size:12px" onclick="downloadCsv()">📤 匯出 CSV</button>
  <button class="btn btn-sm" style="background:#2a2a3e;color:#aaa;border:1px solid #3a3a4e;padding:8px 12px;border-radius:6px;font-size:12px" onclick="loadTrainStatus();loadExamples();loadTranslationLog()" title="重新整理">↻</button>
</div>

<div id="exWarning" style="display:none;background:rgba(250,166,26,.12);border:1px solid rgba(250,166,26,.3);border-radius:8px;padding:10px;margin-top:10px;font-size:12px;color:#faa61a"></div>

<!-- 批次匯入彈窗 -->
<div id="importDialog" style="display:none;background:#0d0d1a;border:1px solid #7c6fef;border-radius:8px;padding:12px;margin-top:10px">
  <div style="font-weight:600;font-size:13px;margin-bottom:6px">📥 批次匯入範例</div>
  <div style="font-size:11px;color:#8a8a9a;margin-bottom:6px">
    每行一筆，格式：<code style="background:#2a2a3e;padding:1px 4px;border-radius:3px">方向 | 中文 | 印尼文</code><br>
    方向：<code style="background:#2a2a3e;padding:1px 4px;border-radius:3px">zh2id</code> 或 <code style="background:#2a2a3e;padding:1px 4px;border-radius:3px">id2zh</code>
  </div>
  <textarea id="importText" rows="6" placeholder="zh2id | 砂輪要換了 | Batu gerinda harus diganti
id2zh | 料件後端損傷 | Barang rusak dari belakang" style="width:100%;padding:8px;border-radius:6px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px;font-family:monospace;resize:vertical"></textarea>
  <div style="margin-top:6px;display:flex;gap:6px">
    <button class="btn btn-primary btn-sm" onclick="doImport()">匯入</button>
    <button class="btn btn-sm" style="background:#2a2a3e;color:#aaa;border:1px solid #3a3a4e;padding:6px 10px;border-radius:6px;font-size:12px" onclick="document.getElementById('importDialog').style.display='none'">取消</button>
  </div>
  <div id="importResult" style="font-size:11px;margin-top:6px"></div>
</div>
</div>

<!-- ③ 手動新增單筆範例 -->
<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:8px">＋ 手動新增單筆</div>
<div style="font-size:11px;color:#8a8a9a;margin-bottom:8px">用於補登歷史錯誤,或加入專業術語表</div>
<div style="margin-bottom:6px">
<select id="exDir" onchange="updateExampleInputOrder()" style="width:100%;padding:8px 10px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
<option value="zh2id">中文 → 印尼（zh→id）</option>
<option value="id2zh">印尼 → 中文（id→zh）</option>
</select>
</div>
<!-- v3.9.4: 兩個 input 用獨立容器,JS 依方向切換順序與 placeholder -->
<div id="exInputContainer">
<div id="exSrcWrap" data-role="src" style="margin-bottom:6px">
<div style="font-size:11px;color:#8a8a9a;margin-bottom:3px"><span id="exSrcLabel">原文（中文)</span></div>
<input id="exSrc" type="text" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
</div>
<div id="exTgtWrap" data-role="tgt" style="margin-bottom:6px">
<div style="font-size:11px;color:#8a8a9a;margin-bottom:3px"><span id="exTgtLabel">正確翻譯（印尼文)</span></div>
<input id="exTgt" type="text" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
</div>
</div>
<button class="btn btn-primary btn-sm" onclick="addExample()">＋ 新增範例</button>
<div id="exAddResult" style="font-size:12px;margin-top:4px"></div>
</div>

<!-- ④ 既有範例列表 -->
<div class="card" style="margin-top:12px">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
<div>
<div style="font-weight:700;font-size:15px">📋 範例列表</div>
<span id="exCountBadge" style="font-size:11px;color:#8a8a9a">0 / 5000</span>
</div>
<div style="display:flex;gap:6px">
<input id="exSearch" type="text" placeholder="搜尋..." oninput="renderExamples()" style="padding:4px 8px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px;width:120px">
<select id="exFilter" style="padding:4px 8px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px" onchange="renderExamples()">
<option value="all">全部</option>
<option value="zh2id">中→印尼</option>
<option value="id2zh">印尼→中文</option>
</select>
</div>
</div>
<div id="exList"><div class="empty">載入中...</div></div>
</div>

</div>

<!-- Forms Panel -->
<div class="panel" id="panel-forms">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">📋 表單管理</div>
<button class="btn btn-primary btn-sm" onclick="showCreateForm()">＋ 建立表單</button>
<div id="formsCreateArea" style="display:none;margin-top:12px;padding:12px;background:#0d0d1a;border-radius:8px;border:1px solid #3a3a4e">
<div style="margin-bottom:8px"><input type="text" id="formTitleZh" placeholder="表單標題（中文）" style="width:100%;padding:8px;border-radius:6px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:13px;margin-bottom:6px"><input type="text" id="formTitleId" placeholder="Judul formulir (印尼文)" style="width:100%;padding:8px;border-radius:6px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:13px"></div>
<div style="font-size:13px;font-weight:600;margin:8px 0 4px">欄位 Fields:</div>
<div id="formFieldsList"></div>
<button class="btn btn-sm" style="background:#2a2a3e;color:#aaa;border:1px solid #3a3a4e;margin-top:6px;padding:6px 10px;border-radius:6px;font-size:12px" onclick="addFormField()">＋ 新增欄位</button>
<div style="margin-top:10px;display:flex;gap:8px"><button class="btn btn-primary btn-sm" onclick="submitCreateForm()">建立</button><button class="btn btn-sm" style="background:#3a3a4e;color:#ccc;border:none;padding:6px 12px;border-radius:6px;font-size:12px" onclick="document.getElementById(&apos;formsCreateArea&apos;).style.display=&apos;none&apos;">取消</button></div>
</div>
</div>
<div id="formsList" style="margin-top:10px"></div>
<div id="formDetailArea" style="display:none;margin-top:10px"></div>
</div>

<!-- Settings Panel -->
<div class="panel" id="panel-settings">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">⚙️ 功能設定</div>
<div class="ch-select-wrap" style="margin-bottom:12px">
<select class="ch-select" id="settingsGroupSelect" style="font-size:13px;padding:10px" onchange="loadFeatureSettingsForGroup()">
<option value="">全域預設</option>
</select>
</div>
<div id="settingsCustomBadge" style="display:none;margin-bottom:10px"><span class="badge badge-on" style="font-size:11px">已自訂</span> <span style="font-size:12px;color:#8a8a9a;cursor:pointer;text-decoration:underline" onclick="resetGroupSettings()">重設為預設</span></div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">👋 歡迎訊息</span><br><span style="font-size:12px;color:#8a8a9a">新成員加入時自動發送</span></div>
<label class="toggle"><input type="checkbox" id="welcomeToggle" onchange="toggleFeatureSetting('welcome_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div style="padding:12px 0">
<div style="font-size:13px;color:#8a8a9a;margin-bottom:6px">中文歡迎詞</div>
<textarea id="welcomeZh" rows="2" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical" onblur="saveWelcomeText()"></textarea>
<div style="font-size:13px;color:#8a8a9a;margin:8px 0 6px">印尼文歡迎詞</div>
<textarea id="welcomeId" rows="2" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical" onblur="saveWelcomeText()"></textarea>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🎨 Flex 翻譯卡片</span><br><span style="font-size:12px;color:#8a8a9a">關閉後用純文字顯示</span></div>
<label class="toggle"><input type="checkbox" id="flexToggle" onchange="toggleFeatureSetting('flex_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">⚡ Quick Reply 按鈕</span><br><span style="font-size:12px;color:#8a8a9a">翻譯後顯示快捷操作</span></div>
<label class="toggle"><input type="checkbox" id="qrToggle" onchange="toggleFeatureSetting('quick_reply_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🔇 靜音模式</span><br><span style="font-size:12px;color:#8a8a9a">翻譯訊息不震動手機</span></div>
<label class="toggle"><input type="checkbox" id="silentToggle" onchange="toggleFeatureSetting('silent_mode',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🎬 影片 OCR 翻譯</span><br><span style="font-size:12px;color:#8a8a9a">影片截圖自動 OCR 翻譯</span></div>
<label class="toggle"><input type="checkbox" id="videoToggle" onchange="toggleFeatureSetting('video_ocr_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">📍 位置訊息翻譯</span><br><span style="font-size:12px;color:#8a8a9a">翻譯地點名稱和地址</span></div>
<label class="toggle"><input type="checkbox" id="locationToggle" onchange="toggleFeatureSetting('location_translate_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">👁️ 標記已讀</span><br><span style="font-size:12px;color:#8a8a9a">處理訊息時顯示已讀標記</span></div>
<label class="toggle"><input type="checkbox" id="markReadToggle" onchange="toggleFeatureSetting('mark_read_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🔄 防重複發送</span><br><span style="font-size:12px;color:#8a8a9a">X-Line-Retry-Key 冪等性</span></div>
<label class="toggle"><input type="checkbox" id="retryKeyToggle" onchange="toggleFeatureSetting('retry_key_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">📷 拍照快捷鈕</span><br><span style="font-size:12px;color:#8a8a9a">Quick Reply 加入拍照按鈕</span></div>
<label class="toggle"><input type="checkbox" id="cameraQrToggle" onchange="toggleFeatureSetting('camera_qr_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">📋 複製快捷鈕</span><br><span style="font-size:12px;color:#8a8a9a">Quick Reply 加入複製指令按鈕</span></div>
<label class="toggle"><input type="checkbox" id="clipboardQrToggle" onchange="toggleFeatureSetting('clipboard_qr_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🖼️ 相簿快捷鈕</span><br><span style="font-size:12px;color:#8a8a9a">Quick Reply 加入開啟相簿按鈕</span></div>
<label class="toggle"><input type="checkbox" id="cameraRollQrToggle" onchange="toggleFeatureSetting('camera_roll_qr_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">📍 位置快捷鈕</span><br><span style="font-size:12px;color:#8a8a9a">Quick Reply 加入分享位置按鈕</span></div>
<label class="toggle"><input type="checkbox" id="locationQrToggle" onchange="toggleFeatureSetting('location_qr_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🗣️ 翻譯口吻</span><br><span style="font-size:12px;color:#8a8a9a">控制翻譯的語氣風格</span></div>
<select id="toneSelect" style="padding:6px 10px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px" onchange="toggleFeatureSetting('translation_tone',this.value)">
<option value="factory">🏭 工廠現場（建議）</option>
<option value="casual">日常口語</option>
<option value="natural">母語自然風格</option>
<option value="formal">正式書面</option>
</select>
</div>
<div style="padding:4px 0 12px">
<div style="font-size:12px;color:#8a8a9a;margin-bottom:6px">自訂語氣指令（追加在上方選項之後，作為微調補充）</div>
<textarea id="toneCustom" rows="3" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical" placeholder="可選填。例如：「最近主管脾氣差，語氣再緩和」、「補充某個專有詞翻譯」。留空則完全使用上方選項的預設規則。" onblur="toggleFeatureSetting('translation_tone_custom',this.value)"></textarea>
</div>

<div style="border-top:1px solid #2a2a3e;padding-top:12px;margin-top:4px">
<div style="font-weight:600;margin-bottom:8px">🤖 AI 模型自動切換</div>
<div class="card-sub" style="margin-bottom:8px">訊息超過指定字數自動升級為 GPT-4o（翻譯更流暢但較貴）。設為 0 表示全部用預設模型。</div>
<div style="display:flex;gap:8px;margin-bottom:8px;align-items:center">
<div style="font-size:13px;color:#8a8a9a;white-space:nowrap">字數門檻</div>
<input id="modelThreshold" type="number" min="0" value="0" style="width:80px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;text-align:center">
<div style="font-size:12px;color:#8a8a9a">字</div>
</div>
<div style="display:flex;gap:8px;margin-bottom:8px">
<div style="flex:1">
<div style="font-size:12px;color:#8a8a9a;margin-bottom:4px">預設模型（短訊息）</div>
<select id="modelDefault" onchange="onModelChange()" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px">
<option value="gpt-4.1-mini">⭐ gpt-4.1-mini（$0.40 / $1.60）</option>
<option value="gpt-5-mini">gpt-5-mini（$0.25 / $2.00）</option>
<option value="gpt-5-nano">gpt-5-nano（$0.05 / $0.40）</option>
<option value="gpt-4.1-nano">gpt-4.1-nano（$0.10 / $0.40）</option>
<option value="gpt-5.4-mini">gpt-5.4-mini（$0.75 / $4.50）</option>
<option value="gpt-5.4-nano">gpt-5.4-nano（$0.20 / $1.25）</option>
<option value="gpt-4o-mini">gpt-4o-mini（$0.15 / $0.60，舊）</option>
<option value="gpt-4.1">gpt-4.1（$2.00 / $8.00）</option>
<option value="gpt-5">gpt-5（$1.25 / $10.00，需 Tier2+）</option>
<option value="gpt-5.4">gpt-5.4（$2.50 / $15.00）</option>
<option value="gpt-4o">gpt-4o（$2.50 / $10.00，舊）</option>
<option value="gpt-5.5">gpt-5.5🆕（$5.00 / $30.00，貴）</option>
<option value="o4-mini">o4-mini（推理模型）</option>
<option value="o3-mini">o3-mini（推理模型）</option>
</select>
</div>
<div style="flex:1">
<div style="font-size:12px;color:#8a8a9a;margin-bottom:4px">升級模型（長訊息）</div>
<select id="modelUpgrade" onchange="onModelChange()" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px">
<option value="gpt-4.1">⭐ gpt-4.1（$2.00 / $8.00）</option>
<option value="gpt-4.1-mini">gpt-4.1-mini（$0.40 / $1.60）</option>
<option value="gpt-5-mini">gpt-5-mini（$0.25 / $2.00）</option>
<option value="gpt-5.4-mini">gpt-5.4-mini（$0.75 / $4.50）</option>
<option value="gpt-5">gpt-5（$1.25 / $10.00，需 Tier2+）</option>
<option value="gpt-5.4">gpt-5.4（$2.50 / $15.00）</option>
<option value="gpt-4o">gpt-4o（$2.50 / $10.00，舊）</option>
<option value="gpt-5.5">gpt-5.5🆕（$5.00 / $30.00，貴）</option>
<option value="gpt-4o-mini">gpt-4o-mini（$0.15 / $0.60，舊）</option>
<option value="o4-mini">o4-mini（推理模型）</option>
<option value="o3-mini">o3-mini（推理模型）</option>
</select>
</div>
</div>
<button class="btn btn-primary btn-sm" onclick="saveModelSettings()">儲存模型設定</button>
<div id="modelSaveResult" style="font-size:12px;color:#8a8a9a;margin-top:4px"></div>
<div style="font-size:11px;color:#666;margin-top:6px;padding:6px 8px;background:#0d0d1a;border-radius:6px;border:1px solid #2a2a3e;line-height:1.6">
<b>📋 模型選擇指引(v3.9.14 後皆已最佳化)</b><br>
🔹 <b>gpt-4.1-mini</b>($0.40 / $1.60)— 經典款,翻譯特化,Intento 2025 評測 #1。便宜、快、對範例順從。<br>
🔹 <b>gpt-5-mini</b>($0.25 / $2.00)— <b>更便宜</b>,有 reasoning 即使 minimal 模式也對複雜句更好。<br>
🔹 <b>gpt-5.4-mini</b>($0.75 / $4.50)— 最新一代,複雜被動句、長公告處理更強。<br>
🔹 <b>gpt-4.1</b>($2.00 / $8.00)— 4.1 家族旗艦,需穩定再現性時用。<br>
<b>翻譯任務</b>:GPT-5 系列自動套用 <code>reasoning_effort=minimal</code> + <code>verbosity=low</code>(防止 GPT-5 自由發揮拆條列、加 ⚠️)。<br>
<b>OCR / 視覺任務</b>:GPT-5 系列自動套用 <code>reasoning_effort=low</code>,<b>不</b>套用 verbosity(避免輸出被壓縮成摘要)。<br>
價錢格式:每百萬 input/output token (USD)。
</div>

<div style="border-top:1px solid #2a2a3e;padding-top:12px;margin-top:12px">
<div style="font-weight:600;margin-bottom:6px">📷 照片分析模型（Vision / OCR）</div>
<div class="card-sub" style="margin-bottom:8px">處理工單照片、班表、文字截圖。失敗自動 fallback 到 gpt-4o-mini。</div>
<select id="visionModel" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px;margin-bottom:8px">
<option value="gpt-5-mini">⭐ gpt-5-mini（$0.25 / $2.00，視覺強+便宜）</option>
<option value="gpt-5-nano">gpt-5-nano（$0.05 / $0.40，最便宜）</option>
<option value="gpt-4o-mini">gpt-4o-mini（$0.15 / $0.60，舊但穩）</option>
<option value="gpt-5.4-nano">gpt-5.4-nano（$0.20 / $1.25）</option>
<option value="gpt-4.1-mini">gpt-4.1-mini（$0.40 / $1.60）</option>
<option value="gpt-5.4-mini">gpt-5.4-mini（$0.75 / $4.50）</option>
<option value="gpt-5">gpt-5（$1.25 / $10.00）</option>
<option value="gpt-4.1">gpt-4.1（$2.00 / $8.00）</option>
<option value="gpt-5.4">gpt-5.4（$2.50 / $15.00）</option>
<option value="gpt-4o">gpt-4o（$2.50 / $10.00，舊）</option>
<option value="gpt-5.5">gpt-5.5🆕（$5.00 / $30.00，最貴）</option>
</select>
<button class="btn btn-primary btn-sm" onclick="saveVisionModel()">儲存照片模型</button>
<div id="visionSaveResult" style="font-size:12px;color:#8a8a9a;margin-top:4px"></div>
<div style="font-size:11px;color:#666;margin-top:6px;padding:6px 8px;background:#0d0d1a;border-radius:6px;border:1px solid #2a2a3e;line-height:1.6">
⭐ <b>gpt-5-mini</b> 是 OCR 甜蜜點(OpenAI 官方:多模態評分超越 gpt-4o)。<br>
🆕 <b>gpt-5.4-mini</b> 比 gpt-5-mini 快 2 倍,接近 gpt-5.4 旗艦品質,但貴 3 倍。<br>
v3.9.14 後 vision call 已根據官方文件最佳化:<code>reasoning_effort=low</code> + 移除 verbosity 限制 + 3x token budget。<br>
注意:<b>gpt-5.5-mini 目前 OpenAI 未推出</b>(只有 5.4-mini)。
</div>
</div>
</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🎯 翻譯品質進階設定（v3.2-0426e）</div>
<div style="font-size:12px;color:#8a8a9a;margin-bottom:10px">控制翻譯穩定度、雙重檢查、Few-shot 格式</div>

<div style="margin-bottom:12px">
<label style="font-size:13px;color:#e0e0e0;display:block;margin-bottom:4px">🌡️ Temperature（隨機度，0=最穩定）</label>
<input id="ttemp" type="number" step="0.1" min="0" max="1" value="0" style="width:80px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;text-align:center"> <span style="font-size:12px;color:#8a8a9a">建議翻譯用 0~0.3</span>
</div>

<div style="margin-bottom:12px">
<label style="font-size:13px;color:#e0e0e0;display:block;margin-bottom:4px">🎲 Top_p（核採樣，預設 1.0）</label>
<input id="ttopp" type="number" step="0.05" min="0" max="1" value="1.0" style="width:80px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;text-align:center">
</div>

<div style="margin-bottom:12px">
<label style="font-size:13px;color:#e0e0e0;display:block;margin-bottom:4px">🌱 Seed（可重現種子，0=隨機）</label>
<input id="tseed" type="number" min="0" value="0" style="width:120px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;text-align:center"> <span style="font-size:12px;color:#8a8a9a">非 0 時相同輸入會得相同結果</span>
</div>

<div style="margin-bottom:12px;padding:10px;background:#0d0d1a;border-radius:8px;border:1px solid #2a2a3e">
<div style="font-weight:600;margin-bottom:6px">🔍 雙重檢查（反譯比對）</div>
<div style="font-size:12px;color:#8a8a9a;margin-bottom:8px">翻譯後反譯回原文比對相似度，差異大會在訊息加 ⚠️。可避免「對不起」翻成「麻煩你了」這種致命誤譯。</div>
<select id="dcMode" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;margin-bottom:8px">
<option value="off">❌ 關閉</option>
<option value="smart">🧠 智能（長句+關鍵詞，建議）</option>
<option value="keywords_only">🔑 只查關鍵詞</option>
<option value="all">🔁 全部訊息（成本高）</option>
</select>
<label style="font-size:12px;color:#8a8a9a;display:block;margin-bottom:4px">相似度門檻（0.55 建議；越低越寬鬆）</label>
<input id="dcThr" type="number" step="0.05" min="0" max="1" value="0.55" style="width:80px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;text-align:center;margin-bottom:8px">
<label style="font-size:12px;color:#8a8a9a;display:block;margin-bottom:4px">關鍵詞（逗號分隔）</label>
<textarea id="dcKw" rows="2" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px"></textarea>
</div>

<div style="margin-bottom:12px">
<label style="font-size:13px;color:#e0e0e0;display:block;margin-bottom:4px">📚 Few-shot 範例格式</label>
<select id="fsMode" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
<option value="messages">messages 模式（OpenAI 標準，較好）</option>
<option value="system_prompt">system_prompt 模式（舊版）</option>
</select>
</div>

<button class="btn btn-primary btn-sm" onclick="saveAdvancedSettings()">儲存進階設定</button>
<div id="advSaveResult" style="font-size:12px;color:#8a8a9a;margin-top:4px"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🛡️ 官方 API 進階參數（v3.2-0426e）</div>
<div style="font-size:12px;color:#8a8a9a;margin-bottom:10px">OpenAI 官方功能,合規追蹤、防止 GPT 加註解、推理模型專用參數</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="ssEn" type="checkbox" checked onchange="saveOfficialFeatures()"> 啟用 stop sequences + logit_bias（防止 GPT 加「註:」「Translation:」「Catatan:」等多餘文字）
</label>

<div style="margin-bottom:12px">
<label style="font-size:13px;color:#e0e0e0;display:block;margin-bottom:4px">🧠 推理深度（僅 o3/o4/o1 模型有效）</label>
<select id="reEf" onchange="saveOfficialFeatures()" style="width:140px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
<option value="low">low（最快最便宜）</option>
<option value="medium" selected>medium（建議）</option>
<option value="high">high（最深入但最貴）</option>
</select>
</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="suid" type="checkbox" checked onchange="saveOfficialFeatures()"> 傳 user_id 給 OpenAI（雜湊後,合規追蹤)
</label>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="smeta" type="checkbox" checked onchange="saveOfficialFeatures()"> 傳 metadata（group_id, 語言對）給 OpenAI 後台
</label>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🎓 Batch C: 信心度 & 結構化輸出</div>
<div style="font-size:12px;color:#8a8a9a;margin-bottom:10px">利用 logprobs 計算翻譯信心度，低信心翻譯加 ⚠️</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="lpEn" type="checkbox" onchange="saveBatchCD()"> 啟用 Logprobs（信心度計算）
</label>

<div style="margin-bottom:12px">
<label style="font-size:13px;color:#e0e0e0;display:block;margin-bottom:4px">信心度警告門檻（0~1，建議 0.85）</label>
<input id="cfThr" type="number" step="0.05" min="0" max="1" value="0.85" onchange="saveBatchCD()" style="width:80px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;text-align:center">
</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="soEn" type="checkbox" onchange="saveBatchCD()"> 強制 JSON 結構化輸出（實驗中）
</label>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="pcEn" type="checkbox" onchange="saveBatchCD()"> 啟用 Prompt Caching（自動，省 75% input 費用）
</label>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🎯 v3.4 印尼文→中文翻譯品質強化</div>
<div style="font-size:12px;color:#8a8a9a;margin-bottom:10px">針對 ID→ZH 常翻錯的問題,套用三項論文驗證的技巧。建議 CoD + CoT 預設開啟。</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="idZhCoD" type="checkbox" checked onchange="saveIdZhEnhance()"> 啟用 CoD 字典注入（強烈推薦,+5% token）
</label>
<div style="font-size:11px;color:#6a6a7a;margin-bottom:10px;padding-left:24px">自動偵測句中高風險誤譯詞,把正確中文塞到 prompt 開頭(barang→料件、biar→為了)</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="idZhCoT" type="checkbox" checked onchange="saveIdZhEnhance()"> 啟用 CoT 二階段思考（強烈推薦,+20% token）
</label>
<div style="font-size:11px;color:#6a6a7a;margin-bottom:10px;padding-left:24px">讓 GPT 先「展開思考」再翻譯(識別句型→列關鍵資訊點→處理俚語)</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="idZhPivot" type="checkbox" onchange="saveIdZhEnhance()"> 啟用 Pivot via English（成本翻倍,慎用）
</label>
<div style="font-size:11px;color:#6a6a7a;margin-bottom:10px;padding-left:24px">先 ID→EN 再 EN→ZH,長句品質更好,但會多一次 API 呼叫</div>

<div style="margin-bottom:12px">
<label style="font-size:13px;color:#e0e0e0;display:block;margin-bottom:4px">Pivot 觸發字數門檻</label>
<input id="idZhPivotThr" type="number" min="20" max="500" value="80" onchange="saveIdZhEnhance()" style="width:80px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;text-align:center">
<span style="font-size:11px;color:#6a6a7a">字以上才用 pivot(短句不需要)</span>
</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🔤 v3.6 印尼文預處理 + 多路徑反譯</div>
<div style="font-size:12px;color:#8a8a9a;margin-bottom:10px">把 gak/udah/bgt 等簡寫先還原為標準印尼文,翻譯品質提升 8-12%(論文驗證)</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="idPrep" type="checkbox" checked onchange="saveV36Settings()"> 啟用印尼俚語/簡寫標準化(0 API 呼叫,推薦)
</label>
<div style="font-size:11px;color:#6a6a7a;margin-bottom:10px;padding-left:24px">純詞表替換,即時生效。例如 gak→tidak, udah→sudah, biar→agar</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="idPrepNano" type="checkbox" onchange="saveV36Settings()"> 進階:nano 模型語意級規範化(成本 +5%)
</label>
<div style="font-size:11px;color:#6a6a7a;margin-bottom:10px;padding-left:24px">處理詞表抓不到的特殊用法、混雜爪哇語</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="multiPath" type="checkbox" onchange="saveV36Settings()"> 啟用多路徑反譯(成本翻倍,只用於長句)
</label>
<div style="font-size:11px;color:#6a6a7a;margin-bottom:10px;padding-left:24px">平行兩條反譯路徑(直譯 + 經英語),兩條都通過才算 OK</div>

<div style="margin-bottom:12px">
<label style="font-size:13px;color:#e0e0e0;display:block;margin-bottom:4px">多路徑反譯觸發字數</label>
<input id="multiPathMin" type="number" min="20" max="500" value="60" onchange="saveV36Settings()" style="width:80px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;text-align:center">
<span style="font-size:11px;color:#6a6a7a">字以上才用多路徑(短句單路徑就夠)</span>
</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📐 v3.7 段落結構保留</div>
<div style="font-size:12px;color:#8a8a9a;margin-bottom:10px">問題:長公告原文有分段,翻譯後變成連續一坨文字。解法:雙層保護(prompt + 分段翻譯)</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="paraPreserve" type="checkbox" checked onchange="saveV37Settings()"> 在 prompt 加入段落保留指示(0 成本,推薦)
</label>
<div style="font-size:11px;color:#6a6a7a;margin-bottom:10px;padding-left:24px">告訴 GPT「原文有空行就要保留空行」+ 給範例</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="paraSplit" type="checkbox" checked onchange="saveV37Settings()"> 分段獨立翻譯(雙保險,長訊息才走)
</label>
<div style="font-size:11px;color:#6a6a7a;margin-bottom:10px;padding-left:24px">把訊息以空行切段,每段獨立翻譯,然後拼回。100% 保證段落結構正確。</div>

<div style="margin-bottom:12px">
<label style="font-size:13px;color:#e0e0e0;display:block;margin-bottom:4px">分段翻譯觸發字數</label>
<input id="paraThr" type="number" min="20" max="500" value="50" onchange="saveV37Settings()" style="width:80px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;text-align:center">
<span style="font-size:11px;color:#6a6a7a">字以上且含分段的訊息才走分段翻譯</span>
</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📈 翻譯品質監控儀表板</div>
<div style="font-size:12px;color:#8a8a9a;margin-bottom:10px">五大關鍵指標 + 各群組品質分數 + 最近的問題訊息</div>

<div style="display:flex;gap:6px;margin-bottom:10px">
<button class="btn btn-dark btn-sm" onclick="loadQualityStats(7)">📅 過去 7 天</button>
<button class="btn btn-dark btn-sm" onclick="loadQualityStats(1)">📅 今天</button>
<button class="btn btn-dark btn-sm" onclick="loadQualityStats(30)">📅 過去 30 天</button>
</div>

<div id="qsMetrics" style="font-size:13px;color:#8a8a9a;margin-bottom:12px;padding:10px;background:#0d0d1a;border-radius:6px">點上方按鈕載入</div>

<div style="font-weight:600;font-size:13px;color:#e0e0e0;margin-bottom:6px">各語言對</div>
<div id="qsByLang" style="font-size:12px;margin-bottom:12px"></div>

<div style="font-weight:600;font-size:13px;color:#e0e0e0;margin-bottom:6px">各群組品質分數(差的排前面)</div>
<div id="qsByGroup" style="font-size:12px;margin-bottom:12px"></div>

<div style="font-weight:600;font-size:13px;color:#e0e0e0;margin-bottom:6px">最近的問題訊息</div>
<div id="qsIssues" style="font-size:12px;max-height:300px;overflow-y:auto"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📊 Batch D: 翻譯日誌 & 監控</div>
<div style="font-size:12px;color:#8a8a9a;margin-bottom:10px">記錄每筆翻譯，可標記錯誤自動加入修正範例</div>

<label style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
<input id="tlEn" type="checkbox" onchange="saveBatchCD()"> 啟用翻譯日誌
</label>

<div id="tlStats" style="font-size:13px;color:#8a8a9a;margin-bottom:10px;padding:8px;background:#0d0d1a;border-radius:6px">載入中...</div>

<div style="display:flex;gap:6px;margin-bottom:8px;flex-wrap:wrap">
<button class="btn btn-dark btn-sm" onclick="loadTransLog('all')">📋 全部</button>
<button class="btn btn-dark btn-sm" onclick="loadTransLog('warned')">⚠️ 只看警告</button>
<button class="btn btn-dark btn-sm" onclick="loadTransLog('wrong')">❌ 已標錯</button>
<button class="btn btn-dark btn-sm" onclick="clearTransLog()">🗑️ 清空</button>
</div>

<div id="tlList" style="max-height:400px;overflow-y:auto;font-size:12px"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📊 LINE 配額 & 統計</div>
<div id="lineQuotaInfo" style="font-size:13px;color:#8a8a9a">載入中...</div>
<div id="lineInsight" style="font-size:13px;color:#8a8a9a;margin-top:8px"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🔗 Webhook 狀態</div>
<div id="webhookInfo" style="font-size:13px;color:#8a8a9a">載入中...</div>
<button class="btn btn-dark btn-sm" style="margin-top:8px" onclick="testWebhook()">🧪 測試 Webhook</button>
<div id="webhookTestResult" style="font-size:12px;color:#8a8a9a;margin-top:6px"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📈 送出統計（昨日）</div>
<div id="deliveryStats" style="font-size:13px;color:#8a8a9a">載入中...</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">👥 好友清單</div>
<button class="btn btn-dark btn-sm" onclick="loadFollowers()">載入好友列表</button>
<div id="followersList" style="font-size:13px;color:#8a8a9a;margin-top:8px;max-height:200px;overflow-y:auto"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🎨 Rich Menu 管理</div>
<div id="richMenuList" style="font-size:13px;color:#8a8a9a;margin-bottom:8px"></div>
<div id="richMenuDefault" style="font-size:13px;color:#8a8a9a;margin-bottom:8px"></div>
<div style="margin-bottom:8px">
<span style="font-size:13px;color:#8a8a9a">查詢用戶綁定選單：</span>
<div style="display:flex;gap:6px;margin-top:4px">
<input id="rmUserIdInput" type="text" placeholder="user ID" style="flex:1;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px">
<button class="btn btn-dark btn-sm" onclick="checkUserMenu()">查詢</button>
</div>
<div id="rmUserResult" style="font-size:12px;color:#8a8a9a;margin-top:4px"></div>
</div>
<div style="font-size:13px;font-weight:600;margin-bottom:6px">Alias 列表</div>
<div id="rmAliasList" style="font-size:13px;color:#8a8a9a"></div>
<button class="btn btn-dark btn-sm" style="margin-top:6px" onclick="loadAliases()">重新載入 Alias</button>
<div style="margin-top:12px;border-top:1px solid #2a2a3e;padding-top:12px">
<div style="font-size:13px;font-weight:600;margin-bottom:6px">📤 上傳選單圖片</div>
<div class="card-sub" style="margin-bottom:8px">選擇 Rich Menu 後上傳圖片（建議 2500x1686 或 2500x843）</div>
<div style="display:flex;gap:6px;margin-bottom:6px;flex-wrap:wrap">
<select id="rmUploadSelect" style="flex:1;min-width:120px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px"></select>
<label class="btn btn-dark btn-sm" style="cursor:pointer;display:inline-flex;align-items:center">
選擇圖片 <input type="file" id="rmImageFile" accept="image/png,image/jpeg" style="display:none" onchange="previewRmImage(this)">
</label>
</div>
<div id="rmImagePreview" style="display:none;margin-bottom:6px"><img id="rmImagePreviewImg" style="max-width:100%;max-height:120px;border-radius:6px;border:1px solid #3a3a4e"></div>
<button class="btn btn-primary btn-sm" onclick="uploadRmImage()" id="rmUploadBtn" style="display:none">⬆️ 上傳</button>
<div id="rmUploadResult" style="font-size:12px;color:#8a8a9a;margin-top:4px"></div>
</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🗺️ Imagemap 圖片選單</div>
<div class="card-sub" style="margin-bottom:10px">發送可點擊區域的圖片到群組（圖片需為 HTTPS URL，建議 1040px 寬）</div>
<div class="ch-select-wrap" style="margin-bottom:8px">
<select class="ch-select" id="imapGroupSelect" style="font-size:13px;padding:10px"></select>
</div>
<div style="margin-bottom:8px">
<div style="font-size:12px;color:#8a8a9a;margin-bottom:4px">圖片 Base URL（不含 /1040 等尺寸後綴）</div>
<input id="imapBaseUrl" type="text" placeholder="https://example.com/images/menu" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
</div>
<div style="display:flex;gap:8px;margin-bottom:8px">
<div style="flex:1"><div style="font-size:12px;color:#8a8a9a;margin-bottom:4px">寬度</div><input id="imapW" type="number" value="1040" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px"></div>
<div style="flex:1"><div style="font-size:12px;color:#8a8a9a;margin-bottom:4px">高度</div><input id="imapH" type="number" value="1040" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px"></div>
</div>
<div style="font-size:13px;font-weight:600;margin-bottom:6px">點擊區域</div>
<div id="imapActions"></div>
<button class="btn btn-dark btn-sm" style="margin-bottom:8px" onclick="addImapAction()">＋ 新增區域</button>
<div style="margin-top:8px">
<button class="btn btn-primary btn-sm" onclick="sendImap()">📤 發送 Imagemap</button>
</div>
<div id="imapResult" style="font-size:12px;color:#8a8a9a;margin-top:6px"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🤖 Bot 顯示設定</div>
<div style="font-size:13px;color:#8a8a9a;margin-bottom:6px">名稱</div>
<div style="display:flex;gap:8px;margin-bottom:8px">
<input id="senderNameInput" type="text" placeholder="翻譯小助手" style="flex:1;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
<button class="btn btn-primary btn-sm" onclick="saveSenderSettings()">儲存</button>
</div>
<div style="font-size:13px;color:#8a8a9a;margin-bottom:6px">圖示 URL（選填）</div>
<input id="senderIconInput" type="text" placeholder="https://example.com/icon.png" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📢 推送訊息</div>
<div class="ch-select-wrap" style="margin-bottom:8px">
<select class="ch-select" id="pushGroupSelect" style="font-size:13px;padding:10px"></select>
</div>
<textarea id="pushText" rows="3" placeholder="輸入要推送的訊息..." style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical;margin-bottom:8px"></textarea>
<div style="display:flex;gap:8px">
<button class="btn btn-primary btn-sm" onclick="pushMessage()">📤 推送到群組</button>
<button class="btn btn-dark btn-sm" onclick="broadcastMessage()">📣 推送全體好友</button>
</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📋 Rich Menu 選單</div>
<div style="display:flex;gap:8px;margin-bottom:8px">
<button class="btn btn-primary btn-sm" onclick="createRichMenu()">建立選單</button>
<button class="btn btn-red btn-sm" onclick="deleteRichMenu()">刪除選單</button>
</div>
<div id="richMenuListOld" style="font-size:12px;color:#8a8a9a;margin-top:6px">LINE 底部常駐按鈕</div>
</div>

</div>

</div><!-- mainPage -->
</div><!-- app -->

<div class="toast" id="toast"></div>

<script>
window.onerror=function(msg,url,line,col,err){
  document.body.innerHTML='<div style="color:red;font:16px monospace;padding:20px;white-space:pre-wrap">JS ERROR:\\n'+msg+'\\nLine: '+line+'\\nCol: '+col+'</div>';
  return false;
};
var KEY=window._ADMIN_KEY||'';
var API=window.location.origin+'/api/admin';

function toast(msg){var t=document.getElementById('toast');if(!t)return;t.textContent=msg;t.classList.add('show');setTimeout(function(){t.classList.remove('show')},2000)}

function api(path,method,body){
  method=method||'GET';
  var h={'Content-Type':'application/json'};
  if(KEY) h['X-Admin-Key']=KEY;
  if(window._MANAGER_ID) h['X-Manager-Id']=window._MANAGER_ID;
  var opts={method:method,headers:h};
  if(body)opts.body=JSON.stringify(body);
  return fetch(API+path,opts).then(function(r){
    if(r.status===403){toast('權限不足');return null}
    var ct=r.headers.get('content-type')||'';
    if(!r.ok||ct.indexOf('application/json')<0){toast('伺服器錯誤('+r.status+')');return null}
    return r.json();
  }).catch(function(e){toast('連線失敗: '+e.message);return null});
}

function doLogin(){
  KEY=document.getElementById('pwInput').value.trim();
  if(!KEY){toast('請輸入密碼');return}
  api('/status').then(function(d){
    if(!d)return;
    document.getElementById('loginPage').style.display='none';
    document.getElementById('mainPage').style.display='block';
    localStorage.setItem('bot_admin_key',KEY);
    if(typeof loadAll==='function') loadAll();
  });
}
</script>
<script>
var FEAT_KEYS=['translation_on','image_on','voice_on','work_order_on'];

var TAB_KEYS=['overview','groups','skip','users','names','storage','packaging','passwords','scrap','insight','examples','forms','settings'];
function switchTab(name){
  document.querySelectorAll('.tab').forEach(function(t,i){
    t.classList.toggle('active',TAB_KEYS[i]===name);
  });
  document.querySelectorAll('.panel').forEach(function(p){p.classList.remove('active')});
  document.getElementById('panel-'+name).classList.add('active');
  if(name==='overview') loadStats();
  if(name==='groups'){loadGroups();loadDM();}
  if(name==='skip') loadGroupSelect();
  if(name==='users'){loadUsersGroupSelect();loadUsers();}
  if(name==='names') loadNames();
  if(name==='storage') loadStorageStats();
  if(name==='packaging') loadPackagingStats();
  if(name==='passwords') loadPasswords();
  if(name==='scrap') loadScrap();
  if(name==='insight') loadInsightTab();
  if(name==='examples'){loadExamples();loadTranslationLog();}
  if(name==='forms') loadFormsTab();
  if(name==='settings') loadFeatureSettings();
}

function loadAll(){loadStats();loadGroups();loadDM();loadGroupSelect();loadUsersGroupSelect();loadUsers();loadNames();loadStorageStats()}

async function loadStats(){
  var d=await api('/stats');
  if(!d)return;
  var h=Math.floor((d.uptime_seconds||0)/3600);
  var m=Math.floor(((d.uptime_seconds||0)%3600)/60);
  document.getElementById('st-uptime').textContent=h+'h '+m+'m';
  setStatVal('st-text',d.text_translations||0);
  setStatVal('st-image',d.image_translations||0);
  setStatVal('st-voice',d.voice_translations||0);
  setStatVal('st-wo',d.work_order_detections||0);
  setStatVal('st-cmd',d.commands||0);
  setStatVal('st-cust',d.customers||0);
  setStatVal('st-groups',d.groups||0);
  setStatVal('st-dm-users',d.dm_users||0);
  var tt=d.tokens_total||0;
  document.getElementById('st-tokens').textContent=tt.toLocaleString();
  document.getElementById('st-cost').textContent='$'+(d.estimated_cost_usd||0).toFixed(4);
}
function setStatVal(id,val){
  var el=document.getElementById(id);
  el.textContent=val;
  if(val>0)el.classList.add('highlight');
  else el.classList.remove('highlight');
}

var _groupList=[];
async function loadGroups(){
  var d=await api('/groups');
  if(!d)return;
  _groupList=d.groups||[];
  var el=document.getElementById('groupList');
  if(!_groupList.length){el.innerHTML='<div class="empty">尚無群組紀錄<br>Bot 收到群組訊息後會自動記錄</div>';return}
  var html='';
  for(var i=0;i<_groupList.length;i++){
    var g=_groupList[i];
    var skipCt=g.skip_count||0;
    var memberCt=g.member_count?g.member_count+'人':'--';
    html+='<div class="card">'+
      '<div class="card-title"><div><span style="font-weight:700;font-size:16px">#'+(g.name||'(未知群組)')+'</span><span style="font-size:12px;color:#8a8a9a;margin-left:8px">👥'+memberCt+'</span></div>'+
      '<span class="badge '+(g.translation_on?'badge-on':'badge-off')+'" style="cursor:pointer" onclick="toggleFeat('+i+',0)">'+(g.translation_on?'翻譯開':'翻譯關')+'</span></div>'+
      '<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin:8px 0">'+
      '<span class="card-sub">中文 ⇄ 🇮🇩 印尼文</span>'+
      '<span class="card-sub">｜跳過: '+skipCt+'人</span></div>'+
      '<div class="feat-badges">'+
      // v3.9.10: 圖片按鈕變 3 段循環:開 → 詢問 → 關
      // 顯示文字 / 樣式根據 (image_on, image_ask_mode) 兩個欄位決定
      '<span class="feat-badge '+(g.image_on?'on':(g.image_ask_mode?'ask':'off'))+'" style="cursor:pointer" onclick="cycleImageMode('+i+')" title="點按循環:開→詢問→關">'+
        '🖼️ '+(g.image_on?'圖片自動翻':(g.image_ask_mode?'圖片詢問':'圖片關'))+'</span>'+
      '<span class="feat-badge '+(g.voice_on?'on':'off')+'" style="cursor:pointer" onclick="toggleFeat('+i+',2)">🎤 '+(g.voice_on?'語音開':'語音關')+'</span>'+
      '<span class="feat-badge '+(g.work_order_on?'on':'off')+'" style="cursor:pointer" onclick="toggleFeat('+i+',3)">📋 '+(g.work_order_on?'工單開':'工單關')+'</span></div>'+
      buildCmdBadges(g, i)+
      '<div style="display:flex;align-items:center;justify-content:space-between;margin:10px 0;padding:10px 12px;background:rgba(124,111,239,.08);border-radius:8px;border:1px solid rgba(124,111,239,.2)">'+
      '<div><span style="font-size:12px;color:#8a8a9a">累計花費</span><br><span style="font-size:18px;font-weight:700;color:#7c6fef">NT$'+(g.cost_twd||0).toFixed(1)+'</span></div>'+
      '<button class="btn btn-dark btn-sm" style="font-size:12px" onclick="resetCost('+i+')">歸零</button></div>'+
      '<button class="btn btn-red btn-sm" onclick="leaveGroup('+i+')">退出群組: '+(g.name||g.id.substring(0,12))+'</button></div>';
  }
  el.innerHTML=html;
}

function toggleFeat(idx,keyIdx){
  var g=_groupList[idx];if(!g)return;
  var key=FEAT_KEYS[keyIdx];
  var cur=g[key];
  var body={group_id:g.id};body[key]=!cur;
  api('/groups/settings','POST',body).then(function(d){if(d){toast('已更新');loadGroups()}});
}

// v3.9.10: 圖片模式 3 段循環(開 → 詢問 → 關 → 開 ...)
function cycleImageMode(idx){
  var g=_groupList[idx];if(!g)return;
  var body={group_id:g.id};
  if(g.image_on){
    // 目前=自動翻 → 切到詢問模式
    body.image_ask_mode=true;
  }else if(g.image_ask_mode){
    // 目前=詢問 → 切到完全關閉
    body.image_on=false;
    body.image_ask_mode=false;
  }else{
    // 目前=關閉 → 切到自動翻
    body.image_on=true;
  }
  api('/groups/settings','POST',body).then(function(d){if(d){toast('已更新');loadGroups()}});
}
var CMD_DEFS=[["pw1","🔑密碼1"],["pw2","🏭密碼2"],["pkg","📦包裝"],["scrap","🎨廢料"],["qry","🔍儲區"],["notice","📢公告"]];
function buildCmdBadges(g,idx){
  var ce=g.cmd_enabled||{};
  var h='<div class="feat-badges" style="margin-top:4px">';
  for(var c=0;c<CMD_DEFS.length;c++){
    var key=CMD_DEFS[c][0],label=CMD_DEFS[c][1];
    var on=ce[key]!==false;
    h+='<span class="feat-badge '+(on?'on':'off')+'" style="cursor:pointer" onclick="toggleCmd('+idx+',&apos;'+key+'&apos;,'+(!on)+')">'+label+(on?'開':'關')+'</span>';
  }
  return h+'</div>';
}
function toggleCmd(idx,key,val){
  var g=_groupList[idx];if(!g)return;
  api('/groups/settings','POST',{group_id:g.id,cmd_toggle:key,cmd_val:val}).then(function(d){if(d){toast('已更新');loadGroups()}});
}
function leaveGroup(idx){
  var g=_groupList[idx];if(!g)return;
  if(!confirm('確定退出「'+(g.name||g.id)+'」？'))return;
  api('/groups/leave','POST',{group_id:g.id}).then(function(d){if(d){toast(d.message||'已退出');loadGroups();loadGroupSelect()}});
}
function resetCost(idx){
  var g=_groupList[idx];if(!g)return;
  if(!confirm('確定歸零累計花費？'))return;
  api('/groups/reset-cost','POST',{group_id:g.id}).then(function(d){if(d){toast('已歸零');loadGroups()}});
}

var _dmUsers=[];
async function loadDM(){
  var d=await api('/dm');
  if(!d)return;
  document.getElementById('dmToggle').checked=d.master_enabled;
  _dmUsers=d.known_users||[];
  var el=document.getElementById('dmWlList');
  if(!_dmUsers.length){el.innerHTML='<div style="padding:8px 0;font-size:13px;color:#5a5a6a">尚無人私訊過 Bot</div>';return}
  var html='';
  for(var i=0;i<_dmUsers.length;i++){
    var u=_dmUsers[i];
    html+='<div class="wl-item" style="border-color:#2a2a3e"><span>'+esc(u.name||'')+'</span>'+
    '<label class="toggle"><input type="checkbox" '+(u.whitelisted?'checked':'')+
    ' onchange="toggleDmWl('+i+',this.checked)"><span class="slider"></span></label></div>';
  }
  el.innerHTML=html;
}
async function toggleDM(){
  var on=document.getElementById('dmToggle').checked;
  var d=await api('/dm','POST',{master_enabled:on});
  if(d) toast(on?'DM 已開啟':'DM 已關閉');
}
function toggleDmWl(idx,on){
  var u=_dmUsers[idx];if(!u)return;
  api('/dm/whitelist','POST',{user_id:u.user_id,action:on?'add':'remove'}).then(function(d){
    if(d) toast(on?'已加入白名單':'已移出白名單');
  });
}

async function loadGroupSelect(){
  var d=await api('/groups');
  if(!d)return;
  var sel=document.getElementById('skipGroupSelect');
  var cur=sel.value;
  sel.innerHTML='<option value="">選擇群組...</option>';
  var groups=d.groups||[];
  for(var i=0;i<groups.length;i++){
    var g=groups[i];
    var opt=document.createElement('option');
    opt.value=g.id;opt.textContent='#'+(g.name||g.id.substring(0,16));
    sel.appendChild(opt);
  }
  if(cur)sel.value=cur;
}

var _skipUsers=[];
async function loadSkipList(){
  var gid=document.getElementById('skipGroupSelect').value;
  var el=document.getElementById('skipListContent');
  if(!gid){el.innerHTML='<div class="empty">請先選擇群組</div>';return}
  var d=await api('/skip?group_id='+gid);
  if(!d)return;
  _skipUsers=d.users||[];
  if(!_skipUsers.length){
    el.innerHTML='<div class="empty">尚無成員紀錄<br>成員在群組發訊息後會自動出現</div>';
    return;
  }
  var html='';
  for(var i=0;i<_skipUsers.length;i++){
    var u=_skipUsers[i];
    html+='<div class="wl-item"><span style="font-size:15px">'+esc(u.name||'')+'</span>'+
    '<label class="toggle"><input type="checkbox" '+(u.skipped?'checked':'')+
    ' onchange="toggleSkip('+i+',this.checked)"><span class="slider"></span></label></div>';
  }
  el.innerHTML=html;
}
function toggleSkip(idx,on){
  var gid=document.getElementById('skipGroupSelect').value;
  var u=_skipUsers[idx];if(!u)return;
  api('/skip','POST',{group_id:gid,user_id:u.user_id,action:on?'add':'remove'}).then(function(d){
    if(d) toast(on?'已跳過翻譯':'已恢復翻譯');
  });
}

async function loadUsersGroupSelect(){
  var d=await api('/groups');
  if(!d)return;
  var sel=document.getElementById('usersGroupSelect');
  var cur=sel.value;
  sel.innerHTML='<option value="">全部使用者</option>';
  var groups=d.groups||[];
  for(var i=0;i<groups.length;i++){
    var g=groups[i];
    var opt=document.createElement('option');
    opt.value=g.id;opt.textContent='#'+(g.name||g.id.substring(0,16));
    sel.appendChild(opt);
  }
  if(cur)sel.value=cur;
}

var _allUsers=[];
async function loadUsers(){
  var gid=document.getElementById('usersGroupSelect').value;
  var path=gid?'/users?group_id='+gid:'/users';
  var d=await api(path);
  if(!d)return;
  _allUsers=d.users||[];
  var el=document.getElementById('usersList');
  if(!_allUsers.length){el.innerHTML='<div class="empty">尚無使用者紀錄<br>使用者互動後會自動出現</div>';return}
  var html='';
  var TAB_OPTS=[['overview','總覽'],['groups','群組'],['skip','白名單'],['users','使用者'],['names','保護名單'],['storage','儲區'],['packaging','包裝碼'],['passwords','密碼'],['scrap','廢料色'],['insight','數據'],['examples','翻譯範例'],['forms','表單'],['settings','設定']];
  for(var i=0;i<_allUsers.length;i++){
    var u=_allUsers[i];
    var langBadge=u.line_lang?'<span class="badge badge-on" style="font-size:11px">'+u.line_lang+'</span>':'';
    var atabs=u.allowed_tabs||[];
    html+='<div class="user-card">'+
      '<div style="display:flex;justify-content:space-between;align-items:flex-start">'+
      '<div><div class="user-name">'+esc(u.name||'')+'</div><div class="user-id">ID: '+u.user_id+'</div></div>'+langBadge+'</div>'+
      '<div class="user-admin-row">'+
      '<span class="admin-label">🔑 管理員</span>'+
      '<label class="toggle"><input type="checkbox" '+(u.is_admin?'checked':'')+
      ' onchange="toggleAdmin('+i+',this.checked)"><span class="slider"></span></label>'+
      '</div>';
    if(u.is_admin){
      html+='<div style="margin-top:6px;padding:8px;background:#0d0d1a;border-radius:6px;border:1px solid #2a2a3e">';
      html+='<div style="font-size:11px;color:#8a8a9a;margin-bottom:4px">📧 Google 信箱（登入用）：</div>';
      html+='<div style="display:flex;gap:4px;margin-bottom:8px"><input type="email" value="'+esc(u.google_email||'')+'" placeholder="example@gmail.com" data-uid="'+u.user_id+'" style="flex:1;padding:5px 8px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px" onchange="saveUserEmail(this)"><span style="font-size:10px;color:#43b581;align-self:center">'+(u.google_email?'✓':'')+'</span></div>';
      html+='<div style="font-size:11px;color:#8a8a9a;margin-bottom:4px">可用功能：</div>';
      html+='<div style="display:flex;flex-wrap:wrap;gap:4px">';
      for(var t=0;t<TAB_OPTS.length;t++){
        var tk=TAB_OPTS[t][0],tl=TAB_OPTS[t][1];
        var ck=atabs.indexOf(tk)>=0;
        html+='<label style="font-size:11px;color:#ccc;display:flex;align-items:center;gap:2px;padding:2px 4px;background:'+(ck?'#1a2a3a':'#1a1a2e')+';border-radius:4px;border:1px solid '+(ck?'#3a5a7a':'#2a2a3e')+'"><input type="checkbox" data-uid="'+u.user_id+'" data-tab="'+tk+'" '+(ck?'checked':'')+' onchange="toggleUserTab(this)" style="width:12px;height:12px">'+tl+'</label>';
      }
      html+='</div></div>';
    }
    html+='</div>';
  }
  el.innerHTML=html;
}
function toggleAdmin(idx,on){
  var u=_allUsers[idx];if(!u)return;
  api('/users/admin','POST',{user_id:u.user_id,is_admin:on}).then(function(d){
    if(d){toast(on?'已設為管理員':'已取消管理員');u.is_admin=on;renderUsersList()}
  });
}
function saveUserEmail(el){
  var uid=el.getAttribute('data-uid');
  var email=el.value.trim();
  api('/users/email','POST',{user_id:uid,google_email:email}).then(function(d){
    if(d)toast(email?'已綁定: '+email:'已清除信箱');
  });
}
function toggleUserTab(el){
  var uid=el.getAttribute('data-uid');
  var tab=el.getAttribute('data-tab');
  var u=null;
  for(var i=0;i<_allUsers.length;i++){if(_allUsers[i].user_id===uid){u=_allUsers[i];break}}
  if(!u)return;
  var tabs=u.allowed_tabs||[];
  if(el.checked){if(tabs.indexOf(tab)<0)tabs.push(tab)}
  else{tabs=tabs.filter(function(t){return t!==tab})}
  u.allowed_tabs=tabs;
  api('/users/tabs','POST',{user_id:uid,allowed_tabs:tabs}).then(function(d){
    if(d)toast('已更新');
  });
}
function renderUsersList(){
  var el=document.getElementById('usersListContainer');
  if(!el)return;
  loadUsers();
}

var _protectedNames=[];
async function loadNames(){
  var d=await api('/names');
  if(!d)return;
  _protectedNames=d.names||[];
  var el=document.getElementById('namesList');
  document.getElementById('namesCount').textContent='共 '+_protectedNames.length+' 個保護名稱';
  if(!_protectedNames.length){el.innerHTML='<div style="padding:8px 0;font-size:13px;color:#5a5a6a">尚無保護名稱</div>';return}
  var html='<div style="display:flex;flex-wrap:wrap;gap:8px">';
  for(var i=0;i<_protectedNames.length;i++){
    html+='<span style="display:inline-flex;align-items:center;gap:6px;padding:6px 12px;background:#2a2a3e;border:1px solid #3a3a4e;border-radius:8px;font-size:13px">'+
    _protectedNames[i]+'<span style="cursor:pointer;color:#f04747;font-weight:700;font-size:15px" onclick="removeName('+i+')"> ×</span></span>';
  }
  html+='</div>';
  el.innerHTML=html;
}
async function addName(){
  var inp=document.getElementById('newNameInput');
  var name=inp.value.trim();
  if(!name){toast('請輸入名字');return}
  var d=await api('/names','POST',{action:'add',name:name});
  if(d){toast('已新增: '+name);inp.value='';loadNames()}
}
function removeName(idx){
  var name=_protectedNames[idx];
  if(!name)return;
  if(!confirm('確定移除「'+name+'」？'))return;
  api('/names','POST',{action:'remove',name:name}).then(function(d){if(d){toast('已移除: '+name);loadNames()}});
}

async function loadStorageStats(){
  var d=await api('/storage/stats');
  if(!d)return;
  document.getElementById('storageStats').innerHTML='客戶數: <strong style="color:#7c6fef">'+d.count+'</strong>';
}

var storageFileData=null;
function previewStorage(){
  var f=document.getElementById('storageFile').files[0];
  if(!f)return;
  document.getElementById('storageFileName').textContent='📄 '+f.name;
  storageFileData=f;
  document.getElementById('storageActions').style.display='block';
  document.getElementById('storagePreview').innerHTML='<div class="card"><div class="card-sub">點「確認更新」上傳並解析</div></div>';
}

async function uploadStorage(){
  if(!storageFileData){toast('請先選擇檔案');return}
  var fd=new FormData();
  fd.append('file',storageFileData);
  try{
    var r=await fetch(API+'/storage/upload',{method:'POST',headers:{'X-Admin-Key':KEY,'X-Manager-Id':window._MANAGER_ID||''},body:fd});
    if(!r.ok){toast('上傳失敗('+r.status+')');return}
    var d=await r.json();
    if(d.error){toast(d.error);return}
    toast(d.message||'更新成功');
    var ghStatus=d.github?'✅ 已推送 GitHub，Render 將自動部署':'⚠️ GitHub 推送失敗，僅暫時生效';
    document.getElementById('storageActions').style.display='none';
    document.getElementById('storagePreview').innerHTML='<div class="card"><div style="color:#43b581;font-weight:600">✅ 已更新 '+d.count+' 筆客戶資料</div><div class="card-sub" style="margin-top:4px">'+ghStatus+'</div></div>';
    loadStorageStats();
  }catch(e){toast('上傳失敗: '+e)}
}

async function downloadJson(){
  try{
    var r=await fetch(API+'/storage/json',{headers:{'X-Admin-Key':KEY,'X-Manager-Id':window._MANAGER_ID||''}});
    var blob=await r.blob();
    var url=URL.createObjectURL(blob);
    var a=document.createElement('a');
    a.href=url;a.download='storage_data.json';a.click();
    URL.revokeObjectURL(url);
    toast('JSON 已下載');
  }catch(e){toast('下載失敗')}
}

// ─── Packaging ───
async function loadPackagingStats(){
  var d=await api('/packaging/stats');
  if(d) document.getElementById('packagingStats').textContent='共 '+d.count+' 筆包裝碼';
}
function previewPackaging(){
  var f=document.getElementById('packagingFile').files[0];
  if(!f)return;
  document.getElementById('packagingFileName').textContent='📄 '+f.name;
  document.getElementById('packagingActions').style.display='block';
}
async function uploadPackaging(){
  var f=document.getElementById('packagingFile').files[0];
  if(!f){toast('請選擇檔案');return;}
  var fd=new FormData();fd.append('file',f);
  try{
    var r=await fetch(API+'/packaging/upload',{method:'POST',headers:{'X-Admin-Key':KEY,'X-Manager-Id':window._MANAGER_ID||''},body:fd});
    if(!r.ok){toast('上傳失敗('+r.status+')');return}
    var d=await r.json();
    if(d.ok){toast(d.message);loadPackagingStats();document.getElementById('packagingActions').style.display='none';}
    else{toast(d.error||'上傳失敗');}
  }catch(e){toast('上傳失敗: '+e);}
}
async function downloadPackagingJson(){
  try{
    var r=await fetch(API+'/packaging/json',{headers:{'X-Admin-Key':KEY,'X-Manager-Id':window._MANAGER_ID||''}});
    var blob=await r.blob();
    var url=URL.createObjectURL(blob);
    var a=document.createElement('a');
    a.href=url;a.download='packaging_data.json';a.click();
    URL.revokeObjectURL(url);
    toast('JSON 已下載');
  }catch(e){toast('下載失敗')}
}

// ─── Passwords ───
async function loadPasswords(){
  var d=await api('/passwords');
  if(!d)return;
  document.getElementById('pw1Input').value=d.pw1||'';
  document.getElementById('pw2Input').value=d.pw2||'';
}
async function savePasswords(){
  var pw1=document.getElementById('pw1Input').value;
  var pw2=document.getElementById('pw2Input').value;
  var d=await api('/passwords','POST',{pw1:pw1,pw2:pw2});
  document.getElementById('pwSaveResult').textContent=d&&d.ok?'✅ 已儲存':'❌ 儲存失敗';
  if(d&&d.ok)toast('密碼已更新');
}

// ─── Scrap ───
async function loadScrap(){
  var d=await api('/scrap');
  if(!d)return;
  document.getElementById('scrapInput').value=d.text||'';
}
async function saveScrap(){
  var text=document.getElementById('scrapInput').value;
  var d=await api('/scrap','POST',{text:text});
  document.getElementById('scrapSaveResult').textContent=d&&d.ok?'✅ 已儲存':'❌ 儲存失敗';
  if(d&&d.ok)toast('廢料色資訊已更新');
}

// ─── Insight Tab ───
async function loadInsightTab(){
  // Load trend
  var t=await api('/insight/trend?days=7');
  if(t&&t.trend&&t.trend.length){
    var trend=t.trend.reverse();
    var labels=[];var followers=[];var blocks=[];
    for(var i=0;i<trend.length;i++){
      var d=trend[i].date;
      labels.push(d.substring(4,6)+'/'+d.substring(6,8));
      followers.push(trend[i].followers||0);
      blocks.push(trend[i].blocks||0);
    }
    drawTrendChart(labels,followers,blocks);
    var html='<table style="width:100%;font-size:12px;border-collapse:collapse">';
    html+='<tr style="color:#8a8a9a"><td>日期</td><td style="text-align:right">好友數</td><td style="text-align:right">封鎖</td></tr>';
    for(var i=0;i<trend.length;i++){
      html+='<tr><td>'+labels[i]+'</td><td style="text-align:right">'+(trend[i].followers||'-')+'</td><td style="text-align:right;color:#f04747">'+(trend[i].blocks||'-')+'</td></tr>';
    }
    html+='</table>';
    document.getElementById('insightTrendData').innerHTML=html;
  }else{
    document.getElementById('insightTrendData').textContent='無趨勢資料（需至少20名好友且帳號開通超過7天）';
  }
  // Load demographics
  var ins=await api('/insight');
  if(ins){
    var dhtml='';
    if(ins.demographics&&ins.demographics.available){
      if(ins.demographics.genders){
        dhtml+='<div style="font-weight:600;margin-bottom:4px">性別</div>';
        var g=ins.demographics.genders;
        if(Array.isArray(g)){
          for(var i=0;i<g.length;i++){
            var pct=g[i].percentage?Math.round(g[i].percentage*100)+'%':'';
            dhtml+='<span class="badge badge-on" style="font-size:11px;margin:2px">'+(g[i].gender||'-')+' '+pct+'</span> ';
          }
        }else{dhtml+=JSON.stringify(g)}
        dhtml+='<br>';
      }
      if(ins.demographics.ages){
        dhtml+='<div style="font-weight:600;margin:8px 0 4px">年齡</div>';
        var a=ins.demographics.ages;
        if(Array.isArray(a)){
          for(var i=0;i<a.length;i++){
            var pct=a[i].percentage?Math.round(a[i].percentage*100)+'%':'';
            dhtml+='<span class="badge badge-on" style="font-size:11px;margin:2px">'+(a[i].age||'-')+' '+pct+'</span> ';
          }
        }else{dhtml+=JSON.stringify(a)}
        dhtml+='<br>';
      }
      if(ins.demographics.areas){
        dhtml+='<div style="font-weight:600;margin:8px 0 4px">地區</div>';
        var ar=ins.demographics.areas;
        if(Array.isArray(ar)){
          for(var i=0;i<Math.min(ar.length,10);i++){
            var pct=ar[i].percentage?Math.round(ar[i].percentage*100)+'%':'';
            dhtml+='<span class="badge badge-on" style="font-size:11px;margin:2px">'+(ar[i].area||'-')+' '+pct+'</span> ';
          }
        }else{dhtml+=JSON.stringify(ar)}
      }
    }else{dhtml='人口統計需至少20名好友'}
    document.getElementById('insightDemoData').innerHTML=dhtml||'無資料';
    // Delivery
    if(ins.delivery){
      var s=ins.delivery;
      var dv='日期: '+(s.date||'-');
      if(s.reply!==null)dv+='<br>Reply: '+s.reply;
      if(s.push!==null)dv+=' ｜ Push: '+s.push;
      if(s.multicast!==null)dv+=' ｜ Multicast: '+s.multicast;
      if(s.broadcast!==null)dv+='<br>Broadcast: '+s.broadcast;
      document.getElementById('insightDelivery').innerHTML=dv;
    }
  }
}
function drawTrendChart(labels,followers,blocks){
  var canvas=document.getElementById('trendCanvas');
  if(!canvas)return;
  var ctx=canvas.getContext('2d');
  var W=canvas.width;var H=canvas.height;
  ctx.clearRect(0,0,W,H);
  var pad={t:20,r:10,b:30,l:50};
  var cw=W-pad.l-pad.r;var ch=H-pad.t-pad.b;
  var maxF=Math.max.apply(null,followers)||1;
  // Grid
  ctx.strokeStyle='#2a2a3e';ctx.lineWidth=1;
  for(var i=0;i<4;i++){
    var y=pad.t+ch*(i/3);
    ctx.beginPath();ctx.moveTo(pad.l,y);ctx.lineTo(W-pad.r,y);ctx.stroke();
  }
  // Labels
  ctx.fillStyle='#8a8a9a';ctx.font='11px sans-serif';ctx.textAlign='center';
  for(var i=0;i<labels.length;i++){
    var x=pad.l+cw*(i/(labels.length-1||1));
    ctx.fillText(labels[i],x,H-8);
  }
  ctx.textAlign='right';
  for(var i=0;i<4;i++){
    var y=pad.t+ch*(i/3);
    var val=Math.round(maxF*(1-i/3));
    ctx.fillText(val,pad.l-6,y+4);
  }
  // Follower line
  ctx.strokeStyle='#7c6fef';ctx.lineWidth=2;ctx.beginPath();
  for(var i=0;i<followers.length;i++){
    var x=pad.l+cw*(i/(followers.length-1||1));
    var y=pad.t+ch*(1-followers[i]/maxF);
    if(i===0)ctx.moveTo(x,y);else ctx.lineTo(x,y);
  }
  ctx.stroke();
  // Dots
  ctx.fillStyle='#7c6fef';
  for(var i=0;i<followers.length;i++){
    var x=pad.l+cw*(i/(followers.length-1||1));
    var y=pad.t+ch*(1-followers[i]/maxF);
    ctx.beginPath();ctx.arc(x,y,3,0,Math.PI*2);ctx.fill();
  }
}

// ─── Rich Menu Image Upload ───
var _rmImageData=null;
function previewRmImage(input){
  if(!input.files||!input.files[0])return;
  var file=input.files[0];
  var reader=new FileReader();
  reader.onload=function(e){
    _rmImageData=e.target.result;
    document.getElementById('rmImagePreviewImg').src=_rmImageData;
    document.getElementById('rmImagePreview').style.display='block';
    document.getElementById('rmUploadBtn').style.display='inline-block';
  };
  reader.readAsDataURL(file);
}
async function uploadRmImage(){
  var sel=document.getElementById('rmUploadSelect');
  var rmId=sel.value;
  if(!rmId){toast('請選擇 Rich Menu');return}
  if(!_rmImageData){toast('請選擇圖片');return}
  document.getElementById('rmUploadResult').textContent='上傳中...';
  var d=await api('/richmenu/upload/'+rmId,'POST',{image:_rmImageData});
  if(d&&d.ok){
    document.getElementById('rmUploadResult').innerHTML='<span style="color:#43b581">✅ 上傳成功</span>';
    toast('圖片已上傳');
    _rmImageData=null;
    document.getElementById('rmImagePreview').style.display='none';
    document.getElementById('rmUploadBtn').style.display='none';
  }else{
    document.getElementById('rmUploadResult').innerHTML='<span style="color:#f04747">❌ 上傳失敗: '+(d&&d.error||'unknown')+'</span>';
  }
}

// ─── Imagemap Send ───
var _imapActionCount=0;
function addImapAction(){
  var div=document.getElementById('imapActions');
  var idx=_imapActionCount++;
  var html='<div id="imapAct'+idx+'" style="background:#0d0d1a;border:1px solid #3a3a4e;border-radius:8px;padding:8px;margin-bottom:6px;font-size:12px">';
  html+='<div style="display:flex;gap:6px;margin-bottom:4px;align-items:center">';
  html+='<select id="imapType'+idx+'" style="padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px"><option value="message">文字訊息</option><option value="uri">開啟網址</option></select>';
  html+='<input id="imapText'+idx+'" placeholder="訊息文字 或 URL" style="flex:1;padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='<span style="color:#f04747;cursor:pointer" onclick="document.getElementById(&apos;imapAct'+idx+'&apos;).remove()">✕</span>';
  html+='</div>';
  html+='<div style="display:flex;gap:4px">';
  html+='<input id="imapX'+idx+'" type="number" placeholder="X" value="0" style="width:60px;padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='<input id="imapY'+idx+'" type="number" placeholder="Y" value="0" style="width:60px;padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='<input id="imapAW'+idx+'" type="number" placeholder="W" value="1040" style="width:60px;padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='<input id="imapAH'+idx+'" type="number" placeholder="H" value="1040" style="width:60px;padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='</div></div>';
  div.insertAdjacentHTML('beforeend',html);
}
async function sendImap(){
  var sel=document.getElementById('imapGroupSelect');
  var to=sel.value;
  if(!to){toast('請選擇群組');return}
  var baseUrl=document.getElementById('imapBaseUrl').value.trim();
  if(!baseUrl){toast('請輸入圖片 Base URL');return}
  var w=parseInt(document.getElementById('imapW').value)||1040;
  var h=parseInt(document.getElementById('imapH').value)||1040;
  var actions=[];
  for(var i=0;i<_imapActionCount;i++){
    var el=document.getElementById('imapAct'+i);
    if(!el)continue;
    var type=document.getElementById('imapType'+i).value;
    var text=document.getElementById('imapText'+i).value.trim();
    var x=parseInt(document.getElementById('imapX'+i).value)||0;
    var y=parseInt(document.getElementById('imapY'+i).value)||0;
    var aw=parseInt(document.getElementById('imapAW'+i).value)||1040;
    var ah=parseInt(document.getElementById('imapAH'+i).value)||1040;
    if(!text)continue;
    var act={type:type,x:x,y:y,w:aw,h:ah};
    if(type==='uri')act.uri=text;else act.text=text;
    actions.push(act);
  }
  if(!actions.length){toast('請新增至少一個點擊區域');return}
  if(!confirm('確定發送 Imagemap 到此群組？'))return;
  var d=await api('/imagemap/send','POST',{to:to,base_url:baseUrl,width:w,height:h,actions:actions});
  if(d&&d.ok){
    document.getElementById('imapResult').innerHTML='<span style="color:#43b581">✅ 已發送</span>';
    toast('Imagemap 已發送');
  }else{
    document.getElementById('imapResult').innerHTML='<span style="color:#f04747">❌ 發送失敗: '+(d&&d.error||'unknown')+'</span>';
  }
}

// ─── Translation Examples ───
var _exData=[];
var _builtinExamples=null;
function _getBuiltinExamples(){
  if(_builtinExamples)return _builtinExamples;
  _builtinExamples=[];
  var raw=[
    ["乾 需不需要提報一下","Aduh, perlu dilaporkan gak nih?"],
    ["UT囤一堆料了","UT udah numpuk banyak material."],
    ["品保還在下班 誇張","QC udah pulang, keterlaluan."],
    ["三米上面放六米","Batang 3 meter ditaruh di atas batang 6 meter."],
    ["來料都短少4-5公斤","Material masuk semuanya kurang 4-5 kilogram."],
    ["已轉達","Sudah disampaikan."],
    ["這批料有問題","Lot material ini ada masalah."],
    ["幫我盯一下","Tolong awasin ya."],
    ["怎麼搞的啦","Kok bisa kayak gini sih."],
    ["人咧","Orangnya mana?"],
    ["辛苦了","Makasih kerja kerasnya."],
    ["靠 又壞了","Astaga, rusak lagi."],
    ["先這樣","Segitu dulu ya."],
    ["砂輪要換了","Batu gerinda harus diganti."],
    ["公差超過了","Toleransinya udah lewat."],
    ["粗拋完已放行","Rough polishing selesai, sudah di-release."],
    ["放了","Sudah di-release."],
    ["放地上","Taruh di lantai."],
    ["料放旁邊","Material taruh di samping."],
    ["幫追料","Tolong kejar materialnya"],
    ["幫追帳","Tolong kejar data administrasinya"],
    ["處長走了","Kepala divisi sudah pergi."],
    ["Saya mau izin besok","我明天要請假"],
    ["Mesinnya rusak","機台壞了"],
    ["Materialnya udah habis","料用完了"],
    ["Saya gak ngerti","我聽不懂"],
    ["Boleh pulang duluan?","可以先下班嗎？"],
    ["Lembur sampai jam berapa?","加班到幾點？"],
    ["Bos, ini udah selesai","老闆，這個好了"],
    ["Ukurannya gak pas","尺寸不對"],
    ["Stoknya masih ada?","庫存還有嗎？"],
    ["Tolong ajarin saya","請教我一下"],
    ["Sudah di-release","放了"],
    ["Tolong bantu release","幫放一下"],
    ["Tolong kejar materialnya","幫追料"],
    ["Kepala divisi sudah pergi","處長走了"],
    ["Besok libur gak?","明天放假嗎？"],
    ["Maaf bos, saya telat","老闆抱歉，我遲到了"],
    ["Mesin udah jalan normal","機台已經恢復正常了"],
    ["Sudah saya laporkan ke QC","我已經跟品保回報了"],
    ["Packing-nya salah, harus bongkar ulang","包裝包錯了，要拆掉重包"],
    ["DILARANG pindahkan batang baja dengan tangan saat mesin jalan","嚴禁運轉中用手搬棒材"],
  ];
  for(var i=0;i<raw.length;i++){
    _builtinExamples.push({zh:raw[i][0].toLowerCase(),id:raw[i][1].toLowerCase()});
  }
  return _builtinExamples;
}
function _checkBuiltinDuplicate(zh,id){
  var bex=_getBuiltinExamples();
  var zl=zh.toLowerCase().trim();
  var il=id.toLowerCase().trim();
  for(var i=0;i<bex.length;i++){
    if(bex[i].zh===zl||bex[i].id===il||bex[i].zh===il||bex[i].id===zl)return true;
  }
  return false;
}
async function loadExamples(){
  var d=await api('/examples');
  if(!d)return;
  _exData=d.examples||[];
  document.getElementById('exCountBadge').textContent=d.count+' / '+d.max;
  if(d.count>=d.max){
    document.getElementById('exWarning').style.display='block';
    document.getElementById('exWarning').textContent='⚠️ 已達上限 '+d.max+' 條，請刪除舊的再新增。';
  }else if(d.count>=d.max*0.8){
    document.getElementById('exWarning').style.display='block';
    document.getElementById('exWarning').textContent='⚠️ 已使用 '+d.count+'/'+d.max+' 條，接近上限。';
  }else{
    document.getElementById('exWarning').style.display='none';
  }
  renderExamples();
  loadTrainStatus();
}
function renderExamples(){
  var filter=document.getElementById('exFilter').value;
  var searchEl=document.getElementById('exSearch');
  var search=searchEl?searchEl.value.trim().toLowerCase():'';
  var el=document.getElementById('exList');
  if(!_exData.length){el.innerHTML='<div class="empty">尚無自訂範例</div>';return}
  var html='';
  var shown=0;
  // Show newest first for usability when list grows large
  for(var i=_exData.length-1;i>=0;i--){
    var ex=_exData[i];
    if(filter!=='all'&&ex.dir!==filter)continue;
    if(search){
      var hay=((ex.zh||'')+' '+(ex.id||'')).toLowerCase();
      if(hay.indexOf(search)<0)continue;
    }
    shown++;
    var dirLabel=ex.dir==='id2zh'?'🇮🇩→🇹🇼':'🇹🇼→🇮🇩';
    html+='<div id="exrow_'+i+'" style="padding:8px 0;border-bottom:1px solid #2a2a3e;font-size:13px">';
    html+='<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:8px">';
    html+='<div style="flex:1;min-width:0">';
    html+='<span class="badge badge-on" style="font-size:10px;margin-right:4px">'+dirLabel+'</span>';
    html+='<span style="color:#e0e0e0">'+esc(ex.zh)+'</span>';
    html+='<br><span style="color:#8a8a9a">→ '+esc(ex.id)+'</span>';
    html+='</div>';
    html+='<div style="display:flex;gap:4px;flex-shrink:0">';
    html+='<span style="color:#7c6fef;cursor:pointer;padding:4px 6px;font-size:13px" onclick="editExample('+i+')" title="編輯">✎</span>';
    html+='<span style="color:#f04747;cursor:pointer;padding:4px 6px;font-size:14px" onclick="deleteExample('+i+')" title="刪除">✕</span>';
    html+='</div>';
    html+='</div></div>';
    if(shown>=300)break; // cap render to avoid choking browser
  }
  if(!shown) html='<div class="empty">無符合篩選的範例</div>';
  if(shown>=300) html+='<div class="empty" style="color:#faa61a">已顯示前 300 筆，請使用篩選/搜尋查看更多</div>';
  el.innerHTML=html;
}
function editExample(idx){
  var ex=_exData[idx];if(!ex)return;
  var row=document.getElementById('exrow_'+idx);if(!row)return;
  var html='<div style="display:flex;flex-direction:column;gap:6px">';
  html+='<select id="edDir_'+idx+'" style="padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='<option value="zh2id"'+(ex.dir==='zh2id'?' selected':'')+'>中文→印尼</option>';
  html+='<option value="id2zh"'+(ex.dir==='id2zh'?' selected':'')+'>印尼→中文</option>';
  html+='</select>';
  html+='<input id="edZh_'+idx+'" type="text" value="'+esc(ex.zh||'')+'" placeholder="中文" style="padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:13px">';
  html+='<input id="edId_'+idx+'" type="text" value="'+esc(ex.id||'')+'" placeholder="印尼文" style="padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:13px">';
  html+='<div style="display:flex;gap:6px"><button class="btn btn-primary btn-sm" onclick="saveEditExample('+idx+')">儲存</button>';
  html+='<button class="btn btn-sm" style="background:#2a2a3e;color:#aaa;border:1px solid #3a3a4e;padding:6px 10px;border-radius:6px;font-size:12px" onclick="renderExamples()">取消</button></div>';
  html+='</div>';
  row.innerHTML=html;
}
async function saveEditExample(idx){
  var dir=document.getElementById('edDir_'+idx).value;
  var zh=document.getElementById('edZh_'+idx).value.trim();
  var id=document.getElementById('edId_'+idx).value.trim();
  if(!zh||!id){toast('請填中文和印尼文');return}
  var d=await api('/examples/edit','POST',{index:idx,zh:zh,id:id,dir:dir});
  if(d&&d.ok){
    toast('✅ 已更新');
    loadExamples();
  }else{
    toast('❌ 更新失敗');
  }
}
async function loadTrainStatus(){
  try{
    var d=await api('/examples/stats');
    if(!d)return;
    document.getElementById('trainTotal').textContent=d.total||0;
    var z=document.getElementById('trainZh2id');if(z)z.textContent=d.zh2id||0;
    var i=document.getElementById('trainId2zh');if(i)i.textContent=d.id2zh||0;
    document.getElementById('trainWrong30').textContent=d.marked_wrong_30d||0;
  }catch(e){}
}
function showImportDialog(){
  var d=document.getElementById('importDialog');
  d.style.display=d.style.display==='none'?'block':'none';
  document.getElementById('importResult').textContent='';
}
async function doImport(){
  var text=document.getElementById('importText').value;
  var resEl=document.getElementById('importResult');
  if(!text.trim()){resEl.innerHTML='<span style="color:#f04747">請輸入資料</span>';return}
  var lines=text.replace(/\\r/g,'').split('\\n');
  var added=0, skipped=0, errors=0;
  for(var i=0;i<lines.length;i++){
    var line=lines[i].trim();
    if(!line||line.startsWith('#'))continue;
    var parts=line.split('|').map(function(p){return p.trim()});
    if(parts.length!==3){errors++;continue}
    var dir=parts[0].toLowerCase();
    if(dir!=='zh2id'&&dir!=='id2zh'){errors++;continue}
    var zh=parts[1], id=parts[2];
    if(!zh||!id){errors++;continue}
    try{
      var r=await api('/examples/add','POST',{zh:zh,id:id,dir:dir});
      if(r&&r.ok)added++;
      else if(r&&r.error==='duplicate')skipped++;
      else errors++;
    }catch(e){errors++}
  }
  resEl.innerHTML='<span style="color:#43b581">✅ 新增 '+added+' 筆</span>'+
    (skipped?'<span style="color:#faa61a;margin-left:8px">⏭ 跳過重複 '+skipped+'</span>':'')+
    (errors?'<span style="color:#f04747;margin-left:8px">❌ 錯誤 '+errors+'</span>':'');
  if(added>0){
    document.getElementById('importText').value='';
    setTimeout(function(){
      document.getElementById('importDialog').style.display='none';
      loadExamples();
    }, 1500);
  }
}
function downloadCsv(){
  if(!_exData||!_exData.length){toast('沒有範例可匯出');return}
  var rows=[['direction','zh','id']];
  for(var i=0;i<_exData.length;i++){
    var ex=_exData[i];
    rows.push([ex.dir||'zh2id', ex.zh||'', ex.id||'']);
  }
  var csv=rows.map(function(r){
    return r.map(function(c){
      var s=String(c).replace(/"/g,'""');
      return '"'+s+'"';
    }).join(',');
  }).join('\\n');
  // Add UTF-8 BOM so Excel opens it correctly
  var blob=new Blob(['\\ufeff'+csv],{type:'text/csv;charset=utf-8'});
  var url=URL.createObjectURL(blob);
  var a=document.createElement('a');
  a.href=url;
  var d=new Date();
  var stamp=d.getFullYear()+(''+(d.getMonth()+1)).padStart(2,'0')+(''+d.getDate()).padStart(2,'0');
  a.download='translation_examples_'+stamp+'.csv';
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
  URL.revokeObjectURL(url);
  toast('✅ CSV 已下載');
}
function downloadFinetuneJsonl(){
  // Kept for backward compatibility / future fine-tune route
  var url=window.location.origin+'/api/admin/examples/export_jsonl';
  var a=document.createElement('a');
  a.href=url;
  a.download='finetune_training_data.jsonl';
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
}
async function loadTranslationLog(){
  var onlyWrongEl=document.getElementById('logOnlyWrong');
  var onlyWrong=onlyWrongEl&&onlyWrongEl.checked?'1':'';
  var qs='?limit=50';
  if(onlyWrong) qs+='&only_wrong=1';
  var el=document.getElementById('tlogList');
  el.innerHTML='<div class="empty">載入中...</div>';
  try{
    var d=await api('/translation_log'+qs);
    if(!d||!d.entries){el.innerHTML='<div class="empty">無法載入</div>';return}
    if(!d.entries.length){el.innerHTML='<div class="empty">沒有翻譯日誌</div>';return}
    var html='';
    for(var i=0;i<d.entries.length;i++){
      var e=d.entries[i];
      var dt=new Date((e.ts||0)*1000);
      var ts=dt.toLocaleString('zh-TW',{month:'2-digit',day:'2-digit',hour:'2-digit',minute:'2-digit'});
      var dirLabel=(e.src_lang==='zh'?'🇹🇼→🇮🇩':(e.src_lang==='id'?'🇮🇩→🇹🇼':e.src_lang||'?'));
      var marked=e.marked_wrong?'<span class="badge" style="background:#5a2020;color:#f88;font-size:10px">已標錯</span> ':'';
      html+='<div style="padding:8px;border-bottom:1px solid #2a2a3e;font-size:12px">';
      html+='<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px">';
      html+='<span style="color:#8a8a9a">'+ts+' '+dirLabel+'</span>';
      html+=marked;
      html+='</div>';
      html+='<div style="color:#e0e0e0;margin-bottom:2px">原: '+esc(e.src||'')+'</div>';
      html+='<div style="color:#aaa">譯: '+esc(e.tgt||'')+'</div>';
      if(e.correct_translation){
        html+='<div style="color:#43b581;margin-top:4px">✓ 正確: '+esc(e.correct_translation)+'</div>';
      }
      if(!e.marked_wrong || !e.correct_translation){
        html+='<div style="margin-top:6px;display:flex;gap:6px">';
        html+='<input type="text" id="fix_'+e.id+'" placeholder="輸入正確翻譯..." style="flex:1;padding:4px 6px;border-radius:4px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px">';
        html+='<button class="btn btn-sm" style="background:#5a2a2a;color:#fff;border:none;padding:4px 10px;border-radius:4px;font-size:11px" onclick="markLogWrong(\\''+e.id+'\\')">標錯+修正</button>';
        html+='</div>';
      }
      html+='</div>';
    }
    el.innerHTML=html;
  }catch(err){
    el.innerHTML='<div class="empty">錯誤：'+err+'</div>';
  }
}
async function markLogWrong(entryId){
  var input=document.getElementById('fix_'+entryId);
  var correct=input?input.value.trim():'';
  if(!correct){
    if(!confirm('未填正確翻譯，只標記錯誤？')) return;
  }
  var d=await api('/translation_log/mark_wrong','POST',{entry_id:entryId,correct:correct});
  if(d&&d.ok){
    if(correct){
      toast('✅ 已加入範例 ('+(d.examples_total||'?')+' 筆) · 下次翻譯立即生效');
    }else{
      toast('✅ 已標記錯誤');
    }
    loadTranslationLog();
    loadExamples();
  }else{
    toast('❌ 失敗：'+(d&&d.info||''));
  }
}
// v3.9.4: 依當前方向更新 input 標籤與 placeholder
function updateExampleInputOrder(){
  var dir = document.getElementById('exDir').value;
  var srcLabel = document.getElementById('exSrcLabel');
  var tgtLabel = document.getElementById('exTgtLabel');
  var srcInput = document.getElementById('exSrc');
  var tgtInput = document.getElementById('exTgt');
  if(dir === 'id2zh'){
    srcLabel.textContent = '原文（印尼文)';
    tgtLabel.textContent = '正確翻譯（中文)';
    srcInput.placeholder = '例：Batu gerinda harus diganti';
    tgtInput.placeholder = '例：砂輪要換了';
  }else{
    srcLabel.textContent = '原文（中文)';
    tgtLabel.textContent = '正確翻譯（印尼文)';
    srcInput.placeholder = '例：砂輪要換了';
    tgtInput.placeholder = '例：Batu gerinda harus diganti';
  }
}

async function addExample(){
  var dir=document.getElementById('exDir').value;
  var src=document.getElementById('exSrc').value.trim();
  var tgt=document.getElementById('exTgt').value.trim();
  var resEl=document.getElementById('exAddResult');
  if(!src||!tgt){resEl.innerHTML='<span style="color:#f04747">請填入原文與正確翻譯</span>';return}
  // v3.9.4: 依方向把 src/tgt 對應回 zh/id 兩個資料庫欄位
  var zh, id;
  if(dir === 'zh2id'){
    zh = src; id = tgt;
  }else{
    id = src; zh = tgt;
  }
  // Check built-in duplicates
  if(_checkBuiltinDuplicate(zh,id)){
    resEl.innerHTML='<span style="color:#faa61a">⚠️ 此範例已存在於核心翻譯範例中，不需重複新增</span>';
    return;
  }
  // Check custom duplicates
  for(var i=0;i<_exData.length;i++){
    if(_exData[i].zh===zh&&_exData[i].id===id){
      resEl.innerHTML='<span style="color:#faa61a">⚠️ 此範例已存在</span>';
      return;
    }
  }
  var d=await api('/examples/add','POST',{zh:zh,id:id,dir:dir});
  if(d&&d.ok){
    resEl.innerHTML='<span style="color:#43b581">✅ 已新增（'+(dir==='zh2id'?'中→印':'印→中')+'）</span>';
    document.getElementById('exSrc').value='';
    document.getElementById('exTgt').value='';
    loadExamples();
  }else if(d&&d.error==='max_reached'){
    resEl.innerHTML='<span style="color:#f04747">❌ 已達上限 '+d.max+' 條</span>';
  }else if(d&&d.error==='duplicate'){
    resEl.innerHTML='<span style="color:#faa61a">⚠️ 此範例已存在</span>';
  }else{
    resEl.innerHTML='<span style="color:#f04747">❌ 新增失敗</span>';
  }
}
async function deleteExample(idx){
  if(!confirm('確定刪除此範例？'))return;
  var d=await api('/examples/delete','POST',{index:idx});
  if(d&&d.ok){toast('已刪除');loadExamples()}
  else toast('刪除失敗');
}

// ─── Forms System ───
function esc(s){if(!s)return '';return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;')}
var _formFields=[];
var _formFieldCounter=0;

function showCreateForm(){
  _formFields=[];_formFieldCounter=0;
  document.getElementById('formFieldsList').innerHTML='';
  document.getElementById('formTitleZh').value='';
  document.getElementById('formTitleId').value='';
  document.getElementById('formsCreateArea').style.display='block';
  addFormField();
}

function addFormField(){
  _formFieldCounter++;
  var fid='fld_'+_formFieldCounter;
  _formFields.push({id:fid,type:'text',label_zh:'',label_id:'',required:true,options:[]});
  renderFormFields();
}

function renderFormFields(){
  var c=document.getElementById('formFieldsList');
  var html='';
  for(var i=0;i<_formFields.length;i++){
    var f=_formFields[i];
    html+='<div style="padding:8px;margin-bottom:6px;background:#1a1a2e;border:1px solid #2a2a3e;border-radius:6px">';
    html+='<div style="display:flex;gap:4px;margin-bottom:4px;align-items:center"><span style="font-size:11px;color:#8a8a9a;min-width:14px">'+(i+1)+'</span>';
    html+='<select onchange="updFldType('+i+',this.value)" style="padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:11px">';
    var types=[['text','文字'],['number','數字'],['date','日期'],['textarea','長文'],['select','下拉'],['checkbox','核取']];
    for(var t=0;t<types.length;t++){html+='<option value="'+types[t][0]+'"'+(f.type===types[t][0]?' selected':'')+'>'+types[t][1]+'</option>'}
    html+='</select>';
    html+='<label style="font-size:11px;color:#8a8a9a;display:flex;align-items:center;gap:2px"><input type="checkbox" '+(f.required?'checked':'')+' onchange="updFldReq('+i+',this.checked)" style="width:14px;height:14px">必填</label>';
    html+='<span style="font-size:11px;color:#f04747;cursor:pointer;margin-left:auto" onclick="delFld('+i+')">✕</span></div>';
    html+='<input type="text" value="'+esc(f.label_zh)+'" placeholder="中文標籤" onchange="updFldLabel('+i+',&apos;zh&apos;,this.value)" style="width:100%;padding:5px;border-radius:4px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px;margin-bottom:3px">';
    html+='<input type="text" value="'+esc(f.label_id)+'" placeholder="Label Indonesia" onchange="updFldLabel('+i+',&apos;id&apos;,this.value)" style="width:100%;padding:5px;border-radius:4px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px">';
    if(f.type==='select'){
      html+='<input type="text" value="'+esc((f.options||[]).map(function(o){return o.zh+'/'+o.id}).join(', '))+'" placeholder="選項: 選項1中/選項1印, 選項2中/選項2印" onchange="updFldOpts('+i+',this.value)" style="width:100%;padding:5px;border-radius:4px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:11px;margin-top:3px">';
    }
    html+='</div>';
  }
  c.innerHTML=html;
}

function updFldType(i,v){_formFields[i].type=v;renderFormFields()}
function updFldReq(i,v){_formFields[i].required=v}
function updFldLabel(i,lang,v){if(lang==='zh')_formFields[i].label_zh=v;else _formFields[i].label_id=v}
function delFld(i){_formFields.splice(i,1);renderFormFields()}
function updFldOpts(i,v){
  var parts=v.split(',');
  _formFields[i].options=parts.map(function(p){
    var s=p.trim().split('/');
    return {zh:s[0]||'',id:s[1]||s[0]||''};
  });
}

async function submitCreateForm(){
  var tz=document.getElementById('formTitleZh').value.trim();
  var ti=document.getElementById('formTitleId').value.trim();
  if(!tz){alert('請填寫中文標題');return}
  if(!ti){alert('Harap isi judul Indonesia');return}
  if(_formFields.length===0){alert('至少新增一個欄位');return}
  for(var i=0;i<_formFields.length;i++){
    if(!_formFields[i].label_zh){alert('欄位 '+(i+1)+' 缺少中文標籤');return}
  }
  var r=await api('/forms/create','POST',{title_zh:tz,title_id:ti,fields:_formFields});
  if(r.ok){
    document.getElementById('formsCreateArea').style.display='none';
    loadFormsTab();
  }
}

async function loadFormsTab(){
  var r=await api('/forms');
  if(!r){document.getElementById('formsList').innerHTML='<div style="text-align:center;padding:20px;color:#f04747">載入失敗</div>';return}
  var forms=r.forms||[];
  var html='';
  if(forms.length===0){html='<div style="text-align:center;padding:20px;color:#8a8a9a">尚無表單</div>'}
  for(var i=0;i<forms.length;i++){
    var f=forms[i];
    var st=f.status==='active'?'<span style="color:#43b581">●啟用</span>':'<span style="color:#f04747">●停用</span>';
    html+='<div class="card" style="margin-bottom:8px">';
    html+='<div style="display:flex;justify-content:space-between;align-items:center">';
    html+='<div><div style="font-weight:600;font-size:14px">'+esc(f.title_zh)+'</div><div style="font-size:12px;color:#8a8a9a">'+esc(f.title_id)+'</div></div>';
    html+='<div style="text-align:right"><div style="font-size:11px">'+st+'</div><div style="font-size:11px;color:#8a8a9a">填寫: '+f.submission_count+'</div></div>';
    html+='</div>';
    html+='<div style="display:flex;gap:6px;margin-top:8px;flex-wrap:wrap">';
    html+='<button class="btn btn-primary btn-sm" style="padding:4px 8px;font-size:11px" onclick="viewFormSubmissions(&apos;'+f.id+'&apos;)">查看填寫</button>';
    html+='<button class="btn btn-sm" style="background:#2a2a3e;color:#aaa;border:1px solid #3a3a4e;padding:4px 8px;font-size:11px;border-radius:6px" onclick="pushFormToGroup(&apos;'+f.id+'&apos;)">推送到群組</button>';
    if(f.status==='active'){
      html+='<button class="btn btn-sm" style="background:#3a2a1a;color:#faa61a;border:1px solid #5a4a2a;padding:4px 8px;font-size:11px;border-radius:6px" onclick="toggleFormStatus(&apos;'+f.id+'&apos;,&apos;closed&apos;)">停用</button>';
    }else{
      html+='<button class="btn btn-sm" style="background:#1a3a2a;color:#43b581;border:1px solid #2a5a3a;padding:4px 8px;font-size:11px;border-radius:6px" onclick="toggleFormStatus(&apos;'+f.id+'&apos;,&apos;active&apos;)">啟用</button>';
    }
    html+='<button class="btn btn-sm" style="background:#3a1a1a;color:#f04747;border:1px solid #5a2a2a;padding:4px 8px;font-size:11px;border-radius:6px" onclick="deleteForm(&apos;'+f.id+'&apos;)">刪除</button>';
    html+='<a href="/api/admin/forms/export/'+f.id+'?key='+KEY+'" target="_blank" class="btn btn-sm" style="background:#2a2a3e;color:#aaa;border:1px solid #3a3a4e;padding:4px 8px;font-size:11px;border-radius:6px;text-decoration:none;display:inline-block">下載Excel</a>';
    html+='</div></div>';
  }
  document.getElementById('formsList').innerHTML=html;
  document.getElementById('formDetailArea').style.display='none';
}

async function toggleFormStatus(fid,status){
  await api('/forms/update','POST',{form_id:fid,status:status});
  loadFormsTab();
}

async function deleteForm(fid){
  if(!confirm('確定刪除此表單？所有填寫資料也會刪除。'))return;
  await api('/forms/delete','POST',{form_id:fid});
  loadFormsTab();
}

async function pushFormToGroup(fid){
  var r=await api('/groups');
  var groups=r.groups||[];
  if(groups.length===0){alert('沒有群組');return}
  var html='<div class="card"><div style="font-weight:600;font-size:14px;margin-bottom:8px">選擇推送群組</div>';
  for(var i=0;i<groups.length;i++){
    var g=groups[i];
    html+='<label style="display:flex;align-items:center;gap:6px;padding:4px 0;font-size:13px"><input type="checkbox" class="pushGrpCb" value="'+g.id+'" style="width:16px;height:16px">'+esc(g.name)+'</label>';
  }
  html+='<button class="btn btn-primary btn-sm" style="margin-top:8px" onclick="doPushForm(&apos;'+fid+'&apos;)">推送</button></div>';
  document.getElementById('formDetailArea').innerHTML=html;
  document.getElementById('formDetailArea').style.display='block';
}

async function doPushForm(fid){
  var cbs=document.querySelectorAll('.pushGrpCb:checked');
  var ids=[];cbs.forEach(function(c){ids.push(c.value)});
  if(ids.length===0){alert('請至少選一個群組');return}
  var r=await api('/forms/push','POST',{form_id:fid,group_ids:ids});
  if(r.ok){alert('已推送到 '+r.pushed+' 個群組');document.getElementById('formDetailArea').style.display='none'}
}

async function viewFormSubmissions(fid){
  var r=await fetch(API+'/forms/submissions/'+fid,{headers:{'X-Admin-Key':KEY,'X-Manager-Id':window._MANAGER_ID||''}});
  var data=await r.json();
  var subs=data.submissions||[];
  var form=data.form||{};
  var fields=form.fields||[];
  var html='<div class="card"><div style="font-weight:600;font-size:14px;margin-bottom:8px">'+esc(form.title_zh)+' — 填寫結果 ('+subs.length+')</div>';
  if(subs.length===0){
    html+='<div style="color:#8a8a9a;font-size:13px">尚無人填寫</div>';
  }else{
    for(var i=0;i<subs.length;i++){
      var s=subs[i];
      var approved=s.approved;
      html+='<div style="padding:8px;margin-bottom:6px;background:#0d0d1a;border-radius:6px;border:1px solid #2a2a3e">';
      html+='<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px"><span style="font-weight:600;font-size:13px">'+esc(s.user_name)+'</span>';
      html+='<span style="font-size:11px;'+(approved?'color:#43b581':'color:#faa61a')+'">'+(approved?'✅已核准':'⏳待核准')+'</span></div>';
      html+='<div style="font-size:11px;color:#666;margin-bottom:4px">'+esc(s.submitted_at)+'</div>';
      for(var j=0;j<fields.length;j++){
        var fd=fields[j];
        var ans=s.answers[fd.id]||'—';
        html+='<div style="font-size:12px;margin-bottom:2px"><span style="color:#8a8a9a">'+esc(fd.label_zh)+':</span> '+esc(ans)+'</div>';
      }
      if(!approved){
        html+='<button class="btn btn-sm" style="background:#1a3a2a;color:#43b581;border:1px solid #2a5a3a;padding:3px 8px;font-size:11px;border-radius:4px;margin-top:4px" onclick="approveSubmission(&apos;'+fid+'&apos;,&apos;'+s.user_id+'&apos;)">核准</button>';
      }
      html+='</div>';
    }
  }
  html+='</div>';
  document.getElementById('formDetailArea').innerHTML=html;
  document.getElementById('formDetailArea').style.display='block';
}

async function approveSubmission(fid,uid){
  await api('/forms/approve','POST',{form_id:fid,user_id:uid});
  viewFormSubmissions(fid);
}

// ─── Feature Settings ───
var _settingsGid='';
async function loadFeatureSettings(){
  await loadSettingsGroupSelects();
  _settingsGid='';
  await _loadFeatures('');
  // Load LINE quota
  var s=await api('/stats');
  if(s){
    var info='';
    if(s.line_quota){
      var q=s.line_quota;
      if(q.quota)info+='配額: '+q.quota.toLocaleString()+' 則/月';
      if(q.used!==null&&q.used!==undefined)info+=(info?' ｜ ':'')+'已用: '+q.used.toLocaleString()+' 則';
    }
    if(s.followers)info+=(info?'<br>':'')+'好友: '+s.followers;
    if(s.unfollowers)info+=' ｜ 封鎖: '+s.unfollowers;
    document.getElementById('lineQuotaInfo').innerHTML=info||'無法取得';
  }
  // Load Insight API data
  var ins=await api('/insight');
  if(ins){
    var ihtml='';
    if(ins.delivery&&ins.delivery.reply)ihtml+='昨日回覆: '+ins.delivery.reply+' 則';
    if(ins.demographics&&ins.demographics.available){
      if(ins.demographics.genders){ihtml+=(ihtml?'<br>':'')+'性別分布: '+JSON.stringify(ins.demographics.genders)}
    }
    document.getElementById('lineInsight').innerHTML=ihtml;
  }
  // Load Webhook info
  try{
    var wh=await api('/webhook');
    if(wh&&wh.webhook){
      var whInfo='URL: '+(wh.webhook.endpoint||'(未設定)');
      if(wh.webhook.active!==null)whInfo+=' ｜ '+(wh.webhook.active?'✅ 啟用':'❌ 停用');
      document.getElementById('webhookInfo').innerHTML=whInfo;
    }
  }catch(e){document.getElementById('webhookInfo').innerHTML='無法取得';}
  // Load Rich Menu list
  var rm=await api('/richmenu/list');
  if(rm&&rm.menus&&rm.menus.length){
    var rmhtml='目前選單: ';
    var rmSel=document.getElementById('rmUploadSelect');
    rmSel.innerHTML='<option value="">選擇 Rich Menu...</option>';
    for(var r=0;r<rm.menus.length;r++){
      rmhtml+='<span class="badge badge-on" style="font-size:11px;margin:2px">'+rm.menus[r].name+'</span> ';
      var opt=document.createElement('option');
      opt.value=rm.menus[r].id;opt.textContent=rm.menus[r].name||rm.menus[r].id.substring(0,16);
      rmSel.appendChild(opt);
    }
    document.getElementById('richMenuList').innerHTML=rmhtml;
  }else{
    document.getElementById('richMenuList').textContent='尚未建立 Rich Menu';
    document.getElementById('rmUploadSelect').innerHTML='<option value="">無選單</option>';
  }
  // Populate imagemap group select
  var imSel=document.getElementById('imapGroupSelect');
  imSel.innerHTML='<option value="">選擇群組...</option>';
  if(_groupList&&_groupList.length){
    for(var g=0;g<_groupList.length;g++){
      var opt=document.createElement('option');
      opt.value=_groupList[g].id;opt.textContent='#'+(_groupList[g].name||_groupList[g].id.substring(0,16));
      imSel.appendChild(opt);
    }
  }
  // Load delivery stats, rich menu default, aliases
  loadDeliveryStats();
  loadRichMenuDefault();
  loadAliases();
}
async function loadFeatureSettingsForGroup(){
  _settingsGid=document.getElementById('settingsGroupSelect').value;
  await _loadFeatures(_settingsGid);
}
async function _loadFeatures(gid){
  var path=gid?'/features?group_id='+encodeURIComponent(gid):'/features';
  var d=await api(path);
  if(!d)return;
  document.getElementById('welcomeToggle').checked=d.welcome_enabled;
  document.getElementById('welcomeZh').value=d.welcome_text_zh||'';
  document.getElementById('welcomeId').value=d.welcome_text_id||'';
  document.getElementById('flexToggle').checked=d.flex_enabled;
  document.getElementById('qrToggle').checked=d.quick_reply_enabled;
  document.getElementById('silentToggle').checked=d.silent_mode;
  document.getElementById('videoToggle').checked=d.video_ocr_enabled!==false;
  document.getElementById('locationToggle').checked=d.location_translate_enabled!==false;
  document.getElementById('markReadToggle').checked=d.mark_read_enabled!==false;
  document.getElementById('retryKeyToggle').checked=d.retry_key_enabled!==false;
  document.getElementById('cameraQrToggle').checked=d.camera_qr_enabled||false;
  document.getElementById('clipboardQrToggle').checked=d.clipboard_qr_enabled||false;
  document.getElementById('cameraRollQrToggle').checked=d.camera_roll_qr_enabled||false;
  document.getElementById('locationQrToggle').checked=d.location_qr_enabled||false;
  document.getElementById('toneSelect').value=d.translation_tone||'casual';
  document.getElementById('toneCustom').value=d.translation_tone_custom||'';
  // Model settings (global only, not per-group)
  if(!gid){
    document.getElementById('modelDefault').value=d.model_default||'gpt-4.1-mini';
    document.getElementById('modelUpgrade').value=d.model_upgrade||'gpt-4.1';
    // v3.9: vision (照片分析) model
    if(document.getElementById('visionModel')) document.getElementById('visionModel').value=d.vision_model||'gpt-5-mini';
    // v3.2-0426d Batch B: load advanced settings
    if(document.getElementById('ttemp')) document.getElementById('ttemp').value=(d.translation_temperature!==undefined?d.translation_temperature:0);
    if(document.getElementById('ttopp')) document.getElementById('ttopp').value=(d.translation_top_p!==undefined?d.translation_top_p:1.0);
    if(document.getElementById('tseed')) document.getElementById('tseed').value=(d.translation_seed!==undefined?d.translation_seed:0);
    if(document.getElementById('dcMode')) document.getElementById('dcMode').value=d.double_check_mode||'smart';
    if(document.getElementById('dcThr')) document.getElementById('dcThr').value=(d.double_check_threshold!==undefined?d.double_check_threshold:0.55);
    if(document.getElementById('dcKw')) document.getElementById('dcKw').value=d.double_check_keywords||'';
    if(document.getElementById('fsMode')) document.getElementById('fsMode').value=d.fewshot_mode||'messages';
    // v3.4 ID→ZH 強化開關
    if(document.getElementById('idZhCoD')) document.getElementById('idZhCoD').checked=d.id_zh_cod_enabled!==false;
    if(document.getElementById('idZhCoT')) document.getElementById('idZhCoT').checked=d.id_zh_cot_enabled!==false;
    if(document.getElementById('idZhPivot')) document.getElementById('idZhPivot').checked=!!d.id_zh_pivot_enabled;
    if(document.getElementById('idZhPivotThr')) document.getElementById('idZhPivotThr').value=(d.id_zh_pivot_threshold||80);
    // v3.6 印尼文預處理 + 多路徑反譯
    if(document.getElementById('idPrep')) document.getElementById('idPrep').checked=d.id_preprocessing_enabled!==false;
    if(document.getElementById('idPrepNano')) document.getElementById('idPrepNano').checked=!!d.id_preprocessing_nano;
    if(document.getElementById('multiPath')) document.getElementById('multiPath').checked=!!d.multi_path_backtrans_enabled;
    if(document.getElementById('multiPathMin')) document.getElementById('multiPathMin').value=(d.multi_path_min_chars||60);
    // v3.7 段落結構保留
    if(document.getElementById('paraPreserve')) document.getElementById('paraPreserve').checked=d.preserve_paragraphs_enabled!==false;
    if(document.getElementById('paraSplit')) document.getElementById('paraSplit').checked=d.paragraph_split_translate!==false;
    if(document.getElementById('paraThr')) document.getElementById('paraThr').value=(d.paragraph_split_threshold||50);
    // Batch C/D load
    if(document.getElementById('lpEn')) document.getElementById('lpEn').checked=d.logprobs_enabled!==false;
    if(document.getElementById('cfThr')) document.getElementById('cfThr').value=(d.confidence_threshold!==undefined?d.confidence_threshold:0.85);
    if(document.getElementById('soEn')) document.getElementById('soEn').checked=!!d.structured_output_enabled;
    if(document.getElementById('pcEn')) document.getElementById('pcEn').checked=d.prompt_caching_enabled!==false;
    if(document.getElementById('tlEn')) document.getElementById('tlEn').checked=d.translation_logging_enabled!==false;
    // v3.2-0426e: load official feature settings
    if(document.getElementById('ssEn')) document.getElementById('ssEn').checked=d.stop_sequences_enabled!==false;
    if(document.getElementById('reEf')) document.getElementById('reEf').value=d.reasoning_effort||'medium';
    if(document.getElementById('suid')) document.getElementById('suid').checked=d.send_user_id_to_openai!==false;
    if(document.getElementById('smeta')) document.getElementById('smeta').checked=d.send_metadata_to_openai!==false;
    if(typeof loadTransLog==='function') setTimeout(function(){loadTransLog('all')},500);
    document.getElementById('modelThreshold').value=d.model_threshold||0;
  }
  document.getElementById('senderNameInput').value=d.sender_name||'翻譯小助手';
  document.getElementById('senderIconInput').value=d.sender_icon||'';
  var cb=document.getElementById('settingsCustomBadge');
  if(gid&&d.is_customized)cb.style.display='block';
  else cb.style.display='none';
  if(d.bot_info&&d.bot_info.name){
    var qi=document.getElementById('lineQuotaInfo');
    if(qi&&!qi.innerHTML.includes('Bot:'))qi.innerHTML+='<br>Bot: '+d.bot_info.name;
  }
  // v3.9.7: 載入完成後立刻檢查相容性 + 標灰不適用設定(silent 模式不顯示提示)
  setTimeout(function(){
    try{
      autoNormalizeIncompatibleSettings({silent:true});
      updateSettingsCompatibility();
    }catch(e){}
  }, 100);
}
function toggleFeatureSetting(key,val){
  var body={};body[key]=val;
  if(_settingsGid)body.group_id=_settingsGid;
  api('/features','POST',body).then(function(d){if(d)toast(_settingsGid?'群組設定已更新':'全域設定已更新')});
}
// v3.9.7: 前端模型能力對照表(鏡像後端 MODEL_CAPABILITIES)
function _modelFamily(name){
  if(!name) return 'unknown';
  var n=name.toLowerCase();
  if(n.indexOf('gpt-5.5')===0||n.indexOf('gpt-5.4')===0||n.indexOf('gpt-5')===0) return 'gpt5';
  if(n.indexOf('o1')===0||n.indexOf('o3')===0||n.indexOf('o4')===0) return 'oseries';
  if(n.indexOf('gpt-4')===0) return 'gpt4';
  return 'unknown';
}
var MODEL_CAPS_FRONT={
  gpt5:    {temperature:false,top_p:false,seed:false,logprobs:false,logit_bias:false,stop:false,structured_output:true, reasoning_effort:true},
  oseries: {temperature:false,top_p:false,seed:false,logprobs:false,logit_bias:false,stop:false,structured_output:false,reasoning_effort:true},
  gpt4:    {temperature:true, top_p:true, seed:true, logprobs:true, logit_bias:true, stop:true, structured_output:true, reasoning_effort:false},
  unknown: {temperature:true, top_p:true, seed:true, logprobs:true, logit_bias:true, stop:true, structured_output:true, reasoning_effort:false},
};
function _modelCap(model,cap){return MODEL_CAPS_FRONT[_modelFamily(model)][cap];}

// v3.9.7 核心:autoNormalizeIncompatibleSettings()
// 切換模型時,自動把對新模型「無效的設定」歸回中性值。
// 這樣使用者切到 GPT-5 不需要再手動調 temperature/logprobs/seed,
// 切回 GPT-4 也不會留著 GPT-5 殘留的 reasoning_effort 為唯一輸出控制。
// 邏輯:檢查兩個翻譯模型(短訊息預設 + 長訊息升級),只要兩者都不支援的能力就歸零。
function autoNormalizeIncompatibleSettings(opts){
  opts=opts||{};
  var silent=opts.silent;  // true 時不顯示提示(初始載入用)
  var modelD=document.getElementById('modelDefault');
  var modelU=document.getElementById('modelUpgrade');
  if(!modelD||!modelU) return;
  var mD=modelD.value, mU=modelU.value;

  // helper:兩模型「都不支援」才歸零(避免單邊支援的設定被誤殺)
  function bothLack(cap){return !_modelCap(mD,cap) && !_modelCap(mU,cap);}
  function bothHave(cap){return _modelCap(mD,cap) && _modelCap(mU,cap);}

  var changed=[];

  // —— Temperature:reasoning model 強制 1.0 ——
  var ttemp=document.getElementById('ttemp');
  if(ttemp && bothLack('temperature') && parseFloat(ttemp.value)!==1.0){
    ttemp.value='1.0';
    changed.push('temperature → 1.0(reasoning model 不支援自訂)');
  }
  // —— Top_p ——
  var ttopp=document.getElementById('ttopp');
  if(ttopp && bothLack('top_p') && parseFloat(ttopp.value)!==1.0){
    ttopp.value='1.0';
    changed.push('top_p → 1.0');
  }
  // —— Seed ——
  var tseed=document.getElementById('tseed');
  if(tseed && bothLack('seed') && parseInt(tseed.value)!==0){
    tseed.value='0';
    changed.push('seed → 0(reasoning model 無法重現種子)');
  }
  // —— Logprobs(信心度計算)——
  var lp=document.getElementById('logprobs');
  if(lp && bothLack('logprobs') && lp.checked){
    lp.checked=false;
    changed.push('Logprobs 信心度 → 關閉(reasoning model 不支援)');
  }
  // —— Stop sequences + logit_bias ——
  var ss=document.getElementById('ssEn');
  if(ss && bothLack('logit_bias') && ss.checked){
    ss.checked=false;
    changed.push('Stop sequences/logit_bias → 關閉(reasoning model 不支援)');
  }
  // —— Structured Output(o-series 不支援,但 gpt-5/gpt-4 都支援)——
  var so=document.getElementById('soutEn');
  if(so && bothLack('structured_output') && so.checked){
    so.checked=false;
    changed.push('結構化輸出 → 關閉(o-series 不支援)');
  }
  // —— Reasoning effort(反向:GPT-4 系列不支援,變成擺設)——
  // 不歸零,因為值留著對 reasoning model 切回時還能用,只在 UI 標灰

  // —— 提示使用者哪些設定被自動還原 ——
  if(changed.length>0 && !silent){
    // (避免在 Python 三引號內寫 backslash-n;改用 br 標籤)
    var msg='⚙️ 已自動調整以下設定以配合新模型:<br>• '+changed.join('<br>• ');
    // 用 toast 而非 alert,不阻塞操作
    if(typeof toast==='function'){
      toast('已自動調整 '+changed.length+' 項設定');
    }
    // 詳細列表寫到模型旁邊的提示卡
    var note=document.getElementById('autoNormalizeNote');
    if(!note){
      note=document.createElement('div');
      note.id='autoNormalizeNote';
      note.style.cssText='font-size:11px;margin-top:6px;padding:6px 8px;background:#1a3a2e;border:1px solid #43b581;border-radius:6px;color:#43b581;line-height:1.5';
      var anchor=document.getElementById('modelSaveResult');
      if(anchor && anchor.parentNode){
        anchor.parentNode.insertBefore(note, anchor.nextSibling);
      }
    }
    note.innerHTML='✅ <b>已自動配合新模型調整這些設定:</b><br>• '+changed.join('<br>• ');
    // 8 秒後自動隱藏
    setTimeout(function(){
      if(note) note.style.display='none';
    }, 10000);
  }
}

// updateSettingsCompatibility:標灰 + 提示哪些欄位對目前模型無效
function updateSettingsCompatibility(){
  var modelD=document.getElementById('modelDefault');
  var modelU=document.getElementById('modelUpgrade');
  if(!modelD||!modelU) return;
  var mD=modelD.value, mU=modelU.value;

  function markField(elementId, capability){
    var el=document.getElementById(elementId);
    if(!el) return;
    var sD=_modelCap(mD,capability), sU=_modelCap(mU,capability);
    var hint=document.getElementById(elementId+'_compat_hint');
    if(!hint){
      hint=document.createElement('div');
      hint.id=elementId+'_compat_hint';
      hint.style.cssText='font-size:11px;margin-top:3px;line-height:1.4';
      if(el.parentNode){el.parentNode.insertBefore(hint, el.nextSibling);}
    }
    if(sD && sU){
      hint.innerHTML=''; hint.style.display='none';
      el.style.opacity='1';
    }else if(!sD && !sU){
      hint.innerHTML='⚠️ <span style="color:#faa61a">此設定對目前兩個模型皆無效,系統會自動忽略</span>';
      hint.style.display='block';
      el.style.opacity='0.45';
    }else{
      hint.innerHTML='💡 <span style="color:#faa61a">僅對 '+(sD?mD:mU)+' 有效</span>';
      hint.style.display='block';
      el.style.opacity='0.7';
    }
  }
  markField('ttemp','temperature');
  markField('ttopp','top_p');
  markField('tseed','seed');
  markField('logprobs','logprobs');
  markField('confTh','logprobs');
  markField('ssEn','logit_bias');
  markField('reEf','reasoning_effort');
  markField('soutEn','structured_output');

  // 模型類型摘要
  var summary=document.getElementById('modelCapSummary');
  if(!summary){
    summary=document.createElement('div');
    summary.id='modelCapSummary';
    summary.style.cssText='font-size:11px;margin-top:6px;padding:6px 8px;background:#0d0d1a;border-radius:6px;border:1px solid #2a2a3e;line-height:1.5';
    var anchor=document.getElementById('modelSaveResult');
    if(anchor && anchor.parentNode){
      anchor.parentNode.insertBefore(summary, anchor);
    }
  }
  var fL=function(f){
    if(f==='gpt5') return 'GPT-5 系列(reasoning model)';
    if(f==='oseries') return 'o-series(reasoning model)';
    if(f==='gpt4') return 'GPT-4 系列(classical chat)';
    return f;
  };
  var fD=_modelFamily(mD), fU=_modelFamily(mU);
  var info='🤖 <b>目前模型類型:</b><br>';
  info+='&nbsp;&nbsp;短訊息: <code style="color:#7c6fef">'+mD+'</code> → '+fL(fD)+'<br>';
  if(mU!==mD) info+='&nbsp;&nbsp;長訊息: <code style="color:#7c6fef">'+mU+'</code> → '+fL(fU)+'<br>';
  if(fD==='gpt5'||fU==='gpt5'||fD==='oseries'||fU==='oseries'){
    info+='<span style="color:#43b581">✅ 不相容的進階設定會自動忽略,翻譯永遠不會中斷</span>';
  }else{
    info+='<span style="color:#43b581">✅ 所有進階設定都生效</span>';
  }
  summary.innerHTML=info;
}

// 模型下拉切換 onchange 統一入口:先自動還原,再標灰提示
function onModelChange(){
  autoNormalizeIncompatibleSettings({silent:false});
  updateSettingsCompatibility();
  // 自動把調整後的進階設定也存回後台,使用者免再點一次儲存
  if(typeof saveAdvancedSettings==='function'){
    try{ saveAdvancedSettings(); }catch(e){}
  }
}

function saveModelSettings(){
  var md=document.getElementById('modelDefault').value;
  var mu=document.getElementById('modelUpgrade').value;
  var mt=parseInt(document.getElementById('modelThreshold').value)||0;
  api('/features','POST',{model_default:md,model_upgrade:mu,model_threshold:mt}).then(function(d){
    if(d){
      toast('模型設定已儲存');
      var info=mt>0?'≥'+mt+'字用 '+mu+'，其餘用 '+md:'全部用 '+md;
      document.getElementById('modelSaveResult').innerHTML='<span style="color:#43b581">✅ '+info+'</span>';
      onModelChange();  // 儲存後立刻自動調整 + 提示
    }
  });
}
function saveVisionModel(){
  var vm=document.getElementById('visionModel').value;
  api('/features','POST',{vision_model:vm}).then(function(d){
    if(d){
      toast('照片分析模型已儲存');
      document.getElementById('visionSaveResult').innerHTML='<span style="color:#43b581">✅ 照片分析改用 '+vm+'（失敗自動 fallback 到 gpt-4o-mini）</span>';
    }
  });
}
function saveAdvancedSettings(){
  var body={
    translation_temperature: parseFloat(document.getElementById('ttemp').value)||0,
    translation_top_p: parseFloat(document.getElementById('ttopp').value)||1.0,
    translation_seed: parseInt(document.getElementById('tseed').value)||0,
    double_check_mode: document.getElementById('dcMode').value,
    double_check_threshold: parseFloat(document.getElementById('dcThr').value)||0.55,
    double_check_keywords: document.getElementById('dcKw').value,
    fewshot_mode: document.getElementById('fsMode').value
  };
  api('/features','POST',body).then(function(d){
    if(d){
      toast('進階設定已儲存');
      document.getElementById('advSaveResult').innerHTML='<span style="color:#43b581">✅ 進階設定已生效</span>';
    }
  });
}
function saveOfficialFeatures(){
  var body={
    stop_sequences_enabled: document.getElementById('ssEn').checked,
    reasoning_effort: document.getElementById('reEf').value,
    send_user_id_to_openai: document.getElementById('suid').checked,
    send_metadata_to_openai: document.getElementById('smeta').checked
  };
  api('/features','POST',body).then(function(d){if(d)toast('官方參數已儲存')});
}
function saveBatchCD(){
  var body={
    logprobs_enabled: document.getElementById('lpEn').checked,
    confidence_threshold: parseFloat(document.getElementById('cfThr').value)||0.85,
    structured_output_enabled: document.getElementById('soEn').checked,
    prompt_caching_enabled: document.getElementById('pcEn').checked,
    translation_logging_enabled: document.getElementById('tlEn').checked
  };
  api('/features','POST',body).then(function(d){if(d)toast('已儲存')});
}
// v3.4 ID→ZH 翻譯強化開關儲存
function saveIdZhEnhance(){
  var body={
    id_zh_cod_enabled: document.getElementById('idZhCoD').checked,
    id_zh_cot_enabled: document.getElementById('idZhCoT').checked,
    id_zh_pivot_enabled: document.getElementById('idZhPivot').checked,
    id_zh_pivot_threshold: parseInt(document.getElementById('idZhPivotThr').value)||80
  };
  api('/features','POST',body).then(function(d){if(d)toast('ID→ZH 強化已儲存')});
}

// v3.6 預處理 + 多路徑反譯設定儲存
function saveV36Settings(){
  var body={
    id_preprocessing_enabled: document.getElementById('idPrep').checked,
    id_preprocessing_nano: document.getElementById('idPrepNano').checked,
    multi_path_backtrans_enabled: document.getElementById('multiPath').checked,
    multi_path_min_chars: parseInt(document.getElementById('multiPathMin').value)||60
  };
  api('/features','POST',body).then(function(d){if(d)toast('v3.6 設定已儲存')});
}

// v3.7 段落結構保留設定儲存
function saveV37Settings(){
  var body={
    preserve_paragraphs_enabled: document.getElementById('paraPreserve').checked,
    paragraph_split_translate: document.getElementById('paraSplit').checked,
    paragraph_split_threshold: parseInt(document.getElementById('paraThr').value)||50
  };
  api('/features','POST',body).then(function(d){if(d)toast('段落保留已儲存')});
}

// v3.6 翻譯品質監控儀表板
async function loadQualityStats(days){
  days = days || 7;
  var d = await api('/translation-stats?days='+days, 'GET');
  if(!d) return;
  
  // === 五大指標 ===
  var m = d.metrics || {};
  var metricsHtml = 
    '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:8px">' +
    '<div style="padding:8px;background:#1a1a2e;border-radius:4px"><div style="font-size:11px;color:#8a8a9a">總翻譯數</div><div style="font-size:18px;font-weight:700;color:#e0e0e0">'+(d.total||0)+'</div></div>' +
    '<div style="padding:8px;background:#1a1a2e;border-radius:4px"><div style="font-size:11px;color:#8a8a9a">平均長度比</div><div style="font-size:18px;font-weight:700;color:'+(m.avg_length_ratio>0.6&&m.avg_length_ratio<1.4?'#10b981':'#f59e0b')+'">'+(m.avg_length_ratio||'N/A')+'</div></div>' +
    '<div style="padding:8px;background:#1a1a2e;border-radius:4px"><div style="font-size:11px;color:#8a8a9a">災難壓縮</div><div style="font-size:18px;font-weight:700;color:'+(m.catastrophic_compression_count?'#ef4444':'#10b981')+'">'+(m.catastrophic_compression_count||0)+'</div></div>' +
    '<div style="padding:8px;background:#1a1a2e;border-radius:4px"><div style="font-size:11px;color:#8a8a9a">術語遺失</div><div style="font-size:18px;font-weight:700;color:'+(m.missing_terms_count?'#ef4444':'#10b981')+'">'+(m.missing_terms_count||0)+'</div></div>' +
    '<div style="padding:8px;background:#1a1a2e;border-radius:4px"><div style="font-size:11px;color:#8a8a9a">GPT 自首未翻完</div><div style="font-size:18px;font-weight:700;color:'+(m.claimed_incomplete_count?'#f59e0b':'#10b981')+'">'+(m.claimed_incomplete_count||0)+'</div></div>' +
    '<div style="padding:8px;background:#1a1a2e;border-radius:4px"><div style="font-size:11px;color:#8a8a9a">多路徑失敗</div><div style="font-size:18px;font-weight:700;color:'+(m.multi_path_fails?'#f59e0b':'#10b981')+'">'+(m.multi_path_fails||0)+'</div></div>' +
    '<div style="padding:8px;background:#1a1a2e;border-radius:4px"><div style="font-size:11px;color:#8a8a9a">⚠️ 警告</div><div style="font-size:18px;font-weight:700;color:#f59e0b">'+(m.warned_total||0)+'</div></div>' +
    '<div style="padding:8px;background:#1a1a2e;border-radius:4px"><div style="font-size:11px;color:#8a8a9a">已標錯</div><div style="font-size:18px;font-weight:700;color:#ef4444">'+(m.wrong_total||0)+'</div></div>' +
    '</div>';
  document.getElementById('qsMetrics').innerHTML = metricsHtml;
  
  // === 各語言對 ===
  var langHtml = '';
  (d.by_lang || []).forEach(function(b){
    langHtml += '<div style="padding:6px 8px;margin-bottom:4px;background:#0d0d1a;border-radius:4px;display:flex;justify-content:space-between">' +
      '<span style="color:#e0e0e0;font-weight:600">'+b.pair+'</span>' +
      '<span style="color:#8a8a9a">'+b.total+' 筆 | 警告率 '+b.warned_rate+'% | 長度比 '+(b.avg_length_ratio||'N/A')+'</span>' +
      '</div>';
  });
  document.getElementById('qsByLang').innerHTML = langHtml || '<div style="color:#6a6a7a;font-size:11px">無資料</div>';
  
  // === 各群組(最差排前面)===
  var grpHtml = '';
  (d.by_group || []).slice(0, 10).forEach(function(g){
    var color = g.quality_score >= 80 ? '#10b981' : (g.quality_score >= 60 ? '#f59e0b' : '#ef4444');
    grpHtml += '<div style="padding:6px 8px;margin-bottom:4px;background:#0d0d1a;border-radius:4px;display:flex;justify-content:space-between">' +
      '<span style="color:#e0e0e0;font-family:monospace;font-size:11px">'+(g.group_id||'unknown').substr(0,32)+'</span>' +
      '<span style="color:'+color+';font-weight:600">'+g.quality_score+' 分 ('+g.total+' 筆,'+g.warned+' 警告)</span>' +
      '</div>';
  });
  document.getElementById('qsByGroup').innerHTML = grpHtml || '<div style="color:#6a6a7a;font-size:11px">無資料</div>';
  
  // === 最近的問題訊息 ===
  var issuesHtml = '';
  (d.issues || []).forEach(function(it){
    var time = new Date(it.ts*1000).toLocaleString();
    var reasons = [];
    if(it.rtc_fail_reason) reasons.push('反譯失敗:'+it.rtc_fail_reason);
    if(it.struct_self_check) reasons.push('Self-check:'+it.struct_self_check);
    if(it.multi_path_fail) reasons.push('多路徑:'+it.multi_path_fail.final_decision);
    if(it.marked_wrong) reasons.push('已手動標錯');
    issuesHtml += '<div style="padding:8px;margin-bottom:6px;background:#0d0d1a;border-left:3px solid #ef4444;border-radius:4px">' +
      '<div style="color:#8a8a9a;font-size:11px">'+time+' | '+(it.src_lang||'?')+'→'+(it.tgt_lang||'?')+' | 長度比 '+(it.length_ratio||'N/A')+'</div>' +
      '<div style="color:#a0a0c0;margin:4px 0">'+escHtml(it.src||'')+'</div>' +
      '<div style="color:#e0e0e0;margin-bottom:4px">→ '+escHtml(it.tgt||'')+'</div>' +
      '<div style="color:#ef4444;font-size:11px">'+reasons.join(' | ')+'</div>' +
      '</div>';
  });
  document.getElementById('qsIssues').innerHTML = issuesHtml || '<div style="color:#6a6a7a;padding:12px">無問題訊息 ✓</div>';
}
async function loadTransLog(filter){
  filter = filter || 'all';
  var url = '/translation-log?limit=100';
  if(filter==='warned') url += '&only_warned=1';
  if(filter==='wrong') url += '&only_wrong=1';
  var d = await api(url, 'GET');
  if(!d) return;
  // Stats
  var s = d.stats || {};
  document.getElementById('tlStats').innerHTML =
    '總計: <b>' + (s.total||0) + '</b> 筆 | ' +
    '警告: <b style="color:#f59e0b">' + (s.warned||0) + '</b> | ' +
    '已標錯: <b style="color:#ef4444">' + (s.wrong||0) + '</b> | ' +
    '平均信心: <b>' + (s.avg_confidence||0) + '</b>';
  // List
  var html = '';
  (d.items||[]).forEach(function(it){
    var d2 = new Date(it.ts*1000);
    var time = d2.getHours()+':'+('0'+d2.getMinutes()).slice(-2);
    var warn = it.tgt && it.tgt.indexOf('⚠️')===0 ? ' style="border-left:3px solid #f59e0b"' : '';
    var wrong = it.marked_wrong ? ' style="border-left:3px solid #ef4444;opacity:0.6"' : '';
    var conf = it.confidence!==null && it.confidence!==undefined ? ' | 信心:'+it.confidence : '';
    var sim = it.similarity!==null && it.similarity!==undefined ? ' | 相似:'+it.similarity : '';
    html += '<div'+(wrong||warn)+' style="padding:8px;margin-bottom:6px;background:#0d0d1a;border-radius:6px">' +
            '<div style="color:#8a8a9a;font-size:11px">'+time+' | '+it.model+conf+sim+'</div>' +
            '<div style="color:#a0a0c0">'+escHtml(it.src||'')+'</div>' +
            '<div style="color:#e0e0e0;margin:4px 0">→ '+escHtml(it.tgt||'')+'</div>' +
            (it.marked_wrong ? '<div style="color:#ef4444;font-size:11px">已標記為錯誤</div>'
                             : '<button class="btn btn-dark btn-sm" onclick="markWrong('+it.ts+')">❌ 標記錯誤</button>') +
            '</div>';
  });
  document.getElementById('tlList').innerHTML = html || '<div style="color:#8a8a9a;padding:12px">無資料</div>';
}
function escHtml(s){return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')}
async function markWrong(ts){
  var correct = prompt('輸入正確翻譯（留空僅標記不加範例）：');
  if(correct === null) return;
  var body = {ts: ts, correct_translation: correct, add_to_examples: !!correct};
  var d = await api('/translation-log','POST',body);
  if(d && d.ok){toast('已標記');loadTransLog('all')}
}
async function clearTransLog(){
  if(!confirm('清空所有翻譯日誌？')) return;
  await api('/translation-log','DELETE',{});
  loadTransLog('all');
}
function saveWelcomeText(){
  var zh=document.getElementById('welcomeZh').value;
  var id=document.getElementById('welcomeId').value;
  var body={welcome_text_zh:zh,welcome_text_id:id};
  if(_settingsGid)body.group_id=_settingsGid;
  api('/features','POST',body).then(function(d){if(d)toast('歡迎詞已儲存')});
}
async function resetGroupSettings(){
  if(!_settingsGid){toast('請先選擇群組');return}
  if(!confirm('確定重設此群組為全域預設？'))return;
  var body={group_id:_settingsGid,reset:true};
  var d=await api('/features/reset','POST',body);
  if(d&&d.ok){toast('已重設');loadFeatureSettingsForGroup()}
  else toast('重設失敗');
}
function saveSenderSettings(){
  var name=document.getElementById('senderNameInput').value.trim();
  var icon=document.getElementById('senderIconInput').value.trim();
  if(!name){toast('請輸入名稱');return}
  api('/features','POST',{sender_name:name,sender_icon:icon}).then(function(d){if(d)toast('已更新')});
}
async function broadcastMessage(){
  var text=document.getElementById('pushText').value.trim();
  if(!text){toast('請輸入訊息');return}
  if(!confirm('確定推送給所有好友？'))return;
  var d=await api('/broadcast','POST',{text:text});
  if(d&&d.ok){toast('已推送全體');document.getElementById('pushText').value=''}
  else toast('推送失敗');
}
async function testWebhook(){
  document.getElementById('webhookTestResult').innerHTML='測試中...';
  var d=await api('/webhook/test','POST',{});
  if(d){
    var r=d.success?'✅ 成功':'❌ 失敗';
    if(d.status_code)r+=' ('+d.status_code+')';
    if(d.reason)r+=' - '+d.reason;
    document.getElementById('webhookTestResult').innerHTML=r;
  }else{document.getElementById('webhookTestResult').innerHTML='測試失敗';}
}
async function loadFollowers(){
  document.getElementById('followersList').innerHTML='載入中...';
  var d=await api('/followers');
  if(d&&d.followers){
    var h='總計: '+d.count+' 人<br>';
    for(var i=0;i<d.followers.length&&i<100;i++){
      var f=d.followers[i];
      h+='<span style="font-size:11px">'+(f.name||f.user_id.substr(0,12)+'...')+'</span> ';
    }
    if(d.count>100)h+='<br>...僅顯示前100人';
    document.getElementById('followersList').innerHTML=h;
  }else{document.getElementById('followersList').innerHTML='無法載入';}
}
async function loadDeliveryStats(){
  var d=await api('/delivery');
  if(d&&d.delivery){
    var s=d.delivery;
    var h='日期: '+(s.date||'-');
    if(s.reply!==null)h+='<br>Reply: '+s.reply;
    if(s.push!==null)h+=' ｜ Push: '+s.push;
    if(s.multicast!==null)h+=' ｜ Multicast: '+s.multicast;
    if(s.broadcast!==null)h+='<br>Broadcast: '+s.broadcast;
    document.getElementById('deliveryStats').innerHTML=h;
  }else{document.getElementById('deliveryStats').innerHTML='無法取得';}
}
async function checkUserMenu(){
  var uid=document.getElementById('rmUserIdInput').value.trim();
  if(!uid){toast('請輸入 user ID');return}
  var d=await api('/richmenu/user/'+uid);
  if(d){
    document.getElementById('rmUserResult').innerHTML=d.rich_menu_id?('綁定: '+d.rich_menu_id):'無綁定（使用預設）';
  }
}
async function loadAliases(){
  var d=await api('/richmenu/alias/list');
  if(d&&d.aliases){
    if(!d.aliases.length){document.getElementById('rmAliasList').innerHTML='（無 alias）';return;}
    var h='';
    for(var i=0;i<d.aliases.length;i++){
      var a=d.aliases[i];
      h+='<div style="padding:4px 0;border-bottom:1px solid #2a2a3e"><b>'+a.alias_id+'</b> → '+a.rich_menu_id+'</div>';
    }
    document.getElementById('rmAliasList').innerHTML=h;
  }
}
async function loadRichMenuDefault(){
  var d=await api('/richmenu/default');
  if(d){
    document.getElementById('richMenuDefault').innerHTML='預設選單 ID: '+(d.default_rich_menu_id||'（未設定）');
  }
}
async function pushMessage(){
  var gid=document.getElementById('pushGroupSelect').value;
  var text=document.getElementById('pushText').value.trim();
  if(!gid){toast('請選擇群組');return}
  if(!text){toast('請輸入訊息');return}
  if(!confirm('確定推送到群組？'))return;
  var d=await api('/push','POST',{group_id:gid,text:text});
  if(d&&d.ok){toast('已推送');document.getElementById('pushText').value=''}
  else toast('推送失敗');
}
function createRichMenu(){
  toast('建立中...');
  api('/richmenu','POST',{action:'create'}).then(function(d){
    if(d&&d.ok){toast('Rich Menu 已建立');loadFeatureSettings();}
    else toast('建立失敗');
  });
}
function deleteRichMenu(){
  if(!confirm('確定刪除所有 Rich Menu？'))return;
  toast('刪除中...');
  api('/richmenu','POST',{action:'delete'}).then(function(d){
    if(d&&d.ok){toast('已刪除 '+(d.deleted||0)+' 個選單');loadFeatureSettings();}
    else toast('刪除失敗');
  });
}
async function loadSettingsGroupSelects(){
  var d=await api('/groups');
  if(!d)return;
  var groups=d.groups||[];
  var sels=['pushGroupSelect','settingsGroupSelect'];
  for(var s=0;s<sels.length;s++){
    var sel=document.getElementById(sels[s]);
    if(!sel)continue;
    var cur=sel.value;
    sel.innerHTML=sels[s]==='settingsGroupSelect'?'<option value="">全域預設</option>':'<option value="">選擇群組...</option>';
    for(var i=0;i<groups.length;i++){
      var g=groups[i];
      var opt=document.createElement('option');
      opt.value=g.id;opt.textContent='#'+(g.name||g.id.substring(0,16));
      sel.appendChild(opt);
    }
    if(cur)sel.value=cur;
  }
}

window.addEventListener('load',function(){
  var k=localStorage.getItem('bot_admin_key');
  if(k){document.getElementById('pwInput').value=k;doLogin()}
});
if('serviceWorker' in navigator){navigator.serviceWorker.register('/sw.js?v=61').catch(function(){})}
</script>
</body>
</html>'''


SW_JS = '''const CACHE='bot-admin-v58';
const URLS=['/admin'];
self.addEventListener('install',e=>{self.skipWaiting();e.waitUntil(caches.open(CACHE).then(c=>c.addAll(URLS)))});
self.addEventListener('activate',e=>{e.waitUntil(caches.keys().then(ks=>Promise.all(ks.filter(k=>k!==CACHE).map(k=>caches.delete(k)))).then(()=>self.clients.claim()))});
self.addEventListener('fetch',e=>{const u=e.request.url;if(u.includes('/api/')||u.includes('/health')){e.respondWith(fetch(e.request));return}e.respondWith(fetch(e.request).then(r=>{if(r.ok){const c=r.clone();caches.open(CACHE).then(ca=>ca.put(e.request,c))}return r}).catch(()=>caches.match(e.request)))});'''

MANIFEST_JSON = json.dumps({
    "name": "翻譯Bot 管理後台",
    "short_name": "Bot管理",
    "start_url": "/admin",
    "display": "standalone",
    "background_color": "#0a0a0a",
    "theme_color": "#7c6fef",
    "icons": [
        {"src": "/icon-192.png", "sizes": "192x192", "type": "image/png"},
        {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png"}
    ]
}, ensure_ascii=False)


def commit_storage_to_github(json_data):
    """Auto-commit storage_data.json to GitHub repo."""
    return _commit_file_to_github("storage_data.json", json_data, "Update storage data via admin panel")


def commit_packaging_to_github(json_data):
    """Auto-commit packaging_data.json to GitHub repo."""
    return _commit_file_to_github("packaging_data.json", json_data, "Update packaging data via admin panel")


def _commit_file_to_github(filename, content_str, message="Auto-update", branch="main"):
    """Generic: commit a file to GitHub repo."""
    if not GITHUB_TOKEN:
        logger.warning("No GITHUB_TOKEN, skipping GitHub commit for %s", filename)
        return False
    try:
        api_url = "https://api.github.com/repos/" + GITHUB_REPO + "/contents/" + filename
        req = urllib.request.Request(api_url + "?ref=" + branch, headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json"
        })
        sha = None
        try:
            with urllib.request.urlopen(req, timeout=8) as resp:
                existing = json.loads(resp.read().decode())
                sha = existing.get("sha")
        except urllib.error.HTTPError as e:
            if e.code != 404:
                raise
        content_b64 = base64.b64encode(content_str.encode("utf-8")).decode("utf-8")
        body = {"message": message, "content": content_b64, "branch": branch}
        if sha:
            body["sha"] = sha
        data = json.dumps(body).encode("utf-8")
        req = urllib.request.Request(api_url, data=data, method="PUT", headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json"
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            logger.info("Committed %s to GitHub (%s)", filename, branch)
            return True
    except Exception as e:
        logger.error("GitHub commit %s failed: %s", filename, e)
        return False


def _ensure_data_branch():
    """Create 'data' branch if it doesn't exist."""
    if not GITHUB_TOKEN:
        return
    try:
        # Check if branch exists
        url = "https://api.github.com/repos/" + GITHUB_REPO + "/branches/data"
        req = urllib.request.Request(url, headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json"
        })
        urllib.request.urlopen(req, timeout=5)
        return  # branch exists
    except urllib.error.HTTPError as e:
        if e.code != 404:
            return
    except Exception:
        return
    try:
        # Get main branch SHA
        url = "https://api.github.com/repos/" + GITHUB_REPO + "/git/refs/heads/main"
        req = urllib.request.Request(url, headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json"
        })
        with urllib.request.urlopen(req, timeout=5) as resp:
            main_data = json.loads(resp.read().decode())
            sha = main_data["object"]["sha"]
        # Create data branch
        url = "https://api.github.com/repos/" + GITHUB_REPO + "/git/refs"
        body = json.dumps({"ref": "refs/heads/data", "sha": sha}).encode("utf-8")
        req = urllib.request.Request(url, data=body, method="POST", headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json"
        })
        urllib.request.urlopen(req, timeout=5)
        logger.info("Created 'data' branch on GitHub")
    except Exception as e:
        logger.error("Failed to create data branch: %s", e)


def _load_file_from_github(filename, branch="main"):
    """Load a JSON file from GitHub repo. Returns parsed dict/list or None."""
    if not GITHUB_TOKEN:
        return None
    try:
        api_url = "https://api.github.com/repos/" + GITHUB_REPO + "/contents/" + filename + "?ref=" + branch
        req = urllib.request.Request(api_url, headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json"
        })
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode())
            content = base64.b64decode(data["content"]).decode("utf-8")
            return json.loads(content)
    except Exception as e:
        logger.warning("Load %s from GitHub (%s) failed: %s", filename, branch, e)
        return None


_last_save_time = 0
_save_lock = _threading.Lock()
_save_scheduled = False

def save_settings():
    """Persist all bot settings to GitHub (background, throttled to max once per 30s).
    If throttled, schedules a delayed save so no changes are lost."""
    global _last_save_time, _save_scheduled
    now = time.time()
    with _save_lock:
        remaining = 30 - (now - _last_save_time)
        if remaining > 0:
            # Throttled — schedule a delayed save if not already scheduled
            if not _save_scheduled:
                _save_scheduled = True
                _threading.Timer(remaining + 1, _flush_pending_save).start()
            return
        _last_save_time = now
        _save_scheduled = False
    _threading.Thread(target=_do_save_impl, daemon=True).start()


def _flush_pending_save():
    """Called by timer after throttle window expires."""
    global _save_scheduled
    with _save_lock:
        _save_scheduled = False
    save_settings()


def _do_save_impl():
    global _last_save_time
    try:
        data = {
            "group_settings": group_settings,
            "group_target_lang": group_target_lang,
            "group_img_settings": group_img_settings,
            "group_img_ask_settings": group_img_ask_settings,
            "group_audio_settings": group_audio_settings,
            "group_wo_settings": group_wo_settings,
            "group_cmd_enabled": group_cmd_enabled,
            "group_skip_users": {k: list(v) for k, v in group_skip_users.items()},
            "group_tracking": group_tracking,
            "group_user_names": group_user_names,
            "dm_master_enabled": dm_master_enabled,
            "dm_whitelist": list(dm_whitelist),
            "dm_known_users": dm_known_users,
            "dm_target_lang": dm_target_lang,
            "admin_users": admin_users,
            "bot_stats": bot_stats,
            "extra_customers": extra_names_by_group,
            "group_api_usage": group_api_usage,
            "user_languages": user_languages,
            "welcome_settings": welcome_settings,
            "group_welcome_settings": group_welcome_settings,
            "flex_enabled": flex_enabled,
            "quick_reply_enabled": quick_reply_enabled,
            "silent_mode": silent_mode,
            "sender_name": sender_name,
            "sender_icon": sender_icon,
            "video_ocr_enabled": video_ocr_enabled,
            "location_translate_enabled": location_translate_enabled,
            "group_flex_settings": group_flex_settings,
            "group_qr_settings": group_qr_settings,
            "group_silent_settings": group_silent_settings,
            "group_video_settings": group_video_settings,
            "group_location_settings": group_location_settings,
            "mark_read_enabled": mark_read_enabled,
            "retry_key_enabled": retry_key_enabled,
            "camera_qr_enabled": camera_qr_enabled,
            "clipboard_qr_enabled": clipboard_qr_enabled,
            "group_mark_read_settings": group_mark_read_settings,
            "group_retry_key_settings": group_retry_key_settings,
            "group_camera_qr_settings": group_camera_qr_settings,
            "group_clipboard_qr_settings": group_clipboard_qr_settings,
            "camera_roll_qr_enabled": camera_roll_qr_enabled,
            "location_qr_enabled": location_qr_enabled,
            "group_camera_roll_qr_settings": group_camera_roll_qr_settings,
            "group_location_qr_settings": group_location_qr_settings,
            "translation_tone": translation_tone,
            "translation_tone_custom": translation_tone_custom,
            "translation_temperature": translation_temperature,
            "translation_top_p": translation_top_p,
            "translation_seed": translation_seed,
            "double_check_mode": double_check_mode,
            "double_check_threshold": double_check_threshold,
            "double_check_keywords": double_check_keywords,
            "fewshot_mode": fewshot_mode,
            "logprobs_enabled": logprobs_enabled,
            "confidence_threshold": confidence_threshold,
            "structured_output_enabled": structured_output_enabled,
            "prompt_caching_enabled": prompt_caching_enabled,
            "translation_logging_enabled": translation_logging_enabled,
            "ab_test_enabled": ab_test_enabled,
            "id_zh_cot_enabled": id_zh_cot_enabled,
            "id_zh_cod_enabled": id_zh_cod_enabled,
            "id_zh_pivot_enabled": id_zh_pivot_enabled,
            "id_zh_pivot_threshold": id_zh_pivot_threshold,
            "id_zh_double_translation": id_zh_double_translation,
            "id_preprocessing_enabled": id_preprocessing_enabled,
            "id_preprocessing_nano": id_preprocessing_nano,
            "multi_path_backtrans_enabled": multi_path_backtrans_enabled,
            "multi_path_min_chars": multi_path_min_chars,
            "quality_metrics_enabled": quality_metrics_enabled,
            "preserve_paragraphs_enabled": preserve_paragraphs_enabled,
            "paragraph_split_translate": paragraph_split_translate,
            "paragraph_split_threshold": paragraph_split_threshold,
            "stop_sequences_enabled": stop_sequences_enabled,
            "reasoning_effort": reasoning_effort,
            "send_user_id_to_openai": send_user_id_to_openai,
            "send_metadata_to_openai": send_metadata_to_openai,
            "group_tone_settings": group_tone_settings,
            "model_default": model_default,
            "model_upgrade": model_upgrade,
            "model_threshold": model_threshold,
            "vision_model": VISION_MODEL,
            "user_pictures": user_pictures,
            "pw1_text": pw1_text,
            "pw2_text": pw2_text,
            "scrap_text": scrap_text,
            "custom_translation_examples": custom_translation_examples,
            "forms_data": forms_data,
            "forms_submissions": forms_submissions,
        }
        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        _commit_file_to_github("bot_settings.json", json_str, "Auto-save bot settings", branch="data")
    except Exception as e:
        logger.error("Background save_settings failed: %s", e)
        # Reset timer so next call will retry instead of being throttled
        with _save_lock:
            _last_save_time = 0


def load_settings():
    """Load bot settings from GitHub on startup."""
    global dm_master_enabled, dm_whitelist, dm_known_users, dm_target_lang
    global group_settings, group_target_lang, group_img_settings, group_img_ask_settings, group_audio_settings
    global group_wo_settings, group_skip_users, group_tracking, group_user_names
    global admin_users, bot_stats
    global EXTRA_CUSTOMERS, group_api_usage, extra_names_by_group, user_languages
    global flex_enabled, quick_reply_enabled, silent_mode, welcome_settings, sender_name, sender_icon, user_pictures, video_ocr_enabled, location_translate_enabled
    global group_flex_settings, group_qr_settings, group_silent_settings, group_video_settings, group_location_settings, group_welcome_settings
    global group_mark_read_settings, group_retry_key_settings, group_camera_qr_settings, group_clipboard_qr_settings
    global group_camera_roll_qr_settings, group_location_qr_settings
    global mark_read_enabled, retry_key_enabled, camera_qr_enabled, clipboard_qr_enabled
    global camera_roll_qr_enabled, location_qr_enabled
    global translation_tone, translation_tone_custom, translation_temperature, translation_top_p, translation_seed, double_check_mode, double_check_threshold, double_check_keywords, fewshot_mode, logprobs_enabled, confidence_threshold, structured_output_enabled, prompt_caching_enabled, translation_logging_enabled, ab_test_enabled, stop_sequences_enabled, forbidden_words_zh, forbidden_words_id, reasoning_effort, send_user_id_to_openai, send_metadata_to_openai
    global id_zh_cot_enabled, id_zh_cod_enabled, id_zh_pivot_enabled, id_zh_pivot_threshold, id_zh_double_translation
    global id_preprocessing_enabled, id_preprocessing_nano, multi_path_backtrans_enabled, multi_path_min_chars, quality_metrics_enabled
    global preserve_paragraphs_enabled, paragraph_split_translate, paragraph_split_threshold
    global model_default, model_upgrade, model_threshold
    global VISION_MODEL
    global pw1_text, pw2_text, scrap_text, PACKAGING_LOOKUP, custom_translation_examples
    global forms_data, forms_submissions
    data = _load_file_from_github("bot_settings.json", branch="data")
    if not data:
        logger.info("No bot_settings.json found on GitHub, starting fresh")
        return
    try:
        group_settings.update(data.get("group_settings", {}))
        group_target_lang.update(data.get("group_target_lang", {}))
        group_img_settings.update(data.get("group_img_settings", {}))
        group_img_ask_settings.update(data.get("group_img_ask_settings", {}))
        group_audio_settings.update(data.get("group_audio_settings", {}))
        group_wo_settings.update(data.get("group_wo_settings", {}))
        group_cmd_enabled.update(data.get("group_cmd_enabled", {}))
        for k, v in data.get("group_skip_users", {}).items():
            group_skip_users[k] = set(v)
        group_tracking.update(data.get("group_tracking", {}))
        group_user_names.update(data.get("group_user_names", {}))
        dm_master_enabled = data.get("dm_master_enabled", True)
        dm_whitelist = set(data.get("dm_whitelist", []))
        dm_known_users.update(data.get("dm_known_users", {}))
        dm_target_lang.update(data.get("dm_target_lang", {}))
        admin_users.update(data.get("admin_users", {}))
        bot_stats.update(data.get("bot_stats", {}))
        if "extra_customers" in data:
            ec = data["extra_customers"]
            if isinstance(ec, dict):
                extra_names_by_group.update(ec)
            elif isinstance(ec, list):
                extra_names_by_group["__all__"] = ec
            rebuild_customer_names()
        group_api_usage.update(data.get("group_api_usage", {}))
        user_languages.update(data.get("user_languages", {}))
        if "welcome_settings" in data:
            welcome_settings.update(data.get("welcome_settings", {}))
        if "flex_enabled" in data:
            flex_enabled = data["flex_enabled"]
        if "quick_reply_enabled" in data:
            quick_reply_enabled = data["quick_reply_enabled"]
        if "silent_mode" in data:
            silent_mode = data["silent_mode"]
        if "sender_name" in data:
            sender_name = data["sender_name"]
        if "sender_icon" in data:
            sender_icon = data["sender_icon"]
        if "video_ocr_enabled" in data:
            video_ocr_enabled = data["video_ocr_enabled"]
        if "location_translate_enabled" in data:
            location_translate_enabled = data["location_translate_enabled"]
        group_flex_settings.update(data.get("group_flex_settings", {}))
        group_qr_settings.update(data.get("group_qr_settings", {}))
        group_silent_settings.update(data.get("group_silent_settings", {}))
        group_video_settings.update(data.get("group_video_settings", {}))
        group_location_settings.update(data.get("group_location_settings", {}))
        if "mark_read_enabled" in data:
            mark_read_enabled = data["mark_read_enabled"]
        if "retry_key_enabled" in data:
            retry_key_enabled = data["retry_key_enabled"]
        if "camera_qr_enabled" in data:
            camera_qr_enabled = data["camera_qr_enabled"]
        if "clipboard_qr_enabled" in data:
            clipboard_qr_enabled = data["clipboard_qr_enabled"]
        group_mark_read_settings.update(data.get("group_mark_read_settings", {}))
        group_retry_key_settings.update(data.get("group_retry_key_settings", {}))
        group_camera_qr_settings.update(data.get("group_camera_qr_settings", {}))
        group_clipboard_qr_settings.update(data.get("group_clipboard_qr_settings", {}))
        if "camera_roll_qr_enabled" in data:
            camera_roll_qr_enabled = data["camera_roll_qr_enabled"]
        if "location_qr_enabled" in data:
            location_qr_enabled = data["location_qr_enabled"]
        group_camera_roll_qr_settings.update(data.get("group_camera_roll_qr_settings", {}))
        group_location_qr_settings.update(data.get("group_location_qr_settings", {}))
        group_welcome_settings.update(data.get("group_welcome_settings", {}))
        if "translation_tone" in data:
            translation_tone = data["translation_tone"]
        # v3.2-0426d Batch B: load advanced translation settings from persisted state
        if "translation_temperature" in data:
            try: translation_temperature = float(data["translation_temperature"])
            except: pass
        if "translation_top_p" in data:
            try: translation_top_p = float(data["translation_top_p"])
            except: pass
        if "translation_seed" in data:
            try: translation_seed = int(data["translation_seed"])
            except: pass
        if "double_check_mode" in data:
            v = str(data["double_check_mode"])
            if v in ("off", "smart", "all", "keywords_only"):
                double_check_mode = v
        if "double_check_threshold" in data:
            try: double_check_threshold = float(data["double_check_threshold"])
            except: pass
        if "double_check_keywords" in data:
            double_check_keywords = str(data["double_check_keywords"])
        if "fewshot_mode" in data:
            v = str(data["fewshot_mode"])
            if v in ("system_prompt", "messages"):
                fewshot_mode = v
        if "logprobs_enabled" in data:
            logprobs_enabled = bool(data["logprobs_enabled"])
        if "confidence_threshold" in data:
            try: confidence_threshold = float(data["confidence_threshold"])
            except: pass
        if "structured_output_enabled" in data:
            structured_output_enabled = bool(data["structured_output_enabled"])
        if "prompt_caching_enabled" in data:
            prompt_caching_enabled = bool(data["prompt_caching_enabled"])
        if "translation_logging_enabled" in data:
            translation_logging_enabled = bool(data["translation_logging_enabled"])
        if "ab_test_enabled" in data:
            ab_test_enabled = bool(data["ab_test_enabled"])
        # ★ v3.4 ID→ZH 強化開關
        if "id_zh_cot_enabled" in data:
            id_zh_cot_enabled = bool(data["id_zh_cot_enabled"])
        if "id_zh_cod_enabled" in data:
            id_zh_cod_enabled = bool(data["id_zh_cod_enabled"])
        if "id_zh_pivot_enabled" in data:
            id_zh_pivot_enabled = bool(data["id_zh_pivot_enabled"])
        if "id_zh_pivot_threshold" in data:
            try: id_zh_pivot_threshold = int(data["id_zh_pivot_threshold"])
            except (ValueError, TypeError): pass
        if "id_zh_double_translation" in data:
            id_zh_double_translation = bool(data["id_zh_double_translation"])
        # ★ v3.6 變數 POST 處理
        if "id_preprocessing_enabled" in data:
            id_preprocessing_enabled = bool(data["id_preprocessing_enabled"])
        if "id_preprocessing_nano" in data:
            id_preprocessing_nano = bool(data["id_preprocessing_nano"])
        if "multi_path_backtrans_enabled" in data:
            multi_path_backtrans_enabled = bool(data["multi_path_backtrans_enabled"])
        if "multi_path_min_chars" in data:
            try: multi_path_min_chars = int(data["multi_path_min_chars"])
            except (ValueError, TypeError): pass
        if "quality_metrics_enabled" in data:
            quality_metrics_enabled = bool(data["quality_metrics_enabled"])
        # ★ v3.7 段落保留變數
        if "preserve_paragraphs_enabled" in data:
            preserve_paragraphs_enabled = bool(data["preserve_paragraphs_enabled"])
        if "paragraph_split_translate" in data:
            paragraph_split_translate = bool(data["paragraph_split_translate"])
        if "paragraph_split_threshold" in data:
            try: paragraph_split_threshold = int(data["paragraph_split_threshold"])
            except (ValueError, TypeError): pass
        if "stop_sequences_enabled" in data:
            stop_sequences_enabled = bool(data["stop_sequences_enabled"])
        if "reasoning_effort" in data:
            v = str(data["reasoning_effort"])
            if v in ("low", "medium", "high"):
                reasoning_effort = v
        if "send_user_id_to_openai" in data:
            send_user_id_to_openai = bool(data["send_user_id_to_openai"])
        if "send_metadata_to_openai" in data:
            send_metadata_to_openai = bool(data["send_metadata_to_openai"])
        if "translation_tone_custom" in data:
            translation_tone_custom = data["translation_tone_custom"]
        if "translation_temperature" in data:
            try: translation_temperature = float(data["translation_temperature"])
            except: pass
        if "translation_top_p" in data:
            try: translation_top_p = float(data["translation_top_p"])
            except: pass
        if "translation_seed" in data:
            try: translation_seed = int(data["translation_seed"])
            except: pass
        if "double_check_mode" in data:
            v = str(data["double_check_mode"])
            if v in ("off", "smart", "all", "keywords_only"):
                double_check_mode = v
        if "double_check_threshold" in data:
            try: double_check_threshold = float(data["double_check_threshold"])
            except: pass
        if "double_check_keywords" in data:
            double_check_keywords = str(data["double_check_keywords"])
        if "fewshot_mode" in data:
            v = str(data["fewshot_mode"])
            if v in ("system_prompt", "messages"):
                fewshot_mode = v
        if "logprobs_enabled" in data:
            logprobs_enabled = bool(data["logprobs_enabled"])
        if "confidence_threshold" in data:
            try: confidence_threshold = float(data["confidence_threshold"])
            except: pass
        if "structured_output_enabled" in data:
            structured_output_enabled = bool(data["structured_output_enabled"])
        if "prompt_caching_enabled" in data:
            prompt_caching_enabled = bool(data["prompt_caching_enabled"])
        if "translation_logging_enabled" in data:
            translation_logging_enabled = bool(data["translation_logging_enabled"])
        if "ab_test_enabled" in data:
            ab_test_enabled = bool(data["ab_test_enabled"])
        # ★ v3.4 ID→ZH 強化開關
        if "id_zh_cot_enabled" in data:
            id_zh_cot_enabled = bool(data["id_zh_cot_enabled"])
        if "id_zh_cod_enabled" in data:
            id_zh_cod_enabled = bool(data["id_zh_cod_enabled"])
        if "id_zh_pivot_enabled" in data:
            id_zh_pivot_enabled = bool(data["id_zh_pivot_enabled"])
        if "id_zh_pivot_threshold" in data:
            try: id_zh_pivot_threshold = int(data["id_zh_pivot_threshold"])
            except (ValueError, TypeError): pass
        if "id_zh_double_translation" in data:
            id_zh_double_translation = bool(data["id_zh_double_translation"])
        # ★ v3.6 變數 POST 處理
        if "id_preprocessing_enabled" in data:
            id_preprocessing_enabled = bool(data["id_preprocessing_enabled"])
        if "id_preprocessing_nano" in data:
            id_preprocessing_nano = bool(data["id_preprocessing_nano"])
        if "multi_path_backtrans_enabled" in data:
            multi_path_backtrans_enabled = bool(data["multi_path_backtrans_enabled"])
        if "multi_path_min_chars" in data:
            try: multi_path_min_chars = int(data["multi_path_min_chars"])
            except (ValueError, TypeError): pass
        if "quality_metrics_enabled" in data:
            quality_metrics_enabled = bool(data["quality_metrics_enabled"])
        # ★ v3.7 段落保留變數
        if "preserve_paragraphs_enabled" in data:
            preserve_paragraphs_enabled = bool(data["preserve_paragraphs_enabled"])
        if "paragraph_split_translate" in data:
            paragraph_split_translate = bool(data["paragraph_split_translate"])
        if "paragraph_split_threshold" in data:
            try: paragraph_split_threshold = int(data["paragraph_split_threshold"])
            except (ValueError, TypeError): pass
        if "stop_sequences_enabled" in data:
            stop_sequences_enabled = bool(data["stop_sequences_enabled"])
        if "reasoning_effort" in data:
            v = str(data["reasoning_effort"])
            if v in ("low", "medium", "high"):
                reasoning_effort = v
        if "send_user_id_to_openai" in data:
            send_user_id_to_openai = bool(data["send_user_id_to_openai"])
        if "send_metadata_to_openai" in data:
            send_metadata_to_openai = bool(data["send_metadata_to_openai"])
        group_tone_settings.update(data.get("group_tone_settings", {}))
        if "model_default" in data:
            model_default = data["model_default"]
        if "model_upgrade" in data:
            model_upgrade = data["model_upgrade"]
        if "model_threshold" in data:
            model_threshold = int(data["model_threshold"])
        # v3.9: vision (照片分析) model
        if "vision_model" in data:
            VISION_MODEL = str(data["vision_model"])
        user_pictures.update(data.get("user_pictures", {}))
        if "pw1_text" in data:
            pw1_text = data["pw1_text"]
        if "pw2_text" in data:
            pw2_text = data["pw2_text"]
        if "scrap_text" in data:
            scrap_text = data["scrap_text"]
        if "custom_translation_examples" in data:
            custom_translation_examples = data["custom_translation_examples"]
        if "forms_data" in data:
            forms_data = data["forms_data"]
        if "forms_submissions" in data:
            forms_submissions = data["forms_submissions"]
        logger.info("Loaded bot settings from GitHub: %d groups, %d DM users, %d protected names, %d custom examples, %d forms",
                     len(group_tracking), len(dm_known_users), len(EXTRA_CUSTOMERS), len(custom_translation_examples), len(forms_data))
    except Exception as e:
        logger.error("Error loading bot settings: %s", e)


# Ensure data branch exists and load settings on startup
try:
    _ensure_data_branch()
    load_settings()
except Exception as e:
    logger.error("Startup settings load failed (non-fatal): %s", e)




def check_admin_key():
    key = request.headers.get("X-Admin-Key", "")
    return key == ADMIN_KEY


def check_manager_access(required_tab=None):
    """Check if request is from super admin or authorized manager."""
    if check_admin_key():
        return True
    uid = request.headers.get("X-Manager-Id", "")
    if not uid or uid not in admin_users:
        return False
    if not admin_users[uid].get("is_admin"):
        return False
    if required_tab:
        allowed = admin_users[uid].get("allowed_tabs", [])
        return required_tab in allowed
    return True


@app.route("/admin")
def admin_page():
    html = ADMIN_HTML.replace("__GOOGLE_CLIENT_ID__", GOOGLE_CLIENT_ID)
    resp = app.response_class(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.route("/debug")
def debug_page():
    """Minimal debug page - no SW, no cache."""
    html = '''<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Debug</title>
<style>body{background:#000;color:#0f0;font:13px monospace;padding:12px}input,button{font:14px sans-serif;padding:8px 12px;margin:4px;border-radius:6px;border:1px solid #555}input{background:#111;color:#fff;width:200px}button{background:#7c6fef;color:#fff;border:none;cursor:pointer}#log{white-space:pre-wrap;margin-top:12px}</style>
</head><body>
<h3>🔧 Bot Debug Panel</h3>
<div>
<input id="pw" type="password" placeholder="管理密碼">
<button onclick="go()">測試登入</button>
<button onclick="testHealth()">Health</button>
<button onclick="clearSW()">清除SW快取</button>
</div>
<div id="log">等待操作...\n</div>
<script>
const L=document.getElementById('log');
const API=window.location.origin+'/api/admin';
function log(s){L.textContent+=new Date().toLocaleTimeString()+' '+s+'\\n';L.scrollTop=L.scrollHeight}

async function testHealth(){
  log('>>> GET /health');
  try{
    const r=await fetch('/health');
    const d=await r.json();
    log('<<< '+r.status+' '+JSON.stringify(d));
  }catch(e){log('!!! '+e.message)}
}

async function go(){
  const key=document.getElementById('pw').value;
  if(!key){log('請輸入密碼');return}
  log('>>> GET /api/admin/status');
  try{
    const r=await fetch(API+'/status',{headers:{'X-Admin-Key':key}});
    log('<<< status: '+r.status);
    const d=await r.json();
    log('<<< body: '+JSON.stringify(d));
    if(d.ok){
      log('✅ 登入成功! 測試其他API...');
      log('>>> GET /api/admin/stats');
      const r2=await fetch(API+'/stats',{headers:{'X-Admin-Key':key}});
      log('<<< stats: '+r2.status);
      const d2=await r2.json();
      log('<<< '+JSON.stringify(d2).substring(0,300));
      log('>>> GET /api/admin/groups');
      const r3=await fetch(API+'/groups',{headers:{'X-Admin-Key':key}});
      log('<<< groups: '+r3.status);
      const d3=await r3.json();
      log('<<< '+JSON.stringify(d3).substring(0,300));
    }
  }catch(e){log('!!! 錯誤: '+e.message)}
}

async function clearSW(){
  log('清除所有 Service Worker...');
  if('serviceWorker' in navigator){
    const regs=await navigator.serviceWorker.getRegistrations();
    log('找到 '+regs.length+' 個 SW');
    for(const r of regs){
      await r.unregister();
      log('已移除: '+r.scope);
    }
    log('清除快取...');
    const keys=await caches.keys();
    for(const k of keys){
      await caches.delete(k);
      log('已刪除快取: '+k);
    }
    log('✅ 全部清除完成! 現在可以回 /admin 了');
  }else{
    log('此瀏覽器不支援 SW');
  }
}

log('Debug頁面載入完成');
log('API: '+API);
</script></body></html>'''
    resp = app.response_class(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["Service-Worker-Allowed"] = "/"
    return resp


@app.route("/manifest.json")
def manifest():
    return app.response_class(MANIFEST_JSON, mimetype="application/manifest+json")


@app.route("/sw.js")
def service_worker():
    resp = app.response_class(SW_JS, mimetype="application/javascript")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp


@app.route("/icon-192.png")
@app.route("/icon-512.png")
def admin_icon():
    # Generate a simple green circle PNG as icon
    import struct, zlib
    size = 192 if "192" in request.path else 512
    # 1x1 green pixel PNG, browser will scale
    png = (b'\\x89PNG\\r\\n\\x1a\\n'
           + struct.pack('>I', 13) + b'IHDR' + struct.pack('>IIBBBBB', 1, 1, 8, 2, 0, 0, 0)
           + struct.pack('>I', zlib.crc32(b'IHDR' + struct.pack('>IIBBBBB', 1, 1, 8, 2, 0, 0, 0)) & 0xffffffff)
           + struct.pack('>I', 12) + b'IDAT' + zlib.compress(b'\\x00\\x06\\xc7\\x55')
           + struct.pack('>I', zlib.crc32(b'IDAT' + zlib.compress(b'\\x00\\x06\\xc7\\x55')) & 0xffffffff)
           + struct.pack('>I', 0) + b'IEND' + struct.pack('>I', zlib.crc32(b'IEND') & 0xffffffff))
    return app.response_class(png, mimetype="image/png")


# ─── Admin API ──────────────────────────────────────────

@app.route("/api/admin/status")
def api_admin_status():
    if check_admin_key():
        return jsonify({"ok": True, "role": "super"})
    if not check_manager_access():
        return jsonify({"error": "forbidden"}), 403
    return jsonify({"ok": True, "role": "manager"})


@app.route("/api/admin/google-config")
def api_admin_google_config():
    """Public debug endpoint - check if Google Client ID is set."""
    return jsonify({
        "client_id_set": bool(GOOGLE_CLIENT_ID),
        "client_id_length": len(GOOGLE_CLIENT_ID),
        "client_id_first10": GOOGLE_CLIENT_ID[:10] if GOOGLE_CLIENT_ID else "",
        "client_id_last10": GOOGLE_CLIENT_ID[-10:] if GOOGLE_CLIENT_ID else "",
        "secret_set": bool(GOOGLE_CLIENT_SECRET),
    })


@app.route("/google-test")
def google_test_page():
    """Visual test page for Google Sign In."""
    html = """<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Google Test</title><script src="https://accounts.google.com/gsi/client" async defer></script>
<style>body{font-family:sans-serif;padding:20px;background:#111;color:#eee}
pre{background:#000;padding:10px;border-radius:6px;word-break:break-all;white-space:pre-wrap;font-size:12px}
#btn{margin:20px 0}</style></head><body>
<h2>Google Sign In Test</h2>
<div>Client ID being used:</div>
<pre id="cid"></pre>
<div>Length: <span id="cidlen"></span></div>
<div id="btn"></div>
<h3>Result:</h3>
<pre id="result">waiting...</pre>
<script>
var CID='""" + GOOGLE_CLIENT_ID + """';
document.getElementById('cid').textContent=CID;
document.getElementById('cidlen').textContent=CID.length;
function handle(response){
  document.getElementById('result').textContent='SUCCESS!\\ncredential length: '+response.credential.length;
}
function init(){
  if(typeof google==='undefined'||!google.accounts){setTimeout(init,500);return}
  try{
    google.accounts.id.initialize({client_id:CID,callback:handle});
    google.accounts.id.renderButton(document.getElementById('btn'),{theme:'outline',size:'large'});
    document.getElementById('result').textContent='Button rendered, click it to test.';
  }catch(e){document.getElementById('result').textContent='ERROR: '+e.message}
}
document.addEventListener('DOMContentLoaded',init);
</script></body></html>"""
    return app.response_class(html, mimetype="text/html")


@app.route("/api/admin/manager-login", methods=["POST"])
def api_admin_manager_login():
    """Login as manager using Google ID token."""
    data = request.get_json(force=True)
    credential = data.get("credential", "")
    if not credential:
        return jsonify({"error": "缺少 Google 憑證"}), 400
    # Verify Google ID token
    try:
        # Decode JWT without verification first to get email (we verify with Google tokeninfo)
        import base64
        parts = credential.split(".")
        if len(parts) < 2:
            return jsonify({"error": "無效的憑證格式"}), 400
        payload = parts[1]
        payload += "=" * (4 - len(payload) % 4)  # pad base64
        token_data = json.loads(base64.urlsafe_b64decode(payload).decode("utf-8"))
        email = token_data.get("email", "").lower().strip()
        if not email:
            return jsonify({"error": "無法取得 Google 信箱"}), 400
        # Verify token with Google
        verify_url = "https://oauth2.googleapis.com/tokeninfo?id_token=" + credential
        req = urllib.request.Request(verify_url)
        with urllib.request.urlopen(req, timeout=5) as resp:
            verify_data = json.loads(resp.read().decode())
            if verify_data.get("aud") != GOOGLE_CLIENT_ID:
                return jsonify({"error": "憑證驗證失敗"}), 403
    except Exception as e:
        logger.error("Google token verify failed: %s", e)
        return jsonify({"error": "Google 驗證失敗: " + str(e)}), 400
    # Find user by email
    found_uid = None
    for uid, info in admin_users.items():
        if info.get("google_email", "").lower().strip() == email:
            found_uid = uid
            break
    if not found_uid:
        return jsonify({"error": "此 Google 帳號未綁定任何管理員 (" + email + ")"}), 403
    if not admin_users[found_uid].get("is_admin"):
        return jsonify({"error": "此帳號沒有管理員權限"}), 403
    allowed = admin_users[found_uid].get("allowed_tabs", [])
    if not allowed:
        return jsonify({"error": "尚未設定可用功能，請聯繫超級管理員"}), 403
    return jsonify({"ok": True, "role": "manager", "tabs": allowed, "email": email, "user_id": found_uid})


@app.route("/api/admin/groups", methods=["GET"])
def api_admin_groups():
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    groups = []
    # Merge from group_tracking + group_settings
    all_gids = set(group_tracking.keys()) | set(group_settings.keys()) | set(group_target_lang.keys())
    for gid in all_gids:
        info = group_tracking.get(gid, {})
        skip_count = len(group_skip_users.get(gid, set()))
        groups.append({
            "id": gid,
            "name": info.get("name", ""),
            "translation_on": group_settings.get(gid, True),
            "target_lang": group_target_lang.get(gid, "id"),
            "image_on": group_img_settings.get(gid, True),
            "image_ask_mode": group_img_ask_settings.get(gid, False),
            "voice_on": group_audio_settings.get(gid, True),
            "work_order_on": group_wo_settings.get(gid, True),
            "cmd_enabled": {k: is_cmd_enabled(gid, k) for k, _, _, _ in CMD_DEFS},
            "skip_count": skip_count,
            "cost_twd": calc_group_cost_twd(gid),
            "member_count": get_group_member_count(gid),
        })
    groups.sort(key=lambda x: x["name"] or x["id"])
    return jsonify({"groups": groups})


@app.route("/api/admin/groups/leave", methods=["POST"])
def api_admin_leave_group():
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    gid = data.get("group_id", "")
    if not gid:
        return jsonify({"error": "missing group_id"}), 400
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.leave_group(gid)
    except Exception as e:
        logger.warning("Leave group failed: %s", e)
        # Try room
        try:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.leave_room(gid)
        except Exception:
            pass
    # Clean up local data
    group_tracking.pop(gid, None)
    group_settings.pop(gid, None)
    group_target_lang.pop(gid, None)
    group_img_settings.pop(gid, None)
    group_audio_settings.pop(gid, None)
    group_wo_settings.pop(gid, None)
    group_skip_users.pop(gid, None)
    group_user_names.pop(gid, None)
    save_settings()
    return jsonify({"message": "已退出群組"})


@app.route("/api/admin/stats")
def api_admin_stats():
    if not check_manager_access():
        return jsonify({"error": "forbidden"}), 403
    uptime = time.time() - bot_start_time
    # Calculate estimated cost (GPT-4o-mini pricing)
    tp = bot_stats.get("tokens_prompt", 0)
    tc = bot_stats.get("tokens_completion", 0)
    cost = (tp * 0.00000015) + (tc * 0.0000006)
    return jsonify({
        "uptime_seconds": int(uptime),
        "text_translations": bot_stats.get("text_translations", 0),
        "image_translations": bot_stats.get("image_translations", 0),
        "voice_translations": bot_stats.get("voice_translations", 0),
        "work_order_detections": bot_stats.get("work_order_detections", 0),
        "commands": bot_stats.get("commands", 0),
        "customers": len(STORAGE_LOOKUP),
        "groups": len(set(group_tracking.keys()) | set(group_settings.keys())),
        "dm_users": len(dm_known_users),
        "tokens_prompt": tp,
        "tokens_completion": tc,
        "tokens_total": tp + tc,
        "estimated_cost_usd": round(cost, 4),
        "followers": bot_stats.get("followers", 0),
        "unfollowers": bot_stats.get("unfollowers", 0),
        "line_quota": get_line_quota(),
        # Feature settings
        "welcome_enabled": welcome_settings.get("enabled", True),
        "welcome_text_zh": welcome_settings.get("text_zh", ""),
        "welcome_text_id": welcome_settings.get("text_id", ""),
        "flex_enabled": flex_enabled,
        "quick_reply_enabled": quick_reply_enabled,
        "silent_mode": silent_mode,
    })


@app.route("/api/admin/translation-log", methods=["GET", "POST", "DELETE"])
def api_translation_log():
    """Get/manage translation log. POST to mark-wrong. DELETE to clear."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    if request.method == "GET":
        # Optional filters
        # v3.9.30c B19: 用 _safe_int 防 ?limit=abc 炸 500
        limit = _safe_int(request.args.get("limit", 100), default=100, min_val=1, max_val=10000)
        only_warned = request.args.get("only_warned") == "1"
        only_wrong = request.args.get("only_wrong") == "1"
        items = list(translation_log)
        if only_warned:
            items = [x for x in items if x.get("warned") or x.get("factory_warning") or x.get("tgt", "").startswith("⚠️")]
        if only_wrong:
            items = [x for x in items if x.get("marked_wrong")]
        items = items[-limit:]
        items.reverse()
        # Stats
        total = len(translation_log)
        warned = sum(1 for x in translation_log if x.get("warned") or x.get("factory_warning") or x.get("tgt", "").startswith("⚠️"))
        wrong = sum(1 for x in translation_log if x.get("marked_wrong"))
        avg_conf = (sum(x.get("confidence") or 1.0 for x in translation_log) / total) if total else 0
        return jsonify({
            "items": items,
            "stats": {
                "total": total,
                "warned": warned,
                "wrong": wrong,
                "avg_confidence": round(avg_conf, 3),
            }
        })
    if request.method == "POST":
        # Mark a log entry as wrong, optionally add to BUILTIN-style override list
        data = request.get_json() or {}
        ts = data.get("ts")
        if ts is None:
            return jsonify({"error": "missing ts"}), 400
        for entry in translation_log:
            if entry.get("ts") == ts:
                entry["marked_wrong"] = True
                entry["correct_translation"] = data.get("correct_translation", "")
                # Optionally add to custom_translation_examples
                if data.get("add_to_examples") and data.get("correct_translation"):
                    src_lang = entry.get("src_lang", "zh")
                    direction = "zh2id" if src_lang == "zh" else "id2zh"
                    new_ex = {
                        "zh": entry.get("src", "") if src_lang == "zh" else data.get("correct_translation", ""),
                        "id": data.get("correct_translation", "") if src_lang == "zh" else entry.get("src", ""),
                        "dir": direction
                    }
                    custom_translation_examples.append(new_ex)
                    if len(custom_translation_examples) > CUSTOM_EXAMPLES_MAX:
                        custom_translation_examples[:] = custom_translation_examples[-CUSTOM_EXAMPLES_MAX:]
                    _save_examples_to_disk()  # v3.9.29: 修補隱性 bug
                _save_translation_log_to_disk()
                return jsonify({"ok": True})
        return jsonify({"error": "not found"}), 404
    if request.method == "DELETE":
        translation_log.clear()
        _save_translation_log_to_disk()
        return jsonify({"ok": True})


@app.route("/api/admin/tlog-diag", methods=["GET"])
def api_tlog_diag():
    """v3.9.33 翻譯品質儀表板診斷端點。
    回傳 translation_log 的記憶體狀態 + 磁碟狀態,協助判斷儀表板為什麼空白。
    用法:GET /api/admin/tlog-diag?key=ADMIN_KEY
    """
    # 接受 query string token(方便手機瀏覽器直接訪問),或標準 admin/manager header
    _key = request.args.get("key", "")
    if _key != ADMIN_KEY and not check_manager_access():
        return jsonify({"error": "forbidden", "hint": "use ?key=YOUR_ADMIN_KEY in URL"}), 403
    import os as _os
    diag = {
        "logging_enabled": translation_logging_enabled,
        "memory_count": len(translation_log),
        "disk_path": TRANSLATION_LOG_FILE,
        "disk_exists": _os.path.exists(TRANSLATION_LOG_FILE) if TRANSLATION_LOG_FILE else False,
        "disk_size_bytes": (_os.path.getsize(TRANSLATION_LOG_FILE)
                            if TRANSLATION_LOG_FILE and _os.path.exists(TRANSLATION_LOG_FILE) else 0),
        "disk_writable": False,
        "now_ts": int(time.time()),
        "now_readable": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
    }
    # 檢測寫入權限
    try:
        folder = _os.path.dirname(_os.path.abspath(TRANSLATION_LOG_FILE)) if TRANSLATION_LOG_FILE else "."
        diag["disk_writable"] = _os.access(folder or ".", _os.W_OK)
        diag["disk_folder"] = folder
    except Exception as _e:
        diag["disk_check_error"] = str(_e)
    # 顯示記憶體中最近 5 筆的時間戳和摘要
    recent = []
    for x in translation_log[-5:]:
        recent.append({
            "ts": x.get("ts"),
            "ts_readable": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(x.get("ts", 0))) if x.get("ts") else "N/A",
            "src_preview": (x.get("src") or "")[:40],
            "tgt_preview": (x.get("tgt") or "")[:40],
            "group_id": (x.get("group_id") or "")[:20],
            "model": x.get("model"),
        })
    diag["recent_5_in_memory"] = recent
    # 檢查 7 天內有多少筆
    cutoff_7d = int(time.time()) - 7 * 86400
    diag["count_last_7_days"] = sum(1 for x in translation_log if x.get("ts", 0) >= cutoff_7d)
    cutoff_1d = int(time.time()) - 86400
    diag["count_last_1_day"] = sum(1 for x in translation_log if x.get("ts", 0) >= cutoff_1d)
    return jsonify(diag)


@app.route("/api/admin/translation-stats", methods=["GET"])
def api_translation_stats():
    """v3.6 翻譯品質監控儀表板 - 五大指標 + 各群組分數"""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    
    # 時間範圍:預設過去 7 天
    # v3.9.30c B19: 用 _safe_int 防炸
    days = _safe_int(request.args.get("days", 7), default=7, min_val=1, max_val=365)
    cutoff_ts = int(time.time()) - days * 86400
    
    items = [x for x in translation_log if x.get("ts", 0) >= cutoff_ts]
    total = len(items)
    
    if total == 0:
        return jsonify({
            "period_days": days,
            "total": 0,
            "metrics": {},
            "by_group": [],
            "by_lang": [],
            "issues": [],
        })
    
    # ===== 指標 1:平均譯文/原文長度比 =====
    length_ratios = [x.get("length_ratio") for x in items if x.get("length_ratio") is not None]
    avg_length_ratio = round(sum(length_ratios) / len(length_ratios), 3) if length_ratios else None
    
    # ===== 指標 2:catastrophic_compression 觸發次數 =====
    catastrophic = sum(
        1 for x in items 
        if x.get("rtc_fail_reason", "").startswith("catastrophic_compression")
    )
    
    # ===== 指標 3:missing_terms (self-check 抓到 GPT 說謊) =====
    missing_terms = sum(
        1 for x in items 
        if "missing_terms" in str(x.get("struct_self_check", ""))
    )
    
    # ===== 指標 4:claimed_incomplete (GPT 自首沒翻完) =====
    claimed_incomplete = sum(
        1 for x in items 
        if "claimed_incomplete" in str(x.get("struct_self_check", ""))
    )
    
    # ===== 指標 5:多路徑反譯失敗次數 =====
    multi_path_fails = sum(1 for x in items if x.get("multi_path_fail"))
    
    # ===== 各群組品質分數 =====
    by_group = {}
    for x in items:
        gid = x.get("group_id", "unknown") or "unknown"
        if gid not in by_group:
            by_group[gid] = {
                "group_id": gid, "total": 0, "warned": 0, "wrong": 0,
                "avg_confidence_sum": 0.0, "confidence_count": 0,
            }
        g = by_group[gid]
        g["total"] += 1
        if x.get("warned") or x.get("factory_warning") or (x.get("tgt", "") or "").startswith("⚠️"):
            g["warned"] += 1
        if x.get("marked_wrong"):
            g["wrong"] += 1
        if x.get("confidence") is not None:
            g["avg_confidence_sum"] += x.get("confidence")
            g["confidence_count"] += 1
    
    by_group_list = []
    for gid, g in by_group.items():
        avg_conf = (g["avg_confidence_sum"] / g["confidence_count"]) if g["confidence_count"] > 0 else None
        # 品質分數 = (1 - warned率) * 100,被標錯扣分
        quality_score = round(
            ((1 - g["warned"] / g["total"]) * 100) - (g["wrong"] / g["total"] * 50)
            if g["total"] > 0 else 0,
            1
        )
        by_group_list.append({
            "group_id": gid,
            "total": g["total"],
            "warned": g["warned"],
            "wrong": g["wrong"],
            "avg_confidence": round(avg_conf, 3) if avg_conf is not None else None,
            "quality_score": max(0, quality_score),
        })
    by_group_list.sort(key=lambda x: x["quality_score"])  # 最差排前面
    
    # ===== 各語言對統計 =====
    by_lang = {}
    for x in items:
        key = f"{x.get('src_lang', '?')}→{x.get('tgt_lang', '?')}"
        if key not in by_lang:
            by_lang[key] = {"pair": key, "total": 0, "warned": 0, 
                            "avg_length_ratio_sum": 0.0, "lr_count": 0}
        b = by_lang[key]
        b["total"] += 1
        if x.get("warned") or (x.get("tgt", "") or "").startswith("⚠️"):
            b["warned"] += 1
        if x.get("length_ratio") is not None:
            b["avg_length_ratio_sum"] += x.get("length_ratio")
            b["lr_count"] += 1
    
    by_lang_list = []
    for key, b in by_lang.items():
        avg_lr = (b["avg_length_ratio_sum"] / b["lr_count"]) if b["lr_count"] > 0 else None
        by_lang_list.append({
            "pair": key,
            "total": b["total"],
            "warned": b["warned"],
            "warned_rate": round(b["warned"] / b["total"] * 100, 1) if b["total"] > 0 else 0,
            "avg_length_ratio": round(avg_lr, 3) if avg_lr is not None else None,
        })
    by_lang_list.sort(key=lambda x: -x["total"])
    
    # ===== 最近的問題訊息(供調試)=====
    recent_issues = []
    for x in reversed(items):
        if len(recent_issues) >= 10:
            break
        is_issue = (
            x.get("rtc_fail_reason") or x.get("struct_self_check") 
            or x.get("multi_path_fail") or x.get("marked_wrong")
        )
        if is_issue:
            recent_issues.append({
                "ts": x.get("ts"),
                "src": (x.get("src", "") or "")[:100],
                "tgt": (x.get("tgt", "") or "")[:100],
                "src_lang": x.get("src_lang"),
                "tgt_lang": x.get("tgt_lang"),
                "rtc_fail_reason": x.get("rtc_fail_reason"),
                "struct_self_check": x.get("struct_self_check"),
                "multi_path_fail": x.get("multi_path_fail"),
                "marked_wrong": x.get("marked_wrong"),
                "length_ratio": x.get("length_ratio"),
            })
    
    return jsonify({
        "period_days": days,
        "total": total,
        "metrics": {
            "avg_length_ratio": avg_length_ratio,
            "catastrophic_compression_count": catastrophic,
            "missing_terms_count": missing_terms,
            "claimed_incomplete_count": claimed_incomplete,
            "multi_path_fails": multi_path_fails,
            "warned_total": sum(1 for x in items if x.get("warned") or (x.get("tgt", "") or "").startswith("⚠️")),
            "wrong_total": sum(1 for x in items if x.get("marked_wrong")),
        },
        "by_group": by_group_list,
        "by_lang": by_lang_list,
        "issues": recent_issues,
    })


@app.route("/api/admin/features", methods=["GET", "POST"])
def api_admin_features():
    """Get/set feature settings. Pass group_id for per-group; omit for global defaults."""
    global flex_enabled, quick_reply_enabled, silent_mode, welcome_settings
    global sender_name, sender_icon, video_ocr_enabled, location_translate_enabled
    global translation_tone, translation_tone_custom, translation_temperature, translation_top_p, translation_seed, double_check_mode, double_check_threshold, double_check_keywords, fewshot_mode, logprobs_enabled, confidence_threshold, structured_output_enabled, prompt_caching_enabled, translation_logging_enabled, ab_test_enabled, stop_sequences_enabled, forbidden_words_zh, forbidden_words_id, reasoning_effort, send_user_id_to_openai, send_metadata_to_openai
    global id_zh_cot_enabled, id_zh_cod_enabled, id_zh_pivot_enabled, id_zh_pivot_threshold, id_zh_double_translation
    global id_preprocessing_enabled, id_preprocessing_nano, multi_path_backtrans_enabled, multi_path_min_chars, quality_metrics_enabled
    global preserve_paragraphs_enabled, paragraph_split_translate, paragraph_split_threshold
    global mark_read_enabled, retry_key_enabled, camera_qr_enabled, clipboard_qr_enabled
    global camera_roll_qr_enabled, location_qr_enabled
    global model_default, model_upgrade, model_threshold
    global VISION_MODEL
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    gid = request.args.get("group_id", "") if request.method == "GET" else (request.get_json() or {}).get("group_id", "")
    if request.method == "POST":
        data = request.get_json() or {}
        if gid:
            # Per-group settings
            _feat_map = {
                "flex_enabled": group_flex_settings,
                "quick_reply_enabled": group_qr_settings,
                "silent_mode": group_silent_settings,
                "video_ocr_enabled": group_video_settings,
                "location_translate_enabled": group_location_settings,
                "mark_read_enabled": group_mark_read_settings,
                "retry_key_enabled": group_retry_key_settings,
                "camera_qr_enabled": group_camera_qr_settings,
                "clipboard_qr_enabled": group_clipboard_qr_settings,
                "camera_roll_qr_enabled": group_camera_roll_qr_settings,
                "location_qr_enabled": group_location_qr_settings,
            }
            for key, d in _feat_map.items():
                if key in data:
                    d[gid] = bool(data[key])
            # Per-group welcome
            if any(k in data for k in ("welcome_enabled", "welcome_text_zh", "welcome_text_id")):
                if gid not in group_welcome_settings:
                    group_welcome_settings[gid] = {}
                if "welcome_enabled" in data:
                    group_welcome_settings[gid]["enabled"] = bool(data["welcome_enabled"])
                if "welcome_text_zh" in data:
                    group_welcome_settings[gid]["text_zh"] = str(data["welcome_text_zh"])
                if "welcome_text_id" in data:
                    group_welcome_settings[gid]["text_id"] = str(data["welcome_text_id"])
            # Per-group tone
            if "translation_tone" in data:
                if gid not in group_tone_settings:
                    group_tone_settings[gid] = {}
                group_tone_settings[gid]["tone"] = str(data["translation_tone"])
            if "translation_tone_custom" in data:
                if gid not in group_tone_settings:
                    group_tone_settings[gid] = {}
                group_tone_settings[gid]["custom"] = str(data["translation_tone_custom"])
        else:
            # Global defaults
            if "welcome_enabled" in data:
                welcome_settings["enabled"] = bool(data["welcome_enabled"])
            if "welcome_text_zh" in data:
                welcome_settings["text_zh"] = str(data["welcome_text_zh"])
            if "welcome_text_id" in data:
                welcome_settings["text_id"] = str(data["welcome_text_id"])
            if "flex_enabled" in data:
                flex_enabled = bool(data["flex_enabled"])
            if "quick_reply_enabled" in data:
                quick_reply_enabled = bool(data["quick_reply_enabled"])
            if "silent_mode" in data:
                silent_mode = bool(data["silent_mode"])
            if "video_ocr_enabled" in data:
                video_ocr_enabled = bool(data["video_ocr_enabled"])
            if "location_translate_enabled" in data:
                location_translate_enabled = bool(data["location_translate_enabled"])
            if "mark_read_enabled" in data:
                mark_read_enabled = bool(data["mark_read_enabled"])
            if "retry_key_enabled" in data:
                retry_key_enabled = bool(data["retry_key_enabled"])
            if "camera_qr_enabled" in data:
                camera_qr_enabled = bool(data["camera_qr_enabled"])
            if "clipboard_qr_enabled" in data:
                clipboard_qr_enabled = bool(data["clipboard_qr_enabled"])
            if "camera_roll_qr_enabled" in data:
                camera_roll_qr_enabled = bool(data["camera_roll_qr_enabled"])
            if "location_qr_enabled" in data:
                location_qr_enabled = bool(data["location_qr_enabled"])
            if "translation_tone" in data:
                translation_tone = str(data["translation_tone"])
            if "translation_tone_custom" in data:
                translation_tone_custom = str(data["translation_tone_custom"])
            if "model_default" in data:
                model_default = str(data["model_default"])
            if "model_upgrade" in data:
                model_upgrade = str(data["model_upgrade"])
            if "model_threshold" in data:
                model_threshold = int(data["model_threshold"])
            # v3.9: vision (照片分析) model selectable from admin panel
            if "vision_model" in data:
                VISION_MODEL = str(data["vision_model"])
        # Sender settings are always global
        if "sender_name" in data:
            sender_name = str(data["sender_name"])[:20]
        if "sender_icon" in data:
            sender_icon = str(data["sender_icon"])
        save_settings()
        return jsonify({"ok": True})
    # GET - return settings for specific group or global
    if gid:
        ws = get_group_welcome(gid)
        return jsonify({
            "group_id": gid,
            "welcome_enabled": ws.get("enabled", True),
            "welcome_text_zh": ws.get("text_zh", ""),
            "welcome_text_id": ws.get("text_id", ""),
            "flex_enabled": get_group_feature(gid, 'flex'),
            "quick_reply_enabled": get_group_feature(gid, 'quick_reply'),
            "silent_mode": get_group_feature(gid, 'silent'),
            "video_ocr_enabled": get_group_feature(gid, 'video_ocr'),
            "location_translate_enabled": get_group_feature(gid, 'location'),
            "mark_read_enabled": get_group_feature(gid, 'mark_read'),
            "retry_key_enabled": get_group_feature(gid, 'retry_key'),
            "camera_qr_enabled": get_group_feature(gid, 'camera_qr'),
            "clipboard_qr_enabled": get_group_feature(gid, 'clipboard_qr'),
            "camera_roll_qr_enabled": get_group_feature(gid, 'camera_roll_qr'),
            "location_qr_enabled": get_group_feature(gid, 'location_qr'),
            "translation_tone": get_group_tone(gid)[0],
            "translation_tone_custom": get_group_tone(gid)[1],
            "sender_name": sender_name,
            "sender_icon": sender_icon,
            "bot_info": get_bot_info(),
            # Include global defaults for reference
            "global_defaults": {
                "welcome_enabled": welcome_settings.get("enabled", True),
                "flex_enabled": flex_enabled,
                "quick_reply_enabled": quick_reply_enabled,
                "silent_mode": silent_mode,
                "video_ocr_enabled": video_ocr_enabled,
                "location_translate_enabled": location_translate_enabled,
                "mark_read_enabled": mark_read_enabled,
                "retry_key_enabled": retry_key_enabled,
                "camera_qr_enabled": camera_qr_enabled,
                "clipboard_qr_enabled": clipboard_qr_enabled,
                "camera_roll_qr_enabled": camera_roll_qr_enabled,
                "location_qr_enabled": location_qr_enabled,
            },
            "is_customized": gid in group_flex_settings or gid in group_qr_settings or gid in group_silent_settings or gid in group_video_settings or gid in group_location_settings or gid in group_welcome_settings or gid in group_tone_settings or gid in group_mark_read_settings or gid in group_retry_key_settings or gid in group_camera_qr_settings or gid in group_clipboard_qr_settings or gid in group_camera_roll_qr_settings or gid in group_location_qr_settings,
        })
    return jsonify({
        "welcome_enabled": welcome_settings.get("enabled", True),
        "welcome_text_zh": welcome_settings.get("text_zh", ""),
        "welcome_text_id": welcome_settings.get("text_id", ""),
        "flex_enabled": flex_enabled,
        "quick_reply_enabled": quick_reply_enabled,
        "silent_mode": silent_mode,
        "video_ocr_enabled": video_ocr_enabled,
        "location_translate_enabled": location_translate_enabled,
        "mark_read_enabled": mark_read_enabled,
        "retry_key_enabled": retry_key_enabled,
        "camera_qr_enabled": camera_qr_enabled,
        "clipboard_qr_enabled": clipboard_qr_enabled,
        "camera_roll_qr_enabled": camera_roll_qr_enabled,
        "location_qr_enabled": location_qr_enabled,
        "translation_tone": translation_tone,
        "translation_tone_custom": translation_tone_custom,
        "model_default": model_default,
        "model_upgrade": model_upgrade,
        "model_threshold": model_threshold,
        "vision_model": VISION_MODEL,
        "sender_name": sender_name,
        "sender_icon": sender_icon,
        "bot_info": get_bot_info(),
    })


@app.route("/api/admin/features/reset", methods=["POST"])
def api_admin_features_reset():
    """Reset per-group feature settings to global defaults."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    gid = data.get("group_id", "")
    if not gid:
        return jsonify({"error": "missing group_id"}), 400
    group_flex_settings.pop(gid, None)
    group_qr_settings.pop(gid, None)
    group_silent_settings.pop(gid, None)
    group_video_settings.pop(gid, None)
    group_location_settings.pop(gid, None)
    group_mark_read_settings.pop(gid, None)
    group_retry_key_settings.pop(gid, None)
    group_camera_qr_settings.pop(gid, None)
    group_clipboard_qr_settings.pop(gid, None)
    group_camera_roll_qr_settings.pop(gid, None)
    group_location_qr_settings.pop(gid, None)
    group_welcome_settings.pop(gid, None)
    group_tone_settings.pop(gid, None)
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/push", methods=["POST"])
def api_admin_push():
    """Push a message to a group."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    gid = data.get("group_id", "")
    text = data.get("text", "").strip()
    if not gid or not text:
        return jsonify({"error": "missing group_id or text"}), 400
    ok = push_message_to_group(gid, text)
    return jsonify({"ok": ok})


@app.route("/api/admin/richmenu", methods=["POST"])
def api_admin_richmenu():
    """Create or delete rich menu."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    action = data.get("action", "create")
    if action == "create":
        rid = setup_rich_menu()
        return jsonify({"ok": bool(rid), "rich_menu_id": rid})
    elif action == "delete":
        count = delete_rich_menu()
        return jsonify({"ok": True, "deleted": count})
    return jsonify({"error": "invalid action"}), 400


@app.route("/api/admin/insight")
def api_admin_insight():
    """Get follower demographics and message delivery stats."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    return jsonify({
        "demographics": get_insight_followers(),
        "delivery": get_message_delivery_stats(),
    })


@app.route("/api/admin/insight/trend")
def api_admin_insight_trend():
    """Get daily follower trend data (7 days)."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    # v3.9.30c B19: 用 _safe_int 防炸
    days = _safe_int(request.args.get("days", 7), default=7, min_val=1, max_val=30)
    trend = get_statistics_per_unit(num_days=days)
    return jsonify({"trend": trend})


@app.route("/api/admin/richmenu/upload/<rm_id>", methods=["POST"])
def api_admin_richmenu_upload(rm_id):
    """Upload a custom image to a rich menu from admin panel."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    img_b64 = data.get("image", "")
    content_type = data.get("content_type", "image/png")
    if not img_b64:
        return jsonify({"error": "missing image data"}), 400
    try:
        # Strip data URI prefix if present
        if "," in img_b64:
            header, img_b64 = img_b64.split(",", 1)
            if "jpeg" in header or "jpg" in header:
                content_type = "image/jpeg"
            elif "png" in header:
                content_type = "image/png"
        img_bytes = base64.b64decode(img_b64)
        if len(img_bytes) > 5 * 1024 * 1024:
            return jsonify({"error": "image too large (max 5MB)"}), 400
        ok = upload_rich_menu_image_custom(rm_id, img_bytes, content_type)
        return jsonify({"ok": ok})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/imagemap/send", methods=["POST"])
def api_admin_imagemap_send():
    """Send an Imagemap message to a group or user."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    to = data.get("to", "")
    base_url = data.get("base_url", "")
    alt_text = data.get("alt_text", "圖片選單")
    width = int(data.get("width", 1040))
    height = int(data.get("height", 1040))
    actions = data.get("actions", [])
    if not to or not base_url:
        return jsonify({"error": "missing to or base_url"}), 400
    if not actions:
        return jsonify({"error": "missing actions"}), 400
    ok = send_imagemap_message(to, base_url, alt_text, width, height, actions)
    return jsonify({"ok": ok})


@app.route("/api/admin/richmenu/list")
def api_admin_richmenu_list():
    """List all rich menus."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    return jsonify({"menus": list_rich_menus()})


@app.route("/api/admin/richmenu/link", methods=["POST"])
def api_admin_richmenu_link():
    """Link/unlink rich menu to a specific user."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    user_id = data.get("user_id", "")
    rm_id = data.get("rich_menu_id", "")
    action = data.get("action", "link")
    if not user_id:
        return jsonify({"error": "missing user_id"}), 400
    if action == "link" and rm_id:
        ok = link_rich_menu_to_user(user_id, rm_id)
    elif action == "unlink":
        ok = unlink_rich_menu_from_user(user_id)
    else:
        return jsonify({"error": "invalid params"}), 400
    return jsonify({"ok": ok})


@app.route("/api/admin/multicast", methods=["POST"])
def api_admin_multicast():
    """Send a message to multiple users."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    user_ids = data.get("user_ids", [])
    text = data.get("text", "").strip()
    if not user_ids or not text:
        return jsonify({"error": "missing user_ids or text"}), 400
    ok = multicast_message(user_ids, text)
    return jsonify({"ok": ok})


@app.route("/api/admin/broadcast", methods=["POST"])
def api_admin_broadcast():
    """Broadcast to all followers."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    text = data.get("text", "").strip()
    if not text:
        return jsonify({"error": "missing text"}), 400
    ok = broadcast_message(text)
    return jsonify({"ok": ok})


@app.route("/api/admin/richmenu/alias", methods=["POST"])
def api_admin_richmenu_alias():
    """Create/delete rich menu alias."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    alias_id = data.get("alias_id", "")
    rm_id = data.get("rich_menu_id", "")
    action = data.get("action", "create")
    ok = manage_rich_menu_alias(alias_id, rm_id, action)
    return jsonify({"ok": ok})


@app.route("/api/admin/richmenu/batch", methods=["POST"])
def api_admin_richmenu_batch():
    """Batch link/unlink rich menu to users."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    user_ids = data.get("user_ids", [])
    rm_id = data.get("rich_menu_id", "")
    action = data.get("action", "link")
    if action == "link" and rm_id:
        ok = batch_link_rich_menu(user_ids, rm_id)
    else:
        ok = batch_unlink_rich_menu(user_ids)
    return jsonify({"ok": ok})


@app.route("/api/admin/groups/settings", methods=["POST"])
def api_admin_group_settings():
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    gid = data.get("group_id", "")
    if not gid:
        return jsonify({"error": "missing group_id"}), 400
    if "target_lang" in data:
        group_target_lang[gid] = data["target_lang"]
    if "translation_on" in data:
        group_settings[gid] = bool(data["translation_on"])
    if "image_on" in data:
        group_img_settings[gid] = bool(data["image_on"])
        # 開了自動翻譯就清掉 ask 旗標
        if bool(data["image_on"]):
            group_img_ask_settings.pop(gid, None)
    if "image_ask_mode" in data:
        # 詢問模式 = image_on=False AND image_ask_mode=True
        if bool(data["image_ask_mode"]):
            group_img_settings[gid] = False
            group_img_ask_settings[gid] = True
        else:
            group_img_ask_settings.pop(gid, None)
    if "voice_on" in data:
        group_audio_settings[gid] = bool(data["voice_on"])
    if "work_order_on" in data:
        group_wo_settings[gid] = bool(data["work_order_on"])
    if "cmd_toggle" in data:
        cmd_key = data["cmd_toggle"]
        cmd_val = bool(data.get("cmd_val", True))
        if gid not in group_cmd_enabled:
            group_cmd_enabled[gid] = {}
        group_cmd_enabled[gid][cmd_key] = cmd_val
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/groups/reset-cost", methods=["POST"])
def api_admin_reset_group_cost():
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    gid = data.get("group_id", "")
    if not gid:
        return jsonify({"error": "missing group_id"}), 400
    group_api_usage.pop(gid, None)
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/dm", methods=["GET", "POST"])
def api_admin_dm():
    global dm_master_enabled
    if not check_manager_access("settings"):
        return jsonify({"error": "forbidden"}), 403
    if request.method == "POST":
        data = request.get_json() or {}
        if "master_enabled" in data:
            dm_master_enabled = bool(data["master_enabled"])
        save_settings()
        return jsonify({"ok": True})
    # Build known users list with whitelist status
    known = []
    for uid, name in dm_known_users.items():
        known.append({"user_id": uid, "name": name, "whitelisted": uid in dm_whitelist})
    return jsonify({
        "master_enabled": dm_master_enabled,
        "whitelist": list(dm_whitelist),
        "known_users": known
    })


@app.route("/api/admin/dm/whitelist", methods=["POST"])
def api_admin_dm_whitelist():
    if not check_manager_access("settings"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    uid = data.get("user_id", "").strip()
    action = data.get("action", "add")
    if not uid:
        return jsonify({"error": "missing user_id"}), 400
    if action == "add":
        dm_whitelist.add(uid)
    elif action == "remove":
        dm_whitelist.discard(uid)
    save_settings()
    return jsonify({"ok": True, "whitelist": list(dm_whitelist)})


@app.route("/api/admin/skip", methods=["GET", "POST"])
def api_admin_skip():
    if not check_manager_access("skip"):
        return jsonify({"error": "forbidden"}), 403
    if request.method == "POST":
        data = request.get_json() or {}
        gid = data.get("group_id", "")
        uid = data.get("user_id", "")
        action = data.get("action", "add")
        if not gid or not uid:
            return jsonify({"error": "missing group_id or user_id"}), 400
        if gid not in group_skip_users:
            group_skip_users[gid] = set()
        if action == "add":
            group_skip_users[gid].add(uid)
            save_settings()
            return jsonify({"ok": True})
        elif action == "remove":
            group_skip_users[gid].discard(uid)
            save_settings()
            return jsonify({"ok": True})
    # GET: return all known users in group with skip status
    gid = request.args.get("group_id", "")
    skipped = group_skip_users.get(gid, set())
    names_cache = group_user_names.get(gid, {})
    users = []
    for uid, dname in names_cache.items():
        users.append({"user_id": uid, "name": dname, "skipped": uid in skipped})
    users.sort(key=lambda x: x["name"])
    return jsonify({"users": users})


@app.route("/api/admin/users", methods=["GET"])
def api_admin_users():
    if not check_manager_access("users"):
        return jsonify({"error": "forbidden"}), 403
    filter_gid = request.args.get("group_id", "")
    users = []
    if filter_gid:
        names_cache = group_user_names.get(filter_gid, {})
        for uid, name in names_cache.items():
            is_admin = admin_users.get(uid, {}).get("is_admin", False)
            users.append({
                "user_id": uid,
                "name": name,
                "is_admin": is_admin,
                "allowed_tabs": admin_users.get(uid, {}).get("allowed_tabs", []),
                "google_email": admin_users.get(uid, {}).get("google_email", ""),
                "line_lang": user_languages.get(uid, ""),
                "picture_url": user_pictures.get(uid, ""),
            })
    else:
        all_users = {}
        for uid, name in dm_known_users.items():
            all_users[uid] = name
        for gid, names in group_user_names.items():
            for uid, name in names.items():
                if uid not in all_users:
                    all_users[uid] = name
        for uid, name in all_users.items():
            is_admin = admin_users.get(uid, {}).get("is_admin", False)
            users.append({
                "user_id": uid,
                "name": name,
                "is_admin": is_admin,
                "allowed_tabs": admin_users.get(uid, {}).get("allowed_tabs", []),
                "google_email": admin_users.get(uid, {}).get("google_email", ""),
                "line_lang": user_languages.get(uid, ""),
                "picture_url": user_pictures.get(uid, ""),
            })
    users.sort(key=lambda x: x["name"])
    return jsonify({"users": users})


@app.route("/api/admin/users/admin", methods=["POST"])
def api_admin_users_toggle_admin():
    if not check_manager_access("users"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    uid = data.get("user_id", "")
    is_admin = bool(data.get("is_admin", False))
    if not uid:
        return jsonify({"error": "missing user_id"}), 400
    if uid not in admin_users:
        admin_users[uid] = {}
    admin_users[uid]["is_admin"] = is_admin
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/users/tabs", methods=["POST"])
def api_admin_users_tabs():
    """Set allowed tabs for a manager user."""
    if not check_manager_access("users"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    uid = data.get("user_id", "")
    tabs = data.get("allowed_tabs", [])
    if not uid:
        return jsonify({"error": "missing user_id"}), 400
    if uid not in admin_users:
        admin_users[uid] = {}
    admin_users[uid]["allowed_tabs"] = tabs
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/users/email", methods=["POST"])
def api_admin_users_email():
    """Set Google email for a manager user."""
    if not check_manager_access("users"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    uid = data.get("user_id", "")
    email = data.get("google_email", "").strip().lower()
    if not uid:
        return jsonify({"error": "missing user_id"}), 400
    if uid not in admin_users:
        admin_users[uid] = {}
    admin_users[uid]["google_email"] = email
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/names", methods=["GET", "POST"])
def api_admin_names():
    if not check_manager_access("names"):
        return jsonify({"error": "forbidden"}), 403
    names_list = extra_names_by_group.setdefault("__all__", [])
    if request.method == "POST":
        data = request.get_json() or {}
        action = data.get("action", "add")
        name = data.get("name", "").strip()
        if not name:
            return jsonify({"error": "missing name"}), 400
        if action == "add":
            if name not in names_list:
                names_list.append(name)
                rebuild_customer_names()
                save_settings()
            return jsonify({"ok": True})
        elif action == "remove":
            if name in names_list:
                names_list.remove(name)
                rebuild_customer_names()
                save_settings()
            return jsonify({"ok": True})
    return jsonify({"names": names_list, "count": len(names_list)})


@app.route("/api/admin/storage/stats")
def api_admin_storage_stats():
    if not check_manager_access("storage"):
        return jsonify({"error": "forbidden"}), 403
    return jsonify({"count": len(STORAGE_LOOKUP)})


@app.route("/api/admin/storage/upload", methods=["POST"])
def api_admin_storage_upload():
    global STORAGE_LOOKUP, CUSTOMER_NAMES
    if not check_manager_access("storage"):
        return jsonify({"error": "forbidden"}), 403
    if 'file' not in request.files:
        return jsonify({"error": "沒有檔案"}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({"error": "沒有檔案"}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "空的 Excel"}), 400
        # Auto-detect format: find header row
        header = [str(c).strip() if c else "" for c in rows[0]]
        new_data = {}
        # Try format: Customer | <=3200 | >3200<=4200 | >4200
        len_cols = {}
        for i, h in enumerate(header):
            hl = h.replace(" ", "")
            if "<=3200" in hl and ">3200" not in hl:
                len_cols["<=3200"] = i
            elif ">3200" in hl and "<=4200" in hl:
                len_cols[">3200<=4200"] = i
            elif ">4200" in hl:
                len_cols[">4200"] = i
        if len(len_cols) >= 2:
            # Detected column-based format
            cust_col = 0  # assume first column is customer
            for _, row in enumerate(rows[1:], 1):
                if not row or not row[cust_col]:
                    continue
                cust = str(row[cust_col]).strip()
                if not cust:
                    continue
                entries = []
                for length_key, col_idx in len_cols.items():
                    if col_idx < len(row) and row[col_idx]:
                        zone = str(row[col_idx]).strip()
                        if zone:
                            entries.append([length_key, zone])
                if entries:
                    new_data[cust] = entries
        else:
            # Try row-based format: Customer | LengthRange | Zone
            for row in rows[1:]:
                if not row or len(row) < 3:
                    continue
                cust = str(row[0]).strip() if row[0] else ""
                length_key = str(row[1]).strip() if row[1] else ""
                zone = str(row[2]).strip() if row[2] else ""
                if cust and length_key and zone:
                    if cust not in new_data:
                        new_data[cust] = []
                    new_data[cust].append([length_key, zone])
        if not new_data:
            return jsonify({"error": "無法解析 Excel，請確認格式：\n欄A=客戶 欄B=<=3200 欄C=>3200<=4200 欄D=>4200"}), 400
        # Update in-memory
        STORAGE_LOOKUP = new_data
        rebuild_customer_names()
        logger.info("Storage updated via admin: %d customers", len(new_data))
        # Auto-commit to GitHub for permanent update
        json_str = json.dumps(new_data, ensure_ascii=False, indent=2)
        gh_ok = commit_storage_to_github(json_str)
        msg = "已更新 " + str(len(new_data)) + " 筆客戶"
        if gh_ok:
            msg += "（已自動推送 GitHub，將永久生效）"
        else:
            msg += "（GitHub 推送失敗，僅暫時生效）"
        return jsonify({"ok": True, "count": len(new_data), "github": gh_ok, "message": msg})
    except Exception as e:
        logger.error("Storage upload error: %s", e)
        return jsonify({"error": "解析失敗: " + str(e)}), 400


@app.route("/api/admin/storage/json")
def api_admin_storage_json():
    if not check_manager_access("storage"):
        return jsonify({"error": "forbidden"}), 403
    json_str = json.dumps(STORAGE_LOOKUP, ensure_ascii=False, indent=2)
    return app.response_class(json_str, mimetype="application/json",
                              headers={"Content-Disposition": "attachment; filename=storage_data.json"})


# ─── Passwords API ──────────────────────────────────
@app.route("/api/admin/passwords", methods=["GET", "POST"])
def api_admin_passwords():
    global pw1_text, pw2_text
    if not check_manager_access("passwords"):
        return jsonify({"error": "forbidden"}), 403
    if request.method == "GET":
        return jsonify({"pw1": pw1_text, "pw2": pw2_text})
    data = request.get_json(force=True)
    if "pw1" in data:
        pw1_text = data["pw1"]
    if "pw2" in data:
        pw2_text = data["pw2"]
    save_settings()
    return jsonify({"ok": True, "pw1": pw1_text, "pw2": pw2_text})


# ─── Scrap Text API ─────────────────────────────────
@app.route("/api/admin/scrap", methods=["GET", "POST"])
def api_admin_scrap():
    global scrap_text
    if not check_manager_access("scrap"):
        return jsonify({"error": "forbidden"}), 403
    if request.method == "GET":
        return jsonify({"text": scrap_text})
    data = request.get_json(force=True)
    if "text" in data:
        scrap_text = data["text"]
    save_settings()
    return jsonify({"ok": True})


# ─── Custom Translation Examples API ──────────────────
@app.route("/api/admin/examples", methods=["GET"])
def api_admin_examples_get():
    """Get all custom translation examples."""
    if not check_manager_access("examples"):
        return jsonify({"error": "forbidden"}), 403
    return jsonify({
        "examples": custom_translation_examples,
        "count": len(custom_translation_examples),
        "max": CUSTOM_EXAMPLES_MAX,
    })


@app.route("/api/admin/examples/add", methods=["POST"])
def api_admin_examples_add():
    """Add a custom translation example."""
    global custom_translation_examples
    if not check_manager_access("examples"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    zh = data.get("zh", "").strip()
    idn = data.get("id", "").strip()
    direction = data.get("dir", "zh2id")
    if not zh or not idn:
        return jsonify({"error": "missing zh or id"}), 400
    if len(custom_translation_examples) >= CUSTOM_EXAMPLES_MAX:
        return jsonify({"error": "max_reached", "max": CUSTOM_EXAMPLES_MAX}), 400
    # Check for duplicates
    for ex in custom_translation_examples:
        if ex.get("zh") == zh and ex.get("id") == idn:
            return jsonify({"error": "duplicate"}), 400
    custom_translation_examples.append({"zh": zh, "id": idn, "dir": direction})
    save_settings()
    return jsonify({"ok": True, "count": len(custom_translation_examples)})


@app.route("/api/admin/examples/delete", methods=["POST"])
def api_admin_examples_delete():
    """Delete a custom translation example by index."""
    global custom_translation_examples
    if not check_manager_access("examples"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    idx = data.get("index")
    if idx is None or idx < 0 or idx >= len(custom_translation_examples):
        return jsonify({"error": "invalid index"}), 400
    custom_translation_examples.pop(idx)
    save_settings()
    return jsonify({"ok": True, "count": len(custom_translation_examples)})


@app.route("/api/admin/examples/edit", methods=["POST"])
def api_admin_examples_edit():
    """Edit a custom translation example by index."""
    global custom_translation_examples
    if not check_manager_access("examples"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    idx = data.get("index")
    if idx is None or idx < 0 or idx >= len(custom_translation_examples):
        return jsonify({"error": "invalid index"}), 400
    zh = data.get("zh", "").strip()
    idn = data.get("id", "").strip()
    direction = data.get("dir", custom_translation_examples[idx].get("dir", "zh2id"))
    if not zh or not idn:
        return jsonify({"error": "missing zh or id"}), 400
    custom_translation_examples[idx] = {"zh": zh, "id": idn, "dir": direction}
    save_settings()
    return jsonify({"ok": True})


# ─── Packaging API ──────────────────────────────────
@app.route("/api/admin/packaging/stats")
def api_admin_packaging_stats():
    if not check_manager_access("packaging"):
        return jsonify({"error": "forbidden"}), 403
    return jsonify({"count": len(PACKAGING_LOOKUP)})


@app.route("/api/admin/packaging/upload", methods=["POST"])
def api_admin_packaging_upload():
    global PACKAGING_LOOKUP
    if not check_manager_access("packaging"):
        return jsonify({"error": "forbidden"}), 403
    if 'file' not in request.files:
        return jsonify({"error": "沒有檔案"}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({"error": "沒有檔案"}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        # Use first sheet (直棒包裝 or whatever)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "空的 Excel"}), 400
        header = [str(c).strip() if c else "" for c in rows[0]]
        # Find the code column: look for header containing 碼/code
        code_col = None
        for i, h in enumerate(header):
            hl = h.lower().replace(" ", "")
            if "包裝碼" in h or "代碼" in h or "代号" in h or "code" in hl:
                code_col = i
                break
        # Fallback: if header contains just "碼" somewhere
        if code_col is None:
            for i, h in enumerate(header):
                if "碼" in h:
                    code_col = i
                    break
        # Last fallback: first non-empty header column
        if code_col is None:
            for i, h in enumerate(header):
                if h:
                    code_col = i
                    break
        if code_col is None:
            return jsonify({"error": "找不到包裝碼欄位"}), 400
        # All other columns with non-empty headers become data fields
        data_cols = []  # [(col_index, header_name), ...]
        for i, h in enumerate(header):
            if i != code_col and h:
                data_cols.append((i, h))
        # Build lookup
        new_data = {}
        for row in rows[1:]:
            if not row:
                continue
            if code_col >= len(row) or not row[code_col]:
                continue
            code = str(row[code_col]).strip()
            if not code:
                continue
            entry = {}
            for col_idx, col_name in data_cols:
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx]).strip()
                    if val:
                        entry[col_name] = val
            if entry:
                new_data[code] = entry
        if not new_data:
            return jsonify({"error": "無法解析 Excel，請確認第一列為標題列，含包裝碼欄位"}), 400
        PACKAGING_LOOKUP = new_data
        logger.info("Packaging updated via admin: %d codes, columns: %s",
                     len(new_data), [c[1] for c in data_cols])
        json_str = json.dumps(new_data, ensure_ascii=False, indent=2)
        gh_ok = commit_packaging_to_github(json_str)
        cols_info = "、".join([c[1] for c in data_cols])
        msg = "已更新 " + str(len(new_data)) + " 筆包裝碼（欄位：" + cols_info + "）"
        if gh_ok:
            msg += "\n已自動推送 GitHub，永久生效"
        else:
            msg += "\nGitHub 推送失敗，僅暫時生效"
        return jsonify({"ok": True, "count": len(new_data), "github": gh_ok, "message": msg})
    except Exception as e:
        logger.error("Packaging upload error: %s", e)
        return jsonify({"error": "解析失敗: " + str(e)}), 400


@app.route("/api/admin/packaging/json")
def api_admin_packaging_json():
    if not check_manager_access("packaging"):
        return jsonify({"error": "forbidden"}), 403
    json_str = json.dumps(PACKAGING_LOOKUP, ensure_ascii=False, indent=2)
    return app.response_class(json_str, mimetype="application/json",
                              headers={"Content-Disposition": "attachment; filename=packaging_data.json"})


@app.route("/api/admin/webhook", methods=["GET", "POST"])
def api_admin_webhook():
    """Get or set webhook endpoint URL."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    if request.method == "POST":
        data = request.get_json() or {}
        url = data.get("url", "").strip()
        if not url:
            return jsonify({"error": "missing url"}), 400
        ok = set_webhook_url(url)
        return jsonify({"ok": ok})
    return jsonify({"webhook": get_webhook_info()})


@app.route("/api/admin/webhook/test", methods=["POST"])
def api_admin_webhook_test():
    """Test webhook endpoint connectivity."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    endpoint = data.get("endpoint", "")
    result = test_webhook(endpoint if endpoint else None)
    return jsonify(result)


@app.route("/api/admin/validate", methods=["POST"])
def api_admin_validate():
    """Validate message objects before sending."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    text = data.get("text", "").strip()
    msg_type = data.get("type", "reply")
    if not text:
        return jsonify({"error": "missing text"}), 400
    msgs = [{"type": "text", "text": text}]
    result = validate_message_objects(msgs, msg_type)
    return jsonify(result)


@app.route("/api/admin/content/preview/<message_id>")
def api_admin_content_preview(message_id):
    """Get content preview for an image/video message."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    content = get_content_preview(message_id)
    if content:
        return app.response_class(content, mimetype="image/jpeg")
    return jsonify({"error": "not found"}), 404


@app.route("/api/admin/content/status/<message_id>")
def api_admin_content_status(message_id):
    """Check preparation status for video/audio content."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    status = check_content_preparation(message_id)
    return jsonify({"message_id": message_id, "status": status})


@app.route("/api/admin/followers")
def api_admin_followers():
    """Get all follower user IDs and resolve names."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    ids = get_all_follower_ids()
    followers = []
    for uid in ids:
        name = ""
        # Try to find in known users
        for gid, names in group_user_names.items():
            if uid in names:
                name = names[uid]
                break
        if not name:
            dm_val = dm_known_users.get(uid, "")
            name = dm_val if isinstance(dm_val, str) else (dm_val.get("name", "") if isinstance(dm_val, dict) else "")
        followers.append({"user_id": uid, "name": name})
    return jsonify({"count": len(ids), "followers": followers})


@app.route("/api/admin/interaction")
def api_admin_interaction():
    """Get message interaction stats by request_id."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    req_id = request.args.get("request_id", "")
    if not req_id:
        return jsonify({"error": "missing request_id"}), 400
    stats = get_message_interaction_stats(req_id)
    return jsonify({"stats": stats})


@app.route("/api/admin/delivery")
def api_admin_delivery():
    """Get full delivery stats (reply/push/multicast/broadcast)."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    date_str = request.args.get("date", "")
    stats = get_message_delivery_stats(date_str if date_str else None)
    return jsonify({"delivery": stats})


@app.route("/api/admin/room/members")
def api_admin_room_members():
    """Get members of a multi-person chat room."""
    if not check_manager_access("insight"):
        return jsonify({"error": "forbidden"}), 403
    room_id = request.args.get("room_id", "")
    if not room_id:
        return jsonify({"error": "missing room_id"}), 400
    count = get_room_member_count(room_id)
    members = fetch_all_room_members(room_id)
    profiles = []
    for uid in members[:50]:  # limit to 50
        p = get_room_member_profile(room_id, uid)
        if p:
            profiles.append(p)
    return jsonify({"count": count, "members": profiles})


@app.route("/api/admin/richmenu/detail/<rm_id>")
def api_admin_richmenu_detail(rm_id):
    """Get detailed info of a single rich menu."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    detail = get_rich_menu_by_id(rm_id)
    return jsonify({"menu": detail})


@app.route("/api/admin/richmenu/default")
def api_admin_richmenu_default():
    """Get the current default rich menu ID."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    rm_id = get_default_rich_menu_id()
    return jsonify({"default_rich_menu_id": rm_id})


@app.route("/api/admin/richmenu/user/<user_id>")
def api_admin_richmenu_user(user_id):
    """Get which rich menu is linked to a specific user."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    rm_id = get_user_rich_menu_id(user_id)
    return jsonify({"user_id": user_id, "rich_menu_id": rm_id})


@app.route("/api/admin/richmenu/validate", methods=["POST"])
def api_admin_richmenu_validate():
    """Validate a rich menu object before creating."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    result = validate_rich_menu_obj(data)
    return jsonify(result)


@app.route("/api/admin/richmenu/image/<rm_id>")
def api_admin_richmenu_image(rm_id):
    """Download/preview the image of a rich menu."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    content = download_rich_menu_image(rm_id)
    if content:
        return app.response_class(content, mimetype="image/png")
    return jsonify({"error": "not found"}), 404


@app.route("/api/admin/richmenu/alias/list")
def api_admin_richmenu_alias_list():
    """Get all rich menu aliases."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    aliases = list_rich_menu_aliases()
    return jsonify({"aliases": aliases})


@app.route("/api/admin/richmenu/alias/update", methods=["POST"])
def api_admin_richmenu_alias_update():
    """Update a rich menu alias to point to a different menu."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    alias_id = data.get("alias_id", "")
    rm_id = data.get("rich_menu_id", "")
    if not alias_id or not rm_id:
        return jsonify({"error": "missing alias_id or rich_menu_id"}), 400
    ok = update_rich_menu_alias(alias_id, rm_id)
    return jsonify({"ok": ok})


@app.route("/api/admin/richmenu/alias/detail/<alias_id>")
def api_admin_richmenu_alias_detail(alias_id):
    """Get info about a specific rich menu alias."""
    if not check_manager_access("groups"):
        return jsonify({"error": "forbidden"}), 403
    info = get_rich_menu_alias(alias_id)
    return jsonify({"alias": info})


# ═══════════════════════════════════════════════════════════
# ─── LIFF Form System ─────────────────────────────────────
# ═══════════════════════════════════════════════════════════

LIFF_FORM_HTML = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>表單系統</title>
<script src="https://static.line-scdn.net/liff/edge/2/sdk.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,sans-serif;background:#0d0d0d;color:#e0e0e0;min-height:100vh}
.header{background:linear-gradient(135deg,#7c6fef,#5b6abf);padding:16px;text-align:center;font-size:18px;font-weight:700}
.container{padding:16px;max-width:600px;margin:0 auto}
.card{background:#1a1a2e;border:1px solid #2a2a3e;border-radius:12px;padding:16px;margin-bottom:12px}
.field-label{font-size:14px;font-weight:600;margin-bottom:6px;color:#ccc}
.field-label-id{font-size:12px;color:#8a8a9a;margin-bottom:8px}
input[type=text],input[type=number],input[type=date],textarea,select{width:100%;padding:10px 12px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:14px;margin-bottom:4px}
textarea{resize:vertical;min-height:60px}
select{appearance:none;-webkit-appearance:none;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%238a8a9a' d='M6 8L1 3h10z'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 10px center}
.checkbox-wrap{display:flex;align-items:center;gap:8px;padding:8px 0}
.checkbox-wrap input[type=checkbox]{width:20px;height:20px;accent-color:#7c6fef}
.checkbox-wrap label{font-size:14px}
.btn{display:block;width:100%;padding:14px;border:none;border-radius:10px;font-size:16px;font-weight:700;cursor:pointer;text-align:center;margin-top:12px}
.btn-primary{background:#7c6fef;color:#fff}
.btn-primary:active{background:#6358d4}
.btn-primary:disabled{background:#3a3a4e;color:#666;cursor:not-allowed}
.required-mark{color:#f04747;margin-left:2px}
.success-box{background:#1a3a2a;border:1px solid #43b581;border-radius:12px;padding:24px;text-align:center;margin-top:20px}
.success-box h2{color:#43b581;font-size:20px;margin-bottom:8px}
.error-box{background:#3a1a1a;border:1px solid #f04747;border-radius:12px;padding:24px;text-align:center;margin-top:20px}
.error-box h2{color:#f04747;font-size:20px;margin-bottom:8px}
.loading{text-align:center;padding:40px;color:#8a8a9a}
.already-box{background:#2a2a1a;border:1px solid #faa61a;border-radius:12px;padding:24px;text-align:center;margin-top:20px}
.already-box h2{color:#faa61a;font-size:20px;margin-bottom:8px}
.form-list .form-item{background:#1a1a2e;border:1px solid #2a2a3e;border-radius:10px;padding:14px;margin-bottom:10px;cursor:pointer}
.form-item:active{background:#252540}
.form-item h3{font-size:15px;margin-bottom:4px}
.form-item .sub{font-size:12px;color:#8a8a9a}
.expired{opacity:.5}
</style>
</head>
<body>
<div class="header">📋 表單系統 Form System</div>
<div class="container" id="app">
<div class="loading" id="loadingBox">載入中 Loading...</div>
</div>
<script>
var LIFF_ID='""" + LIFF_ID + """';
var API_BASE=location.origin;
var currentUser=null;
var formId=null;

async function initLiff(){
  try{
    await liff.init({liffId:LIFF_ID});
    if(!liff.isLoggedIn()){liff.login();return}
    var profile=await liff.getProfile();
    currentUser={userId:profile.userId,displayName:profile.displayName,pictureUrl:profile.pictureUrl||''};
    var params=new URLSearchParams(location.search);
    formId=params.get('id');
    if(formId){
      await loadForm(formId);
    }else{
      await loadFormList();
    }
  }catch(e){
    /* v3.9.30c B12 修補: e.message 用 esc() 防 XSS */
    document.getElementById('app').innerHTML='<div class="error-box"><h2>Error</h2><p>'+esc(e.message||'')+'</p></div>';
  }
}

async function loadFormList(){
  try{
    var r=await fetch(API_BASE+'/api/liff/forms');
    var data=await r.json();
    var forms=data.forms||[];
    var active=forms.filter(function(f){return f.status==='active'});
    if(active.length===0){
      document.getElementById('app').innerHTML='<div style="text-align:center;padding:40px;color:#8a8a9a">目前沒有表單<br>No forms available</div>';
      return;
    }
    var html='<div class="form-list">';
    for(var i=0;i<active.length;i++){
      var f=active[i];
      html+='<div class="form-item" onclick="location.href=location.pathname+&apos;?id='+f.id+'&apos;"><h3>'+esc(f.title_zh)+'</h3><div class="sub">'+esc(f.title_id)+'</div></div>';
    }
    html+='</div>';
    document.getElementById('app').innerHTML=html;
  }catch(e){
    /* v3.9.30c B12 修補: e.message 用 esc() */
    document.getElementById('app').innerHTML='<div class="error-box"><h2>Error</h2><p>'+esc(e.message||'')+'</p></div>';
  }
}

async function loadForm(fid){
  try{
    var r=await fetch(API_BASE+'/api/liff/form/'+fid+'?user_id='+currentUser.userId);
    var data=await r.json();
    if(data.error){
      document.getElementById('app').innerHTML='<div class="error-box"><h2>錯誤</h2><p>'+esc(data.error)+'</p></div>';
      return;
    }
    if(data.already_submitted){
      document.getElementById('app').innerHTML='<div class="already-box"><h2>✅ 已填寫 Sudah diisi</h2><p>你已經提交過此表單<br>Anda sudah mengisi formulir ini</p></div>';
      return;
    }
    renderForm(data.form);
  }catch(e){
    /* v3.9.30c B12 修補: e.message 用 esc() */
    document.getElementById('app').innerHTML='<div class="error-box"><h2>Error</h2><p>'+esc(e.message||'')+'</p></div>';
  }
}

function renderForm(form){
  var html='<div class="card" style="margin-bottom:16px"><h2 style="font-size:18px;margin-bottom:4px">'+esc(form.title_zh)+'</h2><div style="font-size:13px;color:#8a8a9a">'+esc(form.title_id)+'</div></div>';
  var fields=form.fields||[];
  for(var i=0;i<fields.length;i++){
    var f=fields[i];
    html+='<div class="card">';
    html+='<div class="field-label">'+esc(f.label_zh)+(f.required?'<span class="required-mark">*</span>':'')+'</div>';
    html+='<div class="field-label-id">'+esc(f.label_id)+'</div>';
    if(f.type==='text'){
      html+='<input type="text" id="f_'+f.id+'" placeholder="">';
    }else if(f.type==='number'){
      html+='<input type="number" id="f_'+f.id+'" placeholder="">';
    }else if(f.type==='date'){
      html+='<input type="date" id="f_'+f.id+'">';
    }else if(f.type==='textarea'){
      html+='<textarea id="f_'+f.id+'"></textarea>';
    }else if(f.type==='select'){
      html+='<select id="f_'+f.id+'"><option value="">-- 請選擇 Pilih --</option>';
      var opts=f.options||[];
      for(var j=0;j<opts.length;j++){
        html+='<option value="'+esc(opts[j].zh)+'">'+esc(opts[j].zh)+' / '+esc(opts[j].id)+'</option>';
      }
      html+='</select>';
    }else if(f.type==='checkbox'){
      html+='<div class="checkbox-wrap"><input type="checkbox" id="f_'+f.id+'"><label for="f_'+f.id+'">'+esc(f.label_zh)+' / '+esc(f.label_id)+'</label></div>';
    }
    html+='</div>';
  }
  html+='<button class="btn btn-primary" id="submitBtn" onclick="submitForm()">提交 Kirim</button>';
  document.getElementById('app').innerHTML=html;
}

async function submitForm(){
  var btn=document.getElementById('submitBtn');
  btn.disabled=true;btn.textContent='提交中 Mengirim...';
  try{
    var r=await fetch(API_BASE+'/api/liff/form/'+formId+'?user_id='+currentUser.userId);
    var meta=await r.json();
    var fields=meta.form.fields||[];
    var answers={};
    for(var i=0;i<fields.length;i++){
      var f=fields[i];
      var el=document.getElementById('f_'+f.id);
      if(!el)continue;
      var val='';
      if(f.type==='checkbox'){val=el.checked?'yes':'no'}
      else{val=el.value.trim()}
      if(f.required && !val && f.type!=='checkbox'){
        alert('請填寫: '+f.label_zh+'\\nHarap isi: '+f.label_id);
        btn.disabled=false;btn.textContent='提交 Kirim';return;
      }
      answers[f.id]=val;
    }
    var body={user_id:currentUser.userId,user_name:currentUser.displayName,picture_url:currentUser.pictureUrl,answers:answers};
    var r2=await fetch(API_BASE+'/api/liff/form/'+formId+'/submit',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
    var res=await r2.json();
    if(res.ok){
      document.getElementById('app').innerHTML='<div class="success-box"><h2>✅ 提交成功 Berhasil</h2><p>感謝填寫！<br>Terima kasih sudah mengisi!</p></div>';
      setTimeout(function(){if(liff.isInClient())liff.closeWindow()},2000);
    }else{
      alert(res.error||'Error');btn.disabled=false;btn.textContent='提交 Kirim';
    }
  }catch(e){
    alert('Error: '+e.message);btn.disabled=false;btn.textContent='提交 Kirim';
  }
}

function esc(s){if(!s)return '';return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;')}

document.addEventListener('DOMContentLoaded',initLiff);
</script>
</body>
</html>"""


@app.route("/liff/form")
def liff_form_page():
    resp = app.response_class(LIFF_FORM_HTML, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp


@app.route("/api/liff/forms")
def api_liff_forms_list():
    """Public: list active forms (no auth needed, LIFF user sees these)."""
    result = []
    for fid, f in forms_data.items():
        if f.get("status") == "active":
            result.append({"id": fid, "title_zh": f["title_zh"], "title_id": f["title_id"], "field_count": len(f.get("fields", []))})
    return jsonify({"forms": result})


@app.route("/api/liff/form/<form_id>")
def api_liff_form_detail(form_id):
    """Public: get form detail + check if user already submitted."""
    f = forms_data.get(form_id)
    if not f:
        return jsonify({"error": "表單不存在 Form not found"}), 404
    if f.get("status") != "active":
        return jsonify({"error": "表單已關閉 Form closed"}), 403
    user_id = request.args.get("user_id", "")
    already = False
    if user_id and form_id in forms_submissions:
        already = user_id in forms_submissions[form_id]
    return jsonify({"form": f, "already_submitted": already})


@app.route("/api/liff/form/<form_id>/submit", methods=["POST"])
def api_liff_form_submit(form_id):
    """Public: submit form answers."""
    f = forms_data.get(form_id)
    if not f:
        return jsonify({"error": "表單不存在"}), 404
    if f.get("status") != "active":
        return jsonify({"error": "表單已關閉"}), 403
    body = request.get_json(force=True)
    user_id = body.get("user_id", "")
    if not user_id:
        return jsonify({"error": "缺少用戶ID"}), 400
    if form_id not in forms_submissions:
        forms_submissions[form_id] = {}
    if user_id in forms_submissions[form_id]:
        return jsonify({"error": "已填寫過 Sudah diisi"}), 409
    forms_submissions[form_id][user_id] = {
        "user_name": body.get("user_name", ""),
        "picture_url": body.get("picture_url", ""),
        "answers": body.get("answers", {}),
        "submitted_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "approved": False,
    }
    save_settings()
    return jsonify({"ok": True})


# ─── Admin Form Management API ─────────────────────────
@app.route("/api/admin/forms", methods=["GET"])
def api_admin_forms_list():
    if not check_manager_access("forms"):
        return jsonify({"error": "forbidden"}), 403
    result = []
    for fid, f in forms_data.items():
        sub_count = len(forms_submissions.get(fid, {}))
        result.append({**f, "submission_count": sub_count})
    return jsonify({"forms": result})


@app.route("/api/admin/forms/create", methods=["POST"])
def api_admin_forms_create():
    data = request.get_json(force=True)
    if not check_manager_access("forms"):
        return jsonify({"error": "forbidden"}), 403
    import random, string
    fid = "form_" + "".join(random.choices(string.ascii_lowercase + string.digits, k=8))
    form_obj = {
        "id": fid,
        "title_zh": data.get("title_zh", "未命名表單"),
        "title_id": data.get("title_id", "Formulir tanpa nama"),
        "fields": data.get("fields", []),
        "status": "active",
        "created": time.strftime("%Y-%m-%d %H:%M:%S"),
        "target_groups": data.get("target_groups", []),
    }
    forms_data[fid] = form_obj
    save_settings()
    return jsonify({"ok": True, "form": form_obj})


@app.route("/api/admin/forms/update", methods=["POST"])
def api_admin_forms_update():
    data = request.get_json(force=True)
    if not check_manager_access("forms"):
        return jsonify({"error": "forbidden"}), 403
    fid = data.get("form_id")
    if fid not in forms_data:
        return jsonify({"error": "not found"}), 404
    if "title_zh" in data:
        forms_data[fid]["title_zh"] = data["title_zh"]
    if "title_id" in data:
        forms_data[fid]["title_id"] = data["title_id"]
    if "fields" in data:
        forms_data[fid]["fields"] = data["fields"]
    if "status" in data:
        forms_data[fid]["status"] = data["status"]
    if "target_groups" in data:
        forms_data[fid]["target_groups"] = data["target_groups"]
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/forms/delete", methods=["POST"])
def api_admin_forms_delete():
    data = request.get_json(force=True)
    if not check_manager_access("forms"):
        return jsonify({"error": "forbidden"}), 403
    fid = data.get("form_id")
    if fid in forms_data:
        del forms_data[fid]
    if fid in forms_submissions:
        del forms_submissions[fid]
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/forms/submissions/<form_id>")
def api_admin_forms_submissions(form_id):
    if not check_manager_access("forms"):
        return jsonify({"error": "forbidden"}), 403
    subs = forms_submissions.get(form_id, {})
    result = []
    for uid, s in subs.items():
        result.append({"user_id": uid, **s})
    return jsonify({"submissions": result, "form": forms_data.get(form_id)})


@app.route("/api/admin/forms/approve", methods=["POST"])
def api_admin_forms_approve():
    data = request.get_json(force=True)
    if not check_manager_access("forms"):
        return jsonify({"error": "forbidden"}), 403
    fid = data.get("form_id")
    uid = data.get("user_id")
    if fid in forms_submissions and uid in forms_submissions[fid]:
        forms_submissions[fid][uid]["approved"] = True
        save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/forms/push", methods=["POST"])
def api_admin_forms_push():
    """Push form link to target groups via Flex Message."""
    data = request.get_json(force=True)
    if not check_manager_access("forms"):
        return jsonify({"error": "forbidden"}), 403
    fid = data.get("form_id")
    group_ids = data.get("group_ids", [])
    if fid not in forms_data:
        return jsonify({"error": "form not found"}), 404
    f = forms_data[fid]
    liff_url = "https://liff.line.me/" + LIFF_ID + "?id=" + fid
    flex_msg = {
        "type": "flex",
        "altText": "📋 " + f["title_zh"] + " / " + f["title_id"],
        "contents": {
            "type": "bubble",
            "size": "kilo",
            "header": {
                "type": "box", "layout": "vertical", "backgroundColor": "#7c6fef",
                "paddingAll": "14px",
                "contents": [{"type": "text", "text": "📋 " + f["title_zh"], "color": "#ffffff", "weight": "bold", "size": "md"},
                             {"type": "text", "text": f["title_id"], "color": "#ffffffcc", "size": "sm", "margin": "xs"}]
            },
            "body": {
                "type": "box", "layout": "vertical", "paddingAll": "14px",
                "contents": [
                    {"type": "text", "text": "請點擊下方按鈕填寫表單", "size": "sm", "color": "#999999", "wrap": True},
                    {"type": "text", "text": "Silakan klik tombol di bawah untuk mengisi", "size": "sm", "color": "#999999", "wrap": True, "margin": "xs"},
                    {"type": "text", "text": "欄位數 Jumlah kolom: " + str(len(f.get("fields", []))), "size": "xs", "color": "#666666", "margin": "md"},
                ]
            },
            "footer": {
                "type": "box", "layout": "vertical", "paddingAll": "10px",
                "contents": [
                    {"type": "button", "action": {"type": "uri", "label": "填寫表單 Isi Formulir", "uri": liff_url}, "style": "primary", "color": "#7c6fef", "height": "sm"}
                ]
            }
        }
    }
    pushed = 0
    with ApiClient(configuration) as api_client:
        line_api = MessagingApi(api_client)
        for gid in group_ids:
            try:
                line_api.push_message(PushMessageRequest(to=gid, messages=[FlexMessage.from_dict(flex_msg)]))
                pushed += 1
            except Exception as e:
                logger.error("Push form to %s failed: %s", gid, e)
    return jsonify({"ok": True, "pushed": pushed})


@app.route("/api/admin/forms/export/<form_id>")
def api_admin_forms_export(form_id):
    """Export form submissions as Excel."""
    if not check_manager_access("forms") and request.args.get("key") != ADMIN_KEY:
        return jsonify({"error": "forbidden"}), 403
    f = forms_data.get(form_id)
    if not f:
        return jsonify({"error": "not found"}), 404
    subs = forms_submissions.get(form_id, {})
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "submissions"
        fields = f.get("fields", [])
        headers = ["用戶名稱", "提交時間", "已核准"]
        for fd in fields:
            headers.append(fd.get("label_zh", fd["id"]))
        ws.append(headers)
        for uid, s in subs.items():
            row = [s.get("user_name", uid), s.get("submitted_at", ""), "是" if s.get("approved") else "否"]
            for fd in fields:
                row.append(s.get("answers", {}).get(fd["id"], ""))
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        from flask import send_file
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f["title_zh"] + ".xlsx")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/export/finetune/<token>", methods=["GET"])
def public_export_finetune(token):
    """Public download endpoint for fine-tune JSONL files generated via /export jsonl LINE command.
    Tokens are short-lived (1 hour) and one-time-friendly (we don't expire on use, but they auto-clean).
    """
    rec = _export_tokens.get(token)
    if not rec:
        return "找不到或已過期 / Not found or expired", 404
    path, expiry = rec
    if expiry < int(time.time()):
        try:
            os.remove(path)
        except Exception:
            pass
        _export_tokens.pop(token, None)
        return "已過期 / Expired", 410
    if not os.path.exists(path):
        return "檔案遺失 / File missing", 404
    from flask import send_file
    return send_file(path, mimetype="application/jsonl",
                     as_attachment=True,
                     download_name="finetune_training_data.jsonl")


@app.route("/api/admin/examples/export_jsonl", methods=["GET"])
def api_admin_examples_export_jsonl():
    """Admin: download fine-tune JSONL directly from admin panel."""
    if not check_manager_access("examples") and request.args.get("key") != ADMIN_KEY:
        return jsonify({"error": "forbidden"}), 403
    if not custom_translation_examples:
        return jsonify({"error": "no_examples"}), 400
    try:
        from flask import Response
        sys_prompt = ("You are a Taiwan-factory ZH↔ID translator. "
                      "Output only the translation, no commentary.")
        lines = []
        for ex in custom_translation_examples:
            zh = (ex.get("zh") or "").strip()
            idn = (ex.get("id") or "").strip()
            direction = ex.get("dir", "zh2id")
            if not zh or not idn:
                continue
            if direction == "zh2id":
                user_msg, assistant_msg = zh, idn
            else:
                user_msg, assistant_msg = idn, zh
            row = {"messages": [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": user_msg},
                {"role": "assistant", "content": assistant_msg},
            ]}
            lines.append(json.dumps(row, ensure_ascii=False))
        body = "\n".join(lines) + "\n"
        return Response(
            body,
            mimetype="application/jsonl",
            headers={"Content-Disposition": 'attachment; filename="finetune_training_data.jsonl"'}
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/examples/stats", methods=["GET"])
def api_admin_examples_stats():
    """Admin: training-data status summary for the dashboard."""
    if not check_manager_access("examples"):
        return jsonify({"error": "forbidden"}), 403
    total = len(custom_translation_examples)
    zh2id = sum(1 for e in custom_translation_examples if e.get("dir") == "zh2id")
    id2zh = sum(1 for e in custom_translation_examples if e.get("dir") == "id2zh")
    wrong_count = sum(1 for e in translation_log if e.get("marked_wrong"))
    cutoff_30d = int(time.time()) - 30*86400
    wrong_30d = sum(1 for e in translation_log
                    if e.get("marked_wrong") and e.get("ts", 0) > cutoff_30d)
    return jsonify({
        "total": total,
        "zh2id": zh2id,
        "id2zh": id2zh,
        "max": CUSTOM_EXAMPLES_MAX,
        "marked_wrong_total": wrong_count,
        "marked_wrong_30d": wrong_30d,
        "fine_tune_threshold": 300,
        "ready_for_fine_tune": total >= 300,
    })


@app.route("/api/admin/translation_log", methods=["GET"])
def api_admin_translation_log():
    """Admin: list recent translation log entries.
    Optional query params:
      ?group_id=<gid>       filter by group
      ?limit=<int>          default 50, max 500
      ?only_wrong=1         only entries marked wrong
    """
    if not check_manager_access("examples"):
        return jsonify({"error": "forbidden"}), 403
    gid = request.args.get("group_id", "")
    try:
        limit = max(1, min(int(request.args.get("limit", 50)), 500))
    except Exception:
        limit = 50
    only_wrong = request.args.get("only_wrong") in ("1", "true", "yes")
    out = []
    for e in reversed(translation_log):
        if gid and e.get("group_id") != gid:
            continue
        if only_wrong and not e.get("marked_wrong"):
            continue
        out.append({
            "id": e.get("id"),
            "ts": e.get("ts"),
            "src": (e.get("src") or "")[:200],
            "tgt": (e.get("tgt") or "")[:200],
            "src_lang": e.get("src_lang"),
            "tgt_lang": e.get("tgt_lang"),
            "group_id": e.get("group_id"),
            "marked_wrong": bool(e.get("marked_wrong")),
            "correct_translation": e.get("correct_translation", ""),
            "model": e.get("model"),
            "warned": bool(e.get("warned")),
        })
        if len(out) >= limit:
            break
    return jsonify({"entries": out, "total": len(translation_log)})


@app.route("/api/admin/translation_log/mark_wrong", methods=["POST"])
def api_admin_translation_log_mark_wrong():
    """Admin: mark a specific translation_log entry as wrong, optionally with correction.
    Body: { "entry_id": "...", "correct": "...", "add_to_examples": true }
    """
    if not check_manager_access("examples"):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    eid = data.get("entry_id")
    correct = (data.get("correct") or "").strip()
    add = data.get("add_to_examples", True)
    if not eid:
        return jsonify({"error": "missing entry_id"}), 400
    ok, info = mark_translation_wrong(
        group_id=None,
        correct_translation=correct,
        add_to_examples=bool(correct) and add,
        entry_id=eid,
    )
    return jsonify({"ok": ok, "info": info, "examples_total": len(custom_translation_examples)})


@app.route("/health", methods=["GET"])
def health():
    return {"status": "ok", "version": VERSION, "uptime": int(time.time() - bot_start_time)}


@app.route("/admin/apply-best-defaults", methods=["GET", "POST"])
def admin_apply_best_defaults():
    """v3.9.8: One-click apply OpenAI's recommended optimal settings for translation.
    
    Based on:
      - OpenAI GPT-5/5.1/5.2 prompting guides (cookbook.openai.com)
      - GPT-5.5 official prompting guide (April 2026)
      - arxiv 2505.14810 (reasoning vs instruction-following trade-off)
      - Real production debugging (this bot's 5/1 incident)
    
    Sets all toggles to the proven-best combo for ZH↔ID factory translation.
    Usage: GET/POST /admin/apply-best-defaults?key=YOUR_ADMIN_KEY[&dry_run=1]
    """
    if request.args.get("key") != ADMIN_KEY:
        return jsonify({"error": "forbidden"}), 403
    dry_run = request.args.get("dry_run") == "1"
    
    global model_default, model_upgrade, model_threshold
    global translation_temperature, translation_top_p, translation_seed
    global double_check_mode, double_check_threshold
    global fewshot_mode, logprobs_enabled, confidence_threshold, structured_output_enabled
    global prompt_caching_enabled, reasoning_effort
    global send_user_id_to_openai, send_metadata_to_openai
    global id_zh_cot_enabled, id_zh_cod_enabled, id_zh_pivot_enabled
    global id_preprocessing_enabled, multi_path_backtrans_enabled
    global preserve_paragraphs_enabled, paragraph_split_translate, paragraph_split_threshold
    global stop_sequences_enabled, translation_logging_enabled
    
    # The "best defaults" combo (researched May 2026)
    best = {
        # === Models ===
        "model_default": "gpt-4.1-mini",      # ⭐ Stable, follows examples, cheapest fits-purpose
        "model_upgrade": "gpt-4.1",           # Same family for consistency on long messages
        "model_threshold": 150,               # Was 100 — most factory msgs <150 chars, save cost
        
        # === Sampling (only affects gpt-4 family; gpt-5 ignores) ===
        "translation_temperature": 0.0,       # Deterministic
        "translation_top_p": 1.0,             # No nucleus filtering
        "translation_seed": 0,                # Random (no benefit for short translations)
        
        # === Quality controls ===
        "fewshot_mode": "messages",           # OpenAI standard format (best quality)
        "logprobs_enabled": True,             # Confidence calculation (gpt-4 only, auto-skip on gpt-5)
        "confidence_threshold": 0.85,         # ⚠️ flag if below 85%
        "structured_output_enabled": False,   # Off — adds latency, marginal gain for translation
        "double_check_mode": "smart",         # Only re-translate suspect outputs
        "double_check_threshold": 0.55,       # Permissive (lower = more flagged)
        "prompt_caching_enabled": True,       # Free 75% discount when prefixes match
        
        # === Reasoning (auto-overridden per-model anyway, but for safety) ===
        "reasoning_effort": "low",            # If model has no optimal mapping, low > medium
        
        # === ID→ZH quality ===
        "id_zh_cot_enabled": True,            # Chain-of-Thought 二階段 (verified +20%)
        "id_zh_cod_enabled": True,            # Chain-of-Dictionary (verified +5%)
        "id_zh_pivot_enabled": False,         # Off — doubles cost, marginal gain
        "id_preprocessing_enabled": True,     # Free dictionary normalization
        "multi_path_backtrans_enabled": False,# Off — doubles cost
        
        # === Format preservation ===
        "preserve_paragraphs_enabled": True,  # Honor original paragraph breaks
        "paragraph_split_translate": True,    # Per-paragraph translation for long texts
        "paragraph_split_threshold": 50,      # 50+ char + multi-paragraph = split mode
        
        # === Output cleanliness ===
        "stop_sequences_enabled": True,       # Block "Note:", "Translation:" garbage (gpt-4 only)
        
        # === Compliance / monitoring ===
        "translation_logging_enabled": True,  # Track for quality review
        "send_user_id_to_openai": True,       # Hashed, for abuse tracking
        "send_metadata_to_openai": True,      # Filter dashboard by group/lang
    }
    
    # Snapshot current values for diff
    current = {
        "model_default": model_default, "model_upgrade": model_upgrade, "model_threshold": model_threshold,
        "translation_temperature": translation_temperature, "translation_top_p": translation_top_p,
        "translation_seed": translation_seed, "fewshot_mode": fewshot_mode,
        "logprobs_enabled": logprobs_enabled, "confidence_threshold": confidence_threshold,
        "structured_output_enabled": structured_output_enabled,
        "double_check_mode": double_check_mode, "double_check_threshold": double_check_threshold,
        "prompt_caching_enabled": prompt_caching_enabled, "reasoning_effort": reasoning_effort,
        "id_zh_cot_enabled": id_zh_cot_enabled, "id_zh_cod_enabled": id_zh_cod_enabled,
        "id_zh_pivot_enabled": id_zh_pivot_enabled, "id_preprocessing_enabled": id_preprocessing_enabled,
        "multi_path_backtrans_enabled": multi_path_backtrans_enabled,
        "preserve_paragraphs_enabled": preserve_paragraphs_enabled,
        "paragraph_split_translate": paragraph_split_translate,
        "paragraph_split_threshold": paragraph_split_threshold,
        "stop_sequences_enabled": stop_sequences_enabled,
        "translation_logging_enabled": translation_logging_enabled,
        "send_user_id_to_openai": send_user_id_to_openai, "send_metadata_to_openai": send_metadata_to_openai,
    }
    diff = {k: {"old": current.get(k), "new": v} for k, v in best.items() if current.get(k) != v}
    
    if dry_run:
        return jsonify({
            "mode": "dry_run",
            "would_change": diff,
            "would_keep_unchanged": {k: v for k, v in best.items() if k not in diff},
            "total_changes": len(diff),
        })
    
    # Apply
    model_default = best["model_default"]
    model_upgrade = best["model_upgrade"]
    model_threshold = best["model_threshold"]
    translation_temperature = best["translation_temperature"]
    translation_top_p = best["translation_top_p"]
    translation_seed = best["translation_seed"]
    fewshot_mode = best["fewshot_mode"]
    logprobs_enabled = best["logprobs_enabled"]
    confidence_threshold = best["confidence_threshold"]
    structured_output_enabled = best["structured_output_enabled"]
    double_check_mode = best["double_check_mode"]
    double_check_threshold = best["double_check_threshold"]
    prompt_caching_enabled = best["prompt_caching_enabled"]
    reasoning_effort = best["reasoning_effort"]
    id_zh_cot_enabled = best["id_zh_cot_enabled"]
    id_zh_cod_enabled = best["id_zh_cod_enabled"]
    id_zh_pivot_enabled = best["id_zh_pivot_enabled"]
    id_preprocessing_enabled = best["id_preprocessing_enabled"]
    multi_path_backtrans_enabled = best["multi_path_backtrans_enabled"]
    preserve_paragraphs_enabled = best["preserve_paragraphs_enabled"]
    paragraph_split_translate = best["paragraph_split_translate"]
    paragraph_split_threshold = best["paragraph_split_threshold"]
    stop_sequences_enabled = best["stop_sequences_enabled"]
    translation_logging_enabled = best["translation_logging_enabled"]
    send_user_id_to_openai = best["send_user_id_to_openai"]
    send_metadata_to_openai = best["send_metadata_to_openai"]
    
    return jsonify({
        "status": "applied",
        "changed": diff,
        "total_changes": len(diff),
        "note": "All settings restored to OpenAI-recommended best defaults for translation. "
                "Note: For GPT-5 series, reasoning_effort and verbosity are auto-set per request.",
    })


@app.route("/admin/health-check", methods=["GET"])
def admin_health_check():
    """v3.9.7: 完整健康檢查 — 一頁看完所有狀態:
    模型設定、進階設定相容性、環境變數、最近翻譯、資料持久化、群組統計。
    Usage: GET /admin/health-check?key=YOUR_ADMIN_KEY[&format=json]
    """
    if request.args.get("key") != ADMIN_KEY:
        return jsonify({"error": "forbidden, append ?key=YOUR_ADMIN_KEY"}), 403

    fmt = request.args.get("format", "html")
    issues = []

    # ── 環境變數檢查 ──
    env_ok = {
        "LINE_CHANNEL_ACCESS_TOKEN": bool(LINE_TOKEN),
        "LINE_CHANNEL_SECRET": bool(LINE_SECRET),
        "OPENAI_API_KEY": bool(OPENAI_KEY),
        "ADMIN_KEY": bool(ADMIN_KEY) and ADMIN_KEY != "changeme",
    }
    try:
        env_ok["GITHUB_TOKEN"] = bool(GITHUB_TOKEN)
    except NameError:
        env_ok["GITHUB_TOKEN"] = False
    for k, v in env_ok.items():
        if not v:
            issues.append(("error", f"環境變數未設或使用預設值: {k}"))

    # ── 模型相容性檢查 ──
    incompat = []
    if logprobs_enabled and not (model_supports(model_default, "logprobs") or model_supports(model_upgrade, "logprobs")):
        incompat.append("Logprobs(信心度計算)— 目前模型不支援,已自動忽略")
    if translation_temperature != 0 and not (model_supports(model_default, "temperature") or model_supports(model_upgrade, "temperature")):
        incompat.append(f"temperature={translation_temperature} — reasoning model 強制 1.0")
    if translation_seed != 0 and not (model_supports(model_default, "seed") or model_supports(model_upgrade, "seed")):
        incompat.append(f"seed={translation_seed} — reasoning model 不支援可重現種子")
    try:
        if stop_sequences_enabled and not (model_supports(model_default, "logit_bias") or model_supports(model_upgrade, "logit_bias")):
            incompat.append("Stop sequences + logit_bias — reasoning model 不支援")
    except NameError:
        pass
    for s in incompat:
        issues.append(("info", "進階設定自動忽略: " + s))

    # ── 資料狀態 ──
    try:
        log_path = TRANSLATION_LOG_FILE
        log_persistent = log_path.startswith(("/var/data", "/data"))
    except NameError:
        log_path = "?"; log_persistent = False
    if not log_persistent:
        issues.append(("warning", f"翻譯日誌路徑 {log_path},Render 重啟會清空(免費版限制)"))

    last_status = "no_recent_translate"
    last_err = ""
    try:
        if last_translate_debug:
            last_status = last_translate_debug.get("openai_status", "unknown")
            if last_status == "exception":
                last_err = last_translate_debug.get("openai_error", "")[:300]
                issues.append(("error", f"最近一次 OpenAI 翻譯失敗: {last_err}"))
    except NameError:
        pass

    data = {
        "status": "ok" if not any(s == "error" for s, _ in issues) else "issues",
        "version": VERSION,
        "uptime_seconds": int(time.time() - bot_start_time),
        "environment": env_ok,
        "models": {
            "translate_default": {"name": model_default, "family": _model_family(model_default), "threshold": model_threshold},
            "translate_upgrade": {"name": model_upgrade, "family": _model_family(model_upgrade)},
        },
        "advanced_settings_active": {
            "fewshot_mode": fewshot_mode,
            "tone": translation_tone,
            "structured_output": structured_output_enabled,
            "logprobs": logprobs_enabled,
            "double_check": double_check_mode,
            "send_metadata": send_metadata_to_openai,
        },
        "incompatible_auto_ignored": incompat,
        "data_status": {
            "custom_examples": len(custom_translation_examples),
            "translation_log_entries": len(translation_log),
            "translation_log_persistent": log_persistent,
            "translation_cache_entries": len(translation_cache) if 'translation_cache' in globals() else 0,
        },
        "last_translate_status": last_status,
        "last_translate_error": last_err,
        "issues": [{"severity": s, "message": m} for s, m in issues],
    }

    try:
        data["models"]["vision"] = {"name": VISION_MODEL, "family": _model_family(VISION_MODEL)}
    except NameError:
        pass

    if fmt == "json":
        return jsonify(data)

    # ── HTML view ──
    from flask import Response
    h = ['<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">']
    h.append('<title>Health Check</title>')
    h.append('<style>body{font-family:-apple-system,sans-serif;background:#0d0d1a;color:#e0e0e0;padding:12px;font-size:14px;line-height:1.6}')
    h.append('.card{background:#1a1a2e;border:1px solid #3a3a4e;border-radius:8px;padding:12px;margin-bottom:10px}')
    h.append('h2{color:#7c6fef;margin:0 0 10px;font-size:15px;border-bottom:1px solid #3a3a4e;padding-bottom:6px}')
    h.append('.k{color:#7c6fef}.ok{color:#43b581}.warn{color:#faa61a}.err{color:#f04747}.dim{color:#8a8a9a}')
    h.append('table{width:100%;border-collapse:collapse;font-size:13px}td{padding:4px 8px;border-bottom:1px solid #2a2a3e;vertical-align:top}')
    h.append('td:first-child{color:#8a8a9a;width:42%}')
    h.append('.badge{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:bold}')
    h.append('.b-ok{background:#1a4a2e;color:#43b581}.b-warn{background:#4a3a1a;color:#faa61a}.b-err{background:#4a1a1a;color:#f04747}')
    h.append('</style></head><body>')

    overall = data["status"]
    badge = '<span class="badge b-ok">✅ 全部正常</span>' if overall == "ok" else '<span class="badge b-err">⚠️ 有問題</span>'
    h.append(f'<h1 style="font-size:18px;margin:0 0 12px">🩺 Health Check {badge}</h1>')
    h.append(f'<div class="dim" style="font-size:12px;margin-bottom:12px">{VERSION} · uptime {data["uptime_seconds"]}s</div>')

    if issues:
        h.append('<div class="card"><h2>⚠️ 待處理 / 提醒</h2>')
        for sev, msg in issues:
            cls = "err" if sev == "error" else ("warn" if sev == "warning" else "dim")
            icon = "❌" if sev == "error" else ("⚠️" if sev == "warning" else "💡")
            h.append(f'<div class="{cls}" style="margin-bottom:6px">{icon} {msg}</div>')
        h.append('</div>')
    else:
        h.append('<div class="card"><h2>✅ 沒有偵測到問題</h2><div class="ok">所有設定正常運作</div></div>')

    h.append('<div class="card"><h2>🔐 環境變數</h2><table>')
    for k, v in env_ok.items():
        b = '<span class="badge b-ok">已設</span>' if v else '<span class="badge b-err">未設</span>'
        h.append(f'<tr><td>{k}</td><td>{b}</td></tr>')
    h.append('</table></div>')

    fL = {"gpt5_reasoning": "GPT-5 系列(reasoning)", "oseries_reasoning": "o-series(reasoning)",
          "gpt4_classic": "GPT-4 系列(classical)", "unknown": "未知"}
    h.append('<div class="card"><h2>🤖 模型設定</h2><table>')
    for label, key in [("短訊息預設", "translate_default"), ("長訊息升級", "translate_upgrade"), ("照片分析", "vision")]:
        if key in data["models"]:
            m = data["models"][key]
            h.append(f'<tr><td>{label}</td><td><b>{m["name"]}</b><br><span class="dim">{fL.get(m["family"], m["family"])}</span></td></tr>')
    h.append(f'<tr><td>切換門檻</td><td>{model_threshold} 字{("(未啟用自動切換)" if model_threshold == 0 else "")}</td></tr>')
    h.append('</table></div>')

    h.append('<div class="card"><h2>⚙️ 進階設定</h2><table>')
    for k, v in data["advanced_settings_active"].items():
        h.append(f'<tr><td>{k}</td><td>{v}</td></tr>')
    h.append('</table>')
    if incompat:
        h.append('<div class="warn" style="margin-top:8px;font-size:12px;padding:6px 8px;background:#0d0d1a;border-radius:6px"><b>自動忽略中:</b><br>')
        for s in incompat:
            h.append('• ' + s + '<br>')
        h.append('</div>')
    h.append('</div>')

    ds = data["data_status"]
    h.append('<div class="card"><h2>📊 資料狀態</h2><table>')
    h.append(f'<tr><td>自訂範例</td><td>{ds["custom_examples"]} 筆</td></tr>')
    h.append(f'<tr><td>翻譯日誌</td><td>{ds["translation_log_entries"]} 筆 · {("✅ 持久化" if ds["translation_log_persistent"] else "⚠️ 重啟會清空")}</td></tr>')
    h.append(f'<tr><td>翻譯快取</td><td>{ds["translation_cache_entries"]} 筆</td></tr>')
    h.append('</table></div>')

    h.append('<div class="card"><h2>📨 最近翻譯</h2>')
    if last_status == "no_recent_translate":
        h.append('<div class="dim">尚無翻譯紀錄(剛重啟)</div>')
    elif last_status == "success":
        try:
            last = last_translate_debug
            h.append('<table>')
            h.append(f'<tr><td>狀態</td><td><span class="ok">✅ success</span></td></tr>')
            h.append(f'<tr><td>原文</td><td>{last.get("src_text", "")[:80]}</td></tr>')
            h.append(f'<tr><td>方向</td><td>{last.get("src_lang", "")} → {last.get("tgt_lang", "")}</td></tr>')
            h.append(f'<tr><td>使用模型</td><td>{last.get("model_picked", "")}</td></tr>')
            h.append(f'<tr><td>範例對</td><td>{last.get("example_pairs_in_prompt", 0)} 對</td></tr>')
            h.append(f'<tr><td>譯文</td><td>{(last.get("openai_raw_response") or "")[:200]}</td></tr>')
            h.append('</table>')
        except Exception:
            h.append('<div class="dim">資料解析失敗</div>')
    else:
        h.append(f'<div class="err">❌ {last_status}: {last_err[:300]}</div>')
    key_param = request.args.get("key") or ""
    h.append(f'<div style="margin-top:8px;font-size:12px"><a href="/debug/last-translate?key={key_param}&format=html" style="color:#7c6fef">→ 看完整 debug 資料</a></div>')
    h.append('</div>')

    h.append(f'<div class="dim" style="font-size:11px;text-align:center;margin-top:20px">JSON: <a href="?key={key_param}&format=json" style="color:#7c6fef">?format=json</a></div>')
    h.append('</body></html>')
    return Response("\n".join(h), mimetype="text/html; charset=utf-8")


@app.route("/debug/event-log", methods=["GET"])
def debug_event_log():
    """v3.9.19: 看磁碟版 webhook 事件 log(跨 worker)
    
    Usage: GET /debug/event-log?key=KEY[&filter=image][&n=50]
    """
    if request.args.get("key") != ADMIN_KEY:
        return jsonify({"error": "forbidden"}), 403
    
    global _EVENT_LOG_FILE
    if _EVENT_LOG_FILE is None:
        _EVENT_LOG_FILE = _event_log_path()
    
    entries = []
    try:
        if os.path.exists(_EVENT_LOG_FILE):
            with open(_EVENT_LOG_FILE, "r", encoding="utf-8") as f:
                entries = json.load(f)
            if not isinstance(entries, list):
                entries = []
    except Exception as e:
        entries = []
    
    filter_str = request.args.get("filter", "").lower()
    # v3.9.30c B19: 用 _safe_int 防炸
    n = _safe_int(request.args.get("n", "50"), default=50, min_val=1, max_val=1000)
    
    if filter_str:
        entries = [e for e in entries if filter_str in e.get("type", "").lower()]
    entries = entries[-n:]
    entries.reverse()  # 新的在最上
    
    if request.args.get("format") == "json":
        return jsonify({"count": len(entries), "log_file": _EVENT_LOG_FILE, "entries": entries})
    
    # HTML
    html = "<html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>"
    html += "<style>body{font-family:monospace;padding:10px;background:#0d0d1a;color:#e0e0e0;font-size:12px;}"
    html += "h2,h3{color:#7c6fef}"
    html += ".entry{background:#1a1a2e;padding:8px;border-radius:6px;border:1px solid #2a2a3e;margin:6px 0}"
    html += ".ts{color:#8a8a9a;font-size:11px}"
    html += ".type{color:#7FB3FF;font-weight:bold}"
    html += ".webhook_in{border-left:3px solid #43b581}"
    html += ".image_handler_entered{border-left:3px solid #43b581}"
    html += ".image_skipped{border-left:3px solid #f0a020}"
    html += ".image_mode_check{border-left:3px solid #7c6fef}"
    html += "pre{margin:4px 0;white-space:pre-wrap;word-break:break-all}"
    html += "a{color:#7FB3FF;margin-right:10px}"
    html += "</style></head><body>"
    html += "<h2>📋 Event Log</h2>"
    html += f"<p>檔案: {_EVENT_LOG_FILE}</p>"
    html += "<p>篩選: "
    html += f"<a href='?key={ADMIN_KEY}'>全部</a>"
    html += f"<a href='?key={ADMIN_KEY}&filter=image'>圖片</a>"
    html += f"<a href='?key={ADMIN_KEY}&filter=webhook_in'>Webhook 進入</a>"
    html += f"<a href='?key={ADMIN_KEY}&filter=image_skipped'>圖片被略過</a>"
    html += f"<a href='?key={ADMIN_KEY}&filter=vision'>Vision 呼叫</a>"
    html += f"<a href='?key={ADMIN_KEY}&filter=translate'>翻譯呼叫</a>"
    html += f"<a href='?key={ADMIN_KEY}&filter=ImgAsk'>詢問模式</a>"
    html += "</p>"
    
    if not entries:
        html += "<p>(沒有 log 紀錄。可能剛部署或 /tmp 不可寫。)</p>"
    else:
        html += f"<p>共 {len(entries)} 筆,新的在上</p>"
        import datetime as _dt
        for e in entries:
            ts_str = _dt.datetime.fromtimestamp(e.get("ts", 0)).strftime("%H:%M:%S")
            t = e.get("type", "?")
            d = e.get("data", {})
            html += f'<div class="entry {t}">'
            html += f'<div><span class="ts">{ts_str}</span> · <span class="type">{t}</span></div>'
            html += '<pre>' + json.dumps(d, ensure_ascii=False, indent=1) + '</pre>'
            html += '</div>'
    
    html += "</body></html>"
    from flask import Response
    return Response(html, mimetype="text/html")


@app.route("/debug/recent-images", methods=["GET"])
def debug_recent_images():
    """v3.9.18: 列出最近收到的 LINE 圖片,點擊就能用 vision-test 測"""
    if request.args.get("key") != ADMIN_KEY:
        return jsonify({"error": "forbidden"}), 403
    
    html = "<html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>"
    html += "<style>body{font-family:monospace;padding:10px;background:#0d0d1a;color:#e0e0e0;font-size:13px;}"
    html += "h2{color:#7c6fef}a{color:#7FB3FF;display:block;margin:10px 0;padding:10px;background:#1a1a2e;border-radius:6px;text-decoration:none;border:1px solid #2a2a3e}"
    html += "a:hover{background:#2a2a3e}.ts{color:#8a8a9a;font-size:11px}</style></head><body>"
    html += "<h2>📷 最近收到的 LINE 圖片</h2>"
    
    if not _last_image_received_msgs:
        html += "<p>(目前還沒收到任何圖片)<br>請先在 LINE 群組傳一張圖,然後重新整理本頁。</p>"
    else:
        html += f"<p>共 {len(_last_image_received_msgs)} 張(最新在最上)</p>"
        import datetime as _dt
        for img in reversed(_last_image_received_msgs):
            ts_str = _dt.datetime.fromtimestamp(img["ts"]).strftime("%H:%M:%S")
            test_url = f"/debug/vision-test?key={ADMIN_KEY}&msg_id={img['msg_id']}"
            html += f'<a href="{test_url}">'
            html += f'<div>msg_id: {img["msg_id"]}</div>'
            html += f'<div>group: {img["group_id"][:20] if img["group_id"] else "DM"}</div>'
            html += f'<div class="ts">收到時間: {ts_str}</div>'
            html += f'<div style="color:#43b581;margin-top:4px">▶ 點此用此圖測試 vision</div>'
            html += '</a>'
    
    html += "</body></html>"
    from flask import Response
    return Response(html, mimetype="text/html")


@app.route("/debug/vision-test", methods=["GET"])
def debug_vision_test():
    """v3.9.17: 用最近一張真實 LINE 圖片測 vision call
    
    Usage:
      GET /debug/vision-test?key=KEY                      → 用 hardcode 簡單測試圖
      GET /debug/vision-test?key=KEY&msg_id=XXX           → 用指定的 LINE message_id 測
      GET /debug/vision-test?key=KEY&model=gpt-4o-mini   → 指定模型
    """
    if request.args.get("key") != ADMIN_KEY:
        return jsonify({"error": "forbidden"}), 403
    if not oai:
        return jsonify({"error": "OpenAI client not initialized"}), 500
    
    requested_model = request.args.get("model", "")
    line_msg_id = request.args.get("msg_id", "")
    
    result = {
        "ok": False,
        "vision_model_global": VISION_MODEL,
        "vision_fallback_global": VISION_FALLBACK_MODEL,
        "stages": [],
    }
    
    # Stage 0: 取得測試圖片
    img_b64 = None
    img_mime = "image/jpeg"
    if line_msg_id:
        # 用真實 LINE 圖片
        try:
            result["stages"].append({"stage": "downloading_line_image", "msg_id": line_msg_id})
            img_b64, img_raw = download_line_image(line_msg_id)
            if img_b64 and img_raw:
                # 偵測格式
                if img_raw[:3] == b'\xff\xd8\xff':
                    img_mime = "image/jpeg"
                elif img_raw[:8] == b'\x89PNG\r\n\x1a\n':
                    img_mime = "image/png"
                elif img_raw[:6] in (b'GIF87a', b'GIF89a'):
                    img_mime = "image/gif"
                elif img_raw[:4] == b'RIFF' and img_raw[8:12] == b'WEBP':
                    img_mime = "image/webp"
                else:
                    # 看前 12 byte 是什麼
                    img_mime = "unknown (first 12 bytes: " + img_raw[:12].hex() + ")"
                result["stages"].append({
                    "stage": "line_image_downloaded",
                    "size_bytes": len(img_raw),
                    "detected_mime": img_mime,
                    "first_bytes_hex": img_raw[:16].hex(),
                })
            else:
                result["stages"].append({"stage": "line_download_failed"})
                return jsonify(result)
        except Exception as e:
            result["stages"].append({"stage": "line_download_exception", "error": str(e)})
            return jsonify(result)
    else:
        # 用一個確定有效的小 PNG(2x2 紅色方塊)
        # 用 Python 動態產生確保 PNG header 正確
        import struct, zlib
        def _make_simple_png():
            # 2x2 紅色方塊 RGB PNG
            width, height = 2, 2
            raw = b'\x00' + b'\xff\x00\x00' * width  # filter byte + RGB pixels
            raw = raw * height
            compressed = zlib.compress(raw)
            png = b'\x89PNG\r\n\x1a\n'
            # IHDR chunk
            ihdr_data = struct.pack('>IIBBBBB', width, height, 8, 2, 0, 0, 0)
            png += struct.pack('>I', 13) + b'IHDR' + ihdr_data
            png += struct.pack('>I', zlib.crc32(b'IHDR' + ihdr_data))
            # IDAT chunk
            png += struct.pack('>I', len(compressed)) + b'IDAT' + compressed
            png += struct.pack('>I', zlib.crc32(b'IDAT' + compressed))
            # IEND chunk
            png += struct.pack('>I', 0) + b'IEND'
            png += struct.pack('>I', zlib.crc32(b'IEND'))
            return png
        png_bytes = _make_simple_png()
        img_b64 = base64.b64encode(png_bytes).decode("utf-8")
        img_mime = "image/png"
        result["stages"].append({
            "stage": "using_synthetic_png",
            "size_bytes": len(png_bytes),
            "first_bytes_hex": png_bytes[:16].hex(),
        })
    
    # Stage 1: 直接呼叫 OpenAI
    try:
        target_model = requested_model or VISION_MODEL
        result["stages"].append({"stage": "raw_call_attempt", "model": target_model, "mime": img_mime})
        
        kwargs = {
            "model": target_model,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "Describe what you see in this image briefly."},
                        {"type": "image_url", "image_url": {
                            "url": f"data:{img_mime if not img_mime.startswith('unknown') else 'image/jpeg'};base64,{img_b64}"
                        }},
                    ]
                }
            ],
            "timeout": 30,
        }
        # v3.9.28: 用 model_supports 過濾,跟主流程一致
        if model_supports(target_model, "max_completion_tokens"):
            kwargs["max_completion_tokens"] = 3000
        elif model_supports(target_model, "max_tokens"):
            kwargs["max_tokens"] = 500
        if model_supports(target_model, "temperature"):
            kwargs["temperature"] = 0.0
        if model_supports(target_model, "reasoning_effort"):
            _opt = optimal_reasoning_for_translation(target_model)
            if _opt:
                kwargs["reasoning_effort"] = _opt
        if model_supports(target_model, "verbosity"):
            kwargs["verbosity"] = "high"  # vision/OCR 任務
        
        r = oai.chat.completions.create(**kwargs)
        content = r.choices[0].message.content if r.choices else None
        finish = r.choices[0].finish_reason if r.choices else None
        usage = r.usage.model_dump() if hasattr(r, 'usage') and r.usage else None
        
        result["stages"].append({
            "stage": "raw_call_ok",
            "content": content,
            "content_len": len(content) if content else 0,
            "finish_reason": finish,
            "usage": usage,
        })
        result["ok"] = True
    except Exception as e:
        import traceback
        result["stages"].append({
            "stage": "raw_call_failed",
            "error": str(e),
            "error_type": type(e).__name__,
            "traceback": traceback.format_exc()[-1500:],
        })
    
    # 用 HTML 格式回應
    html = "<html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>"
    html += "<style>body{font-family:monospace;padding:10px;background:#0d0d1a;color:#e0e0e0;font-size:12px;}"
    html += "pre{white-space:pre-wrap;word-break:break-all;background:#1a1a2e;padding:8px;border-radius:6px;border:1px solid #2a2a3e;}"
    html += "h2,h3{color:#7c6fef}.ok{color:#43b581}.err{color:#f04747}"
    html += "a{color:#7FB3FF}</style></head><body>"
    html += "<h2>🔍 Vision Debug Report</h2>"
    html += f"<p>VISION_MODEL: <b>{VISION_MODEL}</b> · FALLBACK: <b>{VISION_FALLBACK_MODEL}</b></p>"
    html += f"<p>整體狀態: <span class='{'ok' if result['ok'] else 'err'}'>{'✅ 成功' if result['ok'] else '❌ 失敗'}</span></p>"
    html += "<hr>"
    html += "<p><b>用法:</b><br>"
    html += "&bull; <code>?key=KEY</code> → 用合成的 PNG 測試<br>"
    html += "&bull; <code>?key=KEY&amp;msg_id=LINE_MSG_ID</code> → 用真實 LINE 圖片測試<br>"
    html += "&bull; <code>?key=KEY&amp;model=gpt-4o-mini</code> → 換模型測<br>"
    html += "</p><hr>"
    html += "<h3>各階段:</h3>"
    import json as _json
    html += "<pre>" + _json.dumps(result["stages"], ensure_ascii=False, indent=2) + "</pre>"
    html += "</body></html>"
    
    fmt = request.args.get("format", "html")
    if fmt == "json":
        return jsonify(result)
    from flask import Response
    return Response(html, mimetype="text/html")


@app.route("/debug/last-translate", methods=["GET"])
def debug_last_translate():
    """v3.9.5: Inspect the last translation that hit translate_openai.
    Shows the EXACT messages array that was sent to OpenAI, so we can verify
    custom examples are reaching the model. Requires ADMIN_KEY for security.

    Usage: GET /debug/last-translate?key=YOUR_ADMIN_KEY[&format=html]
    """
    if request.args.get("key") != ADMIN_KEY:
        return jsonify({"error": "forbidden, append ?key=YOUR_ADMIN_KEY"}), 403
    if not last_translate_debug:
        return jsonify({"error": "no translation has been made since restart"}), 404

    fmt = request.args.get("format", "json")
    if fmt == "html":
        # Pretty HTML for mobile reading
        d = last_translate_debug
        msgs = d.get("messages_sent", [])
        from flask import Response
        html = ['<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">']
        html.append('<title>Last Translate Debug</title>')
        html.append('<style>body{font-family:-apple-system,sans-serif;background:#0d0d1a;color:#e0e0e0;padding:12px;font-size:14px;line-height:1.5}')
        html.append('.box{background:#1a1a2e;border:1px solid #3a3a4e;border-radius:8px;padding:10px;margin-bottom:10px}')
        html.append('.k{color:#7c6fef;font-weight:bold}.v{color:#43b581}.warn{color:#faa61a}.err{color:#f04747}')
        html.append('.role{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:bold;margin-right:6px}')
        html.append('.role-system{background:#5a4a8a;color:#fff}.role-user{background:#2a5a8a;color:#fff}.role-assistant{background:#2a8a4a;color:#fff}')
        html.append('pre{white-space:pre-wrap;word-break:break-word;margin:4px 0;font-size:12px}')
        html.append('h2{color:#7c6fef;border-bottom:1px solid #3a3a4e;padding-bottom:4px}</style></head><body>')
        html.append('<h2>📋 Last Translate Debug</h2>')
        html.append('<div class="box">')
        html.append(f'<div><span class="k">原文:</span> <span class="v">{d.get("src_text","")}</span></div>')
        html.append(f'<div><span class="k">方向:</span> {d.get("src_lang","?")} → {d.get("tgt_lang","?")}</div>')
        html.append(f'<div><span class="k">使用模型:</span> <span class="v">{d.get("model_picked","?")}</span></div>')
        html.append(f'<div><span class="k">Few-shot 模式:</span> {d.get("fewshot_mode","?")}</div>')
        html.append(f'<div><span class="k">Tone:</span> {d.get("tone","?")}</div>')
        ec = d.get("example_pairs_in_prompt", 0)
        ec_class = "v" if ec >= 3 else ("warn" if ec >= 1 else "err")
        html.append(f'<div><span class="k">範例對數量:</span> <span class="{ec_class}">{ec}</span> {"⚠️ 太少!" if ec < 3 else "✅"}</div>')
        html.append(f'<div><span class="k">OpenAI 狀態:</span> <span class="{"v" if d.get("openai_status")=="success" else "err"}">{d.get("openai_status","?")}</span></div>')
        if d.get("openai_error"):
            html.append(f'<div><span class="k err">錯誤:</span> <span class="err">{d.get("openai_error")}</span></div>')
            html.append(f'<div><span class="k err">錯誤類型:</span> <span class="err">{d.get("openai_error_type","?")}</span></div>')
        if d.get("openai_raw_response"):
            html.append(f'<div><span class="k">原始回應:</span></div><pre class="v">{d.get("openai_raw_response","")[:1000]}</pre>')
        html.append(f'<div><span class="k">送出的 kwargs keys:</span> <span class="v">{", ".join(d.get("kwargs_keys_sent", []))}</span></div>')
        html.append('</div>')

        html.append(f'<h2>📨 送給 OpenAI 的完整 messages (共 {len(msgs)} 則)</h2>')
        for i, m in enumerate(msgs):
            role = m.get("role", "?")
            name = m.get("name", "")
            content = (m.get("content") or "")
            # truncate long system prompts
            display = content if len(content) < 800 or role != "system" else (content[:800] + f"\n...[truncated {len(content)-800} chars]")
            html.append(f'<div class="box">')
            html.append(f'<span class="role role-{role}">[{i}] {role}{(":"+name) if name else ""}</span>')
            html.append(f'<pre>{display}</pre>')
            html.append('</div>')
        html.append('</body></html>')
        return Response("\n".join(html), mimetype="text/html; charset=utf-8")

    # default: json
    return jsonify(last_translate_debug)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
